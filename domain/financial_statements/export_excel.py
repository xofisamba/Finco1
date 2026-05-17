"""Excel audit export for offline financial statements."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domain.financial_statements.excel_mapping import (
    BS_ROW_MAPPINGS,
    PF_CASH_WATERFALL_ROW_MAPPINGS,
    PNL_ROW_MAPPINGS,
    ExcelRowMapping,
)
from domain.financial_statements.result import FinancialStatementsResult


K_EUR_FORMAT = '#,##0.0;[Red](#,##0.0);"-"'
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
PLACEHOLDER_FILL = PatternFill("solid", fgColor="FFF2CC")
RESIDUAL_FILL = PatternFill("solid", fgColor="FCE4D6")
BALANCE_ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")


def export_financial_statements_audit_workbook(
    statements: FinancialStatementsResult,
    output_path: Path,
    project_name: str | None = None,
    include_source_mapping: bool = True,
) -> Path:
    """Export a human-readable audit workbook from assembled statements."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    _write_summary(summary, statements, project_name)
    _write_statement_sheet(
        workbook.create_sheet("P&L"),
        PNL_ROW_MAPPINGS,
        statements.pnl.periods,
        period_attr="period",
    )
    _write_statement_sheet(
        workbook.create_sheet("Balance Sheet"),
        BS_ROW_MAPPINGS,
        statements.balance_sheet.periods,
        period_attr="period_index",
        balance_check_field="balance_check_keur",
    )
    _write_statement_sheet(
        workbook.create_sheet("PF Cash Waterfall"),
        PF_CASH_WATERFALL_ROW_MAPPINGS,
        statements.pf_cash_waterfall.periods,
        period_attr="period_index",
    )

    if include_source_mapping:
        _write_source_mapping(workbook.create_sheet("Source Mapping"))
    _write_known_gaps(workbook.create_sheet("Known Gaps"))

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    workbook.save(output_path)
    return output_path


def _write_summary(
    sheet: Worksheet,
    statements: FinancialStatementsResult,
    project_name: str | None,
) -> None:
    sheet["A1"] = "Financial Statements Audit Export"
    sheet["A1"].font = Font(bold=True, size=14)

    pf_periods = statements.pf_cash_waterfall.periods
    pnl_periods = statements.pnl.periods

    rows = [
        ("Project name", project_name or "Not specified"),
        ("Generated timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Period count", len(pnl_periods)),
        ("Total revenue", sum(period.revenues_keur for period in pnl_periods)),
        ("Total OPEX", sum(period.operating_expenses_keur for period in pnl_periods)),
        ("Total EBITDA", sum(period.ebitda_cash_keur for period in pf_periods)),
        ("Total senior DS", sum(period.senior_total_ds_keur for period in pf_periods)),
        ("Total cash tax", sum(period.cash_tax_keur for period in pf_periods)),
        (
            "Total SHL service",
            sum(
                period.shl_cash_interest_keur + period.shl_principal_keur
                for period in pf_periods
            ),
        ),
        ("Total distributions", sum(period.net_dividends_keur for period in pf_periods)),
        ("Max BS balance check", statements.balance_sheet.max_abs_balance_check_keur),
        (
            "R99/R102 identity status",
            "OK" if statements.pf_cash_waterfall.r99_r102_identity_holds else "Check",
        ),
        ("Known placeholders count", len(_known_gap_rows())),
    ]

    for row_idx, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_idx, column=1, value=label)
        sheet.cell(row=row_idx, column=2, value=value)
        if isinstance(value, (int, float)) and label != "Period count":
            sheet.cell(row=row_idx, column=2).number_format = K_EUR_FORMAT


def _write_statement_sheet(
    sheet: Worksheet,
    mappings: tuple[ExcelRowMapping, ...],
    periods: Iterable[object],
    *,
    period_attr: str,
    balance_check_field: str | None = None,
) -> None:
    period_list = tuple(periods)
    headers = ["Excel row", "Row label", "Source owner"]
    headers.extend(_period_label(period, period_attr) for period in period_list)
    headers.append("Total")

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, mapping in enumerate(mappings, start=2):
        sheet.cell(row=row_idx, column=1, value=mapping.row_code)
        sheet.cell(row=row_idx, column=2, value=mapping.label)
        sheet.cell(row=row_idx, column=3, value=mapping.source_owner)

        values = [getattr(period, mapping.field_name) for period in period_list]
        for col_idx, value in enumerate(values, start=4):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = K_EUR_FORMAT
        total_cell = sheet.cell(row=row_idx, column=4 + len(values), value=sum(values))
        total_cell.number_format = K_EUR_FORMAT

        status = _mapping_status(sheet.title, mapping)
        if status == "placeholder":
            _fill_row(sheet, row_idx, PLACEHOLDER_FILL)
        elif status == "residual":
            _fill_row(sheet, row_idx, RESIDUAL_FILL)

        if balance_check_field and mapping.field_name == balance_check_field and period_list:
            first_value_col = 4
            last_value_col = 3 + len(period_list)
            value_range = (
                f"{get_column_letter(first_value_col)}{row_idx}:"
                f"{get_column_letter(last_value_col)}{row_idx}"
            )
            sheet.conditional_formatting.add(
                value_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0.01"],
                    fill=BALANCE_ALERT_FILL,
                ),
            )
            sheet.conditional_formatting.add(
                value_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["-0.01"],
                    fill=BALANCE_ALERT_FILL,
                ),
            )


def _write_source_mapping(sheet: Worksheet) -> None:
    headers = ["Sheet", "Excel row", "Label", "Source field", "Source owner", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    mappings_by_sheet = (
        ("P&L", PNL_ROW_MAPPINGS),
        ("Balance Sheet", BS_ROW_MAPPINGS),
        ("PF Cash Waterfall", PF_CASH_WATERFALL_ROW_MAPPINGS),
    )
    row_idx = 2
    for sheet_name, mappings in mappings_by_sheet:
        for mapping in mappings:
            status = _mapping_status(sheet_name, mapping)
            values = [
                sheet_name,
                mapping.row_code,
                mapping.label,
                mapping.field_name,
                mapping.source_owner,
                status,
                mapping.notes,
            ]
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
            if status == "placeholder":
                _fill_row(sheet, row_idx, PLACEHOLDER_FILL)
            elif status == "residual":
                _fill_row(sheet, row_idx, RESIDUAL_FILL)
            row_idx += 1


def _write_known_gaps(sheet: Worksheet) -> None:
    headers = ["Gap", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(_known_gap_rows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        if row[1] in {"placeholder", "residual"}:
            _fill_row(sheet, row_idx, RESIDUAL_FILL if row[1] == "residual" else PLACEHOLDER_FILL)


def _known_gap_rows() -> tuple[tuple[str, str, str], ...]:
    return (
        (
            "Gross fixed assets",
            "placeholder",
            "Uses accumulated depreciation until a statement-safe capex ledger is exposed.",
        ),
        ("Share capital", "placeholder", "Not exposed on WaterfallResult."),
        ("Legal reserve", "placeholder", "Stage 2 P&L placeholder; no runtime source migration."),
        ("Cash", "residual", "Explicit Balance Sheet balancing line, not runtime cash."),
        ("R99/R102", "audit-only", "Shown from existing audit fields; not accepted as runtime source."),
        (
            "Detailed Excel rows",
            "placeholder",
            "Rows without exposed runtime/audit fields are not inferred.",
        ),
    )


def _period_label(period: object, period_attr: str) -> str:
    period_number = getattr(period, period_attr)
    period_date = getattr(period, "date")
    return f"{period_number} | {period_date.isoformat()}"


def _mapping_status(sheet_name: str, mapping: ExcelRowMapping) -> str:
    source = mapping.source_owner.lower()
    field = mapping.field_name.lower()
    notes = mapping.notes.lower()
    if "placeholder" in source or "placeholder" in notes:
        return "placeholder"
    if "residual" in source or "residual" in field:
        return "residual"
    if "audit" in source or mapping.row_code in {"CF_R69", "CF_R84", "CF_R98", "CF_R99", "CF_R100", "CF_R102"}:
        return "audit-only"
    if "assembled" in source or "p&l" in source or "cumulative" in source:
        return "derived"
    if "waterfallperiod" in source:
        return "runtime output"
    if sheet_name == "Balance Sheet" and mapping.field_name in {"total_assets_keur", "total_liabilities_equity_keur", "balance_check_keur"}:
        return "derived"
    return "derived"


def _format_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "D2" if sheet.title in {"P&L", "Balance Sheet", "PF Cash Waterfall"} else "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 34)

    if sheet.title == "Summary":
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 24
        for cell in sheet["A"]:
            if cell.row > 1:
                cell.fill = SUBHEADER_FILL
                cell.font = Font(bold=True)


def _fill_row(sheet: Worksheet, row_idx: int, fill: PatternFill) -> None:
    for cell in sheet[row_idx]:
        cell.fill = fill
