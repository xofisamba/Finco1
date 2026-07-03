"""CSV/Excel export utilities for waterfall results.

Provides functions to export:
- Waterfall period data (CSV)
- Summary metrics (CSV)
- Financial statements (CSV)
- Tax audit bridge (CSV)
- Formula source documentation (CSV)

Usage:
    from utils.export import export_waterfall_csv
    export_waterfall_csv(result, "waterfall_output.csv")
"""
import csv
from pathlib import Path
from typing import Optional

from domain.waterfall.waterfall_engine import WaterfallResult

# ---------------------------------------------------------------------------
# V2 — Formula source documentation
# Maps each CSV column to a short description of where it originates.
# ---------------------------------------------------------------------------
_FORMULA_SOURCES: dict[str, str] = {
    # Core period identification
    "period": "waterfall_engine.py — period counter (0-based)",
    "year_index": "waterfall_engine.py — calendar year of this half-period",
    "period_in_year": "waterfall_engine.py — 1=H1, 2=H2",
    "is_operation": "waterfall_engine.py — True once commercial operation date passed",
    # Revenue / OPEX
    "generation_mwh": "waterfall_engine.py — net generation after degradation curve",
    "revenue_keur": "waterfall_engine.py — generation × blended price (PPA + merchant)",
    "opex_keur": "waterfall_engine.py — opex_params.total_opex_keur half-period",
    "ebitda_keur": "waterfall_engine.py — revenue − opex",
    # P&L items
    "depreciation_keur": "waterfall_engine.py — book depreciation (IFRS straight-line)",
    "interest_senior_keur": "waterfall_engine.py — senior debt interest this half-period",
    "interest_shl_keur": "waterfall_engine.py — SHL PIK interest this half-period",
    "tax_keur": "waterfall_engine.py — accrued CIT this period (Pass-2 result)",
    "cf_after_tax_keur": "waterfall_engine.py — EBITDA − tax_keur (cash basis)",
    # Senior debt service
    "senior_ds_keur": "waterfall_engine.py — total senior debt service (interest + principal)",
    "senior_interest_keur_engine": "waterfall_engine.py — senior interest component",
    "senior_principal_keur_engine": "waterfall_engine.py — senior principal repayment",
    # SHL service
    "shl_service_keur": "domain/shl/canonical_wiring.py — total SHL cash service",
    "shl_interest_keur": "domain/shl/canonical_wiring.py — SHL interest paid",
    "shl_principal_keur": "domain/shl/canonical_wiring.py — SHL principal repaid",
    # Reserves / covenants
    "dsra_contribution_keur": "waterfall_engine.py — DSRA top-up contribution",
    "dsra_balance_keur": "waterfall_engine.py — DSRA closing balance",
    "cf_after_reserves_keur": "waterfall_engine.py — CFADS after DSRA movement",
    "dscr": "waterfall_engine.py — CFADS / senior DS (floor=0, cap=999)",
    "llcr": "waterfall_engine.py — NPV(future CFADS) / senior debt balance",
    "plcr": "waterfall_engine.py — NPV(all CFADS) / senior debt balance",
    # Distribution
    "lockup_active": "waterfall_engine.py — True when any covenant in breach",
    "distribution_keur": "waterfall_engine.py — equity distribution released",
    "cash_sweep_keur": "waterfall_engine.py — excess cash swept to debt prepayment",
    "cum_distribution_keur": "waterfall_engine.py — cumulative equity distributions",
    "cash_balance_keur": "waterfall_engine.py — project cash account closing balance",
    # ── Tax audit fields (V1 additions) ───────────────────────────────────
    "corporate_tax_cash_keur": "waterfall_engine.py — cash CIT paid this period (Pass-2)",
    "cit_accrual_audit_keur": "waterfall_engine.py — H1 CIT carry-forward to H2 settlement",
    "taxable_profit_keur": "waterfall_engine.py — taxable income display field (Pass-2)",
    "taxable_income_before_losses_audit_keur": (
        "domain/tax/ — EBITDA − depr − interest − fiscal_reint (before loss c/f)"
    ),
    "taxable_profit_after_losses_audit_keur": (
        "domain/tax/ — taxable income after loss carryforward consumed"
    ),
    "tax_loss_opening_audit_keur": "waterfall_engine.py — loss c/f opening balance this period",
    "tax_loss_used_audit_keur": "waterfall_engine.py — losses consumed this period",
    "tax_loss_closing_audit_keur": "waterfall_engine.py — losses remaining after this period",
    "fiscal_reintegration_audit_keur": (
        "waterfall_engine.py — IDC/fees reintegrated into taxable base (year 1 only)"
    ),
    "tax_depreciation_audit_keur": (
        "domain/depreciation/canonical_wiring.py — tax depreciation used in CIT calc"
    ),
    "cash_tax_current_period_audit_keur": (
        "waterfall_engine.py — cash CIT attributable to current period only"
    ),
    "cash_tax_excel_style_h2_diagnostic_keur": (
        "waterfall_engine.py — H2 diagnostic: prior H1 accrual + current period tax"
    ),
    "r67_excel_style_cash_tax_diagnostic_keur": (
        "waterfall_engine.py — R67 Excel-style diagnostic cross-check"
    ),
}


def export_waterfall_csv(result: WaterfallResult, filepath: str) -> None:
    """Export waterfall period data to CSV.
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output CSV file path
    """
    fieldnames = [
        "period", "year_index", "period_in_year", "is_operation",
        "generation_mwh", "revenue_keur", "opex_keur", "ebitda_keur",
        "depreciation_keur", "interest_senior_keur", "interest_shl_keur",
        # ── tax column + audit detail (V1) ──
        "tax_keur",
        "corporate_tax_cash_keur",
        "cit_accrual_audit_keur",
        "taxable_profit_keur",
        "taxable_income_before_losses_audit_keur",
        "taxable_profit_after_losses_audit_keur",
        "tax_loss_opening_audit_keur",
        "tax_loss_used_audit_keur",
        "tax_loss_closing_audit_keur",
        "fiscal_reintegration_audit_keur",
        "tax_depreciation_audit_keur",
        "cash_tax_current_period_audit_keur",
        "cash_tax_excel_style_h2_diagnostic_keur",
        "r67_excel_style_cash_tax_diagnostic_keur",
        # ── remaining cash-flow columns ──
        "cf_after_tax_keur",
        "senior_ds_keur", "senior_interest_keur_engine", "senior_principal_keur_engine",
        "shl_service_keur", "shl_interest_keur", "shl_principal_keur",
        "dsra_contribution_keur", "dsra_balance_keur",
        "cf_after_reserves_keur", "dscr", "llcr", "plcr",
        "lockup_active", "distribution_keur", "cash_sweep_keur",
        "cum_distribution_keur", "cash_balance_keur",
    ]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for p in result.periods:
            writer.writerow({
                "period": p.period,
                "year_index": p.year_index,
                "period_in_year": p.period_in_year,
                "is_operation": p.is_operation,
                "generation_mwh": round(p.generation_mwh, 2),
                "revenue_keur": round(p.revenue_keur, 2),
                "opex_keur": round(p.opex_keur, 2),
                "ebitda_keur": round(p.ebitda_keur, 2),
                "depreciation_keur": round(p.depreciation_keur, 2),
                "interest_senior_keur": round(p.interest_senior_keur, 2),
                "interest_shl_keur": round(p.interest_shl_keur, 2),
                "tax_keur": round(p.tax_keur, 2),
                "corporate_tax_cash_keur": round(getattr(p, "corporate_tax_cash_keur", 0.0), 2),
                "cit_accrual_audit_keur": round(getattr(p, "cit_accrual_audit_keur", 0.0), 2),
                "taxable_profit_keur": round(getattr(p, "taxable_profit_keur", 0.0), 2),
                "taxable_income_before_losses_audit_keur": round(getattr(p, "taxable_income_before_losses_audit_keur", 0.0), 2),
                "taxable_profit_after_losses_audit_keur": round(getattr(p, "taxable_profit_after_losses_audit_keur", 0.0), 2),
                "tax_loss_opening_audit_keur": round(getattr(p, "tax_loss_opening_audit_keur", 0.0), 2),
                "tax_loss_used_audit_keur": round(getattr(p, "tax_loss_used_audit_keur", 0.0), 2),
                "tax_loss_closing_audit_keur": round(getattr(p, "tax_loss_closing_audit_keur", 0.0), 2),
                "fiscal_reintegration_audit_keur": round(getattr(p, "fiscal_reintegration_audit_keur", 0.0), 2),
                "tax_depreciation_audit_keur": round(getattr(p, "tax_depreciation_audit_keur", 0.0), 2),
                "cash_tax_current_period_audit_keur": round(getattr(p, "cash_tax_current_period_audit_keur", 0.0), 2),
                "cash_tax_excel_style_h2_diagnostic_keur": round(getattr(p, "cash_tax_excel_style_h2_diagnostic_keur", 0.0), 2),
                "r67_excel_style_cash_tax_diagnostic_keur": round(getattr(p, "r67_excel_style_cash_tax_diagnostic_keur", 0.0), 2),
                "cf_after_tax_keur": round(p.cf_after_tax_keur, 2),
                "senior_ds_keur": round(p.senior_ds_keur, 2),
                "senior_interest_keur_engine": round(p.senior_interest_keur, 2),
                "senior_principal_keur_engine": round(p.senior_principal_keur, 2),
                "shl_service_keur": round(p.shl_service_keur, 2),
                "shl_interest_keur": round(p.shl_interest_keur, 2),
                "shl_principal_keur": round(p.shl_principal_keur, 2),
                "dsra_contribution_keur": round(p.dsra_contribution_keur, 2),
                "dsra_balance_keur": round(p.dsra_balance_keur, 2),
                "cf_after_reserves_keur": round(p.cf_after_reserves_keur, 2),
                "dscr": round(p.dscr, 4) if p.dscr < float('inf') else 999.0,
                "llcr": round(p.llcr, 4) if p.llcr < float('inf') else 999.0,
                "plcr": round(p.plcr, 4) if p.plcr < float('inf') else 999.0,
                "lockup_active": p.lockup_active,
                "distribution_keur": round(p.distribution_keur, 2),
                "cash_sweep_keur": round(p.cash_sweep_keur, 2),
                "cum_distribution_keur": round(p.cum_distribution_keur, 2),
                "cash_balance_keur": round(p.cash_balance_keur, 2),
            })


def export_formula_sources_csv(filepath: str) -> None:
    """V2 — Write a two-column CSV mapping column names to their formula/source.

    Args:
        filepath: Output CSV file path
    """
    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["column_name", "source"])
        for col, source in _FORMULA_SOURCES.items():
            writer.writerow([col, source])


# Tax audit column order for export_tax_audit_csv (logical bridge order)
_TAX_AUDIT_COLUMNS = [
    "period",
    "year_index",
    "ebitda_keur",
    "fiscal_reintegration_audit_keur",
    "taxable_income_before_losses_audit_keur",
    "tax_loss_opening_audit_keur",
    "tax_loss_used_audit_keur",
    "tax_loss_closing_audit_keur",
    "taxable_profit_after_losses_audit_keur",
    "tax_depreciation_audit_keur",
    "tax_rate_pct",
    "tax_accrued_keur",
    "cit_accrual_audit_keur",
    "corporate_tax_cash_keur",
]


def export_tax_audit_csv(result: WaterfallResult, filepath: str) -> None:
    """V3 — Write a complete tax calculation bridge CSV.

    One row per operating period, columns in logical audit order:
    period, year, ebitda, fiscal_reint, taxable_before_losses, losses_used,
    losses_remaining, taxable_after_losses, tax_rate, tax_accrued, h1_carry,
    cash_paid.

    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output CSV file path
    """
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_TAX_AUDIT_COLUMNS)
        writer.writeheader()

        for p in result.periods:
            if not p.is_operation:
                continue
            taxable_after = getattr(p, "taxable_profit_after_losses_audit_keur", 0.0)
            tax_accrued = p.tax_keur
            # Derive implied rate (avoid div-by-zero)
            if taxable_after and taxable_after > 0:
                tax_rate_pct = round(abs(tax_accrued) / taxable_after * 100, 4)
            else:
                tax_rate_pct = 0.0

            writer.writerow({
                "period": p.period,
                "year_index": p.year_index,
                "ebitda_keur": round(p.ebitda_keur, 2),
                "fiscal_reintegration_audit_keur": round(getattr(p, "fiscal_reintegration_audit_keur", 0.0), 2),
                "taxable_income_before_losses_audit_keur": round(getattr(p, "taxable_income_before_losses_audit_keur", 0.0), 2),
                "tax_loss_opening_audit_keur": round(getattr(p, "tax_loss_opening_audit_keur", 0.0), 2),
                "tax_loss_used_audit_keur": round(getattr(p, "tax_loss_used_audit_keur", 0.0), 2),
                "tax_loss_closing_audit_keur": round(getattr(p, "tax_loss_closing_audit_keur", 0.0), 2),
                "taxable_profit_after_losses_audit_keur": round(taxable_after, 2),
                "tax_depreciation_audit_keur": round(getattr(p, "tax_depreciation_audit_keur", 0.0), 2),
                "tax_rate_pct": tax_rate_pct,
                "tax_accrued_keur": round(tax_accrued, 2),
                "cit_accrual_audit_keur": round(getattr(p, "cit_accrual_audit_keur", 0.0), 2),
                "corporate_tax_cash_keur": round(getattr(p, "corporate_tax_cash_keur", 0.0), 2),
            })


def export_summary_csv(result: WaterfallResult, filepath: str) -> None:
    """Export summary metrics to CSV.
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output CSV file path
    """
    sculpt = result.sculpting_result
    
    rows = [
        ("=== REVENUE ===", ""),
        ("Total Revenue (kEUR)", f"{result.total_revenue_keur:,.0f}"),
        ("Total OPEX (kEUR)", f"{result.total_opex_keur:,.0f}"),
        ("Total EBITDA (kEUR)", f"{result.total_ebitda_keur:,.0f}"),
        ("Total Tax (kEUR)", f"{result.total_tax_keur:,.0f}"),
        ("", ""),
        ("=== DEBT ===", ""),
        ("Debt Amount (kEUR)", f"{sculpt.debt_keur:,.0f}"),
        ("Total Senior DS (kEUR)", f"{result.total_senior_ds_keur:,.0f}"),
        ("Total SHL Service (kEUR)", f"{result.total_shl_service_keur:,.0f}"),
        ("", ""),
        ("=== RETURNS ===", ""),
        ("Project IRR", f"{result.project_irr*100:.3f}%"),
        ("Equity IRR", f"{result.equity_irr*100:.3f}%" if result.equity_irr else "N/A"),
        ("Project NPV (kEUR)", f"{result.project_npv:,.0f}"),
        ("Equity NPV (kEUR)", f"{result.equity_npv:,.0f}" if result.equity_npv else "N/A"),
        ("", ""),
        ("=== COVENANTS ===", ""),
        ("Avg DSCR", f"{result.avg_dscr:.3f}"),
        ("Min DSCR", f"{result.min_dscr:.3f}"),
        ("Min LLCR", f"{result.min_llcr:.3f}" if result.min_llcr else "N/A"),
        ("Min PLCR", f"{result.min_plcr:.3f}" if result.min_plcr else "N/A"),
        ("Periods in Lockup", f"{result.periods_in_lockup}"),
        ("", ""),
        ("=== DISTRIBUTIONS ===", ""),
        ("Total Distribution (kEUR)", f"{result.total_distribution_keur:,.0f}"),
    ]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)


def waterfall_to_dataframe(result: WaterfallResult) -> "pd.DataFrame":
    """Convert waterfall result to pandas DataFrame.
    
    Args:
        result: WaterfallResult from run_waterfall
    
    Returns:
        DataFrame with period-level data
    """
    import pandas as pd
    
    data = []
    for p in result.periods:
        data.append({
            "Period": p.period,
            "Year": p.year_index,
            "H": p.period_in_year,
            "Op": p.is_operation,
            "Gen (MWh)": round(p.generation_mwh, 0),
            "Rev (kEUR)": round(p.revenue_keur, 0),
            "OPEX (kEUR)": round(p.opex_keur, 0),
            "EBITDA (kEUR)": round(p.ebitda_keur, 0),
            "Dep (kEUR)": round(p.depreciation_keur, 0),
            "Int Sen (kEUR)": round(p.interest_senior_keur, 0),
            "Tax (kEUR)": round(p.tax_keur, 0),
            "CFAT (kEUR)": round(p.cf_after_tax_keur, 0),
            "Sen DS (kEUR)": round(p.senior_ds_keur, 0),
            "DSCR": round(p.dscr, 2) if p.dscr < float('inf') else None,
            "Dist (kEUR)": round(p.distribution_keur, 0),
            "Sweep (kEUR)": round(p.cash_sweep_keur, 0),
            "Cash Bal (kEUR)": round(p.cash_balance_keur, 0),
        })
    
    return pd.DataFrame(data)


def export_waterfall_excel(result: "WaterfallResult", filepath: str) -> None:
    """Export complete waterfall analysis to formatted Excel workbook.
    
    Creates a multi-sheet workbook suitable for bank/investor presentation:
    - Sheet 1: Summary (key metrics)
    - Sheet 2: Waterfall (period-level cash flows)
    - Sheet 3: Debt Schedule
    - Sheet 4: Covenant Compliance
    
    Args:
        result: WaterfallResult from run_waterfall
        filepath: Output Excel file path
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    number_format = "#,##0"
    pct_format = "0.00%"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    # ===== SHEET 1: Summary =====
    ws = wb.active
    ws.title = "Summary"
    
    sc = result.sculpting_result
    summary_data = [
        ("Project Finance Summary", ""),
        ("", ""),
        ("Debt Sizing", ""),
        ("Debt (kEUR)", f"{sc.debt_keur:,.0f}"),
        ("Avg DSCR", f"{result.avg_dscr:.3f}"),
        ("Min DSCR", f"{result.min_dscr:.3f}"),
        ("", ""),
        ("Returns", ""),
        ("Project IRR", f"{result.project_irr*100:.2f}%"),
        ("Equity IRR", f"{result.equity_irr*100:.2f}%" if result.equity_irr else "N/A"),
        ("", ""),
        ("Cash Flows (kEUR)", ""),
        ("Total Revenue", f"{result.total_revenue_keur:,.0f}"),
        ("Total Distributions", f"{result.total_distribution_keur:,.0f}"),
        ("Total Senior Debt Service", f"{result.total_senior_ds_keur:,.0f}"),
        ("Total Tax", f"{result.total_tax_keur:,.0f}"),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label and not value:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # ===== SHEET 2: Waterfall =====
    ws2 = wb.create_sheet("Waterfall")
    
    wf_headers = ["Period", "Year", "Half", "Gen (MWh)", "Rev (kEUR)", "EBITDA (kEUR)", 
                  "CFAT (kEUR)", "Sen DS (kEUR)", "DSCR", "Dist (kEUR)", "Sweep (kEUR)", "Lockup"]
    wf_data = []
    for p in result.periods:
        wf_data.append([
            p.period, p.year_index, "H1" if p.period_in_year == 1 else "H2",
            round(p.generation_mwh, 0), round(p.revenue_keur, 0), round(p.ebitda_keur, 0),
            round(p.cf_after_tax_keur, 0), round(p.senior_ds_keur, 0),
            round(p.dscr, 3) if p.dscr < float('inf') else 999,
            round(p.distribution_keur, 0), round(p.cash_sweep_keur, 0),
            "Y" if p.lockup_active else "N"
        ])
    
    for col, header in enumerate(wf_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header(ws2)
    for row_idx, row_data in enumerate(wf_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
    
    set_col_widths(ws2, [8, 6, 6, 12, 12, 12, 12, 12, 8, 12, 12, 8])
    
    # ===== SHEET 3: Debt Schedule =====
    ws3 = wb.create_sheet("Debt Schedule")
    
    ds_headers = ["Period", "Year", "Opening Bal", "Interest", "Principal", "Closing Bal", "DSCR"]
    ds_data = []
    for i, (bal, ir, pr) in enumerate(zip(sc.balance_schedule, sc.interest_schedule, sc.principal_schedule)):
        period = i
        year = i // 2 + 1
        half = "H1" if i % 2 == 0 else "H2"
        ds_data.append([period, year, half, round(bal, 0), round(ir, 0), round(pr, 0), round(sc.balance_schedule[i] if i < len(sc.balance_schedule) else 0, 0), round(sc.dscr_schedule[i], 3) if i < len(sc.dscr_schedule) and sc.dscr_schedule[i] < float('inf') else 999])
    
    for col, header in enumerate(ds_headers, 1):
        ws3.cell(row=1, column=col, value=header)
    style_header(ws3)
    for row_idx, row_data in enumerate(ds_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
    
    set_col_widths(ws3, [8, 6, 6, 14, 14, 14, 8])
    
    # ===== SHEET 4: Covenant =====
    ws4 = wb.create_sheet("Covenant")
    
    cov_headers = ["Year", "DSCR", "LLCR", "PLCR", "Lockup", "DSCR OK", "LLCR OK", "PLCR OK"]
    cov_data = []
    for p in result.periods:
        if p.is_operation and p.period_in_year == 2:
            dscr = p.dscr if p.dscr < float('inf') else 999
            llcr = p.llcr if p.llcr < float('inf') else 999
            plcr = p.plcr if p.plcr < float('inf') else 999
            cov_data.append([
                p.year_index, round(dscr, 3), round(llcr, 3), round(plcr, 3),
                "Y" if p.lockup_active else "N",
                "OK" if dscr >= 1.15 else ("WARN" if dscr >= 1.10 else "BREACH"),
                "OK" if llcr >= 1.15 else "BREACH",
                "OK" if plcr >= 1.20 else "BREACH",
            ])
    
    for col, header in enumerate(cov_headers, 1):
        ws4.cell(row=1, column=col, value=header)
    style_header(ws4)
    for row_idx, row_data in enumerate(cov_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # Color code status columns
            if col_idx >= 6 and isinstance(val, str):
                if val == "BREACH":
                    cell.fill = PatternFill("solid", fgColor="FF6B6B")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "WARN":
                    cell.fill = PatternFill("solid", fgColor="FFE066")
    
    set_col_widths(ws4, [8, 8, 8, 8, 8, 10, 10, 10])
    
    wb.save(filepath)
    print(f"Excel export saved to: {filepath}")
