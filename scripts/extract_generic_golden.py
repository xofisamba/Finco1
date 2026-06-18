"""G1A-EXTRACTOR: deterministic Tier-1 anchor extraction for the Generic
Solar / Generic Wind bootstrap reference workbooks.

Scope (see reports/g1a_extractor_report.md for the full scope audit):
  - Reads the two reference workbooks under validation/reference_models/
    using openpyxl ONLY. No formula evaluation library is invoked here and
    no Finco1 runtime code (waterfall_core, project_factories,
    input_adapter, etc.) is imported or executed.
  - The workbooks already carry cached formula values (baked in once, out
    of band, by recalculating the static spreadsheet -- the equivalent of
    opening in Excel/LibreOffice with automatic calculation and pressing
    Ctrl+Alt+F9). This script does not perform that recalculation; it only
    reads what is already cached in the file, exactly as a human auditor
    opening the workbook would see.
  - Output is deterministic: the same workbook bytes always produce the
    same JSON. The only "extraction_date" recorded is the workbook's own
    embedded core-properties timestamp (docProps/core.xml `modified`),
    not the wall-clock time the script happens to run at, so re-running
    this script against an unchanged workbook is byte-for-byte identical.

Usage:
    python scripts/extract_generic_golden.py
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl

EXTRACTOR_VERSION = "1.0.0"
SPEC_VERSION = "0.1"

REPO_ROOT = Path(__file__).resolve().parent.parent
REF_MODEL_DIR = REPO_ROOT / "validation" / "reference_models"
FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures"
MANIFEST_DIR = FIXTURE_DIR / "excel_reference"

SUMMARY_SHEET = "Summary"
SUMMARY_ANCHOR_ROWS = list(range(5, 20))  # C5..C19

# Tolerances per docs/generic_validation_reference_excel_spec.md S8.
TOLERANCES = {
    "total_revenue_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "total_opex_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "total_ebitda_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "total_capex_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "idc_keur": {"type": "relative_pct", "value": 0.01, "min_abs": 1.0},
    "bank_fees_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "senior_debt_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "senior_debt_service_p1_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "senior_debt_service_p2_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "senior_debt_service_p3_keur": {"type": "relative_pct", "value": 0.005, "min_abs": 1.0},
    "avg_dscr": {"type": "absolute", "value": 0.01},
    "min_dscr": {"type": "absolute", "value": 0.01},
    "project_irr": {"type": "absolute", "value": 0.0005},
    "equity_irr": {"type": "absolute", "value": 0.0005},
    "realized_gearing": {"type": "absolute_percentage_points", "value": 0.5},
}

PROJECTS = {
    "generic_solar": REF_MODEL_DIR / "GenericSolar_ReferenceModel.xlsx",
    "generic_wind": REF_MODEL_DIR / "GenericWind_ReferenceModel.xlsx",
}


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def extract_summary_anchors(workbook_path: Path) -> dict:
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

    ws_f = wb_formulas[SUMMARY_SHEET]
    ws_v = wb_values[SUMMARY_SHEET]

    anchors = {}
    for row in SUMMARY_ANCHOR_ROWS:
        name = ws_f.cell(row, 1).value
        if not name:
            continue
        cell_ref = f"C{row}"
        formula = ws_f.cell(row, 3).value
        value = ws_v.cell(row, 3).value
        anchors[name] = {
            "sheet": SUMMARY_SHEET,
            "cell": cell_ref,
            "formula": formula,
            "value": value,
        }

    modified = wb_formulas.properties.modified or wb_formulas.properties.created
    extraction_date = modified.isoformat() if modified else None

    return anchors, extraction_date


def build_fixture(project: str, workbook_path: Path) -> dict:
    anchors, extraction_date = extract_summary_anchors(workbook_path)
    workbook_sha256 = sha256_of(workbook_path)

    return {
        "project": project,
        "source_workbook": str(workbook_path.relative_to(REPO_ROOT)),
        "workbook_sha256": workbook_sha256,
        "extraction_note": (
            "Tier-1 anchors extracted from the bootstrap reference workbook's "
            "cached formula values via openpyxl only; no formula evaluation "
            "engine or Finco1 runtime code was executed by this script. The "
            "source workbook is committed to the repository (it is the "
            "deliverable, not a private artifact)."
        ),
        "extraction_date": extraction_date,
        "spec_version": SPEC_VERSION,
        "extractor_version": EXTRACTOR_VERSION,
        "worksheets": {
            SUMMARY_SHEET: {
                "role": "KPI summary / Tier-1 anchor cells",
            }
        },
        "golden_cells": dict(sorted(anchors.items())),
        "initial_tolerances": TOLERANCES,
    }


def build_manifest(project: str, workbook_path: Path, anchors: list[str], fixture_path: Path) -> dict:
    return {
        "workbook_filename": workbook_path.name,
        "workbook_path": str(workbook_path.relative_to(REPO_ROOT)),
        "workbook_sha256": sha256_of(workbook_path),
        "extractor_version": EXTRACTOR_VERSION,
        "extraction_script": "scripts/extract_generic_golden.py",
        "spec_version": SPEC_VERSION,
        "anchor_list": sorted(anchors),
        "fixture_path": str(fixture_path.relative_to(REPO_ROOT)),
    }


def run() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)

    for project, workbook_path in PROJECTS.items():
        if not workbook_path.exists():
            raise FileNotFoundError(f"Reference workbook not found: {workbook_path}")

        fixture = build_fixture(project, workbook_path)
        fixture_name = f"excel_golden_{project}.json"
        fixture_path = FIXTURE_DIR / fixture_name
        fixture_path.write_text(json.dumps(fixture, indent=2, sort_keys=False) + "\n")

        manifest = build_manifest(project, workbook_path, list(fixture["golden_cells"]), fixture_path)
        manifest_name = f"{project}_manifest.json"
        manifest_path = MANIFEST_DIR / manifest_name
        manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=False) + "\n")

        print(f"wrote {fixture_path.relative_to(REPO_ROOT)}")
        print(f"wrote {manifest_path.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    run()
