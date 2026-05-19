#!/usr/bin/env python3
"""
Phase 6 — 60-Period Model Stack Comparison Export
TUHO Wind 1: Excel vs Python (flag-on) period-by-period comparison.

Report-only diagnostic script. No production runtime changes.

Usage:
    python scripts/export_phase6_model_stack_60_period_comparison.py

Output:
    reports/phase6_model_stack_60_period_comparison.xlsx
    reports/phase6_model_stack_60_period_comparison_long.csv
    reports/phase6_model_stack_60_period_comparison_wide.csv
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Thresholds ────────────────────────────────────────────────────────────────
PASS_TOTAL_THRESHOLD = 100_000   # kEUR
PASS_PERIOD_THRESHOLD = 25_000   # kEUR per period
MINOR_TOTAL_THRESHOLD = 500_000  # kEUR

# ─── Paths ────────────────────────────────────────────────────────────────────
REPO = Path(__file__).parent.parent
FIXTURE_JSON = REPO / "tests/fixtures/excel_tuho_full_model_extract.json"
REPORTS_DIR = REPO / "reports"
OUT_XLSX = REPORTS_DIR / "phase6_model_stack_60_period_comparison.xlsx"
OUT_LONG_CSV = REPORTS_DIR / "phase6_model_stack_60_period_comparison_long.csv"
OUT_WIDE_CSV = REPORTS_DIR / "phase6_model_stack_60_period_comparison_wide.csv"

# ─── Styling constants ─────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="D6DCE4")
SUBHDR_FONT = Font(bold=True, size=10)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
MINOR_FILL = PatternFill("solid", fgColor="FFEB9C")
MATERIAL_FILL = PatternFill("solid", fgColor="FFC7CE")
BLOCKED_FILL = PatternFill("solid", fgColor="D9D9D9")
UNMAPPED_FILL = PatternFill("solid", fgColor="EDEDED")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATA_FONT = Font(size=9)
NUMBER_FMT = "#,##0.0"
PCT_FMT = "0.0%"
STATUS_FILL_MAP = {
    "PASS": PASS_FILL,
    "MINOR": MINOR_FILL,
    "MATERIAL": MATERIAL_FILL,
    "BLOCKED": BLOCKED_FILL,
    "UNMAPPED": UNMAPPED_FILL,
}

# ─── Period helpers ────────────────────────────────────────────────────────────

def period_label(i: int) -> str:
    yr = i // 2 + 1
    h = "H1" if i % 2 == 0 else "H2"
    return f"Y{yr:02d}{h}"

def period_col_header(i: int) -> str:
    return f"P{i+1:02d}"

# ─── Data loading ──────────────────────────────────────────────────────────────

def load_excel_fixture():
    with open(FIXTURE_JSON) as f:
        return json.load(f)

def run_python_model():
    sys.path.insert(0, str(REPO))
    from dataclasses import replace
    from app.project_factories import create_default_tuho_wind1
    from app.ui_runner import _build_period_engine
    from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner

    def _run(project):
        engine = _build_period_engine(project)
        config = WaterfallRunConfig.from_inputs(project, engine)
        return WaterfallRunner(project, engine).run(config)

    tuho = create_default_tuho_wind1()
    tuho = replace(
        tuho,
        info=replace(tuho.info, use_tax_bridge_engine=True),
        tax=replace(tuho.tax, cit_cash_tax_start_operating_index=25),
    )
    return _run(tuho)


def load_revenue_decomposition():
    """Load PR #90 revenue decomposition for revenue split fields."""
    sys.path.insert(0, str(REPO))
    from app.project_factories import create_default_tuho_wind1
    from app.ui_runner import _build_period_engine
    from domain.revenue.generation import revenue_decomposition_schedule

    tuho = create_default_tuho_wind1()
    engine = _build_period_engine(tuho)
    return revenue_decomposition_schedule(tuho, engine)


# ─── Row builder ───────────────────────────────────────────────────────────────

def excel_vals(excel: dict, prefix: str) -> list[float]:
    """Extract column values from Excel fixture by prefix."""
    cols = excel["period_diagnostic_columns"]
    matches = [c for c in cols if c.startswith(prefix)]
    if not matches:
        return [0.0] * 60
    idx = cols.index(matches[0])
    out = []
    for row in excel["period_diagnostics"]:
        v = row[idx] if idx < len(row) else 0.0
        out.append(float(v) if v is not None else 0.0)
    return out


def build_metric_rows(excel: dict, result, rev_decomp: dict) -> dict[str, dict]:
    """Build all metric rows for the comparison workbook.

    Returns dict: metric_key -> {label, category, excel[], python[], ...}
    """
    pdata = result.periods
    N = 60

    def py(field: str) -> list[float]:
        out = []
        for p in pdata:
            v = getattr(p, field, None)
            out.append(float(v) if v is not None else 0.0)
        return out

    def period_rev_field(field: str) -> list[float]:
        out = []
        for i in range(N):
            idx = i + 2  # operating periods start at index 2
            dec = rev_decomp.get(idx, {})
            v = dec.get(field, 0.0)
            out.append(float(v) if v is not None else 0.0)
        return out

    rows = {}

    # ── Revenue ───────────────────────────────────────────────────────────────
    # PR #90 revenue split fields
    rows["revenue_electricity"] = dict(
        label="Electricity Revenue", category="Revenue",
        excel=[0.0]*N,  # not directly in Excel fixture
        python=period_rev_field("electricity_revenue_keur"),
        excel_source="N/A", python_source="decomposition.electricity_revenue_keur",
        confidence="unmapped", note="PR #90 split field",
    )
    rows["revenue_co2"] = dict(
        label="CO2 Certificate Revenue", category="Revenue",
        excel=[0.0]*N,
        python=period_rev_field("co2_certificate_revenue_keur"),
        excel_source="N/A", python_source="decomposition.co2_certificate_revenue_keur",
        confidence="unmapped", note="PR #90 split field",
    )
    rows["revenue_balancing_cost"] = dict(
        label="Balancing Cost (positive)", category="Revenue",
        excel=[0.0]*N,
        python=period_rev_field("balancing_cost_keur"),
        excel_source="N/A", python_source="decomposition.balancing_cost_keur",
        confidence="unmapped", note="PR #90 split field",
    )
    rows["revenue_net"] = dict(
        label="Net Revenue After Balancing", category="Revenue",
        excel=excel_vals(excel, "P&L.total_revenues_keur"),
        python=py("revenue_keur"),
        excel_source="P&L!total_revenues_keur", python_source="result.periods[i].revenue_keur",
        confidence="exact", note="Total P&L revenue",
    )
    rows["revenue_net_alt"] = dict(
        label="Net Revenue (decomposition)", category="Revenue",
        excel=[0.0]*N,
        python=period_rev_field("net_revenue_after_balancing_keur"),
        excel_source="N/A", python_source="decomposition.net_revenue_after_balancing_keur",
        confidence="unmapped", note="PR #90 net from decomposition",
    )
    rows["revenue_ebitda"] = dict(
        label="EBITDA", category="Revenue",
        excel=excel_vals(excel, "CF.ebitda_keur"),
        python=py("ebitda_keur"),
        excel_source="CF!ebitda_keur", python_source="result.periods[i].ebitda_keur",
        confidence="exact", note="",
    )
    rows["revenue_generation"] = dict(
        label="Generation (MWh)", category="Revenue",
        excel=[0.0]*N, python=py("generation_mwh"),
        excel_source="N/A", python_source="result.periods[i].generation_mwh",
        confidence="unmapped", note="Generation from runtime",
    )

    # ── OPEX ─────────────────────────────────────────────────────────────────
    rows["opex_total"] = dict(
        label="Total OPEX", category="OPEX",
        excel=[abs(v) for v in excel_vals(excel, "CF.operating_expenses_after_bank_tax_keur")],
        python=py("opex_keur"),
        excel_source="CF!operating_expenses_after_bank_tax_keur",
        python_source="result.periods[i].opex_keur",
        confidence="exact", note="",
    )

    # ── EBITDA ───────────────────────────────────────────────────────────────
    rows["ebitda"] = dict(
        label="EBITDA", category="EBITDA",
        excel=excel_vals(excel, "CF.ebitda_keur"),
        python=py("ebitda_keur"),
        excel_source="CF!ebitda_keur", python_source="result.periods[i].ebitda_keur",
        confidence="exact", note="",
    )

    # ── Free Cash Flow ────────────────────────────────────────────────────────
    rows["fcf_banks"] = dict(
        label="FCF for Banks", category="Free Cash Flow",
        excel=excel_vals(excel, "CF.free_cash_flow_for_banks_keur"),
        python=py("cf_after_tax_keur"),
        excel_source="CF!free_cash_flow_for_banks_keur",
        python_source="result.periods[i].cf_after_tax_keur",
        confidence="approximate", note="Python uses cf_after_tax_keur as proxy",
    )
    rows["fcf_dist"] = dict(
        label="FCF for Distribution", category="Free Cash Flow",
        excel=excel_vals(excel, "CF.free_cash_flow_for_distribution_keur"),
        python=py("r98_distribution_account_keur"),
        excel_source="CF!free_cash_flow_for_distribution_keur",
        python_source="result.periods[i].r98_distribution_account_keur",
        confidence="approximate", note="",
    )
    rows["fcf_shl"] = dict(
        label="FCF for SHL [BLOCKED]", category="Free Cash Flow",
        excel=[0.0]*N,
        python=py("r102_fcf_for_shl_keur"),
        excel_source="N/A", python_source="result.periods[i].r102_fcf_for_shl_keur",
        confidence="unmapped", note="R102 is BLOCKED — audit-only, not runtime driver",
    )

    # ── Depreciation ─────────────────────────────────────────────────────────
    rows["dep_book"] = dict(
        label="Book Depreciation", category="Depreciation",
        excel=excel_vals(excel, "Dep.depreciation_keur"),
        python=py("depreciation_keur"),
        excel_source="Dep!depreciation_keur",
        python_source="result.periods[i].depreciation_keur",
        confidence="exact", note="",
    )
    rows["dep_unlevered"] = dict(
        label="Unlevered Depreciation", category="Depreciation",
        excel=excel_vals(excel, "Dep.unlevered_depreciation_keur"),
        python=[0.0]*N,
        excel_source="Dep!unlevered_depreciation_keur",
        python_source="N/A",
        confidence="unmapped", note="Python does not compute unlevered depreciation separately",
    )
    rows["dep_tax"] = dict(
        label="Tax Depreciation (audit)", category="Depreciation",
        excel=[0.0]*N,
        python=py("tax_depreciation_audit_keur"),
        excel_source="N/A", python_source="result.periods[i].tax_depreciation_audit_keur",
        confidence="unmapped", note="Tax depreciation from offline engine",
    )

    # ── Senior Debt ──────────────────────────────────────────────────────────
    rows["sd_interest"] = dict(
        label="Senior Interest", category="Senior Debt",
        excel=excel_vals(excel, "DS.senior_net_interest_keur"),
        python=py("interest_senior_keur"),
        excel_source="DS!senior_net_interest_keur",
        python_source="result.periods[i].interest_senior_keur",
        confidence="exact", note="",
    )
    rows["sd_principal"] = dict(
        label="Senior Principal Repayment", category="Senior Debt",
        excel=excel_vals(excel, "DS.senior_principal_keur"),
        python=py("senior_principal_keur"),
        excel_source="DS!senior_principal_keur",
        python_source="result.periods[i].senior_principal_keur",
        confidence="exact", note="",
    )
    rows["sd_balance"] = dict(
        label="Senior Closing Balance", category="Senior Debt",
        excel=[0.0]*N, python=py("senior_balance_keur"),
        excel_source="N/A", python_source="result.periods[i].senior_balance_keur",
        confidence="unmapped", note="Closing balance from waterfall",
    )
    rows["sd_ds"] = dict(
        label="Total Senior Debt Service", category="Senior Debt",
        excel=[0.0]*N, python=py("senior_ds_keur"),
        excel_source="N/A", python_source="result.periods[i].senior_ds_keur",
        confidence="unmapped", note="",
    )
    rows["dscr"] = dict(
        label="DSCR", category="Senior Debt",
        excel=[0.0]*N, python=py("dscr"),
        excel_source="N/A", python_source="result.periods[i].dscr",
        confidence="unmapped", note="",
    )

    # ── SHL ──────────────────────────────────────────────────────────────────
    shl_data = excel.get("shl", [])
    rows["shl_gross_accrued"] = dict(
        label="SHL Gross Accrued Interest [CANDIDATE]", category="SHL",
        excel=[row[3] if len(row) > 3 else 0.0 for row in shl_data[:N]],
        python=py("shl_gross_accrued_interest_keur"),
        excel_source="SHL schedule!gross_interest",
        python_source="result.periods[i].shl_gross_accrued_interest_keur",
        confidence="approximate", note="SHL gross-accrued is candidate driver; treatment may differ",
    )
    rows["shl_balance"] = dict(
        label="SHL Closing Balance", category="SHL",
        excel=[row[2] if len(row) > 2 else 0.0 for row in shl_data[:N]],
        python=py("shl_balance_keur"),
        excel_source="SHL schedule!closing",
        python_source="result.periods[i].shl_balance_keur",
        confidence="approximate", note="",
    )
    rows["shl_pik"] = dict(
        label="SHL PIK / Capitalised Interest", category="SHL",
        excel=[row[5] if len(row) > 5 else 0.0 for row in shl_data[:N]],
        python=py("shl_pik_keur"),
        excel_source="SHL schedule!capitalized_interest",
        python_source="result.periods[i].shl_pik_keur",
        confidence="approximate", note="",
    )
    rows["shl_interest"] = dict(
        label="SHL Paid Interest", category="SHL",
        excel=[0.0]*N, python=py("shl_interest_keur"),
        excel_source="N/A", python_source="result.periods[i].shl_interest_keur",
        confidence="unmapped", note="",
    )

    # ── CIT Tax ──────────────────────────────────────────────────────────────
    rows["r35"] = dict(
        label="R35 Taxable Income (flag-on)", category="CIT Tax",
        excel=excel_vals(excel, "P&L.taxable_income_keur"),
        python=py("taxable_income_before_losses_audit_keur"),
        excel_source="P&L!taxable_income_keur",
        python_source="result.periods[i].taxable_income_before_losses_audit_keur",
        confidence="exact", note="Python R35 uses flag-on tax bridge",
    )
    rows["r41"] = dict(
        label="R41 Taxable Profit After Losses", category="CIT Tax",
        excel=excel_vals(excel, "P&L.taxable_income_keur"),
        python=py("taxable_profit_after_losses_audit_keur"),
        excel_source="P&L!taxable_income_keur (proxy)",
        python_source="result.periods[i].taxable_profit_after_losses_audit_keur",
        confidence="approximate", note="Excel does not separately surface R41",
    )
    rows["cit_accrual"] = dict(
        label="R43 CIT Accrual", category="CIT Tax",
        excel=excel_vals(excel, "P&L.corporate_income_tax_keur"),
        python=py("tax_keur"),
        excel_source="P&L!corporate_income_tax_keur",
        python_source="result.periods[i].tax_keur",
        confidence="approximate", note="Python uses flag-on tax bridge accrual",
    )
    rows["r67"] = dict(
        label="R67 Cash Tax [DIAGNOSTIC]", category="CIT Tax",
        excel=[0.0]*N,
        python=py("r67_excel_style_cash_tax_diagnostic_keur"),
        excel_source="N/A", python_source="result.periods[i].r67_excel_style_cash_tax_diagnostic_keur",
        confidence="unmapped", note="R67 diagnostic; not in Excel fixture",
    )
    rows["cit_accrual_audit"] = dict(
        label="CIT Accrual Audit", category="CIT Tax",
        excel=[0.0]*N,
        python=py("cit_accrual_audit_keur"),
        excel_source="N/A", python_source="result.periods[i].cit_accrual_audit_keur",
        confidence="unmapped", note="",
    )
    rows["tax_loss_opening"] = dict(
        label="Tax Loss Opening (audit)", category="CIT Tax",
        excel=[0.0]*N,
        python=py("tax_loss_opening_audit_keur"),
        excel_source="N/A", python_source="result.periods[i].tax_loss_opening_audit_keur",
        confidence="unmapped", note="",
    )
    rows["tax_loss_closing"] = dict(
        label="Tax Loss Closing (audit)", category="CIT Tax",
        excel=[0.0]*N,
        python=py("tax_loss_closing_audit_keur"),
        excel_source="N/A", python_source="result.periods[i].tax_loss_closing_audit_keur",
        confidence="unmapped", note="",
    )
    rows["fiscal_reintegration"] = dict(
        label="Fiscal Reintegration (audit)", category="CIT Tax",
        excel=[0.0]*N,
        python=py("fiscal_reintegration_audit_keur"),
        excel_source="N/A", python_source="result.periods[i].fiscal_reintegration_audit_keur",
        confidence="unmapped", note="",
    )

    return rows


def compute_delta_flags(rows: dict) -> list[dict]:
    """Compute sorted material deltas."""
    flags = []
    for key, row in rows.items():
        evals = row["excel"]
        pvals = row["python"]
        for i in range(60):
            ev = evals[i] if i < len(evals) else 0.0
            pv = pvals[i] if i < len(pvals) else 0.0
            d = pv - ev
            if abs(d) > 50:  # material threshold for flagging
                pct = d / abs(ev) if ev != 0 else 0.0
                flags.append({
                    "key": key, "label": row["label"], "category": row["category"],
                    "period_idx": i, "period_label": period_label(i),
                    "excel": ev, "python": pv, "delta": d, "delta_pct": pct,
                    "excel_source": row["excel_source"], "python_source": row["python_source"],
                    "confidence": row["confidence"], "note": row["note"],
                    "status": _status(row["confidence"], d, ev, i),
                })
    flags.sort(key=lambda x: -abs(x["delta"]))
    return flags


def _status(confidence: str, delta: float, excel_val: float, period_idx: int) -> str:
    if confidence == "unmapped":
        return "UNMAPPED"
    if confidence == "approximate":
        return "APPROXIMATE"
    if abs(delta) <= 25 and abs(delta / abs(excel_val)) < 0.01 if excel_val else False:
        return "PASS"
    if abs(delta) <= 500:
        return "MINOR"
    return "MATERIAL"


def _aggregate_status(total_delta: float, max_period_delta: float, confidence: str) -> str:
    if confidence in ("unmapped",):
        return "UNMAPPED"
    if confidence in ("approximate",):
        return "APPROXIMATE"
    if abs(total_delta) <= PASS_TOTAL_THRESHOLD and abs(max_period_delta) <= PASS_PERIOD_THRESHOLD:
        return "PASS"
    if abs(total_delta) <= MINOR_TOTAL_THRESHOLD:
        return "MINOR"
    return "MATERIAL"


# ─── CSV writers ───────────────────────────────────────────────────────────────

def write_long_csv(rows: dict, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("category,metric,row_type,period_idx,period_label,"
                "excel_value,python_value,delta,delta_pct,excel_source,python_source,confidence,status,notes\n")
        for key, row in rows.items():
            for i in range(60):
                ev = row["excel"][i] if i < len(row["excel"]) else 0.0
                pv = row["python"][i] if i < len(row["python"]) else 0.0
                d = pv - ev
                pct = d / abs(ev) if ev != 0 else 0.0
                status = _status(row["confidence"], d, ev, i)
                f.write(f"{row['category']},{row['label']},detail,{i},{period_label(i)},"
                        f"{ev:.1f},{pv:.1f},{d:.1f},{pct:.2%},"
                        f"{row['excel_source']},{row['python_source']},"
                        f"{row['confidence']},{status},{row['note']}\n")
            # totals row
            ev_tot = sum(row['excel'])
            pv_tot = sum(row['python'])
            d_tot = pv_tot - ev_tot
            pct_tot = d_tot / abs(ev_tot) if ev_tot else 0.0
            f.write(f"{row['category']},{row['label']},total,,-1,,"
                    f"{ev_tot:.1f},{pv_tot:.1f},{d_tot:.1f},{pct_tot:.2%},"
                    f"{row['excel_source']},{row['python_source']},"
                    f"{row['confidence']},{_aggregate_status(d_tot, 0, row['confidence'])},"
                    f"{row['note']}\n")
    print(f"  {path} ({path.stat().st_size:,} bytes)")


def write_wide_csv(rows: dict, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        # header
        hdr = ["category", "metric", "row_type", "source"] + [f"P{i+1:02d}" for i in range(60)] + ["total", "status", "notes"]
        f.write(",".join(hdr) + "\n")
        for key, row in rows.items():
            for row_type in ["excel", "python", "delta", "delta_pct"]:
                vals = []
                for i in range(60):
                    ev = row["excel"][i] if i < len(row["excel"]) else 0.0
                    pv = row["python"][i] if i < len(row["python"]) else 0.0
                    if row_type == "excel":
                        vals.append(f"{ev:.1f}")
                    elif row_type == "python":
                        vals.append(f"{pv:.1f}")
                    elif row_type == "delta":
                        vals.append(f"{pv-ev:.1f}")
                    elif row_type == "delta_pct":
                        d = pv - ev
                        pct = d / abs(ev) if ev != 0 else 0.0
                        vals.append(f"{pct:.2%}")
                total_val = sum(row["excel"]) if row_type == "excel" else sum(row["python"]) if row_type == "python" else sum(row["python"]) - sum(row["excel"])
                vals.append(f"{total_val:.1f}")
                status = _aggregate_status(total_val, 0, row["confidence"]) if row_type in ("excel", "python") else ""
                note = row["note"] if row_type == "excel" else ""
                f.write(f"{row['category']},{row['label']},{row_type},{row['excel_source'] if row_type == 'excel' else row['python_source']}," + ",".join(vals) + f",{status},{note}\n")
    print(f"  {path} ({path.stat().st_size:,} bytes)")


# ─── XLSX writer ──────────────────────────────────────────────────────────────

def write_xlsx(rows: dict, flags: list, path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Helper ────────────────────────────────────────────────────────────────
    def style_header(cell, is_subhdr=False):
        cell.font = HDR_FONT if not is_subhdr else SUBHDR_FONT
        cell.fill = HDR_FILL if not is_subhdr else SUBHDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    def style_data(cell, value=None, fmt=NUMBER_FMT, center=False):
        cell.font = DATA_FONT
        cell.number_format = fmt
        cell.border = BORDER
        if center:
            cell.alignment = Alignment(horizontal="center")

    def write_sheet_header(ws, cols: list[str]):
        for j, h in enumerate(cols, 1):
            c = ws.cell(1, j, h)
            style_header(c)
        ws.row_dimensions[1].height = 30

    def period_headers(ws, N=60):
        """Write period column headers."""
        ws.cell(1, 1, "Metric")
        ws.cell(1, 2, "Category")
        for i in range(N):
            c = ws.cell(1, 3 + i, f"P{i+1:02d}\n{period_label(i)}")
            style_header(c)
        ws.cell(1, 3 + N, "Total")
        style_header(ws.cell(1, 3 + N))
        ws.cell(1, 3 + N + 1, "Status")
        ws.cell(1, 3 + N + 2, "Notes")
        ws.freeze_panes = "C2"

    def write_metric_block(ws, row_start: int, metric_key: str, row_data: dict, N=60):
        label = row_data["label"]
        category = row_data["category"]
        excel_vals = row_data["excel"]
        py_vals = row_data["python"]
        conf = row_data["confidence"]
        note = row_data["note"]

        excel_tot = sum(excel_vals)
        py_tot = sum(py_vals)
        delta_tot = py_tot - excel_tot
        delta_pct_tot = delta_tot / abs(excel_tot) if excel_tot else 0.0
        max_abs_delta = max(abs(py_vals[i] - excel_vals[i]) for i in range(N))
        status = _aggregate_status(delta_tot, max_abs_delta, conf)

        # Excel row
        r = row_start
        ws.cell(r, 1, label).font = Font(bold=True, size=9)
        ws.cell(r, 2, category)
        ws.cell(r, 3 + N, f"{excel_tot:.1f}").font = DATA_FONT
        ws.cell(r, 3 + N + 1, "").fill = STATUS_FILL_MAP.get(status, PASS_FILL)
        ws.cell(r, 3 + N + 2, note)
        for i, v in enumerate(excel_vals):
            c = ws.cell(r, 3 + i, f"{v:.1f}")
            style_data(c)
        fill = STATUS_FILL_MAP.get(status, PASS_FILL)
        ws.cell(r, 3 + N, f"{excel_tot:.1f}").fill = fill
        ws.row_dimensions[r].height = 15

        # Python row
        r = row_start + 1
        ws.cell(r, 1, f"  Python").font = DATA_FONT
        ws.cell(r, 3 + N, f"{py_tot:.1f}").font = DATA_FONT
        ws.cell(r, 3 + N + 1, "").fill = fill
        ws.cell(r, 3 + N + 2, conf)
        for i, v in enumerate(py_vals):
            c = ws.cell(r, 3 + i, f"{v:.1f}")
            style_data(c)
        ws.cell(r, 3 + N, f"{py_tot:.1f}").fill = fill
        ws.row_dimensions[r].height = 15

        # Delta row
        r = row_start + 2
        ws.cell(r, 1, f"  Delta").font = DATA_FONT
        delta_tot_v = py_tot - excel_tot
        for i in range(N):
            d = py_vals[i] - excel_vals[i]
            c = ws.cell(r, 3 + i, f"{d:.1f}")
            style_data(c)
        dtot = py_tot - excel_tot
        c = ws.cell(r, 3 + N, f"{dtot:.1f}")
        style_data(c)
        c.fill = fill
        ws.row_dimensions[r].height = 15

        # Delta % row
        r = row_start + 3
        ws.cell(r, 1, f"  Delta %").font = DATA_FONT
        for i in range(N):
            ev = excel_vals[i]
            pv = py_vals[i]
            pct = (pv - ev) / abs(ev) if ev != 0 else 0.0
            c = ws.cell(r, 3 + i, f"{pct:.1%}")
            style_data(c, fmt="0.0%")
        dpct = delta_pct_tot
        c = ws.cell(r, 3 + N, f"{dpct:.1%}")
        style_data(c, fmt="0.0%")
        c.fill = fill
        ws.row_dimensions[r].height = 15

        return row_start + 4

    # ── Sheet 1: Revenue ───────────────────────────────────────────────────────
    ws_rev = wb.create_sheet("Revenue")
    revenue_metrics = ["revenue_electricity", "revenue_co2", "revenue_balancing_cost",
                       "revenue_net", "revenue_net_alt", "revenue_ebitda", "revenue_generation"]
    period_headers(ws_rev)
    r = 2
    for m in revenue_metrics:
        r = write_metric_block(ws_rev, r, m, rows[m])

    ws_rev.column_dimensions["A"].width = 28
    ws_rev.column_dimensions["B"].width = 12
    for i in range(60):
        ws_rev.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_rev.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_rev.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_rev.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45

    # ── Sheet 2: OPEX ──────────────────────────────────────────────────────────
    ws_opex = wb.create_sheet("OPEX")
    period_headers(ws_opex)
    r = write_metric_block(ws_opex, 2, "opex_total", rows["opex_total"])
    ws_opex.column_dimensions["A"].width = 28
    ws_opex.column_dimensions["B"].width = 12
    for i in range(60):
        ws_opex.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_opex.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_opex.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_opex.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_opex.freeze_panes = "C2"

    # ── Sheet 3: EBITDA ────────────────────────────────────────────────────────
    ws_ebitda = wb.create_sheet("EBITDA")
    period_headers(ws_ebitda)
    r = write_metric_block(ws_ebitda, 2, "ebitda", rows["ebitda"])
    ws_ebitda.column_dimensions["A"].width = 28
    ws_ebitda.column_dimensions["B"].width = 12
    for i in range(60):
        ws_ebitda.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_ebitda.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_ebitda.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_ebitda.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_ebitda.freeze_panes = "C2"

    # ── Sheet 4: Free Cash Flow ────────────────────────────────────────────────
    ws_fcf = wb.create_sheet("Free Cash Flow")
    fcf_metrics = ["fcf_banks", "fcf_dist", "fcf_shl"]
    period_headers(ws_fcf)
    r = 2
    for m in fcf_metrics:
        r = write_metric_block(ws_fcf, r, m, rows[m])
    ws_fcf.column_dimensions["A"].width = 28
    ws_fcf.column_dimensions["B"].width = 12
    for i in range(60):
        ws_fcf.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_fcf.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_fcf.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_fcf.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_fcf.freeze_panes = "C2"

    # ── Sheet 5: Depreciation ───────────────────────────────────────────────────
    ws_dep = wb.create_sheet("Depreciation")
    dep_metrics = ["dep_book", "dep_unlevered", "dep_tax"]
    period_headers(ws_dep)
    r = 2
    for m in dep_metrics:
        r = write_metric_block(ws_dep, r, m, rows[m])
    ws_dep.column_dimensions["A"].width = 28
    ws_dep.column_dimensions["B"].width = 12
    for i in range(60):
        ws_dep.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_dep.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_dep.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_dep.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_dep.freeze_panes = "C2"

    # ── Sheet 6: Senior Debt ────────────────────────────────────────────────────
    ws_sd = wb.create_sheet("Senior Debt")
    sd_metrics = ["sd_interest", "sd_principal", "sd_balance", "sd_ds", "dscr"]
    period_headers(ws_sd)
    r = 2
    for m in sd_metrics:
        r = write_metric_block(ws_sd, r, m, rows[m])
    ws_sd.column_dimensions["A"].width = 28
    ws_sd.column_dimensions["B"].width = 12
    for i in range(60):
        ws_sd.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_sd.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_sd.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_sd.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_sd.freeze_panes = "C2"

    # ── Sheet 7: SHL ───────────────────────────────────────────────────────────
    ws_shl = wb.create_sheet("SHL")
    shl_metrics = ["shl_gross_accrued", "shl_balance", "shl_pik", "shl_interest"]
    period_headers(ws_shl)
    r = 2
    for m in shl_metrics:
        r = write_metric_block(ws_shl, r, m, rows[m])
    ws_shl.column_dimensions["A"].width = 28
    ws_shl.column_dimensions["B"].width = 12
    for i in range(60):
        ws_shl.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_shl.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_shl.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_shl.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_shl.freeze_panes = "C2"

    # ── Sheet 8: CIT Tax ──────────────────────────────────────────────────────
    ws_cit = wb.create_sheet("CIT Tax")
    cit_metrics = ["r35", "r41", "cit_accrual", "r67", "cit_accrual_audit",
                   "tax_loss_opening", "tax_loss_closing", "fiscal_reintegration"]
    period_headers(ws_cit)
    r = 2
    for m in cit_metrics:
        r = write_metric_block(ws_cit, r, m, rows[m])
    ws_cit.column_dimensions["A"].width = 28
    ws_cit.column_dimensions["B"].width = 12
    for i in range(60):
        ws_cit.column_dimensions[get_column_letter(3 + i)].width = 9
    ws_cit.column_dimensions[get_column_letter(3 + 60)].width = 12
    ws_cit.column_dimensions[get_column_letter(3 + 60 + 1)].width = 10
    ws_cit.column_dimensions[get_column_letter(3 + 60 + 2)].width = 45
    ws_cit.freeze_panes = "C2"

    # ── Sheet 9: Summary ───────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    sum_hdrs = ["Category", "Metric", "Total Excel (kEUR)", "Total Python (kEUR)",
                "Total Delta (kEUR)", "Delta %", "Max Period Delta (kEUR)",
                "Max Period", "Status", "Notes"]
    for j, h in enumerate(sum_hdrs, 1):
        c = ws_sum.cell(1, j, h)
        style_header(c)
    ws_sum.row_dimensions[1].height = 30

    r = 2
    for key, row in rows.items():
        evals = row["excel"]
        pvals = row["python"]
        total_excel = sum(evals)
        total_py = sum(pvals)
        total_delta = total_py - total_excel
        total_pct = total_delta / abs(total_excel) if total_excel else 0.0
        max_delta = 0.0
        max_period = 0
        for i in range(60):
            d = abs(pvals[i] - evals[i])
            if d > max_delta:
                max_delta = d
                max_period = i
        status = _aggregate_status(total_delta, max_delta, row["confidence"])

        fill = STATUS_FILL_MAP.get(status, PASS_FILL)
        for j, v in enumerate([row["category"], row["label"],
                               f"{total_excel:.1f}", f"{total_py:.1f}",
                               f"{total_delta:.1f}", f"{total_pct:.1%}",
                               f"{max_delta:.1f}", period_label(max_period),
                               status, row["note"]], 1):
            c = ws_sum.cell(r, j, v)
            c.font = DATA_FONT
            c.border = BORDER
            if j >= 3:
                c.number_format = NUMBER_FMT if j != 6 else PCT_FMT
            if j == 9:
                c.fill = fill
        r += 1

    ws_sum.column_dimensions["A"].width = 14
    ws_sum.column_dimensions["B"].width = 32
    ws_sum.column_dimensions["C"].width = 18
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 18
    ws_sum.column_dimensions["F"].width = 10
    ws_sum.column_dimensions["G"].width = 22
    ws_sum.column_dimensions["H"].width = 10
    ws_sum.column_dimensions["I"].width = 12
    ws_sum.column_dimensions["J"].width = 50
    ws_sum.freeze_panes = "A2"

    # ── Sheet 10: Delta Flags ──────────────────────────────────────────────────
    ws_flags = wb.create_sheet("Delta Flags")
    flag_hdrs = ["Category", "Metric", "Period", "Period Label",
                 "Excel Value", "Python Value", "Delta (kEUR)", "Delta %",
                 "Likely Cause", "Status", "Recommended Action", "Excel Source", "Python Source"]
    for j, h in enumerate(flag_hdrs, 1):
        c = ws_flags.cell(1, j, h)
        style_header(c)
    ws_flags.row_dimensions[1].height = 30

    r = 2
    for flag in flags[:80]:  # top 80 deltas
        cause = "Minor timing or rounding" if abs(flag["delta"]) <= 500 else "Construction-period mismatch or source basis difference"
        action = "Monitor" if abs(flag["delta"]) <= 500 else "Investigate"
        fill = STATUS_FILL_MAP.get(flag["status"], PASS_FILL)
        for j, v in enumerate([flag["category"], flag["label"],
                               str(flag["period_idx"]), flag["period_label"],
                               f"{flag['excel']:.1f}", f"{flag['python']:.1f}",
                               f"{flag['delta']:.1f}", f"{flag['delta_pct']:.1%}",
                               cause, flag["status"], action,
                               flag["excel_source"], flag["python_source"]], 1):
            c = ws_flags.cell(r, j, v)
            c.font = DATA_FONT
            c.border = BORDER
            if j >= 5 and j != 8:
                c.number_format = NUMBER_FMT if j != 8 else PCT_FMT
            if j == 10:
                c.fill = fill
        r += 1

    ws_flags.column_dimensions["A"].width = 14
    ws_flags.column_dimensions["B"].width = 32
    ws_flags.column_dimensions["C"].width = 8
    ws_flags.column_dimensions["D"].width = 10
    for col in ["E", "F", "G", "H"]:
        ws_flags.column_dimensions[col].width = 14
    ws_flags.column_dimensions["I"].width = 45
    ws_flags.column_dimensions["J"].width = 14
    ws_flags.column_dimensions["K"].width = 22
    ws_flags.column_dimensions["L"].width = 40
    ws_flags.column_dimensions["M"].width = 45
    ws_flags.freeze_panes = "A2"

    # ── Sheet 11: Source Mapping ───────────────────────────────────────────────
    ws_src = wb.create_sheet("Source Mapping")
    src_hdrs = ["Category", "Metric", "Excel Sheet", "Excel Row/Cell",
                "Python Object", "Python Field", "Confidence", "Notes"]
    for j, h in enumerate(src_hdrs, 1):
        c = ws_src.cell(1, j, h)
        style_header(c)
    ws_src.row_dimensions[1].height = 30

    r = 2
    for key, row in rows.items():
        parts = row["excel_source"].split("!") if "!" in row["excel_source"] else [row["excel_source"], "N/A"]
        for j, v in enumerate([row["category"], row["label"],
                               parts[0] if len(parts) > 0 else "N/A",
                               parts[1] if len(parts) > 1 else row["excel_source"],
                               "result.periods[i]", row["python_source"],
                               row["confidence"], row["note"]], 1):
            c = ws_src.cell(r, j, v)
            c.font = DATA_FONT
            c.border = BORDER
            if row["confidence"] == "unmapped":
                c.fill = UNMAPPED_FILL
        r += 1

    ws_src.column_dimensions["A"].width = 14
    ws_src.column_dimensions["B"].width = 32
    ws_src.column_dimensions["C"].width = 14
    ws_src.column_dimensions["D"].width = 22
    ws_src.column_dimensions["E"].width = 20
    ws_src.column_dimensions["F"].width = 45
    ws_src.column_dimensions["G"].width = 14
    ws_src.column_dimensions["H"].width = 55
    ws_src.freeze_panes = "A2"

    wb.save(path)
    print(f"  {path} ({path.stat().st_size:,} bytes)")


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading Excel fixture...")
    excel = load_excel_fixture()

    print("Running Python TUHO flag-on model...")
    result = run_python_model()

    print("Loading PR #90 revenue decomposition...")
    rev_decomp = load_revenue_decomposition()

    print("Building comparison rows...")
    rows = build_metric_rows(excel, result, rev_decomp)

    print("Computing delta flags...")
    flags = compute_delta_flags(rows)

    print("Writing long CSV...")
    write_long_csv(rows, OUT_LONG_CSV)

    print("Writing wide CSV...")
    write_wide_csv(rows, OUT_WIDE_CSV)

    print("Writing XLSX...")
    write_xlsx(rows, flags, OUT_XLSX)

    # Summary stats
    mapped = sum(1 for r in rows.values() if r["confidence"] in ("exact", "approximate"))
    unmapped = sum(1 for r in rows.values() if r["confidence"] == "unmapped")
    print(f"\n  Mapped metrics: {mapped}")
    print(f"  Unmapped metrics: {unmapped}")
    print(f"  Delta flags (material): {len(flags)}")
    print("\nAll outputs written.")

if __name__ == "__main__":
    main()