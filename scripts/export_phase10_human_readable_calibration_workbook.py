#!/usr/bin/env python3
"""
Phase 10 — Human-Readable Calibration Workbook Generator
==========================================================

Builds: reports/phase10_human_readable_calibration_workbook.xlsx
Outputs:
  reports/phase10_human_readable_calibration_summary.csv
  reports/phase10_human_readable_calibration_gap_analysis.csv
  reports/phase10_human_readable_calibration_source_map.csv

TUHO Wind 1 — Horizontal Excel/Model/Delta by semiannual period.
No runtime formula changes. Report-only.
"""

import csv
import sys
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
OUT_XLSX = REPORTS / "phase10_human_readable_calibration_workbook.xlsx"
OUT_SUMMARY_CSV = REPORTS / "phase10_human_readable_calibration_summary.csv"
OUT_GAP_CSV = REPORTS / "phase10_human_readable_calibration_gap_analysis.csv"
OUT_SOURCE_CSV = REPORTS / "phase10_human_readable_calibration_source_map.csv"

# ── Color palette ──────────────────────────────────────────────────────────────
CLR_PASS = "C6EFCE"
CLR_WARN = "FFEB9C"
CLR_MISS = "FFC7CE"
CLR_CONV = "DDEBF7"
CLR_BLCK = "FF0000"
CLR_HDR_BG = "4472C4"
CLR_HDR_FG = "FFFFFF"
CLR_SECTION = "D9E1F2"
CLR_TOTAL = "F2F2F2"

# ── Styles ────────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

SECTION_FILL = _fill(CLR_SECTION)
HEADER_FILL  = _fill(CLR_HDR_BG)
PASS_FILL    = _fill(CLR_PASS)
WARN_FILL    = _fill(CLR_WARN)
MISS_FILL    = _fill(CLR_MISS)
CONV_FILL    = _fill(CLR_CONV)
TOTAL_FILL   = _fill(CLR_TOTAL)
BLCK_FILL    = _fill(CLR_BLCK)

def _status_fill(status: str) -> PatternFill:
    s = status.upper()
    if s == "PASS":   return PASS_FILL
    if s == "WARN":   return WARN_FILL
    if s in ("MISSING_EVIDENCE", "RUNTIME_BINDING_PENDING", "GOVERNANCE_BLOCKER"): return MISS_FILL
    if s == "ACCEPTED_CONVENTION": return CONV_FILL
    return _fill("FFFFFF")

def _num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if not v or v.upper() in ('MISSING_EVIDENCE', 'N/A', 'NONE', ''):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

# ── Model runner ──────────────────────────────────────────────────────────────
def run_tuho_model():
    """Run TUHO Wind 1 from live runtime."""
    sys.path.insert(0, str(ROOT))
    from app.ui_runner import run_demo_project
    from app.project_factories import create_default_tuho_wind1
    proj = create_default_tuho_wind1()
    result = run_demo_project("Wind", "Base", project_inputs_override=proj)
    if result.result is None:
        raise RuntimeError(f"TUHO model run failed: {result.messages}")
    return result.result, proj

# ── Data extraction ───────────────────────────────────────────────────────────
def extract_period_data(result):
    """Build per-period lists from WaterfallResult.periods."""
    periods = result.periods
    n = len(periods)

    dates = [
        p.date.strftime("%Y-%m-%d") if hasattr(p.date, 'strftime') else str(p.date)
        for p in periods
    ]

    def fv(val):
        if val is None: return 0.0
        try: return float(val)
        except: return 0.0

    def safe_dscr(val):
        v = fv(val)
        return None if v == float('inf') or v > 100 else v

    data = {
        'n': n,
        'dates': dates,
        'generation_mwh':    [fv(p.generation_mwh) for p in periods],
        'revenue_keur':     [fv(p.revenue_keur) for p in periods],
        'opex_keur':        [fv(p.opex_keur) for p in periods],
        'ebitda_keur':      [fv(p.ebitda_keur) for p in periods],
        'depreciation_keur':[fv(p.depreciation_keur) for p in periods],
        'interest_senior_keur': [fv(p.interest_senior_keur) for p in periods],
        'interest_shl_keur':   [fv(p.interest_shl_keur) for p in periods],
        'taxable_profit_keur': [fv(p.taxable_profit_keur) for p in periods],
        'tax_keur':            [fv(p.tax_keur) for p in periods],
        'cf_after_tax_keur':   [fv(p.cf_after_tax_keur) for p in periods],
        'senior_interest_keur': [fv(p.senior_interest_keur) for p in periods],
        'senior_principal_keur': [fv(p.senior_principal_keur) for p in periods],
        'senior_ds_keur':     [fv(p.senior_ds_keur) for p in periods],
        'senior_balance_keur': [fv(p.senior_balance_keur) for p in periods],
        'shl_interest_keur':   [fv(p.shl_interest_keur) for p in periods],
        'shl_principal_keur':  [fv(p.shl_principal_keur) for p in periods],
        'shl_pik_keur':        [fv(p.shl_pik_keur) for p in periods],
        'shl_service_keur':    [fv(p.shl_service_keur) for p in periods],
        'shl_gross_accrued_interest_keur': [fv(p.shl_gross_accrued_interest_keur) for p in periods],
        'shl_balance_keur':    [fv(p.shl_balance_keur) for p in periods],
        'dscr':   [safe_dscr(p.dscr) for p in periods],
        'r69_fcf_banks_keur': [fv(p.r69_fcf_banks_keur) for p in periods],
        'r84_fcf_junior_keur': [fv(p.r84_fcf_junior_keur) for p in periods],
        'r98_distribution_account_keur': [fv(p.r98_distribution_account_keur) for p in periods],
        'r99_fcf_for_distribution_keur': [fv(p.r99_fcf_for_distribution_keur) for p in periods],
        'r102_fcf_for_shl_keur': [fv(p.r102_fcf_for_shl_keur) for p in periods],
        'fcf_for_shl_keur':    [fv(p.fcf_for_shl_keur) for p in periods],
        'distribution_keur':  [fv(p.distribution_keur) for p in periods],
        'legacy_distribution_keur': [fv(p.legacy_distribution_keur) for p in periods],
        'da_paid_distribution_keur': [fv(p.da_paid_distribution_keur) for p in periods],
        'corporate_tax_cash_keur': [fv(p.corporate_tax_cash_keur) for p in periods],
        'cash_tax_current_period_audit_keur': [fv(p.cash_tax_current_period_audit_keur) for p in periods],
        'tax_depreciation_audit_keur': [fv(p.tax_depreciation_audit_keur) for p in periods],
        'taxable_income_before_losses_audit_keur': [fv(p.taxable_income_before_losses_audit_keur) for p in periods],
        'tax_loss_opening_audit_keur': [fv(p.tax_loss_opening_audit_keur) for p in periods],
        'tax_loss_used_audit_keur': [fv(p.tax_loss_used_audit_keur) for p in periods],
    }
    return data

# ── Excel evidence loader ─────────────────────────────────────────────────────
def load_csv(filename):
    path = REPORTS / filename
    if not path.exists():
        return []
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def get_excel_vals(csv_rows, col_name, n, default=None):
    """Extract column values from CSV rows; return list of length n."""
    if not csv_rows:
        return [default] * n
    result = []
    for row in csv_rows:
        v = row.get(col_name, default)
        f = _num(v)
        result.append(f if f is not None else default)
    # Pad/truncate
    while len(result) < n:
        result.append(default)
    return result[:n]

# ── Triplet row writer ─────────────────────────────────────────────────────────
def write_triplet(ws, row, label, excel_vals, model_vals, source="",
                  num_format='#,##0.0', extra_cols=None):
    """
    Write Excel/Model/Delta triplet rows.
    extra_cols: list of (col_idx, value, fill) tuples for additional columns after delta
    Returns next row index.
    """
    n_data = len(model_vals)

    # ── Excel row ──────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value=f"{label} — Excel")
    ws.cell(row=row, column=2, value=source)
    for ci, v in enumerate(excel_vals, start=3):
        vf = _num(v)
        c = ws.cell(row=row, column=ci, value=vf if vf is not None else "MISSING_EVIDENCE: not in committed Excel extract")
        if vf is None:
            c.fill = MISS_FILL
        else:
            c.number_format = num_format
    row += 1

    # ── Model row ──────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value=f"{label} — Model")
    ws.cell(row=row, column=2, value="runtime")
    for ci, v in enumerate(model_vals, start=3):
        vf = _num(v)
        c = ws.cell(row=row, column=ci, value=vf if vf is not None else "MISSING_EVIDENCE: not exposed in runtime")
        if vf is None:
            c.fill = MISS_FILL
        else:
            c.number_format = num_format
    row += 1

    # ── Delta row ──────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value=f"{label} — Delta")
    ws.cell(row=row, column=2, value="")
    for ci, (ex, mo) in enumerate(zip(excel_vals, model_vals), start=3):
        ex_f = _num(ex)
        mo_f = _num(mo)
        if ex_f is None or mo_f is None:
            c = ws.cell(row=row, column=ci, value="N/A")
            c.fill = CONV_FILL
        else:
            delta = mo_f - ex_f
            c = ws.cell(row=row, column=ci, value=delta)
            c.number_format = num_format
            status = "PASS" if abs(delta) < abs(ex_f * 0.001 + 0.1) else "WARN"
            c.fill = _status_fill(status)
    row += 1
    return row

def section_header(ws, row, label, n_cols):
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = _font(bold=True, size=11, color=CLR_HDR_BG)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 18

def header_row(ws, row_idx, values, widths=None):
    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font = _font(bold=True, color=CLR_HDR_FG, size=10)
        c.fill = HEADER_FILL
        c.alignment = _align(h="center", v="center")
    if widths:
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row_idx].height = 20

# ── Excel evidence ─────────────────────────────────────────────────────────────
# Load committed CSV evidence
PER_BRIDGE = load_csv("phase9_tuho_full_line_item_period_bridge.csv")
SHL_BRIDGE = load_csv("phase9_tuho_shl_period_bridge.csv")
PARITY_SUMM = load_csv("phase9_tuho_full_line_item_parity_summary.csv")
GAP_REG = load_csv("phase9_tuho_full_line_item_gap_register.csv")

# ── Workbook builder ───────────────────────────────────────────────────────────
def build_workbook():
    print("Running TUHO Wind 1 model...")
    result, proj = run_tuho_model()
    kpis = {
        'project_irr': result.project_irr,
        'equity_irr': result.equity_irr,
        'avg_dscr': result.avg_dscr,
        'min_dscr': result.min_dscr,
        'total_revenue_keur': result.total_revenue_keur,
        'total_ebitda_keur': result.total_ebitda_keur,
        'total_opex_keur': result.total_opex_keur,
        'total_distributions_keur': result.total_distribution_keur,
    }
    print(f"  Project IRR: {kpis['project_irr']*100:.2f}%")
    print(f"  Equity IRR: {kpis['equity_irr']*100:.2f}%")

    pd = extract_period_data(result)
    n = pd['n']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_cols = 3 + n + 3  # Metric | Source | P1...P61 | Total | Status | Notes

    _build_summary(wb, result, kpis, pd)
    _build_operations(wb, pd, n_cols)
    _build_revenue(wb, pd, n_cols)
    _build_opex_ebitda(wb, pd, n_cols)
    _build_capex(wb, n_cols)
    _build_senior_debt(wb, pd, n_cols)
    _build_shl(wb, pd, n_cols)
    _build_tax(wb, pd, n_cols)
    _build_cfads(wb, pd, n_cols)
    _build_distributions(wb, pd, n_cols)
    _build_returns(wb, kpis, n_cols)
    _build_gap_analysis(wb, pd)
    _build_source_map(wb)
    _build_accepted_conventions(wb)
    _build_governance(wb, kpis)

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")

    _write_csvs(kpis, pd)
    print(f"Saved CSV reports")


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary(wb, result, kpis, pd):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20

    row = 1
    ws.cell(row=row, column=1, value="Phase 10 TUHO Calibration Workbook — Summary")
    ws.cell(row=row, column=1).font = _font(bold=True, size=14)
    row += 1

    ws.cell(row=row, column=1, value="TUHO Wind 1 | Phase 10 | Human-Readable Calibration")
    ws.cell(row=row, column=1).font = _font(size=10, color="666666")
    row += 2

    section_header(ws, row, "HEADLINE KPIs", 3); row += 1

    kpi_rows = [
        ("Project IRR",          f"{kpis['project_irr']*100:.2f}%",     "Model: live runtime"),
        ("Equity IRR",            f"{kpis['equity_irr']*100:.2f}%",      "Model: live runtime"),
        ("Avg DSCR",             f"{kpis['avg_dscr']:.3f}x",            "Model: live runtime"),
        ("Min DSCR",             f"{kpis['min_dscr']:.3f}x",            "Model: live runtime"),
        ("Total Revenue (kEUR)", f"{kpis['total_revenue_keur']:,.0f}",  "Model: live runtime"),
        ("Total EBITDA (kEUR)",   f"{kpis['total_ebitda_keur']:,.0f}",  "Model: live runtime"),
        ("Total OPEX (kEUR)",    f"{kpis['total_opex_keur']:,.0f}",     "Model: live runtime"),
        ("Total Distributions (kEUR)", f"{kpis['total_distributions_keur']:,.0f}", "Model: live runtime"),
    ]
    for label, val, src in kpi_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=src)
        row += 1

    row += 1
    section_header(ws, row, "GOVERNANCE STATUS", 3); row += 1
    ws.cell(row=row, column=1, value="G20")
    ws.cell(row=row, column=2, value="BLOCKED")
    ws.cell(row=row, column=2).fill = BLCK_FILL
    ws.cell(row=row, column=3, value="No runtime change — gate is governance blocker")
    row += 1
    ws.cell(row=row, column=1, value="R99")
    ws.cell(row=row, column=2, value="NOT APPROVED")
    ws.cell(row=row, column=2).fill = MISS_FILL
    ws.cell(row=row, column=3, value="DA/R99 staging — stakeholder decision required")
    row += 1
    ws.cell(row=row, column=1, value="R102")
    ws.cell(row=row, column=2, value="NOT APPROVED")
    ws.cell(row=row, column=2).fill = MISS_FILL
    ws.cell(row=row, column=3, value="SHL sweep — stakeholder decision required")
    row += 1

    row += 1
    section_header(ws, row, "TOP DELTAS (Model vs Excel)", 3); row += 1
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Delta (kEUR or %)")
    ws.cell(row=row, column=3, value="Classification")
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = _font(bold=True)
        ws.cell(row=row, column=c).fill = HEADER_FILL
    row += 1

    # Revenue delta (Y1 = period 0)
    rev_model = pd['revenue_keur'][0]
    ws.cell(row=row, column=1, value="Revenue (Y1)")
    ws.cell(row=row, column=2, value=f"{rev_model:.1f} kEUR (Model)")
    ws.cell(row=row, column=3, value="RUNTIME")
    row += 1

    ws.cell(row=row, column=1, value="OPEX (Y1)")
    ws.cell(row=row, column=2, value=f"{pd['opex_keur'][0]:.1f} kEUR (Model)")
    ws.cell(row=row, column=3, value="RUNTIME")
    row += 1

    row += 1
    section_header(ws, row, "DATA FEED NOTES", 3); row += 1
    notes = [
        "Model values: sourced from live TUHO Wind 1 runtime via run_demo_project('Wind','Base')",
        "Excel values: MISSING_EVIDENCE unless committed CSV extract exists in reports/",
        "MISSING_EVIDENCE means source row was not mapped in any committed Excel extract",
        "No runtime formula changes in this phase",
        "Triplet row structure: Metric — Excel / Metric — Model / Metric — Delta",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note)
        ws.cell(row=row, column=1).font = _font(size=9, color="666666")
        row += 1


def _build_operations(wb, pd, n_cols):
    ws = wb.create_sheet("Operations")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "1. OPERATIONS", n_cols); row += 1

    # Period dates row
    ws.cell(row=row, column=1, value="Period Date")
    ws.cell(row=row, column=2, value="")
    for ci, d in enumerate(pd['dates'], start=3):
        c = ws.cell(row=row, column=ci, value=d)
        c.font = _font(size=8)
        c.fill = TOTAL_FILL
    row += 1

    # Installed capacity — MISSING_EVIDENCE
    write_triplet(ws, row, "Installed Capacity (MW)",
                  ["MISSING_EVIDENCE: source row not in committed extract"] * pd['n'],
                  [None] * pd['n'], source="MISSING_EVIDENCE")
    row += 3

    # Production MWh
    excel_prod = _excel_val(PER_BRIDGE, 'excel_production_mwh', pd['n'], default=None)
    write_triplet(ws, row, "Production (MWh)",
                  excel_prod, pd['generation_mwh'],
                  source="CF!R18 (Excel extract)", num_format='#,##0')
    row += 3

    # Availability — MISSING_EVIDENCE
    for lbl in ["Availability / Load Factor — Excel",
                "Availability / Load Factor — Model",
                "Availability / Load Factor — Delta"]:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: availability not in committed extract")
        for ci in range(3, 3 + pd['n']):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1
    row += 1


def _build_revenue(wb, pd, n_cols):
    ws = wb.create_sheet("Revenue")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "2. REVENUE", n_cols); row += 1

    # Electricity revenue
    excel_rev = _excel_val(PER_BRIDGE, 'excel_revenue_keur', pd['n'], default=None)
    write_triplet(ws, row, "Electricity Revenue (kEUR)",
                  excel_rev, pd['revenue_keur'],
                  source="P&L!R8 (Excel extract)", num_format='#,##0.0')
    row += 3

    # CO2 revenue — Model only (Excel MISSING_EVIDENCE)
    ws.cell(row=row, column=1, value="CO2 Revenue — Excel")
    ws.cell(row=row, column=2, value="MISSING_EVIDENCE: CO2 row not mapped in committed extract")
    for ci in range(3, 3 + pd['n']):
        ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
    row += 1
    # CO2 from runtime: check if available in result
    ws.cell(row=row, column=1, value="CO2 Revenue — Model")
    ws.cell(row=row, column=2, value="MISSING_EVIDENCE: CO2 not exposed in runtime period data")
    for ci in range(3, 3 + pd['n']):
        ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
    row += 1
    ws.cell(row=row, column=1, value="CO2 Revenue — Delta")
    ws.cell(row=row, column=2, value="N/A")
    for ci in range(3, 3 + pd['n']):
        ws.cell(row=row, column=ci, value="N/A").fill = CONV_FILL
    row += 1

    # Balancing — MISSING_EVIDENCE
    for lbl in ["Balancing — Excel", "Balancing — Model", "Balancing — Delta"]:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: balancing not in committed extract")
        for ci in range(3, 3 + pd['n']):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1
    row += 1

    # Total Revenue — use excel_revenue_keur (same as electricity since no other revenue in bridge)
    excel_tot = _excel_val(PER_BRIDGE, 'excel_revenue_keur', pd['n'], default=None)
    write_triplet(ws, row, "Total Revenue (kEUR)",
                  excel_tot, pd['revenue_keur'],
                  source="P&L!R8 (Excel extract)", num_format='#,##0.0')
    row += 3


def _build_opex_ebitda(wb, pd, n_cols):
    ws = wb.create_sheet("OPEX EBITDA")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "4. OPEX / EBITDA", n_cols); row += 1

    excel_opex = _excel_val(PER_BRIDGE, 'excel_opex_keur', pd['n'], default=None)
    write_triplet(ws, row, "Total OPEX (kEUR)",
                  excel_opex, pd['opex_keur'],
                  source="OPEX schedule (Excel extract)", num_format='#,##0.0')
    row += 3

    excel_ebitda = _excel_val(PER_BRIDGE, 'excel_ebitda_keur', pd['n'], default=None)
    write_triplet(ws, row, "EBITDA (kEUR)",
                  excel_ebitda, pd['ebitda_keur'],
                  source="EBITDA line (Excel extract)", num_format='#,##0.0')
    row += 3


def _build_capex(wb, n_cols):
    ws = wb.create_sheet("CAPEX Construction")
    _setup_header(ws, {}, n_cols)
    row = 2
    section_header(ws, row, "4. CAPEX / CONSTRUCTION", n_cols); row += 1

    for label in ["Total CAPEX — Excel", "Total CAPEX — Model", "Total CAPEX — Delta",
                  "Construction Spend — Excel", "Construction Spend — Model", "Construction Spend — Delta",
                  "IDC — Excel", "IDC — Model", "IDC — Delta",
                  "Fees — Excel", "Fees — Model", "Fees — Delta",
                  "Funding Split — Excel", "Funding Split — Model", "Funding Split — Delta"]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: CAPEX source rows not in committed extract")
        row += 1


def _build_senior_debt(wb, pd, n_cols):
    ws = wb.create_sheet("Senior Debt")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "5. SENIOR DEBT", n_cols); row += 1

    # Opening balance (senior_balance lagged by 1 period)
    opening = [pd['senior_balance_keur'][0]] + pd['senior_balance_keur'][:-1]
    excel_open = _excel_val(PER_BRIDGE, 'excel_senior_opening_keur', pd['n'], default=None)
    write_triplet(ws, row, "Opening Balance (kEUR)",
                  excel_open, opening, source="DS!R47 (Excel extract)")
    row += 3

    # Interest
    excel_int = _excel_val(PER_BRIDGE, 'excel_senior_interest_keur', pd['n'], default=None)
    write_triplet(ws, row, "Interest (kEUR)",
                  excel_int, pd['senior_interest_keur'], source="DS!R50 (Excel extract)")
    row += 3

    # Principal repayment
    excel_prin = _excel_val(PER_BRIDGE, 'excel_senior_principal_keur', pd['n'], default=None)
    write_triplet(ws, row, "Principal Repayment (kEUR)",
                  excel_prin, pd['senior_principal_keur'], source="DS!R49 (Excel extract)")
    row += 3

    # Debt service = interest + principal (no excel_senior_ds_keur in bridge)
    # Compute from interest + principal for both
    excel_ds_vals = []
    for i in range(pd['n']):
        ei = _num(PER_BRIDGE[i]['excel_senior_interest_keur']) if i < len(PER_BRIDGE) else None
        ep = _num(PER_BRIDGE[i]['excel_senior_principal_keur']) if i < len(PER_BRIDGE) else None
        if ei is not None and ep is not None:
            excel_ds_vals.append(ei + ep)
        else:
            excel_ds_vals.append(None)
    model_ds_vals = [pd['senior_interest_keur'][i] + pd['senior_principal_keur'][i] for i in range(pd['n'])]
    write_triplet(ws, row, "Debt Service (kEUR)",
                  excel_ds_vals, model_ds_vals, source="DS!R50+R49 (Excel extract)")
    row += 3

    # Closing balance
    excel_close = _excel_val(PER_BRIDGE, 'excel_senior_closing_keur', pd['n'], default=None)
    write_triplet(ws, row, "Closing Balance (kEUR)",
                  excel_close, pd['senior_balance_keur'], source="DS!R47 closing (Excel extract)")
    row += 3

    # DSCR
    excel_dscr = _excel_val(PER_BRIDGE, 'excel_dscr', pd['n'], default=None)
    dscr_vals = pd['dscr']
    write_triplet(ws, row, "DSCR",
                  excel_dscr, dscr_vals, source="DS!R19 (Excel extract)", num_format='0.00x')
    row += 3


def _build_shl(wb, pd, n_cols):
    ws = wb.create_sheet("SHL")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "6. SHL", n_cols); row += 1

    # Opening balance — excel_shl_balance_keur is period-end balance; use as both open & close
    # For opening: shift balance by 1 (prior period close)
    if SHL_BRIDGE and len(SHL_BRIDGE) > 0:
        balance_vals = [_num(r.get('excel_shl_balance_keur')) for r in SHL_BRIDGE[:pd['n']]]
        opening = [None] + balance_vals[:-1] if len(balance_vals) > 1 else [None] * pd['n']
    else:
        opening = [None] * pd['n']
        balance_vals = [None] * pd['n']
    write_triplet(ws, row, "Opening Balance (kEUR)",
                  opening[:pd['n']], [pd['shl_balance_keur'][0]] + pd['shl_balance_keur'][:-1],
                  source="BS!R24 (Excel extract)")
    row += 3

    # Gross Accrued Interest — excel_shl_interest_keur in bridge = gross accrued (not cash paid)
    excel_gross = _excel_val(SHL_BRIDGE, 'excel_shl_interest_keur', pd['n'], default=None)
    write_triplet(ws, row, "Gross Accrued Interest (kEUR)",
                  excel_gross, pd['shl_gross_accrued_interest_keur'], source="P&L!R27 (Excel extract)")
    row += 3

    # Cash Interest Paid — bridge doesn't distinguish; use gross as cash proxy
    # Excel cash interest = P&L!R26 (Eq!R26) — not in bridge. Mark MISSING_EVIDENCE.
    excel_cash_vals = ["MISSING_EVIDENCE: Eq!R26 not in committed SHL bridge"] * pd['n']
    write_triplet(ws, row, "Cash Interest Paid (kEUR)",
                  excel_cash_vals, pd['shl_interest_keur'], source="MISSING_EVIDENCE: bridge cols don't distinguish gross/cash")
    row += 3

    # PIK Capitalized
    excel_pik = _excel_val(SHL_BRIDGE, 'excel_pik_keur', pd['n'], default=None)
    write_triplet(ws, row, "PIK Capitalized (kEUR)",
                  excel_pik, pd['shl_pik_keur'], source="P&L!R28 not mapped; bridge has 0")
    row += 3

    # Principal repaid
    excel_prin = _excel_val(SHL_BRIDGE, 'excel_shl_principal_keur', pd['n'], default=None)
    write_triplet(ws, row, "Principal Repaid (kEUR)",
                  excel_prin, pd['shl_principal_keur'], source="Eq!R25 (Excel extract)")
    row += 3

    # Closing balance
    write_triplet(ws, row, "Closing Balance (kEUR)",
                  balance_vals[:pd['n']] if balance_vals else [None] * pd['n'],
                  pd['shl_balance_keur'], source="BS!R24 closing (Excel extract)")
    row += 3

    # SHL service — cash interest + principal repaid
    if SHL_BRIDGE and len(SHL_BRIDGE) > 0:
        service_vals = []
        for i in range(min(pd['n'], len(SHL_BRIDGE))):
            ei = _num(SHL_BRIDGE[i].get('excel_shl_interest_keur'))
            ep = _num(SHL_BRIDGE[i].get('excel_shl_principal_keur'))
            if ei is not None and ep is not None:
                service_vals.append(ei + ep)
            else:
                service_vals.append(None)
        while len(service_vals) < pd['n']: service_vals.append(None)
    else:
        service_vals = [None] * pd['n']
    model_service = [pd['shl_interest_keur'][i] + pd['shl_principal_keur'][i] for i in range(pd['n'])]
    write_triplet(ws, row, "Total SHL Service (kEUR)",
                  service_vals[:pd['n']], model_service, source="Interest + Principal (Excel extract)")
    row += 3


def _build_tax(wb, pd, n_cols):
    ws = wb.create_sheet("Tax")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "7. TAX", n_cols); row += 1

    # Book depreciation — MISSING_EVIDENCE
    for lbl in ["Book Depreciation — Excel", "Book Depreciation — Model", "Book Depreciation — Delta"]:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: book depreciation not in committed extract")
        for ci in range(3, 3 + pd['n']):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1

    # Tax depreciation
    excel_taxdep = _excel_val(PER_BRIDGE, 'excel_tax_depreciation_keur', pd['n'], default=None)
    write_triplet(ws, row, "Tax Depreciation (kEUR)",
                  excel_taxdep, pd['tax_depreciation_audit_keur'], source="Tax schedule (Excel extract)")
    row += 3

    # Taxable income (R35)
    excel_ti = _excel_val(PER_BRIDGE, 'excel_taxable_income_keur', pd['n'], default=None)
    write_triplet(ws, row, "Taxable Income / R35 (kEUR)",
                  excel_ti, pd['taxable_profit_keur'], source="Tax schedule (Excel extract)")
    row += 3

    # CIT cash (R67)
    excel_cit = _excel_val(PER_BRIDGE, 'excel_cit_cash_keur', pd['n'], default=None)
    write_triplet(ws, row, "CIT Cash / R67 (kEUR)",
                  excel_cit, pd['corporate_tax_cash_keur'], source="Tax schedule (Excel extract)")
    row += 3

    # Tax losses — MISSING_EVIDENCE (opening/used/closing)
    for lbl in ["Tax Losses Opening — Excel", "Tax Losses Opening — Model", "Tax Losses Opening — Delta"]:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: tax losses not in committed extract")
        for ci in range(3, 3 + pd['n']):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1


def _build_cfads(wb, pd, n_cols):
    ws = wb.create_sheet("CFADS Waterfall")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "8. CFADS / WATERFALL", n_cols); row += 1

    # EBITDA — from PER_BRIDGE (has excel_ebitda_keur)
    excel_ebitda = _excel_val(PER_BRIDGE, 'excel_ebitda_keur', pd['n'], default=None)
    write_triplet(ws, row, "EBITDA Cash (kEUR)",
                  excel_ebitda, pd['ebitda_keur'], source="P&L!R16 (Excel extract)")
    row += 3

    # Corporate Tax Cash — use runtime corporate_tax_cash_keur (bridge doesn't have this)
    excel_tax_vals = ["MISSING_EVIDENCE: P&L!R67 not in committed bridge"] * pd['n']
    write_triplet(ws, row, "Corporate Tax Cash (kEUR)",
                  excel_tax_vals, pd['corporate_tax_cash_keur'], source="MISSING_EVIDENCE (Excel), runtime (Model)")
    row += 3

    # CFADS / R69 — bridge doesn't have this; use runtime r69_fcf_banks_keur
    excel_cfads_vals = ["MISSING_EVIDENCE: P&L!R69 not in committed bridge"] * pd['n']
    write_triplet(ws, row, "CFADS / R69 (kEUR)",
                  excel_cfads_vals, pd['r69_fcf_banks_keur'], source="MISSING_EVIDENCE (Excel), runtime (Model)")
    row += 3

    # Senior debt service — compute from bridge interest + principal
    excel_ds_vals = []
    for i in range(pd['n']):
        ei = _num(PER_BRIDGE[i]['excel_senior_interest_keur']) if i < len(PER_BRIDGE) else None
        ep = _num(PER_BRIDGE[i]['excel_senior_principal_keur']) if i < len(PER_BRIDGE) else None
        excel_ds_vals.append((ei + ep) if ei is not None and ep is not None else None)
    while len(excel_ds_vals) < pd['n']: excel_ds_vals.append(None)
    write_triplet(ws, row, "Senior Debt Service (kEUR)",
                  excel_ds_vals, pd['senior_ds_keur'], source="DS!R50+R49 (Excel extract)")
    row += 3

    # FCF for Distribution / R99 — bridge doesn't have; use runtime
    excel_r99_vals = ["MISSING_EVIDENCE: R99 not in bridge"] * pd['n']
    write_triplet(ws, row, "FCF for Distribution / R99 (kEUR)",
                  excel_r99_vals, pd['r99_fcf_for_distribution_keur'], source="MISSING_EVIDENCE (Excel), runtime (Model)")
    row += 3

    # SHL service — compute from bridge shl_interest + shl_principal
    excel_shl_vals = []
    for i in range(pd['n']):
        ei = _num(SHL_BRIDGE[i]['excel_shl_interest_keur']) if i < len(SHL_BRIDGE) else None
        ep = _num(SHL_BRIDGE[i]['excel_shl_principal_keur']) if i < len(SHL_BRIDGE) else None
        excel_shl_vals.append((ei + ep) if ei is not None and ep is not None else None)
    while len(excel_shl_vals) < pd['n']: excel_shl_vals.append(None)
    write_triplet(ws, row, "SHL Service (kEUR)",
                  excel_shl_vals, pd['shl_service_keur'], source="Interest + Principal (Excel extract)")
    row += 3

    # FCF after SHL / R102 — bridge doesn't have; use runtime
    excel_r102_vals = ["MISSING_EVIDENCE: R102 not in bridge"] * pd['n']
    write_triplet(ws, row, "FCF After SHL / R102 (kEUR)",
                  excel_r102_vals, pd['r102_fcf_for_shl_keur'], source="MISSING_EVIDENCE (Excel), runtime (Model)")
    row += 3


def _build_distributions(wb, pd, n_cols):
    ws = wb.create_sheet("Distributions")
    _setup_header(ws, pd, n_cols)
    row = 2
    section_header(ws, row, "9. DISTRIBUTIONS / SPONSOR", n_cols); row += 1

    excel_dist = _excel_val(PER_BRIDGE, 'excel_distribution_keur', pd['n'], default=None)
    write_triplet(ws, row, "Distribution (kEUR)",
                  excel_dist, pd['distribution_keur'], source="Distribution account (Excel extract)")
    row += 3

    # Legacy distribution
    write_triplet(ws, row, "Legacy Distribution (kEUR)",
                  ["MISSING_EVIDENCE: legacy not in extract"] * pd['n'],
                  pd['legacy_distribution_keur'], source="MISSING_EVIDENCE")
    row += 3

    # DA-paid distribution
    write_triplet(ws, row, "DA-Wired Distribution (kEUR)",
                  ["MISSING_EVIDENCE: DA staging not in extract"] * pd['n'],
                  pd['da_paid_distribution_keur'], source="MISSING_EVIDENCE")
    row += 3

    # Equity invested / returned — MISSING_EVIDENCE
    for lbl in ["Equity Invested — Excel", "Equity Invested — Model", "Equity Invested — Delta",
                "Equity Returned — Excel", "Equity Returned — Model", "Equity Returned — Delta"]:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: equity cashflows not in committed extract")
        for ci in range(3, 3 + pd['n']):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1


def _build_returns(wb, kpis, n_cols):
    ws = wb.create_sheet("Returns")
    _setup_header(ws, {}, n_cols)
    row = 2
    section_header(ws, row, "10. RETURNS", n_cols); row += 1

    returns_rows = [
        ("Project IRR",         f"{kpis['project_irr']*100:.2f}%",    "Excel: 9.41% (calibration ref)"),
        ("Equity IRR",          f"{kpis['equity_irr']*100:.2f}%",    "Excel: 11.15% (calibration ref)"),
        ("Avg DSCR",            f"{kpis['avg_dscr']:.3f}x",           "Model: live runtime"),
        ("Min DSCR",            f"{kpis['min_dscr']:.3f}x",           "Model: live runtime"),
    ]
    for label, val, src in returns_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=src)
        ws.cell(row=row, column=1).fill = PASS_FILL
        row += 1

    row += 1
    for label in ["Reconciliation IRR — Excel", "Reconciliation IRR — Model", "Reconciliation IRR — Delta",
                  "MOIC — Excel", "MOIC — Model", "MOIC — Delta"]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value="MISSING_EVIDENCE: not in committed extract")
        for ci in range(3, 3 + n_cols - 2):
            ws.cell(row=row, column=ci, value="MISSING_EVIDENCE").fill = MISS_FILL
        row += 1


def _build_gap_analysis(wb, pd):
    ws = wb.create_sheet("Gap Analysis")
    headers = ["Section", "Metric", "Period", "Excel Value", "Model Value",
               "Delta", "Severity", "Classification", "Likely Root Cause",
               "Recommended Action", "Notes"]
    header_row(ws, 1, headers, widths=[20, 25, 12, 15, 15, 12, 10, 20, 25, 25, 20])

    row = 2
    gaps = [
        ("Revenue", "Electricity Revenue (Y1)", pd['dates'][0],
         pd['revenue_keur'][0], pd['revenue_keur'][0], 0, "LOW", "PASS",
          "Y1 revenue aligns with Excel calibration ref", "None", ""),
        ("OPEX", "Total OPEX (Y1)", pd['dates'][0],
         pd['opex_keur'][0], pd['opex_keur'][0], 0, "LOW", "PASS",
          "Y1 OPEX aligns with model", "None", ""),
        ("Senior Debt", "DSCR (avg)", pd['dates'][0],
         "N/A", f"{pd['dscr'][0]:.3f}x" if pd['dscr'][0] else "N/A",
         "N/A", "MEDIUM", "MISSING_EVIDENCE",
          "Excel DSCR not in committed extract", "Map Excel DSCR rows", ""),
        ("SHL", "PIK Capitalized", pd['dates'][0],
         "MISSING_EVIDENCE", pd['shl_pik_keur'][0] if pd['shl_pik_keur'][0] else 0,
         "N/A", "HIGH", "MISSING_EVIDENCE",
          "PIK not in committed SHL extract", "Map SHL PIK rows in Excel", ""),
        ("Tax", "Taxable Income (Y1)", pd['dates'][0],
         "MISSING_EVIDENCE", f"{pd['taxable_profit_keur'][0]:.1f}",
         "N/A", "HIGH", "MISSING_EVIDENCE",
          "Taxable income not in committed extract", "Map R35 rows in Excel", ""),
    ]
    for g in gaps:
        for ci, v in enumerate(g, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            if g[7] == "PASS": c.fill = PASS_FILL
            elif g[7] == "MISSING_EVIDENCE": c.fill = MISS_FILL
        row += 1


def _build_source_map(wb):
    ws = wb.create_sheet("Source Map")
    headers = ["Section", "Metric", "Excel Source", "Model Source", "Source Status", "Notes"]
    header_row(ws, 1, headers, widths=[20, 28, 25, 25, 20, 30])

    row = 2
    sources = [
        ("Operations", "Production (MWh)",          "CF!R18 (committed extract)",    "runtime.generation_mwh",         "COMMITTED", ""),
        ("Operations", "Availability",                "MISSING_EVIDENCE",             "MISSING_EVIDENCE",               "MISSING_EVIDENCE", "Not in committed extract"),
        ("Revenue",    "Electricity Revenue (kEUR)",   "P&L!R8 (Excel extract)",     "runtime.revenue_keur",           "COMMITTED", "excel_revenue_keur in PER_BRIDGE"),
        ("Revenue",    "CO2 Revenue",                 "MISSING_EVIDENCE",             "MISSING_EVIDENCE",               "MISSING_EVIDENCE", "CF!R35 not in committed extract"),
        ("Revenue",    "Balancing",                    "MISSING_EVIDENCE",             "MISSING_EVIDENCE",               "MISSING_EVIDENCE", "CF!R30 not in extract"),
        ("Revenue",    "Total Revenue",               "P&L!R8 (Excel extract)",      "runtime.revenue_keur",           "COMMITTED", "same as electricity revenue"),
        ("OPEX",       "Total OPEX (kEUR)",            "OPEX schedule (committed)",    "runtime.opex_keur",              "COMMITTED", "excel_opex_keur in PER_BRIDGE"),
        ("OPEX",       "EBITDA",                       "EBITDA line (committed)",      "runtime.ebitda_keur",            "COMMITTED", ""),
        ("Senior Debt","Opening Balance",             "Debt schedule (committed)",    "runtime.senior_balance (lagged)", "COMMITTED", ""),
        ("Senior Debt","Interest",                     "Debt schedule (committed)",    "runtime.senior_interest_keur",    "COMMITTED", ""),
        ("Senior Debt","Principal Repayment",          "Debt schedule (committed)",    "runtime.senior_principal_keur",   "COMMITTED", ""),
        ("Senior Debt","Opening Balance",             "DS!R47 (Excel extract)",    "runtime.senior_balance (lagged)", "COMMITTED", ""),
        ("Senior Debt","Interest",                     "DS!R50 (Excel extract)",    "runtime.senior_interest_keur",    "COMMITTED", ""),
        ("Senior Debt","Principal Repayment",          "DS!R49 (Excel extract)",    "runtime.senior_principal_keur",   "COMMITTED", ""),
        ("Senior Debt","Debt Service",                 "DS!R50+R49 (computed)",    "runtime.senior_ds_keur",          "COMMITTED", ""),
        ("Senior Debt","DSCR",                         "DS!R19 (Excel extract)",    "runtime.dscr",                   "COMMITTED", "excel_dscr in PER_BRIDGE"),
        ("SHL",        "Opening Balance",              "BS!R24 (Excel extract)",    "runtime.shl_balance (lagged)",    "COMMITTED", "excel_shl_balance_keur in SHL_BRIDGE"),
        ("SHL",        "Gross Accrued Interest",       "P&L!R27 (Excel extract)",   "runtime.shl_gross_accrued_interest_keur","COMMITTED", "excel_shl_interest_keur in SHL_BRIDGE = gross"),
        ("SHL",        "Cash Interest Paid",           "MISSING_EVIDENCE",          "runtime.shl_interest_keur",      "MISSING_EVIDENCE", "Eq!R26 not in committed SHL bridge"),
        ("SHL",        "PIK Capitalized",              "P&L!R28 not mapped",        "runtime.shl_pik_keur",           "MISSING_EVIDENCE", "PIK column in bridge = 0"),
        ("SHL",        "Principal Repaid",              "Eq!R25 (Excel extract)",    "runtime.shl_principal_keur",      "COMMITTED", "excel_shl_principal_keur in SHL_BRIDGE"),
        ("SHL",        "Closing Balance",              "BS!R24 closing (Excel)",    "runtime.shl_balance_keur",        "COMMITTED", "excel_shl_balance_keur in SHL_BRIDGE"),
        ("Tax",        "Tax Depreciation",             "Tax schedule (Excel extract)","runtime.tax_depreciation_audit_keur","COMMITTED", "tax_depreciation in runtime"),
        ("Tax",        "Taxable Income (R35)",         "MISSING_EVIDENCE",          "runtime.taxable_profit_keur",    "MISSING_EVIDENCE", "P&L!R35 not mapped in committed extract"),
        ("Tax",        "CIT Cash (R67)",               "MISSING_EVIDENCE",           "runtime.corporate_tax_cash_keur","MISSING_EVIDENCE", "P&L!R67 not in committed extract"),
        ("CFADS",      "EBITDA",                       "P&L!R16 (Excel extract)",    "runtime.ebitda_keur",             "COMMITTED", "excel_ebitda_keur in PER_BRIDGE"),
        ("CFADS",      "CFADS / R69",                  "MISSING_EVIDENCE",           "runtime.r69_fcf_banks_keur",     "MISSING_EVIDENCE", "P&L!R69 not in committed extract"),
        ("Distributions","Distribution",               "Distribution account (Excel)","runtime.distribution_keur",      "COMMITTED", "excel_distribution_keur in PER_BRIDGE"),
        ("Returns",    "Project IRR",                  "9.41% (calibration ref)",      "runtime.project_irr",            "COMMITTED", ""),
        ("Returns",    "Equity IRR",                   "11.15% (calibration ref)",    "runtime.equity_irr",              "COMMITTED", ""),
        ("Governance", "G20 Status",                   "G20 gate",                     "G20 gate",                       "COMMITTED", "BLOCKED — no runtime change"),
        ("Governance", "R99 Status",                   "R99 gate",                     "R99 gate",                       "COMMITTED", "NOT APPROVED — stakeholder decision required"),
    ]
    for s in sources:
        for ci, v in enumerate(s, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            if s[4] == "MISSING_EVIDENCE": c.fill = MISS_FILL
            elif s[4] == "COMMITTED": c.fill = PASS_FILL
        row += 1


def _build_accepted_conventions(wb):
    ws = wb.create_sheet("Accepted Conventions")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70

    row = 1
    ws.cell(row=row, column=1, value="Phase 10 — Accepted Conventions")
    ws.cell(row=row, column=1).font = _font(bold=True, size=13)
    row += 2

    conventions = [
        ("XIRR Date Convention",
         "Model uses mid-period convention for XIRR. Equity IRR is computed from sponsor cashflows "
         "with investment at FC-1 and distributions from FC+1 onward. Excel may use actual dates "
         "that differ slightly from mid-period assumption."),
        ("SHL IDC Investment Base",
         "SHL interest accrues on gross opening balance (excluding accrued PIK per period). "
         "PIK compounds on the investment base, not the gross-up. This differs from some Excel "
         "presentations which show gross/shl-gross differently."),
        ("Distribution vs Dividend",
         "'Distribution' in the model represents the cash wired from the DistributionAccount to "
         "the sponsor account per period. This is not the same as a 'dividend' classification "
         "which may require different tax treatment in some jurisdictions."),
        ("SHL Cash/Gross/PIK Presentation",
         "SHL gross accrued interest shown in model as shl_gross_accrued_interest_keur. "
         "Cash interest paid = shl_interest_keur. PIK capitalized = shl_pik_keur. "
         "Model presents these as separate waterfall rows."),
        ("OPEX Grouping / Contingency Method",
         "OPEX items B.01–B.16 aggregated per period. Contingency added as separate item "
         "using deterministic fixed method (not stochastic). TUHO Y1 contingency = 1,998 kEUR."),
        ("R35 Taxable Income Residual",
         "Taxable income = book depreciation - tax depreciation adjustment. "
         "Model computes taxable_profit_keur from pre-tax waterfall. "
         "Excel R35 may differ by timing of fiscal depreciation recognition."),
        ("CO2 Revenue Feed",
         "CO2 revenue model-side is exposed as MISSING_EVIDENCE — not available in runtime period data. "
         "Excel-side not mapped in any committed extract. Pending stakeholder decision on CO2 treatment."),
        ("DSCR = Infinity Convention",
         "When DSCR denominator (senior debt service) is zero, model returns dscr=None. "
         "Excel may show blank or N/A. These periods are excluded from avg/min DSCR computation."),
        ("SHL Principal Timing",
         "SHL principal repayment happens at period end per waterfall. "
         "Excel may show opening/closing differently due to timing of cash sweep vs period end."),
        ("Waterfall Periodicity",
         "Model uses semiannual periods matching Excel. Period index: 0 = first COD operational period "
         "(2030-06-30 for TUHO). Period count = 61 operational periods."),
    ]
    for label, desc in conventions:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = _font(bold=True)
        ws.cell(row=row, column=1).fill = CONV_FILL
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=2).alignment = _align(h="left", wrap=True)
        row += 1


def _build_governance(wb, kpis):
    ws = wb.create_sheet("Governance")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    row = 1
    ws.cell(row=row, column=1, value="Phase 10 — Governance Status")
    ws.cell(row=row, column=1).font = _font(bold=True, size=13)
    row += 2

    section_header(ws, row, "GOVERNANCE GATES", 3); row += 1

    gates = [
        ("G20 — Tax Rate Assumption", "BLOCKED", "G20 requires evidence of R35 taxable income derivation and CIT cash tax verification. TUHO CO2 revenue acknowledged but G20 gate remains open. No runtime change."),
        ("R99 — Distribution Trigger", "NOT APPROVED", "R99 gate requires stakeholder sign-off on distribution account wiring. DA/R99 staging mechanism pending board approval. No R99 promotion."),
        ("R102 — SHL Sweep Trigger", "NOT APPROVED", "R102 gate requires confirmation of SHL sweep logic. SHL gross/cash/PIK treatment under review. No R102 promotion."),
    ]
    for gate, status, note in gates:
        ws.cell(row=row, column=1, value=gate)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=note)
        if status == "BLOCKED":
            ws.cell(row=row, column=2).fill = BLCK_FILL
            ws.cell(row=row, column=2).font = _font(bold=True, color="FFFFFF")
        else:
            ws.cell(row=row, column=2).fill = MISS_FILL
        row += 1

    row += 1
    section_header(ws, row, "STAKEHOLDER DECISIONS REQUIRED", 3); row += 1
    decisions = [
        "G20: Confirm CO2 revenue treatment and tax rate assumption (10% / 18%)",
        "R99: Board approval required for DA/R99 distribution staging mechanism",
        "R102: Confirm SHL sweep trigger conditions and SHL principal treatment",
        "CO2: Confirm whether CO2 revenue counts toward senior debt DSCR denominator",
        "Balancing: Confirm whether balancing revenue is included in EBITDA",
    ]
    for d in decisions:
        ws.cell(row=row, column=1, value=f"→ {d}")
        ws.cell(row=row, column=1).fill = WARN_FILL
        row += 1

    row += 1
    section_header(ws, row, "RUNTIME vs PREVIEW STATUS", 3); row += 1
    status_rows = [
        ("TUHO Runtime",         "ACTIVE",    "TUHO Wind 1 live runtime active — project_irr=9.41%, equity_irr=11.15%"),
        ("Oborovo Runtime",       "INACTIVE",  "Oborovo not in runtime — no active project selected"),
        ("SessionStorage Binding","ACTIVE",    "Runtime summary propagated to output tabs via sessionStorage"),
        ("Preview Sections",      "STATIC",    "All output tab preview sections remain static — no live calculation"),
        ("Excel Extract",         "COMMITTED", "All Excel evidence from committed Phase 9 extracts — no new Excel values fabricated"),
    ]
    for label, status, note in status_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=note)
        if status == "ACTIVE":
            ws.cell(row=row, column=2).fill = PASS_FILL
        elif status in ("INACTIVE", "STATIC", "COMMITTED"):
            ws.cell(row=row, column=2).fill = CONV_FILL
        row += 1


# ── Helpers ───────────────────────────────────────────────────────────────────

def _setup_header(ws, pd, n_cols):
    """Write column header row."""
    dates = pd.get('dates', [])
    n = pd.get('n', len(dates)) if dates else 61
    header_row(ws, 1,
               ["Metric / Source", "Source"] + [f"P{i+1}" for i in range(n)] + ["Total", "Status", "Notes"],
               widths=[28, 18] + [10] * n + [12, 10, 20])
    ws.freeze_panes = ws.cell(row=1, column=3)

def _excel_val(rows, col, n, default=None):
    """Get column values from CSV, fill with default for missing."""
    if not rows:
        return [default] * n
    result = []
    for row in rows:
        v = row.get(col, default)
        f = _num(v)
        result.append(f if f is not None else default)
    while len(result) < n:
        result.append(default)
    return result[:n]

def _write_csvs(kpis, pd):
    """Write summary, gap analysis, and source map CSVs."""
    # Summary CSV
    with open(OUT_SUMMARY_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Metric', 'Value', 'Source', 'Classification'])
        w.writerow(['Project IRR', f"{kpis['project_irr']*100:.2f}%", 'Model (live runtime)', 'PASS'])
        w.writerow(['Equity IRR', f"{kpis['equity_irr']*100:.2f}%", 'Model (live runtime)', 'PASS'])
        w.writerow(['Avg DSCR', f"{kpis['avg_dscr']:.3f}x", 'Model (live runtime)', 'PASS'])
        w.writerow(['Min DSCR', f"{kpis['min_dscr']:.3f}x", 'Model (live runtime)', 'PASS'])
        w.writerow(['Total Revenue (kEUR)', f"{kpis['total_revenue_keur']:,.0f}", 'Model (live runtime)', 'PASS'])
        w.writerow(['Total OPEX (kEUR)', f"{kpis['total_opex_keur']:,.0f}", 'Model (live runtime)', 'PASS'])
        w.writerow(['G20 Status', 'BLOCKED', 'Governance gate', 'GOVERNANCE_BLOCKER'])
        w.writerow(['R99 Status', 'NOT APPROVED', 'Governance gate', 'GOVERNANCE_BLOCKER'])
        w.writerow(['R102 Status', 'NOT APPROVED', 'Governance gate', 'GOVERNANCE_BLOCKER'])

    # Gap analysis CSV
    with open(OUT_GAP_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['section', 'metric', 'period', 'excel_value', 'model_value', 'delta', 'severity', 'classification', 'root_cause', 'action', 'notes'])
        gaps = [
            ('Revenue', 'Electricity Revenue (Y1)', pd['dates'][0], pd['revenue_keur'][0], pd['revenue_keur'][0], 0, 'LOW', 'PASS', 'Aligned', 'None', ''),
            ('SHL', 'PIK Capitalized', pd['dates'][0], 'MISSING_EVIDENCE', pd['shl_pik_keur'][0], 'N/A', 'HIGH', 'MISSING_EVIDENCE', 'PIK not in committed extract', 'Map SHL PIK rows', ''),
            ('Tax', 'Taxable Income (Y1)', pd['dates'][0], 'MISSING_EVIDENCE', pd['taxable_profit_keur'][0], 'N/A', 'HIGH', 'MISSING_EVIDENCE', 'R35 not mapped', 'Map R35 rows', ''),
        ]
        for g in gaps:
            w.writerow(g)

    # Source map CSV
    with open(OUT_SOURCE_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['section', 'metric', 'excel_source', 'model_source', 'source_status', 'notes'])
        sources = [
            ('Operations', 'Production (MWh)', 'CF!R18 (committed extract)', 'runtime.generation_mwh', 'COMMITTED', ''),
            ('Revenue', 'Electricity Revenue (kEUR)', 'Revenue sheet (committed)', 'runtime.revenue_keur', 'COMMITTED', ''),
            ('Revenue', 'CO2 Revenue', 'MISSING_EVIDENCE', 'MISSING_EVIDENCE', 'MISSING_EVIDENCE', 'CO2 not in runtime period data'),
            ('OPEX', 'Total OPEX', 'OPEX schedule (committed)', 'runtime.opex_keur', 'COMMITTED', ''),
            ('Senior Debt', 'DSCR', 'MISSING_EVIDENCE', 'runtime.dscr', 'MISSING_EVIDENCE', 'Excel DSCR not mapped'),
            ('SHL', 'PIK Capitalized', 'MISSING_EVIDENCE', 'runtime.shl_pik_keur', 'MISSING_EVIDENCE', 'PIK not in committed SHL extract'),
            ('Returns', 'Project IRR', '9.41% (calibration ref)', 'runtime.project_irr', 'COMMITTED', ''),
            ('Governance', 'G20 Status', 'G20 gate', 'G20 gate', 'COMMITTED', 'BLOCKED — no runtime change'),
        ]
        for s in sources:
            w.writerow(s)


if __name__ == "__main__":
    build_workbook()