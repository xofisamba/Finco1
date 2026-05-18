#!/usr/bin/env python3
"""
Phase 6 — Model Stack Comparison Export
TUHO Wind 1: Excel vs Python (flag-on) period-by-period comparison.

Report-only diagnostic script. No production runtime changes.

Usage:
    python scripts/export_phase6_model_stack_comparison.py

Output:
    reports/phase6_model_stack_comparison.xlsx
    reports/phase6_model_stack_comparison_long.csv
    reports/phase6_model_stack_comparison_wide.csv
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Thresholds ────────────────────────────────────────────────────────────────
PASS_TOTAL_THRESHOLD = 100_000   # kEUR
PASS_PERIOD_THRESHOLD = 25_000   # kEUR per period
MINOR_TOTAL_THRESHOLD = 500_000  # kEUR

# ─── Paths ────────────────────────────────────────────────────────────────────
REPO = Path(__file__).parent.parent
FIXTURE_JSON = REPO / "tests/fixtures/excel_tuho_full_model_extract.json"
REPORTS_DIR = REPO / "reports"
OUT_XLSX = REPORTS_DIR / "phase6_model_stack_comparison.xlsx"
OUT_LONG_CSV = REPORTS_DIR / "phase6_model_stack_comparison_long.csv"
OUT_WIDE_CSV = REPORTS_DIR / "phase6_model_stack_comparison_wide.csv"

# ─── Period helpers ────────────────────────────────────────────────────────────

def period_label(i: int) -> str:
    yr = i // 2 + 1
    h = "H1" if i % 2 == 0 else "H2"
    return f"Y{yr:02d}{h}"

# ─── Data loading ─────────────────────────────────────────────────────────────

def load_excel_fixture():
    with open(FIXTURE_JSON) as f:
        return json.load(f)

def run_python_model():
    sys.path.insert(0, str(REPO))
    from dataclasses import replace
    from app.project_factories import create_default_tuho_wind1
    from app.ui_runner import _build_period_engine
    from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner

    def _run(project):
        engine = _build_period_engine(project)
        config = WaterfallRunConfig.from_inputs(project, engine)
        return WaterfallRunner(project, engine).run(config)

    tuho = replace(
        create_default_tuho_wind1(),
        info=replace(create_default_tuho_wind1().info, use_tax_bridge_engine=True),
        tax=replace(create_default_tuho_wind1().tax, cit_cash_tax_start_operating_index=25),
    )
    return _run(tuho)

# ─── Row builder ───────────────────────────────────────────────────────────────

def make_rows(excel: dict, result) -> dict:
    cols = excel["period_diagnostic_columns"]
    edata = excel["period_diagnostics"]
    pdata = result.periods
    N = 60

    def excel_vals(prefix: str):
        """Extract column values from Excel fixture by prefix."""
        matches = [c for c in cols if c.startswith(prefix)]
        if not matches:
            return [0.0] * N
        idx = cols.index(matches[0])
        out = []
        for row in edata:
            v = row[idx] if idx < len(row) else 0.0
            out.append(float(v) if v is not None else 0.0)
        return out

    def py(field: str):
        out = []
        for p in pdata:
            v = getattr(p, field, None)
            out.append(float(v) if v is not None else 0.0)
        return out

    rows = {}

    # Revenue
    rows["revenue"] = dict(
        label="Electricity Revenue", category="Revenue",
        excel=excel_vals("P&L.total_revenues_keur"),
        python=py("revenue_keur"),
        excel_source="P&L!total_revenues_keur",
        python_source="result.periods[i].revenue_keur",
        confidence="exact",
        note="",
    )

    # OPEX
    rows["opex"] = dict(
        label="Total OPEX", category="OPEX",
        excel=[abs(v) for v in excel_vals("CF.operating_expenses_after_bank_tax_keur")],
        python=py("opex_keur"),
        excel_source="CF!operating_expenses_after_bank_tax_keur",
        python_source="result.periods[i].opex_keur",
        confidence="exact",
        note="",
    )

    # EBITDA
    rows["ebitda"] = dict(
        label="EBITDA", category="EBITDA",
        excel=excel_vals("CF.ebitda_keur"),
        python=py("ebitda_keur"),
        excel_source="CF!ebitda_keur",
        python_source="result.periods[i].ebitda_keur",
        confidence="exact",
        note="",
    )

    # FCF for banks
    rows["fcf_banks"] = dict(
        label="FCF for Banks", category="Free Cash Flow",
        excel=excel_vals("CF.free_cash_flow_for_banks_keur"),
        python=py("cf_after_tax_keur"),
        excel_source="CF!free_cash_flow_for_banks_keur",
        python_source="result.periods[i].cf_after_tax_keur",
        confidence="approximate",
        note="Python uses cf_after_tax_keur as proxy",
    )

    # FCF for distribution
    rows["fcf_dist"] = dict(
        label="FCF for Distribution", category="Free Cash Flow",
        excel=excel_vals("CF.free_cash_flow_for_distribution_keur"),
        python=py("r98_distribution_account_keur"),
        excel_source="CF!free_cash_flow_for_distribution_keur",
        python_source="result.periods[i].r98_distribution_account_keur",
        confidence="approximate",
        note="",
    )

    # Book depreciation
    rows["dep_book"] = dict(
        label="Book Depreciation", category="Depreciation",
        excel=excel_vals("Dep.depreciation_keur"),
        python=py("depreciation_keur"),
        excel_source="Dep!depreciation_keur",
        python_source="result.periods[i].depreciation_keur",
        confidence="exact",
        note="",
    )

    # Unlevered depreciation
    rows["dep_unlevered"] = dict(
        label="Unlevered Depreciation", category="Depreciation",
        excel=excel_vals("Dep.unlevered_depreciation_keur"),
        python=[0.0] * N,
        excel_source="Dep!unlevered_depreciation_keur",
        python_source="N/A",
        confidence="unmapped",
        note="Python does not compute unlevered depreciation separately",
    )

    # Senior interest
    rows["senior_interest"] = dict(
        label="Senior Interest", category="Senior Debt",
        excel=excel_vals("DS.senior_net_interest_keur"),
        python=py("interest_senior_keur"),
        excel_source="DS!senior_net_interest_keur",
        python_source="result.periods[i].interest_senior_keur",
        confidence="exact",
        note="",
    )

    # Senior principal repayment
    rows["senior_principal"] = dict(
        label="Senior Principal Repayment", category="Senior Debt",
        excel=excel_vals("DS.senior_principal_keur"),
        python=py("senior_principal_keur"),
        excel_source="DS!senior_principal_keur",
        python_source="result.periods[i].senior_principal_keur",
        confidence="exact",
        note="",
    )

    # Senior closing balance (cumulative from opening - principal)
    excel_bal = []
    bal = 0.0
    for i in range(N):
        principal = edata[i][cols.index("DS.senior_principal_keur")] if "DS.senior_principal_keur" in cols else 0.0
        bal = bal - float(principal)
        excel_bal.append(bal)
    rows["senior_balance"] = dict(
        label="Senior Closing Balance", category="Senior Debt",
        excel=excel_bal,
        python=py("senior_balance_keur"),
        excel_source="DS!senior_principal_keur (cumulative)",
        python_source="result.periods[i].senior_balance_keur",
        confidence="approximate",
        note="",
    )

    # SHL gross-accrued interest
    shl_data = excel.get("shl", [])
    rows["shl_gross_accrued"] = dict(
        label="SHL Gross Accrued Interest", category="SHL",
        excel=[row[3] if len(row) > 3 else 0.0 for row in shl_data[:N]],
        python=py("shl_gross_accrued_interest_keur"),
        excel_source="SHL schedule!gross_interest",
        python_source="result.periods[i].shl_gross_accrued_interest_keur",
        confidence="approximate",
        note="SHL gross-accrued is candidate driver; treatment may differ from Excel",
    )

    # SHL closing balance
    rows["shl_balance"] = dict(
        label="SHL Closing Balance", category="SHL",
        excel=[row[2] if len(row) > 2 else 0.0 for row in shl_data[:N]],
        python=py("shl_balance_keur"),
        excel_source="SHL schedule!closing",
        python_source="result.periods[i].shl_balance_keur",
        confidence="approximate",
        note="",
    )

    # SHL PIK / capitalized interest
    rows["shl_pik"] = dict(
        label="SHL PIK / Capitalised Interest", category="SHL",
        excel=[row[5] if len(row) > 5 else 0.0 for row in shl_data[:N]],
        python=py("shl_pik_keur"),
        excel_source="SHL schedule!capitalized_interest",
        python_source="result.periods[i].shl_pik_keur",
        confidence="approximate",
        note="",
    )

    # Taxable income / R35
    rows["r35"] = dict(
        label="Taxable Income / R35 (Python flag-on)", category="CIT / Tax",
        excel=excel_vals("P&L.taxable_income_keur"),
        python=py("taxable_income_before_losses_audit_keur"),
        excel_source="P&L!taxable_income_keur",
        python_source="result.periods[i].taxable_income_before_losses_audit_keur",
        confidence="exact",
        note="Python R35 uses flag-on tax bridge; Excel may differ in construction-period treatment",
    )

    # R41 taxable profit after losses
    rows["r41"] = dict(
        label="Taxable Profit After Losses / R41", category="CIT / Tax",
        excel=excel_vals("P&L.taxable_income_keur"),
        python=py("taxable_income_after_losses_keur"),
        excel_source="P&L!taxable_income_keur (proxy)",
        python_source="result.periods[i].taxable_income_after_losses_keur",
        confidence="approximate",
        note="Excel does not separately surface R41; P&L.taxable_income_keur is closest proxy",
    )

    # CIT accrual
    rows["cit_accrual"] = dict(
        label="CIT Accrual / R43", category="CIT / Tax",
        excel=excel_vals("P&L.corporate_income_tax_keur"),
        python=py("corporate_tax_accrual_keur"),
        excel_source="P&L!corporate_income_tax_keur",
        python_source="result.periods[i].corporate_tax_accrual_keur",
        confidence="approximate",
        note="Python uses flag-on tax bridge accrual; Excel CIT includes construction period",
    )

    # R67 cash tax
    rows["r67"] = dict(
        label="Cash Tax / R67", category="CIT / Tax",
        excel=[0.0] * N,
        python=py("r67_excel_style_cash_tax_diagnostic_keur"),
        excel_source="N/A (not in Excel fixture)",
        python_source="result.periods[i].r67_excel_style_cash_tax_diagnostic_keur",
        confidence="unmapped",
        note="R67 not in Excel fixture; see test_r67_yrs13to30_residual fixture",
    )

    # R99 audit-only
    rows["r99"] = dict(
        label="R99 FCF for Distribution [AUDIT ONLY]", category="CIT / Tax",
        excel=[0.0] * N,
        python=py("r99_fcf_for_distribution_keur"),
        excel_source="N/A",
        python_source="result.periods[i].r99_fcf_for_distribution_keur",
        confidence="unmapped",
        note="R99 is audit-only / BLOCKED; not a runtime driver",
    )

    # R102 audit-only
    rows["r102"] = dict(
        label="R102 FCF for SHL [AUDIT ONLY]", category="CIT / Tax",
        excel=[0.0] * N,
        python=py("r102_fcf_for_shl_keur"),
        excel_source="N/A",
        python_source="result.periods[i].r102_fcf_for_shl_keur",
        confidence="unmapped",
        note="R102 is audit-only / BLOCKED; not a runtime driver",
    )

    return rows

# ─── CSV writers ───────────────────────────────────────────────────────────────

def write_long_csv(rows: dict, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("category,metric,row_type,period_idx,period_label,"
                "excel_value,python_value,delta,delta_pct,excel_source,python_source,confidence,notes\n")
        for key, row in rows.items():
            for i in range(60):
                ev = row["excel"][i] if i < len(row["excel"]) else 0.0
                pv = row["python"][i] if i < len(row["python"]) else 0.0
                delta = pv - ev
                dpct = delta / abs(ev) if ev != 0 else 0.0
                for rt, v in [("Excel", ev), ("Python", pv), ("Delta", delta)]:
                    pct = (delta / abs(ev)) if rt == "Delta" and ev != 0 else 0.0
                    if rt == "Delta %":
                        v_fmt = f"{pct:.4f}"
                    else:
                        v_fmt = f"{v:.1f}"
                    f.write(
                        f"{row['category']},{row['label']},{rt},{i},{period_label(i)},"
                        f"{ev:.1f},{pv:.1f},{delta:.1f},{pct:.4f},"
                        f"{row['excel_source']},{row['python_source']},"
                        f"{row['confidence']},{row['note'][:80]}\n"
                    )

def write_wide_csv(rows: dict, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("category,metric,row_type,source,")
        for i in range(60):
            f.write(f"P{i+1:02d},")
        f.write("total,notes\n")

        for key, row in rows.items():
            for rt in ["Excel", "Python", "Delta"]:
                vals = []
                total = 0.0
                for i in range(60):
                    ev = row["excel"][i] if i < len(row["excel"]) else 0.0
                    pv = row["python"][i] if i < len(row["python"]) else 0.0
                    if rt == "Excel":
                        v = ev
                    elif rt == "Python":
                        v = pv
                    else:
                        v = pv - ev
                        total += abs(v)
                    vals.append(v)
                if rt != "Delta":
                    total = sum(vals)
                f.write(f"{row['category']},{row['label']},{rt},{row['python_source'][:40]},")
                for v in vals:
                    f.write(f"{v:.1f},")
                f.write(f"{total:.1f},{row['note'][:80]}\n")

# ─── XLSX writer ───────────────────────────────────────────────────────────────

def make_status(total_delta: float, max_period_delta: float) -> str:
    if abs(total_delta) <= PASS_TOTAL_THRESHOLD and abs(max_period_delta) <= PASS_PERIOD_THRESHOLD:
        return "PASS"
    if abs(total_delta) <= MINOR_TOTAL_THRESHOLD:
        return "MINOR"
    return "MATERIAL"

def write_xlsx(rows: dict, path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    delta_fill = PatternFill("solid", fgColor="FFF2CC")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    minor_fill = PatternFill("solid", fgColor="FFEB9C")
    mat_fill = PatternFill("solid", fgColor="FFC7CE")

    # ── Summary ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.cell(1, 1, "Phase 6 Model Stack Comparison — Summary").font = Font(bold=True, size=14)
    ws.cell(2, 1, "TUHO Wind 1 | Excel vs Python (flag-on) | Y01-Y30 | All values in kEUR")
    ws.cell(3, 1, "")

    hdrs = ["Metric", "Category", "Total Excel", "Total Python", "Total Delta",
            "Delta %", "Max Period Delta", "Status", "Notes"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = hdr_font
        c.fill = hdr_fill

    summary_keys = ["revenue", "opex", "ebitda", "fcf_banks", "fcf_dist",
                    "dep_book", "senior_interest", "senior_principal",
                    "r35", "cit_accrual"]

    r = 5
    for key in summary_keys:
        row = rows[key]
        evals = [v for v in row["excel"] if isinstance(v, (int, float))]
        pvals = row["python"]
        tot_e = sum(evals)
        tot_p = sum(pvals)
        tot_d = tot_p - tot_e
        max_d = max((abs(pvals[i] - evals[i]) for i in range(60)), default=0.0)
        pct = tot_d / abs(tot_e) if tot_e != 0 else 0.0
        status = make_status(tot_d, max_d)

        fill = pass_fill if status == "PASS" else (minor_fill if status == "MINOR" else mat_fill)
        data = [row["label"], row["category"],
                f"{tot_e:,.1f}", f"{tot_p:,.1f}", f"{tot_d:+,.1f}",
                f"{pct:.1%}", f"{max_d:,.1f}", status, row["note"][:60]]
        for j, v in enumerate(data, 1):
            c = ws.cell(r, j, v)
            c.fill = fill
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14

    # ── Category sheets ────────────────────────────────────────────────────────
    cat_sheets = {
        "Revenue":         ["revenue"],
        "OPEX":            ["opex"],
        "EBITDA":          ["ebitda"],
        "Free Cash Flow":  ["fcf_banks", "fcf_dist"],
        "Depreciation":    ["dep_book", "dep_unlevered"],
        "Senior Debt":     ["senior_interest", "senior_principal", "senior_balance"],
        "SHL":             ["shl_gross_accrued", "shl_balance", "shl_pik"],
        "CIT Tax":         ["r35", "r41", "cit_accrual", "r67", "r99", "r102"],
    }

    for sheet_name, keys in cat_sheets.items():
        ws = wb.create_sheet(sheet_name)
        # Row 1: P01..P60
        ws.cell(1, 1, "")
        for i in range(60):
            ws.cell(1, i + 2, f"P{i+1:02d}")
        # Row 2: Y01H1..Y30H2
        ws.cell(2, 1, "")
        for i in range(60):
            ws.cell(2, i + 2, period_label(i))
        ws.row_dimensions[3].height = 6  # spacer

        rn = 4
        for key in keys:
            row = rows[key]
            excel_vals = row["excel"]
            python_vals = row["python"]

            for label, values, is_delta, is_pct in [
                (row["label"] + " (Excel)", excel_vals, False, False),
                (row["label"] + " (Python)", python_vals, False, False),
                (row["label"] + " (Delta)", [python_vals[i] - (excel_vals[i] if i < len(excel_vals) else 0) for i in range(60)], True, False),
                (row["label"] + " (Delta %)", [0.0] * 60, True, True),
            ]:
                if is_pct:
                    for i in range(60):
                        ev = excel_vals[i] if i < len(excel_vals) else 0
                        pv = python_vals[i] if i < len(python_vals) else 0
                        values[i] = (pv - ev) / abs(ev) if ev != 0 else 0.0

                c = ws.cell(rn, 1, label)
                if "Delta" in label:
                    c.font = Font(bold=True)
                    for i in range(60):
                        ws.cell(rn, i + 2).fill = delta_fill

                for i in range(60):
                    cell = ws.cell(rn, i + 2, values[i])
                    cell.number_format = "0.0%" if is_pct else "#,##0.0"
                rn += 1

            # blank row
            ws.row_dimensions[rn].height = 4
            rn += 1

        ws.freeze_panes = "B4"
        ws.column_dimensions["A"].width = 30
        for i in range(60):
            ws.column_dimensions[get_column_letter(i + 2)].width = 10

    # ── Delta Flags ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Delta Flags")
    fhdrs = ["Metric", "Category", "Period", "Excel", "Python", "Delta",
             "Delta %", "Likely Cause", "Status", "Recommended Action"]
    for j, h in enumerate(fhdrs, 1):
        c = ws.cell(1, j, h)
        c.font = hdr_font
        c.fill = hdr_fill

    all_deltas = []
    for key, row in rows.items():
        if row["confidence"] == "unmapped":
            continue
        evals = row["excel"]
        pvals = row["python"]
        for i in range(60):
            ev = evals[i] if i < len(evals) else 0.0
            pv = pvals[i] if i < len(pvals) else 0.0
            d = pv - ev
            pct = d / abs(ev) if ev != 0 else 0.0
            if abs(d) > 1.0:
                all_deltas.append((key, row, i, ev, pv, d, pct))

    all_deltas.sort(key=lambda x: -abs(x[5]))
    r = 2
    for key, row, i, ev, pv, d, pct in all_deltas[:80]:
        cause = "Minor timing or rounding"
        status = "MINOR"
        action = "Monitor"
        if abs(d) > 500:
            cause = "Construction-period mismatch or source basis difference"
            status = "MATERIAL"
            action = "Investigate"
        elif abs(d) > 100:
            cause = "Moderate difference"
            status = "MINOR"
            action = "Document"

        fill = mat_fill if status == "MATERIAL" else minor_fill
        for j, v in enumerate([row["label"], row["category"], period_label(i),
                               f"{ev:.1f}", f"{pv:.1f}", f"{d:.1f}", f"{pct:.1%}",
                               cause, status, action], 1):
            c = ws.cell(r, j, v)
            if j >= 4:
                c.fill = fill
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["H"].width = 45
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 20

    # ── Source Mapping ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Source Mapping")
    shdrs = ["Category", "Metric", "Excel Sheet", "Excel Row/Cell",
             "Python Object", "Python Field", "Confidence", "Notes"]
    for j, h in enumerate(shdrs, 1):
        c = ws.cell(1, j, h)
        c.font = hdr_font
        c.fill = hdr_fill

    r = 2
    for key, row in rows.items():
        parts = row["excel_source"].split("!") if "!" in row["excel_source"] else [row["excel_source"], "N/A"]
        for j, v in enumerate([row["category"], row["label"],
                               parts[0], parts[1] if len(parts) > 1 else "N/A",
                               "result.periods[i]", row["python_source"],
                               row["confidence"], row["note"]], 1):
            ws.cell(r, j, v)
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 55

    wb.save(path)
    print(f"Saved: {path} ({path.stat().st_size:,} bytes)")

# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading Excel fixture...")
    excel = load_excel_fixture()

    print("Running Python TUHO flag-on model...")
    result = run_python_model()

    print("Building comparison rows...")
    rows = make_rows(excel, result)

    print("Writing long CSV...")
    write_long_csv(rows, OUT_LONG_CSV)
    print(f"  {OUT_LONG_CSV} ({OUT_LONG_CSV.stat().st_size:,} bytes)")

    print("Writing wide CSV...")
    write_wide_csv(rows, OUT_WIDE_CSV)
    print(f"  {OUT_WIDE_CSV} ({OUT_WIDE_CSV.stat().st_size:,} bytes)")

    print("Writing XLSX...")
    write_xlsx(rows, OUT_XLSX)

    print("\nAll outputs written.")

if __name__ == "__main__":
    main()