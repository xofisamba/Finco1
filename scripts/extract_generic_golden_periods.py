"""G2-PERIOD-LEVEL-VALIDATION: deterministic period-level extraction for the
Generic Solar / Generic Wind bootstrap reference workbooks.

This is an additive companion to scripts/extract_generic_golden.py (which
extracts only the Tier-1 Summary anchors). It extracts the per-period rows
needed to validate the runtime period-by-period, not just at the
total/KPI level:

  - Revenue: generation (MWh), PPA tariff, market price, energy revenue,
    total revenue, by period.
  - OpEx: total OpEx by period.
  - P&L: EBITDA by period.
  - Debt Service: opening balance, interest, principal, senior debt
    service, DSCR (sizing proxy), by period.
  - Cash Flow: DSCR (true), by period.
  - Equity: distributions and total equity cash flow, by period.

Like the Tier-1 extractor, this reads only the workbook's cached formula
values via openpyxl; it does not evaluate formulas and does not import or
execute any Finco1 runtime code. Output is deterministic: re-running this
script against an unchanged workbook produces byte-for-byte identical
JSON.

Usage:
    python scripts/extract_generic_golden_periods.py
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl

EXTRACTOR_VERSION = "1.0.0"
SPEC_VERSION = "0.1"

REPO_ROOT = Path(__file__).resolve().parent.parent
REF_MODEL_DIR = REPO_ROOT / "validation" / "reference_models"
FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures"

PROJECTS = {
    "generic_solar": REF_MODEL_DIR / "GenericSolar_ReferenceModel.xlsx",
    "generic_wind": REF_MODEL_DIR / "GenericWind_ReferenceModel.xlsx",
}

# (sheet, row_label, output_key) -- row_label must match column A (or B for
# OpEx, which has a units column in B) exactly, as written in the workbook.
ROW_SPECS = [
    ("Revenue", "Generation (MWh)", "generation_mwh"),
    ("Revenue", "PPA tariff this year (EUR/MWh)", "ppa_tariff_eur_per_mwh"),
    ("Revenue", "Market price this year (EUR/MWh)", "market_price_eur_per_mwh"),
    ("Revenue", "Energy revenue (PPA or merchant, kEUR)", "energy_revenue_keur"),
    ("Revenue", "Total Revenue", "revenue_keur"),
    ("OpEx", "Total OpEx", "opex_keur"),
    ("P&L", "EBITDA", "ebitda_keur"),
    ("Debt Service", "Opening balance (kEUR)", "senior_opening_balance_keur"),
    ("Debt Service", "Interest (kEUR)", "senior_interest_keur"),
    ("Debt Service", "Senior debt service (kEUR)", "senior_debt_service_keur"),
    ("Debt Service", "Principal (kEUR)", "senior_principal_keur"),
    ("Debt Service", "Closing balance (kEUR)", "senior_closing_balance_keur"),
    ("Debt Service", "DSCR (sizing proxy)", "dscr_sizing_proxy"),
    ("Cash Flow", "DSCR (true)", "dscr_true"),
    ("Equity", "Distributions (CFADS - debt service, 100% sweep)", "distribution_keur"),
    ("Equity", "Total equity cash flow", "equity_cash_flow_keur"),
]

PERIOD_LABEL_ROW = 3  # "Period" row: P1, P2, P3, ... in row 3 of every sheet above
PERIOD_FIRST_COL = 3  # column C


def sha256_of(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _find_row(ws, label: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            if ws.cell(row, col).value == label:
                return row
    return None


def _periods_for_sheet(ws) -> list[str]:
    periods = []
    col = PERIOD_FIRST_COL
    while True:
        value = ws.cell(PERIOD_LABEL_ROW, col).value
        if value is None:
            break
        periods.append(str(value))
        col += 1
    return periods


def extract_periods(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    period_labels = _periods_for_sheet(wb["Revenue"])
    n_periods = len(period_labels)

    series: dict[str, list] = {}
    rows_found: dict[str, str] = {}
    rows_missing: list[str] = []

    for sheet_name, row_label, key in ROW_SPECS:
        ws = wb[sheet_name]
        row = _find_row(ws, row_label)
        if row is None:
            rows_missing.append(f"{sheet_name}!{row_label}")
            continue
        rows_found[key] = f"{sheet_name}!row{row}"
        values = [ws.cell(row, PERIOD_FIRST_COL + i).value for i in range(n_periods)]
        series[key] = values

    modified = wb.properties.modified or wb.properties.created
    extraction_date = modified.isoformat() if modified else None

    return {
        "period_labels": period_labels,
        "n_periods": n_periods,
        "series": series,
        "rows_found": dict(sorted(rows_found.items())),
        "rows_missing": sorted(rows_missing),
        "extraction_date": extraction_date,
    }


def build_fixture(project: str, workbook_path: Path) -> dict:
    extracted = extract_periods(workbook_path)
    return {
        "project": project,
        "source_workbook": str(workbook_path.relative_to(REPO_ROOT)),
        "workbook_sha256": sha256_of(workbook_path),
        "extraction_note": (
            "Period-level series extracted from the bootstrap reference "
            "workbook's cached formula values via openpyxl only; no formula "
            "evaluation engine or Finco1 runtime code was executed by this "
            "script. Companion to excel_golden_<project>.json (Tier-1 "
            "Summary anchors); see docs/g2_period_level_validation.md for "
            "which series are tightly validated vs. methodology-caveated."
        ),
        "extraction_date": extracted["extraction_date"],
        "spec_version": SPEC_VERSION,
        "extractor_version": EXTRACTOR_VERSION,
        "period_labels": extracted["period_labels"],
        "n_periods": extracted["n_periods"],
        "series": extracted["series"],
        "rows_found": extracted["rows_found"],
        "rows_missing": extracted["rows_missing"],
    }


def run() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    for project, workbook_path in PROJECTS.items():
        if not workbook_path.exists():
            raise FileNotFoundError(f"Reference workbook not found: {workbook_path}")
        fixture = build_fixture(project, workbook_path)
        out_path = FIXTURE_DIR / f"excel_golden_{project}_periods.json"
        out_path.write_text(json.dumps(fixture, indent=2, sort_keys=False) + "\n")
        print(f"Wrote {out_path}")


if __name__ == "__main__":
    run()
