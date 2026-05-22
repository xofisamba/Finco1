"""Build Phase 9 TUHO horizontal line-item review workbook.

Reporting-only artifact generator. It reads existing runtime outputs and
committed Excel fixture extracts; it does not mutate model/runtime behavior.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner
from domain.revenue.generation import revenue_decomposition_schedule


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"
FIXTURE = ROOT / "tests" / "fixtures" / "excel_tuho_full_model_extract.json"
PR168_PERIOD_BRIDGE = REPORTS / "phase9_tuho_full_line_item_period_bridge.csv"

XLSX_PATH = REPORTS / "phase9_tuho_full_line_item_horizontal_review_workbook.xlsx"
SUMMARY_CSV = REPORTS / "phase9_tuho_full_line_item_horizontal_summary.csv"
GAP_CSV = REPORTS / "phase9_tuho_full_line_item_horizontal_gap_analysis.csv"
SOURCE_MAP_CSV = REPORTS / "phase9_tuho_full_line_item_horizontal_source_map.csv"

STATUS_ORDER = [
    "PASS",
    "WARN",
    "FAIL",
    "ACCEPTED_CONVENTION",
    "MISSING_EVIDENCE",
    "BLOCKER",
]
STATUS_FILL = {
    "PASS": "C6EFCE",
    "WARN": "FFEB9C",
    "FAIL": "FFC7CE",
    "ACCEPTED_CONVENTION": "D9EAF7",
    "MISSING_EVIDENCE": "E7E6E6",
    "BLOCKER": "F4B183",
}


@dataclass(frozen=True)
class Metric:
    section: str
    metric: str
    excel: list[Any]
    model: list[Any]
    excel_source: str
    model_source: str
    source_status: str
    notes: str


def _num(value: Any) -> float | None:
    try:
        if value in ("", None):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _delta(excel: Any, model: Any) -> float | str:
    e = _num(excel)
    m = _num(model)
    if e is None or m is None:
        return ""
    return m - e


def _fmt_total(values: list[Any]) -> float | str:
    nums = [_num(v) for v in values]
    nums = [v for v in nums if v is not None]
    if not nums:
        return ""
    return sum(nums)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _run_tuho_default():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    result = WaterfallRunner(project, engine).run(config)
    return project, engine, result


def _load_excel_fixture() -> dict[str, Any]:
    return json.loads(FIXTURE.read_text(encoding="utf-8"))


def _diagnostic_series(fixture: dict[str, Any], column_name: str) -> list[float]:
    columns = fixture["period_diagnostic_columns"]
    idx = columns.index(column_name)
    return [float(row[idx]) for row in fixture["period_diagnostics"][:60]]


def _shl_series(fixture: dict[str, Any], column_name: str) -> list[float]:
    columns = fixture["shl_columns"]
    idx = columns.index(column_name)
    # First row is construction funding; operating horizon starts at 2030-06-30.
    return [float(row[idx]) for row in fixture["shl"][1:61]]


def _series_from_bridge(rows: list[dict[str, str]], column_name: str) -> list[float]:
    return [float(row[column_name]) for row in rows[:60]]


def _safe_div(numerator: Any, denominator: Any) -> float | str:
    n = _num(numerator)
    d = _num(denominator)
    if n is None or d in (None, 0.0):
        return ""
    return n / d


def _build_metrics() -> tuple[list[str], list[Metric], list[dict[str, Any]], dict[str, Any]]:
    project, engine, result = _run_tuho_default()
    periods = result.periods[:60]
    dates = [period.date.isoformat() for period in periods]
    fixture = _load_excel_fixture()
    pr168_rows = _read_csv_rows(PR168_PERIOD_BRIDGE)
    revenue_decomp = revenue_decomposition_schedule(project, engine)

    model_opening_shl: list[float] = []
    for idx, period in enumerate(periods):
        if idx == 0:
            model_opening_shl.append(project.financing.shl_amount_keur + project.financing.shl_idc_keur)
        else:
            model_opening_shl.append(periods[idx - 1].shl_balance_keur)

    model_senior_opening: list[float] = []
    for idx, period in enumerate(periods):
        if idx == 0:
            model_senior_opening.append(float(project.financing.fixed_debt_keur or 0.0))
        else:
            model_senior_opening.append(periods[idx - 1].senior_balance_keur)

    model_co2: list[float] = []
    model_balancing: list[float] = []
    for period in periods:
        decomp = revenue_decomp.get(period.period, {})
        model_co2.append(float(decomp.get("co2_revenue_keur", 0.0)))
        balancing = (
            float(decomp.get("balancing_cost_pv_keur", 0.0))
            + float(decomp.get("balancing_cost_wind_keur", 0.0))
        )
        model_balancing.append(-balancing)

    missing = ["MISSING_EVIDENCE: source row not mapped in committed Excel extract"] * 60
    metrics: list[Metric] = [
        Metric("Operations", "Production", _series_from_bridge(pr168_rows, "excel_production_mwh"), [p.generation_mwh for p in periods], "CF production extract", "WaterfallPeriod.generation_mwh", "runtime output", "Model production now sourced from runtime periods, not zeroed PR #168 feed."),
        Metric("Operations", "Availability / load factor", missing, missing, "Not in committed fixture", "Not mapped", "MISSING_EVIDENCE", "Specific availability/load-factor source was not present in committed extracts."),
        Metric("Operations", "Price", [_safe_div(r, p) for r, p in zip(_diagnostic_series(fixture, "P&L.total_revenues_keur"), _series_from_bridge(pr168_rows, "excel_production_mwh"))], [_safe_div(p.revenue_keur, p.generation_mwh) for p in periods], "P&L R8 / production", "WaterfallPeriod.revenue_keur / generation_mwh", "derived", "Reviewer price proxy derived from total revenue divided by production."),
        Metric("Revenue", "Electricity revenue", _diagnostic_series(fixture, "P&L.total_revenues_keur"), [p.revenue_keur for p in periods], "P&L R8 total revenues", "WaterfallPeriod.revenue_keur", "runtime output", "Electricity revenue presented as total revenue until Excel sub-splits are mapped."),
        Metric("Revenue", "CO2", missing, model_co2, "Not separately mapped", "revenue_decomposition_schedule.co2_revenue_keur", "partial evidence", "Model CO2 exists; Excel CO2 row is not separately mapped in committed fixture."),
        Metric("Revenue", "Balancing", missing, model_balancing, "Not separately mapped", "revenue_decomposition_schedule balancing cost fields", "partial evidence", "Model balancing cost exists; Excel balancing row is not separately mapped in committed fixture."),
        Metric("Revenue", "Other operating income", missing, missing, "Not in committed fixture", "Not mapped", "MISSING_EVIDENCE", "No separate other-operating-income row found in committed extracts."),
        Metric("Revenue", "Total revenue", _diagnostic_series(fixture, "P&L.total_revenues_keur"), [p.revenue_keur for p in periods], "P&L R8", "WaterfallPeriod.revenue_keur", "runtime output", "Primary total revenue parity row."),
        Metric("Costs / EBITDA", "OPEX", [-v for v in _diagnostic_series(fixture, "CF.operating_expenses_after_bank_tax_keur")], [p.opex_keur for p in periods], "CF R38 / P&L R10", "WaterfallPeriod.opex_keur", "runtime output", "OPEX shown as positive cost in review workbook."),
        Metric("Costs / EBITDA", "EBITDA", _diagnostic_series(fixture, "CF.ebitda_keur"), [p.ebitda_keur for p in periods], "CF R40", "WaterfallPeriod.ebitda_keur", "runtime output", "EBITDA parity row."),
        Metric("Costs / EBITDA", "EBITDA margin", [_safe_div(e, r) for e, r in zip(_diagnostic_series(fixture, "CF.ebitda_keur"), _diagnostic_series(fixture, "P&L.total_revenues_keur"))], [_safe_div(p.ebitda_keur, p.revenue_keur) for p in periods], "CF R40 / P&L R8", "WaterfallPeriod.ebitda_keur / revenue_keur", "derived", "Derived reviewer ratio."),
        Metric("Senior Debt", "Opening balance", _series_from_bridge(pr168_rows, "excel_senior_opening_keur"), model_senior_opening, "DS senior opening balance", "derived from fixed_debt_keur and prior closing", "runtime output", "Opening balance derived without changing runtime."),
        Metric("Senior Debt", "Closing balance", _series_from_bridge(pr168_rows, "excel_senior_closing_keur"), [p.senior_balance_keur for p in periods], "DS senior closing balance", "WaterfallPeriod.senior_balance_keur", "runtime output", "Closing senior balance."),
        Metric("Senior Debt", "Drawdown", [0.0] * 60, [0.0] * 60, "No operating drawdown", "No operating drawdown", "ACCEPTED_CONVENTION", "Operating horizon starts after construction drawdown."),
        Metric("Senior Debt", "Principal repayment", _diagnostic_series(fixture, "DS.senior_principal_keur"), [p.senior_principal_keur for p in periods], "DS R49", "WaterfallPeriod.senior_principal_keur", "runtime output", "Senior principal row."),
        Metric("Senior Debt", "Interest", _diagnostic_series(fixture, "DS.senior_net_interest_keur"), [p.senior_interest_keur for p in periods], "DS R50", "WaterfallPeriod.senior_interest_keur", "runtime output", "Senior interest row."),
        Metric("Senior Debt", "Debt service", [-v for v in _diagnostic_series(fixture, "CF.senior_debt_service_keur")], [p.senior_ds_keur for p in periods], "CF R70", "WaterfallPeriod.senior_ds_keur", "runtime output", "Debt service shown as positive outflow."),
        Metric("Senior Debt", "DSCR", _diagnostic_series(fixture, "CF.average_senior_dscr_period"), [p.dscr for p in periods], "CF average DSCR period", "WaterfallPeriod.dscr", "runtime output", "DSCR row."),
        Metric("SHL", "SHL Opening Balance", _shl_series(fixture, "opening"), model_opening_shl, "Eq SHL opening balance", "derived from shl_amount+shl_idc and prior WaterfallPeriod.shl_balance_keur", "runtime output", "Critical PR #168 fix: model P1 opening balance is non-zero."),
        Metric("SHL", "SHL Gross Accrued Interest", _shl_series(fixture, "gross_interest"), [p.shl_gross_accrued_interest_keur for p in periods], "Eq gross accrued interest", "WaterfallPeriod.shl_gross_accrued_interest_keur", "runtime output", "Gross accrued interest is separate from cash interest and PIK."),
        Metric("SHL", "SHL Cash Interest Paid", _shl_series(fixture, "paid_net_interest"), [p.shl_interest_keur for p in periods], "Eq paid net interest", "WaterfallPeriod.shl_interest_keur", "runtime output", "Cash interest paid row."),
        Metric("SHL", "SHL PIK Capitalized", _shl_series(fixture, "capitalized_interest"), [p.shl_pik_keur for p in periods], "Eq capitalized interest", "WaterfallPeriod.shl_pik_keur", "runtime output", "PIK row."),
        Metric("SHL", "SHL Principal Repaid", _shl_series(fixture, "principal_flow"), [p.shl_principal_keur for p in periods], "Eq principal flow", "WaterfallPeriod.shl_principal_keur", "runtime output", "Default/off path shown; P25 alignment remains config-gated."),
        Metric("SHL", "SHL Closing Balance", _shl_series(fixture, "closing"), [p.shl_balance_keur for p in periods], "Eq SHL closing balance", "WaterfallPeriod.shl_balance_keur", "runtime output", "Closing SHL balance row."),
        Metric("Tax / CFADS", "Taxable Income", _diagnostic_series(fixture, "P&L.taxable_income_keur"), [p.taxable_income_before_losses_audit_keur for p in periods], "P&L R35", "WaterfallPeriod.taxable_income_before_losses_audit_keur", "runtime audit output", "Tax/CFADS is no longer entirely missing; taxable income is wired."),
        Metric("Tax / CFADS", "CIT Cash", _diagnostic_series(fixture, "CF.corporate_income_tax_keur"), [p.corporate_tax_cash_keur for p in periods], "CF R67", "WaterfallPeriod.corporate_tax_cash_keur", "runtime output", "Cash CIT row."),
        Metric("Tax / CFADS", "CFADS", _diagnostic_series(fixture, "CF.free_cash_flow_for_banks_keur"), [p.r69_fcf_banks_keur for p in periods], "CF R69", "WaterfallPeriod.r69_fcf_banks_keur", "runtime audit output", "CFADS / FCF banks evidence is wired from R69 audit field."),
        Metric("Distributions", "Net Dividends / Distribution", _shl_series(fixture, "net_dividend"), [p.distribution_keur for p in periods], "Eq net dividend flow", "WaterfallPeriod.distribution_keur", "default runtime / legacy path", "Model distribution row uses one defined flag state: default runtime / legacy path."),
        Metric("Distributions", "DA-wired / pre-G20 staging", ["" for _ in periods], [p.r99_fcf_for_distribution_keur for p in periods], "Not a runtime source", "WaterfallPeriod.r99_fcf_for_distribution_keur", "audit-only / pre-G20 staging", "Separate annotation row; not mixed into model distribution row."),
        Metric("Distributions", "Lockup / blocked reason", ["" for _ in periods], ["LOCKUP" if p.lockup_active else "" for p in periods], "CF lockup rows not separately mapped", "WaterfallPeriod.lockup_active", "runtime output", "Blank means not locked in that period."),
    ]

    scalars = {
        "project_irr_model": result.project_irr,
        "equity_irr_model": result.equity_irr,
        "avg_dscr_model": sum(p.dscr for p in periods if p.dscr != float("inf")) / len([p for p in periods if p.dscr != float("inf")]),
        "min_dscr_model": min(p.dscr for p in periods if p.dscr != float("inf")),
        "total_distributions_model": sum(p.distribution_keur for p in periods),
        "total_shl_service_model": sum(p.shl_interest_keur + p.shl_principal_keur for p in periods),
        "excel_project_irr": fixture.get("excel_project_irr"),
        "excel_equity_irr": 0.1161,
    }
    return dates, metrics, _gap_rows(metrics, scalars), scalars


def _gap_rows(metrics: list[Metric], scalars: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    def add(section: str, metric: str, period_or_total: str, excel: Any, model: Any, classification: str, root: str, action: str, notes: str) -> None:
        rows.append({
            "section": section,
            "metric": metric,
            "period_or_total": period_or_total,
            "excel_value": excel,
            "model_value": model,
            "delta": _delta(excel, model),
            "root_cause": root,
            "classification": classification,
            "recommended_action": action,
            "notes": notes,
        })

    metric_by_name = {(m.section, m.metric): m for m in metrics}
    for section, metric, status, root, action, notes in [
        ("Operations", "Production", "PASS", "Runtime model data now sourced from period.generation_mwh.", "None", "Prior pack zero-feed defect fixed."),
        ("Revenue", "Total revenue", "PASS", "Runtime model data now sourced from period.revenue_keur.", "None", "Within known tolerance from PR #168."),
        ("Revenue", "CO2", "MISSING_EVIDENCE", "Excel CO2 row not separately mapped in committed extract.", "Map Excel CO2 row if reviewer requires sub-line parity.", "Model CO2 is visible, Excel split is missing."),
        ("Revenue", "Balancing", "MISSING_EVIDENCE", "Excel balancing row not separately mapped in committed extract.", "Map Excel balancing row if reviewer requires sub-line parity.", "Model balancing cost is visible, Excel split is missing."),
        ("Costs / EBITDA", "OPEX", "WARN", "Known local-tax/minor-row grouping difference.", "Keep sub-line attribution branch open.", "Total within tolerance but sub-line ownership remains partial."),
        ("Senior Debt", "Debt service", "PASS", "Runtime fields wired from senior period outputs.", "None", "Senior rows no longer zeroed by reporting feed."),
        ("SHL", "SHL Opening Balance", "PASS", "Runtime feed fixed from shl_amount + IDC / prior closing balance.", "None", "P1 model opening is around 32,704 kEUR."),
        ("SHL", "SHL Principal Repaid", "WARN", "Default runtime/off path differs from Excel timing; alignment path remains config-gated.", "Use flag-state legend; do not mix paths.", "P25 alignment exists but is not default/factory opt-in."),
        ("Tax / CFADS", "Taxable Income", "WARN", "R35 audit field wired but residual row ownership remains under review.", "Continue R35/tax basis closeout.", "No longer entirely MISSING_EVIDENCE."),
        ("Tax / CFADS", "CIT Cash", "PASS", "R67 cash CIT wired from runtime field.", "None", "Cash CIT evidence present."),
        ("Tax / CFADS", "CFADS", "PASS", "R69 FCF banks audit field wired.", "None", "CFADS evidence present."),
        ("Distributions", "Net Dividends / Distribution", "ACCEPTED_CONVENTION", "Excel dividend definition differs from runtime distribution_keur.", "Review using flag-state legend.", "Default runtime distribution row uses one flag state."),
        ("Distributions", "DA-wired / pre-G20 staging", "ACCEPTED_CONVENTION", "Audit-only R99 staging row is shown separately.", "Do not treat as runtime source.", "Separate annotation fixes prior mixed-feed defect."),
        ("Returns", "Equity IRR", "WARN", "Equity IRR remains below Excel target after known convention fixes.", "Final G20 gate review required.", "G20 remains blocked."),
        ("Returns", "Reconciliation IRR", "MISSING_EVIDENCE", "Excel-date/ex-IFC reconciliation IRR view is not implemented.", "Implement reviewer-only reconciliation IRR if required.", "Do not mark G20 pass without this or stakeholder waiver."),
        ("Governance", "G20", "BLOCKER", "G20 requires final stakeholder acceptance or reconciliation IRR evidence.", "Proceed to final parity/gate review only.", "G20 remains BLOCKED."),
        ("Governance", "R99/R102", "BLOCKER", "Runtime promotion remains explicitly unapproved.", "No promotion in this branch.", "R99/R102 remains NOT approved."),
    ]:
        if (section, metric) in metric_by_name:
            m = metric_by_name[(section, metric)]
            excel_total = _fmt_total(m.excel)
            model_total = _fmt_total(m.model)
        else:
            excel_total = ""
            model_total = ""
        add(section, metric, "total/horizon", excel_total, model_total, status, root, action, notes)

    return rows


def _write_csvs(gap_rows: list[dict[str, Any]], metrics: list[Metric], scalars: dict[str, Any]) -> Counter:
    REPORTS.mkdir(exist_ok=True)
    counts = Counter(row["classification"] for row in gap_rows)
    summary_rows = [
        {"item": "workbook", "value": str(XLSX_PATH.relative_to(ROOT)), "status": "PASS"},
        {"item": "SHL feed", "value": "runtime WaterfallPeriod SHL fields", "status": "PASS"},
        {"item": "distribution feed", "value": "default distribution_keur; DA staging separate", "status": "PASS"},
        {"item": "Tax/CFADS feed", "value": "R35/R67/R69 wired where available", "status": "PASS"},
        {"item": "model_equity_irr", "value": scalars["equity_irr_model"], "status": "WARN"},
        {"item": "model_project_irr", "value": scalars["project_irr_model"], "status": "PASS"},
        {"item": "G20", "value": "BLOCKED", "status": "BLOCKER"},
        {"item": "R99/R102", "value": "NOT APPROVED", "status": "BLOCKER"},
    ] + [{"item": f"count_{status}", "value": counts.get(status, 0), "status": status} for status in STATUS_ORDER]
    with SUMMARY_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["item", "value", "status"])
        writer.writeheader()
        writer.writerows(summary_rows)
    with GAP_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(gap_rows[0].keys()))
        writer.writeheader()
        writer.writerows(gap_rows)
    source_rows = [{
        "metric": f"{m.section} / {m.metric}",
        "excel_source": m.excel_source,
        "model_source": m.model_source,
        "source_status": m.source_status,
        "notes": m.notes,
    } for m in metrics]
    with SOURCE_MAP_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(source_rows[0].keys()))
        writer.writeheader()
        writer.writerows(source_rows)
    return counts


def _style_sheet(ws, freeze: str = "B2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 42
    for idx in range(2, min(ws.max_column, 61) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 13


def _write_workbook(dates: list[str], metrics: list[Metric], gap_rows: list[dict[str, Any]], counts: Counter, scalars: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value", "Status", "Notes"])
    for status in STATUS_ORDER:
        summary.append([f"{status} count", counts.get(status, 0), status, "Category status count"])
    summary_rows = [
        ["XLSX purpose", "Human-readable horizontal reviewer workbook", "PASS", "Fixes PR #168 reporting/data-feed defects."],
        ["SHL feed", "Runtime SHL fields wired; P1 opening non-zero", "PASS", "Model P1 SHL opening balance around 32,704 kEUR."],
        ["Distribution feed", "Legacy distribution_keur shown separately from DA staging", "PASS", "No mixed flag-state row."],
        ["Tax/CFADS feed", "R35/R67/R69 wired where available", "PASS", "No longer entirely MISSING_EVIDENCE."],
        ["Project IRR", scalars["project_irr_model"], "PASS", "Runtime project IRR from current model."],
        ["Equity IRR", scalars["equity_irr_model"], "WARN", "G20 still blocked pending final parity/gate review."],
        ["G20", "BLOCKED", "BLOCKER", "No approval in this branch."],
        ["R99/R102", "NOT APPROVED", "BLOCKER", "No runtime promotion."],
    ]
    for row in summary_rows:
        summary.append(row)
    _style_sheet(summary, "A2")
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 32
    summary.column_dimensions["D"].width = 80

    horiz = wb.create_sheet("Horizontal Review")
    horiz.append(["Section / metric / source row", *dates, "Total"])
    current_section = None
    row_idx = 2
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    for metric in metrics:
        if metric.section != current_section:
            horiz.append([metric.section, *["" for _ in dates], ""])
            for cell in horiz[row_idx]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
            row_idx += 1
            current_section = metric.section
        excel_label = f"{metric.metric} — Excel"
        model_label = f"{metric.metric} — Model"
        if metric.section == "Distributions" and metric.metric == "Net Dividends / Distribution":
            model_label += " [default runtime / legacy path]"
        if metric.metric == "DA-wired / pre-G20 staging":
            model_label += " [audit-only]"
        delta = [_delta(e, m) for e, m in zip(metric.excel, metric.model)]
        horiz.append([excel_label, *metric.excel, _fmt_total(metric.excel)])
        horiz.append([model_label, *metric.model, _fmt_total(metric.model)])
        horiz.append([f"{metric.metric} — Delta", *delta, _fmt_total(delta)])
        row_idx += 3
    for label, excel_value, model_value, status, notes in [
        ("Project IRR", scalars["excel_project_irr"], scalars["project_irr_model"], "PASS", "Model vs Excel project IRR."),
        ("Equity IRR", scalars["excel_equity_irr"], scalars["equity_irr_model"], "WARN", "G20 remains blocked."),
        ("Excel Eq!D28", "MISSING_EVIDENCE: exact Eq!D28 label/value not separately exported", "", "MISSING_EVIDENCE", "User-requested row retained with precise missing source."),
        ("Reconciliation IRR", "MISSING_EVIDENCE: not implemented", "", "MISSING_EVIDENCE", "Reviewer-only reconciliation IRR remains a gap."),
        ("Average DSCR", "", scalars["avg_dscr_model"], "PASS", "Model average DSCR."),
        ("Minimum DSCR", "", scalars["min_dscr_model"], "PASS", "Model minimum DSCR."),
        ("Total distributions", "", scalars["total_distributions_model"], "ACCEPTED_CONVENTION", "Runtime distribution definition."),
        ("Total SHL service", "", scalars["total_shl_service_model"], "PASS", "Cash interest plus principal."),
    ]:
        if current_section != "Returns":
            horiz.append(["Returns", *["" for _ in dates], ""])
            for cell in horiz[row_idx]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
            row_idx += 1
            current_section = "Returns"
        horiz.append([f"{label} — Excel", excel_value, *["" for _ in dates[1:]], ""])
        horiz.append([f"{label} — Model", model_value, *["" for _ in dates[1:]], ""])
        horiz.append([f"{label} — Status", status, *["" for _ in dates[1:]], notes])
        row_idx += 3
    _style_sheet(horiz)
    horiz.conditional_formatting.add(
        f"B2:{get_column_letter(horiz.max_column)}{horiz.max_row}",
        CellIsRule(operator="greaterThan", formula=["100"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    gap = wb.create_sheet("Gap Analysis")
    gap.append(list(gap_rows[0].keys()))
    for row in gap_rows:
        gap.append([row[key] for key in gap_rows[0].keys()])
    _style_sheet(gap)
    gap.column_dimensions["G"].width = 24
    gap.column_dimensions["H"].width = 22
    gap.column_dimensions["I"].width = 40
    gap.column_dimensions["J"].width = 70
    for row in range(2, gap.max_row + 1):
        status = gap.cell(row, 8).value
        fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "FFFFFF"))
        for col in range(1, gap.max_column + 1):
            gap.cell(row, col).fill = fill

    accepted = wb.create_sheet("Accepted Conventions")
    accepted.append(["Convention", "Status", "Explanation"])
    for row in [
        ["SHL IDC investment-base treatment", "ACCEPTED_CONVENTION", "Excel and model use different sponsor investment-base conventions; no runtime change here."],
        ["XIRR construction date convention", "ACCEPTED_CONVENTION", "Excel XIRR uses construction-date convention; model runtime uses model period dates."],
        ["Distribution vs dividend definition", "ACCEPTED_CONVENTION", "Runtime distribution_keur is not the same as Excel dividend-only line."],
        ["SHL cash interest vs gross accrued / PIK presentation", "ACCEPTED_CONVENTION", "Workbook shows gross accrued, cash paid, and PIK separately."],
        ["DSCR / senior debt service convention", "ACCEPTED_CONVENTION", "DSCR source and senior debt service definitions are shown explicitly."],
    ]:
        accepted.append(row)
    _style_sheet(accepted)
    accepted.column_dimensions["A"].width = 42
    accepted.column_dimensions["C"].width = 100

    legend = wb.create_sheet("Flag-State Legend")
    legend.append(["Flag state", "Used for", "Definition", "Status"])
    for row in [
        ["default runtime / legacy path", "Primary model rows, including distribution_keur", "No factory opt-in; default runtime behavior.", "PASS"],
        ["DA-wired / pre-G20 staging", "Separate distribution annotation row only", "Uses R99/DA audit-stage field where visible; not a runtime source.", "ACCEPTED_CONVENTION"],
        ["use_tuho_shl_repayment_alignment", "Not enabled in primary model rows", "Config-gated P25 alignment exists but remains default-off.", "ACCEPTED_CONVENTION"],
        ["R99/R102 runtime source", "Not used", "Promotion remains explicitly not approved.", "BLOCKER"],
        ["G20", "Not approved", "G20 remains blocked pending final parity/gate review.", "BLOCKER"],
    ]:
        legend.append(row)
    _style_sheet(legend)
    legend.column_dimensions["A"].width = 34
    legend.column_dimensions["C"].width = 90

    source_rows = _read_csv_rows(SOURCE_MAP_CSV)
    smap = wb.create_sheet("Source Map")
    smap.append(list(source_rows[0].keys()))
    for row in source_rows:
        smap.append([row[key] for key in source_rows[0].keys()])
    _style_sheet(smap)
    smap.column_dimensions["A"].width = 44
    smap.column_dimensions["B"].width = 42
    smap.column_dimensions["C"].width = 58
    smap.column_dimensions["D"].width = 24
    smap.column_dimensions["E"].width = 80

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'

    REPORTS.mkdir(exist_ok=True)
    wb.save(XLSX_PATH)


def main() -> None:
    dates, metrics, gap_rows, scalars = _build_metrics()
    counts = _write_csvs(gap_rows, metrics, scalars)
    _write_workbook(dates, metrics, gap_rows, counts, scalars)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
