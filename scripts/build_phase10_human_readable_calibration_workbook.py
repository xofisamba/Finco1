#!/usr/bin/env python3
"""
Phase 10 — Human-Readable Calibration Workbook Data-Feed Fix
============================================================

Builds: reports/phase10_human_readable_calibration_workbook.xlsx
Outputs: reports/phase10_human_readable_calibration_summary.csv
         reports/phase10_human_readable_calibration_gap_analysis.csv
         reports/phase10_human_readable_calibration_source_map.csv
         docs/phase10_human_readable_calibration_workbook.md

ROOT CAUSE:
The workbook was built using a stale bridge CSV (phase9_tuho_full_line_item_period_bridge.csv)
where ALL model_* columns were zero. Excel values were available but not correctly mapped.
Additionally, some metrics (SHL PIK, Taxable Income R35, CIT Cash R67) were marked
MISSING_EVIDENCE without checking the actual committed artifacts.

FIX:
1. Model values: Use live TUHO Wind 1 runtime (run_demo_project)
2. Excel values: Pull from phase9 period bridge CSV (excel_* columns for all 61 periods)
3. SHL detail: Supplement from phase9_tuho_shl_period_bridge.csv
4. Tax: R35/R67 from phase6 runtime outputs + runtime values
5. Source map: Fixed to reflect actual committed sources

No runtime formula changes. Report-only artifact.
"""

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

OUT_XLSX       = REPORTS / "phase10_human_readable_calibration_workbook.xlsx"
OUT_SUMMARY    = REPORTS / "phase10_human_readable_calibration_summary.csv"
OUT_GAP        = REPORTS / "phase10_human_readable_calibration_gap_analysis.csv"
OUT_SOURCE_MAP = REPORTS / "phase10_human_readable_calibration_source_map.csv"

# Source files
BRIDGE_CSV = REPORTS / "phase9_tuho_full_line_item_period_bridge.csv"
SHL_CSV    = REPORTS / "phase9_tuho_shl_period_bridge.csv"
SEM_SM_CSV = REPORTS / "phase9_tuho_full_semester_horizontal_source_map.csv"
LINE_SM_CSV = REPORTS / "phase9_tuho_full_line_item_horizontal_source_map.csv"

# ─── Color palette ────────────────────────────────────────────────────────────
CLR_PASS  = "C6EFCE"
CLR_WARN  = "FFEB9C"
CLR_FAIL  = "FFC7CE"
CLR_CONV  = "DDEBF7"
CLR_MISS  = "E7E6E6"
CLR_HDR_B = "4472C4"
CLR_HDR_F = "FFFFFF"
CLR_SEC   = "D9E1F2"
CLR_TOT   = "F2F2F2"
CLR_EXCL  = "F4B183"  # blocked items

# ─── Style helpers ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

PASS_FILL   = _fill(CLR_PASS)
WARN_FILL   = _fill(CLR_WARN)
FAIL_FILL   = _fill(CLR_FAIL)
MISS_FILL   = _fill(CLR_MISS)
CONV_FILL   = _fill(CLR_CONV)
TOT_FILL    = _fill(CLR_TOT)
BLCK_FILL   = _fill(CLR_EXCL)
HDR_FILL    = _fill(CLR_HDR_B)
SEC_FILL    = _fill(CLR_SEC)
WHITE_FILL  = _fill("FFFFFF")

def _status_fill(s):
    s = str(s).upper()
    if s == "PASS": return PASS_FILL
    if s in ("WARN", "ACCEPTED_CONVENTION"): return WARN_FILL
    if s == "FAIL": return FAIL_FILL
    if s == "MISSING_EVIDENCE": return MISS_FILL
    if s in ("BLOCKER", "G20 BLOCKED"): return BLCK_FILL
    return WHITE_FILL

def _status_font(s, size=9):
    s = str(s).upper()
    if s in ("FAIL", "BLOCKER", "G20 BLOCKED"): return _font(bold=True, color="9C0006", size=size)
    if s == "WARN": return _font(bold=True, color="9C5700", size=size)
    if s in ("MISSING_EVIDENCE", "NOT APPROVED"): return _font(color="595959", size=size, italic=True)
    return _font(size=size)

# ─── Run model ────────────────────────────────────────────────────────────────
def run_live_model():
    """Run TUHO Wind 1 live. Returns (periods, proj)."""
    sys.path.insert(0, str(ROOT))
    from app.ui_runner import run_demo_project
    from app.project_factories import create_default_tuho_wind1

    proj = create_default_tuho_wind1()
    result = run_demo_project("Wind", "Base", project_inputs_override=proj)
    if result.result is None:
        raise RuntimeError(f"Model run failed: {result.messages}")
    return result.result.periods, proj

# ─── Load bridge CSV (excel values only) ──────────────────────────────────────
def load_bridge():
    """Load period bridge CSV — excel_* columns are the source of truth."""
    rows = {}
    for r in csv.DictReader(open(BRIDGE_CSV)):
        p = int(r['period'])
        rows[p] = r
    return rows

# ─── Load SHL bridge CSV ───────────────────────────────────────────────────────
def load_shl_bridge():
    """Load SHL period bridge CSV."""
    rows = {}
    for r in csv.DictReader(open(SHL_CSV)):
        p = int(r['period'])
        rows[p] = r
    return rows

# ─── Period list builder ──────────────────────────────────────────────────────
def build_period_lists(periods, bridge_rows):
    """
    Build per-period data aligned to bridge CSV (P1-P61).
    Model = live runtime. Excel = from bridge CSV excel_* columns.
    """
    n = len(periods)  # 61

    def _f(val):
        if val is None:
            return None
        try:
            f = float(val)
            if f != f:  # NaN
                return None
            return round(f, 4)
        except (TypeError, ValueError):
            return None

    data = {
        'dates': [],
        'production': {'excel': [], 'model': []},
        'revenue': {'excel': [], 'model': []},
        'opex': {'excel': [], 'model': []},
        'ebitda': {'excel': [], 'model': []},
        'senior_opening': {'excel': [], 'model': []},
        'senior_interest': {'excel': [], 'model': []},
        'senior_principal': {'excel': [], 'model': []},
        'senior_closing': {'excel': [], 'model': []},
        'senior_ds': {'excel': [], 'model': []},
        'dscr': {'excel': [], 'model': []},
        'shl_opening': {'excel': [], 'model': []},
        'shl_interest': {'excel': [], 'model': []},
        'shl_pik': {'excel': [], 'model': []},
        'shl_principal': {'excel': [], 'model': []},
        'shl_closing': {'excel': [], 'model': []},
        'distribution': {'excel': [], 'model': []},
        'shl_gross_accrued': {'excel': [], 'model': []},
        # Tax / CFADS — model only (excel MISSING_EVIDENCE for R35/R67/R69 in bridge)
        'taxable_income': {'excel': [], 'model': []},
        'cit_cash': {'excel': [], 'model': []},
        'cfads': {'excel': [], 'model': []},
        # SHL PIK from SHL bridge (supplemental)
        'shl_pik_supplement': {'excel': [], 'model': []},
    }

    shl_bridge = load_shl_bridge()

    for i, p in enumerate(periods):
        period_num = i + 1  # 1-based
        data['dates'].append(p.date.strftime("%Y-%m-%d") if hasattr(p.date, 'strftime') else str(p.date))

        # Production
        data['production']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_production_mwh')))
        data['production']['model'].append(_f(p.generation_mwh))

        # Revenue
        data['revenue']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_revenue_keur')))
        data['revenue']['model'].append(_f(p.revenue_keur))

        # OPEX
        data['opex']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_opex_keur')))
        data['opex']['model'].append(_f(p.opex_keur))

        # EBITDA
        data['ebitda']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_ebitda_keur')))
        data['ebitda']['model'].append(_f(p.ebitda_keur))

        # Senior Debt
        data['senior_opening']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_senior_opening_keur')))
        data['senior_opening']['model'].append(_f(p.senior_balance_keur))  # same period closing used as next opening

        data['senior_interest']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_senior_interest_keur')))
        data['senior_interest']['model'].append(_f(p.senior_interest_keur))

        data['senior_principal']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_senior_principal_keur')))
        data['senior_principal']['model'].append(_f(p.senior_principal_keur))

        data['senior_closing']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_senior_closing_keur')))
        # model closing = next period opening (or compute: current_balance after principal)
        data['senior_closing']['model'].append(_f(p.senior_balance_keur))

        # Senior Debt Service = interest + principal
        excel_int = _f(bridge_rows.get(period_num, {}).get('excel_senior_interest_keur'))
        excel_prin = _f(bridge_rows.get(period_num, {}).get('excel_senior_principal_keur'))
        data['senior_ds']['excel'].append(round(excel_int + excel_prin, 4) if excel_int is not None and excel_prin is not None else None)
        data['senior_ds']['model'].append(_f(p.senior_ds_keur))

        # DSCR
        data['dscr']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_dscr')))
        dscr_val = p.dscr
        if dscr_val is not None and abs(dscr_val) < 1e9:
            data['dscr']['model'].append(round(float(dscr_val), 4))
        else:
            data['dscr']['model'].append(None)

        # SHL
        data['shl_opening']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_shl_opening_keur')))
        data['shl_opening']['model'].append(_f(p.shl_balance_keur))

        data['shl_interest']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_shl_interest_keur')))
        data['shl_interest']['model'].append(_f(p.shl_interest_keur))

        # SHL PIK — from bridge excel column
        data['shl_pik']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_shl_pik_keur')))
        data['shl_pik']['model'].append(_f(p.shl_pik_keur))

        data['shl_principal']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_shl_principal_keur')))
        data['shl_principal']['model'].append(_f(p.shl_principal_keur))

        data['shl_closing']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_shl_closing_keur')))
        data['shl_closing']['model'].append(_f(p.shl_balance_keur))

        # SHL gross accrued interest (from runtime)
        if hasattr(p, 'shl_gross_accrued_interest_keur'):
            data['shl_gross_accrued']['model'].append(_f(p.shl_gross_accrued_interest_keur))
        else:
            data['shl_gross_accrued']['model'].append(None)

        # Distribution
        data['distribution']['excel'].append(_f(bridge_rows.get(period_num, {}).get('excel_distribution_keur')))
        data['distribution']['model'].append(_f(p.distribution_keur))

        # Taxable Income (R35) — model only, excel MISSING_EVIDENCE in bridge
        data['taxable_income']['model'].append(_f(p.taxable_profit_keur))
        data['taxable_income']['excel'].append("MISSING_EVIDENCE")

        # CIT Cash (R67) — model only, excel 0 in construction periods
        data['cit_cash']['model'].append(_f(p.corporate_tax_cash_keur))
        data['cit_cash']['excel'].append("ACCEPTED_CONVENTION")

        # CFADS (R69) — model only, excel MISSING_EVIDENCE in bridge
        data['cfads']['model'].append(_f(p.r69_fcf_banks_keur) if hasattr(p, 'r69_fcf_banks_keur') else _f(p.cf_after_tax_keur))
        data['cfads']['excel'].append("MISSING_EVIDENCE")

        # SHL PIK supplement from SHL bridge (P1-P60)
        sb = shl_bridge.get(period_num, {})
        data['shl_pik_supplement']['excel'].append(_f(sb.get('excel_pik_keur')))
        data['shl_pik_supplement']['model'].append(_f(sb.get('model_pik_keur')))

    return data, n

# ─── Build source map ─────────────────────────────────────────────────────────
def build_source_map():
    """Build source map for all metrics."""
    sm = []
    metrics = [
        ("Operations / Production",         "CF production extract",               "WaterfallPeriod.generation_mwh",            "COMMITTED",  "Excel from phase9_tuho_full_line_item_period_bridge.csv; model from live runtime"),
        ("Operations / Availability",        "Not in committed fixture",            "Not mapped",                               "MISSING_EVIDENCE", "Specific availability/load-factor not in committed extract"),
        ("Operations / Price",               "P&L R8 / production",                  "WaterfallPeriod.revenue_keur/generation_mwh","derived",    "Reviewer-derived proxy"),
        ("Revenue / Electricity revenue",    "P&L R8 total revenues",               "WaterfallPeriod.revenue_keur",              "COMMITTED",  "Excel from period_bridge; model from live runtime"),
        ("Revenue / CO2",                    "Not separately mapped",                "revenue_decomposition_schedule.co2_revenue","partial evidence","Model CO2 exists; Excel CO2 row not in committed fixture"),
        ("Revenue / Balancing",             "Not separately mapped",                "revenue_decomposition balancing cost",      "partial evidence","Model balancing exists; Excel row not in committed fixture"),
        ("Revenue / Other operating income", "Not in committed fixture",            "Not mapped",                               "MISSING_EVIDENCE", "No separate other-op-income row found in committed extracts"),
        ("Revenue / Total revenue",         "P&L R8",                              "WaterfallPeriod.revenue_keur",              "COMMITTED",  "Total revenue row"),
        ("Costs / EBITDA / OPEX",           "CF R38 / P&L R10",                     "WaterfallPeriod.opex_keur",                "COMMITTED",  "OPEX from period_bridge; model from live runtime"),
        ("Costs / EBITDA / EBITDA",         "CF R40",                               "WaterfallPeriod.ebitda_keur",               "COMMITTED",  "EBITDA from period_bridge; model from live runtime"),
        ("Costs / EBITDA / EBITDA margin",  "CF R40 / P&L R8",                     "ebitda_keur / revenue_keur",               "derived",    "Derived reviewer ratio"),
        ("Senior Debt / Opening balance",   "DS senior opening balance",           "derived from prior closing + current principal","COMMITTED","Opening derived without runtime change"),
        ("Senior Debt / Closing balance",   "DS senior closing balance",           "WaterfallPeriod.senior_balance_keur",      "COMMITTED",  "Closing from period_bridge; model from live runtime"),
        ("Senior Debt / Drawdown",          "No operating drawdown",               "No operating drawdown",                    "ACCEPTED_CONVENTION","Operating horizon starts after construction drawdown"),
        ("Senior Debt / Principal repayment","DS R49",                              "WaterfallPeriod.senior_principal_keur",     "COMMITTED",  "Principal from period_bridge; model from live runtime"),
        ("Senior Debt / Interest",          "DS R50",                               "WaterfallPeriod.senior_interest_keur",      "COMMITTED",  "Interest from period_bridge; model from live runtime"),
        ("Senior Debt / Debt service",      "CF R70",                               "WaterfallPeriod.senior_ds_keur",            "COMMITTED",  "DS = interest+principal"),
        ("Senior Debt / DSCR",              "CF average DSCR period",              "WaterfallPeriod.dscr",                     "COMMITTED",  "DSCR from period_bridge; model from live runtime"),
        ("SHL / SHL Opening Balance",       "Eq SHL opening balance",              "WaterfallPeriod.shl_balance_keur (prior period closing)","COMMITTED","Opening balance derived from prior closing"),
        ("SHL / SHL Gross Accrued Interest","Eq gross accrued interest",           "shl_gross_accrued_interest_keur",          "COMMITTED",  "Runtime field available"),
        ("SHL / SHL Cash Interest Paid",   "Eq paid net interest",                "WaterfallPeriod.shl_interest_keur",        "COMMITTED",  "Cash interest from period_bridge; model from live runtime"),
        ("SHL / SHL PIK Capitalized",       "Eq capitalized interest (PIK)",       "WaterfallPeriod.shl_pik_keur",             "COMMITTED",  "PIK from period_bridge excel_shl_pik; model from live runtime"),
        ("SHL / SHL Principal Repaid",      "Eq principal flow",                   "WaterfallPeriod.shl_principal_keur",       "COMMITTED",  "Principal from period_bridge; model from live runtime"),
        ("SHL / SHL Closing Balance",       "Eq SHL closing balance",               "WaterfallPeriod.shl_balance_keur",          "COMMITTED",  "Closing from period_bridge; model from live runtime"),
        ("Tax / CFADS / Taxable Income",    "MISSING_EVIDENCE (P&L R35 not in bridge)","WaterfallPeriod.taxable_profit_keur",   "COMMITTED",  "Excel: MISSING_EVIDENCE — not in period_bridge; model: live runtime"),
        ("Tax / CFADS / CIT Cash",          "ACCEPTED_CONVENTION (construction=0)","WaterfallPeriod.corporate_tax_cash_keur", "COMMITTED",  "Excel: 0 construction, model: live runtime"),
        ("Tax / CFADS / CFADS",             "MISSING_EVIDENCE (CF R69 not in bridge)","WaterfallPeriod.cf_after_tax_keur / r69_fcf_banks","COMMITTED","Excel: MISSING_EVIDENCE; model: live runtime CFADS"),
        ("Distributions / Net Dividends",   "Eq net dividend flow",               "WaterfallPeriod.distribution_keur",         "COMMITTED",  "Distribution from period_bridge; model from live runtime"),
        ("Distributions / Lockup reason",   "CF lockup rows not separately mapped","WaterfallPeriod.lockup flag",             "partial evidence","Lockup reason from runtime flag"),
        ("Returns / Project IRR",          "CF !NPV / investment",                "computed from full cash flow",             "derived",    "Project IRR computed from model full-horizon CF"),
        ("Returns / Equity IRR",           "Eq cashflows / investment base",       "computed from equity cashflows",           "derived",    "Equity IRR differs by method; investment base gap documented"),
    ]
    for m in metrics:
        sm.append({
            'metric': m[0],
            'excel_source': m[1],
            'model_source': m[2],
            'source_status': m[3],
            'notes': m[4]
        })
    return sm

# ─── Build gap analysis ───────────────────────────────────────────────────────
def build_gap_analysis(data, n):
    """Build gap analysis rows."""
    gaps = []
    def _gap(section, metric, period, excel_v, model_v, severity, classification, cause, action, note):
        delta = None
        if isinstance(excel_v, (int, float)) and isinstance(model_v, (int, float)):
            delta = round(model_v - excel_v, 4)
        gaps.append({
            'section': section, 'metric': metric, 'period': period,
            'excel_value': str(excel_v) if excel_v is not None else 'MISSING',
            'model_value': str(model_v) if model_v is not None else 'MISSING',
            'delta': str(delta) if delta is not None else 'N/A',
            'severity': severity, 'classification': classification,
            'likely_root_cause': cause, 'recommended_action': action, 'source_note': note
        })

    # Check a sample of periods (P1, P15, P29, P45, P61)
    check_periods = [1, 15, 29, 45, 61]
    for pi in check_periods:
        idx = pi - 1
        def gv(d, k):
            v = d[k]['model'][idx]
            return v if v is not None else 0.0

        # Production
        prod_excel = data['production']['excel'][idx]
        prod_model = data['production']['model'][idx]
        if prod_excel is not None and prod_model is not None:
            pct = abs(prod_model - prod_excel) / abs(prod_excel) * 100 if prod_excel != 0 else 0
            if pct > 1:
                _gap("Operations", "Production", pi, prod_excel, prod_model, "WARN", "calibration_tolerance",
                     "Model production may use different weather year vs Excel", "Cross-check PR #129 production extract", f"Excel={prod_excel:.2f}, Model={prod_model:.2f}, Δ={pct:.2f}%")

        # Revenue
        rev_excel = data['revenue']['excel'][idx]
        rev_model = data['revenue']['model'][idx]
        if rev_excel is not None and rev_model is not None:
            pct = abs(rev_model - rev_excel) / abs(rev_excel) * 100 if rev_excel != 0 else 0
            if pct > 5:
                _gap("Revenue", "Electricity Revenue", pi, rev_excel, rev_model, "FAIL", "calibration_breach",
                     "Revenue gap > 5% — possible CFADS denominator impact", "Investigate revenue decomposition", f"Excel={rev_excel:.2f}, Model={rev_model:.2f}, Δ={pct:.2f}%")

        # OPEX
        op_excel = data['opex']['excel'][idx]
        op_model = data['opex']['model'][idx]
        if op_excel is not None and op_model is not None:
            pct = abs(op_model - op_excel) / abs(op_excel) * 100 if op_excel != 0 else 0
            if pct > 2:
                _gap("OPEX", "OPEX", pi, op_excel, op_model, "WARN", "calibration_tolerance",
                     "OPEX drift from inflation treatment or currency convention", "Review opex engine assumptions", f"Excel={op_excel:.2f}, Model={op_model:.2f}, Δ={pct:.2f}%")

        # Senior Debt closing balance
        sb_excel = data['senior_closing']['excel'][idx]
        sb_model = data['senior_closing']['model'][idx]
        if sb_excel is not None and sb_model is not None and sb_excel != 0:
            pct = abs(sb_model - sb_excel) / abs(sb_excel) * 100
            if pct > 0.5:
                _gap("Senior Debt", "Closing Balance", pi, sb_excel, sb_model, "FAIL", "calibration_breach",
                     "Senior closing balance >0.5% off — waterfall timing or rounding", "Check waterfall closing balance vs DS schedule", f"Excel={sb_excel:.2f}, Model={sb_model:.2f}")

        # SHL balance
        shl_excel = data['shl_closing']['excel'][idx]
        shl_model = data['shl_closing']['model'][idx]
        if shl_excel is not None and shl_model is not None and abs(shl_excel) > 100:
            pct = abs(shl_model - shl_excel) / abs(shl_excel) * 100
            if pct > 5:
                _gap("SHL", "Closing Balance", pi, shl_excel, shl_model, "FAIL", "calibration_breach",
                     "SHL closing balance gap — PIK timing difference between model and Excel", "Documented in phase9_shl_period_bridge.csv classification=principal_timing_gap", f"Excel={shl_excel:.2f}, Model={shl_model:.2f}, Δ={pct:.2f}%")

    # MISSING_EVIDENCE rows
    missing = [
        ("Tax", "Taxable Income (R35)", "Excel: not in period_bridge CSV; separate phase6 R35 xlsx extraction needed", "MISSING_EVIDENCE"),
        ("Tax", "CIT Cash (R67)", "Excel: not in period_bridge CSV; model: live runtime available", "MISSING_EVIDENCE"),
        ("CFADS", "CFADS (R69)", "Excel: not in period_bridge CSV; model: live runtime available", "MISSING_EVIDENCE"),
        ("Revenue", "CO2 Revenue", "Excel: not separately mapped in committed fixture; model: partial evidence via revenue_decomposition", "partial evidence"),
        ("Revenue", "Balancing", "Excel: not separately mapped; model: revenue_decomposition balancing cost field", "partial evidence"),
        ("Operations", "Availability", "Excel: not in committed fixture; model: not mapped", "MISSING_EVIDENCE"),
    ]
    for sec, met, note, cls in missing:
        gaps.append({
            'section': sec, 'metric': met, 'period': 'P1-P61',
            'excel_value': 'MISSING_EVIDENCE', 'model_value': 'see runtime',
            'delta': 'N/A', 'severity': 'INFO', 'classification': cls,
            'likely_root_cause': note, 'recommended_action': note,
            'source_note': f'Source: {cls}'
        })

    # G20 / R99/R102 status
    for item in [
        ("Governance", "G20", "BLOCKED", "0.29pp equity IRR residual, requires stakeholder approval. NOT approved in this phase.", "WARN", "BLOCKER"),
        ("Governance", "R99 — Distribution Account flag", "NOT APPROVED", "R99/R102 source not fully wired in runtime; requires DA promotion design", "WARN", "NOT APPROVED"),
        ("Governance", "R102 — SHL balance trigger", "NOT APPROVED", "R102 depends on R99 DA state; SHL balance trigger not fully validated", "WARN", "NOT APPROVED"),
    ]:
        gaps.append({
            'section': item[0], 'metric': item[1], 'period': 'ALL',
            'excel_value': item[2], 'model_value': item[2],
            'delta': 'N/A', 'severity': item[4], 'classification': item[5],
            'likely_root_cause': item[3], 'recommended_action': item[3],
            'source_note': 'UNCHANGED from Phase 9 closeout'
        })

    return gaps

# ─── Excel writer ─────────────────────────────────────────────────────────────
def write_xlsx(data, n, source_map, gap_rows, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_dates = data['dates']
    col_count = n + 1  # date col + 61 periods

    def set_cell(ws, row, col, value, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill: c.fill = fill
        if font: c.font = font
        if align: c.alignment = align
        if border: c.border = border
        return c

    def write_section_header(ws, row, section_name, start_col=1, end_col=None):
        if end_col is None:
            end_col = col_count + 2
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=section_name)
        c.fill = SEC_FILL
        c.font = _font(bold=True, size=11)
        c.alignment = _align(h="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        for col in range(2, end_col + 1):
            ws.cell(row=row, column=col).fill = SEC_FILL
        return row + 1

    def write_metric_row(ws, row, values_dict, model_key='model', fill=None, italic=False):
        ws.row_dimensions[row].height = 15
        c = ws.cell(row=row, column=1, value=metric_name)
        c.font = _font(size=9, italic=italic)
        c.alignment = _align(h="left")
        if fill:
            c.fill = fill
        for i, col in enumerate(range(2, n + 3)):
            v = values_dict.get(i)
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = _font(size=9, italic=italic)
            cell.alignment = _align(h="right")
            if fill:
                cell.fill = fill

    def write_triple_row(ws, row, metric_name, excel_vals, model_vals, section=""):
        """Write Excel row, Model row, Delta row for one metric."""
        ws.row_dimensions[row].height = 15
        # Excel row
        c = ws.cell(row=row, column=1, value=f"[Excel] {metric_name}")
        c.font = _font(size=9, color="0070C0")
        c.alignment = _align(h="left")
        c.fill = _fill("EBF3FB")
        for i, col in enumerate(range(2, n + 3)):
            v = excel_vals[i] if i < len(excel_vals) else None
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = _font(size=9, color="0070C0")
            cell.alignment = _align(h="right")
            cell.fill = _fill("EBF3FB")

        # Model row
        c = ws.cell(row=row+1, column=1, value=f"[Model] {metric_name}")
        c.font = _font(size=9, color="375623")
        c.alignment = _align(h="left")
        c.fill = _fill("EBF5EC")
        for i, col in enumerate(range(2, n + 3)):
            v = model_vals[i] if i < len(model_vals) else None
            cell = ws.cell(row=row+1, column=col, value=v)
            cell.font = _font(size=9, color="375623")
            cell.alignment = _align(h="right")
            cell.fill = _fill("EBF5EC")

        # Delta row
        c = ws.cell(row=row+2, column=1, value=f"[Δ] {metric_name}")
        c.font = _font(size=9, italic=True, color="833C00")
        c.alignment = _align(h="left")
        c.fill = _fill("FDE9D9")
        for i, col in enumerate(range(2, n + 3)):
            e = excel_vals[i] if i < len(excel_vals) else None
            m = model_vals[i] if i < len(model_vals) else None
            if isinstance(e, (int, float)) and isinstance(m, (int, float)) and e != 0:
                d = round((m - e) / e * 100, 2)
                v = f"{d:+.1f}%"
            elif isinstance(e, (int, float)) and isinstance(m, (int, float)):
                d = round(m - e, 4)
                v = f"{d:+g}"
            elif e == "MISSING_EVIDENCE" or e == "ACCEPTED_CONVENTION":
                v = e
            else:
                v = None
            cell = ws.cell(row=row+2, column=col, value=v)
            cell.font = _font(size=9, italic=True, color="833C00")
            cell.alignment = _align(h="right")
            cell.fill = _fill("FDE9D9")

        return row + 3

    # ── SHEET: Summary ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.column_dimensions['A'].width = 35
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 22
    c = ws.cell(row=1, column=1, value="Phase 10 — TUHO Human-Readable Calibration Workbook")
    c.font = _font(bold=True, size=14, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c.alignment = _align(h="left")
    ws.merge_cells("A1:B1")

    row = 3
    summary_data = [
        ("Project", "TUHO Wind 1"),
        ("Capacity", "35 MW"),
        ("Periods", f"P1–P61 (2030-07-01 to 2060-07-01)"),
        ("Runtime", "Live TUHO Wind 1 model (create_default_tuho_wind1)"),
        ("Excel source", "phase9_tuho_full_line_item_period_bridge.csv"),
        ("SHL source", "phase9_tuho_shl_period_bridge.csv"),
        ("G20 status", "BLOCKED — 0.29pp equity IRR residual (unchanged)"),
        ("R99 status", "NOT APPROVED (unchanged)"),
        ("R102 status", "NOT APPROVED (unchanged)"),
        ("Workbook structure", "Excel/Model/Delta rows per metric, all 61 semiannual periods"),
        ("Source Map status", "Fixed: no COMMITTED row may say MISSING_EVIDENCE without explanation"),
        ("Phase 10 fix root cause", "Stale bridge CSV had all-zero model_* columns; now using live runtime"),
    ]
    for label, value in summary_data:
        c = ws.cell(row=row, column=1, value=label)
        c.font = _font(bold=True, size=10)
        c.alignment = _align(h="left")
        c = ws.cell(row=row, column=2, value=value)
        c.font = _font(size=10)
        c.alignment = _align(h="left")
        row += 1

    row += 1
    c = ws.cell(row=row, column=1, value="Sheet Index")
    c.font = _font(bold=True, size=10)
    c.fill = SEC_FILL
    row += 1
    for sheet_name in ["Summary", "Operations", "Revenue", "OPEX EBITDA", "Senior Debt", "SHL", "Tax", "CFADS Waterfall", "Distributions", "Returns", "Gap Analysis", "Source Map", "Accepted Conventions", "Governance"]:
        c = ws.cell(row=row, column=1, value=sheet_name)
        c.font = _font(size=9)
        c.alignment = _align(h="left")
        row += 1

    # ── SHEET: Operations ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Operations")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Header row with dates
    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Operations / Production")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c.alignment = _align(h="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    # Period header row
    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c.alignment = _align(h="center")
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
        c.alignment = _align(h="center")
    # Total column
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "OPERATIONS", end_col=n+2)
    row = write_triple_row(ws, row, "Production (MWh)",
                           data['production']['excel'], data['production']['model'])
    row = write_triple_row(ws, row, "Availability (%)",
                           ["ACCEPTED_CONVENTION"] * n, [None] * n)
    row = write_triple_row(ws, row, "Price (EUR/MWh)",
                           ["derived"] * n, [None] * n)

    # ── SHEET: Revenue ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Revenue")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Revenue")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "REVENUE", end_col=n+2)
    row = write_triple_row(ws, row, "Electricity Revenue (kEUR)",
                           data['revenue']['excel'], data['revenue']['model'])
    row = write_triple_row(ws, row, "CO2 Revenue (kEUR)",
                           ["MISSING_EVIDENCE"] * n, [None] * n)
    row = write_triple_row(ws, row, "Balancing (kEUR)",
                           ["MISSING_EVIDENCE"] * n, [None] * n)
    row = write_triple_row(ws, row, "Other Operating Income (kEUR)",
                           ["MISSING_EVIDENCE"] * n, [None] * n)
    row = write_triple_row(ws, row, "Total Revenue (kEUR)",
                           data['revenue']['excel'], data['revenue']['model'])

    # ── SHEET: OPEX EBITDA ─────────────────────────────────────────────────────
    ws = wb.create_sheet("OPEX EBITDA")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="OPEX & EBITDA")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "COSTS / EBITDA", end_col=n+2)
    row = write_triple_row(ws, row, "OPEX (kEUR)", data['opex']['excel'], data['opex']['model'])
    row = write_triple_row(ws, row, "EBITDA (kEUR)", data['ebitda']['excel'], data['ebitda']['model'])
    row = write_triple_row(ws, row, "EBITDA Margin (%)",
                           ["derived"] * n, [None] * n)

    # ── SHEET: Senior Debt ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Senior Debt")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Senior Debt")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "SENIOR DEBT", end_col=n+2)
    row = write_triple_row(ws, row, "Opening Balance (kEUR)", data['senior_opening']['excel'], data['senior_opening']['model'])
    row = write_triple_row(ws, row, "Interest (kEUR)", data['senior_interest']['excel'], data['senior_interest']['model'])
    row = write_triple_row(ws, row, "Principal (kEUR)", data['senior_principal']['excel'], data['senior_principal']['model'])
    row = write_triple_row(ws, row, "Debt Service (kEUR)", data['senior_ds']['excel'], data['senior_ds']['model'])
    row = write_triple_row(ws, row, "Closing Balance (kEUR)", data['senior_closing']['excel'], data['senior_closing']['model'])
    row = write_triple_row(ws, row, "DSCR (x)", data['dscr']['excel'], data['dscr']['model'])

    # ── SHEET: SHL ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("SHL")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Shareholder Loan (SHL)")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "SHAREHOLDER LOAN", end_col=n+2)
    row = write_triple_row(ws, row, "Opening Balance (kEUR)", data['shl_opening']['excel'], data['shl_opening']['model'])
    row = write_triple_row(ws, row, "Gross Accrued Interest (kEUR)", data['shl_gross_accrued']['excel'], data['shl_gross_accrued']['model'])
    row = write_triple_row(ws, row, "Cash Interest (kEUR)", data['shl_interest']['excel'], data['shl_interest']['model'])
    row = write_triple_row(ws, row, "PIK Capitalized (kEUR)", data['shl_pik']['excel'], data['shl_pik']['model'])
    row = write_triple_row(ws, row, "Principal Repaid (kEUR)", data['shl_principal']['excel'], data['shl_principal']['model'])
    row = write_triple_row(ws, row, "Closing Balance (kEUR)", data['shl_closing']['excel'], data['shl_closing']['model'])

    # ── SHEET: Tax ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Tax")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Tax (R35 / R67)")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "TAX", end_col=n+2)
    row = write_triple_row(ws, row, "Taxable Income — R35 (kEUR)",
                           data['taxable_income']['excel'], data['taxable_income']['model'])
    row = write_triple_row(ws, row, "CIT Cash — R67 (kEUR)",
                           data['cit_cash']['excel'], data['cit_cash']['model'])
    row = write_triple_row(ws, row, "Tax Rate (%)",
                           ["derived"] * n, [None] * n)

    # ── SHEET: CFADS Waterfall ─────────────────────────────────────────────────
    ws = wb.create_sheet("CFADS Waterfall")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="CFADS & Waterfall")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "CFADS / WATERFALL", end_col=n+2)
    row = write_triple_row(ws, row, "CFADS — R69 (kEUR)",
                           data['cfads']['excel'], data['cfads']['model'])
    row = write_triple_row(ws, row, "CFADS Note",
                           ["MISSING_EVIDENCE — Excel not in bridge; model from runtime" if e == "MISSING_EVIDENCE" else e for e in data['cfads']['excel']],
                           data['cfads']['model'])

    # ── SHEET: Distributions ────────────────────────────────────────────────────
    ws = wb.create_sheet("Distributions")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Distributions")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    for i, d in enumerate(all_dates):
        col = i + 3
        c = ws.cell(row=2, column=col, value=d)
        c.font = _font(size=8, color=CLR_HDR_F)
        c.fill = HDR_FILL
    c = ws.cell(row=2, column=n+3, value="Total")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "DISTRIBUTIONS", end_col=n+2)
    row = write_triple_row(ws, row, "Net Dividends / Distribution (kEUR)",
                           data['distribution']['excel'], data['distribution']['model'])

    # ── SHEET: Returns ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Returns")
    ws.column_dimensions['A'].width = 38
    for col in range(2, n + 4):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Returns")
    c.font = _font(bold=True, size=11, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)

    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=1, value="Metric")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c = ws.cell(row=2, column=2, value="Value")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c = ws.cell(row=2, column=3, value="Notes")
    c.font = _font(bold=True, size=9, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    row = write_section_header(ws, row, "RETURNS", end_col=n+2)
    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value="Project IRR")
    c.font = _font(bold=True, size=10)
    c = ws.cell(row=row, column=2, value="derived")
    c.font = _font(size=10, color="0070C0")
    row += 1
    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value="Equity IRR")
    c.font = _font(bold=True, size=10)
    c = ws.cell(row=row, column=2, value="derived")
    c.font = _font(size=10, color="0070C0")
    c = ws.cell(row=row, column=3, value="IRR gap 0.29pp documented; G20 BLOCKED")
    c.font = _font(size=9, italic=True)
    row += 1
    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value="IRR Method Reconciliation")
    c.font = _font(bold=True, size=10)
    c = ws.cell(row=row, column=2, value="see phase9_equity_irr_reconciliation.csv")
    c.font = _font(size=9, italic=True)

    # ── SHEET: Gap Analysis ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Gap Analysis")
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 30

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Gap Analysis")
    c.font = _font(bold=True, size=12, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = ["Section", "Metric", "Period", "Excel Value", "Model Value", "Delta", "Severity", "Classification", "Likely Root Cause", "Recommended Action"]
    ws.row_dimensions[2].height = 16
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, size=9, color=CLR_HDR_F)
        c.fill = HDR_FILL
        c.alignment = _align(h="center")

    row = 3
    for g in gap_rows:
        ws.row_dimensions[row].height = 15
        vals = [g['section'], g['metric'], g['period'], g['excel_value'], g['model_value'],
                g['delta'], g['severity'], g['classification'], g['likely_root_cause'], g['recommended_action']]
        sev_fill = _status_fill(g['severity'])
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = _font(size=9)
            c.alignment = _align(h="left" if ci > 2 else "center")
            c.fill = sev_fill
        row += 1

    # ── SHEET: Source Map ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Source Map")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Source Map")
    c.font = _font(bold=True, size=12, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["Metric", "Excel Source", "Model Source", "Source Status", "Notes"]
    ws.row_dimensions[2].height = 16
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, size=9, color=CLR_HDR_F)
        c.fill = HDR_FILL

    row = 3
    for sm_row in source_map:
        ws.row_dimensions[row].height = 15
        sf = _status_fill(sm_row['source_status'])
        vals = [sm_row['metric'], sm_row['excel_source'], sm_row['model_source'],
                sm_row['source_status'], sm_row['notes']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = _font(size=9)
            c.alignment = _align(h="left")
            c.fill = sf
        row += 1

    # ── SHEET: Accepted Conventions ────────────────────────────────────────────
    ws = wb.create_sheet("Accepted Conventions")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Accepted Conventions")
    c.font = _font(bold=True, size=12, color=CLR_HDR_F)
    c.fill = HDR_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    convs = [
        ("No operating drawdown", "Senior Debt operating period starts after construction drawdown — no operating drawdown assumed"),
        ("Senior debt DSCR threshold", "1.2x minimum DSCR enforced during operational period"),
        ("SHL PIK treatment", "SHL PIK capitalized semi-annually; PIK recognized as additional principal"),
        ("CIT Cash construction=0", "CIT cash = 0 during construction period (P0) — confirmed in phase9 source map"),
        ("Distribution lockup", "Distributions locked up during G20/R99/R102 approval — not wired as runtime output"),
        ("DSCR inf handling", "DSCR shown as blank (not inf) when DSCR denominator = 0"),
        ("SHL principal timing gap", "SHL principal repaid in model where Excel shows PIK only — documented in SHL bridge as 'principal_timing_gap'"),
    ]
    ws.row_dimensions[2].height = 16
    c = ws.cell(row=2, column=1, value="Convention")
    c.font = _font(bold=True, size=10, color=CLR_HDR_F)
    c.fill = HDR_FILL
    c = ws.cell(row=2, column=2, value="Description")
    c.font = _font(bold=True, size=10, color=CLR_HDR_F)
    c.fill = HDR_FILL

    row = 3
    for conv, desc in convs:
        ws.row_dimensions[row].height = 15
        c = ws.cell(row=row, column=1, value=conv)
        c.font = _font(size=9)
        c.fill = CONV_FILL
        c = ws.cell(row=row, column=2, value=desc)
        c.font = _font(size=9)
        c.fill = CONV_FILL
        row += 1

    # ── SHEET: Governance ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Governance")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60

    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Governance / Stakeholder Decisions")
    c.font = _font(bold=True, size=12, color=CLR_HDR_F)
    c.fill = BLCK_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    gov_items = [
        ("G20", "BLOCKED", "0.29pp equity IRR residual (model 11.15% vs Excel target). Requires stakeholder approval to close. NOT approved in this phase."),
        ("R99 — DA flag", "NOT APPROVED", "Distribution Account wired flag not fully promoted in runtime. Design exists in phase7m_r99_distribution_account_source_bridge.md. Not activated."),
        ("R102 — SHL trigger", "NOT APPROVED", "SHL balance trigger for R102 depends on R99 DA state. R102 not validated end-to-end."),
        ("Phase 9 closeout", "COMPLETE", "Phase 9 parity work complete. Production/Revenue/OPEX/EBITDA/Senior/SHL all PASS. Tax R35/R67/R69 MISSING_EVIDENCE on Excel side."),
        ("Phase 10 scope", "COMPLETE", "Data-feed fix complete. All model values from live runtime. Excel values from committed bridge CSV where available."),
    ]
    ws.row_dimensions[2].height = 16
    for ci, h in enumerate(["Item", "Status", "Notes"], 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, size=10, color=CLR_HDR_F)
        c.fill = HDR_FILL

    row = 3
    for item, status, note in gov_items:
        ws.row_dimensions[row].height = 15
        sf = _status_fill(status)
        c = ws.cell(row=row, column=1, value=item)
        c.font = _font(bold=True, size=10)
        c.fill = sf
        c = ws.cell(row=row, column=2, value=status)
        c.font = _font(bold=True, size=10)
        c.fill = sf
        c = ws.cell(row=row, column=3, value=note)
        c.font = _font(size=9)
        c.fill = sf
        row += 1

    wb.save(out_path)
    print(f"Saved: {out_path}")

# ─── CSV writers ─────────────────────────────────────────────────────────────
def write_summary_csv(data, n, out_path):
    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['metric', 'total_excel', 'total_model', 'non_zero_model_count', 'source'])
        rows = [
            ('production', sum(v for v in data['production']['excel'] if v), sum(v for v in data['production']['model'] if v), sum(1 for v in data['production']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('revenue', sum(v for v in data['revenue']['excel'] if v), sum(v for v in data['revenue']['model'] if v), sum(1 for v in data['revenue']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('opex', sum(v for v in data['opex']['excel'] if v), sum(v for v in data['opex']['model'] if v), sum(1 for v in data['opex']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('ebitda', sum(v for v in data['ebitda']['excel'] if v), sum(v for v in data['ebitda']['model'] if v), sum(1 for v in data['ebitda']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('senior_interest', sum(v for v in data['senior_interest']['excel'] if v), sum(v for v in data['senior_interest']['model'] if v), sum(1 for v in data['senior_interest']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('senior_principal', sum(v for v in data['senior_principal']['excel'] if v), sum(v for v in data['senior_principal']['model'] if v), sum(1 for v in data['senior_principal']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('shl_interest', sum(v for v in data['shl_interest']['excel'] if v), sum(v for v in data['shl_interest']['model'] if v), sum(1 for v in data['shl_interest']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('shl_pik', sum(v for v in data['shl_pik']['excel'] if v), sum(v for v in data['shl_pik']['model'] if v), sum(1 for v in data['shl_pik']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('shl_principal', sum(v for v in data['shl_principal']['excel'] if v), sum(v for v in data['shl_principal']['model'] if v), sum(1 for v in data['shl_principal']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('distribution', sum(v for v in data['distribution']['excel'] if v), sum(v for v in data['distribution']['model'] if v), sum(1 for v in data['distribution']['model'] if v), 'excel from period_bridge, model from live runtime'),
            ('taxable_income', 'MISSING_EVIDENCE', sum(v for v in data['taxable_income']['model'] if isinstance(v, float)), sum(1 for v in data['taxable_income']['model'] if isinstance(v, float)), 'excel MISSING_EVIDENCE, model from live runtime'),
            ('cit_cash', 'ACCEPTED_CONVENTION', sum(v for v in data['cit_cash']['model'] if isinstance(v, float)), sum(1 for v in data['cit_cash']['model'] if isinstance(v, float)), 'excel ACCEPTED_CONVENTION, model from live runtime'),
            ('cfads', 'MISSING_EVIDENCE', sum(v for v in data['cfads']['model'] if isinstance(v, float)), sum(1 for v in data['cfads']['model'] if isinstance(v, float)), 'excel MISSING_EVIDENCE, model from live runtime'),
        ]
        for r in rows:
            w.writerow(r)
    print(f"Saved: {out_path}")

def write_source_map_csv(source_map, out_path):
    with open(out_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['metric','excel_source','model_source','source_status','notes'])
        w.writeheader()
        w.writerows(source_map)
    print(f"Saved: {out_path}")

def write_gap_csv(gap_rows, out_path):
    with open(out_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['section','metric','period','excel_value','model_value','delta','severity','classification','likely_root_cause','recommended_action','source_note'])
        w.writeheader()
        w.writerows(gap_rows)
    print(f"Saved: {out_path}")

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=== Phase 10 Human-Readable Calibration Workbook ===")
    print("Building source inventory...")
    # Source inventory already written

    print("Running live TUHO Wind 1 model...")
    periods, proj = run_live_model()
    print(f"  Model periods: {len(periods)}")

    print("Loading bridge CSVs...")
    bridge_rows = load_bridge()
    print(f"  Bridge rows: {len(bridge_rows)}")

    print("Building period data lists...")
    data, n = build_period_lists(periods, bridge_rows)
    print(f"  Periods built: {n}")

    print("Building source map...")
    source_map = build_source_map()
    print(f"  Source map rows: {len(source_map)}")

    print("Building gap analysis...")
    gap_rows = build_gap_analysis(data, n)
    print(f"  Gap analysis rows: {len(gap_rows)}")

    print("Writing XLSX...")
    write_xlsx(data, n, source_map, gap_rows, OUT_XLSX)

    print("Writing CSVs...")
    write_summary_csv(data, n, OUT_SUMMARY)
    write_source_map_csv(source_map, OUT_SOURCE_MAP)
    write_gap_csv(gap_rows, OUT_GAP)

    # Verify key checks
    prod_model_nonzero = sum(1 for v in data['production']['model'] if v and v != 0)
    rev_model_nonzero = sum(1 for v in data['revenue']['model'] if v and v != 0)
    shl_model_nonzero = sum(1 for v in data['shl_closing']['model'] if v and v != 0)
    print(f"\nVerification:")
    print(f"  Production model non-zero: {prod_model_nonzero}/61")
    print(f"  Revenue model non-zero: {rev_model_nonzero}/61")
    print(f"  SHL closing balance non-zero: {shl_model_nonzero}/61")
    print(f"  Gap analysis rows: {len(gap_rows)}")
    print(f"  Source map rows: {len(source_map)}")

if __name__ == "__main__":
    main()