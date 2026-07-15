"""
build_source_extraction.py
==========================

Stage 1 of the v4 mapping pipeline.

Reads the (untracked) agent package from a path passed in
explicitly, and produces **sanitized** source extraction JSON
files in ``docs/model_mapping/source/``.

**Stage 1 is portable** — it does not contain any hard-coded
absolute path. The package root is supplied at run time via
either:

* ``--package-root <path>`` command-line flag, or
* the ``FINCO_MAPPING_PACKAGE_ROOT`` environment variable.

The script errors out with a clear message if no package root
is supplied. After stage 1 runs once, the artifacts under
``docs/model_mapping/source/`` are the source of truth for every
subsequent step. **Stage 2 (``build_artifacts.py``) does not need
the package** — it reads only the committed, sanitized JSONs
plus the live Registry.

**Confidential source files are never committed.** This script
never copies the original XLSM, the package ZIP, the preliminary
XLSX, the preliminary PDF, or any client-confidential value into
the committed artifacts. The output is structural evidence
(coordinates, kinds, policy flags, dependencies) and the redacted
*kind* of each value (numeric / text / date / formula / empty);
it is never the original value, formula text, or scenario
sample.

Usage
-----

::

    python3 build_source_extraction.py --package-root /path/to/package
    # or
    FINCO_MAPPING_PACKAGE_ROOT=/path/to/package \
        python3 build_source_extraction.py
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

HERE = Path(__file__).resolve().parent
ARTIFACT_DIR = HERE.parent.parent
SOURCE_DIR = ARTIFACT_DIR / "source"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)


def _resolve_package_root() -> Path:
    """Resolve the package root from CLI flag, env var, or error out."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--package-root",
        type=str,
        default=None,
        help="Path to the untracked agent package directory.",
    )
    args = parser.parse_args()

    pkg_root: Optional[str] = args.package_root or os.environ.get(
        "FINCO_MAPPING_PACKAGE_ROOT"
    )
    if not pkg_root:
        sys.stderr.write(
            "ERROR: package root is required. Pass it via "
            "--package-root <path> or set the "
            "FINCO_MAPPING_PACKAGE_ROOT environment variable.\n"
            "The agent package itself is confidential and is "
            "never committed; only the sanitized output of this "
            "script is committed.\n"
        )
        sys.exit(2)

    pkg_path = Path(pkg_root).expanduser().resolve()
    if not pkg_path.is_dir():
        sys.stderr.write(
            f"ERROR: package root does not exist or is not a directory: {pkg_path}\n"
        )
        sys.exit(2)

    input_csv = pkg_path / "mapping_pack" / "Finco1_Input_Row_Map.csv"
    scenario_csv = pkg_path / "mapping_pack" / "Finco1_Scenario_Row_Map.csv"
    if not input_csv.is_file():
        sys.stderr.write(
            f"ERROR: expected file not found: {input_csv}\n"
            "The package root must point to the unzipped agent "
            "package directory (which contains mapping_pack/).\n"
        )
        sys.exit(2)
    if not scenario_csv.is_file():
        sys.stderr.write(
            f"ERROR: expected file not found: {scenario_csv}\n"
            "The package root must point to the unzipped agent "
            "package directory (which contains mapping_pack/).\n"
        )
        sys.exit(2)

    return pkg_path


PACKAGE_ROOT = _resolve_package_root()
INPUT_ROW_MAP = PACKAGE_ROOT / "mapping_pack" / "Finco1_Input_Row_Map.csv"
SCENARIO_ROW_MAP = PACKAGE_ROOT / "mapping_pack" / "Finco1_Scenario_Row_Map.csv"
SOURCE_MODELS_DIR = PACKAGE_ROOT / "source_models"

MODEL_ID_TO_SHORT = {
    "TUHO_2026_03_30": "TUHO",
    "OBOROVO_2026_04_14": "OBOROVO",
}

# Active value / formula redaction:
# we keep the *kind* (numeric / text / date / formula / empty)
# but never the original payload.

_NUMERIC_RE = re.compile(r"^[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?$")
_DATE_RE = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")
_TEXT_RE = re.compile(r"^[A-Za-z]")


def _kind_of(value: str) -> str:
    """Return the kind of a value without revealing the value itself."""
    if value is None:
        return "empty"
    s = str(value).strip()
    if s == "":
        return "empty"
    if s.startswith("="):
        return "formula"
    if _NUMERIC_RE.match(s):
        return "numeric"
    if _DATE_RE.match(s):
        return "date"
    if _TEXT_RE.match(s):
        return "text"
    return "other"


def _redact(value: str) -> Dict[str, str]:
    """Return a redacted form of a value: only the kind, never the value."""
    return {"kind": _kind_of(value)}


def _load_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    return rows


def _split_cells(value: str) -> List[str]:
    if not value:
        return []
    cells: List[str] = []
    for raw in str(value).split(","):
        token = raw.strip()
        if not token:
            continue
        if ":" in token:
            try:
                from openpyxl.utils.cell import get_column_letter, range_boundaries

                min_col, min_row, max_col, max_row = range_boundaries(token)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cells.append(f"{get_column_letter(col)}{row}")
            except ValueError:
                cells.append(token)
        else:
            cells.append(token)
    return cells


def _merge_cells(*groups: List[str]) -> str:
    cells: List[str] = []
    for group in groups:
        cells.extend(group)
    return ", ".join(cells)


def _open_source_workbooks() -> Dict[str, Any]:
    """Open source workbooks read-only and keep only storage-kind evidence."""
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write(
            "ERROR: openpyxl is required for workbook storage verification.\n"
        )
        sys.exit(2)

    if not SOURCE_MODELS_DIR.is_dir():
        sys.stderr.write(
            "ERROR: source workbook directory not found under the supplied package root.\n"
        )
        sys.exit(2)

    paths = list(SOURCE_MODELS_DIR.glob("*." + "xlsm"))
    by_model: Dict[str, Path] = {}
    for path in paths:
        name = path.name.upper()
        if "TUHO" in name:
            by_model["TUHO"] = path
        elif "OBOROVO" in name:
            by_model["OBOROVO"] = path

    missing = sorted(set(MODEL_ID_TO_SHORT.values()) - set(by_model))
    if missing:
        sys.stderr.write(
            "ERROR: expected source workbook(s) were not found for model(s): "
            + ", ".join(missing)
            + "\n"
        )
        sys.exit(2)

    return {
        model: openpyxl.load_workbook(path, data_only=False, read_only=False)
        for model, path in by_model.items()
    }


SOURCE_WORKBOOKS = _open_source_workbooks()


def _storage_kind(model_id: str, sheet_name: str, cell: str) -> str:
    wb = SOURCE_WORKBOOKS[model_id]
    if sheet_name not in wb.sheetnames or not cell or not re.match(r"^[A-Z]+[0-9]+$", cell):
        return "empty"
    value = wb[sheet_name][cell].value
    if value is None:
        return "empty"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "hardcode"


def _storage_fields(model_id: str, sheet_name: str, cells: List[str]) -> Dict[str, str]:
    hard: List[str] = []
    formula: List[str] = []
    empty: List[str] = []
    label: List[str] = []
    for cell in cells:
        kind = _storage_kind(model_id, sheet_name, cell)
        if kind == "formula":
            formula.append(cell)
        elif kind == "hardcode":
            hard.append(cell)
            if not re.match(r"^[A-Z]+[0-9]+$", cell):
                label.append(cell)
        else:
            empty.append(cell)
    return {
        "verified_hardcode_cells": _merge_cells(hard),
        "verified_formula_cells": _merge_cells(formula),
        "verified_empty_cells": _merge_cells(empty),
        "verified_label_or_presentation_cells": _merge_cells(label),
    }


def _add_label_cells(fields: Dict[str, str], model_id: str, sheet_name: str, cells: List[str]) -> None:
    labels = _split_cells(fields.get("verified_label_or_presentation_cells", ""))
    for cell in cells:
        if _storage_kind(model_id, sheet_name, cell) in {"hardcode", "formula"}:
            labels.append(cell)
    fields["verified_label_or_presentation_cells"] = _merge_cells(labels)


def _add_formula_cells(fields: Dict[str, str], model_id: str, sheet_name: str, cells: List[str]) -> None:
    formulas = _split_cells(fields.get("verified_formula_cells", ""))
    for cell in cells:
        if _storage_kind(model_id, sheet_name, cell) == "formula":
            formulas.append(cell)
    fields["verified_formula_cells"] = _merge_cells(formulas)


def _active_storage_kind(model_id: str, sheet_name: str, active_cell: str) -> str:
    kind = _storage_kind(model_id, sheet_name, active_cell)
    return {"formula": "FORMULA", "hardcode": "HARDCODE", "empty": "EMPTY"}[kind]


def _active_formula_kind(model_id: str, sheet_name: str, active_cell: str) -> str:
    return "formula" if _storage_kind(model_id, sheet_name, active_cell) == "formula" else "empty"


# ---------------------------------------------------------------------------
# Inputs source
# ---------------------------------------------------------------------------


def _build_inputs_source() -> Dict[str, Any]:
    raw = _load_csv(INPUT_ROW_MAP)
    out_rows: List[Dict[str, Any]] = []
    for r in raw:
        mid = r["model_id"]
        if mid not in MODEL_ID_TO_SHORT:
            continue
        model_id = MODEL_ID_TO_SHORT[mid]
        sheet = r.get("sheet", "Inputs")
        active_cell = r.get("primary_cell", "")
        package_hardcodes = r.get("hardcode_cells", "")
        package_formulas = r.get("formula_cells", "")
        label_cells = _split_cells(r.get("cell_span", ""))
        candidates = (
            ([active_cell] if active_cell else [])
            + _split_cells(package_hardcodes)
            + _split_cells(package_formulas)
            + label_cells
        )
        verified = _storage_fields(model_id, sheet, candidates)
        _add_label_cells(verified, model_id, sheet, label_cells)
        if r.get("field_id_candidate", "") == "reserves.dsra":
            if model_id == "TUHO":
                _add_formula_cells(verified, model_id, sheet, ["D330", "D331", "D332"])
            elif model_id == "OBOROVO":
                _add_formula_cells(verified, model_id, sheet, ["D347", "D348", "D349"])

        out_rows.append({
            "model_id": MODEL_ID_TO_SHORT[mid],
            "source_model_id": mid,
            "sheet": sheet,
            "row": int(r["row"]),
            "section": r.get("section", ""),
            "owner": r.get("owner", ""),
            "team": r.get("team", ""),
            "domain": r.get("domain", ""),
            "technology": r.get("technology", ""),
            "label": r.get("label", ""),
            "field_id_candidate": r.get("field_id_candidate", ""),
            "active_cell": active_cell,
            "active_value_kind": _kind_of(r.get("primary_value", "")),
            "active_formula_kind": _active_formula_kind(model_id, sheet, active_cell),
            "kind": r.get("kind", ""),
            "source_type": r.get("source_type", ""),
            "dependencies": r.get("dependencies", ""),
            "cell_span": r.get("cell_span", ""),
            "formula_cells": verified["verified_formula_cells"],
            "hardcode_cells": verified["verified_hardcode_cells"],
            "data_type": r.get("data_type", ""),
            "unit": r.get("unit", ""),
            "editable_policy": r.get("editable_policy", ""),
            "scenario_policy": r.get("scenario_policy", ""),
            "notes": r.get("notes", ""),
            "package_claim_hardcode_cells": package_hardcodes,
            "package_claim_formula_cells": package_formulas,
            "package_claim_active_formula_kind": _kind_of(r.get("primary_formula", "")),
            **verified,
            "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
            "verified_active_cell_storage_kind": _active_storage_kind(model_id, sheet, active_cell),
        })
    out_rows.sort(key=lambda r: (r["model_id"], r["row"]))
    return {
        "source_artifact": "Finco1_Input_Row_Map.csv (sanitized)",
        "row_count": len(out_rows),
        "models": sorted(set(r["model_id"] for r in out_rows)),
        "rows": out_rows,
    }


# ---------------------------------------------------------------------------
# Scenarios source
# ---------------------------------------------------------------------------


def _build_scenarios_source() -> Dict[str, Any]:
    raw = _load_csv(SCENARIO_ROW_MAP)
    out_rows: List[Dict[str, Any]] = []
    for r in raw:
        mid = r["model_id"]
        if mid not in MODEL_ID_TO_SHORT:
            continue
        model_id = MODEL_ID_TO_SHORT[mid]
        sheet = r.get("sheet", "Scenarios")
        active_cell = r.get("active_cell", "")
        package_formula_kind = _kind_of(r.get("active_formula", ""))
        scenario_cells = r.get("scenario_cells", "")
        candidates = ([active_cell] if active_cell else []) + _split_cells(scenario_cells)
        verified = _storage_fields(model_id, sheet, candidates)

        out_rows.append({
            "model_id": MODEL_ID_TO_SHORT[mid],
            "source_model_id": mid,
            "sheet": sheet,
            "row": int(r["row"]),
            "section": r.get("section", ""),
            "owner": r.get("owner", ""),
            "team": r.get("team", ""),
            "domain": r.get("domain", ""),
            "technology": r.get("technology", ""),
            "label": r.get("label", ""),
            "field_id_candidate": r.get("field_id_candidate", ""),
            "active_cell": active_cell,
            "active_value_kind": _kind_of(r.get("active_value", "")),
            "active_formula_kind": _active_formula_kind(model_id, sheet, active_cell),
            "scenario_cells": scenario_cells,
            "scenario_value_kind": _kind_of(r.get("scenario_values_sample", "")),
            "scenario_formula_kind": _kind_of(r.get("scenario_formula_sample", "")),
            "source_type": r.get("source_type", ""),
            "dependencies": r.get("dependencies", ""),
            "data_type": r.get("data_type", ""),
            "unit": r.get("unit", ""),
            "override_policy": r.get("override_policy", ""),
            "notes": r.get("notes", ""),
            "package_claim_hardcode_cells": "",
            "package_claim_formula_cells": scenario_cells,
            "package_claim_active_formula_kind": package_formula_kind,
            **verified,
            "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
            "verified_active_cell_storage_kind": _active_storage_kind(model_id, sheet, active_cell),
        })
    out_rows.sort(key=lambda r: (r["model_id"], r["row"]))
    return {
        "source_artifact": "Finco1_Scenario_Row_Map.csv (sanitized)",
        "row_count": len(out_rows),
        "models": sorted(set(r["model_id"] for r in out_rows)),
        "rows": out_rows,
    }


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------


def main() -> int:
    sys.stderr.write(
        f"[build_source_extraction] package_root = {PACKAGE_ROOT}\n"
    )
    inputs = _build_inputs_source()
    scenarios = _build_scenarios_source()

    (SOURCE_DIR / "tuho_inputs_source_v2.json").write_text(
        json.dumps(
            {
                "source_artifact": inputs["source_artifact"],
                "model_id": "TUHO",
                "source_model_id": "TUHO_2026_03_30",
                "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
                "row_count": sum(1 for r in inputs["rows"] if r["model_id"] == "TUHO"),
                "rows": [r for r in inputs["rows"] if r["model_id"] == "TUHO"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (SOURCE_DIR / "oborovo_inputs_source_v2.json").write_text(
        json.dumps(
            {
                "source_artifact": inputs["source_artifact"],
                "model_id": "OBOROVO",
                "source_model_id": "OBOROVO_2026_04_14",
                "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
                "row_count": sum(1 for r in inputs["rows"] if r["model_id"] == "OBOROVO"),
                "rows": [r for r in inputs["rows"] if r["model_id"] == "OBOROVO"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (SOURCE_DIR / "tuho_scenarios_source_v2.json").write_text(
        json.dumps(
            {
                "source_artifact": scenarios["source_artifact"],
                "model_id": "TUHO",
                "source_model_id": "TUHO_2026_03_30",
                "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
                "row_count": sum(1 for r in scenarios["rows"] if r["model_id"] == "TUHO"),
                "rows": [r for r in scenarios["rows"] if r["model_id"] == "TUHO"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (SOURCE_DIR / "oborovo_scenarios_source_v2.json").write_text(
        json.dumps(
            {
                "source_artifact": scenarios["source_artifact"],
                "model_id": "OBOROVO",
                "source_model_id": "OBOROVO_2026_04_14",
                "storage_verification_basis": "PROGRAMMATIC_WORKBOOK_INSPECTION",
                "row_count": sum(1 for r in scenarios["rows"] if r["model_id"] == "OBOROVO"),
                "rows": [r for r in scenarios["rows"] if r["model_id"] == "OBOROVO"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    sys.stderr.write(
        "[build_source_extraction] wrote:\n"
        f"  tuho_inputs_source_v2.json         ({sum(1 for r in inputs['rows'] if r['model_id']=='TUHO')} rows)\n"
        f"  oborovo_inputs_source_v2.json      ({sum(1 for r in inputs['rows'] if r['model_id']=='OBOROVO')} rows)\n"
        f"  tuho_scenarios_source_v2.json      ({sum(1 for r in scenarios['rows'] if r['model_id']=='TUHO')} rows)\n"
        f"  oborovo_scenarios_source_v2.json   ({sum(1 for r in scenarios['rows'] if r['model_id']=='OBOROVO')} rows)\n"
    )
    sys.stderr.write(
        "\n[build_source_extraction] REMINDER: confidential source files "
        "(XLSM, ZIP, preliminary XLSX/PDF) must NEVER be committed. Only "
        "the sanitized JSONs above are committed.\n"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
