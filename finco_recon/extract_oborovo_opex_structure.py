"""finco_recon.extract_oborovo_opex_structure — Deterministic OPEX structural truth extraction.

Reads the authoritative Oborovo workbook (SHA256-verified) TWICE:
  A) data_only=True  — cached Excel values (formula results)
  B) data_only=False — exact Excel formula strings

Emits tests/fixtures/excel_oborovo_opex_structural_truth.json.

This script is SOURCE TRUTH ONLY.  It performs zero financial computation
and changes no runtime code.  Run it only when the authoritative workbook
is updated and the fixture must be refreshed.

Usage:
    python3 -m finco_recon.extract_oborovo_opex_structure \\
        --workbook <path/to/workbook.xlsm> \\
        [--out tests/fixtures/excel_oborovo_opex_structural_truth.json]

Exit codes:
    0 — fixture written (or --dry-run: printed to stdout).
    1 — SHA256 mismatch or structural assertion failure.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

EXPECTED_SHA256 = "15a621c4d6b79024980766e00ebc79d7235fd56f00567be7bf345c769ce57920"

# ---------------------------------------------------------------------------
# OpEx sheet column indices (1-based)
# ---------------------------------------------------------------------------
_COL_CODE = 1        # A
_COL_NAME = 2        # B
_COL_BUDGET = 3      # C
_COL_INFLATION = 4   # D  (contingency rate for B.13)
_COL_WHT = 5         # E
_COL_Y1 = 6          # F
_COL_Y30 = 35        # AE
_N_YEARS = 30

# Shared CPI reference cell on OpEx sheet
_OPEX_CPI_REF_ROW = 102   # C102 = "=Inputs!D85"

# Category header rows (code, row); subitem rows are auto-discovered as rows
# between consecutive category headers.  B.13 has no subitems (formula category).
_CAT_HEADER_ROWS: list[tuple[str, int]] = [
    ("B.01", 3),
    ("B.02", 8),
    ("B.03", 26),
    ("B.04", 31),
    ("B.05", 35),
    ("B.06", 39),
    ("B.07", 45),
    ("B.08", 48),
    ("B.09", 53),
    ("B.10", 58),
    ("B.11", 65),
    ("B.12", 70),
    ("B.13", 76),
]
# Row after B.13 where Claims section begins (not a B.xx category)
_FIRST_NON_OPEX_ROW = 77

# Totals rows
_ROW_TOTAL_EXCL = 95
_ROW_TOTAL_INCL = 96

# Inputs cells (row, col)
_INPUTS_D85  = (85,  4)
_INPUTS_D93  = (93,  4)
_INPUTS_D196 = (196, 4)
_INPUTS_D259 = (259, 4)

# Scenarios sheet scenario-value columns (col_index → column_letter)
_SCEN_VALUE_COLS = {8: "H", 9: "I", 10: "J", 11: "K"}
_SCEN_NAME_ROW   = 3
_SCEN_INDEX_ROW  = 4
_SCEN_SELECTOR_COL = 5   # E4 = currently selected scenario index


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _formula_text(ws_f, row: int, col: int) -> str | None:
    """Return the Excel formula string if the cell contains one, else None."""
    v = ws_f.cell(row=row, column=col).value
    if isinstance(v, ArrayFormula):
        return v.text
    if isinstance(v, str) and v.startswith("="):
        return v
    return None


def _cell_formula_or_scalar(ws_f, row: int, col: int):
    v = ws_f.cell(row=row, column=col).value
    if isinstance(v, ArrayFormula):
        return v.text
    return v


def _annual_cached(ws_v, row: int) -> list:
    return [ws_v.cell(row=row, column=c).value for c in range(_COL_Y1, _COL_Y30 + 1)]


def _annual_formulas(ws_f, row: int) -> list:
    result = []
    for c in range(_COL_Y1, _COL_Y30 + 1):
        v = ws_f.cell(row=row, column=c).value
        result.append(v.text if isinstance(v, ArrayFormula) else v)
    return result


def _budget_rec(ws_v, ws_f, row: int) -> dict:
    cached = ws_v.cell(row=row, column=_COL_BUDGET).value
    ftext  = _formula_text(ws_f, row, _COL_BUDGET)
    rec = {"formula_cell": f"OpEx!C{row}", "cached_value": cached}
    if ftext:
        rec["formula_text"] = ftext
    return rec


def _inflation_rec(ws_v, ws_f, row: int) -> dict:
    cached = ws_v.cell(row=row, column=_COL_INFLATION).value
    ftext  = _formula_text(ws_f, row, _COL_INFLATION)
    rec = {"formula_cell": f"OpEx!D{row}", "cached_value": cached}
    if ftext:
        rec["formula_text"] = ftext
    return rec


# ---------------------------------------------------------------------------
# Scenarios sheet helpers
# ---------------------------------------------------------------------------

def _read_scenario_metadata(ws_scen_v, ws_scen_f) -> tuple[dict, str | None, int | None]:
    """Return (columns_dict, selected_column_letter, selected_index)."""
    columns: dict[str, dict] = {}
    for col_idx, letter in _SCEN_VALUE_COLS.items():
        name = ws_scen_v.cell(_SCEN_NAME_ROW,  col_idx).value
        idx  = ws_scen_v.cell(_SCEN_INDEX_ROW, col_idx).value
        columns[letter] = {"index": idx, "name": name}

    selected_index = ws_scen_v.cell(_SCEN_INDEX_ROW, _SCEN_SELECTOR_COL).value
    selected_col   = next(
        (l for l, info in columns.items() if info["index"] == selected_index),
        None,
    )
    return columns, selected_col, selected_index


def _scen_row_lineage(
    ws_scen_v,
    ws_scen_f,
    scen_row: int,
    columns: dict,
    selected_col: str | None,
) -> dict:
    """Build Scenarios sheet lineage for one opex subitem row."""
    label   = ws_scen_v.cell(scen_row, 3).value
    e_val   = ws_scen_v.cell(scen_row, _SCEN_SELECTOR_COL).value
    e_f     = _cell_formula_or_scalar(ws_scen_f, scen_row, _SCEN_SELECTOR_COL)

    per_scenario: dict = {}
    for col_idx, letter in _SCEN_VALUE_COLS.items():
        v = ws_scen_v.cell(scen_row, col_idx).value
        f = _cell_formula_or_scalar(ws_scen_f, scen_row, col_idx)
        entry: dict = {"cached_value": v}
        if isinstance(f, str) and f.startswith("="):
            entry["formula_text"] = f
        per_scenario[letter] = entry

    rec: dict = {
        "scenarios_row": scen_row,
        "label": label,
        "cached_value": e_val,
        "per_scenario_values": per_scenario,
    }
    if isinstance(e_f, str) and e_f.startswith("="):
        rec["selector_formula"] = e_f
    if selected_col:
        rec["selected_column"] = selected_col
    return rec


# ---------------------------------------------------------------------------
# B.13 contingency block
# ---------------------------------------------------------------------------

def _build_b13_contingency(ws_v, ws_f, b13_row: int, cat_headers: list[tuple[str, int]]) -> dict:
    budget_ftext  = _formula_text(ws_f, b13_row, _COL_BUDGET)
    annual_ftext  = _formula_text(ws_f, b13_row, _COL_Y1)
    rate_cached   = ws_v.cell(b13_row, _COL_INFLATION).value  # D76 = 0.04

    # Parse base rows from formula =SUM(C82,C90,...)*$D$76
    base_rows_sorted = sorted(int(r) for r in re.findall(r"C(\d+)", budget_ftext or ""))

    # Look up code+name for each base row
    base_entries = []
    for r in base_rows_sorted:
        code = ws_v.cell(r, _COL_CODE).value
        name = ws_v.cell(r, _COL_NAME).value
        base_entries.append({"opex_row": r, "code": code, "name": name})

    base_codes = [e["code"] for e in base_entries]

    # Excluded section (Claims = C, rows 78-81)
    excluded = []
    for r in range(78, 82):
        code = ws_v.cell(r, _COL_CODE).value
        name = ws_v.cell(r, _COL_NAME).value
        if code or name:
            excluded.append({"opex_row": r, "code": code, "name": name})

    return {
        "calculation_type": "PERCENTAGE_OF_SELECTED_CATEGORIES",
        "contingency_rate": rate_cached,
        "rate_cell": f"OpEx!D{b13_row}",
        "budget_formula_cell": f"OpEx!C{b13_row}",
        "budget_formula_text": budget_ftext,
        "annual_formula_template_y1": annual_ftext,
        "contingency_base_rows": base_rows_sorted,
        "contingency_base_codes": base_codes,
        "contingency_base": base_entries,
        "excluded_categories": excluded,
    }


# ---------------------------------------------------------------------------
# Label / flag mismatch detection
# ---------------------------------------------------------------------------

def _detect_label_flag_mismatch(si_name: str | None, flags: list) -> str | None:
    """Return a mismatch note if the workbook label disagrees with actual flags."""
    if not si_name:
        return None
    active_years = [i + 1 for i, v in enumerate(flags) if v == 1]
    if "Y1-2" in si_name and active_years and active_years != [1, 2]:
        return (
            f"WORKBOOK LABEL / YEAR-INDEX SEMANTIC MISMATCH: "
            f"label says 'Y1-2' but actual active years: {active_years}"
        )
    if "Y3-30" in si_name and active_years and active_years[0] != 3:
        return (
            f"WORKBOOK LABEL / YEAR-INDEX SEMANTIC MISMATCH: "
            f"label says 'Y3-30' but actual active from Y{active_years[0]} "
            f"({len(active_years)} years: {active_years})"
        )
    return None


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------

def extract(workbook_path: Path) -> dict:
    """Extract OPEX structure from workbook.  Raises ValueError on SHA256 mismatch."""
    digest = _sha256(workbook_path)
    if digest != EXPECTED_SHA256:
        raise ValueError(
            f"SHA256 mismatch.\n"
            f"  expected: {EXPECTED_SHA256}\n"
            f"  actual:   {digest}\n"
            "Update EXPECTED_SHA256 only after confirming the new file is authoritative."
        )

    wb_v = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    wb_f = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)

    ws_opex_v   = wb_v["OpEx"]
    ws_opex_f   = wb_f["OpEx"]
    ws_inputs_v = wb_v["Inputs"]
    ws_inputs_f = wb_f["Inputs"]
    ws_scen_v   = wb_v["Scenarios"]
    ws_scen_f   = wb_f["Scenarios"]

    # ------------------------------------------------------------------
    # Shared parameters
    # ------------------------------------------------------------------
    cpi_opex_ref   = _formula_text(ws_opex_f, _OPEX_CPI_REF_ROW, _COL_BUDGET)
    d85_formula    = _formula_text(ws_inputs_f, *_INPUTS_D85)
    d93_value      = ws_inputs_v.cell(*_INPUTS_D93).value
    inflation_cached = ws_inputs_v.cell(*_INPUTS_D85).value

    d196_formula   = _formula_text(ws_inputs_f, *_INPUTS_D196)
    debt_tenor_1   = ws_inputs_v.cell(*_INPUTS_D196).value
    debt_tenor_2   = ws_inputs_v.cell(*_INPUTS_D259).value

    b11_3_y1_formula = _formula_text(ws_opex_f, 68, _COL_Y1)

    inputs_block = {
        "inflation_rate": {
            "formula_cell": "OpEx!C102",
            "formula_text": cpi_opex_ref,
            "chain": {
                "Inputs!D85": d85_formula,
                "Inputs!D93": d93_value,
            },
            "cached_value": inflation_cached,
        },
        "senior_debt_tenor": {
            "formula_cell": "Inputs!D196",
            "formula_text": d196_formula,
            "cached_value": debt_tenor_1,
        },
        "second_debt_tenor": {
            "formula_cell": "Inputs!D259",
            "cached_value": debt_tenor_2,
        },
        "b11_3_activation_formula": {
            "formula_cell": "OpEx!F68",
            "formula_text": b11_3_y1_formula,
            "driver_cell": "Inputs!D196",
            "driver_cached_value": debt_tenor_1,
        },
    }

    # ------------------------------------------------------------------
    # Scenarios metadata
    # ------------------------------------------------------------------
    scen_columns, selected_col, selected_index = _read_scenario_metadata(ws_scen_v, ws_scen_f)
    scenarios_meta = {
        "selector_cell": "Scenarios!E4",
        "selected_index": selected_index,
        "selected_column": selected_col,
        "selection_mechanism": "=INDEX(H<row>:K<row>,0,MATCH($E$4,$H$4:$K$4,0))",
        "columns": scen_columns,
    }

    # ------------------------------------------------------------------
    # Categories
    # ------------------------------------------------------------------
    categories: dict[str, dict] = {}
    cat_rows_list = _CAT_HEADER_ROWS

    for i, (cat_code, cat_row) in enumerate(cat_rows_list):
        next_cat_row = (
            cat_rows_list[i + 1][1] if i + 1 < len(cat_rows_list)
            else _FIRST_NON_OPEX_ROW
        )

        cat_entry: dict = {
            "code": cat_code,
            "opex_row": cat_row,
            "name": ws_opex_v.cell(cat_row, _COL_NAME).value,
            "budget": _budget_rec(ws_opex_v, ws_opex_f, cat_row),
            "inflation": _inflation_rec(ws_opex_v, ws_opex_f, cat_row),
            "wht_cached": ws_opex_v.cell(cat_row, _COL_WHT).value,
            "annual": {
                "formula_template_y1": _formula_text(ws_opex_f, cat_row, _COL_Y1),
                "formula_template_y2": _formula_text(ws_opex_f, cat_row, _COL_Y1 + 1),
                "cached_values_y1_y30": _annual_cached(ws_opex_v, cat_row),
            },
        }

        if cat_code == "B.13":
            cat_entry["contingency"] = _build_b13_contingency(
                ws_opex_v, ws_opex_f, cat_row, cat_rows_list
            )
        else:
            subitems: dict[str, dict] = {}
            for si_row in range(cat_row + 1, next_cat_row):
                si_code     = ws_opex_v.cell(si_row, _COL_CODE).value
                si_name     = ws_opex_v.cell(si_row, _COL_NAME).value
                si_bud_cached = ws_opex_v.cell(si_row, _COL_BUDGET).value
                flags_cached  = _annual_cached(ws_opex_v, si_row)
                flags_formula = _annual_formulas(ws_opex_f, si_row)

                # Skip rows that are entirely empty (no code, name, budget, or flags)
                if (si_code is None and si_name is None and si_bud_cached is None
                        and all(v is None for v in flags_cached)):
                    continue

                si_bud_formula = _formula_text(ws_opex_f, si_row, _COL_BUDGET)

                bud_rec: dict = {
                    "formula_cell": f"OpEx!C{si_row}",
                    "cached_value": si_bud_cached,
                }
                if si_bud_formula:
                    bud_rec["formula_text"] = si_bud_formula
                    m = re.match(r"=Scenarios!E(\d+)$", si_bud_formula)
                    if m:
                        bud_rec["scenarios_lineage"] = _scen_row_lineage(
                            ws_scen_v, ws_scen_f,
                            int(m.group(1)),
                            scen_columns, selected_col,
                        )

                has_formula_flags = any(
                    isinstance(f, str) and f.startswith("=")
                    for f in flags_formula
                )

                si_entry: dict = {
                    "opex_row": si_row,
                    "code": si_code,
                    "name": si_name,
                    "budget": bud_rec,
                    "activation_cached_y1_y30": flags_cached,
                }
                if has_formula_flags:
                    si_entry["activation_formula_y1"] = flags_formula[0]

                mismatch = _detect_label_flag_mismatch(si_name, flags_cached)
                if mismatch:
                    si_entry["label_flag_mismatch"] = mismatch

                dict_key = si_code if si_code is not None else f"row_{si_row}"
                subitems[dict_key] = si_entry

            cat_entry["subitems"] = subitems

        categories[cat_code] = cat_entry

    # ------------------------------------------------------------------
    # Totals
    # ------------------------------------------------------------------
    totals = {
        "total_opex_excl_contingencies": {
            "opex_row": _ROW_TOTAL_EXCL,
            "budget_cached": ws_opex_v.cell(_ROW_TOTAL_EXCL, _COL_BUDGET).value,
            "y1_cached": ws_opex_v.cell(_ROW_TOTAL_EXCL, _COL_Y1).value,
        },
        "total_opex_incl_contingencies": {
            "opex_row": _ROW_TOTAL_INCL,
            "budget_cached": ws_opex_v.cell(_ROW_TOTAL_INCL, _COL_BUDGET).value,
            "y1_cached": ws_opex_v.cell(_ROW_TOTAL_INCL, _COL_Y1).value,
        },
    }

    return {
        "_meta": {
            "description": "Authoritative OPEX structural truth extracted from Oborovo sensitivity workbook",
            "source_file": workbook_path.name,
            "source_sha256": EXPECTED_SHA256,
            "sheet": "OpEx",
            "inputs_sheet": "Inputs",
            "scenarios_sheet": "Scenarios",
            "dual_load": "data_only=True for cached values; data_only=False for formulas",
        },
        "inputs": inputs_block,
        "scenarios": scenarios_meta,
        "categories": categories,
        "totals": totals,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).parents[1] / "tests" / "fixtures" / "excel_oborovo_opex_structural_truth.json",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        data = extract(args.workbook)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    output = json.dumps(data, indent=2, ensure_ascii=False)

    if args.dry_run:
        print(output)
    else:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(output, encoding="utf-8")
        print(f"Fixture written: {args.out}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
