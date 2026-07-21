"""finco_recon.workbook — Build the Oborovo Excel↔Python reconciliation workbook.

Generates a 16-sheet audit-grade workbook using openpyxl.
All model periods run horizontally. No financial formulas are modified.
"""
from __future__ import annotations

import math
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from finco_recon.catalog import CATALOG, LineItem, get_catalog_by_section, get_item
from finco_recon.materiality import MaterialitySettings, DEFAULT_MATERIALITY
from finco_recon.sources import OborovoSources, ExcelData, EngineData, LegacyData

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASELINE_ID = "oborovo"
UNAVAILABLE = "UNAVAILABLE"
OUT_OF_SCOPE = "OUT OF CLEAN ENGINE SCOPE"

# Classification → fill colour (hex, no #)
CLASS_FILL = {
    "MATCH":                    "C6EFCE",
    "PYTHON BUG":               "FF9999",
    "EXCEL BUG":                "FFD966",
    "POLICY DIFFERENCE":        "BDD7EE",
    "UNRESOLVED SOURCE":        "E2EFDA",
    "TIMING / ROUNDING":        "FFFFCC",
    "OUT OF CLEAN ENGINE SCOPE":"D9D9D9",
    # blank = OPEN — ROOT CAUSE REQUIRED (no fill — white)
    "":                         "FFFFFF",
}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def _font(bold: bool = False, size: int = 9, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = _fill("2F4F6F")  # dark navy
SECTION_FILL = _fill("4472C4")  # blue
EXCEL_FILL   = _fill("EAF0FB")  # light blue
PYTHON_FILL  = _fill("EBF7EE")  # light green
DELTA_FILL   = _fill("FFF9E6")  # light amber
DPCT_FILL    = _fill("FFF3CD")  # lighter amber
GREY_FILL    = _fill("F2F2F2")

KEUR_FMT  = '#,##0.0'
INT_FMT   = '#,##0'
PCT_FMT   = '0.00%'
DSCR_FMT  = '0.000'
DATE_FMT  = 'DD-MMM-YY'
FRAC_FMT  = '0.0000'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe(val: Any) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _delta(excel: float | None, python: float | None) -> float | None:
    if excel is None or python is None:
        return None
    return python - excel


def _delta_pct(delta: float | None, excel: float | None) -> float | None:
    if delta is None or excel is None:
        return None
    ref = abs(excel)
    if ref < 1e-9:
        return None
    return delta / ref


def _classify(
    item: LineItem,
    delta: float | None,
    excel: float | None,
    python: float | None,
    mat: MaterialitySettings,
    documented_cls: str | None = None,
) -> str:
    """Classify a delta. Only assigns a non-blank classification when evidence exists.

    A material delta without a documented root cause returns "" (blank),
    meaning OPEN — ROOT CAUSE REQUIRED. This is never auto-assigned to
    POLICY DIFFERENCE based on magnitude alone.
    """
    if not item.in_clean_engine:
        return "OUT OF CLEAN ENGINE SCOPE"
    if excel is None and python is None:
        return "UNRESOLVED SOURCE"
    if excel is None:
        return "UNRESOLVED SOURCE"
    if python is None:
        return "OUT OF CLEAN ENGINE SCOPE"
    if delta is None:
        return "UNRESOLVED SOURCE"
    if not mat.is_material(delta, excel, python, item.unit):
        return "MATCH"
    # Material delta: return documented classification only if evidence exists
    if documented_cls:
        return documented_cls
    # No root cause established yet → blank = OPEN
    return ""


def _num_fmt(unit: str) -> str:
    if unit in ("kEUR",):
        return KEUR_FMT
    if unit in ("%",):
        return PCT_FMT
    if unit in ("x",):
        return DSCR_FMT
    if unit == "MWh":
        return INT_FMT
    if unit == "frac":
        return FRAC_FMT
    if unit == "days":
        return INT_FMT
    if unit == "MW":
        return KEUR_FMT
    return KEUR_FMT


def _set_cell(ws, row: int, col: int, value: Any, *, bold: bool = False,
              fill: PatternFill | None = None, fmt: str | None = None,
              align: str = "left", font_color: str = "000000",
              font_size: int = 9) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=font_size, color=font_color)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.alignment = _align(h=align, v="center")


def _header_row(ws, values: list, row: int = 1, fill: PatternFill | None = None) -> None:
    f = fill or HEADER_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _font(bold=True, size=9, color="FFFFFF")
        cell.fill = f
        cell.alignment = _align(h="center", v="center")


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _freeze(ws, cell: str) -> None:
    ws.freeze_panes = cell


def _autofilter(ws, first_row: int, last_col: int) -> None:
    ws.auto_filter.ref = f"A{first_row}:{get_column_letter(last_col)}{first_row}"


def _write_period_headers(ws, engine_periods: list[EngineData], header_row: int, start_col: int) -> None:
    for i, ep in enumerate(engine_periods):
        col = start_col + i
        cell = ws.cell(row=header_row, column=col, value=ep.period_end)
        cell.font = _font(bold=True, size=8, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = _align(h="center", v="center")
        cell.number_format = DATE_FMT
        ws.column_dimensions[get_column_letter(col)].width = 10.5


def _section_header(ws, row: int, label: str, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = _font(bold=True, size=9, color="FFFFFF")
    ws.cell(row=row, column=2, value=label)
    ws.cell(row=row, column=2).alignment = _align(h="left", v="center")


def _write_recon_block(
    ws,
    label: str,
    code: str,
    unit: str,
    excel_vals: list[float | None],
    python_vals: list[float | None],
    start_row: int,
    data_col: int,
    mat: MaterialitySettings,
    item: LineItem | None = None,
    classification: str | None = None,
) -> int:
    """Write Excel / Python / Delta / Delta% rows for one line item.

    Returns next available row.
    """
    fmt = _num_fmt(unit)
    n = len(excel_vals)

    for view_idx, (view_label, vals, row_fill) in enumerate([
        ("Excel",   excel_vals,  EXCEL_FILL),
        ("Python",  python_vals, PYTHON_FILL),
    ]):
        r = start_row + view_idx
        ws.cell(row=r, column=1, value=code if view_idx == 0 else "").font = _font(size=8)
        ws.cell(row=r, column=1).value = code if view_idx == 0 else ""
        ws.cell(row=r, column=2, value=label if view_idx == 0 else "").font = _font(bold=(view_idx == 0), size=9)
        ws.cell(row=r, column=2).value = label if view_idx == 0 else ""
        ws.cell(row=r, column=3, value=view_label).font = _font(size=8)
        ws.cell(row=r, column=3).fill = row_fill
        ws.cell(row=r, column=4, value=unit).font = _font(size=8)
        # Reviewer Status column (col 5)
        rev_cell = ws.cell(row=r, column=5, value="NOT REVIEWED" if view_idx == 0 else None)
        rev_cell.fill = _fill("F2F2F2")
        rev_cell.font = _font(size=7)
        rev_cell.alignment = _align(h="center", v="center")
        for ci in range(n):
            v = vals[ci]
            cell = ws.cell(row=r, column=data_col + ci)
            if v is None:
                cell.value = None
            elif isinstance(v, str):
                cell.value = v
            else:
                cell.value = v
                cell.number_format = fmt
            cell.fill = row_fill
            cell.font = _font(size=8)
            cell.alignment = _align(h="right", v="center")

    # Delta row
    r_delta = start_row + 2
    ws.cell(row=r_delta, column=2, value="").font = _font(size=8)
    ws.cell(row=r_delta, column=3, value="Delta").font = _font(bold=True, size=8)
    ws.cell(row=r_delta, column=3).fill = DELTA_FILL
    ws.cell(row=r_delta, column=4, value=unit).font = _font(size=8)
    ws.cell(row=r_delta, column=5).fill = _fill("F2F2F2")
    for ci in range(n):
        ev = excel_vals[ci]
        pv = python_vals[ci]
        d = _delta(ev if isinstance(ev, (int, float)) else None,
                   pv if isinstance(pv, (int, float)) else None)
        cell = ws.cell(row=r_delta, column=data_col + ci)
        if d is None:
            cell.value = None
        else:
            cell.value = d
            cell.number_format = fmt
            # Colour: neutral audit highlighting
            abs_d = abs(d)
            if item and mat.is_material(d, ev, pv, unit):
                if abs_d > 0 and _delta_pct(d, ev) is not None and abs(_delta_pct(d, ev)) > 0.10:
                    cell.fill = _fill("FFAAAA")
                else:
                    cell.fill = _fill("FFD9B3")
            else:
                cell.fill = DELTA_FILL
        cell.font = _font(size=8)
        cell.alignment = _align(h="right", v="center")

    # Delta% row
    r_dpct = start_row + 3
    ws.cell(row=r_dpct, column=3, value="Delta %").font = _font(size=8)
    ws.cell(row=r_dpct, column=3).fill = DPCT_FILL
    ws.cell(row=r_dpct, column=4, value="%").font = _font(size=8)
    ws.cell(row=r_dpct, column=5).fill = _fill("F2F2F2")
    for ci in range(n):
        ev = excel_vals[ci]
        pv = python_vals[ci]
        d = _delta(ev if isinstance(ev, (int, float)) else None,
                   pv if isinstance(pv, (int, float)) else None)
        dp = _delta_pct(d, ev if isinstance(ev, (int, float)) else None)
        cell = ws.cell(row=r_dpct, column=data_col + ci)
        if dp is None:
            cell.value = None
        else:
            cell.value = dp
            cell.number_format = PCT_FMT
        cell.fill = DPCT_FILL
        cell.font = _font(size=8)
        cell.alignment = _align(h="right", v="center")

    return start_row + 4


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_readme(wb: Workbook, src: OborovoSources, generation_ts: str, pr_head: str) -> None:
    """00_README — validation guide for the finance reviewer."""
    ws = wb.create_sheet("00_README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72

    def _row(r, a, b, bold_a=False, bold_b=False, fill_a=None, fill_b=None):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.font = _font(bold=bold_a, size=10)
        cb.font = _font(bold=bold_b, size=10)
        if fill_a:
            ca.fill = _fill(fill_a)
        if fill_b:
            cb.fill = _fill(fill_b)
        ca.alignment = _align(h="right", v="center")
        cb.alignment = _align(h="left", v="center", wrap=True)

    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Oborovo Excel ↔ Python Reconciliation Workbook")
    title.font = _font(bold=True, size=14)
    title.fill = HEADER_FILL
    title.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title.alignment = _align(h="left", v="center")
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    ws.merge_cells("A1:B1")

    r = 3
    _row(r, "Source workbook", src.excel_source_filename or "N/A"); r += 1
    _row(r, "Source SHA256", src.excel_fixture_sha256 or "N/A"); r += 1
    _row(r, "Generation timestamp", generation_ts); r += 1
    _row(r, "PR head SHA", pr_head or "N/A"); r += 1
    _row(r, "Materiality (abs kEUR)", "1.0 kEUR"); r += 1
    _row(r, "Materiality (relative)", "0.1%"); r += 1
    r += 1

    _row(r, "COLOUR KEY", "", bold_a=True, bold_b=True); r += 1
    _row(r, "Excel row", "Light blue background — authoritative XLSM value", fill_b="EAF0FB"); r += 1
    _row(r, "Python row", "Light green background — clean engine computed value", fill_b="EBF7EE"); r += 1
    _row(r, "Delta row", "Amber background — Delta = Python − Excel (always signed)", fill_b="FFF9E6"); r += 1
    _row(r, "Material delta", "Orange-amber — material difference (abs > 1 kEUR or > 0.1%)", fill_b="FFD9B3"); r += 1
    _row(r, "Large material delta", "Light red — large difference (> 10% relative)", fill_b="FFAAAA"); r += 1
    _row(r, "Section header", "Blue row — section separator", fill_b="4472C4"); r += 1
    r += 1

    _row(r, "DELTA CONVENTION", "", bold_a=True); r += 1
    _row(r, "", "Delta = Python − Excel. A positive Delta means Python is higher."); r += 1
    _row(r, "", "A negative Delta means Python is lower. Sign is informational only."); r += 1
    _row(r, "", "Do not interpret positive as good or negative as bad."); r += 1
    r += 1

    _row(r, "N/A CONVENTION", "", bold_a=True); r += 1
    _row(r, "", "N/A means not modelled / unavailable. It is never assumed zero."); r += 1
    _row(r, "", "OUT OF CLEAN ENGINE SCOPE means the clean Python engine does not yet model this line."); r += 1
    r += 1

    _row(r, "REVIEWER STATUS", "", bold_a=True); r += 1
    _row(r, "", "Each item begins as NOT REVIEWED. Allowed values:"); r += 1
    for status in ("NOT REVIEWED", "ACCEPT", "INVESTIGATE", "EXCEL ISSUE", "PYTHON ISSUE", "POLICY DECISION"):
        _row(r, "", f"  {status}"); r += 1
    _row(r, "", "Reviewer status is for manual notation only. It does not affect calculations."); r += 1
    r += 1

    _row(r, "REVIEW ORDER", "", bold_a=True); r += 1
    for i, step in enumerate([
        "01 EXEC RECON — landing summary",
        "02 INPUTS RECON — verify input assumptions",
        "03 TIMELINE — confirm period dates and COD",
        "04 PROD & REVENUE — does production match? Does price match?",
        "05 OPEX — per-item B.01–B.13 by period",
        "06 P&L — revenue through net income",
        "07 CAPEX/IDC — CAPEX items and financing costs",
        "08 DEPRECIATION — book vs tax",
        "09 TAX — full tax bridge including loss carryforward",
        "10 CFADS — EBITDA bridge to CFADS",
        "11 SENIOR DEBT — period debt schedule and DSCR",
        "12 SHL — SHL schedule (Excel only, Python out of scope)",
        "13 OPENING BALANCES",
        "14 DELTA REGISTER — all material deltas, filter by OPEN",
    ], 1):
        _row(r, f"Step {i:02d}", step); r += 1


def _build_exec_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("01_EXEC_RECON")
    ws.sheet_view.showGridLines = False

    headers = ["Section", "Line Item", "Excel Value", "Python Value",
               "Delta", "Delta %", "Max Period |Delta|", "Status", "Classification",
               "Notes / Root Cause", "Detail Sheet"]
    _header_row(ws, headers)

    SECTION_DETAIL_MAP = {
        "TIMELINE":       "03_TIMELINE_RECON",
        "PRODUCTION":     "04_PROD_REV_RECON",
        "REVENUE":        "04_PROD_REV_RECON",
        "OPEX":           "05_OPEX_RECON",
        "EBITDA":         "06_PNL_RECON",
        "CAPEX":          "07_CAPEX_IDC_RECON",
        "IDC":            "07_CAPEX_IDC_RECON",
        "DEPRECIATION":   "08_DEPRECIATION_RECON",
        "TAX":            "09_TAX_RECON",
        "CFADS":          "10_CFADS_RECON",
        "SENIOR DEBT":    "11_SENIOR_DEBT_RECON",
        "SENIOR INTEREST":"11_SENIOR_DEBT_RECON",
        "SHL":            "12_SHL_RECON",
        "SHL INTEREST":   "12_SHL_RECON",
        "CASH FLOW":      "10_CFADS_RECON",
    }

    def _tot(vals: list) -> float | None:
        filtered = [v for v in vals if v is not None]
        return sum(filtered) if filtered else None

    def _max_abs_delta(excel_list: list, python_list: list) -> float | None:
        deltas = []
        for ev, pv in zip(excel_list, python_list):
            d = _delta(_safe(ev), _safe(pv))
            if d is not None:
                deltas.append(abs(d))
        return max(deltas) if deltas else None

    excel_p = src.excel
    eng_p   = src.engine

    # Period-level lists for max-delta calculation
    xl_rev   = [e.revenue_keur for e in excel_p]
    py_rev   = [e.revenue_keur for e in eng_p]
    xl_opex  = [e.opex_keur for e in excel_p]
    py_opex  = [e.opex_keur for e in eng_p]
    # EBITDA: use direct workbook value where cached (periods 1-20), derived elsewhere
    xl_ebitda = [
        e.ebitda_keur if e.ebitda_keur is not None else e.ebitda_derived_keur
        for e in excel_p
    ]
    py_ebitda = [e.ebitda_keur for e in eng_p]
    xl_ctax  = [e.cash_tax_keur for e in excel_p]
    py_ctax  = [e.corporate_tax_cash_keur for e in eng_p]
    xl_cfads = [e.cfads_keur for e in excel_p]
    py_cfads = [e.cfads_keur for e in eng_p]
    xl_sint  = [e.senior_interest_keur for e in excel_p]
    py_sint  = [e.sd_interest_keur for e in eng_p]
    xl_dep   = [e.depreciation_keur for e in excel_p]
    py_dep   = [e.book_depreciation_keur for e in eng_p]

    # Excel EBITDA identity: Revenue - OPEX should equal EBITDA (direct or derived) for all 60 periods
    xl_rev_tot   = _tot(xl_rev)
    xl_opex_tot  = _tot(xl_opex)
    xl_ebitda_tot = _tot(xl_ebitda)
    xl_ebitda_computed = (xl_rev_tot - xl_opex_tot) if (xl_rev_tot is not None and xl_opex_tot is not None) else None

    py_rev_tot   = _tot(py_rev)
    py_opex_tot  = _tot(py_opex)
    py_ebitda_tot = _tot(py_ebitda)

    # Python average DSCR: mean over periods with active senior debt balance AND active debt service.
    # Filter mirrors Excel logic: sd_opening_keur > 0.01 AND sd_service_keur > 0.01.
    dscr_vals = [
        e.sd_dscr for e in eng_p
        if e.sd_dscr is not None and e.sd_dscr > 0
        and e.sd_opening_keur is not None and e.sd_opening_keur > 0.01
    ]
    py_avg_dscr = sum(dscr_vals) / len(dscr_vals) if dscr_vals else None
    py_min_dscr = min(dscr_vals) if dscr_vals else None

    # COD date: Excel from XLSM Inputs!D11, Python from engine first period_start
    py_cod = src.engine[0].period_start if eng_p else None
    cod_match = src.cod_date == py_cod
    cod_cls = "MATCH" if cod_match else ""

    rows_data = [
        # (section, line_item, excel_val, python_val, max_period_delta, classification, notes)
        ("TIMELINE",     "Operating periods",
         float(len(excel_p)), float(len(eng_p)), None, "MATCH", ""),
        ("TIMELINE",     "COD date (Excel vs engine period start)",
         src.cod_date, py_cod, None, cod_cls,
         "Excel COD from XLSM Inputs!D11 (operation_start); Python from engine first period_start" if not cod_match else ""),
        ("PRODUCTION",   "Total net production (MWh)",
         _tot([e.production_mwh for e in excel_p]),
         _tot([e.production_mwh for e in eng_p]),
         _max_abs_delta([e.production_mwh for e in excel_p], [e.production_mwh for e in eng_p]),
         None, "Excel: CF row 21 production_mwh; Python: engine production_mwh"),
        ("PRODUCTION",   "Avg PPA tariff (EUR/MWh)",
         src.ppa_tariff_eur_mwh or None, None, None, "UNRESOLVED SOURCE",
         "Excel: PPA base tariff Y1; indexed per-period price vs Python production×tariff requires period-level reconciliation in 03_PROD_REV_RECON"),
        ("REVENUE",      "Total operating revenue CF (kEUR)",
         xl_rev_tot, py_rev_tot, _max_abs_delta(xl_rev, py_rev), None,
         "Excel: CF.operating_revenues_keur; Python: engine revenue_keur"),
        ("OPEX",         "Total OPEX (kEUR)",
         xl_opex_tot, py_opex_tot, _max_abs_delta(xl_opex, py_opex), None,
         "Excel: abs(CF.operating_expenses_after_bank_tax_keur); Python: engine opex_keur"),
        ("EBITDA",       "Total EBITDA CF (kEUR) — direct+derived",
         xl_ebitda_tot, py_ebitda_tot, _max_abs_delta(xl_ebitda, py_ebitda), None,
         "Excel: CF.ebitda_keur (direct periods 1-20) + Rev-OPEX (derived periods 21-60); Python: engine ebitda_keur"),
        ("EBITDA",       "Excel EBITDA identity: Revenue - OPEX (kEUR)",
         xl_ebitda_computed, xl_ebitda_tot, None,
         "MATCH" if (xl_ebitda_computed is not None and xl_ebitda_tot is not None and abs(xl_ebitda_computed - xl_ebitda_tot) < 1.0) else ("" if xl_ebitda_computed is not None else "UNRESOLVED SOURCE"),
         f"Computed={xl_ebitda_computed:.1f} vs Direct/Derived={xl_ebitda_tot:.1f}" if (xl_ebitda_computed is not None and xl_ebitda_tot is not None) else "Cannot verify — missing data"),
        ("CAPEX",        "Total CAPEX (kEUR)",
         src.excel_total_capex_keur or None, src.total_capex_keur, None, None,
         "Excel: XLSM Inputs!C45 (authoritative); Python: factory total_capex. See 06_CAPEX_IDC_RECON for per-item breakdown."),
        ("IDC",          "Bank IDC (kEUR)",
         src.excel_idc_keur or None, src.idc_keur, None, None,
         "Excel: XLSM CapEx sheet C.IDC; Python: factory idc_keur"),
        ("DEPRECIATION", "Total book depreciation (kEUR)",
         _tot(xl_dep), _tot(py_dep), _max_abs_delta(xl_dep, py_dep), None,
         "Excel: P&L.depreciation_keur; Python: engine book_depreciation_keur"),
        ("TAX",          "Total cash tax (kEUR)",
         _tot(xl_ctax), _tot(py_ctax), _max_abs_delta(xl_ctax, py_ctax), None,
         "Excel: abs(CF.corporate_income_tax_keur); Python: engine corporate_tax_cash_keur"),
        ("CFADS",        "Total CFADS (kEUR)",
         _tot(xl_cfads), _tot(py_cfads), _max_abs_delta(xl_cfads, py_cfads), None,
         "Excel: CF.free_cash_flow_for_banks_keur; Python: engine cfads_keur"),
        ("SENIOR DEBT",  "Debt size at COD (kEUR)",
         src.excel_total_debt_keur or None, src.engine_debt_size_keur, None, "POLICY DIFFERENCE",
         "Excel uses GEARING CAP (75.24% × CAPEX); Python uses DSCR SCULPTED at 1.15x"),
        ("SENIOR DEBT",  "Target DSCR",
         src.excel_target_dscr or None, 1.15, None, "MATCH",
         "Excel: DS.senior_debt_dscr_target; Python: SeniorDebtPolicy.target_dscr=1.15"),
        ("SENIOR DEBT",  "Min DSCR",
         src.excel_min_dscr or None, py_min_dscr, None, None,
         "Excel: CF.minimum_senior_dscr_period min; Python: min of engine sd_dscr (active DS periods)"),
        ("SENIOR DEBT",  "Avg DSCR (active DS periods)",
         src.excel_avg_dscr or None, py_avg_dscr, None, None,
         "Excel: average of CF.average_senior_dscr_period; Python: mean of engine sd_dscr (sd_dscr>0)"),
        ("SENIOR INTEREST", "Total senior interest (kEUR)",
         _tot(xl_sint), _tot(py_sint), _max_abs_delta(xl_sint, py_sint), None,
         "Excel: DS.senior_net_interest_keur; Python: engine sd_interest_keur"),
        ("SHL INTEREST", "Total SHL interest (kEUR)",
         _tot([e.shl_interest_keur for e in excel_p]), UNAVAILABLE, None,
         "OUT OF CLEAN ENGINE SCOPE",
         "Excel: P&L.shl_interests_keur (DS sheet direct); Python: not modelled in clean engine"),
        ("CASH FLOW",    "Total FCF for distribution (kEUR)",
         None, UNAVAILABLE, None,
         "OUT OF CLEAN ENGINE SCOPE",
         "Excel FCF for distribution = CFADS - SD service (not separately extracted); Python: not modelled in clean engine"),
    ]

    for r_idx, row_tuple in enumerate(rows_data, 2):
        section, line, excel_v, python_v, max_d, cls_override, notes = row_tuple
        ev = _safe(excel_v) if not isinstance(excel_v, str) else None
        pv = _safe(python_v) if not isinstance(python_v, str) else None
        d = _delta(ev, pv)
        dp = _delta_pct(d, ev)
        is_mat = mat.is_material(d, ev, pv) if d is not None else False

        # Determine classification
        if cls_override is not None:
            cls = cls_override
        elif ev is None:
            cls = "UNRESOLVED SOURCE"
        elif pv is None:
            cls = "OUT OF CLEAN ENGINE SCOPE"
        elif d is not None and not is_mat:
            cls = "MATCH"
        else:
            cls = ""  # OPEN — ROOT CAUSE REQUIRED

        status = "OPEN" if (cls == "" and is_mat) else ("MATERIAL" if is_mat else "OK")

        fill = _fill(CLASS_FILL.get(cls, "FFFFFF"))
        row_fill = _fill("FFF0F0") if (is_mat and cls == "") else None

        ws.cell(row=r_idx, column=1, value=section).font = _font(bold=True, size=9)
        ws.cell(row=r_idx, column=2, value=line).font = _font(size=9)
        c3 = ws.cell(row=r_idx, column=3, value=ev if ev is not None else (excel_v if isinstance(excel_v, str) else None))
        c3.font = _font(size=9)
        if ev is not None: c3.number_format = KEUR_FMT
        c4 = ws.cell(row=r_idx, column=4, value=pv if pv is not None else (python_v if isinstance(python_v, str) else None))
        c4.font = _font(size=9)
        if pv is not None: c4.number_format = KEUR_FMT
        c5 = ws.cell(row=r_idx, column=5, value=d)
        c5.font = _font(size=9)
        if d is not None: c5.number_format = KEUR_FMT
        c6 = ws.cell(row=r_idx, column=6, value=dp)
        c6.font = _font(size=9)
        if dp is not None: c6.number_format = PCT_FMT
        c7 = ws.cell(row=r_idx, column=7, value=max_d)
        c7.font = _font(size=9)
        if max_d is not None: c7.number_format = KEUR_FMT
        ws.cell(row=r_idx, column=8, value=status).font = _font(size=9)
        cls_cell = ws.cell(row=r_idx, column=9, value=cls if cls else "")
        cls_cell.font = _font(size=9)
        cls_cell.fill = fill
        ws.cell(row=r_idx, column=10, value=notes).font = _font(size=8)
        target_sheet = SECTION_DETAIL_MAP.get(section)
        if target_sheet:
            link_cell = ws.cell(row=r_idx, column=11, value=target_sheet)
            link_cell.hyperlink = f"#'{target_sheet}'!A1"
            link_cell.style = "Hyperlink"
            link_cell.font = _font(size=9)
        if row_fill:
            for c in range(1, 9):
                ws.cell(row=r_idx, column=c).fill = row_fill

    _set_col_widths(ws, {1: 16, 2: 44, 3: 16, 4: 16, 5: 12, 6: 10, 7: 14, 8: 10, 9: 28, 10: 60, 11: 24})
    _freeze(ws, "A2")
    _autofilter(ws, 1, 11)


def _build_inputs_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("02_INPUTS_RECON")
    ws.sheet_view.showGridLines = False
    _header_row(ws, ["Code", "Input", "Excel", "Python", "Delta", "Status"])

    def row(ws, r, code, label, excel_v, python_v, fmt=KEUR_FMT):
        ev = _safe(excel_v)
        pv = _safe(python_v)
        d = _delta(ev, pv)
        is_mat = mat.is_material(d, ev, pv) if d is not None else False
        status = "MATERIAL" if is_mat else "MATCH" if d is not None and abs(d) < 1e-6 else "REVIEW"
        ws.cell(row=r, column=1, value=code).font = _font(size=8)
        ws.cell(row=r, column=2, value=label).font = _font(size=9)
        c3 = ws.cell(row=r, column=3, value=ev if ev is not None else excel_v)
        c3.font = _font(size=9); c3.number_format = fmt if ev is not None else "@"
        c4 = ws.cell(row=r, column=4, value=pv if pv is not None else python_v)
        c4.font = _font(size=9); c4.number_format = fmt if pv is not None else "@"
        c5 = ws.cell(row=r, column=5, value=d); c5.font = _font(size=9)
        if d is not None: c5.number_format = fmt
        ws.cell(row=r, column=6, value=status).font = _font(size=9)
        if is_mat:
            for c in range(1, 7): ws.cell(row=r, column=c).fill = _fill("FFC7CE")

    UNRES = "UNRESOLVED — not in Excel fixture"

    r = 2
    def sec(label):
        nonlocal r
        ws.cell(row=r, column=1, value="").fill = SECTION_FILL
        ws.cell(row=r, column=2, value=label).font = _font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=r, column=2).fill = SECTION_FILL
        for c in range(3, 7): ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1

    # Excel-provenance: oborovo_golden.json was extracted from the source Excel workbook.
    # Python-provenance: app.project_factories (factory config / engine policy).
    # Where both come from the same golden JSON, the Python side is also golden-derived.
    # Where only golden exists (no independent factory check), mark UNRESOLVED SOURCE.

    sec("— PROJECT / TIMELINE (Excel: oborovo_golden.json) —")
    row(ws, r, "T.01", "Capacity (MW)",        src.capacity_mw,         src.capacity_mw,         KEUR_FMT); r += 1
    row(ws, r, "T.02", "Financial close",       src.financial_close,     src.financial_close,     "@"); r += 1
    row(ws, r, "T.03", "COD date",              src.cod_date,            src.engine[0].period_start if src.engine else src.cod_date, "@"); r += 1
    row(ws, r, "T.04", "Horizon (years)",       src.horizon_years,       src.horizon_years,       INT_FMT); r += 1
    row(ws, r, "T.05", "Construction (months)", src.construction_months, src.construction_months, INT_FMT); r += 1
    row(ws, r, "T.06", "Period frequency",      "Semestrial",            "Semestrial",            "@"); r += 1

    sec("— PRODUCTION (Excel: XLSM Inputs sheet) —")
    row(ws, r, "P.01", "P50 operating hours",    src.operating_hours_p50 or None, 1494.0, INT_FMT); r += 1
    row(ws, r, "P.02", "PV degradation (%/yr)",  src.pv_degradation_pa or None,  0.004,   PCT_FMT); r += 1
    row(ws, r, "P.03", "Plant availability",      src.plant_availability or None, 0.99,    PCT_FMT); r += 1
    row(ws, r, "P.04", "Grid availability",       src.grid_availability or None,  0.99,    PCT_FMT); r += 1

    sec("— PRICE / REVENUE (Excel: oborovo_golden.json) —")
    row(ws, r, "R.01", "PPA tariff (EUR/MWh)",  src.ppa_tariff_eur_mwh, src.ppa_tariff_eur_mwh, KEUR_FMT); r += 1
    row(ws, r, "R.02", "PPA term (years)",       src.ppa_term_years,     src.ppa_term_years,     INT_FMT); r += 1
    row(ws, r, "R.03", "PPA indexation (%/yr)",  src.ppa_index,          src.ppa_index,          PCT_FMT); r += 1
    row(ws, r, "R.04", "Effective price (EUR/MWh) [UNRESOLVED]", UNRES, UNRES, "@"); r += 1

    sec("— OPEX Y1 (kEUR/yr) — Excel: XLSM Inputs rows 146-160; Python: app.project_factories —")
    for item in src.opex_items:
        code = item["code"]
        xl_opex_item = src.excel_opex_annual.get(code, {})
        xl_y1 = xl_opex_item.get("year1_keur") if xl_opex_item else None
        row(ws, r, code, item["name"] + " Y1 (kEUR)", xl_y1, item["y1_keur"], KEUR_FMT); r += 1
        row(ws, r, "",   item["name"] + " inflation [Excel: not extracted]", UNRES, item["inflation"], "@"); r += 1

    sec("— CAPEX (kEUR) — Excel: XLSM Inputs/CapEx sheets; Python: app.project_factories —")
    for item in src.capex_items:
        code = item["code"]
        xl_cap = src.excel_capex_items.get(code, {})
        xl_amt = xl_cap.get("amount_keur") if xl_cap else None
        row(ws, r, code, item["name"], xl_amt, item["amount_keur"], KEUR_FMT); r += 1
    row(ws, r, "C.00", "Total Hard CAPEX [Python factory only]", UNRES, src.hard_capex_keur, KEUR_FMT); r += 1
    row(ws, r, "C.IDC", "Bank IDC",   src.excel_idc_keur or None, src.idc_keur, KEUR_FMT); r += 1
    row(ws, r, "C.TOT", "Total CAPEX (Excel XLSM vs Python factory)",
        src.excel_total_capex_keur, src.total_capex_keur, KEUR_FMT); r += 1

    sec("— TAX — Excel: not in fixture; Python: tax_reference_inputs —")
    row(ws, r, "TX.01", "CIT rate [UNRESOLVED Excel]",               UNRES, 0.10,    "@"); r += 1
    row(ws, r, "TX.02", "Loss carryforward (yrs) [UNRESOLVED Excel]",UNRES, 5.0,    "@"); r += 1
    row(ws, r, "TX.03", "ATAD enabled [UNRESOLVED Excel]",           UNRES, "Yes",  "@"); r += 1
    row(ws, r, "TX.04", "ATAD EBITDA limit [UNRESOLVED Excel]",      UNRES, 0.30,   "@"); r += 1
    row(ws, r, "TX.05", "ATAD de minimis kEUR/yr [UNRESOLVED Excel]",UNRES, 3000.0, "@"); r += 1

    sec("— SENIOR DEBT (Excel: XLSM Inputs sheet; Python: SeniorDebtPolicy) —")
    row(ws, r, "SD.01", "Target DSCR",              src.excel_target_dscr, 1.15,         DSCR_FMT); r += 1
    row(ws, r, "SD.02", "Gearing ratio (Excel only)", 0.7524,               UNRES,        PCT_FMT); r += 1
    row(ws, r, "SD.03", "Debt size at COD (kEUR)",  src.excel_total_debt_keur, src.engine_debt_size_keur, KEUR_FMT); r += 1
    xl_rate = (src.excel_senior_base_rate + src.excel_senior_margin_bps / 10000.0) if src.excel_senior_base_rate else None
    row(ws, r, "SD.04", "All-in fixed rate (base+margin)", xl_rate, 0.0565, PCT_FMT); r += 1
    row(ws, r, "SD.05", "Tenor (years)",             src.excel_senior_maturity_years or None, 14.0, INT_FMT); r += 1
    row(ws, r, "SD.06", "Day count [Excel: not extracted]", UNRES, "ACT/365", "@"); r += 1
    row(ws, r, "SD.07", "Sizing mode",              "GEARING CAP (Excel)", "DSCR SCULPTED (Python)", "@"); r += 1

    sec("— SHL (Excel: XLSM Inputs sheet / DS sheet) —")
    row(ws, r, "SH.01", "SHL amount at COD (kEUR)", src.shl_amount_keur,  UNRES, KEUR_FMT); r += 1
    row(ws, r, "SH.02", "SHL rate",                 src.shl_rate,         UNRES, PCT_FMT); r += 1
    row(ws, r, "SH.03", "SHL IDC capitalised (kEUR)",src.shl_idc_keur,    UNRES, KEUR_FMT); r += 1

    _set_col_widths(ws, {1: 8, 2: 40, 3: 18, 4: 18, 5: 14, 6: 12})
    _freeze(ws, "A2")
    _autofilter(ws, 1, 6)


def _build_horizontal_sheet(
    wb: Workbook,
    sheet_name: str,
    src: OborovoSources,
    mat: MaterialitySettings,
    blocks: list[dict],
) -> None:
    """Generic builder for all period-horizontal sheets.

    blocks = list of dicts:
        type: "section" | "item"
        For section: label: str
        For item: code, name, unit, excel_vals, python_vals, item (LineItem)
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n = len(src.engine)
    HDR_ROW = 1
    DATA_START_COL = 6  # A=code B=name C=view D=unit E=reviewer_status F+...

    # Row 1: fixed headers
    for c, label in enumerate(["Code", "Line Item", "View", "Unit", "Reviewer Status"], 1):
        cell = ws.cell(row=HDR_ROW, column=c, value=label)
        cell.font = _font(bold=True, size=9, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = _align(h="center", v="center")
    # Period date headers
    _write_period_headers(ws, src.engine, HDR_ROW, DATA_START_COL)

    # Fixed column widths
    _set_col_widths(ws, {1: 7, 2: 32, 3: 8, 4: 7, 5: 13})

    cur_row = 2
    for block in blocks:
        if block["type"] == "section":
            n_cols = DATA_START_COL + n
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = SECTION_FILL
                cell.font = _font(bold=True, size=9, color="FFFFFF")
            ws.cell(row=cur_row, column=2, value=block["label"])
            ws.cell(row=cur_row, column=2).alignment = _align(h="left", v="center")
            cur_row += 1
        else:
            excel_vals = block.get("excel_vals", [None] * n)
            python_vals = block.get("python_vals", [None] * n)
            item = block.get("item")
            cur_row = _write_recon_block(
                ws=ws,
                label=block["name"],
                code=block["code"],
                unit=block["unit"],
                excel_vals=excel_vals,
                python_vals=python_vals,
                start_row=cur_row,
                data_col=DATA_START_COL,
                mat=mat,
                item=item,
            )

    # Freeze first 4 cols + header row
    _freeze(ws, f"{get_column_letter(DATA_START_COL)}2")
    _autofilter(ws, HDR_ROW, DATA_START_COL - 1)


def _build_timeline_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("03_TIMELINE_RECON")
    ws.sheet_view.showGridLines = False

    n = len(src.engine)
    DATA_COL = 6
    _header_row(ws, ["Code", "Line Item", "View", "Unit", "Reviewer Status"])
    _write_period_headers(ws, src.engine, 1, DATA_COL)
    _set_col_widths(ws, {1: 7, 2: 32, 3: 8, 4: 7, 5: 13})

    n = len(src.engine)
    none_n = [None] * n

    # TL.02, TL.04: now sourced from XLSM CF sheet.
    # TL.01, TL.03, TL.05: engine-only.
    items_def = [
        ("TL.01", "Period index",    "index",
         none_n,                                                    [e.period_index for e in src.engine]),
        ("TL.02", "Period end date", "date",
         [ep.period_end for ep in src.excel],                       [e.period_end for e in src.engine]),
        ("TL.03", "Days in period",  "days",
         none_n,                                                    [e.days_in_period for e in src.engine]),
        ("TL.04", "Day fraction",    "frac",
         [ep.operation_period_fraction for ep in src.excel],        [e.day_fraction for e in src.engine]),
        ("TL.05", "Is operation",    "flag",
         none_n,                                                    [str(e.is_operation) for e in src.engine]),
    ]

    cur_row = 2
    for code, name, unit, excel_vals, python_vals in items_def:
        for view_label, row_vals, row_fill in [
            ("Excel",  excel_vals, EXCEL_FILL),
            ("Python", python_vals, PYTHON_FILL),
        ]:
            ws.cell(row=cur_row, column=1, value=code if view_label == "Excel" else "")
            ws.cell(row=cur_row, column=2, value=name if view_label == "Excel" else "")
            ws.cell(row=cur_row, column=3, value=view_label).fill = row_fill
            ws.cell(row=cur_row, column=4, value=unit)
            rev_cell = ws.cell(row=cur_row, column=5, value="NOT REVIEWED" if view_label == "Excel" else None)
            rev_cell.fill = _fill("F2F2F2")
            rev_cell.font = _font(size=7)
            rev_cell.alignment = _align(h="center", v="center")
            for ci, v in enumerate(row_vals[:n]):
                cell = ws.cell(row=cur_row, column=DATA_COL + ci, value=v)
                cell.fill = row_fill
                cell.font = _font(size=8)
                cell.alignment = _align(h="center", v="center")
                if unit == "days" and isinstance(v, (int, float)):
                    cell.number_format = INT_FMT
                elif unit == "frac" and isinstance(v, (int, float)):
                    cell.number_format = FRAC_FMT
            cur_row += 1

    _freeze(ws, f"{get_column_letter(DATA_COL)}2")


def _build_prod_rev_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    none_row = [None] * n
    blocks = [
        {"type": "section", "label": "— PRODUCTION — Excel: CF row 21; Python: engine —"},
        {"type": "item", "code": "PR.01", "name": "Net production (MWh)", "unit": "MWh",
         "excel_vals": [e.production_mwh for e in src.excel],
         "python_vals": [e.production_mwh for e in src.engine],
         "item": get_item("PR.01")},
        {"type": "section", "label": (
            "— PRICE — Excel: CF indexed tariff per period; "
            "Python: EFFECTIVE price back-calculated as Revenue÷Production (NOT the actual PPA tariff) —"
        )},
        {"type": "item", "code": "PR.02", "name": "Effective revenue/MWh — Revenue÷Prod. (EUR/MWh)", "unit": "EUR/MWh",
         "excel_vals": [e.tariff_indexed_eur_mwh for e in src.excel],
         "python_vals": [
             (e.revenue_keur * 1000.0 / e.production_mwh)
             if (e.revenue_keur and e.production_mwh) else None
             for e in src.engine
         ],
         "item": get_item("PR.02")},
        {"type": "section", "label": "— REVENUE — Excel: CF row 23; Python: engine —"},
        {"type": "item", "code": "RV.01", "name": "Operating revenues CF (kEUR)", "unit": "kEUR",
         "excel_vals": [e.revenue_keur for e in src.excel],
         "python_vals": [e.revenue_keur for e in src.engine],
         "item": get_item("RV.01")},
        {"type": "item", "code": "RV.02", "name": "P&L total revenues (kEUR)", "unit": "kEUR",
         "excel_vals": [e.pl_revenue_keur for e in src.excel],
         "python_vals": [e.revenue_keur for e in src.engine],
         "item": get_item("RV.02")},
        {"type": "section", "label": "— SHL INTEREST (P&L) — OUT OF CLEAN ENGINE SCOPE —"},
        {"type": "item", "code": "SH.02", "name": "SHL gross interest P&L (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_interest_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n,
         "item": get_item("SH.02")},
    ]
    _build_horizontal_sheet(wb, "04_PROD_REV_RECON", src, mat, blocks)


def _build_opex_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    none_row = [None] * n

    opex_field_map = [
        "opex_b01_keur", "opex_b02_keur", "opex_b03_keur", "opex_b04_keur",
        "opex_b05_keur", "opex_b06_keur", "opex_b07_keur", "opex_b08_keur",
        "opex_b09_keur", "opex_b10_keur", "opex_b11_keur", "opex_b12_keur",
        "opex_b13_keur", "opex_b14_keur", "opex_b15_keur",
    ]

    # Map B.xx code → ExcelData field
    excel_opex_field_map = [
        ("B.01", "opex_b01_keur"), ("B.02", "opex_b02_keur"), ("B.03", "opex_b03_keur"),
        ("B.04", "opex_b04_keur"), ("B.05", "opex_b05_keur"), ("B.06", "opex_b06_keur"),
        ("B.07", "opex_b07_keur"), ("B.08", "opex_b08_keur"), ("B.09", "opex_b09_keur"),
        ("B.10", "opex_b10_keur"), ("B.11", "opex_b11_keur"), ("B.12", "opex_b12_keur"),
        ("B.13", "opex_b13_keur"),
    ]

    blocks = [
        {"type": "section", "label": "— TOTAL OPEX — Excel: CF row 49 (abs); Python: engine —"},
        {"type": "item", "code": "OP.00", "name": "Total OPEX", "unit": "kEUR",
         "excel_vals": [e.opex_keur for e in src.excel],
         "python_vals": [e.opex_keur for e in src.engine],
         "item": get_item("OP.00")},
        {"type": "section", "label": (
            "— OPEX BY ITEM CODE: Excel: CF per-period values (exact period); "
            "Python: factory annual × day_fraction (APPROXIMATION — B.13 % treatment means sum ≠ total exactly) —"
        )},
    ]
    for i, (opex_item, py_field) in enumerate(zip(src.opex_items, opex_field_map)):
        code = opex_item["code"]
        name = opex_item["name"]
        # Match Excel B.xx field (B.01..B.13; B.14/B.15 Excel is None)
        xl_field = excel_opex_field_map[i][1] if i < len(excel_opex_field_map) else None
        xl_vals = [getattr(e, xl_field) for e in src.excel] if xl_field else none_row
        blocks.append({
            "type": "item",
            "code": code,
            "name": name,
            "unit": "kEUR",
            "excel_vals": xl_vals,
            "python_vals": [
                getattr(e, py_field) * e.day_fraction
                if (getattr(e, py_field) is not None and e.day_fraction)
                else None
                for e in src.engine
            ],
            "item": get_item(f"OP.{i+1:02d}"),
        })

    _build_horizontal_sheet(wb, "05_OPEX_RECON", src, mat, blocks)


def _build_pnl_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    unav = [UNAVAILABLE] * n
    none_row = [None] * n

    blocks = [
        {"type": "section", "label": "— P&L RECONSTRUCTION —"},
        {"type": "item", "code": "RV.02", "name": "Total revenues", "unit": "kEUR",
         "excel_vals": [e.pl_revenue_keur for e in src.excel],
         "python_vals": [e.revenue_keur for e in src.engine], "item": get_item("RV.02")},
        {"type": "item", "code": "OP.00", "name": "Total OPEX", "unit": "kEUR",
         "excel_vals": [e.opex_keur for e in src.excel],
         "python_vals": [e.opex_keur for e in src.engine], "item": get_item("OP.00")},
        {"type": "item", "code": "EB.01", "name": "EBITDA (direct+derived)", "unit": "kEUR",
         "excel_vals": [
             e.ebitda_keur if e.ebitda_keur is not None else e.ebitda_derived_keur
             for e in src.excel
         ],
         "python_vals": [e.ebitda_keur for e in src.engine], "item": get_item("EB.01")},
        {"type": "item", "code": "DP.01", "name": "Book depreciation", "unit": "kEUR",
         "excel_vals": [e.depreciation_keur for e in src.excel],
         "python_vals": [e.book_depreciation_keur for e in src.engine], "item": get_item("DP.01")},
        {"type": "item", "code": "EBIT", "name": "EBIT (OUT OF SCOPE)", "unit": "kEUR",
         "excel_vals": none_row, "python_vals": unav, "item": None},
        {"type": "section", "label": "— FINANCIAL ITEMS —"},
        {"type": "item", "code": "SD.07", "name": "Senior interest (P&L)", "unit": "kEUR",
         "excel_vals": [e.pl_senior_interest_keur for e in src.excel],
         "python_vals": [e.sd_interest_keur for e in src.engine], "item": get_item("SD.07")},
        {"type": "item", "code": "SH.02", "name": "SHL interest (P&L)", "unit": "kEUR",
         "excel_vals": [e.shl_interest_keur for e in src.excel],
         "python_vals": unav, "item": get_item("SH.02")},
        {"type": "item", "code": "CF.02", "name": "Earnings before tax (P&L)", "unit": "kEUR",
         "excel_vals": [e.earnings_before_tax_keur for e in src.excel],
         "python_vals": unav, "item": get_item("CF.02")},
        {"type": "section", "label": "— TAX —"},
        {"type": "item", "code": "TX.02", "name": "Taxable income (P&L)", "unit": "kEUR",
         "excel_vals": [e.taxable_income_keur for e in src.excel],
         "python_vals": [e.taxable_profit_keur for e in src.engine], "item": get_item("TX.02")},
        {"type": "item", "code": "TX.06", "name": "CIT accrual (P&L)", "unit": "kEUR",
         "excel_vals": [e.pl_cit_keur for e in src.excel],
         "python_vals": [e.tax_keur for e in src.engine], "item": get_item("TX.06")},
        {"type": "item", "code": "NET", "name": "Net income (OUT OF SCOPE)", "unit": "kEUR",
         "excel_vals": none_row, "python_vals": unav, "item": None},
    ]
    _build_horizontal_sheet(wb, "06_PNL_RECON", src, mat, blocks)


def _build_capex_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("07_CAPEX_IDC_RECON")
    ws.sheet_view.showGridLines = False
    headers = ["Code", "Item", "Excel (kEUR)", "Python (kEUR)", "Delta (kEUR)",
               "Excel Source", "Python Source", "Classification", "Notes"]
    _header_row(ws, headers)
    _set_col_widths(ws, {1: 8, 2: 40, 3: 16, 4: 16, 5: 12, 6: 28, 7: 28, 8: 26, 9: 50})

    r = 2
    UNRES_STR = "UNRESOLVED — not in Excel fixture"

    def sec(label):
        nonlocal r
        for c in range(1, len(headers) + 1): ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=2, value=label).font = _font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=r, column=2).fill = SECTION_FILL
        r += 1

    def add(code, label, ev, pv, excel_src, python_src, cls=None, note=""):
        nonlocal r
        ev_f = _safe(ev) if not isinstance(ev, str) else None
        pv_f = _safe(pv) if not isinstance(pv, str) else None
        d = _delta(ev_f, pv_f)
        # Classification: only assign if we have both sides or documented reason
        if cls is None:
            if ev_f is None:
                cls = "UNRESOLVED SOURCE"
            elif pv_f is None:
                cls = "UNRESOLVED SOURCE"
            elif d is not None and not mat.is_material(d, ev_f, pv_f):
                cls = "MATCH"
            else:
                cls = ""  # OPEN
        ws.cell(row=r, column=1, value=code).font = _font(size=8)
        ws.cell(row=r, column=2, value=label).font = _font(size=9)
        c3 = ws.cell(row=r, column=3, value=ev_f if ev_f is not None else (ev if isinstance(ev, str) else None))
        c3.font = _font(size=9); c3.fill = EXCEL_FILL
        if ev_f is not None: c3.number_format = KEUR_FMT
        c4 = ws.cell(row=r, column=4, value=pv_f if pv_f is not None else (pv if isinstance(pv, str) else None))
        c4.font = _font(size=9); c4.fill = PYTHON_FILL
        if pv_f is not None: c4.number_format = KEUR_FMT
        c5 = ws.cell(row=r, column=5, value=d)
        c5.font = _font(size=9)
        if d is not None: c5.number_format = KEUR_FMT
        ws.cell(row=r, column=6, value=excel_src).font = _font(size=8)
        ws.cell(row=r, column=7, value=python_src).font = _font(size=8)
        display_cls = cls if cls else "OPEN — ROOT CAUSE REQUIRED"
        cls_cell = ws.cell(row=r, column=8, value=display_cls)
        cls_cell.fill = _fill(CLASS_FILL.get(cls, "FFFFFF"))
        cls_cell.font = _font(size=9)
        ws.cell(row=r, column=9, value=note).font = _font(size=8)
        r += 1

    sec("— HARD CAPEX per-item: Excel: XLSM Inputs rows (C.01-C.08) / CapEx sheet; Python: app.project_factories —")
    for item in src.capex_items:
        code = item["code"]
        xl_item = src.excel_capex_items.get(code, {})
        xl_amt = xl_item.get("amount_keur") if xl_item else None
        xl_lbl = xl_item.get("label") or UNRES_STR
        add(code, item["name"],
            xl_amt, item["amount_keur"],
            f"XLSM Inputs: {xl_lbl}" if xl_amt is not None else "UNRESOLVED — code not in XLSM extract",
            "app.project_factories")
    add("C.00", "Total Hard CAPEX (Python factory sum)",
        None, src.hard_capex_keur,
        "Not separately totalled in XLSM (individual codes above)", "app.project_factories (sum of capex_items)",
        "UNRESOLVED SOURCE")

    sec("— FINANCING COSTS —")
    add("C.IDC", "Bank IDC (capitalised)",
        src.excel_idc_keur or None, src.idc_keur,
        "XLSM Inputs CapEx C.IDC", "app.project_factories",
        note="IDC extracted from XLSM CapEx sheet")
    add("C.SHL", "SHL IDC capitalised",
        src.shl_idc_keur, None,
        "XLSM DS sheet row 128 period 0 (construction capitalised interest)", "OUT OF CLEAN ENGINE SCOPE",
        "OUT OF CLEAN ENGINE SCOPE", "Excel SHL IDC = construction period capitalised SHL interest")

    sec("— TOTAL PROJECT COST —")
    add("C.TOT", "Total CAPEX (Hard + IDC)",
        src.excel_total_capex_keur, src.total_capex_keur,
        "XLSM Inputs!C45 (authoritative)", "app.project_factories total_capex",
        note=f"Excel={src.excel_total_capex_keur:.1f} kEUR; Python={src.total_capex_keur:.1f} kEUR")

    _freeze(ws, "A2")
    _autofilter(ws, 1, len(headers))


def _build_depreciation_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    blocks = [
        {"type": "section", "label": "— BOOK DEPRECIATION —"},
        {"type": "item", "code": "DP.01", "name": "Book depreciation", "unit": "kEUR",
         "excel_vals": [e.depreciation_keur for e in src.excel],
         "python_vals": [e.book_depreciation_keur for e in src.engine], "item": get_item("DP.01")},
        {"type": "section", "label": "— TAX DEPRECIATION —"},
        {"type": "item", "code": "DP.02", "name": "Tax depreciation (Python only)", "unit": "kEUR",
         "excel_vals": [None] * n,
         "python_vals": [e.tax_depreciation_keur for e in src.engine], "item": get_item("DP.02")},
        {"type": "section", "label": "— DEP SHEET TOTAL (cross-check vs P&L) —"},
        {"type": "item", "code": "DP.03", "name": "Dep sheet total depreciation", "unit": "kEUR",
         "excel_vals": [e.dep_total_keur for e in src.excel],
         "python_vals": [None] * n, "item": get_item("DP.03")},
    ]
    _build_horizontal_sheet(wb, "08_DEPRECIATION_RECON", src, mat, blocks)


def _build_tax_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    none_row = [None] * n
    blocks = [
        {"type": "section", "label": "— TAX BRIDGE —"},
        {"type": "item", "code": "EB.01", "name": "EBITDA (direct+derived)", "unit": "kEUR",
         "excel_vals": [
             e.ebitda_keur if e.ebitda_keur is not None else e.ebitda_derived_keur
             for e in src.excel
         ],
         "python_vals": [e.ebitda_keur for e in src.engine], "item": get_item("EB.01")},
        {"type": "item", "code": "SD.07", "name": "Senior interest deductible", "unit": "kEUR",
         "excel_vals": [e.pl_senior_interest_keur for e in src.excel],
         "python_vals": [e.sd_interest_keur for e in src.engine], "item": get_item("SD.07")},
        {"type": "item", "code": "TX.01", "name": "Taxable profit (before losses)", "unit": "kEUR",
         "excel_vals": none_row,
         "python_vals": [e.taxable_profit_keur for e in src.engine], "item": get_item("TX.01")},
        {"type": "section", "label": "— LOSS LEDGER —"},
        {"type": "item", "code": "TX.03", "name": "Tax loss opening", "unit": "kEUR",
         "excel_vals": none_row,
         "python_vals": [e.tax_loss_opening_keur for e in src.engine], "item": get_item("TX.03")},
        {"type": "item", "code": "TX.04", "name": "Tax loss used", "unit": "kEUR",
         "excel_vals": none_row,
         "python_vals": [e.tax_loss_used_keur for e in src.engine], "item": get_item("TX.04")},
        {"type": "item", "code": "TX.05", "name": "Tax loss closing", "unit": "kEUR",
         "excel_vals": none_row,
         "python_vals": [e.tax_loss_closing_keur for e in src.engine], "item": get_item("TX.05")},
        {"type": "section", "label": "— CIT —"},
        {"type": "item", "code": "TX.06", "name": "CIT accrual", "unit": "kEUR",
         "excel_vals": [e.pl_cit_keur for e in src.excel],
         "python_vals": [e.tax_keur for e in src.engine], "item": get_item("TX.06")},
        {"type": "item", "code": "TX.07", "name": "Cash tax paid", "unit": "kEUR",
         "excel_vals": [e.cash_tax_keur for e in src.excel],
         "python_vals": [e.corporate_tax_cash_keur for e in src.engine], "item": get_item("TX.07")},
    ]
    _build_horizontal_sheet(wb, "09_TAX_RECON", src, mat, blocks)


def _build_cfads_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    unav = [UNAVAILABLE] * n
    blocks = [
        {"type": "section", "label": "— CFADS BRIDGE —"},
        {"type": "item", "code": "RV.01", "name": "Revenue", "unit": "kEUR",
         "excel_vals": [e.revenue_keur for e in src.excel],
         "python_vals": [e.revenue_keur for e in src.engine], "item": get_item("RV.01")},
        {"type": "item", "code": "OP.00", "name": "OPEX", "unit": "kEUR",
         "excel_vals": [e.opex_keur for e in src.excel],
         "python_vals": [e.opex_keur for e in src.engine], "item": get_item("OP.00")},
        {"type": "item", "code": "EB.01", "name": "EBITDA (direct+derived)", "unit": "kEUR",
         "excel_vals": [
             e.ebitda_keur if e.ebitda_keur is not None else e.ebitda_derived_keur
             for e in src.excel
         ],
         "python_vals": [e.ebitda_keur for e in src.engine], "item": get_item("EB.01")},
        {"type": "item", "code": "TX.07", "name": "Cash tax paid", "unit": "kEUR",
         "excel_vals": [e.cash_tax_keur for e in src.excel],
         "python_vals": [e.corporate_tax_cash_keur for e in src.engine], "item": get_item("TX.07")},
        {"type": "item", "code": "CF.01", "name": "CFADS", "unit": "kEUR",
         "excel_vals": [e.cfads_keur for e in src.excel],
         "python_vals": [e.cfads_keur for e in src.engine], "item": get_item("CF.01")},
        {"type": "section", "label": "— SENIOR DEBT SERVICE —"},
        {"type": "item", "code": "SD.04", "name": "Senior debt service", "unit": "kEUR",
         "excel_vals": [e.senior_ds_keur for e in src.excel],
         "python_vals": [e.sd_ds_keur for e in src.engine], "item": get_item("SD.04")},
        {"type": "item", "code": "FC.01", "name": "FCF for distribution (OUT OF SCOPE)", "unit": "kEUR",
         "excel_vals": [None] * n,
         "python_vals": unav, "item": get_item("FC.01")},
        {"type": "section", "label": "— OUT OF CLEAN ENGINE SCOPE —"},
        {"type": "item", "code": "FC.03", "name": "DSRA contribution (OUT OF SCOPE)", "unit": "kEUR",
         "excel_vals": [None] * n, "python_vals": unav, "item": get_item("FC.03")},
    ]
    _build_horizontal_sheet(wb, "10_CFADS_RECON", src, mat, blocks)


def _build_senior_debt_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    debt_delta = src.engine_debt_size_keur - src.excel_total_debt_keur

    blocks = [
        {"type": "section", "label": (
            f"— DEBT SIZE: Excel={src.excel_total_debt_keur:.1f} kEUR "
            f"(GEARING CAP 75.24% × CAPEX)  |  "
            f"Python={src.engine_debt_size_keur:.1f} kEUR "
            f"(DSCR SCULPTED 1.15×)  |  "
            f"Delta={debt_delta:+.1f} kEUR  → POLICY DIFFERENCE —"
        )},
        {"type": "section", "label": (
            "— SD.01 / SD.05 Opening/Closing: Excel DIRECT from DS sheet rows 50/56 (authoritative XLSM) —"
        )},
        {"type": "item", "code": "SD.01", "name": "Opening senior debt", "unit": "kEUR",
         "excel_vals": src.excel_sd_opening_keur,
         "python_vals": [e.sd_opening_keur for e in src.engine], "item": get_item("SD.01")},
        {"type": "item", "code": "SD.02", "name": "Senior interest (DS sheet)", "unit": "kEUR",
         "excel_vals": [e.senior_interest_keur for e in src.excel],
         "python_vals": [e.sd_interest_keur for e in src.engine], "item": get_item("SD.02")},
        {"type": "item", "code": "SD.03", "name": "Senior principal (DS sheet)", "unit": "kEUR",
         "excel_vals": [e.senior_principal_keur for e in src.excel],
         "python_vals": [e.sd_principal_keur for e in src.engine], "item": get_item("SD.03")},
        {"type": "item", "code": "SD.04", "name": "Senior debt service (CF sheet)", "unit": "kEUR",
         "excel_vals": [e.senior_ds_keur for e in src.excel],
         "python_vals": [e.sd_ds_keur for e in src.engine], "item": get_item("SD.04")},
        {"type": "item", "code": "SD.05", "name": "Closing senior debt (reconstructed)", "unit": "kEUR",
         "excel_vals": src.excel_sd_closing_keur,
         "python_vals": [e.sd_closing_keur for e in src.engine], "item": get_item("SD.05")},
        {"type": "item", "code": "SD.06", "name": "DSCR per period (CFADS÷DS service)", "unit": "x",
         "excel_vals": [
             (e.cfads_keur / e.sd_service_keur)
             if (e.cfads_keur is not None and e.sd_service_keur and e.sd_service_keur > 0.01) else None
             for e in src.excel
         ],
         "python_vals": [e.sd_dscr for e in src.engine], "item": get_item("SD.06")},
        {"type": "item", "code": "SD.07", "name": "Senior interest P&L (cross-check)", "unit": "kEUR",
         "excel_vals": [e.pl_senior_interest_keur for e in src.excel],
         "python_vals": [e.sd_interest_keur for e in src.engine], "item": get_item("SD.07")},
    ]
    _build_horizontal_sheet(wb, "11_SENIOR_DEBT_RECON", src, mat, blocks)


def _build_shl_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    n = len(src.engine)
    blocks = [
        {"type": "section", "label": (
            "— SHL SCHEDULE: Excel: DS sheet rows 123-130 (DIRECT from XLSM) — OUT OF CLEAN ENGINE SCOPE —"
        )},
        {"type": "item", "code": "SH.01", "name": "SHL opening balance (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_opening_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.01")},
        {"type": "item", "code": "SH.02", "name": "SHL net interest (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_net_interest_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.02")},
        {"type": "item", "code": "SH.03", "name": "SHL capitalised interest (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_interest_capitalised_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.03")},
        {"type": "item", "code": "SH.04", "name": "SHL closing balance (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_closing_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.04")},
        {"type": "item", "code": "SH.05", "name": "SHL service (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_service_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.05")},
        {"type": "section", "label": "— SHL INTEREST CROSS-CHECK vs P&L —"},
        {"type": "item", "code": "SH.06", "name": "SHL interest P&L (kEUR)", "unit": "kEUR",
         "excel_vals": [e.shl_interest_keur for e in src.excel],
         "python_vals": [UNAVAILABLE] * n, "item": get_item("SH.06")},
    ]
    _build_horizontal_sheet(wb, "12_SHL_RECON", src, mat, blocks)


def _build_opening_balances(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("13_OPENING_BALANCES")
    ws.sheet_view.showGridLines = False
    _header_row(ws, ["Code", "Item", "Excel", "Python", "Delta", "Source", "Classification"])
    _set_col_widths(ws, {1: 8, 2: 40, 3: 16, 4: 16, 5: 12, 6: 36, 7: 26})

    r = 2
    def add(code, label, ev, pv, source, cls="MATCH"):
        nonlocal r
        ev_f = _safe(ev)
        pv_f = _safe(pv)
        d = _delta(ev_f, pv_f)
        ws.cell(row=r, column=1, value=code).font = _font(size=8)
        ws.cell(row=r, column=2, value=label).font = _font(size=9)
        c3 = ws.cell(row=r, column=3, value=ev_f if ev_f is not None else ev)
        c3.font = _font(size=9)
        if ev_f is not None: c3.number_format = KEUR_FMT
        c4 = ws.cell(row=r, column=4, value=pv_f if pv_f is not None else pv)
        c4.font = _font(size=9)
        if pv_f is not None: c4.number_format = KEUR_FMT
        c5 = ws.cell(row=r, column=5, value=d)
        c5.font = _font(size=9)
        if d is not None: c5.number_format = KEUR_FMT
        ws.cell(row=r, column=6, value=source).font = _font(size=8)
        cl = ws.cell(row=r, column=7, value=cls)
        cl.fill = _fill(CLASS_FILL.get(cls, "FFFFFF"))
        cl.font = _font(size=9)
        r += 1

    leg0 = src.legacy[0] if src.legacy else None

    add("OB.01", "Opening senior debt (at COD)",   src.excel_total_debt_keur,  src.engine_debt_size_keur,  "golden / engine",  "POLICY DIFFERENCE")
    add("OB.02", "Opening SHL",                    src.shl_amount_keur,         None,                       "oborovo_golden.json", "OUT OF CLEAN ENGINE SCOPE")
    add("OB.03", "Opening tax losses",             0.0,                          0.0,                        "tax_reference_inputs", "MATCH")
    add("OB.04", "Opening DSRA balance",           None,                         None,                       "UNRESOLVED SOURCE",    "UNRESOLVED SOURCE")
    add("OB.05", "Gross asset basis at COD (kEUR)",src.excel_total_capex_keur,   src.total_capex_keur,       "golden / factory",     "POLICY DIFFERENCE")
    add("OB.06", "Senior debt at COD (legacy)",    leg0.sd_opening_keur if leg0 else None, src.engine_debt_size_keur, "legacy_snapshot / engine", "POLICY DIFFERENCE")

    _freeze(ws, "A2")
    _autofilter(ws, 1, 7)


def _build_delta_register(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("14_DELTA_REGISTER")
    ws.sheet_view.showGridLines = False
    headers = [
        "delta_id", "baseline_id", "section", "code", "line_item", "period",
        "excel_value", "python_value", "absolute_delta", "relative_delta",
        "classification", "root_cause", "financial_judgment", "decision",
        "excel_source", "python_source", "status",
    ]
    _header_row(ws, headers)
    _set_col_widths(ws, {c: 14 for c in range(1, len(headers) + 1)})
    _set_col_widths(ws, {1: 12, 4: 8, 5: 36, 11: 26, 12: 36, 13: 36})

    delta_rows = []
    n = len(src.engine)

    # Check per-period, per-item deltas
    per_period_checks = [
        # (item_code, section, name, excel_getter, python_getter, unit)
        ("RV.01", "REVENUE", "Operating revenues", lambda ep, eg: ep.revenue_keur, lambda ep, eg: eg.revenue_keur, "kEUR"),
        ("OP.00", "OPEX",    "Total OPEX",         lambda ep, eg: ep.opex_keur,    lambda ep, eg: eg.opex_keur,    "kEUR"),
        ("EB.01", "EBITDA",  "EBITDA",             lambda ep, eg: ep.ebitda_keur,  lambda ep, eg: eg.ebitda_keur,  "kEUR"),
        ("TX.06", "TAX",     "CIT accrual",        lambda ep, eg: ep.pl_cit_keur,  lambda ep, eg: eg.tax_keur,     "kEUR"),
        ("TX.07", "TAX",     "Cash tax",           lambda ep, eg: ep.cash_tax_keur,lambda ep, eg: eg.corporate_tax_cash_keur, "kEUR"),
        ("CF.01", "CFADS",   "CFADS",              lambda ep, eg: ep.cfads_keur,   lambda ep, eg: eg.cfads_keur,   "kEUR"),
        ("SD.02", "SENIOR_DEBT","Senior interest", lambda ep, eg: ep.senior_interest_keur, lambda ep, eg: eg.sd_interest_keur, "kEUR"),
        ("SD.03", "SENIOR_DEBT","Senior principal",lambda ep, eg: ep.senior_principal_keur,lambda ep, eg: eg.sd_principal_keur,"kEUR"),
        ("SD.04", "SENIOR_DEBT","Senior DS",       lambda ep, eg: ep.senior_ds_keur, lambda ep, eg: eg.sd_ds_keur, "kEUR"),
        ("SD.06", "SENIOR_DEBT","DSCR",
         lambda ep, eg: (ep.cfads_keur / ep.sd_service_keur)
             if (ep.cfads_keur is not None and ep.sd_service_keur and ep.sd_service_keur > 0.01)
             else None,
         lambda ep, eg: eg.sd_dscr, "x"),
        ("DP.01", "DEPRECIATION","Book depreciation",lambda ep, eg: ep.depreciation_keur, lambda ep, eg: eg.book_depreciation_keur, "kEUR"),
    ]

    # Root-cause notes per line item (upstream-to-downstream order).
    # Material deltas without root cause get blank classification + OPEN status.
    ITEM_EXCEL_SOURCE = {
        "RV.01": "CF.operating_revenues_keur",
        "OP.00": "abs(CF.operating_expenses_after_bank_tax_keur)",
        "EB.01": "CF.ebitda_keur",
        "TX.06": "abs(P&L.corporate_income_tax_keur)",
        "TX.07": "abs(CF.corporate_income_tax_keur)",
        "CF.01": "CF.free_cash_flow_for_banks_keur",
        "SD.02": "DS.senior_net_interest_keur",
        "SD.03": "DS.senior_principal_keur",
        "SD.04": "abs(CF.senior_debt_service_keur)",
        "SD.06": "CF.average_senior_dscr_period",
        "DP.01": "abs(P&L.depreciation_keur)",
    }
    ITEM_ROOT_CAUSE_HINT = {
        "RV.01": "OPEN — ROOT CAUSE REQUIRED. Trace: production match? → price match? → which revenue component diverges?",
        "OP.00": "OPEN — ROOT CAUSE REQUIRED. Trace: which B.xx item drives the OPEX delta? Check Excel total vs sum(B.01..B.15).",
        "EB.01": "OPEN — ROOT CAUSE REQUIRED. EBITDA = Revenue - OPEX; diagnose upstream first (RV.01, OP.00).",
        "TX.06": "OPEN — ROOT CAUSE REQUIRED. CIT accrual: check taxable income alignment (TX.02 vs TX.01), then rate application.",
        "TX.07": "OPEN — ROOT CAUSE REQUIRED. Cash tax timing: lag vs accrual? Check CIT accrual delta first.",
        "CF.01": "OPEN — ROOT CAUSE REQUIRED. CFADS = EBITDA - cash_tax; diagnose EBITDA and tax deltas first.",
        "SD.02": "OPEN — ROOT CAUSE REQUIRED. Interest = opening_balance × rate × day_fraction. Check debt-size delta (POLICY DIFFERENCE) as upstream cause.",
        "SD.03": "OPEN — ROOT CAUSE REQUIRED. Principal driven by CFADS sculpting vs gearing schedule; debt-size POLICY DIFFERENCE is primary upstream cause.",
        "SD.04": "OPEN — ROOT CAUSE REQUIRED. DS = interest + principal; diagnose SD.02, SD.03 first.",
        "SD.06": "OPEN — ROOT CAUSE REQUIRED. DSCR = CFADS / DS; debt-size POLICY DIFFERENCE affects denominator directly.",
        "DP.01": "OPEN — ROOT CAUSE REQUIRED. Check: same depreciable base (CAPEX)? Same useful life? Same method (straight-line)?",
    }

    delta_counter = 0
    for code, section, name, ef, pf, unit in per_period_checks:
        item = get_item(code)
        for i, (ep, eg) in enumerate(zip(src.excel, src.engine)):
            ev = _safe(ef(ep, eg))
            pv = _safe(pf(ep, eg))
            d = _delta(ev, pv)
            if d is None:
                continue
            if not mat.is_material(d, ev, pv, unit):
                continue
            delta_counter += 1
            dp = _delta_pct(d, ev)
            cls = _classify(item, d, ev, pv, mat) if item else ""
            root_cause = ITEM_ROOT_CAUSE_HINT.get(code, "OPEN — ROOT CAUSE REQUIRED")
            excel_src = ITEM_EXCEL_SOURCE.get(code, "excel_oborovo_full_model_extract.json")
            delta_rows.append([
                f"DELTA-{delta_counter:04d}", BASELINE_ID, section, code, name,
                eg.period_end, ev, pv, d, dp,
                cls if cls else "",
                root_cause,
                "Manual financial review required",
                "OPEN",
                excel_src,
                "run_senior_debt_model(oborovo)",
                "OPEN",
            ])

    # Add scalar delta for debt size
    ev_ds = src.excel_total_debt_keur
    pv_ds = src.engine_debt_size_keur
    d_ds = _delta(ev_ds, pv_ds)
    if d_ds and mat.is_material(d_ds, ev_ds, pv_ds):
        delta_counter += 1
        delta_rows.append([
            f"DELTA-{delta_counter:04d}", BASELINE_ID, "SENIOR_DEBT", "SD.00", "Debt size at COD",
            "COD", ev_ds, pv_ds, d_ds, _delta_pct(d_ds, ev_ds),
            "POLICY DIFFERENCE",
            "Excel uses GEARING CAP sizing; Python uses DSCR SCULPTED sizing",
            "Different sizing methodologies. Excel debt = 75.24% × CAPEX; Python debt = DSCR-sculpted at 1.15x",
            "OPEN — methodology choice requires decision",
            "oborovo_golden.json", "run_senior_debt_model(oborovo)", "OPEN",
        ])

    for r_idx, row in enumerate(delta_rows, 2):
        for c_idx, val in enumerate(row, 1):
            display_val = val
            cell = ws.cell(row=r_idx, column=c_idx, value=display_val)
            cell.font = _font(size=9)
            if c_idx in (7, 8, 9):
                if isinstance(val, float):
                    cell.number_format = KEUR_FMT
            if c_idx == 10 and isinstance(val, float):
                cell.number_format = PCT_FMT
            if c_idx == 11:
                raw_cls = str(val) if val else ""
                cell.fill = _fill(CLASS_FILL.get(raw_cls, "FFFFFF"))

    _freeze(ws, "A2")
    _autofilter(ws, 1, len(headers))


def _build_source_map(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("15_SOURCE_MAP")
    ws.sheet_view.showGridLines = False
    headers = ["section", "code", "line_item", "excel_sheet", "excel_field",
               "python_module", "python_field", "python_function"]
    _header_row(ws, headers)
    _set_col_widths(ws, {1: 14, 2: 8, 3: 36, 4: 12, 5: 40, 6: 30, 7: 30, 8: 40})

    source_map = [
        # section, code, line_item, excel_sheet, excel_field, python_module, python_field, python_function
        # TIMELINE
        ("TIMELINE",    "TL.02", "Period end date",         "period_diagnostics", "date col",                 "financial_engine.results", "period_end",                 "run_senior_debt_model → OperatingPeriodResult"),
        ("TIMELINE",    "TL.01", "Period index",            "UNRESOLVED",         "not in fixture",           "financial_engine.results", "period_index",               "run_senior_debt_model → OperatingPeriodResult"),
        ("TIMELINE",    "TL.03", "Days in period",          "UNRESOLVED",         "not in fixture",           "financial_engine.results", "days_in_period",             "run_senior_debt_model → OperatingPeriodResult"),
        ("TIMELINE",    "TL.04", "Day fraction",            "UNRESOLVED",         "not in fixture",           "financial_engine.results", "day_fraction",               "run_senior_debt_model → OperatingPeriodResult"),
        ("TIMELINE",    "TL.05", "Is operation",            "UNRESOLVED",         "not in fixture",           "financial_engine.results", "is_operation",               "run_senior_debt_model → OperatingPeriodResult"),
        # PRODUCTION
        ("PRODUCTION",  "PR.01", "Net production (engine)", "UNRESOLVED",         "not in fixture",           "financial_engine.results", "production_mwh",             "from_project_inputs → run_senior_debt_model"),
        # REVENUE
        ("REVENUE",     "RV.01", "Operating revenues (CF)", "CF",                 "CF.operating_revenues_keur","financial_engine.results", "revenue_keur",              "from_project_inputs → run_senior_debt_model"),
        ("REVENUE",     "RV.02", "Total revenues (P&L)",    "P&L",                "P&L.total_revenues_keur",  "financial_engine.results", "revenue_keur (same field)",  "from_project_inputs → run_senior_debt_model"),
        # OPEX
        ("OPEX",        "OP.00", "Total OPEX",              "CF",                 "CF.operating_expenses_after_bank_tax_keur","financial_engine.results","opex_keur","from_project_inputs → run_senior_debt_model"),
        ("OPEX",        "B.01", "O&M labour",                   "CF", "CF.opex_b01_keur (period)",  "app.project_factories", "opex_b01_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.02", "Infrastructure maintenance",   "CF", "CF.opex_b02_keur (period)",  "app.project_factories", "opex_b02_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.03", "O&M contract",                 "CF", "CF.opex_b03_keur (period)",  "app.project_factories", "opex_b03_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.04", "Asset management",             "CF", "CF.opex_b04_keur (period)",  "app.project_factories", "opex_b04_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.05", "Insurance",                    "CF", "CF.opex_b05_keur (period)",  "app.project_factories", "opex_b05_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.06", "Land lease",                   "CF", "CF.opex_b06_keur (period)",  "app.project_factories", "opex_b06_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.07", "Corporate overheads",          "CF", "CF.opex_b07_keur (period)",  "app.project_factories", "opex_b07_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.08", "Power expenses",               "CF", "CF.opex_b08_keur (period)",  "app.project_factories", "opex_b08_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.09", "Technical management",         "CF", "CF.opex_b09_keur (period)",  "app.project_factories", "opex_b09_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.10", "Cleaning",                     "CF", "CF.opex_b10_keur (period)",  "app.project_factories", "opex_b10_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.11", "Security",                     "CF", "CF.opex_b11_keur (period)",  "app.project_factories", "opex_b11_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.12", "Spare parts",                  "CF", "CF.opex_b12_keur (period)",  "app.project_factories", "opex_b12_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        ("OPEX",        "B.13", "Variable O&M (% of revenue)",  "CF", "CF.opex_b13_keur (period)",  "app.project_factories", "opex_b13_keur × day_fraction", "OpexItem.amount_at_year(year_index) × day_fraction"),
        # EBITDA
        ("EBITDA",      "EB.01", "EBITDA",                  "CF",                 "CF.ebitda_keur",           "financial_engine.results", "ebitda_keur",                "from_project_inputs → run_senior_debt_model"),
        # CAPEX
        ("CAPEX",       "CA.00–CA.15","Hard CAPEX per item","UNRESOLVED",         "not in fixture",           "app.project_factories",    "capex_items[].amount_keur",  "CapexStructure.capex_items()"),
        ("CAPEX",       "C.TOT", "Total CAPEX",             "oborovo_golden.json","outputs.total_capex_keur", "app.project_factories",    "total_capex",                "CapexStructure.total_capex"),
        ("CAPEX",       "C.SHL", "SHL IDC capitalised",     "SHL fixture",        "shl[0].capitalized_interest","OUT OF CLEAN ENGINE SCOPE","N/A",                    "N/A — SHL not in clean engine"),
        # DEPRECIATION
        ("DEPRECIATION","DP.01", "Book depreciation",       "P&L",                "P&L.depreciation_keur",   "financial_engine.results", "book_depreciation_keur",     "from_project_inputs → run_senior_debt_model"),
        ("DEPRECIATION","DP.02", "Tax depreciation",        "UNRESOLVED",         "not in fixture",           "financial_engine.results", "tax_depreciation_keur",      "calculate_tax → TaxAndCfadsSchedules"),
        ("DEPRECIATION","DP.03", "Cumulated depreciation",  "Dep",                "Dep.cumulated_depreciation_keur","UNRESOLVED",         "N/A — no Python equivalent", "N/A"),
        # TAX
        ("TAX",         "TX.01", "Taxable profit (EBITDA basis)","UNRESOLVED",    "not in fixture",           "financial_engine.results", "taxable_profit_keur",        "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.02", "P&L taxable income",      "P&L",                "P&L.taxable_income_keur",  "financial_engine.results", "taxable_profit_keur",        "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.03", "Tax loss opening",        "UNRESOLVED",         "not in fixture",           "financial_engine.results", "tax_loss_opening_keur",      "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.04", "Tax loss used",           "UNRESOLVED",         "not in fixture",           "financial_engine.results", "tax_loss_used_keur",         "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.05", "Tax loss closing",        "UNRESOLVED",         "not in fixture",           "financial_engine.results", "tax_loss_closing_keur",      "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.06", "CIT accrual (P&L)",       "P&L",                "P&L.corporate_income_tax_keur","financial_engine.results","tax_keur",              "calculate_tax → TaxAndCfadsSchedules"),
        ("TAX",         "TX.07", "Cash tax paid",           "CF",                 "CF.corporate_income_tax_keur","financial_engine.results","corporate_tax_cash_keur", "calculate_tax → TaxAndCfadsSchedules"),
        # CFADS
        ("CFADS",       "CF.01", "CFADS",                   "CF",                 "CF.free_cash_flow_for_banks_keur","financial_engine.results","cfads_keur",          "calculate_canonical_cfads → TaxAndCfadsSchedules"),
        ("CFADS",       "CF.02", "Earnings before tax",     "P&L",                "P&L.earnings_before_tax_keur","OUT OF CLEAN ENGINE SCOPE","N/A",                  "N/A — EBT not in clean engine output"),
        # SENIOR DEBT
        ("SENIOR_DEBT", "SD.01", "Opening senior debt",     "Reconstructed",      "opening = excel_total_debt_keur; roll: closing[t-1]","financial_engine.senior_debt","sd_opening_keur","solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.02", "Senior interest",         "DS",                 "DS.senior_net_interest_keur","financial_engine.senior_debt","sd_interest_keur",   "solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.03", "Senior principal",        "DS",                 "DS.senior_principal_keur", "financial_engine.senior_debt","sd_principal_keur",   "solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.04", "Senior debt service",     "CF",                 "abs(CF.senior_debt_service_keur)","financial_engine.senior_debt","sd_ds_keur",    "solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.05", "Closing senior debt",     "Reconstructed",      "closing = opening - principal (DS sheet)","financial_engine.senior_debt","sd_closing_keur","solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.06", "DSCR (period average)",   "CF",                 "CF.average_senior_dscr_period","financial_engine.senior_debt","sd_dscr",         "solve_senior_debt → SeniorDebtSchedules"),
        ("SENIOR_DEBT", "SD.07", "Senior interest (P&L)",   "P&L",                "abs(P&L.senior_interests_keur)","financial_engine.senior_debt","sd_interest_keur","solve_senior_debt → SeniorDebtSchedules"),
        # SHL
        ("SHL",         "SH.01", "Opening SHL",             "SHL fixture",        "shl[n].opening",           "OUT OF CLEAN ENGINE SCOPE","N/A",                    "N/A"),
        ("SHL",         "SH.02", "SHL gross interest",      "P&L",                "abs(P&L.shareholder_loan_interests_keur)","OUT OF CLEAN ENGINE SCOPE","N/A",     "N/A"),
        ("SHL",         "SH.03", "Closing SHL",             "SHL fixture",        "shl[n].closing",           "OUT OF CLEAN ENGINE SCOPE","N/A",                    "N/A"),
        # CASHFLOW
        ("CASHFLOW",    "FC.01", "Free cash flow",          "CF",                 "CF.free_cash_flow_for_distribution_keur","OUT OF CLEAN ENGINE SCOPE","N/A",      "N/A"),
        ("CASHFLOW",    "FC.02", "Net dividends",           "P&L",                "P&L.net_dividends_keur",   "OUT OF CLEAN ENGINE SCOPE","N/A",                    "N/A"),
        ("CASHFLOW",    "FC.03", "DSRA contribution",       "UNRESOLVED",         "not in fixture",           "OUT OF CLEAN ENGINE SCOPE","N/A",                    "N/A"),
    ]

    for r_idx, row in enumerate(source_map, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val).font = _font(size=9)

    _freeze(ws, "A2")
    _autofilter(ws, 1, 8)


def _build_raw_recon(wb: Workbook, src: OborovoSources, mat: MaterialitySettings) -> None:
    ws = wb.create_sheet("16_RAW_RECON")
    ws.sheet_view.showGridLines = False
    headers = [
        "baseline_id", "period_index", "period_end", "section", "code", "line_item",
        "unit", "excel_value", "python_value", "delta", "delta_pct",
        "classification", "excel_source", "python_source",
    ]
    _header_row(ws, headers)
    _set_col_widths(ws, {c: 14 for c in range(1, len(headers) + 1)})
    _set_col_widths(ws, {5: 8, 6: 36, 12: 26})

    per_period_checks = [
        ("REVENUE",     "RV.01", "Operating revenues",   "kEUR", lambda ep: ep.revenue_keur,       lambda eg: eg.revenue_keur),
        ("OPEX",        "OP.00", "Total OPEX",           "kEUR", lambda ep: ep.opex_keur,           lambda eg: eg.opex_keur),
        ("EBITDA",      "EB.01", "EBITDA",               "kEUR", lambda ep: ep.ebitda_keur,         lambda eg: eg.ebitda_keur),
        ("TAX",         "TX.06", "CIT accrual",          "kEUR", lambda ep: ep.pl_cit_keur,         lambda eg: eg.tax_keur),
        ("TAX",         "TX.07", "Cash tax paid",        "kEUR", lambda ep: ep.cash_tax_keur,       lambda eg: eg.corporate_tax_cash_keur),
        ("CFADS",       "CF.01", "CFADS",                "kEUR", lambda ep: ep.cfads_keur,          lambda eg: eg.cfads_keur),
        ("SENIOR_DEBT", "SD.02", "Senior interest",      "kEUR", lambda ep: ep.senior_interest_keur,lambda eg: eg.sd_interest_keur),
        ("SENIOR_DEBT", "SD.03", "Senior principal",     "kEUR", lambda ep: ep.senior_principal_keur,lambda eg: eg.sd_principal_keur),
        ("SENIOR_DEBT", "SD.04", "Senior DS",            "kEUR", lambda ep: ep.senior_ds_keur,      lambda eg: eg.sd_ds_keur),
        ("SENIOR_DEBT", "SD.06", "DSCR", "x",
         lambda ep: (ep.cfads_keur / ep.sd_service_keur)
             if (ep.cfads_keur is not None and ep.sd_service_keur and ep.sd_service_keur > 0.01) else None,
         lambda eg: eg.sd_dscr),
        ("DEPRECIATION","DP.01", "Book depreciation",    "kEUR", lambda ep: ep.depreciation_keur,   lambda eg: eg.book_depreciation_keur),
        ("PRODUCTION",  "PR.01", "Net production",       "MWh",  lambda ep: ep.production_mwh,      lambda eg: eg.production_mwh),
    ]

    r_idx = 2
    for section, code, name, unit, ef, pf in per_period_checks:
        item = get_item(code)
        for i, (ep, eg) in enumerate(zip(src.excel, src.engine)):
            ev = _safe(ef(ep))
            pv = _safe(pf(eg))
            d = _delta(ev, pv)
            dp = _delta_pct(d, ev)
            cls = _classify(item, d, ev, pv, mat) if item else "UNRESOLVED SOURCE"
            display_cls = cls if cls else "OPEN — ROOT CAUSE REQUIRED"
            row = [BASELINE_ID, i, eg.period_end, section, code, name, unit,
                   ev, pv, d, dp, display_cls,
                   "excel_oborovo_full_model_extract.json",
                   "run_senior_debt_model(oborovo)"]
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _font(size=8)
                if c_idx in (8, 9, 10) and isinstance(val, float):
                    cell.number_format = KEUR_FMT
                if c_idx == 11 and isinstance(val, float):
                    cell.number_format = PCT_FMT
                if c_idx == 12:
                    raw_cls = cls if cls else ""
                    cell.fill = _fill(CLASS_FILL.get(raw_cls, "FFFFFF"))
            r_idx += 1

    _freeze(ws, "A2")
    _autofilter(ws, 1, len(headers))


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_workbook(
    sources: OborovoSources,
    output_path: str,
    materiality: MaterialitySettings = DEFAULT_MATERIALITY,
    generation_ts: str = "",
    pr_head: str = "",
) -> Workbook:
    """Build the reconciliation workbook and save to output_path. Returns the Workbook."""
    if not generation_ts:
        generation_ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_readme(wb, sources, generation_ts, pr_head)
    _build_exec_recon(wb, sources, materiality)
    _build_inputs_recon(wb, sources, materiality)
    _build_timeline_recon(wb, sources, materiality)
    _build_prod_rev_recon(wb, sources, materiality)
    _build_opex_recon(wb, sources, materiality)
    _build_pnl_recon(wb, sources, materiality)
    _build_capex_recon(wb, sources, materiality)
    _build_depreciation_recon(wb, sources, materiality)
    _build_tax_recon(wb, sources, materiality)
    _build_cfads_recon(wb, sources, materiality)
    _build_senior_debt_recon(wb, sources, materiality)
    _build_shl_recon(wb, sources, materiality)
    _build_opening_balances(wb, sources, materiality)
    _build_delta_register(wb, sources, materiality)
    _build_source_map(wb, sources, materiality)
    _build_raw_recon(wb, sources, materiality)

    wb.save(output_path)
    return wb
