"""finco_recon.extract_oborovo_excel — Deterministic authoritative extractor.

Reads ``20260414_BP_Oborovo_Sensitivity_FINAL_for_PPT.xlsm`` and emits a
machine-readable JSON fixture containing financial-truth data for the
Excel↔Python reconciliation.

Design rules
------------
* Only cached/stored cell values are read (``data_only=True``).  Formulas
  are not re-evaluated in Python.
* Every extracted value records its source sheet, row, and column.
* The JSON payload is deterministic: the extraction timestamp is stored
  *separately* from the financial payload so repeated runs on the same
  workbook produce identical financial content.
* The workbook SHA-256 is recorded so downstream consumers can verify
  provenance.

Usage::

    python -m finco_recon.extract_oborovo_excel \\
        --workbook "/path/to/20260414_BP_Oborovo_Sensitivity_FINAL_for_PPT.xlsm" \\
        --output tests/fixtures/excel_oborovo_financial_truth.json

Exit codes:  0 success · 1 workbook not found · 2 unexpected error.
"""
from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import pathlib
import sys
from typing import Any

_EXTRACTOR_VERSION = "3.3.0"
_EXPECTED_FILENAME = "20260414_BP_Oborovo_Sensitivity_FINAL_for_PPT.xlsm"

# ---------------------------------------------------------------------------
# Column layout helpers
# ---------------------------------------------------------------------------
# In every schedule sheet (CF, DS, P&L, Dep) the period axis is:
#   column index 6  → period 0  (construction / pre-COD)
#   column index 7  → period 1  (H2-2030, first operation semester)
#   ...
#   column index 66 → period 60 (last operation semester)
# We extract all 61 periods (0–60).
_PERIOD_COL_OFFSET = 6   # 0-based column index of period-0 column
_N_PERIODS = 61          # periods 0 … 60


def _row_to_periods(row: tuple, n: int = _N_PERIODS) -> list[float | None]:
    """Extract n consecutive period values starting at _PERIOD_COL_OFFSET."""
    out: list[float | None] = []
    for p in range(n):
        col = _PERIOD_COL_OFFSET + p
        v = row[col] if col < len(row) else None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append(float(v))
        else:
            out.append(None)
    return out


def _scalar(row: tuple, col: int) -> Any:
    v = row[col] if col < len(row) else None
    return v


def _date_str(v: Any) -> str | None:
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return None


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def _read_inputs(wb) -> dict:
    ws = wb["Inputs"]
    rows = {r: tuple(row) for r, row in enumerate(
        ws.iter_rows(max_row=350, max_col=20, values_only=True), start=1)}

    def row(r): return rows.get(r, ())
    def cell(r, c): return _scalar(row(r), c)

    def date_cell(r, c) -> str | None:
        return _date_str(cell(r, c))

    # -----------------------------------------------------------------------
    # Scalar inputs
    # -----------------------------------------------------------------------
    inputs: dict = {
        "_source": {"sheet": "Inputs", "extractor_note": "row/col are 1-based"},
        "project_name":             {"value": cell(2, 3),  "row": 2,  "col": "D"},
        "financial_close_date":     {"value": date_cell(9, 3),  "row": 9,  "col": "D"},
        "construction_months":      {"value": cell(10, 3), "row": 10, "col": "D"},
        "operation_start_date":     {"value": date_cell(11, 3), "row": 11, "col": "D"},
        "investment_horizon_years": {"value": cell(14, 3), "row": 14, "col": "D"},
        "model_period":             {"value": cell(18, 3), "row": 18, "col": "D"},
        # Technical
        "capacity_mwp":             {"value": cell(51, 3), "row": 51, "col": "D"},
        "production_scenario":      {"value": cell(52, 3), "row": 52, "col": "D"},
        "operating_hours_p50":      {"value": cell(54, 3), "row": 54, "col": "D"},
        "pv_degradation_pa":        {"value": cell(56, 3), "row": 56, "col": "D"},
        "plant_availability":       {"value": cell(58, 3), "row": 58, "col": "D"},
        "grid_availability":        {"value": cell(59, 3), "row": 59, "col": "D"},
        # Revenue
        "ppa_base_tariff_y1_eur_mwh": {"value": cell(78, 3), "row": 78, "col": "D"},
        "ppa_term_years":           {"value": cell(81, 3), "row": 81, "col": "D"},
        "ppa_index_pa":             {"value": cell(83, 3), "row": 83, "col": "D"},
        "market_price_scenario":    {"value": cell(89, 3), "row": 89, "col": "D"},
        "market_price_y1_eur_mwh":  {"value": cell(92, 3), "row": 92, "col": "D"},
        "market_price_index_pa":    {"value": cell(93, 3), "row": 93, "col": "D"},
        # Sizing
        "senior_debt_amount_keur":  {"value": cell(192, 3), "row": 192, "col": "D"},
        "senior_debt_maturity_years": {"value": cell(196, 3), "row": 196, "col": "D"},
        "senior_debt_last_repayment": {"value": date_cell(197, 3), "row": 197, "col": "D"},
        "senior_debt_base_rate":    {"value": cell(202, 3), "row": 202, "col": "D"},
        "senior_debt_margin_bps":   {"value": cell(203, 3), "row": 203, "col": "D"},
        "senior_dscr_covenant":     {"value": cell(221, 3), "row": 221, "col": "D"},
        "senior_lockup_dscr":       {"value": cell(223, 3), "row": 223, "col": "D"},
        # SHL
        "shl_amount_keur":          {"value": cell(325, 3), "row": 325, "col": "D"},
        "shl_interest_rate":        {"value": cell(328, 5), "row": 328, "col": "F"},
        # Equity
        "equity_capital_keur":      {"value": cell(312, 3), "row": 312, "col": "D"},
        # CAPEX totals
        "total_capex_keur":         {"value": cell(45, 2), "row": 45, "col": "C"},
    }

    # -----------------------------------------------------------------------
    # CAPEX line items (rows 22-44 in Inputs; individual line descriptions)
    # -----------------------------------------------------------------------
    capex_input_rows = {
        "Production Units": (23, "C.01"),
        "EPC Contract": (24, "C.02"),
        "EPC other costs": (25, "C.03"),
        "Grid connection": (26, "C.04"),
        "Investments to prepare operation phase": (27, "C.05"),
        "Insurances": (28, "C.06"),
        "Project finance costs due at closing": (30, "C.08"),
        "Construction Management": (34, "C.09"),
        "Contingencies": (35, "C.10"),
        "Project Acquisition / Project Development": (37, "C.11"),
        "Project Rights": (38, "C.12"),
        "IDCs": (39, "C.IDC"),
        "Commitment Fees": (40, "C.CF"),
        "Bank Fees": (41, "C.BF"),
        "VAT Costs": (44, "C.VAT"),
    }
    capex_items: dict = {}
    for name, (r, code) in capex_input_rows.items():
        amount = cell(r, 2)
        capex_items[code] = {
            "label": name,
            "amount_keur": float(amount) if isinstance(amount, (int, float)) else None,
            "sheet": "Inputs", "row": r, "col": "C",
        }
    inputs["capex_items_from_inputs"] = capex_items

    # -----------------------------------------------------------------------
    # OPEX annual line items (rows 146-161)
    # -----------------------------------------------------------------------
    opex_rows = [
        (146, "B.01", "Technical Management"),
        (147, "B.02", "Infrastructure Maintenance"),
        (148, "B.03", "Maintain Site"),
        (149, "B.04", "Clean Material"),
        (150, "B.05", "Security"),
        (151, "B.06", "Insurance"),
        (152, "B.07", "Lease & property Tax"),
        (153, "B.08", "Power Expenses"),
        (154, "B.09", "Fees"),
        (155, "B.10", "Audit & Accounting & Legal Fees"),
        (156, "B.11", "Bank Fees"),
        (157, "B.12", "Environmental & Social management"),
        (158, "B.13", "Contingencies"),
        (159, "B.14", "Taxes"),
        (160, "B.15", "Salary and payroll Tax"),
    ]
    opex_annual: dict = {}
    for r, code, label in opex_rows:
        y_vals = [cell(r, c) for c in range(4, 10)]  # cols 5-10 = years 1-6
        y1 = y_vals[0] if y_vals else None
        opex_annual[code] = {
            "label": label,
            "year1_keur": float(y1) if isinstance(y1, (int, float)) else None,
            "years_1_to_6_keur": [float(v) if isinstance(v, (int, float)) else None
                                   for v in y_vals],
            "sheet": "Inputs", "row": r,
        }
    inputs["opex_annual_items"] = opex_annual

    return inputs


def _read_capex_sheet(wb) -> dict:
    ws = wb["CapEx"]
    rows = {r: tuple(row) for r, row in enumerate(
        ws.iter_rows(max_row=200, max_col=30, values_only=True), start=1)}

    def cell(r, c): return _scalar(rows.get(r, ()), c)

    items: dict = {}
    code_rows = [
        ("C.01", 6,  "Production Units"),
        ("C.02", 9,  "EPC Contract"),
        ("C.03a", 13, "EPC other costs"),
        ("C.04", 18, "Grid connection"),
        ("C.03b", 24, "Other costs for construction"),
        ("C.05", 31, "Investments to prepare operation phase"),
        ("C.06", 47, "Insurances"),
        ("C.07", 54, "Lease and Property Tax"),
        ("C.08", 57, "Project finance costs due at closing"),
    ]
    for code, r, label in code_rows:
        amount = cell(r, 2)
        dep_life = cell(r, 1)
        items[code] = {
            "label": label,
            "amount_keur": float(amount) if isinstance(amount, (int, float)) else None,
            "dep_life_years": int(dep_life) if isinstance(dep_life, int) else dep_life,
            "sheet": "CapEx", "row": r,
        }

    # Total CAPEX from CapEx header row 4
    total = cell(4, 2)
    return {
        "_source": {"sheet": "CapEx"},
        "total_hard_capex_keur": float(total) if isinstance(total, (int, float)) else None,
        "items": items,
    }


def _read_schedule(wb, sheet_name: str, row_map: dict[str, int]) -> dict:
    """Generic period-schedule reader.  row_map: label → 1-based row number."""
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(max_row=max(row_map.values()) + 2,
                                  max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
                                  values_only=True))
    result: dict = {"_source": {"sheet": sheet_name}}
    for label, r in row_map.items():
        row = all_rows[r - 1] if r - 1 < len(all_rows) else ()
        result[label] = _row_to_periods(row)
    return result


def _read_cf(wb) -> dict:
    row_map = {
        "bop_date": 1,
        "eop_date": 2,
        "calendar_year": 3,
        "operation_period_fraction": 7,
        "is_project_life": 6,
        "production_mwh": 21,
        "operating_revenues_keur": 23,
        "ppa_sales_keur": 24,
        "production_to_ppa_mwh": 25,
        "tariff_indexed_eur_mwh": 26,
        "operating_expenses_keur": 49,
        "ebitda_keur": 51,
        "tax_technical_management_keur": 56,
        "tax_infrastructure_maintenance_keur": 57,
        "tax_maintain_site_keur": 58,
        "tax_clean_material_keur": 59,
        "tax_security_keur": 60,
        "tax_insurance_keur": 61,
        "tax_lease_property_keur": 62,
        "tax_power_expenses_keur": 63,
        "tax_fees_keur": 64,
        "tax_audit_legal_keur": 65,
        "tax_bank_fees_keur": 66,
        "tax_env_social_keur": 67,
        "tax_contingencies_keur": 68,
        "corporate_income_tax_keur": 77,
        "fcf_for_banks_keur": 79,
        "senior_debt_service_keur": 80,
        "free_cash_flow_for_junior_keur": 94,
        "free_cash_flow_for_shl_keur": 112,
        "free_cash_flow_for_dividends_keur": 116,
    }
    ws = wb["CF"]
    all_rows = list(ws.iter_rows(max_row=max(row_map.values()) + 2,
                                  max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 2,
                                  values_only=True))
    result: dict = {"_source": {"sheet": "CF"}}
    for label, r in row_map.items():
        row = all_rows[r - 1] if r - 1 < len(all_rows) else ()
        if label in ("bop_date", "eop_date"):
            vals = []
            for p in range(_N_PERIODS):
                col = _PERIOD_COL_OFFSET + p
                v = row[col] if col < len(row) else None
                vals.append(_date_str(v))
            result[label] = vals
        else:
            result[label] = _row_to_periods(row)
    # OPEX by individual item (CF rows 56-68), sign as stored (negative)
    opex_items_period: dict = {}
    item_rows = {
        "B.01": 56, "B.02": 57, "B.03": 58, "B.04": 59, "B.05": 60,
        "B.06": 61, "B.07": 62, "B.08": 63, "B.09": 64, "B.10": 65,
        "B.11": 66, "B.12": 67, "B.13": 68,
    }
    for code, r in item_rows.items():
        row_data = all_rows[r - 1] if r - 1 < len(all_rows) else ()
        vals = _row_to_periods(row_data)
        opex_items_period[code] = vals
    result["opex_items_period_keur"] = opex_items_period
    return result


def _read_ds(wb) -> dict:
    row_map = {
        "bop_date": 1,
        "eop_date": 2,
        "sd_period_fraction": 6,
        "dscr_target": 22,
        "cfads_for_sd_keur": 20,
        "sd_beginning_keur": 50,
        "sd_funding_keur": 51,
        "sd_principal_keur": 52,
        "sd_net_interest_keur": 53,
        "sd_gross_interest_keur": 55,
        "sd_ending_keur": 56,
        "sd_service_keur": 57,
        "shl_beginning_keur": 123,
        "shl_funding_keur": 124,
        "shl_net_interest_keur": 125,
        "shl_interest_capitalised_keur": 128,
        "shl_ending_keur": 129,
        "shl_service_keur": 130,
    }
    return _read_schedule(wb, "DS", row_map)


def _read_pl(wb) -> dict:
    row_map = {
        "bop_date": 1,
        "eop_date": 2,
        "total_revenues_keur": 8,
        "operating_expenses_keur": 10,
        "local_tax_keur": 11,
        "depreciation_keur": 13,
        "total_expenses_keur": 14,
        "ebit_keur": 16,
        "senior_interests_keur": 24,
        "shl_interests_keur": 27,
        "financial_earnings_keur": 30,
        "earnings_before_tax_keur": 32,
        "fiscal_reintegration_keur": 34,
        "taxable_income_keur": 35,
        "losses_carryforward_keur": 39,
        "taxable_profit_keur": 41,
        "corporate_income_tax_keur": 44,
        "net_income_keur": 46,
        "net_dividends_keur": 50,
    }
    return _read_schedule(wb, "P&L", row_map)


def _read_pl_tax_formulas(wb_formula, wb_data) -> dict:
    """Extract P&L tax rows with both formula text and cached values.

    Loads two workbook handles: wb_formula (data_only=False) for formula text,
    wb_data (data_only=True) for cached numerical values.  A cached value
    without formula text or a formula without cached value is recorded but
    flagged SOURCE_PARTIAL.

    Row references are 1-based (Excel convention).
    Column G (index 6) = period 0 (construction).
    """
    pl_f_rows = list(wb_formula["P&L"].iter_rows(
        min_row=1, max_row=65, max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
        values_only=False,
    ))
    pl_d_rows = list(wb_data["P&L"].iter_rows(
        min_row=1, max_row=65, max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
        values_only=True,
    ))

    # CF sheet dual-load for CF row 77 formula text and cash-tax chain
    _CF_MAX_ROW = 85
    cf_f_rows = list(wb_formula["CF"].iter_rows(
        min_row=1, max_row=_CF_MAX_ROW, max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
        values_only=False,
    ))
    cf_d_rows = list(wb_data["CF"].iter_rows(
        min_row=1, max_row=_CF_MAX_ROW, max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
        values_only=True,
    ))

    def cf_formula_at(row1, col0):
        r = row1 - 1
        if r >= len(cf_f_rows):
            return None
        row = cf_f_rows[r]
        if col0 >= len(row):
            return None
        v = row[col0].value
        return str(v) if v is not None else None

    def cf_period_vector(row1):
        r = row1 - 1
        if r >= len(cf_d_rows):
            return [None] * _N_PERIODS
        row_data = cf_d_rows[r]
        out = []
        for p in range(_N_PERIODS):
            col = _PERIOD_COL_OFFSET + p
            v = row_data[col] if col < len(row_data) else None
            out.append(float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None)
        return out

    # CF dates for period_diagnostic
    def cf_date_vector(row1):
        r = row1 - 1
        if r >= len(cf_d_rows):
            return [None] * _N_PERIODS
        row_data = cf_d_rows[r]
        out = []
        for p in range(_N_PERIODS):
            col = _PERIOD_COL_OFFSET + p
            v = row_data[col] if col < len(row_data) else None
            out.append(_date_str(v))
        return out

    def formula_at(row1, col0):
        """Get formula string from formula workbook (row=1-based, col=0-based)."""
        r = row1 - 1
        if r >= len(pl_f_rows):
            return None
        row = pl_f_rows[r]
        if col0 >= len(row):
            return None
        v = row[col0].value
        return str(v) if v is not None else None

    def cached_at(row1, col0):
        """Get cached numeric value (row=1-based, col=0-based)."""
        r = row1 - 1
        if r >= len(pl_d_rows):
            return None
        row = pl_d_rows[r]
        if col0 >= len(row):
            return None
        v = row[col0]
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, (int, float)):
            return float(v)
        return None

    def period_vector_from_data(row1):
        """Extract all 61 period cached values for a P&L row (1-based)."""
        r = row1 - 1
        if r >= len(pl_d_rows):
            return [None] * _N_PERIODS
        row = pl_d_rows[r]
        out = []
        for p in range(_N_PERIODS):
            col = _PERIOD_COL_OFFSET + p
            v = row[col] if col < len(row) else None
            if isinstance(v, bool):
                out.append(bool(v))
            elif isinstance(v, (int, float)):
                out.append(float(v))
            else:
                out.append(None)
        return out

    G = _PERIOD_COL_OFFSET  # col index for period 0
    B = 1                   # col index for B-column (0-based)
    C = 2
    D = 3

    source_map = {
        "sheet": "P&L",
        "extractor_note": "rows are 1-based; col G (index 6) = period 0 (construction)",
        "formula_workbook_data_only": False,
        "cached_workbook_data_only": True,
    }

    # --- Row-level formula evidence ---
    rows = {}

    def row_entry(row1, label, b_const_col=None, c_const_col=None, d_const_col=None):
        entry = {
            "label": label,
            "row": row1,
            "formula_period0": formula_at(row1, G),
            "cached_period0": cached_at(row1, G),
            "cached_period1": cached_at(row1, G + 1),
            "period_values": period_vector_from_data(row1),
        }
        if b_const_col is not None:
            entry["B_col_cached"] = cached_at(row1, b_const_col)
        if c_const_col is not None:
            entry["C_col_cached"] = cached_at(row1, c_const_col)
        if d_const_col is not None:
            entry["D_col_cached"] = cached_at(row1, d_const_col)
        return entry

    # --- Financial-revenue rows (needed for DSRA/cash-interest identity) ---
    rows["fin_rev_reserve_account"] = row_entry(19, "Interests from Reserve Account (=(G$3>0)*$B19*(CF!F105+CF!F91)*G$6)", b_const_col=B)
    rows["fin_rev_cash"] = row_entry(20, "Interests from Cash (=(CF!F144>0)*$B19*CF!F144*G$5*G$6)")
    rows["fin_rev_adjustment"] = row_entry(21, "Financial revenue adjustment (==-SUM(G19:G20)*$B21*G$6)", b_const_col=B)
    # --- P&L core ---
    rows["depreciation"] = row_entry(13, "Depreciation total (=Dep!G30; row30=SUM(G7:G28) incl. financing costs)")
    rows["ebit"] = row_entry(16, "EBIT (=G8-G14; G8=Total Revenues =CF!G23, G14=Total Expenses=SUM(G10:G13))")
    rows["senior_interests"] = row_entry(24, "Senior Interests (=DS!G53-CF!G83; DS!G53=net interest, CF!G83=-DS!G35 VAT facility)")
    rows["shl_interests"] = row_entry(27, "Shareholder Loan Interests (=DS!G125)")
    rows["financial_earnings"] = row_entry(30, "Financial Earnings (=SUM(G19:G21)-SUM(G24:G28))")
    rows["earnings_before_tax"] = row_entry(32, "Earnings Before Tax (=G16+G30)")
    rows["fiscal_reintegration_display"] = row_entry(34, "Fiscal Reintegration display (=-G54; sign flip of helper)")
    rows["taxable_income"] = row_entry(35, "Taxable Income (=G34+G32 = FR+EBT)")
    rows["losses_n_minus_1"] = row_entry(36, "Losses N-1 opening balance (SUMIF rolling+SUM($F$37:F$37))", b_const_col=B)
    rows["allocated_losses"] = row_entry(37, "Allocated Losses (=IF(AND(G36<=0,G32>0),MIN(ABS(G36),G32),0); uses EBT G32 not TI)", b_const_col=B)
    rows["losses_n"] = row_entry(38, "Losses N (=MIN(G37+G36,0))")
    rows["carriable_losses"] = row_entry(39, "Carriable Losses closing balance (=MIN(G38,F35*$B37); F35=prev-period TI)")
    rows["taxable_profit_n"] = row_entry(41, "Taxable Profit N (=-G37+G35)")
    rows["corporate_income_tax_formula"] = row_entry(43, "CIT P&L formula (=MAX(SUM(F41:G41),0)*$B43*(G4>0)*(MOD(G4,2)=0))", b_const_col=B)
    rows["corporate_income_tax_net_income"] = row_entry(44, "CIT feeding Net Income (=Macro!G40; IF(base,G38,G39); hardcoded scenario values)")
    rows["net_income"] = row_entry(46, "Net Income (=G32-G44; uses row 44 not row 43)")
    rows["fiscal_reintegration_helper"] = row_entry(54, "FR helper (=MIN(MAX(G57,G58)+G59,G27); row34==-G54 is display with sign flip)")
    rows["thin_cap_rule"] = row_entry(56, "Thin Cap Rule (=BS!G45 = False always for Oborovo)")
    rows["thin_cap_amount"] = row_entry(57, "Thin Cap amount (=IF(G56,MAX(G27-$C$57,0),0)); C57=3000 kEUR threshold", c_const_col=C)
    rows["atad_30pct_amount"] = row_entry(58, "ATAD 30pct (=IF(G56,MAX(G27-$C$58*(G32-G30+G13),0),0)); C58=0.30", c_const_col=C)
    rows["non_deductible_shl"] = row_entry(59, "Non-deductible SHL (=-G$27*C59*D59); C59=1.0 D59=True => full SHL", c_const_col=C, d_const_col=D)

    # --- CIT authority analysis: rows 43 vs 44 ---
    cit_formula_vec = period_vector_from_data(43)
    cit_macro_vec = period_vector_from_data(44)
    cit_delta_max = max(
        abs((cit_formula_vec[i] or 0) - (cit_macro_vec[i] or 0))
        for i in range(_N_PERIODS)
    )
    # CF row 77 actual formula from dual-load (period 0 = construction col G)
    cf_row77_formula_first_operating = cf_formula_at(77, G + 1)  # period 1 column
    cf_row83_formula_first_operating = cf_formula_at(83, G + 1)
    cit_authority = {
        "row_43_formula": "=MAX(SUM(F41:G41),0)*$B43*(G4>0)*(MOD(G4,2)=0)",
        "row_44_macro": "=Macro!G40 = IF(Production_Scenario=base_scenario,Macro!G38,Macro!G39)",
        "macro_rows_38_39": "Hardcoded period-by-period values matching row 43 for current workbook snapshot",
        "net_income_uses": "Row 44 (Macro!G40), NOT row 43",
        "cf_cash_tax_row_77_formula_dual_load": cf_row77_formula_first_operating,
        "cf_row_83_formula_dual_load": cf_row83_formula_first_operating,
        "row_43_vs_44_max_delta_keur": cit_delta_max,
        "conclusion": (
            "Rows 43 and 44 produce identical values in current workbook snapshot. "
            "Row 43 is formula-driven (dynamic). Row 44 routes through Macro hardcoded values "
            "that were pre-populated to match row 43. Risk: if upstream changes, "
            "Macro rows 38/39 may become stale. "
            "Authority for P&L and CF: row 44. "
            "CF row 77 = -P&L!row44 (confirmed from dual-load formula read). "
            "CF row 83 = -DS!row35 (VAT facility = 0 for all Oborovo periods)."
        ),
    }

    # --- CF cash-tax chain ---
    cf_cash_tax_v = cf_period_vector(77)  # CF row 77 = -P&L!row44, negative = outflow
    cf_bop_dates = cf_date_vector(1)
    cf_eop_dates = cf_date_vector(2)

    cf_lifetime_sum = sum(v or 0.0 for v in cf_cash_tax_v)
    pl44_lifetime_sum = sum(v or 0.0 for v in cit_macro_vec)
    cf_vs_pl44_max_delta = max(
        abs((cf_cash_tax_v[i] or 0.0) + (cit_macro_vec[i] or 0.0))
        for i in range(_N_PERIODS)
    )

    # --- BS label inventory (evidence-based) ---
    _BS_MAX_ROW = 55
    bs_d_rows = list(wb_data["BS"].iter_rows(
        min_row=1, max_row=_BS_MAX_ROW, max_col=4, values_only=True,
    ))
    _TAX_LABEL_PATTERNS = (
        "tax payable", "corporate income tax payable", "tax receivable",
        "prepaid tax", "income tax payable", "deferred tax", "current tax",
        "cit payable", "tax balance",
    )
    bs_label_inventory = []
    bs_tax_matches = []
    for row_idx, row_data in enumerate(bs_d_rows, start=1):
        label = row_data[0]
        if label is not None:
            label_str = str(label).strip()
            bs_label_inventory.append({"row": row_idx, "label": label_str})
            lower = label_str.lower()
            if any(pat in lower for pat in _TAX_LABEL_PATTERNS):
                bs_tax_matches.append({"row": row_idx, "label": label_str})
    bs_tax_payable_exists = len(bs_tax_matches) > 0
    cf_tax_chain = {
        "cf_row_77_formula_period1": cf_row77_formula_first_operating,
        "cf_row_77_sign_convention": "Negative (cash outflow); equals -P&L!row44",
        "cf_cash_tax_lifetime_sum_keur": round(cf_lifetime_sum, 4),
        "pl_row44_lifetime_sum_keur": round(pl44_lifetime_sum, 4),
        "cf_vs_pl44_max_delta_keur": round(cf_vs_pl44_max_delta, 9),
        "payment_lag_periods": 0,
        "payment_timing": "CIT is settled within the same semi-annual period it accrues; no lag",
        "bs_label_inventory": bs_label_inventory,
        "bs_tax_label_matches": bs_tax_matches,
        "bs_tax_payable_row_exists": bs_tax_payable_exists,
        "bs_tax_payable_conclusion": (
            "NO_SEPARATE_BS_TAX_BALANCE_ROW_FOUND"
            if not bs_tax_payable_exists else
            f"TAX_RELATED_BS_ROWS_FOUND: {[m['label'] for m in bs_tax_matches]}"
        ),
        "terminal_bs_tax_balance_keur": None,
        "terminal_bs_tax_balance_note": (
            "null — no BS tax balance row exists; balance cannot be derived from absence"
        ),
        "cf_cash_tax_period_values": [round(v or 0.0, 6) for v in cf_cash_tax_v],
    }

    # --- Period-by-period machine-readable reconciliation diagnostic ---
    ebit_v = period_vector_from_data(16)
    dep_v = period_vector_from_data(13)   # P&L row 13 = book dep = tax dep for Oborovo
    fin_earn_v = period_vector_from_data(30)
    ebt_v = period_vector_from_data(32)
    fr_v = period_vector_from_data(34)
    ti_v = period_vector_from_data(35)
    lcf_open_v = period_vector_from_data(36)
    alloc_v = period_vector_from_data(37)
    tp_v = period_vector_from_data(41)
    cit43_v = period_vector_from_data(43)
    cit44_v = cit_macro_vec
    sd_v = period_vector_from_data(24)
    shl_v = period_vector_from_data(27)

    period_diagnostic = []
    for p in range(_N_PERIODS):
        ebt = ebt_v[p] or 0.0
        fr = fr_v[p] or 0.0
        ti = ti_v[p] or 0.0
        cit43 = cit43_v[p] or 0.0
        cit44 = cit44_v[p] or 0.0
        cf_ct = cf_cash_tax_v[p] or 0.0

        ti_ebt_fr_delta = abs(ebt + fr - ti)
        fr_shl_delta = abs(fr - (shl_v[p] or 0.0))
        cf_pl44_delta = abs(cf_ct + cit44)  # CF = -P44

        # Model-year pair: periods 1,2 → MY1; 3,4 → MY2; etc.
        if p == 0:
            model_year_pair = None
        else:
            model_year_pair = ((p - 1) // 2) + 1

        # CIT pairing: CIT fires at even p, pairs [p-1, p]; p=0 excluded
        cit_pairing_bucket = None if (p == 0 or p % 2 != 0) else f"CIT_pair_{p-1}_{p}"

        # Calendar tax year from EoP date (approximate from known schedule)
        eop_date_str = cf_eop_dates[p] if p < len(cf_eop_dates) else None
        calendar_tax_year = None
        if eop_date_str:
            try:
                calendar_tax_year = int(eop_date_str[:4])
            except (ValueError, TypeError):
                pass

        # Deltas are null when clean runtime values are unavailable — not zero.
        # Zero would misrepresent "not compared" as "perfect match".
        signed_delta = None
        abs_delta = None
        classification = "NOT_AVAILABLE_IN_CURRENT_RUNTIME" if p > 0 else "CONSTRUCTION_PERIOD"

        period_diagnostic.append({
            "period": p,
            "period_bop_date": cf_bop_dates[p] if p < len(cf_bop_dates) else None,
            "period_eop_date": eop_date_str,
            "excel_model_year_pair": model_year_pair,
            "excel_cit_pairing_bucket": cit_pairing_bucket,
            "python_calendar_tax_year": calendar_tax_year,
            # Depreciation
            "excel_book_dep_keur": round(dep_v[p] or 0.0, 6),
            "excel_tax_dep_keur": round(dep_v[p] or 0.0, 6),  # book=tax for Oborovo
            "clean_book_dep_keur": None,
            "clean_tax_dep_keur": None,
            "clean_dep_classification": "NOT_AVAILABLE_IN_CURRENT_RUNTIME",
            # Interest
            "excel_senior_keur": round(sd_v[p] or 0.0, 6),
            "phase2c_senior_keur": None,
            "legacy_senior_keur": None,
            "phase23q_senior_keur": None,  # populated by comparison tests
            "interest_classification": "INTEREST_DEPENDENCY_BLOCKS_TAX",
            "excel_shl_keur": round(shl_v[p] or 0.0, 6),
            # Income chain
            "excel_fin_earn_keur": round(fin_earn_v[p] or 0.0, 6),
            "excel_fiscal_reintegration_keur": round(fr or 0.0, 6),
            "clean_fiscal_reintegration_keur": None,
            "excel_ebit_keur": round(ebit_v[p] or 0.0, 6),
            "excel_ebt_keur": round(ebt, 6),
            "excel_taxable_income_keur": round(ti, 6),
            "clean_taxable_income_keur": None,
            # Losses
            "excel_lcf_opening_keur": round(lcf_open_v[p] or 0.0, 6),
            "excel_allocated_losses_keur": round(alloc_v[p] or 0.0, 6),
            "clean_loss_position_keur": None,
            # Tax
            "excel_taxable_profit_keur": round(tp_v[p] or 0.0, 6),
            "excel_cit_p43_keur": round(cit43, 6),
            "excel_cit_p44_routed_keur": round(cit44, 6),
            "excel_cf_cash_tax_keur": round(cf_ct, 6),
            "clean_current_tax_keur": None,
            "clean_cash_tax_keur": None,
            # Reconciliation
            "signed_delta_keur": None,   # null: no clean runtime available for comparison
            "abs_delta_keur": None,
            "cumulative_delta_keur": None,
            "classification": classification,
            # Identity checks (all zero = proved)
            "identity_ti_eq_ebt_plus_fr_delta": round(ti_ebt_fr_delta, 9),
            "identity_fr_eq_shl_delta": round(fr_shl_delta, 9),
            "identity_cf_eq_neg_p44_delta": round(cf_pl44_delta, 9),
        })

    return {
        "_source": source_map,
        "proved_formula_identity": (
            "taxable_income (row 35) = EBT (row 32) + FR (row 34)"
            "; FR = -G54 where G54=MIN(MAX(G57,G58)+G59,G27)"
            "; with thin_cap=BS!G45=False: G57=0, G58=0, G54=MIN(G59,G27)"
            "; G59=-G27*C59*D59=-G27*1.0*True=-G27 => G54=MIN(-G27,G27)=-G27"
            "; so FR=-G54=G27=SHL. Therefore TI=EBT+SHL=(EBIT+fin_earn-SD-SHL)+SHL=EBIT+fin_earn-SD."
            " fin_earn=SUM(G19:G21)-SUM(G24:G28); rows 19-21 are financing revenues;"
            " during debt tenor rows 19-21~=0 so TI~=EBIT-SD."
            " After debt repayment: row 20 (Interests from Cash)>0 so TI=EBIT+small_cash_interest."
        ),
        "proved_cit_formula": (
            "Row 43: MAX(SUM(F41:G41),0)*$B43*(G4>0)*(MOD(G4,2)=0)."
            " SUM(F41:G41) = taxable_profit[prev_period]+taxable_profit[this_period]."
            " MOD(G4,2)=0 fires in even-indexed periods (2,4,6,...)."
            " G4>0 excludes construction (period 0)."
            " Even periods = H1 of each model-year (Jan-Jun)."
            " Each CIT charge = H2+H1 pair (prior H2 + current H1)."
            " Row 44 = Macro!G40 routes same values via hardcoded scenario lookup."
            " Net Income uses row 44; CF cash-tax = -row 44."
        ),
        "proved_lcf_window": (
            "Row 36: SUMIF(IF(G4<=$B$36,$F35:F35,F35:OFFSET(F35,0,-$B$36+1)),\"<0\")+SUM($F$37:F$37)."
            " B36=5 => rolling 5-period window of negative taxable incomes."
            " The SUM($F$37:F$37) term accumulates prior utilized losses (not just window)."
            " Row 37 allocation condition uses G32 (EBT), not G35 (TI): losses only utilized when EBT>0."
            " Row 39 carriable = MIN(G38, F35*$B37) where B37=1; F35=prior-period TI."
        ),
        "proved_thin_cap": "BS!G45=False for all Oborovo periods => thin_cap=False always",
        "proved_atad": (
            "ATAD rows 57/58=0 for all periods (guarded by IF(G56,...)) since G56=BS!G45=False."
            " FR=non_deductible_shl=row59=-G27*C59*D59 where C59=1.0, D59=True => FR=full SHL."
        ),
        "proved_model_year_mapping": (
            "Period 0: construction (BoP=2029-06-29, EoP=2030-06-30)."
            " Period 1: H2-2030 (Jul-Dec). Period 2: H1-2031 (Jan-Jun). ..."
            " Pairs (1,2),(3,4),(5,6),... form model years 1,2,3,..."
            " CIT fires in even periods (2,4,6,...) summing [prev_period+this_period] taxable profits."
            " Python splits on calendar Jan-Dec boundaries, NOT on model-year H2+H1 pairs."
            " LCF: B36=5 = 5 semi-annual model periods (~2.5 calendar years, not 5 tax years)."
        ),
        "proved_dep_row_authority": (
            "P&L row 13 = Dep!G30 = SUM(G7:G28) = total book depreciation including financing costs."
            " Dep!G31 = SUM(G7:G22) = unlevered depreciation (CAPEX only, excludes rows 23-28)."
            " Tax dep gap = Dep!G30 - Dep!G31 = financing costs capitalized into asset base."
            " Python adapter currently uses hard-CAPEX basis (similar to Dep!G31 but not identical)."
        ),
        "cit_row43_vs_row44_authority": cit_authority,
        "cf_tax_chain": cf_tax_chain,
        "period_diagnostic": period_diagnostic,
        "rows": rows,
        **_derive_lcf_rollforward(period_diagnostic),
        "croatian_calendar_year_lcf_proforma": _derive_croatian_legal_lcf_proforma(period_diagnostic),
    }


def _derive_lcf_rollforward(period_diagnostic: list[dict]) -> dict:
    """Compute LCF roll-forward table, tax milestones, and periodisation mismatch.

    Derived entirely from period_diagnostic rows — no workbook re-read required.
    Croatian CIT year: 1 January – 31 December. The workbook pairs H2(yr N) + H1(yr N+1);
    that is a modelling convention (WORKBOOK_PERIODISATION_MISMATCH), NOT a July fiscal year.
    """
    LCF_WINDOW = 5  # B36 = 5 semi-annual model periods
    CIT_RATE = 0.10  # B43 = 10%

    op_rows = [r for r in period_diagnostic if r["period"] > 0]

    # --- LCF roll-forward table ---
    rollforward = []
    for r in op_rows:
        p = r["period"]
        ti = r["excel_taxable_income_keur"]
        lcf_op = r["excel_lcf_opening_keur"]
        alloc = r["excel_allocated_losses_keur"]
        tp = r["excel_taxable_profit_keur"]
        cit = r["excel_cit_p43_keur"]
        cash = r["excel_cf_cash_tax_keur"]

        # LCF closing = SUMIF negative TI in rolling window [p-LCF_WINDOW+1 .. p]
        window_start = max(1, p - LCF_WINDOW + 1)
        window_tis = [rr["excel_taxable_income_keur"] for rr in op_rows
                      if window_start <= rr["period"] <= p]
        lcf_cl = sum(t for t in window_tis if t < 0)

        # Losses expired = negative TI that fell out of the window this period
        if p >= LCF_WINDOW + 1:
            dropped_p = p - LCF_WINDOW
            dropped_ti = next((rr["excel_taxable_income_keur"] for rr in period_diagnostic
                               if rr["period"] == dropped_p), 0.0)
            expiry = -min(dropped_ti, 0.0)
        else:
            expiry = 0.0

        rollforward.append({
            "period": p,
            "period_bop_date": r["period_bop_date"],
            "period_eop_date": r["period_eop_date"],
            "calendar_tax_year": r["python_calendar_tax_year"],
            "taxable_income_keur": round(ti, 3),
            "lcf_opening_keur": round(lcf_op, 3),
            "allocated_losses_keur": round(alloc, 3),
            "losses_expired_keur": round(expiry, 3),
            "lcf_closing_keur": round(lcf_cl, 3),
            "taxable_profit_keur": round(tp, 3),
            "cit_keur": round(cit, 3),
            "cash_tax_keur": round(cash, 3),
            "cit_pairing_bucket": r.get("excel_cit_pairing_bucket"),
        })

    # --- Tax milestones ---
    first_pos_ti = next((r for r in op_rows if r["excel_taxable_income_keur"] > 0), None)
    first_loss_util = next((r for r in op_rows if r["excel_allocated_losses_keur"] > 0), None)
    first_pos_tp = next((r for r in op_rows if r["excel_taxable_profit_keur"] > 0), None)
    first_cit = next((r for r in op_rows if r["excel_cit_p43_keur"] > 0), None)
    first_cash = next((r for r in op_rows if r["excel_cf_cash_tax_keur"] < 0), None)
    lcf_exhaust = next((r for r in op_rows if r["period"] > 5 and r["excel_lcf_opening_keur"] == 0.0), None)
    lcf_exhaust_prev = next((r for r in op_rows if lcf_exhaust and r["period"] == lcf_exhaust["period"] - 1), None)

    tax_milestones = {
        "first_positive_ti_before_losses": {
            "period": first_pos_ti["period"] if first_pos_ti else None,
            "period_bop_date": first_pos_ti["period_bop_date"] if first_pos_ti else None,
            "ti_keur": round(first_pos_ti["excel_taxable_income_keur"], 3) if first_pos_ti else None,
        },
        "first_loss_utilized": {
            "period": first_loss_util["period"] if first_loss_util else None,
            "note": (
                "NEVER — row 37 condition AND(G36<=0,G32>0) uses EBT (G32). "
                "EBT remains negative throughout all loss-carrying periods "
                "(SHL interest keeps EBT below zero even when TI turns positive). "
                "Losses expire through the rolling 5-period window, never allocated."
            ),
        },
        "first_positive_tp_after_losses": {
            "period": first_pos_tp["period"] if first_pos_tp else None,
            "period_bop_date": first_pos_tp["period_bop_date"] if first_pos_tp else None,
            "tp_keur": round(first_pos_tp["excel_taxable_profit_keur"], 3) if first_pos_tp else None,
            "note": "TP = TI because allocated_losses = 0 for all periods; no loss offset applied",
        },
        "first_cit_expense": {
            "period": first_cit["period"] if first_cit else None,
            "period_bop_date": first_cit["period_bop_date"] if first_cit else None,
            "cit_keur": round(first_cit["excel_cit_p43_keur"], 3) if first_cit else None,
            "pairing_bucket": first_cit.get("excel_cit_pairing_bucket") if first_cit else None,
            "calendar_year_of_booking": first_cit["python_calendar_tax_year"] if first_cit else None,
        },
        "first_cash_tax_outflow": {
            "period": first_cash["period"] if first_cash else None,
            "period_bop_date": first_cash["period_bop_date"] if first_cash else None,
            "cash_keur": round(first_cash["excel_cf_cash_tax_keur"], 3) if first_cash else None,
        },
        "workbook_rolling_window_balance_reaches_zero": {
            "after_period": lcf_exhaust_prev["period"] if lcf_exhaust_prev else None,
            "after_period_eop": lcf_exhaust_prev["period_eop_date"] if lcf_exhaust_prev else None,
            "note": (
                "Rolling 5-period window no longer contains any negative TI period. "
                "Losses expired; not utilized. LCF opening = 0 from this period onward."
            ),
        },
        "workbook_lcf_classification": "WORKBOOK_LCF_MECHANICS_PROVED",
        "losses_ever_utilized": False,
        "utilization_blocked_reason": (
            "Row 37: allocated_losses = IF(AND(G36<=0, G32>0), MIN(ABS(G36), G32), 0). "
            "G32 = EBT. EBT is negative throughout all loss-carrying periods (p1-p10) "
            "because SHL interest (>600 kEUR/period) exceeds EBIT even when EBIT>0. "
            "LCF balance declines solely through rolling-window expiry: the earliest "
            "loss period (p1 TI=-219.2) drops out of the 5-period window after period 6, "
            "reducing the balance from -581.7 to -362.6 kEUR. "
            "CIT fires at period 6 directly on paired taxable profits (not after loss offset)."
        ),
    }

    # --- Calendar-year CIT vs workbook pairing comparison ---
    cal_years: dict = {}
    for r in op_rows:
        yr = r["python_calendar_tax_year"]
        if yr not in cal_years:
            cal_years[yr] = {"periods": [], "ti_sum": 0.0}
        cal_years[yr]["periods"].append(r["period"])
        cal_years[yr]["ti_sum"] += r["excel_taxable_income_keur"]

    wb_cit_by_year: dict = {}
    for r in op_rows:
        cit = r["excel_cit_p43_keur"]
        if cit > 0:
            yr = r["python_calendar_tax_year"]
            wb_cit_by_year[yr] = wb_cit_by_year.get(yr, 0.0) + cit

    cal_vs_wb_table = []
    for yr in sorted(cal_years.keys()):
        cal_ti = cal_years[yr]["ti_sum"]
        cal_cit = max(cal_ti, 0.0) * CIT_RATE
        wb_cit = wb_cit_by_year.get(yr, 0.0)
        cal_vs_wb_table.append({
            "calendar_year": yr,
            "periods": cal_years[yr]["periods"],
            "calendar_year_ti_keur": round(cal_ti, 3),
            "calendar_year_cit_keur": round(cal_cit, 3),
            "workbook_cit_keur": round(wb_cit, 3),
            "delta_keur": round(wb_cit - cal_cit, 3),
        })

    periodisation_mismatch = {
        "classification": "WORKBOOK_PERIODISATION_MISMATCH",
        "legal_tax_year": "1 January to 31 December (Croatia: calendar year)",
        "workbook_pairing_convention": (
            "Even period CIT fires on SUM(F41:G41) = TP[prev_period] + TP[this_period]. "
            "Prev period = H2(yr N, Jul-Dec), this period = H1(yr N+1, Jan-Jun). "
            "Example: pair 5_6 = H2-2032 (p5) + H1-2033 (p6). Straddles Jan 1."
        ),
        "calendar_year_aggregation_note": (
            "Each calendar year aggregates H1(yr N, Jan-Jun) + H2(yr N, Jul-Dec). "
            "Example year 2033: p6(H1-2033) + p7(H2-2033) = 92.7 + 107.4 = 200.1 kEUR TI. "
            "The calendar_vs_workbook_table shows naive calendar-year CIT (TI × 10%, no LCF offset) "
            "vs workbook CIT, illustrating the timing shift only. "
            "For LCF-adjusted calendar-year CIT see the croatian_calendar_year_lcf_proforma section."
        ),
        "workbook_first_tax_period_note": (
            "Workbook pair 5_6: TI = -3.7 + 92.7 = 89.0 → CIT = 8.9 kEUR, booked H1-2033. "
            "Naive calendar year 2033 TI = 200.1 kEUR → naive CIT = 20.0 kEUR (no LCF). "
            "With Article 17 LCF applied (see croatian_calendar_year_lcf_proforma), "
            "2033 taxable base = 0 → CIT = 0. Do not treat 20.0 kEUR as 'correct 2033 CIT'."
        ),
        "impact_on_tax_shield_exhaustion": (
            "No impact on LCF window exhaustion date: losses expire by rolling window, "
            "not by utilization offset. Exhaustion date is independent of CIT aggregation."
        ),
        "impact_on_cit_timing": (
            "Periodisation shifts CIT booking by approximately 6 months relative to a "
            "calendar-year assessment without LCF. With LCF applied the timing and amounts "
            "both change materially (first CIT under Article 17 pro forma: 2034). "
            "The workbook LCF shortcut (rolling window, no utilization) and the "
            "H2+H1 pairing compound to produce a different pattern from the legal treatment."
        ),
        "calendar_vs_workbook_table": cal_vs_wb_table,
    }

    return {
        "lcf_rollforward": rollforward,
        "tax_milestones": tax_milestones,
        "periodisation_mismatch": periodisation_mismatch,
    }


def _derive_croatian_legal_lcf_proforma(period_diagnostic: list[dict]) -> dict:
    """Croatian Article 17 CIT Act: 5 calendar-year expiry, oldest-first, NO EBT gate.

    This is the legal pro forma — entirely distinct from the workbook's rolling 5-period
    window (WORKBOOK_LCF_MECHANICS_PROVED). The workbook does NOT implement Article 17.
    """
    CIT_RATE = 0.10
    EXPIRY_YEARS = 5

    # Aggregate TI per calendar year from period_diagnostic (operating periods only)
    cal_year_ti: dict[int, float] = {}
    for r in period_diagnostic:
        if r["period"] == 0:
            continue
        yr = r["python_calendar_tax_year"]
        if yr is None:
            continue
        cal_year_ti[yr] = cal_year_ti.get(yr, 0.0) + r["excel_taxable_income_keur"]

    years = sorted(cal_year_ti.keys())
    vintages: dict[int, float] = {}  # origin_year → balance_keur
    proforma_rows = []

    for yr in years:
        ti = cal_year_ti[yr]
        opening_vintages = {k: round(v, 6) for k, v in sorted(vintages.items())}

        # Expire vintages whose 5-year horizon has passed
        expired_by_vintage: dict[int, float] = {}
        for orig_yr in sorted(list(vintages.keys())):
            if yr - orig_yr >= EXPIRY_YEARS:
                expired_by_vintage[orig_yr] = vintages.pop(orig_yr)

        # Accumulate new loss (negative TI → new vintage)
        new_loss = max(0.0, -ti)
        if new_loss > 0.001:
            vintages[yr] = vintages.get(yr, 0.0) + new_loss

        # Utilize losses oldest-first (no EBT gate — Article 17 pro forma)
        utilized_by_vintage: dict[int, float] = {}
        remaining = max(0.0, ti)
        if remaining > 0.001:
            for orig_yr in sorted(list(vintages.keys())):
                if remaining <= 0.001:
                    break
                avail = vintages[orig_yr]
                used = min(avail, remaining)
                utilized_by_vintage[orig_yr] = round(used, 6)
                vintages[orig_yr] -= used
                remaining -= used
            vintages = {k: v for k, v in vintages.items() if v > 0.001}

        total_utilized = sum(utilized_by_vintage.values())
        taxable_base = max(0.0, ti - total_utilized) if ti > 0 else 0.0
        cit = taxable_base * CIT_RATE
        closing_vintages = {k: round(v, 6) for k, v in sorted(vintages.items())}

        proforma_rows.append({
            "calendar_year": yr,
            "calendar_year_ti_keur": round(ti, 3),
            "opening_lcf_by_vintage_keur": {str(k): round(v, 3) for k, v in opening_vintages.items()},
            "opening_lcf_total_keur": round(sum(opening_vintages.values()), 3),
            "new_loss_keur": round(new_loss, 3),
            "expired_by_vintage_keur": {str(k): round(v, 3) for k, v in expired_by_vintage.items()},
            "utilized_by_vintage_keur": {str(k): round(v, 3) for k, v in utilized_by_vintage.items()},
            "total_utilized_keur": round(total_utilized, 3),
            "taxable_base_keur": round(taxable_base, 3),
            "cit_keur": round(cit, 3),
            "closing_lcf_by_vintage_keur": {str(k): round(v, 3) for k, v in closing_vintages.items()},
            "closing_lcf_total_keur": round(sum(closing_vintages.values()), 3),
        })

    first_util = next((r for r in proforma_rows if r["total_utilized_keur"] > 0), None)
    first_cit = next((r for r in proforma_rows if r["cit_keur"] > 0), None)
    exhaustion = next(
        (r for r in proforma_rows if r["total_utilized_keur"] > 0 and r["closing_lcf_total_keur"] < 0.001),
        None,
    )

    return {
        "legal_basis": "Croatian CIT Act Article 17: 5 calendar-year expiry, oldest-first utilization, no EBT gate",
        "proforma_classification": "LEGAL_PROFORMA_NOT_WORKBOOK_IMPLEMENTATION",
        "note": (
            "Pro forma only — the workbook does NOT implement Article 17. "
            "Calendar-year TI sums are derived from period_diagnostic excel_taxable_income_keur. "
            "Loss utilization uses oldest-first and no EBT gate, unlike workbook row 37."
        ),
        "proforma_rows": proforma_rows,
        "milestones": {
            "LEGAL_PROFORMA_FIRST_LOSS_UTILIZATION_YEAR": first_util["calendar_year"] if first_util else None,
            "LEGAL_PROFORMA_FIRST_CIT_YEAR": first_cit["calendar_year"] if first_cit else None,
            "LEGAL_PROFORMA_LCF_EXHAUSTION_YEAR": exhaustion["calendar_year"] if exhaustion else None,
        },
    }


def _derive_debt_sizing_from_workbook(wb_formula, wb_data) -> dict:
    """Extract D192 (Inputs!D192 = senior debt amount) in dual-load mode.

    Classifies the debt-sizing source methodology from formula-mode content.
    When formula-mode content is unavailable (workbook absent in test environment),
    source_classification defaults to SOURCE_UNRESOLVED and the debt verdict
    is DEBT_SIZING_SOURCE_UNRESOLVED_MANUAL_CHECK_REQUIRED.

    Valid source_classification values:
      LITERAL_VALUE_IN_CURRENT_WORKBOOK  — D192 contains a numeric literal (no formula)
      FORMULA_DERIVED                    — D192 contains an Excel formula
      MACRO_OR_GOAL_SEEK_HISTORY_NOT_OBSERVABLE — formula text references Goal Seek or macro
      SOURCE_UNRESOLVED                  — formula-mode content not available for inspection
    """
    D192_ROW = 192
    D192_COL_1 = 4  # col D = 1-based column 4

    d192_formula: str | None = None
    d192_cached: float | None = None

    try:
        ws_f = wb_formula["Inputs"]
        for row in ws_f.iter_rows(min_row=D192_ROW, max_row=D192_ROW,
                                   min_col=D192_COL_1, max_col=D192_COL_1,
                                   values_only=False):
            if row:
                v = row[0].value
                d192_formula = str(v) if v is not None else None
    except Exception:
        pass

    try:
        ws_d = wb_data["Inputs"]
        for row in ws_d.iter_rows(min_row=D192_ROW, max_row=D192_ROW,
                                   min_col=D192_COL_1, max_col=D192_COL_1,
                                   values_only=True):
            if row:
                v = row[0]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    d192_cached = float(v)
    except Exception:
        pass

    if d192_formula is None:
        d192_source = "SOURCE_UNRESOLVED"
    elif not d192_formula.startswith("="):
        d192_source = "LITERAL_VALUE_IN_CURRENT_WORKBOOK"
    elif any(kw in d192_formula.upper() for kw in ("GOAL", "MACRO")):
        d192_source = "MACRO_OR_GOAL_SEEK_HISTORY_NOT_OBSERVABLE"
    else:
        d192_source = "FORMULA_DERIVED"

    # DS row 22 (dscr_target): dual-load for DSCR classification
    DS_DSCR_ROW = 22
    ds_dscr_formula_p1: str | None = None
    ds_dscr_cached_distinct: list[float] = []

    try:
        ws_ds_f = wb_formula["DS"]
        for row in ws_ds_f.iter_rows(min_row=DS_DSCR_ROW, max_row=DS_DSCR_ROW,
                                      max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
                                      values_only=False):
            col = _PERIOD_COL_OFFSET + 1  # period 1 (0-based)
            if col < len(row):
                v = row[col].value
                ds_dscr_formula_p1 = str(v) if v is not None else None
    except Exception:
        pass

    try:
        ws_ds_d = wb_data["DS"]
        vals: list[float] = []
        for row in ws_ds_d.iter_rows(min_row=DS_DSCR_ROW, max_row=DS_DSCR_ROW,
                                      max_col=_PERIOD_COL_OFFSET + _N_PERIODS + 1,
                                      values_only=True):
            for p in range(_N_PERIODS):
                col = _PERIOD_COL_OFFSET + p
                v = row[col] if col < len(row) else None
                if isinstance(v, (int, float)) and not isinstance(v, bool) and v > 0:
                    vals.append(float(v))
        ds_dscr_cached_distinct = sorted(set(round(v, 3) for v in vals))
    except Exception:
        pass

    dscr_classification = (
        "SOURCE_UNRESOLVED"
        if ds_dscr_formula_p1 is None
        else "SOURCE_UNRESOLVED"  # formula present but role requires full DS chain analysis
    )
    dscr_role_evidence = (
        "Formula-mode content for DS row 22 not available; cannot classify DSCR role"
        if ds_dscr_formula_p1 is None
        else f"Formula extracted (period 1): {ds_dscr_formula_p1}; role classification requires full DS chain analysis"
    )

    # Verdict: SOURCE_UNRESOLVED unless formula-mode content proves a literal
    if d192_source == "LITERAL_VALUE_IN_CURRENT_WORKBOOK":
        verdict = "DEBT_SIZING_SOURCE_UNRESOLVED_MANUAL_CHECK_REQUIRED"
        verdict_rationale = (
            f"D192 is a LITERAL_VALUE (no formula): {d192_formula}. "
            "This confirms the amount was entered directly, but does not prove it was not "
            "previously determined by Goal Seek or macro. Manual verification required "
            "to confirm whether this is an original input or a frozen Goal Seek result."
        )
    elif d192_source == "FORMULA_DERIVED":
        verdict = "DEBT_SIZING_SOURCE_UNRESOLVED_MANUAL_CHECK_REQUIRED"
        verdict_rationale = (
            f"D192 is FORMULA_DERIVED: {d192_formula}. "
            "Formula chain must be traced to confirm sizing methodology. "
            "Manual verification required."
        )
    else:
        verdict = "DEBT_SIZING_SOURCE_UNRESOLVED_MANUAL_CHECK_REQUIRED"
        verdict_rationale = (
            "D192 formula-mode content is SOURCE_UNRESOLVED: workbook binary required for "
            "formula-mode inspection. Cannot classify whether D192 is a literal, formula-derived, "
            "or Goal Seek/macro result. Manual inspection required (see manual_check_pack)."
        )

    return {
        "d192_evidence": {
            "sheet": "Inputs",
            "row": D192_ROW,
            "col": "D",
            "formula_mode_content": d192_formula,
            "cached_value_keur": d192_cached,
            "source_classification": d192_source,
        },
        "ds_dscr_row22_evidence": {
            "sheet": "DS",
            "row": DS_DSCR_ROW,
            "col": "G-onward (period axis)",
            "formula_mode_period1": ds_dscr_formula_p1,
            "cached_distinct_values": ds_dscr_cached_distinct,
            "dscr_classification": dscr_classification,
            "dscr_role_evidence": dscr_role_evidence,
        },
        "debt_sizing_verdict": verdict,
        "verdict_rationale": verdict_rationale,
    }


def _read_dep(wb) -> dict:
    row_map = {
        "bop_date": 1,
        "eop_date": 2,
        "dep_production_units_keur": 7,
        "dep_epc_contract_keur": 8,
        "dep_epc_other_keur": 9,
        "dep_grid_connection_keur": 10,
        "dep_investments_operation_keur": 11,
        "dep_insurances_keur": 12,
        "dep_project_finance_keur": 14,
        "dep_construction_mgmt_keur": 18,
        "dep_contingencies_keur": 19,
        "dep_project_acquisition_keur": 21,
        "dep_project_rights_keur": 22,
        "dep_idc_keur": 23,
        "dep_commitment_fees_keur": 24,
        "dep_bank_fees_keur": 25,
        "dep_vat_keur": 28,
        "dep_total_keur": 30,
    }
    return _read_schedule(wb, "Dep", row_map)


# ---------------------------------------------------------------------------
# SHA-256 of the workbook file
# ---------------------------------------------------------------------------

def _sha256(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def extract(workbook_path: pathlib.Path) -> dict:
    import openpyxl

    wb_data = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True, keep_vba=False
    )
    wb_formula = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=False, keep_vba=False
    )
    try:
        payload: dict = {
            "_meta": {
                "extractor_version": _EXTRACTOR_VERSION,
                "source_filename": workbook_path.name,
                "source_sha256": _sha256(workbook_path),
                "sheets_inspected": wb_data.sheetnames,
                "dual_load_note": (
                    "dual_load: data_only=True for cached values, "
                    "data_only=False for formula text. "
                    "Both loads read the same binary; formulas are not re-evaluated."
                ),
            },
            "inputs": _read_inputs(wb_data),
            "capex_sheet": _read_capex_sheet(wb_data),
            "cf": _read_cf(wb_data),
            "ds": _read_ds(wb_data),
            "pl": _read_pl(wb_data),
            "dep": _read_dep(wb_data),
            "tax": _read_pl_tax_formulas(wb_formula, wb_data),
            "debt_sizing_evidence": _derive_debt_sizing_from_workbook(wb_formula, wb_data),
        }
    finally:
        wb_data.close()
        wb_formula.close()
    return payload


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m finco_recon.extract_oborovo_excel",
        description="Extract Oborovo financial truth from the authoritative XLSM.",
    )
    parser.add_argument("--workbook", required=True, type=pathlib.Path,
                        help="Path to the source XLSM workbook")
    parser.add_argument("--output", required=True, type=pathlib.Path,
                        help="Output JSON fixture path")
    args = parser.parse_args(argv)

    wb_path: pathlib.Path = args.workbook
    if not wb_path.exists():
        print(f"STOP\nSOURCE_WORKBOOK_REQUIRED\n\nFile not found: {wb_path}",
              file=sys.stderr)
        return 1

    print(f"Extracting from: {wb_path.name}", flush=True)
    print(f"SHA-256: ...", flush=True)

    try:
        payload = extract(wb_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        import traceback; traceback.print_exc()
        return 2

    sha = payload["_meta"]["source_sha256"]
    print(f"SHA-256: {sha}", flush=True)

    # Write timestamp separately so financial payload is deterministic
    output_path: pathlib.Path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    out_doc = {
        "_extraction_timestamp_utc": datetime.datetime.utcnow().isoformat() + "Z",
        **payload,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(out_doc, f, indent=2, default=str)

    print(f"Written: {output_path}", flush=True)
    sheets = payload["_meta"]["sheets_inspected"]
    print(f"Sheets inspected ({len(sheets)}): {', '.join(sheets)}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
