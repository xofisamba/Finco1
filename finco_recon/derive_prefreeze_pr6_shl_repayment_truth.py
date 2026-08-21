"""Extract PR-6 SHL repayment evidence from authoritative source workbooks.

This utility is evidence-only. Its output is a test fixture and must never be
loaded by production runtime code.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SOURCES = {
    "TUHO": {
        "filename": "20260330_TUHO_BP.xlsm",
        "sha256": "780779eba4278ccc2b8546a9411ccee24917d388f411ba60c88aa342cb5c727a",
        "rows": {"opening": 120, "funding": 121, "gross": 122, "wht": 123,
                 "principal": 124, "pik": 125, "closing": 126, "cash": 102},
        "rate_cell": "Inputs!F311",
        "construction_interest_classification": "SOURCE_IDC_HANDOFF_UNPROMOTED",
    },
    "OBOROVO": {
        "filename": "20260414_BP_Oborovo_Sensitivity_FINAL_for_PPT.xlsm",
        "sha256": "15a621c4d6b79024980766e00ebc79d7235fd56f00567be7bf345c769ce57920",
        "rows": {"opening": 123, "funding": 124, "gross": 125, "wht": 126,
                 "principal": 127, "pik": 128, "closing": 129, "cash": 112},
        "rate_cell": "Inputs!F328",
        "construction_interest_classification": "SIMPLE_DCF_1_SOURCE_PROVEN",
    },
    "KUPI": {
        "filename": "20260422_KUPI_BP_NEW.xlsm",
        "sha256": "111178fb21109f55df45c0cc1ea108104ac8b6ed60f010ba75b6c498795f5954",
        "rows": {"opening": 120, "funding": 121, "gross": 122, "wht": 123,
                 "principal": 124, "pik": 125, "closing": 126, "cash": 102},
        "rate_cell": "Inputs!F311",
        "construction_interest_classification": "COMPOUND_PERIODIC_SOURCE_PROVEN",
    },
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _cell_value(workbook, qualified_cell: str):
    sheet, cell = qualified_cell.split("!", 1)
    return workbook[sheet][cell].value


def _extract_source(source_dir: Path, project: str, config: dict) -> dict:
    path = source_dir / config["filename"]
    actual_sha = _sha256(path)
    if actual_sha != config["sha256"]:
        raise ValueError(
            f"{project} source hash mismatch: expected {config['sha256']}, got {actual_sha}"
        )

    values = load_workbook(path, data_only=True, keep_vba=True, read_only=False)
    formulas = load_workbook(path, data_only=False, keep_vba=True, read_only=False)
    ds_values = values["DS"]
    cf_values = values["CF"]
    ds_formulas = formulas["DS"]
    cf_formulas = formulas["CF"]
    rows = config["rows"]

    last_active_col = 7
    first_principal_col = None
    for col in range(7, 130):
        vector = [float(ds_values.cell(rows[key], col).value or 0.0) for key in (
            "opening", "funding", "gross", "principal", "pik", "closing"
        )]
        if any(abs(value) > 1e-9 for value in vector):
            last_active_col = col
        if first_principal_col is None and abs(vector[3]) > 1e-9:
            first_principal_col = col
    if first_principal_col is None:
        raise ValueError(f"{project}: source contains no SHL principal repayment")

    periods = []
    for col in range(7, last_active_col + 1):
        gross = float(ds_values.cell(rows["gross"], col).value or 0.0)
        pik = float(ds_values.cell(rows["pik"], col).value or 0.0)
        periods.append({
            "period_index": col - 7,
            "excel_column": get_column_letter(col),
            "day_count_fraction": float(ds_values.cell(14, col).value or 0.0),
            "opening_balance_keur": float(ds_values.cell(rows["opening"], col).value or 0.0),
            "drawdown_keur": float(ds_values.cell(rows["funding"], col).value or 0.0),
            "gross_interest_keur": gross,
            "cash_interest_keur": gross - pik,
            "pik_interest_keur": pik,
            "principal_keur": float(ds_values.cell(rows["principal"], col).value or 0.0),
            "closing_balance_keur": float(ds_values.cell(rows["closing"], col).value or 0.0),
            "cash_available_for_shl_keur": float(cf_values.cell(rows["cash"], col).value or 0.0),
            "withholding_tax_keur": float(ds_values.cell(rows["wht"], col).value or 0.0),
        })

    evidence_columns = sorted({8, first_principal_col, last_active_col})
    formula_lock = {}
    for col in evidence_columns:
        letter = get_column_letter(col)
        for key in ("opening", "gross", "principal", "pik", "closing"):
            coordinate = f"{letter}{rows[key]}"
            formula_lock[f"DS!{coordinate}"] = ds_formulas[coordinate].value
        cash_coordinate = f"{letter}{rows['cash']}"
        formula_lock[f"CF!{cash_coordinate}"] = cf_formulas[cash_coordinate].value

    return {
        "workbook_filename": config["filename"],
        "workbook_sha256": actual_sha,
        "annual_rate": float(_cell_value(values, config["rate_cell"])),
        "annual_rate_cell": config["rate_cell"],
        "repayment_mode": "CASH_SWEEP",
        "repayment_start_period_index": first_principal_col - 7,
        "maturity_period_index": last_active_col - 7,
        "construction_interest_classification": config["construction_interest_classification"],
        "formula_lock": formula_lock,
        "periods": periods,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = {
        "_meta": {
            "classification": "SOURCE_TYPED_SHL_CASH_SWEEP_AUTHORITY",
            "runtime_use": "FORBIDDEN_TEST_EVIDENCE_ONLY",
            "derivation_script": "finco_recon/derive_prefreeze_pr6_shl_repayment_truth.py",
        },
        "projects": {
            project: _extract_source(args.source_dir, project, config)
            for project, config in SOURCES.items()
        },
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
