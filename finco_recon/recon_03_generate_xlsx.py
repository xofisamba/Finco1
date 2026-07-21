"""finco_recon.recon_03_generate_xlsx — Generate the Oborovo Excel↔Python
reconciliation workbook.

Writes ``artifacts/reconciliation/oborovo_excel_python_reconciliation.xlsx``.

Usage::

    python finco_recon/recon_03_generate_xlsx.py

Requires openpyxl only (no app imports).
"""
from __future__ import annotations

import datetime
import json
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from finco_recon.recon_03_oborovo_full import (
    build_delta_register,
    load_data,
    MATCH,
    PYTHON_BUG,
    EXCEL_BUG,
    POLICY_DIFFERENCE,
    TIMING_ROUNDING,
    PERIOD_CONVENTION,
    UNRESOLVED_SOURCE,
    OUT_OF_CLEAN_ENGINE_SCOPE,
    RESOLVED,
    OPEN,
    OPEN_CASCADE,
)

_OUT_PATH = pathlib.Path("artifacts/reconciliation/oborovo_excel_python_reconciliation.xlsx")

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED = PatternFill("solid", fgColor="FFC7CE")
_ORANGE = PatternFill("solid", fgColor="FFCC99")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_BLUE_HDR = PatternFill("solid", fgColor="1F4E79")
_LIGHT_GREY = PatternFill("solid", fgColor="F2F2F2")
_WHITE = PatternFill("solid", fgColor="FFFFFF")

_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BOLD = Font(bold=True, name="Calibri", size=10)
_NORMAL = Font(name="Calibri", size=10)

_THIN = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill_for_row(row: dict) -> PatternFill | None:
    cl = row["classification"]
    st = row["status"]
    if cl == MATCH:
        return _GREEN
    if cl == PYTHON_BUG:
        return _RED
    if st in (OPEN, OPEN_CASCADE):
        return _ORANGE
    if cl == PERIOD_CONVENTION:
        return _YELLOW
    if cl == POLICY_DIFFERENCE:
        return _YELLOW
    return None


def _hdr(ws, row_num: int, cols: list[str], start_col: int = 1) -> None:
    for i, text in enumerate(cols):
        cell = ws.cell(row=row_num, column=start_col + i, value=text)
        cell.font = _HDR_FONT
        cell.fill = _BLUE_HDR
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_readme(wb: openpyxl.Workbook, excel_meta: dict) -> None:
    ws = wb.create_sheet("00_README")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    rows = [
        ("Reconciliation Title", "Oborovo — Excel ↔ Python Full Financial Reconciliation"),
        ("Version", "03"),
        ("Generated", datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Workbook SHA256", excel_meta.get("_meta", {}).get("source_sha256", "N/A")),
        ("Source filename", excel_meta.get("_meta", {}).get("source_filename", "N/A")),
        ("Extractor version", excel_meta.get("_meta", {}).get("extractor_version", "N/A")),
        ("", ""),
        ("Methodology", ""),
        ("Period alignment",
         "Excel has 61 periods (0=construction, 1-60=operating). "
         "Python has 60 operating periods (0-59). Excel period i+1 ↔ Python period i."),
        ("Materiality — period", "> 1.0 kEUR per period"),
        ("Materiality — cumulative", "> 10.0 kEUR cumulative"),
        ("", ""),
        ("Classifications", ""),
        (MATCH, "Excel and Python agree within 0.5 kEUR"),
        (PYTHON_BUG, "Python engine produces incorrect result vs Excel"),
        (EXCEL_BUG, "Excel contains an error"),
        (POLICY_DIFFERENCE, "Documented accounting policy difference"),
        (PERIOD_CONVENTION, "Calendar day fraction convention (Python=actual, Excel=nominal semi-annual)"),
        (TIMING_ROUNDING, "Sub-5 kEUR calendar rounding"),
        (UNRESOLVED_SOURCE, "Source of delta unclear — investigation required"),
        (OUT_OF_CLEAN_ENGINE_SCOPE, "Not applicable / not extracted"),
        ("", ""),
        ("Known residuals", ""),
        ("OPEX total delta", "~-3.98 kEUR (PERIOD_CONVENTION — nominal vs actual day count)"),
        ("Book dep delta", "PYTHON_BUG — IDC/commitment_fees/bank_fees absent from Python dep basis (20y/12y lives)"),
        ("Revenue delta", "UNRESOLVED_SOURCE — +1,047.9492 kEUR cumulative; arithmetic bridge incomplete"),
        ("", ""),
        ("Stage A status", "SOURCE COMPLETION IN PROGRESS — DO NOT MERGE"),
        ("DSCR actual-vs-target", "UNRESOLVED_SOURCE — CF row 138 not extracted from Excel fixture"),
        ("Avg DSCR", "returns.actual_avg_dscr is Python waterfall engine result, NOT Excel AVERAGEIF(CF!G138:DW138,\"<10\")"),
        ("Material OPEN count", "893 rows > 1 kEUR (all open statuses combined)"),
    ]

    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=k).font = _BOLD if k and not k.startswith(" ") else _NORMAL
        ws.cell(row=r, column=2, value=v).font = _NORMAL


def _write_exec_summary(wb: openpyxl.Workbook, register: list[dict], python_snap: dict) -> None:
    ws = wb.create_sheet("01_EXEC_SUMMARY")
    _hdr(ws, 1, ["Metric", "Value", "Note"])
    _set_col_widths(ws, {"A": 35, "B": 20, "C": 50})

    ret = python_snap["returns"]
    total = len(register)
    open_rrc = sum(1 for r in register if r["status"] == OPEN)
    open_cascade = sum(1 for r in register if r["status"] == OPEN_CASCADE)
    total_open = open_rrc + open_cascade
    mat_open = sum(
        1 for r in register
        if r["status"] in (OPEN, OPEN_CASCADE) and r["materiality"] == "MATERIAL"
    )
    by_cl: dict[str, int] = {}
    by_st: dict[str, int] = {}
    for row in register:
        by_cl[row["classification"]] = by_cl.get(row["classification"], 0) + 1
        by_st[row["status"]] = by_st.get(row["status"], 0) + 1

    rows = [
        ("Total reconciliation rows", total, ""),
        ("Total OPEN (all statuses)", total_open, "OPEN__ROOT_CAUSE_REQUIRED + OPEN__CASCADE_CONFIRMATION_REQUIRED"),
        ("OPEN__ROOT_CAUSE_REQUIRED", open_rrc, "Source missing / root cause unconfirmed"),
        ("OPEN__CASCADE_CONFIRMATION_REQUIRED", open_cascade, "Downstream cascade — awaiting Stage B A/B proof"),
        ("Material OPEN items (>1 kEUR)", mat_open, "Priority — all open statuses"),
        ("", "", ""),
        ("Status: RESOLVED", by_st.get(RESOLVED, 0), ""),
        ("Status: OPEN__ROOT_CAUSE_REQUIRED", by_st.get(OPEN, 0), ""),
        ("Status: OPEN__CASCADE_CONFIRMATION_REQUIRED", by_st.get(OPEN_CASCADE, 0), ""),
        ("", "", ""),
        ("Python project IRR", f"{ret.get('project_irr', 0)*100:.3f}%", ""),
        ("Python equity IRR", f"{ret.get('equity_irr', 0)*100:.3f}%", ""),
        ("Python avg DSCR", f"{ret.get('avg_dscr', 0):.4f}", ""),
        ("Python actual avg DSCR", f"{ret.get('actual_avg_dscr', 0):.4f}", ""),
        ("Python min DSCR", f"{ret.get('min_dscr', 0):.4f}", ""),
        ("", "", ""),
        ("Total revenue (Python kEUR)", f"{ret.get('total_revenue_keur', 0):.1f}", ""),
        ("Total OPEX (Python kEUR)", f"{ret.get('total_opex_keur', 0):.1f}", ""),
        ("Total EBITDA (Python kEUR)", f"{ret.get('total_ebitda_keur', 0):.1f}", ""),
        ("Total CIT cash (Python kEUR)", f"{ret.get('total_tax_keur', 0):.1f}", ""),
        ("", "", ""),
    ]
    for cl, cnt in sorted(by_cl.items()):
        rows.append((f"Classification: {cl}", cnt, ""))

    for r, (k, v, n) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k).font = _BOLD if k else _NORMAL
        ws.cell(row=r, column=2, value=v).font = _NORMAL
        ws.cell(row=r, column=3, value=n).font = _NORMAL


def _write_section_grid(
    wb: openpyxl.Workbook,
    sheet_name: str,
    section_filter: str | list[str],
    register: list[dict],
    line_filter: str | None = None,
) -> None:
    """Generic period grid: lines as rows, periods 0-59 as columns."""
    ws = wb.create_sheet(sheet_name)

    sections = [section_filter] if isinstance(section_filter, str) else section_filter
    rows_for_section = [
        r for r in register
        if r["financial_section"] in sections
        and r["period_index"] is not None
        and (line_filter is None or r["financial_line"] == line_filter)
    ]

    # Collect unique lines preserving order
    seen: dict[str, bool] = {}
    lines = []
    for r in rows_for_section:
        key = r["financial_line"]
        if key not in seen:
            seen[key] = True
            lines.append(key)

    # Header: line | E/P | periods 0-59
    hdr = ["Line", "E/P"] + [str(p) for p in range(60)]
    _hdr(ws, 1, hdr)

    for line in lines:
        line_rows = [r for r in rows_for_section if r["financial_line"] == line]
        by_period = {r["period_index"]: r for r in line_rows}

        for ep, key in [("Excel", "excel_value"), ("Python", "python_value"), ("Delta", "delta")]:
            row_data = [line, ep]
            for p in range(60):
                r = by_period.get(p)
                row_data.append(r[key] if r else None)
            row_num = ws.max_row + 1
            for c, v in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.font = _NORMAL
                if ep == "Delta" and c > 2 and v is not None:
                    # colour delta cell
                    sample_row = by_period.get(c - 3)
                    if sample_row:
                        fill = _fill_for_row(sample_row)
                        if fill:
                            cell.fill = fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 8
    ws.freeze_panes = "C2"


def _write_delta_register(wb: openpyxl.Workbook, register: list[dict]) -> None:
    ws = wb.create_sheet("16_DELTA_REGISTER")
    cols = list(register[0].keys()) if register else []
    _hdr(ws, 1, cols)

    for r_num, row in enumerate(register, start=2):
        fill = _fill_for_row(row)
        for c_num, col in enumerate(cols, start=1):
            cell = ws.cell(row=r_num, column=c_num, value=row.get(col))
            cell.font = _NORMAL
            if fill:
                cell.fill = fill

    # Freeze and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Column widths
    widths = {
        "A": 18, "B": 20, "C": 28, "D": 10, "E": 12, "F": 12,
        "G": 14, "H": 14, "I": 10, "J": 12, "K": 12,
        "L": 12, "M": 25, "N": 10, "O": 45, "P": 30, "Q": 30, "R": 30,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_source_map(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("17_SOURCE_MAP")
    _hdr(ws, 1, ["Source ID", "Description", "Sheet/Key", "Notes"])
    sources = [
        ("Excel.CF", "Cash flow schedule", "CF", "61 periods, index 0=construction"),
        ("Excel.DS", "Debt service schedule", "DS", "61 periods"),
        ("Excel.PL", "P&L schedule", "P&L", "61 periods"),
        ("Excel.Dep", "Depreciation schedule", "Dep", "61 periods"),
        ("Python.operating_schedules", "Operating aggregates", "canonical JSON", "60 periods"),
        ("Python.tax_and_cfads", "Tax and CFADS", "canonical JSON", "60 periods"),
        ("Python.financing.senior_debt", "Senior debt waterfall", "canonical JSON", "60 periods"),
        ("Python.financing.shl", "SHL waterfall", "canonical JSON", "60 periods"),
        ("Python.financial_statements.pnl", "P&L periods", "canonical JSON", "60 periods"),
        ("Python.returns", "Scalar return metrics", "canonical JSON", "project/equity IRR, DSCR"),
    ]
    for r, row in enumerate(sources, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v).font = _NORMAL
    _set_col_widths(ws, {"A": 30, "B": 35, "C": 20, "D": 40})


def _write_raw_recon(wb: openpyxl.Workbook, excel: dict, python_snap: dict) -> None:
    ws = wb.create_sheet("18_RAW_RECON")
    _hdr(ws, 1, [
        "Period", "BOP (Excel)", "EOP (Excel)",
        "Prod MWh (E)", "Prod MWh (P)",
        "Rev kEUR (E)", "Rev kEUR (P)",
        "OPEX kEUR (E)", "OPEX kEUR (P)",
        "EBITDA kEUR (E)", "EBITDA kEUR (P)",
        "BookDep kEUR (E)", "BookDep kEUR (P)",
        "SD Opening (E)", "SD Opening (P)",
        "SD Closing (E)", "SD Closing (P)",
        "DSCR (E)", "DSCR (P)",
    ])

    cf = excel["cf"]
    dep = excel["dep"]
    ds = excel["ds"]
    os_ = python_snap["operating_schedules"]
    sd = python_snap["financing"]["senior_debt"]
    py_dscr = sd["dscr"]

    for i in range(60):
        ei = i + 1
        e_cfads_p = ds["cfads_for_sd_keur"][ei]
        e_ds_p = ds["sd_service_keur"][ei]
        e_dscr = (e_cfads_p / e_ds_p) if (e_ds_p and e_ds_p != 0 and e_cfads_p is not None) else None

        row = [
            i,
            cf["bop_date"][ei],
            cf["eop_date"][ei],
            cf["production_mwh"][ei],
            os_["production_mwh"][i],
            cf["operating_revenues_keur"][ei],
            os_["revenue_keur"][i],
            cf["operating_expenses_keur"][ei],
            -os_["opex_keur"][i],
            cf["ebitda_keur"][ei],
            os_["ebitda_keur"][i],
            dep["dep_total_keur"][ei],
            os_["book_depreciation_keur"][i],
            ds["sd_beginning_keur"][ei],
            sd["opening_keur"][i],
            ds["sd_ending_keur"][ei],
            sd["closing_keur"][i],
            e_dscr,
            py_dscr[i],
        ]
        for c, v in enumerate(row, start=1):
            ws.cell(row=i + 2, column=c, value=v).font = _NORMAL

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 8
    for c in range(2, 20):
        ws.column_dimensions[get_column_letter(c)].width = 15


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate() -> None:
    print("Loading data...")
    excel, python_snap = load_data()
    print("Building delta register...")
    register = build_delta_register()
    print(f"  {len(register)} rows")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Writing sheets...")
    _write_readme(wb, excel)
    _write_exec_summary(wb, register, python_snap)

    # Period grids
    _write_section_grid(wb, "02_TIMELINE", "TIMELINE", register)
    _write_section_grid(wb, "03_PRODUCTION", "PRODUCTION", register)
    _write_section_grid(wb, "04_REVENUE", "REVENUE", register)
    _write_section_grid(wb, "05_OPEX_SUMMARY", "OPEX", register, line_filter="total_opex_keur")
    _write_section_grid(wb, "06_OPEX_DETAIL", "OPEX", register)
    _write_section_grid(wb, "07_CAPEX", "CAPEX", register)  # placeholder
    _write_section_grid(wb, "08_BOOK_DEPRECIATION", "BOOK_DEPRECIATION", register, line_filter="book_depreciation_total_keur")
    _write_section_grid(wb, "09_TAX_DEPRECIATION", "TAX_DEPRECIATION", register)
    _write_section_grid(wb, "10_PNL", "PNL", register)
    _write_section_grid(wb, "11_TAX_LCF", "TAX_LCF", register)
    _write_section_grid(wb, "12_CFADS", "CFADS", register)
    _write_section_grid(wb, "13_SENIOR_DEBT", "SENIOR_DEBT", register)
    _write_section_grid(wb, "14_DSCR", "DSCR", register)

    # Equity returns (scalar)
    ws_ret = wb.create_sheet("15_EQUITY_RETURNS")
    _hdr(ws_ret, 1, ["recon_id", "line", "excel_value", "python_value", "delta", "classification", "status", "root_cause"])
    ret_rows = [r for r in register if r["financial_section"] == "EQUITY_RETURNS"]
    for i, row in enumerate(ret_rows, start=2):
        vals = [
            row["recon_id"], row["financial_line"],
            row["excel_value"], row["python_value"], row["delta"],
            row["classification"], row["status"], row["root_cause"],
        ]
        fill = _fill_for_row(row)
        for c, v in enumerate(vals, start=1):
            cell = ws_ret.cell(row=i, column=c, value=v)
            cell.font = _NORMAL
            if fill:
                cell.fill = fill
    ws_ret.column_dimensions["A"].width = 18
    ws_ret.column_dimensions["H"].width = 60

    _write_delta_register(wb, register)
    _write_source_map(wb)
    _write_raw_recon(wb, excel, python_snap)

    _OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_OUT_PATH)
    print(f"Saved: {_OUT_PATH}")


if __name__ == "__main__":
    generate()
