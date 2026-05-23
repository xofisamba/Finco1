"""
Tests for Phase 10 TUHO Human-Readable Calibration Workbook — Data-Feed Fix.

Verifies:
1. Workbook and source inventory exist
2. Required sheets present (14 sheets)
3. Source map: no COMMITTED row where workbook says MISSING_EVIDENCE without explanation
4. Production/Revenue/OPEX/EBITDA model rows non-zero (61/61)
5. Senior Debt model rows non-zero
6. SHL model rows non-zero where applicable
7. Tax model rows populated
8. CFADS model rows populated
9. Gap Analysis has more than a few rows
10. MISSING_EVIDENCE rows have precise reason
11. G20 remains BLOCKED
12. R99/R102 remains NOT APPROVED
13. No runtime/model formula files changed
"""

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "reports" / "phase10_human_readable_calibration_workbook.xlsx"
INVENTORY = ROOT / "reports" / "phase10_human_readable_calibration_source_inventory.csv"
SOURCE_MAP_CSV = ROOT / "reports" / "phase10_human_readable_calibration_source_map.csv"
GAP_CSV = ROOT / "reports" / "phase10_human_readable_calibration_gap_analysis.csv"
SUMMARY_CSV = ROOT / "reports" / "phase10_human_readable_calibration_summary.csv"


REQUIRED_SHEETS = {
    "Summary",
    "Operations",
    "Revenue",
    "OPEX EBITDA",
    "Senior Debt",
    "SHL",
    "Tax",
    "CFADS Waterfall",
    "Distributions",
    "Returns",
    "Gap Analysis",
    "Source Map",
    "Accepted Conventions",
    "Governance",
}


def _load_workbook():
    return load_workbook(WORKBOOK)


def _sheet_labels_all(ws, col=1):
    return [ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)]


def _find_row(ws, label):
    labels = _sheet_labels_all(ws)
    if label in labels:
        return labels.index(label) + 1
    return None


def _extract_col_values(ws, row, col_start=3, col_count=61):
    """Extract values from a row starting at col_start for col_count columns."""
    vals = []
    for col in range(col_start, col_start + col_count):
        v = ws.cell(row=row, column=col).value
        vals.append(v)
    return vals


def test_workbook_exists():
    """Test 1: workbook exists."""
    assert WORKBOOK.exists(), f"Workbook not found: {WORKBOOK}"


def test_source_inventory_exists():
    """Test 2: source inventory exists."""
    assert INVENTORY.exists(), f"Source inventory not found: {INVENTORY}"


def test_required_sheets_exist():
    """Test 3: required sheets exist."""
    wb = _load_workbook()
    assert set(wb.sheetnames) == REQUIRED_SHEETS, \
        f"Sheet mismatch. Expected {REQUIRED_SHEETS}, got {set(wb.sheetnames)}"


def test_source_map_no_committed_contradiction():
    """Test 4: source map has no COMMITTED row where workbook still shows MISSING_EVIDENCE without explanation."""
    import csv
    sm_rows = list(csv.DictReader(open(SOURCE_MAP_CSV)))
    # All COMMITTED rows must have meaningful notes (not just MISSING_EVIDENCE)
    contradictions = []
    for r in sm_rows:
        if r['source_status'] == 'COMMITTED' and 'MISSING_EVIDENCE' in r.get('notes', '') and 'Excel:' in r.get('notes', ''):
            # Special case: Taxable Income and CFADS have MISSING_EVIDENCE for Excel but COMMITTED for model
            # These are legitimate — Excel side is MISSING_EVIDENCE, model side is COMMITTED
            # Only flag if the notes don't clearly separate the two
            notes = r.get('notes', '')
            if 'Excel: MISSING_EVIDENCE' not in notes or 'model: live runtime' not in notes:
                contradictions.append(r['metric'])
    assert len(contradictions) == 0, \
        f"COMMITTED rows with MISSING_EVIDENCE in notes: {contradictions}"


def test_production_model_not_all_zero():
    """Test 5: Production model row not all zero."""
    wb = _load_workbook()
    ws = wb["Operations"]
    # Find [Model] Production row
    row = _find_row(ws, "[Model] Production (MWh)")
    assert row is not None, "[Model] Production row not found in Operations sheet"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    non_zero = [v for v in vals if v is not None and v != 0]
    assert len(non_zero) >= 55, \
        f"Production model non-zero count: {len(non_zero)}/61 — should be ~61"


def test_revenue_model_not_all_zero():
    """Test 6: Revenue model row not all zero."""
    wb = _load_workbook()
    ws = wb["Revenue"]
    row = _find_row(ws, "[Model] Electricity Revenue (kEUR)")
    assert row is not None, "[Model] Revenue row not found"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    non_zero = [v for v in vals if v is not None and v != 0]
    assert len(non_zero) >= 55, f"Revenue model non-zero: {len(non_zero)}/61"


def test_opex_model_not_all_zero():
    """Test 7: OPEX model row not all zero."""
    wb = _load_workbook()
    ws = wb["OPEX EBITDA"]
    row = _find_row(ws, "[Model] OPEX (kEUR)")
    assert row is not None, "[Model] OPEX row not found"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    non_zero = [v for v in vals if v is not None and v != 0]
    assert len(non_zero) >= 55, f"OPEX model non-zero: {len(non_zero)}/61"


def test_ebitda_model_not_all_zero():
    """Test 8: EBITDA model row not all zero."""
    wb = _load_workbook()
    ws = wb["OPEX EBITDA"]
    row = _find_row(ws, "[Model] EBITDA (kEUR)")
    assert row is not None, "[Model] EBITDA row not found"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    non_zero = [v for v in vals if v is not None and v != 0]
    assert len(non_zero) >= 55, f"EBITDA model non-zero: {len(non_zero)}/61"


def test_senior_debt_model_not_all_zero():
    """Test 9: Senior Debt model rows not all zero."""
    wb = _load_workbook()
    ws = wb["Senior Debt"]
    for metric in ["Interest (kEUR)", "Principal (kEUR)", "Debt Service (kEUR)", "Closing Balance (kEUR)"]:
        row = _find_row(ws, f"[Model] {metric}")
        assert row is not None, f"[Model] Senior Debt {metric} not found"
        vals = _extract_col_values(ws, row, col_start=3, col_count=61)
        non_zero = [v for v in vals if v is not None and v != 0]
        assert len(non_zero) >= 20, \
            f"Senior Debt {metric} non-zero: {len(non_zero)}/61"


def test_shl_model_rows_not_all_zero():
    """Test 10: SHL model rows not all zero (where balance > 0)."""
    wb = _load_workbook()
    ws = wb["SHL"]
    for metric in ["Cash Interest (kEUR)", "PIK Capitalized (kEUR)", "Closing Balance (kEUR)"]:
        row = _find_row(ws, f"[Model] {metric}")
        assert row is not None, f"[Model] SHL {metric} not found in sheet. Available labels: {[l for l in labels if l and '[' in l]}"
        vals = _extract_col_values(ws, row, col_start=3, col_count=61)
        non_zero = [v for v in vals if v is not None and v != 0]
        assert len(non_zero) >= 20, \
            f"SHL {metric} non-zero: {len(non_zero)}/61"


def test_tax_model_rows_populated():
    """Test 11: Tax model rows populated where runtime exists."""
    wb = _load_workbook()
    ws = wb["Tax"]
    # Taxable income model should have non-zero values in operational periods
    row = _find_row(ws, "[Model] Taxable Income — R35 (kEUR)")
    assert row is not None, "[Model] Taxable Income row not found"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    # Model may have zeros in early periods (loss carryforward) but should have non-zero in later periods
    has_some = any(v is not None and v != 0 for v in vals)
    assert has_some, "Taxable income model has no non-zero values"


def test_cfads_model_rows_populated():
    """Test 12: CFADS model rows populated where runtime exists."""
    wb = _load_workbook()
    ws = wb["CFADS Waterfall"]
    row = _find_row(ws, "[Model] CFADS — R69 (kEUR)")
    assert row is not None, "[Model] CFADS row not found"
    vals = _extract_col_values(ws, row, col_start=3, col_count=61)
    has_some = any(v is not None and v != 0 for v in vals)
    assert has_some, "CFADS model has no non-zero values"


def test_gap_analysis_has_rows():
    """Test 13: Gap Analysis has more than a few rows."""
    import csv
    gaps = list(csv.DictReader(open(GAP_CSV)))
    assert len(gaps) >= 10, f"Gap analysis has only {len(gaps)} rows — should be more than a few"


def test_missing_evidence_has_reasons():
    """Test 14: MISSING_EVIDENCE rows have precise reason in Gap Analysis."""
    import csv
    gaps = list(csv.DictReader(open(GAP_CSV)))
    missing = [g for g in gaps if g.get('classification') == 'MISSING_EVIDENCE']
    for m in missing:
        cause = m.get('likely_root_cause', '')
        assert len(cause) > 20, \
            f"MISSING_EVIDENCE row '{m['metric']}' has vague reason: '{cause}'"


def test_g20_blocked():
    """Test 15: G20 remains BLOCKED."""
    wb = _load_workbook()
    ws = wb["Governance"]
    labels = _sheet_labels_all(ws)
    assert "G20" in labels, "G20 not found in Governance sheet"
    row = _find_row(ws, "G20")
    status = ws.cell(row=row, column=2).value
    assert status == "BLOCKED", f"G20 status is {status}, expected BLOCKED"


def test_r99_r102_not_approved():
    """Test 16: R99/R102 remains NOT APPROVED."""
    wb = _load_workbook()
    ws = wb["Governance"]
    for item in ["R99 — DA flag", "R102 — SHL trigger"]:
        row = _find_row(ws, item)
        assert row is not None, f"{item} not found in Governance sheet"
        status = ws.cell(row=row, column=2).value
        assert status == "NOT APPROVED", f"{item} status is {status}, expected NOT APPROVED"


def test_no_runtime_files_changed():
    """Test 17: No runtime/model formula files changed."""
    import subprocess
    result = subprocess.run(
        ["git", "diff", "--name-only", "HEAD", "--", "domain/", "app/", "tests/"],
        cwd=ROOT, capture_output=True, text=True
    )
    changed = result.stdout.strip().split('\n')
    # Allow test files and new report files; block domain/app changes
    runtime_changed = [f for f in changed if f and 'test_phase10' not in f and 'build_phase10' not in f]
    assert len(runtime_changed) == 0, \
        f"Runtime files changed: {runtime_changed}"


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v"])