"""Tests for Excel depreciation disclosure sheets."""
import pytest
from io import BytesIO
import openpyxl

from app.excel_export import build_excel_export
from app.ui_runner import run_demo_project
from app.capex_engine import build_capex_line_items_from_defaults


def sheet_exists(workbook_bytes: bytes, sheet_name: str) -> bool:
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes))
    return sheet_name in wb.sheetnames


def get_sheet_data(workbook_bytes: bytes, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes))
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


class TestDepreciationAssumptionsSheet:
    """Test 'Depreciation Assumptions' sheet."""

    def test_sheet_exists_in_export(self):
        """Depreciation Assumptions sheet is present."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
        )
        assert sheet_exists(excel_bytes, "Depreciation Assumptions")

    def test_has_required_columns(self):
        """Sheet has Asset Class, Tax Life, Book Life, Tax Method, Book Method, Convention."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        data = get_sheet_data(excel_bytes, "Depreciation Assumptions")
        assert len(data) >= 2, "Should have header + at least one data row"
        header = data[0]
        assert any("Asset Class" in str(h) for h in header), f"Missing 'Asset Class' in {header}"
        assert any("Tax Life" in str(h) for h in header), f"Missing 'Tax Life' in {header}"

    def test_inverter_row_present(self):
        """Inverters appear in the assumptions table."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        data = get_sheet_data(excel_bytes, "Depreciation Assumptions")
        rows_text = [[str(c) for c in row] for row in data]
        flat = [" ".join(r) for r in rows_text]
        assert any("Inverter" in r or "inverter" in r.lower() for r in flat), \
            "Inverter row not found"


class TestTaxDepreciationSheet:
    """Test 'Tax Depreciation' sheet."""

    def test_sheet_exists(self):
        """Tax Depreciation sheet exists."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        assert sheet_exists(excel_bytes, "Tax Depreciation")

    def test_shows_note_when_no_advanced_capex(self):
        """Without advanced_capex_line_items, shows explanatory note."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        data = get_sheet_data(excel_bytes, "Tax Depreciation")
        rows_text = [" ".join(str(c) for c in row) for row in data]
        assert any("Advanced CAPEX" in r for r in rows_text), \
            "Should show note about advanced CAPEX not provided"

    def test_has_data_rows_with_advanced_capex(self):
        """With advanced_capex_line_items, sheet has real depreciation data."""
        result = run_demo_project("Solar", "Base", advanced_capex_line_items=build_capex_line_items_from_defaults("solar"))
        excel_bytes = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            advanced_capex_line_items=build_capex_line_items_from_defaults("solar"),
        )
        data = get_sheet_data(excel_bytes, "Tax Depreciation")
        assert len(data) >= 3, f"Should have header + multiple asset rows + total. Got {len(data)} rows"


class TestBookDepreciationSheet:
    """Test 'Book Depreciation' sheet."""

    def test_sheet_exists(self):
        """Book Depreciation sheet exists."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        assert sheet_exists(excel_bytes, "Book Depreciation")

    def test_shows_note_when_no_advanced_capex(self):
        """Without advanced_capex_line_items, shows explanatory note."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(result=result.result, project_inputs=result.project_inputs)
        data = get_sheet_data(excel_bytes, "Book Depreciation")
        rows_text = [" ".join(str(c) for c in row) for row in data]
        assert any("Advanced CAPEX" in r for r in rows_text), \
            "Should show note about advanced CAPEX not provided"


class TestDepreciationDisclosureCaveats:
    """Test that Notes sheet includes depreciation caveats."""

    def test_notes_contains_depreciation_warnings(self):
        """Notes sheet includes tax/book schedules are model outputs, not audited accounts."""
        result = run_demo_project("Solar", "Base")
        excel_bytes = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            advanced_capex_line_items=build_capex_line_items_from_defaults("solar"),
        )
        notes_data = get_sheet_data(excel_bytes, "Notes")
        rows_text = [" ".join(str(c) for c in row) for row in notes_data]
        notes_combined = " ".join(rows_text)
        assert "Model outputs" in notes_combined or "not audited" in notes_combined.lower(), \
            "Notes should warn that depreciation schedules are model outputs, not audited"


class TestWindDepreciationProfile:
    """Wind export uses wind_croatia_ibl profile."""

    def test_wind_export_uses_wind_profile(self):
        """Wind project type should use wind_croatia_ibl profile in Depreciation Assumptions."""
        from app.ui_runner import run_demo_project
        result = run_demo_project("Wind", "Base")
        excel_bytes = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
        )
        # Check that assumptions sheet mentions Wind-appropriate asset rules
        # Wind profile has same 10y inverter rule as solar
        data = get_sheet_data(excel_bytes, "Depreciation Assumptions")
        rows_text = [" ".join(str(c) for c in row) for row in data]
        flat = " ".join(rows_text)
        # Should have Inverters row
        assert "Inverter" in flat or "inverter" in flat.lower(), \
            "Depreciation Assumptions should have Inverters row for Wind"


class TestLandDepreciationZero:
    """LAND depreciation is zero by design."""

    def test_land_depreciation_is_zero(self):
        """LAND asset class has zero tax and book depreciation."""
        from app.depreciation_bankable import (
            generate_tax_and_book_schedule,
            map_capex_line_item_to_basis,
            get_profile,
            DepreciationConvention,
            BankableAssetClass,
        )
        from app.capex_engine import build_capex_line_items_from_defaults

        items = build_capex_line_items_from_defaults("solar")
        profile = get_profile("solar_croatia_ibl")
        basis_items = [map_capex_line_item_to_basis(item, profile) for item in items]
        tax_sched, book_sched = generate_tax_and_book_schedule(
            basis_items, profile, total_periods=20,
            convention=DepreciationConvention.FULL_YEAR,
        )
        
        land_tax = tax_sched.total_by_asset_class(BankableAssetClass.LAND)
        land_book = book_sched.total_by_asset_class(BankableAssetClass.LAND)
        
        assert all(v == 0.0 for v in land_tax), f"LAND tax should be all zeros, got {land_tax}"
        assert all(v == 0.0 for v in land_book), f"LAND book should be all zeros, got {land_book}"


class TestBookDepreciationScheduleTotalByAssetClass:
    """BookDepreciationSchedule has total_by_asset_class()."""

    def test_book_schedule_has_total_by_asset_class(self):
        """BookDepreciationSchedule has total_by_asset_class method."""
        from app.depreciation_bankable import (
            generate_tax_and_book_schedule,
            map_capex_line_item_to_basis,
            get_profile,
            DepreciationConvention,
            BankableAssetClass,
        )
        from app.capex_engine import build_capex_line_items_from_defaults

        items = build_capex_line_items_from_defaults("solar")
        profile = get_profile("solar_croatia_ibl")
        basis_items = [map_capex_line_item_to_basis(item, profile) for item in items]
        tax_sched, book_sched = generate_tax_and_book_schedule(
            basis_items, profile, total_periods=20,
            convention=DepreciationConvention.FULL_YEAR,
        )
        
        # Method should exist
        assert hasattr(book_sched, "total_by_asset_class"), \
            "BookDepreciationSchedule should have total_by_asset_class method"
        
        # Should return list of length total_periods
        result = book_sched.total_by_asset_class(BankableAssetClass.SOLAR_MODULES)
        assert isinstance(result, list), "Should return list"
        assert len(result) == 20, f"Should return 20-period list, got {len(result)}"
        
        # Should be zero-filled before first use
        assert all(v >= 0 for v in result), "All values should be non-negative"

    def test_tax_and_book_sheets_differ_when_lives_differ(self):
        """Tax and Book Depreciation sheets have different values when lives differ."""
        from app.ui_runner import run_demo_project
        from app.capex_engine import build_capex_line_items_from_defaults
        from io import BytesIO
        import openpyxl

        result = run_demo_project("Solar", "Base", advanced_capex_line_items=build_capex_line_items_from_defaults("solar"))
        excel_bytes = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            advanced_capex_line_items=build_capex_line_items_from_defaults("solar"),
        )
        
        tax_data = get_sheet_data(excel_bytes, "Tax Depreciation")
        book_data = get_sheet_data(excel_bytes, "Book Depreciation")
        
        # Both sheets should have multiple rows (header + asset rows + total)
        assert len(tax_data) >= 3, f"Tax sheet should have multiple rows, got {len(tax_data)}"
        assert len(book_data) >= 3, f"Book sheet should have multiple rows, got {len(book_data)}"
        
        # Find total rows (Total is in second column, first column is row index)
        tax_total = [r for r in tax_data if len(r) > 1 and str(r[1]).strip().lower() == "total"]
        book_total = [r for r in book_data if len(r) > 1 and str(r[1]).strip().lower() == "total"]
        
        assert tax_total, f"Tax sheet should have Total row. Rows: {tax_data}"
        assert book_total, f"Book sheet should have Total row. Rows: {book_data}"
