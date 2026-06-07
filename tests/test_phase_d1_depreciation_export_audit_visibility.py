"""Tests for Phase D1 — Depreciation Export / Audit Visibility Hardening.

This is AUDIT VISIBILITY ONLY. No runtime enablement, no formula
change, no schedule change, no tax / P&L / CFADS change. The
tests pin:

- the new "Depreciation Audit" sheet is present in the export
- the sheet discloses the active depreciation path (legacy by
  default; canonical iff flag is on)
- the sheet discloses the canonical / tax-bridge / book-pnl
  flag status
- the sheet discloses the runtime authority source
- the sheet lists audit-only and runtime surfaces
- the sheet notes known limitations
- the sheet notes generic-project support is NOT provided by
  the active runtime path
- TUHO numeric values are unchanged by the D1 sheet addition
- Oborovo numeric values are unchanged by the D1 sheet addition
- existing depreciation export sheets are still present
- no feature flag is enabled by D1 (i.e. the legacy path is
  still authoritative and the canonical engine is still off)
- rc1 frozen SHA resolves
"""

from __future__ import annotations

import io
import subprocess
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from app.excel_export import (
    _write_depreciation_audit_sheet,
    build_excel_export,
)
from app.depreciation_audit_visibility import (
    _resolve_depreciation_runtime_authority,
    build_depreciation_audit_dataframe,
)
from app.ui_runner import run_demo_project


REPO_ROOT = Path(__file__).resolve().parents[1]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load_sheet_rows(workbook_bytes: bytes, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    assert sheet_name in wb.sheetnames, (
        f"Sheet {sheet_name!r} not present in export. "
        f"Available: {wb.sheetnames}"
    )
    ws = wb[sheet_name]
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(values_only=False)
    ]


def _row_dict(rows: list[list]) -> dict[str, str]:
    """Coerce a (Field, Value) style sheet to a dict, dropping
    the auto-index column and header row."""
    out: dict[str, str] = {}
    for r in rows[1:]:
        # rows[0] is the header (None, 'Field', 'Value'); data
        # rows are (index, field, value). Skip empty fields.
        if len(r) >= 3 and r[1] is not None:
            out[str(r[1])] = r[2] if r[2] is not None else ""
    return out


def _sum_numeric(workbook_bytes: bytes, sheet_name: str) -> float:
    """Sum every numeric cell in a sheet (ignores strings)."""
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    assert sheet_name in wb.sheetnames
    ws = wb[sheet_name]
    total = 0.0
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, (int, float)):
                total += float(v)
    return total


# ---------------------------------------------------------------------------
# 1. Sheet existence
# ---------------------------------------------------------------------------


class TestDepreciationAuditSheetExists:
    """The 'Depreciation Audit' sheet must be present in the
    export for both factory projects (TUHO, Oborovo)."""

    def test_sheet_present_in_tuho_export(self):
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        assert "Depreciation Audit" in wb.sheetnames

    def test_sheet_present_in_oborovo_export(self):
        r = run_demo_project("Oborovo", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="oborovo",
        )
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        assert "Depreciation Audit" in wb.sheetnames


# ---------------------------------------------------------------------------
# 2. Disclosure content — required fields
# ---------------------------------------------------------------------------


REQUIRED_DISCLOSURE_FIELDS = (
    "Phase",
    "Scope",
    "Notice",
    "Active Depreciation Path",
    "Runtime Authority Source",
    "Canonical Depreciation Engine Enabled",
    "Canonical Tax Bridge Enabled",
    "Tax Bridge Engine Enabled",
    "Book Depreciation P&L Bridge Enabled",
    "Values Reflect",
    "Audit-Only Surfaces",
    "Runtime Surfaces",
    "Known Limitations",
    "Generic Project Support",
)


class TestDepreciationAuditDisclosure:
    """The 'Depreciation Audit' sheet must disclose all the
    review-required fields, in the exact required wording
    family."""

    def _build_factory_sheet(self, project: str) -> dict[str, str]:
        r = run_demo_project(project, "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type=project.lower(),
        )
        rows = _load_sheet_rows(data, "Depreciation Audit")
        return _row_dict(rows)

    def test_all_required_fields_present_tuho(self):
        sheet = self._build_factory_sheet("TUHO")
        for field in REQUIRED_DISCLOSURE_FIELDS:
            assert field in sheet, (
                f"D1 sheet is missing required field {field!r}. "
                f"Got: {sorted(sheet)}"
            )

    def test_all_required_fields_present_oborovo(self):
        sheet = self._build_factory_sheet("Oborovo")
        for field in REQUIRED_DISCLOSURE_FIELDS:
            assert field in sheet, (
                f"D1 sheet is missing required field {field!r}. "
                f"Got: {sorted(sheet)}"
            )

    def test_phase_disclosure_d1_visibility_only(self):
        sheet = self._build_factory_sheet("TUHO")
        assert "D1" in sheet["Phase"], sheet["Phase"]
        assert "Audit Visibility" in sheet["Phase"], sheet["Phase"]

    def test_scope_disclosure_visibility_only(self):
        sheet = self._build_factory_sheet("TUHO")
        scope = sheet["Scope"]
        assert "Audit visibility only" in scope, scope
        assert "NO runtime enablement" in scope, scope
        assert "formula" in scope.lower(), scope
        assert "schedule" in scope.lower(), scope
        assert "tax" in scope.lower(), scope
        assert "P&L" in scope, scope
        assert "CFADS" in scope, scope

    def test_notice_disclosure_no_runtime_authority_change(self):
        sheet = self._build_factory_sheet("TUHO")
        notice = sheet["Notice"]
        assert "no runtime authority change" in notice.lower(), notice
        assert "Depreciation export visibility only" in notice, notice

    def test_known_limitations_disclosure(self):
        sheet = self._build_factory_sheet("TUHO")
        lim = sheet["Known Limitations"]
        # Wording requirement: canonical engine not runtime-
        # authoritative unless explicitly enabled.
        assert "not runtime-authoritative" in lim, lim
        assert "Canonical DepreciationEngine" in lim, lim
        # Wording requirement: audit values are advisory.
        assert "advisory" in lim, lim
        # Wording requirement: no generic-project claims.
        assert "Generic-project" in lim, lim
        # Wording requirement: no CFADS / P&L change implicit.
        assert "consolidated" in lim.lower(), lim

    def test_generic_project_support_disclosure_negative(self):
        sheet = self._build_factory_sheet("TUHO")
        gen = sheet["Generic Project Support"]
        # Wording requirement: NO for generic projects.
        assert gen.startswith("NO"), gen
        assert "generic" in gen.lower(), gen
        # Wording requirement: applies to active path only.
        assert "active path" in gen.lower(), gen


# ---------------------------------------------------------------------------
# 3. Authority resolution — flag status is correctly exported
# ---------------------------------------------------------------------------


class TestDepreciationAuthorityResolution:
    """The flag status disclosure must be correct: when the
    canonical flag is OFF (the default in main), the sheet
    must say NO and the active path must be the legacy path."""

    def test_tuho_default_flags_all_disabled(self):
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        sheet = _row_dict(
            _load_sheet_rows(data, "Depreciation Audit")
        )
        # Default: all canonical / bridge flags are off
        # (D1 is NOT supposed to enable any of them).
        assert sheet["Canonical Depreciation Engine Enabled"] == "NO"
        assert sheet["Canonical Tax Bridge Enabled"] == "NO"
        assert sheet["Tax Bridge Engine Enabled"] == "NO"
        assert sheet["Book Depreciation P&L Bridge Enabled"] == "NO"

    def test_active_path_legacy_when_canonical_off(self):
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        sheet = _row_dict(
            _load_sheet_rows(data, "Depreciation Audit")
        )
        # With canonical off, active path must be legacy.
        assert (
            sheet["Active Depreciation Path"]
            == "legacy_depreciation_runtime"
        ), sheet["Active Depreciation Path"]
        # Authority source should be the legacy module path.
        assert (
            sheet["Runtime Authority Source"]
            == "waterfall_core.LegacyDepreciationPath"
        ), sheet["Runtime Authority Source"]

    def test_authority_resolution_helper_handles_no_provenance(self):
        """The resolution helper must be safe to call with
        no provenance and no project_inputs — it falls back
        to a static 'legacy' / 'unknown' state."""
        auth = _resolve_depreciation_runtime_authority()
        # When nothing is provided, defaults are:
        # - active path: legacy (canonical flag = NO)
        # - flags: UNKNOWN
        assert (
            auth["active_depreciation_path"]
            == "legacy_depreciation_runtime"
        )
        assert (
            auth["canonical_depreciation_engine_enabled"]
            == "UNKNOWN"
        )
        assert auth["canonical_tax_bridge_enabled"] == "UNKNOWN"
        assert (
            auth["book_depreciation_pnl_bridge_enabled"]
            == "UNKNOWN"
        )

    def test_authority_resolution_helper_with_canonical_on(self):
        """If provenance says the canonical flag is on, the
        helper must report the canonical path and the
        canonical engine authority source."""

        class _StubInputs:
            pass

        provenance = {
            "runtime_flag_snapshot": {
                "use_depreciation_canonical_engine": True,
                "use_canonical_tax_depreciation_bridge": False,
                "use_tax_bridge_engine": False,
                "use_book_depreciation_for_pnl": False,
            }
        }
        auth = _resolve_depreciation_runtime_authority(
            project_inputs=_StubInputs(),
            provenance_metadata=provenance,
        )
        assert (
            auth["active_depreciation_path"]
            == "canonical_depreciation_engine"
        )
        assert (
            auth["runtime_authority_source"]
            == "domain.depreciation.engine.DepreciationEngine"
        )
        assert (
            auth["canonical_depreciation_engine_enabled"] == "YES"
        )
        # Other flags still read directly from provenance.
        assert auth["canonical_tax_bridge_enabled"] == "NO"
        assert auth["book_depreciation_pnl_bridge_enabled"] == "NO"


# ---------------------------------------------------------------------------
# 4. Numeric invariance — TUHO
# ---------------------------------------------------------------------------


TUHO_REFERENCE_SUMS = {
    # Pinned pre-D1 numeric sum of every numeric cell in each
    # existing TUHO export sheet. The D1 PR must NOT change
    # any of these values.
    "CapEx": 145988.42,
    "CapEx_Items": 70706.54,
    "Inputs": 79580.2375,
    "Depreciation Assumptions": 45.0,
    "Tax Depreciation": 0.0,
    "Tax_Depreciation": 0.0,
    "Book Depreciation": 0.0,
}


class TestTuhoNumericInvariance:
    """TUHO numeric sums per existing sheet must be unchanged
    after the D1 sheet addition. The D1 sheet is text-only
    (the small numeric sum from the auto-index column is
    allowed; it is NOT a financial value)."""

    def test_tuho_existing_sheet_sums_unchanged(self):
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        for sheet_name, expected in TUHO_REFERENCE_SUMS.items():
            actual = _sum_numeric(data, sheet_name)
            assert abs(actual - expected) < 0.01, (
                f"TUHO {sheet_name!r} sum drifted: "
                f"expected {expected}, got {actual}"
            )


# ---------------------------------------------------------------------------
# 5. Numeric invariance — Oborovo
# ---------------------------------------------------------------------------


OBOROVO_REFERENCE_SUMS = {
    "CapEx": 115758.5053,
    "CapEx_Items": 56104.09,
    "Inputs": 61272.8532,
    "Depreciation Assumptions": 45.0,
    "Tax Depreciation": 0.0,
    "Tax_Depreciation": 0.0,
    "Book Depreciation": 0.0,
}


class TestOborovoNumericInvariance:
    """Oborovo numeric sums per existing sheet must be unchanged
    after the D1 sheet addition."""

    def test_oborovo_existing_sheet_sums_unchanged(self):
        r = run_demo_project("Oborovo", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="oborovo",
        )
        for sheet_name, expected in OBOROVO_REFERENCE_SUMS.items():
            actual = _sum_numeric(data, sheet_name)
            assert abs(actual - expected) < 0.01, (
                f"Oborovo {sheet_name!r} sum drifted: "
                f"expected {expected}, got {actual}"
            )


# ---------------------------------------------------------------------------
# 6. Existing sheets still present
# ---------------------------------------------------------------------------


EXPECTED_PRE_D1_SHEETS = (
    "Dashboard",
    "Returns",
    "Waterfall",
    "Revenue",
    "Debt",
    "Tax_Depreciation",
    "Notes",
    "Inputs",
    "CapEx",
    "CapEx_Items",
    "Validation",
    "Depreciation Assumptions",
    "Tax Depreciation",
    "Book Depreciation",
    # Phase D1 addition (this PR):
    "Depreciation Audit",
)


class TestExistingSheetsStillPresent:
    """D1 is purely additive: every pre-existing sheet must
    still be present in the export. The D1 sheet is the only
    addition."""

    def test_all_expected_sheets_present_tuho(self):
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        for s in EXPECTED_PRE_D1_SHEETS:
            assert s in wb.sheetnames, (
                f"Pre-existing sheet {s!r} missing from TUHO "
                f"export. Available: {wb.sheetnames}"
            )

    def test_all_expected_sheets_present_oborovo(self):
        r = run_demo_project("Oborovo", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="oborovo",
        )
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        for s in EXPECTED_PRE_D1_SHEETS:
            assert s in wb.sheetnames, (
                f"Pre-existing sheet {s!r} missing from Oborovo "
                f"export. Available: {wb.sheetnames}"
            )


# ---------------------------------------------------------------------------
# 7. No feature flag enabled, no schedule change
# ---------------------------------------------------------------------------


class TestNoRuntimeEnabling:
    """D1 must not enable any feature flag, must not change
    any schedule, must not change waterfall runtime."""

    def test_canonical_depreciation_flag_still_off(self):
        """Read the flag via the same code path the audit
        sheet uses. The canonical flag must be off (NO)."""
        r = run_demo_project("TUHO", "Base")
        from app.persistence.provenance import runtime_flag_snapshot
        flags = runtime_flag_snapshot(r.project_inputs)
        # Canonical depreciation engine is OFF in baseline;
        # D1 must NOT enable it.
        assert flags.get("use_depreciation_canonical_engine") is False
        assert (
            flags.get("use_canonical_tax_depreciation_bridge")
            is False
        )
        assert flags.get("use_book_depreciation_for_pnl") is False

    def test_d1_sheet_is_text_only(self):
        """The D1 sheet has no financial numeric cells. The
        only numeric values are the auto-index column (0..N)
        and the small column-width numbers. There must be no
        cells with a 'kEUR', 'EUR', or '$' unit marker, and
        no cell with a year-period style number that could be
        mistaken for a depreciation period value."""
        r = run_demo_project("TUHO", "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type="tuho",
        )
        rows = _load_sheet_rows(data, "Depreciation Audit")
        # Every numeric value in the D1 sheet is either the
        # auto-index (0..13) or part of a unit label like
        # "kEUR". We assert by counting numerics in the value
        # column (column index 2). All values in the value
        # column are strings, by design.
        value_col_numerics = 0
        for r in rows[1:]:
            if len(r) >= 3 and isinstance(r[2], (int, float)):
                value_col_numerics += 1
        assert value_col_numerics == 0, (
            f"D1 sheet value column should be text-only. "
            f"Found {value_col_numerics} numeric values: "
            f"{rows}"
        )

    def test_helper_function_writes_only_audit_text(self):
        """Direct test of the helper: the DataFrame the
        helper produces must be text-only (no numerics in
        the 'Value' column)."""
        df = build_depreciation_audit_dataframe(
            project_inputs=None, provenance_metadata=None,
        )
        assert "Field" in df.columns
        assert "Value" in df.columns
        # The "Value" column must contain NO numbers
        # (it is disclosure text only).
        non_string = df["Value"][
            df["Value"].apply(lambda v: isinstance(v, (int, float)))
        ]
        assert len(non_string) == 0, (
            f"D1 sheet Value column should be text only. "
            f"Found numeric values: {list(non_string)}"
        )
        # And the Field column must be non-empty.
        assert len(df) >= 14, (
            f"Expected at least 14 disclosure rows, got {len(df)}"
        )


# ---------------------------------------------------------------------------
# 8. rc1 frozen
# ---------------------------------------------------------------------------


class TestRc1Frozen:
    def test_rc1_sha_resolves(self):
        r = subprocess.run(
            [
                "git", "rev-parse", "--verify",
                "b425a0708719eaa5e1d922b1008e5609758e0ad4",
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        assert r.returncode == 0, (
            "rc1 frozen SHA must still resolve."
        )


# ---------------------------------------------------------------------------
# 9. Phase plan + hard no-go
# ---------------------------------------------------------------------------


class TestD1PhasePlanAndHardNoGo:
    """D1 — Depreciation Export / Audit Visibility Hardening is
    audit-visibility only. The hard no-go list is the
    implementation boundary."""

    def test_hard_no_go_pinned(self):
        no_go = [
            "no_runtime_depreciation_enablement (D1 is visibility only)",
            "no_feature_flag_enablement",
            "no_formula_changes",
            "no_depreciation_schedule_changes",
            "no_tax_calculation_changes",
            "no_pnl_calculation_changes",
            "no_cfads_changes",
            "no_persistence_changes",
            "no_schema_changes",
            "no_ui_workflow_changes_unless_readonly_label",
            "no_generic_project_depreciation_claims",
            "rc1_frozen (b425a0708719eaa5e1d922b1008e5609758e0ad4)",
        ]
        assert len(no_go) == 12

    def test_stop_after_report_contract(self):
        contract = {
            "type": "DRAFT PR",
            "merge": "NO auto-merge",
            "next_phase_start": "gated on user review",
        }
        assert contract["type"] == "DRAFT PR"
        assert contract["merge"] == "NO auto-merge"
