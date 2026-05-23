"""Phase 12 audit replay metadata hardening tests."""

from __future__ import annotations

import csv
import os
import sys
import uuid
from io import BytesIO
from pathlib import Path
import types

import pytest
from openpyxl import load_workbook

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

if "bcrypt" not in sys.modules:
    fake_bcrypt = types.SimpleNamespace(
        gensalt=lambda rounds=12: b"salt",
        hashpw=lambda password, salt: b"stub-hash",
        checkpw=lambda password, hashed: True,
    )
    sys.modules["bcrypt"] = fake_bcrypt

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack
from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.export.runtime_summary import build_runtime_summary_rows
from app.persistence.repository import (
    build_export_lineage,
    list_exports,
    save_project,
    save_run,
    save_scenario,
    record_export,
)


ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "phase12_audit_replay_metadata_hardening.md"
INVENTORY = ROOT / "reports" / "phase12_audit_replay_metadata_inventory.csv"
GAP_MATRIX = ROOT / "reports" / "phase12_replay_gap_matrix.csv"
FLAG_REPORT = ROOT / "reports" / "phase12_runtime_flag_provenance.csv"
TRACEABILITY = ROOT / "reports" / "phase12_export_replay_traceability.csv"


@pytest.fixture
def test_db():
    import app.persistence.db as db_mod

    old_path = os.environ.get("FINCO_DB_PATH")
    db_file = ROOT / f"phase12_replay_{uuid.uuid4().hex[:8]}.db"
    os.environ["FINCO_DB_PATH"] = str(db_file)
    db_mod.DB_PATH = str(db_file)

    yield db_file

    if db_file.exists():
        db_file.unlink()
    if old_path:
        os.environ["FINCO_DB_PATH"] = old_path
        db_mod.DB_PATH = old_path
    else:
        os.environ.pop("FINCO_DB_PATH", None)


def _sheet_values(sheet) -> list[object]:
    return [
        sheet.cell(row=row, column=column).value
        for row in range(1, sheet.max_row + 1)
        for column in range(1, sheet.max_column + 1)
    ]


def test_runtime_summary_rows_surface_replay_metadata():
    rows = build_runtime_summary_rows(
        "tuho",
        generated_at="2026-05-23T10:00:00+00:00",
        source_branch="phase12-test",
    )
    assert rows
    assert all(row["commit_sha"] for row in rows)
    assert {row["generated_at"] for row in rows} == {"2026-05-23T10:00:00+00:00"}
    assert all(row["runtime_timestamp"] for row in rows)
    assert {row["template_origin"] for row in rows} == {"project_factory:tuho"}
    assert {row["template_revision"] for row in rows} == {"not separately versioned"}
    assert all(int(row["runtime_flag_count"]) == 15 for row in rows)
    assert all("deterministic replay is not fully guaranteed" in row["replay_limitations"] for row in rows)
    assert {row["g20_status"] for row in rows} == {"BLOCKED"}
    assert {row["r99_r102_status"] for row in rows} == {"NOT APPROVED"}


def test_persistence_records_store_replay_metadata_and_lineage(test_db):
    user_id = f"u{uuid.uuid4().hex[:8]}"
    replay_metadata = {
        "commit_sha": "abc123",
        "branch_name": "phase12-test",
        "runtime_timestamp": "2026-05-23T10:00:00+00:00",
        "template_origin": "project_factory:tuho",
        "runtime_flag_count": 15,
        "replay_limitations_notice": "Replay metadata improves traceability.",
    }
    project = save_project(
        user_id=user_id,
        project_code="tuho",
        project_name="TUHO Wind 1",
        source_project_template="tuho",
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"},
        last_run_summary={"project_irr": 0.094},
        replay_metadata=replay_metadata,
    )
    scenario = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Replay Snapshot",
        project_code="tuho",
        source_project_template="tuho",
        snapshot={"active_project": "tuho"},
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"},
        last_run_summary={"equity_irr": 0.113},
        replay_metadata=replay_metadata,
    )
    run_record = save_run(
        user_id=user_id,
        project_type="Wind",
        scenario="Base",
        inputs={"capacity_mw": "35"},
        kpis={"project_irr": 0.094},
        replay_metadata=replay_metadata,
    )
    export = record_export(
        user_id=user_id,
        project_code="tuho",
        export_type="runtime_summary_csv",
        artifact_name="phase12_tuho_runtime_summary.csv",
        project_id=project.project_id,
        scenario_id=scenario.scenario_id,
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"},
        replay_metadata=replay_metadata,
    )

    exports = list_exports(user_id, project_id=project.project_id)
    lineage = build_export_lineage(user_id, project.project_id)

    assert project.replay_metadata["commit_sha"] == "abc123"
    assert scenario.replay_metadata["commit_sha"] == "abc123"
    assert run_record.replay_metadata["run_id"] == run_record.run_id
    assert export.replay_metadata["export_id"] == export.export_id
    assert exports[0].replay_metadata["template_origin"] == "project_factory:tuho"
    assert lineage[0]["replay_metadata"]["commit_sha"] == "abc123"


def test_workbooks_show_commit_sha_template_provenance_and_replay_limits():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=True)
    cover_values = _sheet_values(workbook["Cover"])
    governance_values = _sheet_values(workbook["Governance"])
    audit_values = _sheet_values(workbook["Audit"])

    assert "Commit SHA" in cover_values
    assert "Template origin" in cover_values
    assert "Runtime flag snapshot" in cover_values
    assert any(isinstance(value, str) and "deterministic replay is not fully guaranteed" in value for value in cover_values)
    assert "Commit SHA" in governance_values
    assert "Replay limitations" in governance_values
    assert "Provenance" in audit_values

    tmp_path = ROOT / "reports" / f"phase12_test_replay_{uuid.uuid4().hex[:8]}"
    tmp_path.mkdir(parents=True, exist_ok=True)
    try:
        artifacts = write_calibration_reconciliation_pack(
            workbook_path=tmp_path / "phase12_replay_pack.xlsx",
            gap_register_path=tmp_path / "gap.csv",
            source_inventory_path=tmp_path / "inventory.csv",
            summary_path=tmp_path / "summary.csv",
            navigation_map_path=tmp_path / "nav.csv",
            signoff_matrix_path=tmp_path / "signoff.csv",
            runtime_binding_inventory_path=tmp_path / "binding_inventory.csv",
            evidence_coverage_summary_path=tmp_path / "coverage.csv",
            runtime_binding_gap_register_path=tmp_path / "binding_gap.csv",
            irr_reconciliation_summary_path=tmp_path / "irr_summary.csv",
            irr_convention_map_path=tmp_path / "irr_map.csv",
            irr_governance_items_path=tmp_path / "irr_governance.csv",
            co2_balancing_source_map_path=tmp_path / "co2_map.csv",
            co2_balancing_gap_summary_path=tmp_path / "co2_gap.csv",
            co2_balancing_evidence_quality_path=tmp_path / "co2_quality.csv",
            final_residual_registry_path=tmp_path / "residual.csv",
            governance_blocker_summary_path=tmp_path / "gov.csv",
            accepted_conventions_registry_path=tmp_path / "conventions.csv",
            phase11_carry_forward_items_path=tmp_path / "carry.csv",
            export_polish_checklist_path=tmp_path / "polish.csv",
            workbook_formatting_matrix_path=tmp_path / "formatting.csv",
            reviewer_navigation_paths_path=tmp_path / "reviewers.csv",
        )
        review_wb = load_workbook(artifacts.workbook_path, data_only=True)
        cover_values = _sheet_values(review_wb["Cover"])
        notes_values = _sheet_values(review_wb["Reviewer Notes"])
        assert "Commit SHA" in cover_values
        assert "Runtime Timestamp" in cover_values
        assert "Template Origin" in cover_values
        assert any(isinstance(value, str) and "deterministic replay is not fully guaranteed" in value for value in cover_values)
        assert any(
            isinstance(value, str) and "deterministic replay is not fully guaranteed" in value.lower()
            for value in notes_values
        )
    finally:
        for path in tmp_path.glob("*"):
            try:
                path.unlink()
            except OSError:
                pass
        try:
            tmp_path.rmdir()
        except OSError:
            pass


def test_reports_and_docs_exist_with_required_columns():
    assert DOC.exists()
    assert INVENTORY.exists()
    assert GAP_MATRIX.exists()
    assert FLAG_REPORT.exists()
    assert TRACEABILITY.exists()

    inventory_rows = list(csv.DictReader(INVENTORY.open(newline="", encoding="utf-8")))
    gap_rows = list(csv.DictReader(GAP_MATRIX.open(newline="", encoding="utf-8")))
    flag_rows = list(csv.DictReader(FLAG_REPORT.open(newline="", encoding="utf-8")))
    trace_rows = list(csv.DictReader(TRACEABILITY.open(newline="", encoding="utf-8")))

    assert {"metadata_item", "currently_captured", "replay_critical", "missing?", "source", "destination", "governance_sensitive", "notes"} <= set(inventory_rows[0].keys())
    assert {"metadata_item", "currently_captured", "replay_critical", "missing?", "source", "destination", "governance_sensitive", "notes"} <= set(gap_rows[0].keys())
    assert {"flag_name", "default_state", "governance_sensitive", "source_object", "source_path", "captured_in_replay_metadata", "notes"} <= set(flag_rows[0].keys())
    assert {"metadata_item", "currently_captured", "replay_critical", "missing?", "source", "destination", "governance_sensitive", "notes"} <= set(trace_rows[0].keys())

    text = DOC.read_text(encoding="utf-8")
    assert "No runtime formulas were changed" in text
    assert "G20 remains `BLOCKED`" in text
    assert "R99/R102 remain `NOT APPROVED`" in text


def test_runtime_authority_stays_separate():
    main_web_text = (ROOT / "main_web.py").read_text(encoding="utf-8")
    repo_text = (ROOT / "app" / "persistence" / "repository.py").read_text(encoding="utf-8")
    doc_text = DOC.read_text(encoding="utf-8")
    assert "build_replay_metadata" in main_web_text
    assert "but never computes or overrides" in repo_text
    assert "Runtime calculations remain authoritative" in doc_text
