"""TUHO Senior Debt Schedule diagnostic — Phase 7F-5 PR B.

Investigates why Python produces a balloon at op_idx 27 / 2043-12-31
while Excel does not.

Run: python tests/test_tuho_senior_debt_diagnostic.py
"""

import openpyxl
from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunner, WaterfallRunConfig


WB_PATH = "/root/.openclaw/workspace/oborovo_model/tuho_excel_1.xlsm"


def _extract_excel_ds():
    wb = openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
    ds = wb["DS"]
    cf = wb["CF"]

    ds_header = list(ds.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cf_header = list(cf.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    def _col(op_idx):
        return 5 + op_idx

    def _row_vals(sheet, row_num, col_start=5, col_end=None):
        rows = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        if not rows:
            return []
        row = rows[0]
        end = col_end if col_end else len(row)
        return row[col_start:end]

    data = {}
    for op_idx in range(20, 31):
        col = _col(op_idx)
        date = ds_header[col] if col < len(ds_header) else None
        data[op_idx] = {"date": date}

        # DS rows
        for rnum, key in [
            (19, "dscr"),
            (44, "max_senior_debt"),  # opening balance
            (47, "sr_beginning"),
            (49, "sr_principal"),
            (50, "sr_net_int"),
            (52, "sr_gross_int"),
            (53, "sr_end"),
            (54, "sr_ds"),
        ]:
            row = _row_vals(ds, rnum)
            data[op_idx][key] = row[col] if col < len(row) else None

        # CF sheet rows
        for rnum, key in [(99, "cf99"), (102, "cf102")]:
            row = _row_vals(cf, rnum)
            data[op_idx][key] = row[col] if col < len(row) else None

    return data


def _run_python():
    proj = create_default_tuho_wind1()
    engine = _build_period_engine(proj)
    config = WaterfallRunConfig.from_inputs(proj, engine)
    runner = WaterfallRunner(inputs=proj, engine=engine)
    result = runner.run(config)
    return result


def _python_balance_schedules(proj):
    """Reconstruct balance/payment/interest/principal schedules for TUHO.

    Mirrors the logic in waterfall_engine.py lines 415-437.
    """
    from domain.financing.sculpting_iterative import closed_form_sculpt

    engine = _build_period_engine(proj)
    periods = list(engine.periods())
    op_periods = [p for p in periods if p.is_operation]
    ebitda_schedule = [p.ebitda_keur for p in op_periods]

    fin = proj.financing
    rate_per_period = fin.senior_rate / 2
    service_periods = fin.senior_tenor_years * 2
    target_dscr = fin.target_dscr

    # Sculpt
    result = closed_form_sculpt(ebitda_schedule, target_dscr, [rate_per_period] * 100, 0.0, 0.0)
    scale = fin.fixed_debt_keur / result.debt_keur if result.debt_keur > 0 else 1.0

    bs = [b * scale for b in result.balance_schedule]
    payments = [p * scale for p in result.payment_schedule]
    interest_schedule = [bs[t] * rate_per_period for t in range(len(bs))]

    # Fixed DS override
    fixed_ds_keur = fin.fixed_ds_keur
    tenor_periods = service_periods
    if fixed_ds_keur and fixed_ds_keur > 0:
        fixed_ds = fixed_ds_keur
        principal_schedule = [max(0.0, fixed_ds - interest_schedule[t]) for t in range(tenor_periods)]
        principal_schedule = [
            min(principal_schedule[t], max(0.0, bs[t]))
            for t in range(tenor_periods)
        ]
        debt = bs[0]
        bs_fixed = [debt]
        bal = debt
        for t in range(tenor_periods):
            bal = max(0.0, bal - principal_schedule[t])
            bs_fixed.append(bal)
        bs = bs_fixed
        interest_schedule = [bs[t] * rate_per_period for t in range(len(bs))]
        principal_schedule = [max(0.0, fixed_ds - interest_schedule[t]) for t in range(tenor_periods)]
        principal_schedule = [
            min(principal_schedule[t], max(0.0, bs[t]))
            for t in range(tenor_periods)
        ]
        payments = [fixed_ds] * tenor_periods

    return {
        "balance_schedule": bs,
        "principal_schedule": principal_schedule,
        "interest_schedule": interest_schedule,
        "payments": payments,
        "tenor_periods": tenor_periods,
    }


def main():
    print("=== PROPAGATION CHECK ===")
    proj = create_default_tuho_wind1()
    fin = proj.financing
    print(f"FinancingParams.amortization_type = {fin.amortization_type!r}")
    print(f"FinancingParams.fixed_ds_keur = {fin.fixed_ds_keur}")
    print(f"FinancingParams.fixed_debt_keur = {fin.fixed_debt_keur}")
    print(f"FinancingParams.senior_tenor_years = {fin.senior_tenor_years}")

    # Check WaterfallRunConfig.from_inputs() — does it pass fixed_ds_keur?
    engine = _build_period_engine(proj)
    config = WaterfallRunConfig.from_inputs(proj, engine)
    print(f"\nWaterfallRunConfig.fixed_ds_keur = {config.fixed_ds_keur}")
    print(f"WaterfallRunConfig.tenor_periods = {config.tenor_periods}")
    print(f"WaterfallRunConfig.amortization_type = {getattr(config, 'amortization_type', 'NOT PRESENT')!r}")

    print("\n=== EXCEL DATA ===")
    excel = _extract_excel_ds()
    for op_idx in range(20, 31):
        ex = excel[op_idx]
        date = str(ex["date"])[:10] if ex["date"] else "?"
        print(f"OpIdx={op_idx} {date}: "
              f"SR_open={ex['sr_beginning']:.1f} "
              f"SR_prin={ex['sr_principal']:.1f} "
              f"SR_int={ex['sr_net_int']:.1f} "
              f"SR_ds={ex['sr_ds']:.1f} "
              f"SR_close={ex['sr_end']:.1f} "
              f"DSCR={ex['dscr']:.3f}")

    print("\n=== PYTHON WATERFALL ===")
    result = _run_python()
    for i in range(20, 31):
        p = result.periods[i]
        print(f"OpIdx={i} {p.date}: "
              f"sr_bal={p.senior_balance_keur:.1f} "
              f"sr_ds={p.senior_ds_keur:.1f} "
              f"sr_prin={p.senior_principal_keur:.1f} "
              f"sr_int={p.senior_interest_keur:.1f} "
              f"dscr={p.dscr:.3f} "
              f"lockup={p.lockup_active}")

    print("\n=== PYTHON BALANCE SCHEDULES ===")
    sched = _python_balance_schedules(proj)
    print(f"tenor_periods = {sched['tenor_periods']}")
    print(f"balance_schedule (indices 19-30):")
    for t in range(19, min(31, len(sched["balance_schedule"]))):
        pmt = sched["payments"][t] if t < len(sched["payments"]) else 0
        ints = sched["interest_schedule"][t] if t < len(sched["interest_schedule"]) else 0
        prin = sched["principal_schedule"][t] if t < len(sched["principal_schedule"]) else 0
        bs_t = sched["balance_schedule"][t] if t < len(sched["balance_schedule"]) else 0
        bs_next = sched["balance_schedule"][t+1] if t+1 < len(sched["balance_schedule"]) else 0
        print(f"  [{t}] bs={bs_t:.1f} bs_next={bs_next:.1f} payment={pmt:.1f} interest={ints:.1f} principal={prin:.1f}")

    print("\n=== KEY QUESTIONS ===")
    print("Q1: amortization_type propagated to run_waterfall? NO — not in WaterfallRunConfig")
    print("Q2: fixed_ds_keur propagated to run_waterfall? YES (via config.fixed_ds_keur)")
    print("Q3: fixed_ds_keur applied for all 28 periods? Only if amortization_type checked")
    print("Q4: final-period balloon forced? YES (line 557: sp = opening_balance if tenor-1)")
    print("Q5: remaining_senior_balance uses opening vs closing? Opening at t, closing at t+1")
    print("Q6: period_in_tenor dates align with Excel? Need to verify")


if __name__ == "__main__":
    main()