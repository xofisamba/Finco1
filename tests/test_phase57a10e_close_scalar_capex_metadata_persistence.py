from __future__ import annotations

import os
import sqlite3
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest


REPO_ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture
def test_db(monkeypatch):
    temp_root = REPO_ROOT / "app" / "data"
    temp_root.mkdir(exist_ok=True)
    db_file = temp_root / f"phase57a10e_close_{uuid.uuid4().hex[:8]}.db"
    monkeypatch.setenv("FINCO_DB_PATH", str(db_file))
    import app.persistence.db as db_mod

    monkeypatch.setattr(db_mod, "DB_PATH", str(db_file))
    db_mod.init_db()
    try:
        yield db_file
    finally:
        for suffix in ("", "-wal", "-shm"):
            try:
                (Path(str(db_file) + suffix)).unlink(missing_ok=True)
            except PermissionError:
                pass


def _snapshot(**overrides):
    data = {
        "active_project": "pilot-capex-metadata",
        "project_name": "Capex Metadata Project",
        "project_type": "Solar",
        "project_origin": "user_created",
        "template_source": "generic_solar",
        "country_market": "Croatia",
        "scenario": "Base",
        "capacity_mw": "50",
        "cod_date": "2027-01-01",
        "construction_months": "12",
        "horizon_years": "25",
        "tariff_eur_mwh": "60",
        "ppa_term_years": "15",
        "p50_hours": "1400",
        "opex_y1_keur": "1000",
        "total_capex_keur": "50000",
        "gearing_pct": "70",
        "interest_rate_pct": "5",
        "tenor_years": "15",
        "target_dscr": "1.30",
    }
    data.update(overrides)
    return data


def _load_wb(excel_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)


def _sheet_rows_by_business_code(ws) -> list[dict[str, object]]:
    headers = None
    found_header = False
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not found_header:
            if "Business Code" in values:
                headers = values
                found_header = True
            continue
        if all(v in (None, "") for v in values):
            continue
        if values[0] == "Note":
            continue
        record = {
            str(headers[idx]): values[idx] if idx < len(values) else None
            for idx in range(len(headers))
            if headers[idx] is not None
        }
        if record.get("Business Code") in (None, ""):
            continue
        rows.append(record)
    assert headers is not None
    return rows


def _notes_dict(ws) -> dict[str, str]:
    notes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = row[1] if len(row) > 1 else None
        if key == "Parent Category":
            break
        if key:
            notes[str(key)] = "" if value is None else str(value)
    return notes


def _make_user_project(user_id: str = "u57a10e-close"):
    from app.persistence.repository import create_project_record

    return create_project_record(
        user_id=user_id,
        project_code="pilot-capex-metadata",
        project_name="Capex Metadata Project",
        project_type="Solar",
        project_origin="user_project",
        template_source="generic_solar",
        baseline_snapshot=_snapshot(),
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
    )


def _metadata():
    return {
        "contingency_pct": 7.5,
        "vat_recoverable_flag": True,
        "vat_rate_pct": 25.0,
        "vat_basis_mode": "base_amount_only",
        "vat_jurisdiction_code": "HR",
        "wht_rate_pct": 10.0,
        "wht_treatment_mode": "withholding",
        "wht_gross_up_flag": False,
        "wht_jurisdiction_code": "HR",
        "depreciation_asset_class": "electrical_equipment",
        "depreciation_useful_life_years": 20,
        "depreciable_flag": True,
        "depreciation_basis_mode": "base_plus_contingency",
    }


def _create_sub_line(project_id: str, *, scalar_metadata: dict | None = None):
    from app.persistence.capex_sub_lines import create_sub_line

    with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
        conn.row_factory = sqlite3.Row
        sub = create_sub_line(
            conn.cursor(),
            project_id=project_id,
            parent_category_code="C.02",
            label="Metadata EPC",
            amount_keur=1000.0,
            comments="Metadata-bearing sub-line",
            scalar_metadata=scalar_metadata or {},
        )
        conn.commit()
        return sub


def _capex_amounts_tuple(capex) -> tuple[tuple[str, float], ...]:
    return tuple(
        (field, getattr(capex, field).amount_keur)
        for field in capex._CAPEX_ITEM_FIELDS
    )


class TestPhase57A10ECloseScalarCapexMetadataPersistence:
    def test_metadata_save_and_load_round_trip(self, test_db):
        from app.persistence.projects_repository import get_project_with_sub_lines

        project = _make_user_project()
        sub = _create_sub_line(project.project_id, scalar_metadata=_metadata())

        record, sub_lines = get_project_with_sub_lines("u57a10e-close", project.project_code)
        assert record is not None
        assert len(sub_lines) == 1
        assert sub_lines[0].sub_line_id == sub.sub_line_id
        assert sub_lines[0].scalar_metadata == _metadata()

    def test_metadata_survives_scenario_duplication_and_copy_integrity(self, test_db):
        from app.persistence.projects_repository import get_project_with_sub_lines
        from app.persistence.repository import add_scenario, duplicate_scenario, seed_scenarios_if_needed

        project = _make_user_project()
        sub = _create_sub_line(project.project_id, scalar_metadata=_metadata())
        base = seed_scenarios_if_needed(
            user_id="u57a10e-close",
            project_id=project.project_id,
            project_code=project.project_code,
            project_type="Solar",
            source_project_template="generic_solar",
            baseline_snapshot=_snapshot(),
            governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
            template_origin="generic_solar",
        )
        scenario_a = add_scenario(
            user_id="u57a10e-close",
            project_id=project.project_id,
            project_code=project.project_code,
            scenario_name="Scenario A",
            parent_scenario_id=base.scenario_id,
            base_input_set=base.base_input_set,
            overrides={"tariff_eur_mwh": "71"},
        )

        copied = duplicate_scenario("u57a10e-close", scenario_a.scenario_id, "Scenario Copy")
        assert copied is not None
        assert copied.copied_from_scenario_id == scenario_a.scenario_id

        _record, sub_lines = get_project_with_sub_lines("u57a10e-close", project.project_code)
        assert len(sub_lines) == 1
        assert sub_lines[0].sub_line_id == sub.sub_line_id
        assert sub_lines[0].scalar_metadata == _metadata()

    def test_metadata_exports_as_audit_only_and_runtime_used_is_no(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project()
        sub = _create_sub_line(project.project_id, scalar_metadata=_metadata())

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        notes = _notes_dict(wb["CapEx_SubLines_Audit"])
        assert notes["Metadata Scope"] == "Metadata only - does not affect Run."
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        row = next(r for r in rows if r["Business Code"] == sub.business_code)
        assert row["Metadata Runtime Used"] == "NO"
        assert row["Contingency %"] == 7.5
        assert row["VAT Rate %"] == 25
        assert row["VAT Basis Mode"] == "base_amount_only"
        assert row["WHT Treatment Mode"] == "withholding"
        assert row["Depreciation Useful Life (Years)"] == 20

    def test_metadata_does_not_affect_run_or_capex_totals(self, test_db):
        from app.persistence.capex_sub_lines import CapexSubLine, replace_sub_lines_for_project
        from app.services.capex_sub_lines_integration import _apply_user_sub_lines_to_capex
        from app.ui_runner import run_demo_project

        project = _make_user_project()
        sub = _create_sub_line(project.project_id, scalar_metadata=_metadata())

        demo = run_demo_project("Solar", "Base")
        folded_a = _apply_user_sub_lines_to_capex(
            demo.project_inputs.capex,
            project_id=project.project_id,
            scenario_overrides={},
        )

        updated = CapexSubLine(
            id=sub.id,
            sub_line_id=sub.sub_line_id,
            project_id=sub.project_id,
            parent_category_code=sub.parent_category_code,
            business_code=sub.business_code,
            display_order=sub.display_order,
            label=sub.label,
            amount_keur=sub.amount_keur,
            comments=sub.comments,
            schedule_json=sub.schedule_json,
            scalar_metadata={
                **_metadata(),
                "vat_rate_pct": 13.0,
                "wht_rate_pct": 3.0,
                "depreciation_useful_life_years": 15,
            },
            source=sub.source,
            is_active=True,
            governance_state=sub.governance_state,
            replay_metadata=sub.replay_metadata,
            created_at=sub.created_at,
            updated_at=sub.updated_at,
        )
        with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
            conn.row_factory = sqlite3.Row
            replace_sub_lines_for_project(
                conn.cursor(),
                project_id=project.project_id,
                sub_lines=[updated],
            )
            conn.commit()

        folded_b = _apply_user_sub_lines_to_capex(
            demo.project_inputs.capex,
            project_id=project.project_id,
            scenario_overrides={},
        )

        assert _capex_amounts_tuple(folded_a) == _capex_amounts_tuple(folded_b)

    def test_cost_per_mw_remains_non_persisted_derived_only_metadata(self, test_db):
        from app.persistence.capex_sub_lines import create_sub_line

        project = _make_user_project()
        with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
            conn.row_factory = sqlite3.Row
            with pytest.raises(ValueError, match="Derived-only CAPEX metadata keys"):
                create_sub_line(
                    conn.cursor(),
                    project_id=project.project_id,
                    parent_category_code="C.02",
                    label="Derived-only cost/MW",
                    amount_keur=1000.0,
                    scalar_metadata={"cost_per_mw": 12.5},
                )
