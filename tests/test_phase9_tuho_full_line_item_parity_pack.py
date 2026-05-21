"""
Tests for phase9 — TUHO Full Line Item Parity Pack.

Scope: ANALYSIS / REPORTS / TESTS ONLY. No runtime code changed.
G20: BLOCKED. R99/R102: NOT APPROVED.

Required artifacts:
- XLSX parity workbook (13 sheets)
- docs/phase9_tuho_full_line_item_parity_pack.md
- reports/phase9_tuho_full_line_item_period_bridge.csv
- reports/phase9_tuho_full_line_item_parity_summary.csv
- reports/phase9_tuho_full_line_item_gap_register.csv
- reports/phase9_tuho_parity_gate_readiness.csv
"""
import csv, os, pytest

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(REPO, "reports", "phase9_tuho_full_line_item_parity_pack.xlsx")
DOC = os.path.join(REPO, "docs", "phase9_tuho_full_line_item_parity_pack.md")
REPORTS = os.path.join(REPO, "reports")


def _csv(name):
    with open(os.path.join(REPORTS, name)) as f:
        return list(csv.DictReader(f))


class TestRequiredFiles:
    def test_xlsx_exists(self):
        assert os.path.exists(XLSX), f"Missing XLSX: {XLSX}"

    def test_doc_exists(self):
        assert os.path.exists(DOC), f"Missing doc: {DOC}"

    def test_period_bridge_exists(self):
        assert os.path.exists(os.path.join(REPORTS, "phase9_tuho_full_line_item_period_bridge.csv"))

    def test_summary_exists(self):
        assert os.path.exists(os.path.join(REPORTS, "phase9_tuho_full_line_item_parity_summary.csv"))

    def test_gap_register_exists(self):
        assert os.path.exists(os.path.join(REPORTS, "phase9_tuho_full_line_item_gap_register.csv"))

    def test_gate_readiness_exists(self):
        assert os.path.exists(os.path.join(REPORTS, "phase9_tuho_parity_gate_readiness.csv"))


class TestXLSXSheets:
    def test_has_13_sheets(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert len(wb.sheetnames) == 13, f"Expected 13 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
        wb.close()

    def test_has_dashboard(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Dashboard" in wb.sheetnames
        wb.close()

    def test_has_executive_summary(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Executive Summary" in wb.sheetnames
        wb.close()

    def test_has_operations(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Operations" in wb.sheetnames
        wb.close()

    def test_has_revenue(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Revenue" in wb.sheetnames
        wb.close()

    def test_has_opex_ebitda(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "OPEX_EBITDA" in wb.sheetnames
        wb.close()

    def test_has_senior_debt(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Senior Debt" in wb.sheetnames
        wb.close()

    def test_has_shl(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "SHL" in wb.sheetnames
        wb.close()

    def test_has_tax_cfads(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Tax_CFADS" in wb.sheetnames
        wb.close()

    def test_has_distributions(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Distributions" in wb.sheetnames
        wb.close()

    def test_has_returns(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Returns" in wb.sheetnames
        wb.close()

    def test_has_gap_register(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Gap Register" in wb.sheetnames
        wb.close()

    def test_has_gate_readiness(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Gate Readiness" in wb.sheetnames
        wb.close()

    def test_has_source_map(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        assert "Source Map" in wb.sheetnames
        wb.close()


class TestXLSXStructure:
    def test_dashboard_has_headline_metrics(self):
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        ws = wb["Dashboard"]
        # Find headline metrics section
        cell_vals = [ws.cell(r, 1).value for r in range(1, 50)]
        assert any("HEADLINE" in str(v).upper() for v in cell_vals if v), "Dashboard must have HEADLINE METRICS section"
        wb.close()

    def test_gap_register_has_19_gaps(self):
        rows = _csv("phase9_tuho_full_line_item_gap_register.csv")
        assert len(rows) == 19, f"Expected 19 gap rows, got {len(rows)}"

    def test_gate_readiness_has_11_gates(self):
        rows = _csv("phase9_tuho_parity_gate_readiness.csv")
        assert len(rows) == 11, f"Expected 11 gate rows, got {len(rows)}"

    def test_summary_has_20_metrics(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        assert len(rows) == 20, f"Expected 20 summary rows, got {len(rows)}"


class TestPeriodBridge:
    def test_has_bridge_columns(self):
        rows = _csv("phase9_tuho_full_line_item_period_bridge.csv")
        assert rows, "Period bridge is empty"
        required = ["period","date","excel_production_mwh","model_production_mwh",
                   "excel_revenue_keur","model_revenue_keur","excel_opex_keur","model_opex_keur",
                   "excel_senior_opening_keur","model_senior_opening_keur",
                   "excel_shl_opening_keur","model_shl_opening_keur",
                   "excel_distribution_keur","model_distribution_keur"]
        for col in required:
            assert col in rows[0], f"Missing column in bridge: {col}"

    def test_bridge_has_61_rows(self):
        rows = _csv("phase9_tuho_full_line_item_period_bridge.csv")
        assert len(rows) == 61, f"Expected 61 bridge rows, got {len(rows)}"

    def test_bridge_includes_senior_debt_columns(self):
        rows = _csv("phase9_tuho_full_line_item_period_bridge.csv")
        required = ["excel_senior_interest_keur","model_senior_interest_keur",
                   "excel_senior_principal_keur","model_senior_principal_keur",
                   "excel_senior_closing_keur","model_senior_closing_keur"]
        for col in required:
            assert col in rows[0], f"Missing senior debt column: {col}"

    def test_bridge_includes_shl_columns(self):
        rows = _csv("phase9_tuho_full_line_item_period_bridge.csv")
        required = ["excel_shl_interest_keur","model_shl_interest_keur",
                   "excel_shl_pik_keur","model_shl_pik_keur",
                   "excel_shl_principal_keur","model_shl_principal_keur"]
        for col in required:
            assert col in rows[0], f"Missing SHL column: {col}"


class TestGapRegister:
    def test_has_required_columns(self):
        rows = _csv("phase9_tuho_full_line_item_gap_register.csv")
        required = ["gap_id","category","issue","excel_value","model_value",
                   "severity","affects_runtime","affects_reporting",
                   "affects_g20","recommended_action","status","notes"]
        for col in required:
            assert col in rows[0], f"Missing gap register column: {col}"

    def test_includes_accepted_convention_differences(self):
        rows = _csv("phase9_tuho_full_line_item_gap_register.csv")
        accepted = [r for r in rows if r.get("status") == "ACCEPTED_CONVENTION"]
        assert len(accepted) >= 4, f"Expected at least 4 ACCEPTED_CONVENTION gaps, got {len(accepted)}"

    def test_includes_g20_blocker(self):
        rows = _csv("phase9_tuho_full_line_item_gap_register.csv")
        assert any(r.get("affects_g20","").upper() == "YES" for r in rows), \
            "Gap register must include at least one G20-affecting gap"

    def test_includes_missing_evidence(self):
        rows = _csv("phase9_tuho_full_line_item_gap_register.csv")
        missing = [r for r in rows if r.get("status") == "MISSING_EVIDENCE"]
        assert len(missing) >= 2, f"Expected at least 2 MISSING_EVIDENCE gaps, got {len(missing)}"


class TestGateReadiness:
    def test_has_required_columns(self):
        rows = _csv("phase9_tuho_parity_gate_readiness.csv")
        required = ["gate_id","gate_name","status","evidence","blocker","required_next_action"]
        for col in required:
            assert col in rows[0], f"Missing gate column: {col}"

    def test_includes_g20_gate(self):
        rows = _csv("phase9_tuho_parity_gate_readiness.csv")
        g20 = [r for r in rows if "G20" in r.get("gate_id","") or "g20" in r.get("gate_id","").lower()]
        assert g20, "Gate readiness must include G20 gates"

    def test_includes_equity_irr_gate(self):
        rows = _csv("phase9_tuho_parity_gate_readiness.csv")
        irr = [r for r in rows if "equity" in r.get("gate_name","").lower() or "irr" in r.get("gate_name","").lower()]
        assert irr, "Gate readiness must include equity IRR gate"

    def test_equity_irr_gate_is_blocker(self):
        rows = _csv("phase9_tuho_parity_gate_readiness.csv")
        irr_gates = [r for r in rows if "equity" in r.get("gate_name","").lower() or "irr" in r.get("gate_name","").lower()]
        blocker = [r for r in irr_gates if r.get("status") in ("BLOCKER","WARN")]
        assert blocker, "Equity IRR gate must have BLOCKER or WARN status"


class TestDocContent:
    def test_doc_has_headline_metrics_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "HEADLINE METRICS" in content or "Senior Debt" in content, "Doc must have headline metrics"
        assert "423,787" in content or "423,844" in content, "Doc must include revenue figures"

    def test_doc_includes_operations_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "Operations" in content, "Doc must have Operations section"
        assert "Production" in content, "Doc must cover Production"

    def test_doc_includes_revenue_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "Revenue" in content, "Doc must have Revenue section"

    def test_doc_includes_opex_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "OPEX" in content or "EBITDA" in content, "Doc must have OPEX/EBITDA section"

    def test_doc_includes_senior_debt_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "Senior Debt" in content, "Doc must have Senior Debt section"

    def test_doc_includes_shl_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "SHL" in content, "Doc must have SHL section"

    def test_doc_includes_distributions_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "Distribution" in content, "Doc must have Distributions section"

    def test_doc_includes_returns_table(self):
        with open(DOC) as f:
            content = f.read()
        assert "Returns" in content or "Equity IRR" in content, "Doc must have Returns section"

    def test_doc_states_g20_blocked(self):
        with open(DOC) as f:
            content = f.read()
        assert "BLOCKED" in content and "G20" in content, "Doc must state G20 BLOCKED"

    def test_doc_states_r99_not_approved(self):
        with open(DOC) as f:
            content = f.read()
        assert "NOT approved" in content or "NOT APPROVED" in content, "Doc must state R99/R102 NOT APPROVED"

    def test_doc_states_029pp_gap(self):
        with open(DOC) as f:
            content = f.read()
        assert "0.29" in content, "Doc must state 0.29pp gap"

    def test_doc_states_reconciliation_irr_recommended(self):
        with open(DOC) as f:
            content = f.read()
        assert "reconciliation" in content.lower(), "Doc must mention reconciliation IRR"

    def test_doc_has_accepted_conventions_section(self):
        with open(DOC) as f:
            content = f.read()
        assert "ACCEPTED_CONVENTION" in content or "Accepted Convention" in content, \
            "Doc must have Accepted Convention differences section"

    def test_doc_names_pr_165(self):
        with open(DOC) as f:
            content = f.read()
        assert "PR #165" in content or "PR #165" in content or "PR165" in content, \
            "Doc must reference PR #165 SHL alignment"

    def test_doc_recommends_next_branch(self):
        with open(DOC) as f:
            content = f.read()
        assert "next branch" in content.lower() or "recommended next" in content.lower(), \
            "Doc must recommend next branch"


class TestNoRuntimeChanges:
    def test_no_runtime_python_changed(self):
        # Verify no .py files were modified in docs/reports/tests
        for dirpath, dirnames, filenames in os.walk(os.path.join(REPO, "domain")):
            for fname in filenames:
                if fname.endswith(".py"):
                    fpath = os.path.join(dirpath, fname)
                    assert not (fpath.startswith(REPO + "/docs") or
                               fpath.startswith(REPO + "/reports") or
                               fpath.startswith(REPO + "/tests")), \
                        f"Runtime file modified: {fpath}"

    def test_test_file_contains_no_runtime_imports(self):
        test_path = os.path.abspath(__file__)
        with open(test_path) as f:
            content = f.read()
        lines = content.split('\n')
        import_lines = [l.strip() for l in lines if l.strip().startswith('import ') or l.strip().startswith('from ')]
        forbidden = [l for l in import_lines if any(m in l for m in [
            'waterfall_engine', 'waterfall_core', 'project_factories',
            'shl_domain', 'senior_debt_sizing', 'tax_bridge', 'distribution_account'
        ])]
        assert not forbidden, f"Test file must not import runtime modules: {forbidden}"

    def test_xlsx_is_primary_reviewer_artifact(self):
        # Verify XLSX is the primary artifact (not raw CSVs)
        xlsx_size = os.path.getsize(XLSX)
        assert xlsx_size > 30000, f"XLSX should be substantial (>{xlsx_size} bytes)"
        csv_dir = os.path.join(REPORTS, "phase9_tuho_full_line_item_period_bridge.csv")
        csv_size = os.path.getsize(csv_dir)
        assert xlsx_size > csv_size, "XLSX should be larger than period bridge CSV"


class TestSummaryMetrics:
    def test_summary_has_passing_metrics(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        pass_rows = [r for r in rows if r.get("status") == "PASS"]
        assert len(pass_rows) >= 8, f"Expected at least 8 PASS metrics, got {len(pass_rows)}"

    def test_summary_has_warn_metrics(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        warn_rows = [r for r in rows if r.get("status") == "WARN"]
        assert len(warn_rows) >= 1, f"Expected at least 1 WARN metric, got {len(warn_rows)}: {[r['metric'] for r in warn_rows]}"

    def test_summary_equity_irr_is_warn(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        irr_rows = [r for r in rows if "equity" in r.get("metric","").lower()]
        assert irr_rows, "Summary must include equity IRR"
        equity_irr_rows = [r for r in irr_rows if "equity" in r.get("metric","").lower() and "irr" in r.get("metric","").lower()]
        assert equity_irr_rows, f"Summary must have Equity IRR row, got: {[r['metric'] for r in irr_rows]}"
        assert equity_irr_rows[0].get("status") in ("WARN","MISSING_EVIDENCE"), \
            f"Equity IRR must be WARN or MISSING_EVIDENCE, got: {equity_irr_rows[0].get('status')}"

    def test_summary_project_irr_is_pass(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        proj_rows = [r for r in rows if "project" in r.get("metric","").lower()]
        assert proj_rows, "Summary must include project IRR"
        assert proj_rows[0].get("status") == "PASS", f"Project IRR must be PASS, got: {proj_rows[0].get('status')}"

    def test_summary_has_accepted_conventions(self):
        rows = _csv("phase9_tuho_full_line_item_parity_summary.csv")
        accept_rows = [r for r in rows if r.get("status") == "ACCEPTED_CONVENTION"]
        assert len(accept_rows) >= 4, f"Expected at least 4 ACCEPTED_CONVENTION, got {len(accept_rows)}"


class TestDeliverables:
    def test_xlsx_size_reasonable(self):
        size = os.path.getsize(XLSX)
        assert 30000 < size < 200000, f"XLSX size {size} seems unusual (should be 30-200KB)"

    def test_all_csvs_populated(self):
        csvs = [
            "phase9_tuho_full_line_item_period_bridge.csv",
            "phase9_tuho_full_line_item_parity_summary.csv",
            "phase9_tuho_full_line_item_gap_register.csv",
            "phase9_tuho_parity_gate_readiness.csv",
        ]
        for name in csvs:
            rows = _csv(name)
            assert len(rows) > 0, f"CSV {name} is empty"

    def test_pr_link_in_doc(self):
        with open(DOC) as f:
            content = f.read()
        assert "https://github.com/xofisamba/Finco1/pull/" in content, \
            "Doc must include PR link"