from __future__ import annotations

import csv
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.export.institutional_workbook import (
    INSTITUTIONAL_SHEET_DEFINITIONS,
    export_institutional_workbook_skeleton,
)
from app.export.registry import default_export_registry


ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "phase10_runtime_to_workbook_binding.md"
SHEET_MAP = ROOT / "reports" / "phase10_workbook_sheet_map.csv"
BINDING_STATUS = ROOT / "reports" / "phase10_runtime_workbook_binding_status.csv"
REGISTRY_REPORT = ROOT / "reports" / "phase10_export_registry.csv"


def _open_workbook(project: str):
    return load_workbook(BytesIO(export_institutional_workbook_skeleton(project)), data_only=True)


def _sheet_values(sheet) -> list[object]:
    return [
        sheet.cell(row=row, column=column).value
        for row in range(1, sheet.max_row + 1)
        for column in range(1, sheet.max_column + 1)
    ]


def _contains(sheet, value: object) -> bool:
    return value in _sheet_values(sheet)


def _value_for_label(sheet, label: str) -> object:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return sheet.cell(row=row, column=2).value
    raise AssertionError(f"Label {label!r} not found in sheet {sheet.title!r}")


def test_workbook_generates_with_required_sheet_names():
    workbook = _open_workbook("tuho")
    assert workbook.sheetnames == [
        definition.sheet_name for definition in INSTITUTIONAL_SHEET_DEFINITIONS
    ]


def test_governance_metadata_is_preserved():
    workbook = _open_workbook("tuho")
    governance = workbook["Governance"]

    values = _sheet_values(governance)
    assert "Governance and provenance" in values
    assert "BLOCKED" in values
    assert "NOT APPROVED" in values
    assert "Project context + runtime result + offline financial statements" in values


def test_runtime_summary_and_bound_sheets_are_populated():
    workbook = _open_workbook("tuho")

    assert _contains(workbook["Inputs"], "Project metadata")
    assert _contains(workbook["OPEX"], "OPEX summary")
    assert _contains(workbook["Revenue"], "Revenue assumptions and totals")
    assert _contains(workbook["Senior Debt"], "Senior debt summary")
    assert _contains(workbook["SHL"], "SHL summary")
    assert _contains(workbook["Tax"], "Tax summary")
    assert _contains(workbook["P&L"], "P&L period table")
    assert _contains(workbook["Cash Flow"], "Project finance cash waterfall")
    assert _contains(workbook["Balance Sheet"], "Balance sheet period table")
    assert _contains(workbook["Audit"], "Audit source map")
    assert _contains(workbook["Gap Register"], "Known gaps and accepted conventions")


def test_runtime_vs_preview_labels_exist():
    workbook = _open_workbook("tuho")

    assert _contains(workbook["Inputs"], "runtime + template assumptions")
    assert _contains(workbook["Construction"], "template + runtime")
    assert _contains(workbook["P&L"], "runtime")
    assert _contains(workbook["Audit"], "review")


def test_tuho_and_oborovo_exports_both_work_and_differ():
    tuho = _open_workbook("tuho")
    oborovo = _open_workbook("oborovo")

    assert _contains(tuho["Inputs"], "TUHO")
    assert _contains(oborovo["Inputs"], "OBOROVO")
    assert tuho["Cover"]["B5"].value != oborovo["Cover"]["B5"].value
    assert _value_for_label(tuho["Runtime Summary"], "Project IRR") != _value_for_label(
        oborovo["Runtime Summary"], "Project IRR"
    )


def test_shl_and_debt_runtime_values_are_not_placeholders():
    workbook = _open_workbook("tuho")
    shl_values = _sheet_values(workbook["SHL"])
    debt_values = _sheet_values(workbook["Senior Debt"])

    assert "SHL period schedule" in shl_values
    assert "Senior debt period table" in debt_values
    assert any(isinstance(value, (int, float)) and abs(value) > 1000 for value in shl_values)
    assert any(isinstance(value, (int, float)) and abs(value) > 1000 for value in debt_values)


def test_reports_and_registry_capture_runtime_binding_status():
    registry = default_export_registry()
    names = {artifact.artifact_name for artifact in registry.artifacts}
    assert "Phase 10 institutional workbook skeleton" in names

    report_rows = list(csv.DictReader(REGISTRY_REPORT.open(newline="", encoding="utf-8")))
    workbook_row = next(
        row for row in report_rows if row["artifact_name"] == "Phase 10 institutional workbook skeleton"
    )
    assert workbook_row["current_status"] == "runtime_binding_in_progress"

    sheet_rows = list(csv.DictReader(SHEET_MAP.open(newline="", encoding="utf-8")))
    binding_rows = list(csv.DictReader(BINDING_STATUS.open(newline="", encoding="utf-8")))
    assert len(sheet_rows) == 16
    assert len(binding_rows) == 16
    assert any(row["sheet_name"] == "P&L" and row["implemented_level"] == "runtime_bound" for row in sheet_rows)
    assert any(row["sheet_name"] == "Tax" and row["runtime_binding_level"] == "runtime_bound" for row in binding_rows)


def test_docs_keep_runtime_scope_and_governance_constraints():
    text = DOC.read_text(encoding="utf-8")
    assert "no runtime changes" in text.lower()
    assert "G20 remains `BLOCKED`" in text
    assert "R99/R102 remains `NOT APPROVED`" in text
    assert "runtime versus template assumptions" in text.lower()

    main_web = (ROOT / "main_web.py").read_text(encoding="utf-8")
    assert '@app.get("/exports/institutional-workbook.xlsx")' in main_web
    assert "build_institutional_workbook_export(runtime_project_code, safe_project=safe_project)" in main_web
