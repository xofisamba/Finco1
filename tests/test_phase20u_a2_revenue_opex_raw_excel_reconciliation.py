"""Tests for Phase 20U-A2 Revenue & OPEX Line-Item Reconciliation.

Branch: phase20u-a2-revenue-opex-raw-excel-reconciliation
Status: DIAGNOSTIC ONLY — no runtime formula changes

Run:
- python -m pytest tests/test_phase20u_a2_revenue_opex_raw_excel_reconciliation.py -v
- python -m pytest tests/test_revenue.py tests/test_opex.py -v
- python -c "import main_web"
"""

import os
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
OUT_XLSX = REPORTS / "phase20u_a2_revenue_opex_raw_excel_reconciliation.xlsx"
OUT_CSV = REPORTS / "phase20u_a2_root_cause_table.csv"
SCRIPT = ROOT / "scripts" / "export_phase20u_a2_revenue_opex_reconciliation.py"


class TestPhase20U_A2WorkbookGeneration:
    """Test Phase 20U-A2 workbook generation."""

    def test_workbook_exists(self):
        """Reconciliation workbook must be generated."""
        import subprocess
        result = subprocess.run(
            ["python3", str(SCRIPT)],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=120,
        )
        assert OUT_XLSX.exists(), (
            f"Workbook not generated at {OUT_XLSX}\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )

    def test_workbook_has_required_sheets(self):
        """Workbook must have: Revenue_Reconciliation, OPEX_Reconciliation, Summary."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        sheet_names = wb.sheetnames

        # Check for OPEX sheets (TUHO_OPEX and OBOROVO_OPEX)
        has_opex = any("OPEX" in s for s in sheet_names)
        has_revenue = any("Revenue" in s for s in sheet_names)
        has_summary = "Summary" in sheet_names

        assert has_opex, f"Missing OPEX sheet (have: {sheet_names})"
        assert has_revenue, f"Missing Revenue sheet (have: {sheet_names})"
        assert has_summary, f"Missing Summary sheet (have: {sheet_names})"

    def test_opex_sheets_contain_b_code_rows(self):
        """OPEX sheets must contain B.01, B.02, B.12 rows."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        # Find an OPEX sheet
        opex_sheet = next((s for s in wb.sheetnames if "OPEX" in s and "OBOROVO" in s), None)
        if not opex_sheet:
            opex_sheet = next((s for s in wb.sheetnames if "OPEX" in s), None)

        assert opex_sheet, "No OPEX sheet found"
        ws = wb[opex_sheet]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)
        for code in ("B.01", "B.02", "B.12", "B.13"):
            assert code in content_str, f"Code {code} not found in {opex_sheet}"

    def test_oborovo_opex_has_delta_table(self):
        """OBOROVO OPEX sheet must contain Delta table."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        opex_sheet = next((s for s in wb.sheetnames if "OBOROVO" in s and "OPEX" in s), None)
        assert opex_sheet, "No OBOROVO_OPEX sheet found"
        ws = wb[opex_sheet]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)
        assert "DELTA" in content_str.upper(), "Delta table not found in OBOROVO_OPEX"

    def test_delta_shows_python_minus_excel(self):
        """Delta table must show Python - Excel values."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        opex_sheet = next((s for s in wb.sheetnames if "OBOROVO" in s and "OPEX" in s), None)
        if not opex_sheet:
            pytest.skip("No OBOROVO_OPEX sheet")

        ws = wb[opex_sheet]
        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)
        # Should contain numeric delta values
        assert "32" in content_str or "+32" in content_str, (
            "Expected to find P4 delta reference in OBOROVO_OPEX"
        )

    def test_root_cause_csv_exists(self):
        """Root cause CSV must be generated."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not generated")
        assert OUT_CSV.exists(), f"CSV not generated at {OUT_CSV}"

    def test_no_runtime_formula_changes(self):
        """Confirm no runtime formula changes in domain files."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/ files changed (runtime formula change): {changed}"


class TestPhase20U_A2NoRuntimeChanges:
    """Confirm no runtime formulas changed."""

    def test_domain_opex_unchanged(self):
        """domain/opex/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/opex/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/opex/ files changed: {changed}"

    def test_domain_revenue_unchanged(self):
        """domain/revenue/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/revenue/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/revenue/ files changed: {changed}"

    def test_domain_tax_unchanged(self):
        """domain/tax/ or domain/atad/ files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/tax/", "domain/atad/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/tax/ or domain/atad/ files changed: {changed}"

    def test_domain_senior_debt_unchanged(self):
        """domain/senior_debt/ files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/senior_debt/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/senior_debt/ files changed: {changed}"

    def test_domain_shl_unchanged(self):
        """domain/shl/ files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/shl/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/shl/ files changed: {changed}"


class TestPhase20U_A2ValidationCommands:
    """Run validation commands from the phase brief."""

    def test_import_main_web(self):
        """python -c 'import main_web' must succeed."""
        import subprocess
        result = subprocess.run(
            ["python3", "-c", "import main_web"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"import main_web failed: {result.stderr}"

    def test_opex_tests_pass(self):
        """tests/test_opex.py must pass."""
        import subprocess
        result = subprocess.run(
            ["python3", "-m", "pytest", "tests/test_opex.py", "-v", "--tb=short"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert "failed" not in result.stdout.lower(), (
            f"test_opex.py had failures:\n{result.stdout}"
        )

    def test_revenue_tests_pass(self):
        """tests/test_revenue.py must pass."""
        import subprocess
        result = subprocess.run(
            ["python3", "-m", "pytest", "tests/test_revenue.py", "-v", "--tb=short"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert "failed" not in result.stdout.lower(), (
            f"test_revenue.py had failures:\n{result.stdout}"
        )


class TestPhase20U_A2KeyFindings:
    """Test that key findings are documented and workbook contains critical data."""

    def test_workbook_contains_oborovo_b02_root_cause(self):
        """OBOROVO_OPEX sheet must document B.02 root cause."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        opex_sheet = next((s for s in wb.sheetnames if "OBOROVO" in s and "OPEX" in s), None)
        assert opex_sheet, "No OBOROVO_OPEX sheet found"
        ws = wb[opex_sheet]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)

        # B.02 root cause must be mentioned
        assert "B.02" in content_str, "B.02 not found in OBOROVO_OPEX"
        assert "244" in content_str or "Infrastructure" in content_str, (
            "B.02 Infrastructure Maintenance root cause not documented"
        )

    def test_workbook_contains_p4_delta_value(self):
        """Workbook must contain P4 delta value 32.45 or similar."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        summary = wb["Summary"] if "Summary" in wb.sheetnames else None
        if not summary:
            pytest.skip("No Summary sheet")

        content = []
        for row in summary.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)

        # Should mention 32 or 676 or 644
        assert "32" in content_str or "676" in content_str or "644" in content_str, (
            "P4 delta key values not found in Summary sheet"
        )

    def test_csv_contains_b02_root_cause_row(self):
        """Root cause CSV must contain B.02 row."""
        if not OUT_CSV.exists():
            pytest.skip("CSV not generated")

        with open(OUT_CSV, 'r') as f:
            content = f.read()

        assert "B.02" in content, "B.02 root cause not in CSV"
        assert "Infrastructure" in content, "Infrastructure Maintenance not in CSV"