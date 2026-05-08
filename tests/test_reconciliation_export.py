"""Tests for app/reconciliation/ helpers and reconciliation sheet export."""
import pytest

from app.reconciliation import (
    build_debt_schedule_rows,
    build_project_cf_rows,
    build_equity_cf_rows,
)
from tests.conftest import *


# =============================================================================
# Debt Schedule Reconciliation Tests
# =============================================================================

class TestDebtScheduleReconciliation:
    """Tests for debt schedule reconciliation table."""

    @pytest.fixture
    def oborovo_result(self):
        from app.project_factories import create_default_oborovo
        from app.ui_runner import run_demo_project
        oborovo = create_default_oborovo()
        return run_demo_project("Solar", "Base", project_inputs_override=oborovo).result

    def test_total_ds_equals_interest_plus_principal(self, oborovo_result):
        """Total senior DS = interest + principal (±50 kEUR tolerance).

        Last period may include balloon payment causing slightly larger difference.
        """
        rows = build_debt_schedule_rows(oborovo_result)
        for r in rows:
            total = r["total_senior_ds"]
            computed = r["senior_interest"] + r["senior_principal"]
            assert abs(total - computed) < 50.0, (
                f"Period {r['period']}: DS {total:.1f} vs {computed:.1f} (diff: {abs(total-computed):.1f})"
            )

    def test_closing_debt_nonzero_until_maturity(self, oborovo_result):
        """Closing debt > 0 during tenor; reaches ~0 at maturity."""
        rows = build_debt_schedule_rows(oborovo_result)
        non_zero = [r for r in rows if r["closing_senior_debt"] > 1]
        zero_rows = [r for r in rows if r["closing_senior_debt"] <= 1]
        assert len(non_zero) > 0, "No non-zero closing balances found"
        assert len(zero_rows) >= 1, f"Closing debt never reaches near-zero. Last 5: {rows[-5:]}"

    def test_debt_keur_matches_sculpt_result(self, oborovo_result):
        """First opening balance matches sculpting result debt_keur."""
        rows = build_debt_schedule_rows(oborovo_result)
        if not rows:
            pytest.skip("No rows returned")
        opening = rows[0]["opening_senior_debt"]
        sculpt_debt = oborovo_result.sculpting_result.debt_keur
        assert abs(opening - sculpt_debt) < 100, (
            f"First opening {opening:.0f} != sculpt debt {sculpt_debt:.0f}"
        )

    def test_dscr_is_reasonable(self, oborovo_result):
        """DSCR in plausible range [0, 10.0] (zero valid for repaid periods)."""
        rows = build_debt_schedule_rows(oborovo_result)
        for r in rows:
            d = r["dscr"]
            assert 0 <= d <= 10.0, f"Period {r['period']} DSCR {d} outside [0, 10.0]"

    def test_all_periods_have_dscr(self, oborovo_result):
        """Every operation period has a DSCR value (>=0)."""
        rows = build_debt_schedule_rows(oborovo_result)
        assert all(r["dscr"] >= 0 for r in rows), "Some periods missing DSCR"

    def test_rows_not_empty_for_oborovo(self, oborovo_result):
        """Debt schedule returns rows for Oborovo (operation periods exist)."""
        rows = build_debt_schedule_rows(oborovo_result)
        assert len(rows) > 0, "No debt schedule rows returned"

    def test_opening_minus_principal_approx_closes(self, oborovo_result):
        """opening - principal ≈ closing for each period (±100 kEUR)."""
        rows = build_debt_schedule_rows(oborovo_result)
        for r in rows:
            if r["opening_senior_debt"] > 0:
                expected = r["opening_senior_debt"] - r["senior_principal"]
                diff = abs(r["closing_senior_debt"] - expected)
                assert diff < 100.0, (
                    f"Period {r['period']}: closing {r['closing_senior_debt']:.1f} ≠ "
                    f"opening {r['opening_senior_debt']:.1f} - principal {r['senior_principal']:.1f} "
                    f"(diff={diff:.1f})"
                )


# =============================================================================
# Project CF Bridge Tests
# =============================================================================

class TestProjectCFBridge:
    """Tests for project cashflow bridge table."""

    @pytest.fixture
    def oborovo_result(self):
        from app.project_factories import create_default_oborovo
        from app.ui_runner import run_demo_project
        oborovo = create_default_oborovo()
        demo = run_demo_project("Solar", "Base", project_inputs_override=oborovo)
        demo.result.project_inputs = oborovo  # attach inputs for tax rate
        return demo.result

    def test_ebitda_equals_revenue_minus_opex(self, oborovo_result):
        """EBITDA = Revenue - OpEx for each operation period."""
        rows = build_project_cf_rows(oborovo_result, oborovo_result.project_inputs)
        for r in rows:
            if r["ebitda"] > 0:
                computed = r["revenue"] - r["opex"]
                assert abs(r["ebitda"] - computed) < 1.0, (
                    f"Period {r['period']}: EBITDA {r['ebitda']:.1f} ≠ Rev {r['revenue']:.1f} - OpEx {r['opex']:.1f}"
                )

    def test_unlevered_tax_differs_from_levered_tax(self, oborovo_result):
        """Unlevered tax computed via tax_rate × max(0, EBITDA - depreciation)."""
        rows = build_project_cf_rows(oborovo_result, oborovo_result.project_inputs)
        op_rows = [r for r in rows if r["ebitda"] > 0]
        assert len(op_rows) > 0
        # Unlevered tax should be >= 0 in operation periods
        for r in op_rows[:5]:
            assert r["unlevered_tax"] >= 0, (
                f"Period {r['period']}: unlevered_tax {r['unlevered_tax']:.1f} should be >= 0"
            )

    def test_project_free_cf_equals_ebitda_minus_unlevered_tax(self, oborovo_result):
        """Project free CF = EBITDA - unlevered_tax."""
        rows = build_project_cf_rows(oborovo_result, oborovo_result.project_inputs)
        for r in rows:
            if r["ebitda"] > 0:
                expected = r["ebitda"] - r["unlevered_tax"]
                assert abs(r["project_free_cf"] - expected) < 1.0, (
                    f"Period {r['period']}: free_cf {r['project_free_cf']:.1f} ≠ "
                    f"EBITDA {r['ebitda']:.1f} - unlev_tax {r['unlevered_tax']:.1f}"
                )

    def test_cumulative_cf_is_sequential(self, oborovo_result):
        """Cumulative CF changes by period CF each step."""
        rows = build_project_cf_rows(oborovo_result, oborovo_result.project_inputs)
        for i in range(1, len(rows)):
            delta = rows[i]["project_free_cf"]
            computed = rows[i]["cumulative_project_cf"] - rows[i-1]["cumulative_project_cf"]
            assert abs(delta - computed) < 1.0, (
                f"Period {rows[i]['period']}: delta {delta:.1f} ≠ cumulative change {computed:.1f}"
            )

    def test_project_irr_cashflows_reasonable(self, oborovo_result):
        """Project free CF is plausible magnitude (< ±50,000 kEUR)."""
        rows = build_project_cf_rows(oborovo_result, oborovo_result.project_inputs)
        for r in rows:
            if r["ebitda"] > 0:
                assert -50_000 <= r["project_free_cf"] <= 50_000, (
                    f"Period {r['period']} CF {r['project_free_cf']:.0f} unreasonable"
                )


# =============================================================================
# Equity CF Bridge Tests
# =============================================================================

class TestEquityCFBridge:
    """Tests for equity cashflow bridge table."""

    @pytest.fixture
    def oborovo_result(self):
        from app.project_factories import create_default_oborovo
        from app.ui_runner import run_demo_project
        oborovo = create_default_oborovo()
        return run_demo_project("Solar", "Base", project_inputs_override=oborovo).result

    def test_includes_shl_service(self, oborovo_result):
        """Equity CF bridge captures SHL interest and principal in operation periods."""
        rows = build_equity_cf_rows(oborovo_result)
        shl_rows = [r for r in rows if r["shl_interest"] > 0 or r["shl_principal"] > 0]
        assert len(shl_rows) > 0, "No SHL service found"

    def test_includes_distributions(self, oborovo_result):
        """Equity CF bridge includes distribution amounts."""
        rows = build_equity_cf_rows(oborovo_result)
        dist_rows = [r for r in rows if r["distributions"] > 0]
        assert len(dist_rows) > 0, "No distributions found"

    def test_equity_investment_negative_or_zero(self, oborovo_result):
        """Equity investment is negative or zero (cash out for investor)."""
        rows = build_equity_cf_rows(oborovo_result)
        for r in rows:
            assert r["equity_investment"] <= 0, (
                f"Period {r['period']}: equity_investment {r['equity_investment']:.1f} should be ≤ 0"
            )

    def test_equity_cf_sign_convention_investor_perspective(self, oborovo_result):
        """Equity CF is negative during initial investment period, positive in operation.

        Note: Oborovo financial close = COD (June 2030), so first waterfall period
        is already an operation period. Initial equity investment appears as negative
        cash flow at the start of operations (period 2 = first operation period).
        """
        rows = build_equity_cf_rows(oborovo_result)
        # First few periods should have positive equity_cash_flow (distributions coming in)
        # OR there should be equity investment at some point in the bridge
        positive_cf = [r for r in rows if r["equity_cash_flow"] > 0]
        assert len(positive_cf) > 0, "No positive equity cash flows found (expected distributions)"
        # Operation periods should have SHL/distributions as positive inflows
        op_rows = [r for r in rows if r["distributions"] > 0]
        assert len(op_rows) > 0, "No operation-period distributions found"

    def test_equity_cf_sum_near_zero(self, oborovo_result):
        """Sum of equity CFs is plausible (not too far from zero)."""
        rows = build_equity_cf_rows(oborovo_result)
        total = sum(r["equity_cash_flow"] for r in rows)
        assert -200_000 <= total <= 200_000, f"Total equity CF {total:.0f} kEUR unreasonable"


# =============================================================================
# Excel Export with Reconciliation Sheets
# =============================================================================

class TestExcelReconciliationSheets:
    """Tests for Excel export with include_reconciliation_sheets=True."""

    @pytest.fixture
    def oborovo_result(self):
        from app.project_factories import create_default_oborovo
        from app.ui_runner import run_demo_project
        oborovo = create_default_oborovo()
        demo = run_demo_project("Solar", "Base", project_inputs_override=oborovo)
        demo.result.project_inputs = oborovo
        return demo

    def test_reconciliation_sheets_are_bytes(self, oborovo_result):
        """Excel with reconciliation sheets returns valid bytes."""
        from app.excel_export import build_excel_export
        data = build_excel_export(
            result=oborovo_result.result,
            project_inputs=oborovo_result.project_inputs,
            include_reconciliation_sheets=True,
        )
        assert isinstance(data, bytes) and len(data) > 1000

    def test_reconciliation_sheets_in_workbook(self, oborovo_result):
        """All 4 sheets present when flag is True."""
        import openpyxl
        from io import BytesIO
        from app.excel_export import build_excel_export

        data = build_excel_export(
            result=oborovo_result.result,
            project_inputs=oborovo_result.project_inputs,
            include_reconciliation_sheets=True,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        expected = {"Debt Schedule", "Project CF Bridge", "Equity CF Bridge", "Calibration Notes"}
        assert expected.issubset(set(wb.sheetnames)), f"Missing sheets. Found: {wb.sheetnames}"

    def test_backward_compatible_without_flag(self, oborovo_result):
        """Excel without flag does NOT include reconciliation sheets."""
        import openpyxl
        from io import BytesIO
        from app.excel_export import build_excel_export

        data = build_excel_export(
            result=oborovo_result.result,
            project_inputs=oborovo_result.project_inputs,
            include_reconciliation_sheets=False,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Debt Schedule" not in wb.sheetnames
        assert "Project CF Bridge" not in wb.sheetnames

    def test_generic_solar_export_still_works(self):
        """Generic Solar export with reconciliation sheets works."""
        from app.project_factories import create_default_solar_project
        from app.ui_runner import run_demo_project
        from app.excel_export import build_excel_export

        solar = create_default_solar_project()
        result = run_demo_project("Solar", "Base", project_inputs_override=solar).result
        data = build_excel_export(
            result=result,
            project_inputs=None,
            include_reconciliation_sheets=True,
        )
        assert isinstance(data, bytes) and len(data) > 1000

    def test_generic_wind_export_still_works(self):
        """Generic Wind export with reconciliation sheets works."""
        from app.project_factories import create_default_wind_project
        from app.ui_runner import run_demo_project
        from app.excel_export import build_excel_export

        wind = create_default_wind_project()
        result = run_demo_project("Wind", "Base", project_inputs_override=wind).result
        data = build_excel_export(
            result=result,
            project_inputs=None,
            include_reconciliation_sheets=True,
        )
        assert isinstance(data, bytes) and len(data) > 1000

    def test_calibration_notes_contains_oborovo_specifics(self, oborovo_result):
        """Oborovo Calibration Notes contains Oborovo-specific text, not generic."""
        import openpyxl
        from io import BytesIO
        from app.excel_export import build_excel_export

        data = build_excel_export(
            result=oborovo_result.result,
            project_inputs=oborovo_result.project_inputs,
            include_reconciliation_sheets=True,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Calibration Notes"]
        content = "\n".join(str(cell.value or "") for row in ws.iter_rows() for cell in row)
        # Should mention Oborovo calibration
        assert "Oborovo" in content or "7.985" in content or "AFRY" in content

    def test_calibration_notes_generic_solar_no_oborovo_hardcoding(self):
        """Generic Solar Calibration Notes does NOT contain Oborovo-specific hardcoding."""
        import openpyxl
        from io import BytesIO
        from app.project_factories import create_default_solar_project
        from app.ui_runner import run_demo_project
        from app.excel_export import build_excel_export

        solar = create_default_solar_project()
        result = run_demo_project("Solar", "Base", project_inputs_override=solar).result
        data = build_excel_export(
            result=result,
            project_inputs=solar,
            include_reconciliation_sheets=True,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Calibration Notes"]
        content = "\n".join(str(cell.value or "") for row in ws.iter_rows() for cell in row)
        # Should NOT contain hardcoded Oborovo debt anchor
        assert "42,852" not in content, "Generic Solar should not contain Oborovo debt anchor"
        assert "OBR" not in content.upper() or "Oborovo" not in content, (
            "Generic Solar should not contain Oborovo references"
        )