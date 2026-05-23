"""Tests for Phase 12 governance semantics cleanup."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
SEMANTICS = REPORTS / "phase12_governance_semantics_dictionary.csv"
USAGE = REPORTS / "phase12_governance_label_usage_matrix.csv"
HIERARCHY = REPORTS / "phase12_governance_escalation_hierarchy.csv"
SPLIT = REPORTS / "phase12_missing_evidence_split_matrix.csv"
DOC = DOCS / "phase12_governance_semantics_cleanup.md"


def _generate() -> None:
    write_calibration_reconciliation_pack()


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def test_governance_label_inventory_exists():
    _generate()
    rows = _csv_rows(USAGE)
    assert rows
    labels = {row["label"] for row in rows}
    required = {
        "PASS",
        "WARN",
        "FAIL",
        "ACCEPTED_CONVENTION",
        "MISSING_EVIDENCE",
        "MISSING_EXCEL_EVIDENCE",
        "MISSING_REVIEW_SCALAR",
        "SOURCE_NOT_AVAILABLE",
        "EVIDENCE_LIMITATION",
        "GROUPED_SOURCE_ONLY",
        "GOVERNANCE_BLOCKER",
        "GOVERNANCE_REVIEW",
        "ENGINEERING_FOLLOWUP",
        "STAKEHOLDER_DECISION",
        "RUNTIME_BINDING_PENDING",
        "BLOCKED",
        "NOT_APPROVED",
    }
    assert required.issubset(labels)


def test_semantics_dictionary_exists_and_covers_split_labels():
    _generate()
    rows = _csv_rows(SEMANTICS)
    assert rows
    labels = {row["label"] for row in rows}
    assert "MISSING_EXCEL_EVIDENCE" in labels
    assert "MISSING_REVIEW_SCALAR" in labels
    assert "SOURCE_NOT_AVAILABLE" in labels
    legacy = next(row for row in rows if row["label"] == "MISSING_EVIDENCE")
    assert "legacy" in legacy["escalation_level"].lower()


def test_escalation_hierarchy_exists():
    _generate()
    rows = _csv_rows(HIERARCHY)
    assert rows
    labels = {row["label"] for row in rows}
    assert {"GOVERNANCE_BLOCKER", "BLOCKED", "NOT_APPROVED"}.issubset(labels)


def test_overloaded_missing_evidence_usage_is_resolved_and_documented():
    _generate()
    rows = _csv_rows(SPLIT)
    replacements = {row["replacement_label"] for row in rows}
    assert {"MISSING_EXCEL_EVIDENCE", "MISSING_REVIEW_SCALAR", "SOURCE_NOT_AVAILABLE", "RUNTIME_BINDING_PENDING"}.issubset(replacements)


def test_workbook_report_semantics_are_consistent():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    navigation = _sheet_text(wb, "Navigation")
    governance = _sheet_text(wb, "Governance")
    notes = _sheet_text(wb, "Reviewer Notes")
    assert "MISSING_EXCEL_EVIDENCE" in navigation
    assert "MISSING_REVIEW_SCALAR" in navigation
    assert "SOURCE_NOT_AVAILABLE" in navigation
    assert "Accepted conventions" in governance
    assert "SOURCE_NOT_AVAILABLE" in notes


def test_accepted_conventions_remain_non_authoritative():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance = _sheet_text(wb, "Governance")
    navigation = _sheet_text(wb, "Navigation")
    assert "not runtime fixes" in governance.lower()
    assert "documentation, not runtime fixes" in navigation.lower()


def test_runtime_vs_governance_distinction_preserved():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance = _sheet_text(wb, "Governance")
    runtime = _sheet_text(wb, "Runtime Summary")
    assert "G20 status BLOCKED" in governance
    assert "R99/R102 status NOT APPROVED" in governance
    assert "Runtime / Preview" in runtime


def test_doc_states_scope_and_philosophy():
    text = DOC.read_text(encoding="utf-8")
    assert "governance semantics philosophy" in text.lower()
    assert "missing-evidence split rationale" in text.lower()
    assert "accepted convention clarification" in text.lower()
    assert "G20" in text and "BLOCKED" in text
    assert "R99/R102" in text and "NOT APPROVED" in text
    assert "No runtime/model formulas are changed" in text


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
