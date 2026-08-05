"""Stage C3B2 — Oborovo Debt Sizing & Interest Source Closure.

Validates the five open questions from C3B1 against the pre-populated fixture
and confirms equal-input/equal-policy verdict classification.

All tests are CI-portable: no actual workbook binary required.
"""
from __future__ import annotations

import json
import pathlib

import pytest

# ---------------------------------------------------------------------------
# Fixture loading
# ---------------------------------------------------------------------------

FIXTURE_PATH = (
    pathlib.Path(__file__).parent / "fixtures" / "excel_oborovo_debt_interest_truth.json"
)


@pytest.fixture(scope="module")
def truth() -> dict:
    with FIXTURE_PATH.open(encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# Extractor version
# ---------------------------------------------------------------------------

class TestExtractorVersion:
    def test_extractor_version_importable(self):
        from finco_recon.extract_oborovo_debt_interest import _EXTRACTOR_VERSION
        assert isinstance(_EXTRACTOR_VERSION, str)
        assert _EXTRACTOR_VERSION == "1.0.0"

    def test_fixture_meta_extractor_version(self, truth):
        assert truth["_meta"]["extractor_version"] == "1.0.0"

    def test_fixture_meta_has_source_filename(self, truth):
        assert "source_filename" in truth["_meta"]
        assert "20260414_BP_Oborovo_Sensitivity_FINAL_for_PPT" in truth["_meta"]["source_filename"]


# ---------------------------------------------------------------------------
# Workstream A — CFADS composition
# ---------------------------------------------------------------------------

class TestWorkstreamA_CFADS:
    def test_section_present(self, truth):
        assert "workstream_a" in truth

    def test_workstream_label(self, truth):
        assert truth["workstream_a"]["workstream"] == "A"

    def test_ds_row20_references_macro(self, truth):
        """DS!row20 formula must reference Macro row (CFADS source)."""
        formula = truth["workstream_a"]["ds_row20"]["formula_h"]
        assert formula is not None
        assert "Macro" in formula or "CF" in formula or "macro" in formula.lower()

    def test_cfads_period1_value(self, truth):
        """DS!H20 period 1 CFADS ≈ 2575.003 kEUR (from workbook exploration)."""
        values = truth["workstream_a"]["ds_row20"]["period_values_keur"]
        p1 = values[1]
        assert p1 is not None
        assert abs(p1 - 2575.003) < 1.0, f"Expected ~2575 kEUR, got {p1}"

    def test_cf_row79_label(self, truth):
        label = truth["workstream_a"]["cf_row79_free_cash_flow_for_banks"]["label"]
        assert label is not None
        assert "Cash Flow" in label or "cash flow" in label.lower() or "FCF" in label

    def test_cf_row79_period1_matches_ds_row20(self, truth):
        """CF!row79 and DS!row20 must carry the same period-1 value."""
        ds_p1 = truth["workstream_a"]["ds_row20"]["period_values_keur"][1]
        cf_p1 = truth["workstream_a"]["cf_row79_free_cash_flow_for_banks"]["period_values_keur"][1]
        assert ds_p1 is not None and cf_p1 is not None
        assert abs(ds_p1 - cf_p1) < 0.01

    def test_classification_is_input_policy_mismatch(self, truth):
        assert truth["workstream_a"]["classification"] == "INPUT_POLICY_MISMATCH"

    def test_finding_mentions_post_tax(self, truth):
        finding = truth["workstream_a"]["finding"]
        assert "POST-TAX" in finding or "post-tax" in finding.lower()

    def test_phase2c_cfads_formula_documented(self, truth):
        formula = truth["workstream_a"]["phase2c_cfads_formula"]
        assert "EBITDA" in formula or "ebitda" in formula.lower()
        assert "cash_tax" in formula

    def test_dscr_target_period1_is_1_15(self, truth):
        """Base DSCR target in the first 24 operating periods must be 1.15."""
        vals = truth["workstream_a"]["ds_row22_dscr_target"]["period_values"]
        p1 = vals[1]
        assert p1 is not None
        assert abs(p1 - 1.15) < 0.001


# ---------------------------------------------------------------------------
# Workstream B — Sculpting algorithm
# ---------------------------------------------------------------------------

class TestWorkstreamB_Sculpting:
    def test_section_present(self, truth):
        assert "workstream_b" in truth

    def test_workstream_label(self, truth):
        assert truth["workstream_b"]["workstream"] == "B"

    def test_ds_row47_formula_references_next_period(self, truth):
        """Backward induction formula references I47 (next period's value)."""
        formula = truth["workstream_b"]["ds_row47_capacity"]["formula_h"]
        assert formula is not None
        assert "I47" in formula or "i47" in formula.lower()

    def test_ds_d51_value(self, truth):
        """DS!D51 total debt ≈ 42852.279 kEUR."""
        val = truth["workstream_b"]["ds_d51_total_debt"]["value_keur"]
        assert abs(val - 42852.279) < 1.0

    def test_convergence_mechanism_documented(self, truth):
        mech = truth["workstream_b"]["convergence_mechanism"]
        assert "EXCEL" in mech or "ITERATIVE" in mech

    def test_phase2c_mechanism_documented(self, truth):
        mech = truth["workstream_b"]["phase2c_mechanism"]
        assert "FORWARD" in mech or "SCULPT" in mech

    def test_algorithm_equivalence_noted(self, truth):
        eq = truth["workstream_b"]["algorithm_equivalence"]
        assert "EQUIVALENT" in eq or "equivalent" in eq.lower()

    def test_ds_b22_base_dscr(self, truth):
        val = truth["workstream_b"]["ds_b22_base_dscr"]["value"]
        assert val is not None
        assert abs(val - 1.15) < 0.01


# ---------------------------------------------------------------------------
# Workstream C — DSRA
# ---------------------------------------------------------------------------

class TestWorkstreamC_DSRA:
    def test_section_present(self, truth):
        assert "workstream_c" in truth

    def test_workstream_label(self, truth):
        assert truth["workstream_c"]["workstream"] == "C"

    def test_dsra_target_is_zero(self, truth):
        val = truth["workstream_c"]["inputs_i348_dsra_target"]["value"]
        assert val == 0 or val is None or (isinstance(val, float) and val == 0.0)

    def test_dsra_not_present(self, truth):
        assert truth["workstream_c"]["dsra_present"] is False

    def test_dsra_aligned_both_zero(self, truth):
        assert truth["workstream_c"]["classification"] == "ALIGNED_BOTH_ZERO"

    def test_dsra_funding_row_all_zero(self, truth):
        rows = truth["workstream_c"]["cf_dsra_rows"]
        if "cf_row87" in rows:
            vals = [v for v in rows["cf_row87"]["period_values_keur"] if v is not None]
            assert all(v == 0.0 for v in vals), "DSRA funding should be zero in all periods"


# ---------------------------------------------------------------------------
# Workstream D — Sizing base (gearing base)
# ---------------------------------------------------------------------------

class TestWorkstreamD_SizingBase:
    def test_section_present(self, truth):
        assert "workstream_d" in truth

    def test_workstream_label(self, truth):
        assert truth["workstream_d"]["workstream"] == "D"

    def test_g171_total_eligible_cost(self, truth):
        val = truth["workstream_d"]["inputs_g171_total_eligible_cost"]["value_keur"]
        assert abs(val - 57973.053) < 1.0

    def test_idc_is_included(self, truth):
        assert truth["workstream_d"]["idc_included_in_gearing_base"] is True

    def test_idc_value_approximately_correct(self, truth):
        components = truth["workstream_d"]["components"]
        idc_row = components.get("inputs_row166_g", {})
        val = idc_row.get("value_keur")
        assert val is not None
        assert abs(val - 1086.032) < 5.0, f"IDC expected ~1086 kEUR, got {val}"

    def test_gearing_cap_not_binding(self, truth):
        assert truth["workstream_d"]["gearing_cap_binding"] is False

    def test_gearing_cap_value(self, truth):
        cap = truth["workstream_d"]["gearing_cap_keur"]
        assert cap is not None
        assert cap > 42852.279, "Gearing cap must exceed the DSCR-sculpted debt"
        assert abs(cap - 46378.442) < 5.0

    def test_phase2c_mapping_documented(self, truth):
        mapping = truth["workstream_d"]["phase2c_mapping"]
        assert "G171" in mapping or "g171" in mapping.lower()
        assert "eligible_project_cost" in mapping or "eligible" in mapping.lower()

    def test_capex_component_present(self, truth):
        components = truth["workstream_d"]["components"]
        capex_row = components.get("inputs_row165_g", {})
        val = capex_row.get("value_keur")
        assert val is not None
        assert val > 50000, "Hard CAPEX should dominate the sizing base"


# ---------------------------------------------------------------------------
# Workstream E — Interest rate
# ---------------------------------------------------------------------------

class TestWorkstreamE_InterestRate:
    def test_section_present(self, truth):
        assert "workstream_e" in truth

    def test_workstream_label(self, truth):
        assert truth["workstream_e"]["workstream"] == "E"

    def test_fixed_fraction_is_080(self, truth):
        val = truth["workstream_e"]["ds_b40_fixed_fraction"]["value"]
        assert val is not None
        assert abs(val - 0.80) < 0.001

    def test_float_fraction_is_020(self, truth):
        val = truth["workstream_e"]["ds_b39_float_fraction"]["value"]
        assert val is not None
        assert abs(val - 0.20) < 0.001

    def test_swap_rate(self, truth):
        val = truth["workstream_e"]["ds_c40_swap_rate"]["value"]
        assert val is not None
        assert abs(val - 0.032) < 0.001

    def test_sculpting_rate_pct_documented(self, truth):
        rate = truth["workstream_e"]["sculpting_rate_pct"]
        assert rate is not None
        assert abs(rate - 5.95) < 0.1, f"Sculpting rate expected ~5.95%, got {rate}"

    def test_phase2c_rate_pct(self, truth):
        rate = truth["workstream_e"]["phase2c_rate_pct"]
        assert abs(rate - 5.65) < 0.01

    def test_rate_mismatch_documented(self, truth):
        bps = truth["workstream_e"]["rate_mismatch_basis_points"]
        assert bps is not None
        assert bps > 0, "Rate mismatch must be positive (Excel > Phase 2C)"

    def test_tranche_interest_formula_uses_sculpting_rate(self, truth):
        formula = truth["workstream_e"]["ds_row64_period_interest"]["formula_h"]
        assert formula is not None
        assert "H44" in formula, "Tranche interest must reference the sculpting rate H44"

    def test_tranche_interest_period1_value(self, truth):
        vals = truth["workstream_e"]["ds_row64_period_interest"]["period_values_keur"]
        p1 = vals[1]
        assert p1 is not None
        assert abs(p1 - 1303.483) < 5.0, f"Expected ~1303 kEUR interest, got {p1}"

    def test_d280_is_fcf_section_only(self, truth):
        note = truth["workstream_e"]["inputs_d280_fcf_rate"]["note"]
        assert "B33" in note or "FCF" in note, "D280 usage note must mention B33 / FCF section"
        assert "NOT" in note or "not" in note.lower()

    def test_inputs_d230_hedge_fraction(self, truth):
        val = truth["workstream_e"]["inputs_d230_hedge_fraction"]["value"]
        assert val is not None
        assert abs(val - 0.80) < 0.001

    def test_classification_is_mismatch(self, truth):
        assert truth["workstream_e"]["classification"] == "INPUT_POLICY_MISMATCH"

    def test_finding_mentions_both_rates(self, truth):
        finding = truth["workstream_e"]["finding"]
        assert "5.65" in finding
        assert ("5.95" in finding or "5.9" in finding)


# ---------------------------------------------------------------------------
# Equal-input / equal-policy verdict
# ---------------------------------------------------------------------------

class TestEqualInputEqualPolicy:
    def test_section_present(self, truth):
        assert "equal_input_equal_policy" in truth

    def test_verdict_is_valid_classification(self, truth):
        verdict = truth["equal_input_equal_policy"]["verdict"]
        valid = {
            "C3B2_EQUAL_INPUT_EQUAL_POLICY_MATCH",
            "C3B2_INPUT_OR_POLICY_MISMATCH_FULLY_EXPLAINED",
            "C3B2_EQUAL_INPUT_EQUAL_POLICY_DIVERGENCE_PROVED",
            "C3B2_SOURCE_TRUTH_PARTIAL_MANUAL_CHECK_REQUIRED",
        }
        assert verdict in valid, f"Unrecognised verdict: {verdict}"

    def test_verdict_is_b_mismatch_explained(self, truth):
        """Based on confirmed exploration: verdict must be MISMATCH_FULLY_EXPLAINED."""
        verdict = truth["equal_input_equal_policy"]["verdict"]
        assert verdict == "C3B2_INPUT_OR_POLICY_MISMATCH_FULLY_EXPLAINED"

    def test_mismatch_root_causes_documented(self, truth):
        causes = truth["equal_input_equal_policy"]["mismatch_root_causes"]
        assert len(causes) >= 2, "At least interest rate and CFADS mismatches must be documented"

    def test_interest_rate_cause_present(self, truth):
        causes = truth["equal_input_equal_policy"]["mismatch_root_causes"]
        cause_ids = [c["cause"] for c in causes]
        assert "INTEREST_RATE_MISMATCH" in cause_ids

    def test_cfads_cause_present(self, truth):
        causes = truth["equal_input_equal_policy"]["mismatch_root_causes"]
        cause_ids = [c["cause"] for c in causes]
        assert "CFADS_COMPOSITION_MISMATCH" in cause_ids

    def test_excel_debt_value(self, truth):
        val = truth["equal_input_equal_policy"]["excel_total_debt_keur"]
        assert abs(val - 42852.279) < 1.0

    def test_verdict_rationale_mentions_dsra_aligned(self, truth):
        rationale = truth["equal_input_equal_policy"]["verdict_rationale"]
        assert "DSRA" in rationale or "dsra" in rationale.lower()

    def test_verdict_rationale_mentions_gearing_not_binding(self, truth):
        rationale = truth["equal_input_equal_policy"]["verdict_rationale"]
        assert (
            "not binding" in rationale.lower()
            or "gearing" in rationale.lower()
        )


# ---------------------------------------------------------------------------
# Synthetic extractor smoke-test (CI-portable, no workbook needed)
# ---------------------------------------------------------------------------

class TestExtractorSynthetic:
    """Run extractor functions against a minimal in-memory openpyxl workbook."""

    def _make_minimal_workbook(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for name in ["DS", "CF", "Inputs"]:
            ws = wb.create_sheet(name)
            # Row 1 header placeholder
            ws.cell(row=1, column=1, value=name + "_placeholder")
        return wb

    def test_extractor_version_constant(self):
        from finco_recon.extract_oborovo_debt_interest import _EXTRACTOR_VERSION
        assert _EXTRACTOR_VERSION == "1.0.0"

    def test_row_to_periods_helper(self):
        from finco_recon.extract_oborovo_debt_interest import _row_to_periods
        row = tuple([None] * 6 + [1.0, 2.0, 3.0] + [None] * 52)
        result = _row_to_periods(row, n=61)
        assert result[0] == 1.0
        assert result[1] == 2.0
        assert result[2] == 3.0
        assert len(result) == 61

    def test_scalar_helper_in_bounds(self):
        from finco_recon.extract_oborovo_debt_interest import _scalar
        row = (10, 20, 30)
        assert _scalar(row, 0) == 10
        assert _scalar(row, 2) == 30

    def test_scalar_helper_out_of_bounds(self):
        from finco_recon.extract_oborovo_debt_interest import _scalar
        row = (10,)
        assert _scalar(row, 99) is None

    def test_formula_helper_detects_formula(self):
        from finco_recon.extract_oborovo_debt_interest import _formula
        row = (None, "=A1+B1", "literal")
        assert _formula(row, 1) == "=A1+B1"
        assert _formula(row, 2) is None

    def test_main_entry_point_importable(self):
        from finco_recon.extract_oborovo_debt_interest import main
        assert callable(main)

    def test_main_returns_1_on_missing_workbook(self, tmp_path):
        from finco_recon.extract_oborovo_debt_interest import main
        out = tmp_path / "out.json"
        code = main([
            "--workbook", str(tmp_path / "nonexistent.xlsm"),
            "--output", str(out),
        ])
        assert code == 1
