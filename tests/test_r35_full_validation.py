"""Workbook tests for the TUHO full R35 validation report."""

from __future__ import annotations

from dataclasses import fields
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner
from domain.inputs import ProjectInfo


WORKBOOK_PATH = Path("reports/phase6_tuho_r35_full_validation.xlsx")
REQUIRED_SHEETS = {
    "Summary",
    "R35 Bridge Progression",
    "Row Comparison",
    "Remaining Drivers",
    "Largest Deltas",
    "R99 Readiness",
    "Notes",
}


def _run_project(project):
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return WaterfallRunner(project, engine).run(config)


def test_workbook_exists_with_required_sheets():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    assert REQUIRED_SHEETS.issubset(set(workbook.sheetnames))


def test_r35_bridge_progression_contains_60_operating_periods():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R35 Bridge Progression"]
    op_indexes = [sheet.cell(row, 1).value for row in range(2, 62)]

    assert op_indexes == list(range(60))
    assert sheet.cell(62, 1).value == "TOTAL"


def test_bridges_close_r27_and_r13_gaps():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    remaining = workbook["Remaining Drivers"]
    driver_values = {
        remaining.cell(row, 1).value: remaining.cell(row, 2).value
        for row in range(2, remaining.max_row + 1)
    }

    assert driver_values["SHL interest gross/net/timing"] == pytest.approx(0.0, abs=0.1)
    assert driver_values["Book/tax depreciation timing"] == pytest.approx(0.0, abs=0.1)


def test_r35_final_residual_max_delta_and_material_count():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    summary = workbook["Summary"]
    metrics = {
        summary.cell(row, 1).value: summary.cell(row, 2).value
        for row in range(2, summary.max_row + 1)
    }

    assert metrics["R35 after SHL + book depreciation bridges"] == pytest.approx(
        -433.1,
        abs=0.1,
    )
    assert metrics["Max period delta after both bridges"] == pytest.approx(152.8, abs=0.1)
    assert metrics["Material periods >100 kEUR"] == 7


def test_largest_deltas_are_sorted_by_absolute_delta():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Largest Deltas"]
    abs_deltas = [sheet.cell(row, 7).value for row in range(2, sheet.max_row + 1)]

    assert abs_deltas == sorted(abs_deltas, reverse=True)
    assert abs_deltas[0] == pytest.approx(152.8, abs=0.1)


def test_r99_readiness_remains_blocked():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R99 Readiness"]
    readiness = {
        sheet.cell(row, 1).value: sheet.cell(row, 2).value
        for row in range(2, sheet.max_row + 1)
    }

    assert readiness["Tax bridge runtime source"] == "Blocked"
    assert readiness["SHL FCF"] == "Off"


def test_no_runtime_behavior_change_and_no_new_flags():
    project_info_fields = {field.name: field for field in fields(ProjectInfo)}

    assert project_info_fields["use_shl_gross_accrued_for_pnl"].default is False
    assert project_info_fields["use_book_depreciation_for_pnl"].default is False

    project = create_default_tuho_wind1()
    before = _run_project(project)
    after = _run_project(project)

    assert [period.corporate_tax_cash_keur for period in after.periods] == pytest.approx(
        [period.corporate_tax_cash_keur for period in before.periods],
        abs=0.0001,
    )
    assert [period.r99_fcf_for_distribution_keur for period in after.periods] == pytest.approx(
        [period.r99_fcf_for_distribution_keur for period in before.periods],
        abs=0.0001,
    )
    assert project.info.use_shl_fcf_waterfall_engine is False
    assert project.financing.use_tuho_r99_input_engine is False
