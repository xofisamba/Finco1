"""Product Gap PR11 -- Export Reality Check tests.

Scope (see docs/PRODUCT_GAP_PR11_EXPORT_REALITY_CHECK.md):
- Export still works (institutional workbook + runtime summary CSV +
  values-only Excel export all generate successfully).
- Workbook still opens; sheet order/count unchanged from the existing
  U5 pin (no sheets removed/added by this PR).
- No worksheets were removed -- this investigation found the export
  pipeline already sources every worksheet from real persisted/
  Run-backed data (unlike PR1-PR9, which found and removed real
  placeholder content elsewhere in the app).
- All genuine worksheets remain exported.
- No fake financial data is exported -- exported Runtime Summary
  values match a direct re-run of the same project.
- No banned internal wording appears in exported cell values.
- Workbook generation unchanged for genuine sheets.
- Guardrail files untouched.
- The orphaned calibration_reconciliation export pipeline is confirmed
  not wired into any live route.
"""
from __future__ import annotations

import io
import os
import subprocess

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
os.environ.setdefault("FINCO_COOKIE_SECURE", "false")

import openpyxl
import pytest

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

BANNED_TERMS = [
    "preview architecture",
    "runtime pipeline",
    "stub",
    "prototype",
    "todo:",
    "fixme",
    "placeholder architecture",
]

EXPECTED_INSTITUTIONAL_SHEET_ORDER = (
    "Export_Metadata",
    "Workbook_Index",
    "Cover",
    "Governance",
    "Runtime Summary",
    "Inputs",
    "Construction",
    "OPEX",
    "CAPEX",
    "Revenue",
    "Senior Debt",
    "SHL",
    "Tax",
    "P&L",
    "Cash Flow",
    "Balance Sheet",
    "Audit",
    "Gap Register",
    "Validation Status",
)

GENUINE_RUNTIME_SHEETS = (
    "Inputs",
    "Construction",
    "OPEX",
    "CAPEX",
    "Revenue",
    "Senior Debt",
    "SHL",
    "Tax",
    "P&L",
    "Cash Flow",
    "Balance Sheet",
    "Runtime Summary",
)


def _all_cell_text(wb) -> str:
    return " ".join(
        str(c.value)
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for c in row
        if c.value
    )


@pytest.fixture(scope="module")
def institutional_workbook_bytes():
    from app.export.institutional_workbook import export_institutional_workbook_skeleton

    return export_institutional_workbook_skeleton("tuho")


@pytest.fixture(scope="module")
def institutional_workbook(institutional_workbook_bytes):
    return openpyxl.load_workbook(io.BytesIO(institutional_workbook_bytes))


@pytest.fixture(scope="module")
def runtime_summary_csv_text():
    from app.export.runtime_summary import build_runtime_summary_csv

    return build_runtime_summary_csv("tuho")


@pytest.fixture(scope="module")
def values_only_workbook():
    from app.excel_export import build_excel_export
    from app.project_factories import create_default_solar_project
    from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner
    from domain.period_engine import PeriodEngine

    project_inputs = create_default_solar_project()
    engine = PeriodEngine(
        financial_close=project_inputs.info.financial_close,
        construction_months=project_inputs.info.construction_months,
        horizon_years=project_inputs.info.horizon_years,
        ppa_years=project_inputs.revenue.ppa_term_years,
    )
    config = WaterfallRunConfig.from_inputs(project_inputs, engine)
    result = WaterfallRunner(project_inputs, engine).run(config)
    data = build_excel_export(result=result, project_inputs=project_inputs, provenance_metadata={})
    return openpyxl.load_workbook(io.BytesIO(data))


# ─── 1: Export still works ──────────────────────────────────────────────────

class TestExportStillWorks:
    def test_institutional_workbook_generates(self, institutional_workbook_bytes):
        assert isinstance(institutional_workbook_bytes, (bytes, bytearray))
        assert len(institutional_workbook_bytes) > 0

    def test_runtime_summary_csv_generates(self, runtime_summary_csv_text):
        assert isinstance(runtime_summary_csv_text, str)
        assert "project_irr" in runtime_summary_csv_text

    def test_values_only_excel_generates(self, values_only_workbook):
        assert len(values_only_workbook.sheetnames) > 0


# ─── 2: Workbook still opens / sheet order unchanged ────────────────────────

class TestWorkbookStillOpensSheetOrderUnchanged:
    def test_institutional_workbook_opens(self, institutional_workbook):
        assert institutional_workbook is not None

    def test_sheet_order_matches_pinned_sequence(self, institutional_workbook):
        assert tuple(institutional_workbook.sheetnames) == EXPECTED_INSTITUTIONAL_SHEET_ORDER

    def test_values_only_workbook_opens(self, values_only_workbook):
        assert values_only_workbook is not None
        assert "Dashboard" in values_only_workbook.sheetnames


# ─── 3: No worksheets were removed by this PR ──────────────────────────────

class TestNoWorksheetsRemoved:
    """This PR's investigation found no fake/placeholder worksheet to
    remove -- every sheet already sources from real data. This locks
    in that the full pre-existing sheet set is still present."""

    def test_all_pinned_sheets_present(self, institutional_workbook):
        for name in EXPECTED_INSTITUTIONAL_SHEET_ORDER:
            assert name in institutional_workbook.sheetnames

    def test_sheet_count_unchanged(self, institutional_workbook):
        assert len(institutional_workbook.sheetnames) == 19


# ─── 4: All genuine worksheets remain exported ─────────────────────────────

class TestGenuineWorksheetsExported:
    @pytest.mark.parametrize("sheet_name", GENUINE_RUNTIME_SHEETS)
    def test_genuine_sheet_present_and_non_empty(self, institutional_workbook, sheet_name):
        sheet = institutional_workbook[sheet_name]
        # Sheet must have more than just the 5-row metadata header block.
        assert sheet.max_row > 6

    def test_pnl_sheet_sources_from_assembled_statements(self, institutional_workbook):
        sheet = institutional_workbook["P&L"]
        text = " ".join(str(c.value) for row in sheet.iter_rows() for c in row if c.value)
        assert "Offline assembled P&L" in text

    def test_runtime_summary_sheet_has_real_metrics(self, institutional_workbook):
        sheet = institutional_workbook["Runtime Summary"]
        text = " ".join(str(c.value) for row in sheet.iter_rows() for c in row if c.value)
        assert "Project IRR" in text
        assert "Average DSCR" in text


# ─── 5: No fake financial data is exported ─────────────────────────────────

class TestNoFakeFinancialData:
    def test_runtime_summary_values_match_direct_run(self, institutional_workbook):
        from app.export.runtime_summary import _run_project

        _, result = _run_project("tuho")
        sheet = institutional_workbook["Runtime Summary"]
        rows = {}
        for row in sheet.iter_rows(min_row=7):
            label = row[0].value
            value = row[1].value
            if label:
                rows[label] = value

        assert rows["Project IRR"] == pytest.approx(result.project_irr, rel=1e-6)
        assert rows["Equity IRR"] == pytest.approx(result.equity_irr, rel=1e-6)
        assert rows["Average DSCR"] == pytest.approx(result.actual_avg_dscr, rel=1e-6)

    def test_pnl_revenue_matches_assembled_statements(self, institutional_workbook):
        from app.export.runtime_summary import _run_project
        from domain.financial_statements import assemble_financial_statements

        _, result = _run_project("tuho")
        statements = assemble_financial_statements(result)
        expected_total_revenue = sum(p.revenues_keur for p in statements.pnl.periods)

        sheet = institutional_workbook["P&L"]
        # Find the "Revenue" row and its "Total" column.
        for row in sheet.iter_rows(min_row=7):
            if row[0].value == "Revenue":
                total_col_idx = None
                header_row = sheet[8]
                for idx, cell in enumerate(header_row):
                    if cell.value == "Total":
                        total_col_idx = idx
                        break
                if total_col_idx is not None:
                    assert row[total_col_idx].value == pytest.approx(expected_total_revenue, rel=1e-6)
                break


# ─── 6: No banned internal wording in exported cell values ────────────────

class TestNoBannedWordingInExports:
    def test_institutional_workbook_has_no_banned_terms(self, institutional_workbook):
        text = _all_cell_text(institutional_workbook).lower()
        for term in BANNED_TERMS:
            assert term not in text, f"banned term {term!r} found in institutional workbook export"

    def test_runtime_summary_csv_has_no_banned_terms(self, runtime_summary_csv_text):
        text = runtime_summary_csv_text.lower()
        for term in BANNED_TERMS:
            assert term not in text, f"banned term {term!r} found in runtime summary CSV export"

    def test_values_only_workbook_has_no_banned_terms(self, values_only_workbook):
        text = _all_cell_text(values_only_workbook).lower()
        for term in BANNED_TERMS:
            assert term not in text, f"banned term {term!r} found in values-only Excel export"

    def test_depreciation_audit_sheet_is_genuinely_audit_only(self, values_only_workbook):
        # The one legitimate "audit-only" exception: explicitly disclosed,
        # and the "Value" column (col C, since col A is a bare row index)
        # contains no numeric depreciation values -- text-only disclosures.
        sheet = values_only_workbook["Depreciation Audit"]
        header = [c.value for c in sheet[1]]
        value_col_idx = header.index("Value") + 1  # 1-based openpyxl column
        for row in sheet.iter_rows(min_row=2):
            cell = row[value_col_idx - 1]
            if cell.value is not None:
                assert not isinstance(cell.value, (int, float)), (
                    "Depreciation Audit sheet's Value column must remain "
                    "text-only to justify its audit-only framing"
                )


# ─── 7: Workbook generation unchanged for genuine sheets ───────────────────

class TestWorkbookGenerationUnchanged:
    def test_institutional_workbook_source_not_modified_in_diff(self):
        result = subprocess.run(
            ["git", "diff", "main", "--name-only"],
            cwd=PROJECT_ROOT, capture_output=True, text=True,
        )
        if result.returncode != 0:
            pytest.skip("git diff against main not available in this sandbox")
        changed = result.stdout.splitlines()
        for forbidden in (
            "app/export/institutional_workbook.py",
            "app/export/workbook_index.py",
            "app/export/runtime_summary.py",
            "app/excel_export.py",
            "main_web.py",
        ):
            assert forbidden not in changed, (
                f"{forbidden} should not need changes -- PR11 investigation found "
                "the export pipeline already honest"
            )


# ─── 8: Guardrail files untouched ──────────────────────────────────────────

class TestGuardrailFilesUntouched:
    RESTRICTED_PATHS = [
        "domain/",
        "app/waterfall_core.py",
        "app/input_adapter.py",
        "app/project_factories.py",
        "static/modelling/runtime-renderer.js",
        "app/services/model_preview.py",
        "app/services/preview_context.py",
        "app/services/previews/",
    ]

    def test_restricted_paths_not_in_diff_against_main(self):
        result = subprocess.run(
            ["git", "diff", "main", "--name-only"],
            cwd=PROJECT_ROOT, capture_output=True, text=True,
        )
        if result.returncode != 0:
            pytest.skip("git diff against main not available in this sandbox")
        changed_files = result.stdout.splitlines()
        for changed in changed_files:
            for restricted in self.RESTRICTED_PATHS:
                assert not changed.startswith(restricted), (
                    f"Restricted path touched: {changed!r} matches {restricted!r}"
                )


# ─── 9: Orphaned calibration_reconciliation pipeline confirmed unreachable ─

class TestOrphanedCalibrationPipelineUnreachable:
    def test_calibration_reconciliation_not_imported_in_main_web(self):
        with open(os.path.join(PROJECT_ROOT, "main_web.py"), "r", encoding="utf-8") as f:
            main_web_source = f.read()
        assert "calibration_reconciliation" not in main_web_source

    def test_export_registry_marks_unbuilt_exports_as_coming_soon(self):
        path = os.path.join(
            PROJECT_ROOT, "app", "templates", "partials", "export_registry.html"
        )
        with open(path, "r", encoding="utf-8") as f:
            src = f.read()
        assert "export-card--disabled" in src
        assert "Coming Soon" in src
        assert "Not yet available" in src


# ─── 10: Other Product Gap areas unaffected ────────────────────────────────

class TestOtherProductGapAreasUnaffected:
    def test_financials_unavailable_panel_still_present(self):
        path = os.path.join(
            PROJECT_ROOT, "app", "templates", "partials", "sheet_financials.html"
        )
        with open(path, "r", encoding="utf-8") as f:
            src = f.read()
        assert "fs-unavailable-panel" in src

    def test_tax_unavailable_panel_still_present(self):
        path = os.path.join(PROJECT_ROOT, "app", "templates", "partials", "sheet_tax.html")
        with open(path, "r", encoding="utf-8") as f:
            src = f.read()
        assert "tax-unavailable-panel" in src
