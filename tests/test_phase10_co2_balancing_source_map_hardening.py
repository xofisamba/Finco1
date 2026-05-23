"""Tests for Phase 10 CO2 / balancing source-map hardening."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
SOURCE_MAP = REPORTS / "phase10_co2_balancing_source_map.csv"
GAP_SUMMARY = REPORTS / "phase10_co2_balancing_gap_summary.csv"
EVIDENCE_QUALITY = REPORTS / "phase10_co2_balancing_evidence_quality.csv"
DOC = DOCS / "phase10_co2_balancing_source_map_hardening.md"


def _generate() -> None:
    write_calibration_reconciliation_pack()


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def test_co2_balancing_sheet_exists():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "CO2 & Balancing Reconciliation" in wb.sheetnames


def test_source_map_reports_exist():
    _generate()
    assert SOURCE_MAP.exists()
    assert GAP_SUMMARY.exists()
    assert EVIDENCE_QUALITY.exists()


def test_grouped_and_evidence_quality_columns_exist():
    _generate()
    rows = _csv_rows(SOURCE_MAP)
    assert rows
    required = {
        "metric",
        "runtime_source",
        "excel_source",
        "grouped_or_separate",
        "evidence_quality",
        "runtime_available",
        "excel_available",
        "workbook_bound",
        "unresolved_reason",
        "notes",
    }
    assert required.issubset(rows[0].keys())


def test_grouped_only_rows_are_clearly_labeled():
    _generate()
    rows = _csv_rows(SOURCE_MAP)
    grouped = [row for row in rows if row["classification"] == "GROUPED_SOURCE_ONLY"]
    assert grouped
    assert all(row["grouped_or_separate"] == "GROUPED_ONLY" for row in grouped)


def test_no_fabricated_split_values_introduced():
    _generate()
    rows = _csv_rows(SOURCE_MAP)
    by_metric = {row["metric"]: row for row in rows}
    assert by_metric["CO2 Revenue"]["runtime_value_available"] == "MISSING_EVIDENCE"
    assert by_metric["CO2 Revenue"]["excel_value_available"] == "MISSING_EVIDENCE"
    assert by_metric["Balancing Revenue"]["runtime_value_available"] == "MISSING_EVIDENCE"
    assert by_metric["Balancing Revenue"]["excel_value_available"] == "MISSING_EVIDENCE"


def test_gap_register_contains_co2_balancing_rows():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Gap Register")
    assert "CO2 / Balancing" in text
    assert "Grouped Revenue Evidence" in text
    assert "Balancing Revenue" in text


def test_executive_dashboard_and_notes_contain_co2_guidance():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    dashboard = _sheet_text(wb, "Executive Dashboard")
    notes = _sheet_text(wb, "Reviewer Notes")
    assert "CO2 / balancing evidence summary" in dashboard
    assert "Grouped vs separated evidence summary" in dashboard
    assert "How to interpret grouped revenue evidence" in notes
    assert "CO2 / balancing reviewer caution" in notes


def test_governance_posture_is_unchanged():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance = _sheet_text(wb, "Governance")
    assert "G20 status BLOCKED" in governance
    assert "R99/R102 status NOT APPROVED" in governance


def test_doc_states_scope_and_limits():
    content = DOC.read_text(encoding="utf-8")
    assert "grouped vs separate evidence policy" in content.lower()
    assert "evidence-quality methodology" in content.lower()
    assert "G20 remains `BLOCKED`" in content
    assert "R99/R102 remain `NOT APPROVED`" in content
    assert "no runtime changes statement" in content.lower()


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
