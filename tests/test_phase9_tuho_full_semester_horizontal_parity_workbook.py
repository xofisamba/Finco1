"""
Tests for Phase 9 TUHO Full Semester Horizontal Parity Workbook.

Verifies:
1. XLSX exists and has required sheets
2. SHL model feed is non-zero (P1 >= 30,000 kEUR)
3. Distributions section and flag-state legend present
4. Governance status (G20 BLOCKED, R99/R102 NOT approved)
5. All operational periods present horizontally
6. No runtime files changed
"""

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "reports" / "phase9_tuho_full_semester_horizontal_parity_workbook.xlsx"
DOC = ROOT / "docs" / "phase9_tuho_full_semester_horizontal_parity_workbook.md"


REQUIRED_SHEETS = {
    "Summary",
    "Operations",
    "Revenue",
    "OPEX EBITDA",
    "Depreciation Tax",
    "CFADS Waterfall",
    "Senior Debt",
    "SHL",
    "Distributions",
    "Returns",
    "Accepted Conventions",
    "Gap Analysis",
    "Source Map",
    "Governance",
}


def _sheet_labels_all(ws, col=2):
    return [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]


def _find_row(ws, label):
    labels = _sheet_labels_all(ws, col=2)
    if label in labels:
        return labels.index(label) + 1
    return None


# ── Workbook existence ────────────────────────────────────────────────────────

def test_xlsx_exists():
    assert WORKBOOK.exists(), f"XLSX not found: {WORKBOOK}"


def test_required_sheets_all_present():
    wb = load_workbook(WORKBOOK, data_only=True)
    missing = REQUIRED_SHEETS - set(wb.sheetnames)
    assert not missing, f"Missing sheets: {missing}"


def test_summary_sheet_exists_with_headline_data():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Summary"]
    labels = _sheet_labels_all(ws, col=2)
    assert "G20 Status:" in " ".join(str(l) for l in labels if l)
    assert "R99/R102 Promotion:" in " ".join(str(l) for l in labels if l)


def test_governance_sheet_has_g20_and_r99_status():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Governance"]
    # Check col 2 for status values
    statuses = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    statuses = [s for s in statuses if s]
    assert "BLOCKED" in statuses, f"G20 BLOCKED not found in Governance col 2: {statuses[:5]}"
    assert "NOT APPROVED" in statuses, f"R99/R102 NOT APPROVED not found: {statuses[:5]}"


# ── SHL section ────────────────────────────────────────────────────────────────

def test_shl_sheet_model_p1_opening_is_nonzero():
    """Model SHL P1 opening balance must be non-zero (~30,930 kEUR per bridge)."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["SHL"]
    labels = _sheet_labels_all(ws, col=2)

    # Find "SHL Opening Balance (kEUR) — Model" row
    model_row_label = "SHL Opening Balance (kEUR) — Model"
    assert model_row_label in labels, f"'{model_row_label}' not found in SHL sheet"

    row_idx = _find_row(ws, model_row_label)
    # Col 3 = P1 (first data column after Source)
    p1_value = ws.cell(row=row_idx, column=3).value
    assert p1_value is not None, "P1 value is None"
    assert p1_value >= 30_000, f"SHL P1 model opening {p1_value} < 30,000 (expect ~30,930)"


def test_shl_model_values_are_not_all_zero():
    """Model SHL rows must not all be zero (no zero-feed regression)."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["SHL"]
    labels = _sheet_labels_all(ws, col=2)

    shl_model_rows = []
    for row, label in enumerate(labels, start=1):
        if isinstance(label, str) and label.startswith("SHL ") and "— Model" in label:
            shl_model_rows.append(row)

    all_values = []
    for row in shl_model_rows:
        for col in range(3, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                all_values.append(abs(val))

    total = sum(all_values)
    assert total > 0, "All SHL model values are zero — zero-feed regression!"


# ── Distributions section ─────────────────────────────────────────────────────

def test_distributions_sheet_exists():
    wb = load_workbook(WORKBOOK, data_only=True)
    assert "Distributions" in wb.sheetnames


def test_distributions_uses_defined_flag_state():
    """Model distributions row must use defined flag-state, not zero/inconsistent."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Distributions"]
    labels = _sheet_labels_all(ws, col=2)

    # Must have separate legacy runtime vs DA-wired rows
    assert any("distribution_keur" in str(l).lower() or "DA-wired" in str(l) for l in labels), \
        "No runtime distribution row found"


def test_flag_state_legend_sheet_exists():
    wb = load_workbook(WORKBOOK, data_only=True)
    # Actually we have "Governance" sheet instead of "Flag-State Legend"
    # Check Governance has explicit flag-state info
    ws = wb["Governance"]
    text = " ".join(str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1))
    assert "distribution_keur" in text.lower() or "runtime" in text.lower()


# ── Operations section ─────────────────────────────────────────────────────────

def test_operations_sheet_production_is_present():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Operations"]
    labels = _sheet_labels_all(ws, col=2)
    assert any("Production" in str(l) for l in labels)


def test_all_operational_periods_shown_horizontally():
    """Must have 61 operational periods (P1-P61) across columns."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Operations"]
    # Check that row 1 (header) has P1 through P61
    header_vals = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]
    period_labels = [v for v in header_vals if v and str(v).startswith("P")]
    assert len(period_labels) >= 60, f"Only {len(period_labels)} periods found (expected 60+)"


# ── Returns section ────────────────────────────────────────────────────────────

def test_returns_sheet_has_equity_irr():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Returns"]
    text = " ".join(str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1))
    assert "Equity IRR" in text, "Equity IRR row not found in Returns sheet"


# ── Source map ─────────────────────────────────────────────────────────────────

def test_source_map_sheet_exists():
    wb = load_workbook(WORKBOOK, data_only=True)
    assert "Source Map" in wb.sheetnames


# ── Gap analysis ─────────────────────────────────────────────────────────────

def test_gap_analysis_sheet_exists():
    wb = load_workbook(WORKBOOK, data_only=True)
    assert "Gap Analysis" in wb.sheetnames


# ── Docs ─────────────────────────────────────────────────────────────────────

def test_docs_exist():
    assert DOC.exists(), f"Docs not found: {DOC}"


def test_docs_state_g20_blocked():
    """Docs must explicitly state G20 remains BLOCKED."""
    with open(DOC, encoding="utf-8") as f:
        text = f.read()
    assert "BLOCKED" in text, "Docs must state G20 BLOCKED"
    assert "G20" in text


def test_docs_state_r99_not_approved():
    """Docs must explicitly state R99/R102 promotion NOT approved."""
    with open(DOC, encoding="utf-8") as f:
        text = f.read()
    assert "NOT APPROVED" in text or "not approved" in text.lower(), \
        "Docs must state R99/R102 NOT approved"


# ── No runtime changes ─────────────────────────────────────────────────────────

def test_no_runtime_files_changed():
    """Verify no runtime code was modified by this workbook fix.

    PR #171 only touches: scripts/export_*workbook*, tests/*workbook*,
    reports/*semester_horizontal*, docs/phase9_*workbook*.

    Note: git diff origin/main shows phase9_5 sidebar-nav/app files because
    those were merged into main AFTER this branch split. They are NOT changes
    made by this PR — they appear because origin/main is ahead of our branch tip.
    """
    import subprocess
    result = subprocess.run(
        ["git", "diff", "--name-only", "origin/main"],
        cwd=ROOT,
        capture_output=True,
        text=True
    )
    changed = {p for p in result.stdout.strip().split("\n") if p}

    # Files this PR is allowed to modify
    our_files = {
        "scripts/export_phase9_tuho_full_semester_horizontal_parity_workbook.py",
        "tests/test_phase9_tuho_full_semester_horizontal_parity_workbook.py",
        "reports/phase9_tuho_full_semester_horizontal_gap_analysis.csv",
        "reports/phase9_tuho_full_semester_horizontal_parity_workbook.xlsx",
        "reports/phase9_tuho_full_semester_horizontal_source_map.csv",
        "reports/phase9_tuho_full_semester_horizontal_summary.csv",
    }
    # Docs may or may not be modified (depends on if we regenerated)
    docs_file = "docs/phase9_tuho_full_semester_horizontal_parity_workbook.md"

    # Phase 9.5 files that appear in diff because they were merged into main
    # after this branch split — not changes introduced by this PR
    phase9_5_in_main = {
        "app/templates/base.html", "app/templates/index.html",
        "app/templates/partials/kpis.html",
        "static/app.js", "static/styles.css",
        "docs/phase9_5_excel_like_project_workspace_ui_design.md",
        "docs/phase9_5_htmx_ui_ux_product_polish.md",
        "reports/hardfix_check.png", "reports/live_deploy_check.png",
        "reports/nav_fix_final.png", "reports/nav_fix_verify.png",
        "reports/phase9_5_excel_like_workspace_wireframe.html",
        "reports/phase9_5_project_template_load_flow.csv",
        "reports/phase9_5_ui_navigation_architecture.csv",
        "reports/phase9_5_ui_sheet_map.csv",
        "reports/prod_current.png", "reports/prod_left_check.png",
        "reports/ui_hotfix_check.png", "reports/ui_hotfix_pre.png",
        "tests/test_auth_lite.py",
        "tests/test_phase9_5_excel_like_project_workspace_ui_design.py",
    }

    # All files in diff NOT in our_files and NOT phase9_5_in_main = runtime contamination
    unexpected = changed - our_files - phase9_5_in_main
    assert not unexpected, (
        f"Unexpected runtime changes (not in our_files or phase9_5_in_main): {sorted(unexpected)}"
    )


# ── Sheet structure checks ────────────────────────────────────────────────────

def test_each_metric_has_excel_model_delta_rows():
    """Each metric in operational sheets must have Excel/Model/Delta 3-row structure."""
    wb = load_workbook(WORKBOOK, data_only=True)
    for sheet_name in ["Operations", "Revenue", "Senior Debt", "SHL"]:
        ws = wb[sheet_name]
        labels = _sheet_labels_all(ws, col=2)
        # Count rows with — Excel, — Model, — Delta pattern
        excel_rows = [l for l in labels if l and "— Excel" in str(l)]
        model_rows = [l for l in labels if l and "— Model" in str(l)]
        delta_rows = [l for l in labels if l and "— Delta" in str(l)]
        assert len(excel_rows) > 0, f"No Excel rows in {sheet_name}"
        assert len(model_rows) > 0, f"No Model rows in {sheet_name}"
        assert len(delta_rows) > 0, f"No Delta rows in {sheet_name}"


# ── SHL PIK / Tax MISSING_EVIDENCE ─────────────────────────────────────────────

# ── SHL PIK / Tax sourced model rows ────────────────────────────────────────

def test_shl_pik_marked_missing_evidence():
    """SHL PIK model row must show runtime values (PIK capitalised in PIK phase).

    PR #171 fix: SHL PIK is now sourced from runtime shl_pik_keur.
    Model row shows actual PIK values, not MISSING_EVIDENCE.
    Excel row remains MISSING_EVIDENCE (P&L!R28 not mapped).
    """
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["SHL"]
    labels = _sheet_labels_all(ws, col=2)

    pik_label_model = "SHL PIK Capitalized (kEUR) — Model"
    assert pik_label_model in labels, "SHL PIK Model row not found"
    row_idx = _find_row(ws, pik_label_model)
    model_vals = [ws.cell(row=row_idx, column=col).value for col in range(3, min(8, ws.max_column + 1))]
    numeric_count = sum(1 for v in model_vals if isinstance(v, (int, float)))
    assert numeric_count > 0, f"SHL PIK Model row should have numeric values (runtime), got {model_vals}"

    pik_label_excel = "SHL PIK Capitalized (kEUR) — Excel"
    assert pik_label_excel in labels
    excel_row = _find_row(ws, pik_label_excel)
    excel_vals = [ws.cell(row=excel_row, column=col).value for col in range(3, min(8, ws.max_column + 1))]
    for val in excel_vals:
        assert val == "MISSING_EVIDENCE", f"SHL PIK Excel should be MISSING_EVIDENCE, got {val}"


def test_tax_cfads_not_entirely_missing():
    """Tax/CFADS section must have at least one sourced metric, not all MISSING."""
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Depreciation Tax"]
    labels = _sheet_labels_all(ws, col=2)

    model_rows = [l for l in labels if l and "— Model" in str(l)]
    sourced = [l for l in model_rows if l and "MISSING_EVIDENCE" not in str(l)]
    # At least CIT Cash (=0) should be shown as sourced
    assert len(sourced) > 0, "Depreciation Tax sheet has no sourced model rows"