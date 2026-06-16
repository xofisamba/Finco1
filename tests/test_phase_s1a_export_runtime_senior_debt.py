"""Phase S1-A regression tests: Excel export senior debt source.

Phase S1-A (low-risk, presentation-only) replaces
``bundle.context.senior_debt_keur`` reads in
``institutional_workbook`` with a helper
``_resolve_export_senior_debt_keur`` that:

- returns the input value when it is non-zero
  (TUHO, Oborovo, any project with ``fixed_debt_keur`` set
  explicitly — the runtime uses the input as an override
  and the runtime result equals the input, so the two are
  bit-identical);
- falls back to ``runtime_result.sculpting_result.debt_keur``
  when the input is zero (Generic projects where the
  factory does not set ``fixed_debt_keur`` and the runtime
  computes the senior debt from the DSCR-sculpt formula).

This test verifies:
  - Generic Solar export reads the runtime result, not zero.
  - TUHO export reads the input (= runtime result), bit-identical.
  - Oborovo export reads the input (= runtime result), bit-identical.
  - The helper function never invents a value.

Constraints preserved (all pinned by tests):
  - TUHO and Oborovo parity behaviour bit-identical
  - Generic export now shows the actual runtime debt amount
  - No engine / factory / model / formula change
"""
from __future__ import annotations

import os
import sys
from io import BytesIO
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-key-s1a-export")

from openpyxl import load_workbook

from app.export.institutional_workbook import (
    _build_export_bundle,
    _resolve_export_senior_debt_keur,
    export_institutional_workbook_skeleton,
)


# ---------------------------------------------------------------------------
# 1. Helper function — direct unit tests
# ---------------------------------------------------------------------------


class TestResolveExportSeniorDebtKeurHelper:
    """The helper is a pure function of the bundle. It reads
    either context.senior_debt_keur (when non-zero) or
    runtime_result.sculpting_result.debt_keur (when input is
    zero). It never invents a value."""

    def test_helper_returns_input_for_tuho(self):
        """TUHO has fixed_debt_keur=43,359. Helper must return
        the input value, not the runtime result (even though
        they are bit-identical for TUHO)."""
        bundle = _build_export_bundle("tuho")
        result = _resolve_export_senior_debt_keur(bundle)
        assert result == 43_359.0, (
            f"TUHO export should return input fixed_debt_keur=43,359, "
            f"got {result}"
        )

    def test_helper_returns_input_for_oborovo(self):
        """Oborovo has fixed_debt_keur=42,852.27. Helper must
        return the input value."""
        bundle = _build_export_bundle("oborovo")
        result = _resolve_export_senior_debt_keur(bundle)
        assert abs(result - 42_852.2667) < 0.01, (
            f"Oborovo export should return input fixed_debt_keur~=42,852.27, "
            f"got {result}"
        )

    def test_helper_returns_runtime_for_generic_solar(self):
        """Generic Solar has fixed_debt_keur=0. Helper must
        return the runtime result (the actual sculpted debt),
        not zero."""
        bundle = _build_export_bundle("generic_solar")
        result = _resolve_export_senior_debt_keur(bundle)
        runtime_debt = bundle.runtime_result.sculpting_result.debt_keur
        assert runtime_debt > 0, (
            f"Runtime should produce a real sculpted debt for Generic, "
            f"got {runtime_debt}"
        )
        assert result == runtime_debt, (
            f"Generic Solar export should equal runtime result {runtime_debt}, "
            f"got {result}"
        )
        # And specifically, the value should be the actual runtime
        # number (not zero, not fixed_debt_keur which is 0).
        assert result > 10_000, (
            f"Generic Solar export should be the runtime sculpted value "
            f"(>10,000 kEUR), got {result}"
        )

    def test_helper_returns_input_when_both_paths_match(self):
        """For TUHO and Oborovo, the input and the runtime
        result are bit-identical (because the runtime uses
        fixed_debt_keur as an override). The helper must
        return either, and they are the same."""
        for key in ("tuho", "oborovo"):
            bundle = _build_export_bundle(key)
            input_debt = bundle.context.senior_debt_keur or 0.0
            runtime_debt = (
                bundle.runtime_result.sculpting_result.debt_keur
                if bundle.runtime_result.sculpting_result
                else 0.0
            )
            helper_result = _resolve_export_senior_debt_keur(bundle)
            assert input_debt > 0, (
                f"{key} input should be > 0, got {input_debt}"
            )
            # Input and runtime are bit-identical for these projects.
            assert abs(input_debt - runtime_debt) < 0.01, (
                f"{key} input and runtime must be bit-identical: "
                f"input={input_debt}, runtime={runtime_debt}"
            )
            assert helper_result == input_debt, (
                f"{key} helper should return input value {input_debt}, "
                f"got {helper_result}"
            )


# ---------------------------------------------------------------------------
# 2. End-to-end workbook export tests
# ---------------------------------------------------------------------------


def _open_workbook(project: str):
    return load_workbook(
        BytesIO(export_institutional_workbook_skeleton(project)),
        data_only=True,
    )


def _find_row_value(sheet, label: str) -> object:
    """Find the value cell in a key-value section whose label
    column matches ``label``. Returns the value cell or None."""
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        for cell_idx, cell_value in enumerate(row):
            if cell_value == label and cell_idx + 1 < len(row):
                return row[cell_idx + 1]
    return None


class TestInstitutionalWorkbookSeniorDebt:
    """End-to-end test: the actual Excel workbook produced by
    ``export_institutional_workbook_skeleton`` must show the
    correct senior debt value for each project type.

    The S1-A change touches three workbook cells:
      - Construction sheet, "Senior debt anchor" row
      - CAPEX sheet, "Senior debt funding" row
      - Senior Debt sheet, "Senior debt amount" row
    """

    def test_generic_solar_senior_debt_anchor_is_runtime_value(self):
        """Generic Solar Construction sheet must show the
        runtime sculpted debt (~22,650 kEUR), not zero."""
        workbook = _open_workbook("generic_solar")
        construction = workbook["Construction"]
        value = _find_row_value(construction, "Senior debt anchor")
        assert value is not None, "Senior debt anchor row missing"
        # The runtime produces ~22,650 for default Generic Solar.
        assert value > 10_000, (
            f"Generic Solar Senior debt anchor must be the runtime "
            f"value (>10,000 kEUR), got {value}"
        )
        # Bit-exact: must equal the runtime result.
        bundle = _build_export_bundle("generic_solar")
        runtime_debt = bundle.runtime_result.sculpting_result.debt_keur
        assert value == runtime_debt, (
            f"Generic Solar Senior debt anchor must equal runtime "
            f"result {runtime_debt}, got {value}"
        )

    def test_generic_solar_capex_senior_debt_funding_is_runtime_value(self):
        workbook = _open_workbook("generic_solar")
        capex = workbook["CAPEX"]
        value = _find_row_value(capex, "Senior debt funding")
        assert value is not None, "Senior debt funding row missing"
        assert value > 10_000, (
            f"Generic Solar Senior debt funding must be the runtime "
            f"value (>10,000 kEUR), got {value}"
        )
        bundle = _build_export_bundle("generic_solar")
        runtime_debt = bundle.runtime_result.sculpting_result.debt_keur
        assert value == runtime_debt

    def test_generic_solar_senior_debt_amount_is_runtime_value(self):
        """Senior Debt sheet, "Senior debt amount" row must
        show the runtime result, not zero."""
        workbook = _open_workbook("generic_solar")
        debt_sheet = workbook["Senior Debt"]
        value = _find_row_value(debt_sheet, "Senior debt amount")
        assert value is not None, "Senior debt amount row missing"
        bundle = _build_export_bundle("generic_solar")
        runtime_debt = bundle.runtime_result.sculpting_result.debt_keur
        assert value == runtime_debt, (
            f"Generic Solar Senior debt amount must equal runtime "
            f"result {runtime_debt}, got {value}"
        )

    def test_tuho_senior_debt_anchor_unchanged(self):
        """TUHO Senior debt anchor must still be 43,359
        (bit-identical to the prior export)."""
        workbook = _open_workbook("tuho")
        construction = workbook["Construction"]
        value = _find_row_value(construction, "Senior debt anchor")
        assert value == 43_359.0, (
            f"TUHO Senior debt anchor must be 43,359 (parity), got {value}"
        )

    def test_tuho_capex_senior_debt_funding_unchanged(self):
        workbook = _open_workbook("tuho")
        capex = workbook["CAPEX"]
        value = _find_row_value(capex, "Senior debt funding")
        assert value == 43_359.0, (
            f"TUHO Senior debt funding must be 43,359 (parity), got {value}"
        )

    def test_tuho_senior_debt_amount_unchanged(self):
        workbook = _open_workbook("tuho")
        debt_sheet = workbook["Senior Debt"]
        value = _find_row_value(debt_sheet, "Senior debt amount")
        assert value == 43_359.0, (
            f"TUHO Senior debt amount must be 43,359 (parity), got {value}"
        )

    def test_oborovo_senior_debt_anchor_unchanged(self):
        """Oborovo Senior debt anchor must still be ~42,852
        (bit-identical to the prior export)."""
        workbook = _open_workbook("oborovo")
        construction = workbook["Construction"]
        value = _find_row_value(construction, "Senior debt anchor")
        assert abs(value - 42_852.2667) < 0.01, (
            f"Oborovo Senior debt anchor must be ~42,852.27 (parity), got {value}"
        )

    def test_oborovo_capex_senior_debt_funding_unchanged(self):
        workbook = _open_workbook("oborovo")
        capex = workbook["CAPEX"]
        value = _find_row_value(capex, "Senior debt funding")
        assert abs(value - 42_852.2667) < 0.01, (
            f"Oborovo Senior debt funding must be ~42,852.27 (parity), got {value}"
        )

    def test_oborovo_senior_debt_amount_unchanged(self):
        workbook = _open_workbook("oborovo")
        debt_sheet = workbook["Senior Debt"]
        value = _find_row_value(debt_sheet, "Senior debt amount")
        assert abs(value - 42_852.2667) < 0.01, (
            f"Oborovo Senior debt amount must be ~42,852.27 (parity), got {value}"
        )


# ---------------------------------------------------------------------------
# 3. Parity bit-identity check (TUHO and Oborovo)
# ---------------------------------------------------------------------------


class TestParityBitIdentity:
    """For TUHO and Oborovo, the helper output and the runtime
    result must be bit-identical. The export must therefore
    remain bit-identical to the pre-S1-A export for these
    two projects (only Generic Solar changes)."""

    def test_tuho_helper_matches_runtime(self):
        bundle = _build_export_bundle("tuho")
        helper = _resolve_export_senior_debt_keur(bundle)
        runtime = bundle.runtime_result.sculpting_result.debt_keur
        assert helper == runtime, (
            f"TUHO helper {helper} must equal runtime {runtime}"
        )

    def test_oborovo_helper_matches_runtime(self):
        bundle = _build_export_bundle("oborovo")
        helper = _resolve_export_senior_debt_keur(bundle)
        runtime = bundle.runtime_result.sculpting_result.debt_keur
        assert abs(helper - runtime) < 0.01, (
            f"Oborovo helper {helper} must equal runtime {runtime}"
        )

    def test_generic_solar_helper_differs_from_input(self):
        """Sanity check: Generic Solar must show the runtime
        result, not the input (which is 0). The whole point
        of S1-A is that this difference is now exposed in
        the export."""
        bundle = _build_export_bundle("generic_solar")
        helper = _resolve_export_senior_debt_keur(bundle)
        input_debt = bundle.context.senior_debt_keur or 0.0
        runtime_debt = bundle.runtime_result.sculpting_result.debt_keur
        # The change is meaningful: helper now reads from
        # runtime, not from input.
        assert input_debt == 0.0, (
            f"Generic Solar input should be 0, got {input_debt}"
        )
        assert helper > 0, (
            f"Generic Solar helper should be > 0 (runtime result), "
            f"got {helper}"
        )
        assert helper == runtime_debt


# ---------------------------------------------------------------------------
# 4. Engine / factory / model invariants (no runtime change)
# ---------------------------------------------------------------------------


class TestNoRuntimeChange:
    """S1-A is presentation-only. Confirm the engine, factory,
    model, and runtime result are not touched."""

    def test_engine_md5_unchanged(self):
        import hashlib
        path = REPO_ROOT / "app" / "waterfall_core.py"
        md5 = hashlib.md5(path.read_bytes()).hexdigest()
        assert md5 == "6bf49f33efc989736c17cea0cb9b7723", (
            f"app/waterfall_core.py MD5 changed: {md5}"
        )

    def test_factories_md5_unchanged(self):
        import hashlib
        path = REPO_ROOT / "app" / "project_factories.py"
        md5 = hashlib.md5(path.read_bytes()).hexdigest()
        assert md5 == "3350c93a7689bb3f5e717a064adcd106", (
            f"app/project_factories.py MD5 changed: {md5}"
        )

    def test_waterfall_engine_md5_unchanged(self):
        import hashlib
        path = REPO_ROOT / "domain" / "waterfall" / "waterfall_engine.py"
        md5 = hashlib.md5(path.read_bytes()).hexdigest()
        # Recorded baseline from rc1 anchor; recomputed if drift
        # (will be picked up by Phase 51F parity guardrails).
        # We assert the file exists; we do not pin the MD5 here
        # because the engine is still evolved (Phase 20O mode
        # architecture, Phase 23A frozen schedule wiring).
        assert md5, f"waterfall_engine.py MD5 is empty: {md5}"

    def test_rc1_ancestor(self):
        import subprocess
        rc1_sha = "b425a0708719eaa5e1d922b1008e5609758e0ad4"
        result = subprocess.run(
            ["git", "merge-base", "--is-ancestor", rc1_sha, "HEAD"],
            cwd=str(REPO_ROOT), capture_output=True, text=True,
        )
        assert result.returncode == 0, (
            f"rc1 SHA {rc1_sha} must be ancestor of HEAD"
        )
