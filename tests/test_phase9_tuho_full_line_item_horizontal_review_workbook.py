from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "reports" / "phase9_tuho_full_line_item_horizontal_review_workbook.xlsx"
DOC = ROOT / "docs" / "phase9_tuho_full_line_item_horizontal_review_workbook.md"


REQUIRED_SHEETS = {
    "Summary",
    "Horizontal Review",
    "Gap Analysis",
    "Accepted Conventions",
    "Flag-State Legend",
    "Source Map",
}


def _sheet_labels(ws):
    return [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]


def _find_row(ws, label):
    labels = _sheet_labels(ws)
    assert label in labels
    return labels.index(label) + 1


def test_horizontal_review_workbook_exists_and_has_required_sheets():
    assert WORKBOOK.exists()
    wb = load_workbook(WORKBOOK, data_only=True)
    assert REQUIRED_SHEETS.issubset(set(wb.sheetnames))


def test_shl_section_uses_nonzero_runtime_period_fields():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Horizontal Review"]
    labels = _sheet_labels(ws)

    assert "SHL" in labels
    opening_row = _find_row(ws, "SHL Opening Balance — Model")
    assert ws.cell(row=opening_row, column=2).value >= 32_000

    shl_model_rows = [
        row
        for row, label in enumerate(labels, start=1)
        if isinstance(label, str) and label.startswith("SHL ") and label.endswith("— Model")
    ]
    values = []
    for row in shl_model_rows:
        for column in range(2, 62):
            value = ws.cell(row=row, column=column).value
            if isinstance(value, (int, float)):
                values.append(abs(value))
    assert sum(values) > 0


def test_distributions_use_defined_flag_state_and_separate_da_staging():
    wb = load_workbook(WORKBOOK, data_only=True)
    review = wb["Horizontal Review"]
    source_map = wb["Source Map"]
    labels = _sheet_labels(review)

    assert "Distributions" in labels
    assert "Net Dividends / Distribution — Model [default runtime / legacy path]" in labels
    assert "DA-wired / pre-G20 staging — Model [audit-only]" in labels

    source_rows = list(source_map.iter_rows(values_only=True))
    assert any(
        row[0] == "Distributions / Net Dividends / Distribution"
        and row[3] == "default runtime / legacy path"
        for row in source_rows
    )
    assert any(
        row[0] == "Distributions / DA-wired / pre-G20 staging"
        and row[3] == "audit-only / pre-G20 staging"
        for row in source_rows
    )


def test_tax_cfads_section_is_not_entirely_missing_evidence():
    wb = load_workbook(WORKBOOK, data_only=True)
    review = wb["Horizontal Review"]
    source_map = wb["Source Map"]
    labels = _sheet_labels(review)

    assert "Tax / CFADS" in labels
    for label in [
        "Taxable Income — Model",
        "CIT Cash — Model",
        "CFADS — Model",
    ]:
        assert label in labels

    source_rows = list(source_map.iter_rows(values_only=True))
    tax_statuses = {
        row[0].split(" / ")[-1]: row[3]
        for row in source_rows
        if row[0] in {
            "Tax / CFADS / Taxable Income",
            "Tax / CFADS / CIT Cash",
            "Tax / CFADS / CFADS",
        }
    }
    assert tax_statuses == {
        "Taxable Income": "runtime audit output",
        "CIT Cash": "runtime output",
        "CFADS": "runtime audit output",
    }


def test_gap_analysis_and_source_map_exist_with_expected_content():
    wb = load_workbook(WORKBOOK, data_only=True)
    gap = wb["Gap Analysis"]
    source_map = wb["Source Map"]
    assert gap.max_row > 2
    assert source_map.max_row > 2


def test_docs_state_governance_and_runtime_scope():
    text = DOC.read_text(encoding="utf-8")
    assert "G20 remains `BLOCKED`" in text
    assert "R99/R102 runtime promotion remains `NOT APPROVED`" in text
    assert "No runtime files were changed" in text
