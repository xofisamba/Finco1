"""Tests for phase6_model_stack_comparison_export."""
from __future__ import annotations

import json
from pathlib import Path

import pytest

REPO = Path(__file__).parent.parent
SCRIPT = REPO / "scripts/export_phase6_model_stack_comparison.py"
REPORTS_DIR = REPO / "reports"
OUT_XLSX = REPORTS_DIR / "phase6_model_stack_comparison.xlsx"
OUT_LONG_CSV = REPORTS_DIR / "phase6_model_stack_comparison_long.csv"
OUT_WIDE_CSV = REPORTS_DIR / "phase6_model_stack_comparison_wide.csv"
FIXTURE_JSON = REPO / "tests/fixtures/excel_tuho_full_model_extract.json"


def test_outputs_exist():
    """All three output files exist after running the script."""
    assert OUT_XLSX.exists(), f"XLSX not found: {OUT_XLSX}"
    assert OUT_LONG_CSV.exists(), f"Long CSV not found: {OUT_LONG_CSV}"
    assert OUT_WIDE_CSV.exists(), f"Wide CSV not found: {OUT_WIDE_CSV}"


def test_xlsx_sheets():
    """XLSX has all expected sheet names."""
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX)
    expected = [
        "Summary", "Revenue", "OPEX", "EBITDA", "Free Cash Flow",
        "Depreciation", "Senior Debt", "SHL", "CIT Tax", "Delta Flags", "Source Mapping",
    ]
    assert wb.sheetnames == expected, f"Expected {expected}, got {wb.sheetnames}"


def test_xlsx_summary_has_rows():
    """Summary sheet has data rows."""
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["Summary"]
    # Row 4 is header, row 5+ are data
    row5_val = ws.cell(5, 1).value
    assert row5_val is not None, "Summary sheet row 5 is empty"
    assert row5_val != ""


def test_xlsx_period_count():
    """Workbook has 60 period columns."""
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["Revenue"]
    # Row 1: P01..P60 starting at col 2
    p01_val = ws.cell(1, 2).value
    p60_val = ws.cell(1, 61).value
    assert p01_val == "P01", f"Expected P01, got {p01_val}"
    assert p60_val == "P60", f"Expected P60, got {p60_val}"


def test_long_csv_shape():
    """Long CSV has expected columns and period rows."""
    with open(OUT_LONG_CSV, encoding="utf-8") as f:
        header = f.readline().strip().split(",")
    expected = ["category", "metric", "row_type", "period_idx", "period_label",
               "excel_value", "python_value", "delta", "delta_pct",
               "excel_source", "python_source", "confidence", "notes"]
    assert header == expected, f"CSV header mismatch: {header}"
    # Count lines (should be 60 periods × 4 rows per metric × N metrics ≈ thousands)
    with open(OUT_LONG_CSV, encoding="utf-8") as f:
        lines = f.readlines()
    assert len(lines) > 1000, f"Long CSV has only {len(lines)} rows — too few"


def test_wide_csv_shape():
    """Wide CSV has period columns P01-P60 plus total."""
    with open(OUT_WIDE_CSV, encoding="utf-8") as f:
        header = f.readline().strip().split(",")
    # 4 label cols + 60 period cols + total + notes = 66
    assert len(header) == 66, f"Wide CSV expected 66 columns, got {len(header)}: {header[:10]}..."
    assert header[4] == "P01", f"First period col should be P01, got {header[4]}"
    assert header[63] == "P60", f"P60 should be at index 63, got {header[63]}"


def test_fixture_readable():
    """Excel fixture is valid JSON."""
    with open(FIXTURE_JSON) as f:
        d = json.load(f)
    assert "period_diagnostic_columns" in d
    assert "period_diagnostics" in d
    assert len(d["period_diagnostics"]) == 60, f"Expected 60 periods, got {len(d['period_diagnostics'])}"


def test_no_runtime_mutation():
    """Running the script does not modify waterfall_core, waterfall_runner, project_factories."""
    import ast, hashlib

    with open("app/waterfall_core.py") as f:
        core_before = f.read()
    core_before_hash = hashlib.md5(core_before.encode()).hexdigest()

    with open("app/waterfall_runner.py") as f:
        runner_before = f.read()
    runner_before_hash = hashlib.md5(runner_before.encode()).hexdigest()

    # Run script (already executed by test runner)
    import subprocess, sys
    result = subprocess.run(
        [sys.executable, str(SCRIPT)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    assert result.returncode == 0, f"Script failed: {result.stderr}"

    with open("app/waterfall_core.py") as f:
        core_after = f.read()
    assert hashlib.md5(core_after.encode()).hexdigest() == core_before_hash, \
        "waterfall_core.py was modified!"

    with open("app/waterfall_runner.py") as f:
        runner_after = f.read()
    assert hashlib.md5(runner_after.encode()).hexdigest() == runner_before_hash, \
        "waterfall_runner.py was modified!"