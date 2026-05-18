"""Validation for the TUHO R35 post-SHL-gross bridge workbook."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner


WORKBOOK_PATH = Path("reports/phase6_tuho_r35_post_shl_gross_bridge.xlsx")
REQUIRED_SHEETS = {
    "Summary",
    "R35 Before After",
    "R27 SHL Interest",
    "Remaining Drivers",
    "Largest Remaining Deltas",
    "R67 Impact",
    "Notes",
}


def _run_default_tuho():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return WaterfallRunner(project, engine).run(config)


def test_post_shl_gross_workbook_exists_with_required_sheets():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    assert REQUIRED_SHEETS.issubset(set(workbook.sheetnames))


def test_r35_before_after_contains_60_operating_periods():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R35 Before After"]

    op_indices = [sheet.cell(row=row, column=1).value for row in range(2, 62)]

    assert op_indices == list(range(60))
    assert sheet.cell(row=62, column=1).value == "TOTAL"


def test_summary_captures_shl_driver_closure_and_remaining_delta():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    summary = workbook["Summary"]
    values = {
        summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
        for row in range(2, summary.max_row + 1)
    }

    assert values["R35 delta before SHL gross bridge"] == pytest.approx(
        12_216.370554918292,
        abs=0.01,
    )
    assert values["R35 delta after SHL gross bridge"] == pytest.approx(
        1_869.0997795006383,
        abs=0.01,
    )
    assert values["Delta reduction"] == pytest.approx(10_347.270775417655, abs=0.01)
    assert values["SHL driver closed?"] == "Yes"
    assert values["R99 readiness"] == "Blocked"
    assert values["Recommended next branch"] == "phase6-depreciation-book-tax-ledger-design"


def test_r27_shl_interest_delta_is_closed_by_flag_on_bridge():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R27 SHL Interest"]

    legacy_deltas = [sheet.cell(row=row, column=5).value for row in range(2, 62)]
    flag_deltas = [sheet.cell(row=row, column=7).value for row in range(2, 62)]
    closed_flags = [sheet.cell(row=row, column=8).value for row in range(2, 62)]

    assert sum(legacy_deltas) == pytest.approx(10_347.270775417655, abs=0.01)
    assert max(abs(delta) for delta in flag_deltas) == pytest.approx(0.0, abs=0.0001)
    assert all(closed_flags)


def test_largest_remaining_deltas_are_depreciation_driven_and_sorted():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Largest Remaining Deltas"]

    abs_deltas = [sheet.cell(row=row, column=7).value for row in range(2, 62)]

    assert abs_deltas == sorted(abs_deltas, reverse=True)
    assert sheet.cell(row=2, column=2).value == 59
    assert sheet.cell(row=2, column=8).value == "book vs tax depreciation"


def test_r67_impact_sheet_marks_bridge_effect_as_zero():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R67 Impact"]

    bridge_effects = [sheet.cell(row=row, column=6).value for row in range(2, 62)]

    assert max(abs(effect) for effect in bridge_effects) == pytest.approx(0.0, abs=0.0001)
    assert sheet.cell(row=62, column=7).value == "R67 unchanged by P&L-only SHL bridge"


def test_workbook_generation_does_not_change_default_runtime_behavior():
    baseline = _run_default_tuho()
    diagnostic = _run_default_tuho()

    assert diagnostic.total_revenue_keur == pytest.approx(baseline.total_revenue_keur, abs=0.0001)
    assert diagnostic.total_opex_keur == pytest.approx(baseline.total_opex_keur, abs=0.0001)
    assert diagnostic.total_tax_keur == pytest.approx(baseline.total_tax_keur, abs=0.0001)
    assert diagnostic.total_distribution_keur == pytest.approx(
        baseline.total_distribution_keur,
        abs=0.0001,
    )
