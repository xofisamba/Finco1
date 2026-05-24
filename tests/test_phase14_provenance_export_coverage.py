from __future__ import annotations

import csv
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.excel_export import build_excel_export
from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.export.runtime_summary import build_runtime_summary_rows
from app.persistence.provenance import build_replay_metadata
from app.ui_runner import run_demo_project


ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "phase14_provenance_export_coverage.md"
PATH_INVENTORY = ROOT / "reports" / "phase14_export_path_inventory.csv"
FIELD_MATRIX = ROOT / "reports" / "phase14_export_provenance_field_matrix.csv"
GAP_REGISTER = ROOT / "reports" / "phase14_export_provenance_gap_register.csv"
PHASE9_DOC = ROOT / "docs" / "phase9_audit_economic_mode_contract_reconciliation.md"


def _sheet_values(sheet) -> list[object]:
    return [
        sheet.cell(row=row, column=column).value
        for row in range(1, sheet.max_row + 1)
        for column in range(1, sheet.max_column + 1)
    ]


def _notes_mapping(sheet) -> dict[str, object]:
    mapping: dict[str, object] = {}
    for row in range(2, sheet.max_row + 1):
        field = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if isinstance(field, str) and field.strip():
            mapping[field] = value
    return mapping


def _first_numeric_values(sheet, limit: int = 8) -> list[float]:
    values: list[float] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                values.append(float(cell.value))
                if len(values) >= limit:
                    return values
    raise AssertionError(f"No numeric values found in sheet {sheet.title!r}")


def test_runtime_summary_rows_include_canonical_provenance_fields():
    rows = build_runtime_summary_rows(
        "tuho",
        generated_at="2026-05-24T10:00:00+00:00",
        source_branch="phase14-provenance-test",
    )
    assert rows
    first = rows[0]

    required_fields = {
        "export_generated_at",
        "runtime_generated_at",
        "branch_name",
        "active_project",
        "scenario_id",
        "scenario_name",
        "scenario_revision",
        "runtime_snapshot_id",
        "runtime_origin",
        "template_origin",
        "template_revision",
        "export_template_version",
        "runtime_flag_count",
        "runtime_flags_json",
        "governance_posture_summary",
    }
    assert required_fields <= set(first.keys())
    assert first["export_generated_at"] == "2026-05-24T10:00:00+00:00"
    assert first["runtime_generated_at"]
    assert first["active_project"] == "tuho"
    assert first["scenario_id"] == "not_applicable"
    assert first["scenario_name"] == "not_applicable"
    assert first["scenario_revision"] == "not_applicable"
    assert first["runtime_snapshot_id"] == "unavailable"
    assert first["runtime_origin"] == "factory_base_runtime"
    assert first["template_origin"] == "project_factory:tuho"
    assert first["export_template_version"] == "not separately versioned"
    assert "G20 remains BLOCKED" in first["governance_posture_summary"]
    assert "R99/R102 remain NOT APPROVED" in first["governance_posture_summary"]


def test_institutional_workbook_contains_reviewer_readable_provenance_rows():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=True)
    cover_values = _sheet_values(workbook["Cover"])
    governance_values = _sheet_values(workbook["Governance"])

    assert "Active project" in cover_values
    assert "Scenario ID" in cover_values
    assert "Runtime snapshot ID" in cover_values
    assert "Runtime origin" in cover_values
    assert "Export template version" in cover_values
    assert "Governance posture summary" in governance_values
    assert "Branch name" in governance_values
    assert "Export generated at" in governance_values
    assert "Scenario revision" in governance_values
    assert "Runtime snapshot ID" in governance_values
    assert "not_applicable" in governance_values
    assert "unavailable" in governance_values
    assert any(
        isinstance(value, str) and "G20 remains BLOCKED; R99/R102 remain NOT APPROVED" in value
        for value in governance_values
    )


def test_values_only_excel_export_notes_sheet_surfaces_provenance_without_changing_numeric_cells():
    demo = run_demo_project("Solar")
    provenance_metadata = build_replay_metadata(
        project_key="oborovo",
        project_inputs=demo.project_inputs,
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"},
        export_type="excel_model_export",
        workbook_type="values_only_excel_export",
        active_project="oborovo",
        scenario_name="Base",
        runtime_origin="factory_base_runtime",
    )

    plain_wb = load_workbook(
        BytesIO(build_excel_export(result=demo.result, project_inputs=demo.project_inputs)),
        data_only=True,
    )
    prov_wb = load_workbook(
        BytesIO(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata=provenance_metadata,
            )
        ),
        data_only=True,
    )

    notes = _notes_mapping(prov_wb["Notes"])
    assert notes["Branch Name"] == provenance_metadata["branch_name"]
    assert notes["Active Project"] == "oborovo"
    assert notes["Scenario ID"] == "not_applicable"
    assert notes["Runtime Snapshot ID"] == "unavailable"
    assert notes["Runtime Origin"] == "factory_base_runtime"
    assert notes["Template Origin"] == "project_factory:oborovo"
    assert "G20 remains BLOCKED" in str(notes["Governance Posture"])
    assert "R99/R102 remain NOT APPROVED" in str(notes["Governance Posture"])
    assert "deterministic replay is not fully guaranteed" in str(notes["Replay Limitations"]).lower()

    assert _first_numeric_values(plain_wb["Dashboard"]) == _first_numeric_values(prov_wb["Dashboard"])
    assert _first_numeric_values(plain_wb["Revenue"]) == _first_numeric_values(prov_wb["Revenue"])


def test_reports_and_docs_capture_active_export_inventory_and_guardrails():
    assert DOC.exists()
    assert PATH_INVENTORY.exists()
    assert FIELD_MATRIX.exists()
    assert GAP_REGISTER.exists()

    inventory_rows = list(csv.DictReader(PATH_INVENTORY.open(newline="", encoding="utf-8")))
    field_rows = list(csv.DictReader(FIELD_MATRIX.open(newline="", encoding="utf-8")))
    gap_rows = list(csv.DictReader(GAP_REGISTER.open(newline="", encoding="utf-8")))

    assert {row["export_type"] for row in inventory_rows} == {
        "excel_model_export",
        "runtime_summary_csv",
        "institutional_workbook",
    }
    assert any(row["field_name"] == "runtime_snapshot_id" for row in field_rows)
    assert any(row["field_name"] == "governance_posture_summary" for row in field_rows)
    assert any(row["status"] == "done" for row in gap_rows)

    doc_text = DOC.read_text(encoding="utf-8")
    phase9_text = PHASE9_DOC.read_text(encoding="utf-8")
    assert "Minor descriptive provenance gaps" in doc_text
    assert "No runtime formulas were changed" in doc_text
    assert "No workbook calculations were changed" in doc_text
    assert "No persistence authority promotion occurred" in doc_text
    assert "No replay engine behavior was added" in doc_text
    assert "`G20` remains `BLOCKED`" in doc_text
    assert "`R99/R102` remain `NOT APPROVED`" in doc_text
    assert "`audit_economic_mode` remains audit/reconciliation-only." in doc_text
    assert "`runtime_economic_mode` remains the only explicit runtime staging path." in doc_text
    assert "Conclusion" in phase9_text
