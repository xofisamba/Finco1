from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.export.runtime_summary import build_runtime_summary_rows


ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "phase15_numeric_workbook_validation.md"
MAPPING = ROOT / "reports" / "phase15_institutional_workbook_numeric_mapping.csv"
GAPS = ROOT / "reports" / "phase15_numeric_workbook_gap_register.csv"
MATRIX = ROOT / "reports" / "phase15_numeric_workbook_validation_matrix.csv"


def _value_for_label(sheet, label: str):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return sheet.cell(row=row, column=2).value
    raise AssertionError(f"Label {label!r} not found in sheet {sheet.title!r}")


def _cell_for_label(sheet, label: str):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return sheet.cell(row=row, column=2)
    raise AssertionError(f"Label {label!r} not found in sheet {sheet.title!r}")


def _runtime_value(rows: list[dict[str, str]], metric: str) -> float:
    row = next(item for item in rows if item["metric"] == metric)
    return float(row["value"])


def _sheet_text(sheet) -> str:
    parts: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return "\n".join(parts)


def test_institutional_workbook_representative_numeric_values_match_backend_runtime():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=True)
    runtime_rows = build_runtime_summary_rows("tuho")

    runtime_summary = workbook["Runtime Summary"]
    opex = workbook["OPEX"]
    governance = workbook["Governance"]
    cover = workbook["Cover"]

    workbook_project_irr = _value_for_label(runtime_summary, "Project IRR")
    workbook_total_revenue = _value_for_label(runtime_summary, "Total revenue")
    workbook_avg_dscr = _value_for_label(runtime_summary, "Average DSCR")
    workbook_total_opex = _value_for_label(opex, "Runtime total OPEX")

    assert isinstance(workbook_project_irr, (int, float))
    assert isinstance(workbook_total_revenue, (int, float))
    assert isinstance(workbook_avg_dscr, (int, float))
    assert isinstance(workbook_total_opex, (int, float))

    assert abs(float(workbook_project_irr) - _runtime_value(runtime_rows, "project_irr")) < 1e-9
    assert abs(float(workbook_total_revenue) - _runtime_value(runtime_rows, "total_revenue_keur")) < 1e-6
    assert abs(float(workbook_avg_dscr) - _runtime_value(runtime_rows, "avg_dscr")) < 1e-9
    assert abs(float(workbook_total_opex) - _runtime_value(runtime_rows, "total_opex_keur")) < 1e-6

    governance_text = _sheet_text(governance)
    cover_text = _sheet_text(cover)
    assert "G20 remains BLOCKED" in governance_text or "G20 remains BLOCKED" in cover_text
    assert "R99/R102 remain NOT APPROVED" in governance_text or "R99/R102 remain NOT APPROVED" in cover_text


def test_institutional_workbook_missing_markers_and_cover_notes_remain_explicit():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=False)
    governance = workbook["Governance"]
    cover = workbook["Cover"]
    runtime_summary = workbook["Runtime Summary"]
    opex = workbook["OPEX"]

    assert _value_for_label(governance, "Scenario ID") == "not_applicable"
    assert _value_for_label(governance, "Runtime snapshot ID") == "unavailable"

    cover_text = _sheet_text(cover).lower()
    governance_text = _sheet_text(governance).lower()
    assert "financial source of truth" in cover_text
    assert "descriptive only" in governance_text or "descriptive only" in cover_text
    assert "not lender-ready without external model review" in governance_text
    assert "not audit-certified" in governance_text
    assert "unavailable" in governance_text
    assert "not_applicable" in governance_text

    # Representative validated cells should remain direct numeric values rather than formulas.
    assert not (isinstance(_cell_for_label(runtime_summary, "Project IRR").value, str) and _cell_for_label(runtime_summary, "Project IRR").value.startswith("="))
    assert not (isinstance(_cell_for_label(opex, "Runtime total OPEX").value, str) and _cell_for_label(opex, "Runtime total OPEX").value.startswith("="))


def test_docs_and_reports_capture_numeric_validation_scope_and_guardrails():
    assert DOC.exists()
    assert MAPPING.exists()
    assert GAPS.exists()
    assert MATRIX.exists()

    doc_text = DOC.read_text(encoding="utf-8")
    mapping_text = MAPPING.read_text(encoding="utf-8")
    gaps_text = GAPS.read_text(encoding="utf-8")
    matrix_text = MATRIX.read_text(encoding="utf-8")

    assert "Project IRR" in doc_text
    assert "Total revenue" in doc_text
    assert "Runtime total OPEX" in doc_text
    assert "Average DSCR" in doc_text
    assert "does **not** make the workbook authoritative" in doc_text
    assert "Missing or unavailable values are not silently treated as zero." in doc_text
    assert "No runtime/model formulas were changed." in doc_text
    assert "No workbook calculations were changed." in doc_text
    assert "No workbook formulas were changed." in doc_text
    assert "No export calculation logic was changed." in doc_text
    assert "`audit_economic_mode` remains audit/reconciliation-only." in doc_text
    assert "`runtime_economic_mode` remains the only explicit runtime staging path." in doc_text
    assert "`G20` remains `BLOCKED`." in doc_text
    assert "`R99/R102` remain `NOT APPROVED`." in doc_text

    assert "Runtime Summary,Project IRR" in mapping_text
    assert "OPEX,Runtime total OPEX" in mapping_text
    assert "Governance,text contains G20 remains BLOCKED and R99/R102 remain NOT APPROVED" in mapping_text
    assert "Senior Debt,only aggregate debt coverage metric is validated today" in gaps_text
    assert "validated_cells_not_formula_driven,confirmed" in matrix_text
    assert "missing_unavailable_markers_not_zero_filled,confirmed" in matrix_text
    assert "workbook_descriptive_only,confirmed" in matrix_text
