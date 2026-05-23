"""Tests for the Phase 10 reconciliation IRR reporting view."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
IRR_SUMMARY = REPORTS / "phase10_irr_reconciliation_summary.csv"
IRR_CONVENTIONS = REPORTS / "phase10_irr_convention_map.csv"
IRR_GOVERNANCE = REPORTS / "phase10_irr_governance_items.csv"
DOC = DOCS / "phase10_reconciliation_irr_reporting_view.md"


def _generate() -> None:
    write_calibration_reconciliation_pack()


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def test_irr_reports_exist():
    _generate()
    assert IRR_SUMMARY.exists()
    assert IRR_CONVENTIONS.exists()
    assert IRR_GOVERNANCE.exists()


def test_irr_reconciliation_sheet_exists():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "IRR Reconciliation" in wb.sheetnames


def test_runtime_and_excel_irr_values_are_surfaced():
    _generate()
    rows = _csv_rows(IRR_SUMMARY)
    by_metric = {row["metric"]: row for row in rows}
    assert float(by_metric["Project IRR"]["runtime_value"]) > 0.0
    assert float(by_metric["Project IRR"]["excel_value"]) > 0.0
    assert float(by_metric["Equity IRR"]["runtime_value"]) > 0.0
    assert float(by_metric["Equity IRR"]["excel_value"]) > 0.0


def test_irr_reports_have_required_columns():
    _generate()
    rows = _csv_rows(IRR_SUMMARY)
    required = {
        "metric",
        "runtime_value",
        "excel_value",
        "delta",
        "classification",
        "root_cause",
        "governance_sensitive",
        "stakeholder_decision_required",
        "notes",
    }
    assert rows
    assert required.issubset(rows[0].keys())


def test_executive_dashboard_contains_irr_summary():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Executive Dashboard")
    assert "IRR parity summary" in text
    assert "IRR convention summary" in text


def test_gap_register_contains_irr_rows():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Gap Register")
    assert "IRR Reporting" in text
    assert "XIRR Date Convention" in text
    assert "Reconciliation IRR" in text


def test_runtime_vs_excel_notes_exist():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "IRR Reconciliation")
    assert "Runtime vs Excel timing note" in text
    assert "distribution-definition" in text or "distribution" in text.lower()


def test_governance_posture_is_preserved():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance_text = _sheet_text(wb, "Governance")
    assert "G20 status BLOCKED" in governance_text
    assert "R99/R102 status NOT APPROVED" in governance_text


def test_doc_states_no_runtime_changes_and_governance_limits():
    content = DOC.read_text(encoding="utf-8")
    assert "no runtime changes statement" in content.lower()
    assert "G20 remains `BLOCKED`" in content
    assert "R99/R102 remain `NOT APPROVED`" in content
    assert "xirr convention discussion" in content.lower()


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
