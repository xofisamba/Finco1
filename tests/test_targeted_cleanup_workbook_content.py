"""Targeted cleanup numeric workbook-content coverage."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.export.runtime_summary import build_runtime_summary_rows


def _value_for_label(sheet, label: str):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return sheet.cell(row=row, column=2).value
    raise AssertionError(f"Label {label!r} not found in sheet {sheet.title!r}")


def _runtime_value(rows: list[dict[str, str]], metric: str) -> float:
    row = next(item for item in rows if item["metric"] == metric)
    return float(row["value"])


def test_runtime_summary_numeric_cells_match_backend_runtime_rows():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=True)
    runtime_rows = build_runtime_summary_rows("tuho")

    runtime_summary_sheet = workbook["Runtime Summary"]
    opex_sheet = workbook["OPEX"]

    workbook_project_irr = float(_value_for_label(runtime_summary_sheet, "Project IRR"))
    workbook_total_revenue = float(_value_for_label(runtime_summary_sheet, "Total revenue"))
    workbook_total_opex = float(_value_for_label(opex_sheet, "Runtime total OPEX"))

    expected_project_irr = _runtime_value(runtime_rows, "project_irr")
    expected_total_revenue = _runtime_value(runtime_rows, "total_revenue_keur")
    expected_total_opex = _runtime_value(runtime_rows, "total_opex_keur")

    assert abs(workbook_project_irr - expected_project_irr) < 1e-9
    assert abs(workbook_total_revenue - expected_total_revenue) < 1e-6
    assert abs(workbook_total_opex - expected_total_opex) < 1e-6
