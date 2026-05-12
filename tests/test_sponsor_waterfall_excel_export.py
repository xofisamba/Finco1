"""Phase 7C-5 — Tests for sponsor waterfall Excel export helpers.

Tests:
- export works with no waterfall results (no-op)
- export works with waterfall results (all 3 sheet types)
- sheet names deterministic
- row counts match result entries
- exported totals reconcile with domain result totals
- no export-side recalculation
- no mutation of input results
"""
from __future__ import annotations

import tempfile
import pandas as pd
import pytest
from openpyxl import load_workbook, Workbook

from domain.sponsor.sponsor_waterfall_tier import TierType, CompoundingConvention, SponsorShare, SponsorWaterfallTier
from domain.sponsor.preferred_return_result import (
    PreferredReturnAccrualEntry,
    PreferredReturnResult,
)
from domain.sponsor.waterfall_allocation_result import (
    TierAllocationEntry,
    PeriodWaterfallResult,
    WaterfallAllocationResult,
)
from domain.sponsor.capital_account_tier_annotation import (
    TierAnnotation,
    TierAnnotatedCapitalAccountEntry,
    TierAnnotatedSponsorCapitalAccount,
    from_waterfall_allocation_result,
)
from domain.sponsor.sponsor_capital_account import CapitalAccountEntry
from app.sponsor_waterfall_excel_export import (
    write_sponsor_waterfall_audit_sheets,
    build_waterfall_allocation_table,
    build_preferred_return_table,
    build_tier_capital_account_table,
)


# ── Fixtures ─────────────────────────────────────────────────────────────────

def make_waterfall_result(periods=3, sponsor_codes=("SPONSOR-1", "LP-1")) -> WaterfallAllocationResult:
    """Build a simple waterfall result for testing."""
    period_results = []
    for p in range(periods):
        per_sp = 5000.0 / len(sponsor_codes)
        tier_entries = [
            TierAllocationEntry(
                tier_index=0,
                tier_type=TierType.RETURN_OF_CAPITAL,
                available_cash_before_tier_keur=5000.0,
                allocated_amount_keur=5000.0,
                allocated_per_sponsor_keur=tuple((s, per_sp) for s in sponsor_codes),
                remaining_cash_after_tier_keur=0.0,
            ),
        ]
        period_results.append(
            PeriodWaterfallResult(
                period_index=p,
                available_cash_keur=5000.0,
                tier_entries=tuple(tier_entries),
                total_allocated_keur=5000.0,
                total_remaining_cash_keur=0.0,
                cumulative_distributions_by_sponsor_keur=tuple(
                    (s, (p + 1) * per_sp) for s in sponsor_codes
                ),
            )
        )
    return WaterfallAllocationResult(
        investor_id="WATERFALL",
        period_results=tuple(period_results),
        total_allocated_keur=5000.0 * periods,
        total_distributions_by_sponsor_keur=tuple(
            (s, periods * per_sp) for s in sponsor_codes
        ),
    )


def make_pref_result() -> PreferredReturnResult:
    entries = []
    for p in range(3):
        entries.append(
            PreferredReturnAccrualEntry(
                period_index=p,
                opening_invested_capital_keur=10000.0,
                invested_capital_delta_keur=0.0,
                hurdle_rate_pa=0.08,
                compounding_convention="SIMPLE",
                accrual_periods=0.5,
                accrued_pref_keur=400.0,
                cumulative_accrued_pref_keur=(p + 1) * 400.0,
                cumulative_distributions_keur=0.0,
                unpaid_pref_balance_keur=(p + 1) * 400.0,
                hurdle_satisfied=False,
            )
        )
    return PreferredReturnResult(
        sponsor_code="SPONSOR-1",
        hurdle_rate_pa=0.08,
        compounding_convention="SIMPLE",
        entries=tuple(entries),
        total_accrued_pref_keur=1200.0,
        total_unpaid_pref_keur=1200.0,
        all_periods_hurdle_satisfied=False,
    )


def _make_tmp_wb():
    """Create a minimal Workbook with one sheet to host audit sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Existing"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


# ── Test: build table functions ──────────────────────────────────────────────

class TestBuildWaterfallAllocationTable:
    def test_row_count_matches_entries(self):
        result = make_waterfall_result(periods=3)
        df = build_waterfall_allocation_table(result)
        # 3 periods × 1 tier = 3 rows
        assert len(df) == 3

    def test_columns_present(self):
        result = make_waterfall_result(periods=1)
        df = build_waterfall_allocation_table(result)
        assert "period_index" in df.columns
        assert "tier_index" in df.columns
        assert "tier_type" in df.columns
        assert "available_cash_before_tier_keur" in df.columns
        assert "allocated_amount_keur" in df.columns

    def test_sponsor_columns_present(self):
        result = make_waterfall_result(periods=1, sponsor_codes=("SPONSOR-1", "LP-1"))
        df = build_waterfall_allocation_table(result)
        assert "alloc_SPONSOR-1_keur" in df.columns
        assert "alloc_LP-1_keur" in df.columns

    def test_tier_type_as_string(self):
        result = make_waterfall_result(periods=1)
        df = build_waterfall_allocation_table(result)
        assert df["tier_type"].iloc[0] == "RETURN_OF_CAPITAL"

    def test_totals_reconcile(self):
        result = make_waterfall_result(periods=2, sponsor_codes=("SPONSOR-1",))
        df = build_waterfall_allocation_table(result)
        # Total allocated across all periods/tiers
        assert df["allocated_amount_keur"].sum() == pytest.approx(result.total_allocated_keur)

    def test_cumulative_distributions_present(self):
        result = make_waterfall_result(periods=2, sponsor_codes=("SPONSOR-1",))
        df = build_waterfall_allocation_table(result)
        assert "cumdist_SPONSOR-1_keur" in df.columns


class TestBuildPreferredReturnTable:
    def test_row_count_matches_entries(self):
        result = make_pref_result()
        df = build_preferred_return_table(result)
        assert len(df) == 3

    def test_columns_present(self):
        result = make_pref_result()
        df = build_preferred_return_table(result)
        assert "period_index" in df.columns
        assert "accrued_pref_keur" in df.columns
        assert "unpaid_pref_balance_keur" in df.columns
        assert "hurdle_satisfied" in df.columns

    def test_accrued_pref_values(self):
        result = make_pref_result()
        df = build_preferred_return_table(result)
        assert df["accrued_pref_keur"].iloc[0] == pytest.approx(400.0)
        assert df["cumulative_accrued_pref_keur"].iloc[2] == pytest.approx(1200.0)


class TestBuildTierCapitalAccountTable:
    def test_row_count_from_accounts(self):
        wf = make_waterfall_result(periods=2, sponsor_codes=("SPONSOR-1",))
        cap = from_waterfall_allocation_result(wf, "SPONSOR-1")
        df = build_tier_capital_account_table((cap,))
        assert len(df) == 2  # 2 periods

    def test_columns_present(self):
        wf = make_waterfall_result(periods=1, sponsor_codes=("SPONSOR-1",))
        cap = from_waterfall_allocation_result(wf, "SPONSOR-1")
        df = build_tier_capital_account_table((cap,))
        assert "investor_id" in df.columns
        assert "period_index" in df.columns
        assert "amount_keur" in df.columns
        assert "tier_index" in df.columns
        assert "tier_type" in df.columns
        assert "tier_source_note" in df.columns

    def test_tier_type_string(self):
        wf = make_waterfall_result(periods=1, sponsor_codes=("SPONSOR-1",))
        cap = from_waterfall_allocation_result(wf, "SPONSOR-1")
        df = build_tier_capital_account_table((cap,))
        assert df["tier_type"].iloc[0] == "RETURN_OF_CAPITAL"


# ── Test: write_sponsor_waterfall_audit_sheets ─────────────────────────────

class TestWriteSheetsNoOp:
    def test_no_results_no_sheets(self):
        """No waterfall/pref/capital results → no new sheets created."""
        tmp = _make_tmp_wb()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer)

        wb = load_workbook(tmp)
        # Only the original "Existing" sheet should remain
        assert wb.sheetnames == ["Existing"]

    def test_no_waterfall_result_no_waterfall_sheet(self):
        """waterfall_result=None → no Sponsor Waterfall Allocation sheet."""
        tmp = _make_tmp_wb()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(
                writer,
                waterfall_result=None,
                preferred_return_result=make_pref_result(),
            )

        wb = load_workbook(tmp)
        assert "Sponsor Waterfall Allocation" not in wb.sheetnames


class TestWriteSheetsBasic:
    def test_waterfall_sheet_created(self):
        tmp = _make_tmp_wb()
        result = make_waterfall_result(periods=2, sponsor_codes=("SPONSOR-1",))
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=result)

        wb = load_workbook(tmp)
        assert "Sponsor Waterfall Allocation" in wb.sheetnames

    def test_preferred_return_sheet_created(self):
        tmp = _make_tmp_wb()
        result = make_pref_result()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, preferred_return_result=result)

        wb = load_workbook(tmp)
        assert "Preferred Return Accrual" in wb.sheetnames

    def test_tier_capital_account_sheet_created(self):
        tmp = _make_tmp_wb()
        wf = make_waterfall_result(periods=1, sponsor_codes=("SPONSOR-1",))
        cap = from_waterfall_allocation_result(wf, "SPONSOR-1")
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, tier_annotated_accounts=(cap,))

        wb = load_workbook(tmp)
        assert "Tier Capital Account" in wb.sheetnames

    def test_audit_note_in_row_1(self):
        tmp = _make_tmp_wb()
        result = make_waterfall_result(periods=1)
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=result)

        wb = load_workbook(tmp)
        ws = wb["Sponsor Waterfall Allocation"]
        assert "AUDIT-ONLY" in str(ws.cell(1, 1).value)

    def test_headers_in_row_2(self):
        tmp = _make_tmp_wb()
        result = make_waterfall_result(periods=1)
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=result)

        wb = load_workbook(tmp)
        ws = wb["Sponsor Waterfall Allocation"]
        # Row 2 should have column headers
        assert ws.cell(2, 1).value == "period_index"


class TestSheetRowCounts:
    def test_waterfall_row_count(self):
        tmp = _make_tmp_wb()
        result = make_waterfall_result(periods=3)  # 3 rows expected
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=result)

        wb = load_workbook(tmp)
        ws = wb["Sponsor Waterfall Allocation"]
        # Rows: 1=audit note, 2=headers, 3..5 = data (3 periods)
        assert ws.max_row == 5

    def test_preferred_return_row_count(self):
        tmp = _make_tmp_wb()
        result = make_pref_result()  # 3 entries
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, preferred_return_result=result)

        wb = load_workbook(tmp)
        ws = wb["Preferred Return Accrual"]
        # Rows: 1=audit note, 2=headers, 3..5 = 3 data rows
        assert ws.max_row == 5

    def test_tier_capital_account_row_count(self):
        tmp = _make_tmp_wb()
        wf = make_waterfall_result(periods=2, sponsor_codes=("SPONSOR-1",))
        cap = from_waterfall_allocation_result(wf, "SPONSOR-1")
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, tier_annotated_accounts=(cap,))

        wb = load_workbook(tmp)
        ws = wb["Tier Capital Account"]
        # Rows: 1=audit note, 2=headers, 3..4 = 2 data rows
        assert ws.max_row == 4


class TestTotalsReconciliation:
    def test_waterfall_allocated_total_matches_result(self):
        result = make_waterfall_result(periods=2)
        df = build_waterfall_allocation_table(result)
        assert df["allocated_amount_keur"].sum() == pytest.approx(result.total_allocated_keur)

    def test_preferred_return_total_matches_result(self):
        result = make_pref_result()
        df = build_preferred_return_table(result)
        assert df["accrued_pref_keur"].sum() == pytest.approx(result.total_accrued_pref_keur)


class TestNoMutation:
    def test_input_result_not_mutated(self):
        result = make_waterfall_result(periods=2)
        original = str(result)
        tmp = _make_tmp_wb()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=result)
        assert str(result) == original

    def test_input_pref_result_not_mutated(self):
        result = make_pref_result()
        original = str(result)
        tmp = _make_tmp_wb()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, preferred_return_result=result)
        assert str(result) == original


class TestAllTierTypes:
    def test_multi_tier_waterfall_export(self):
        """Export handles multiple tier types in same period."""
        tier_entries = [
            TierAllocationEntry(
                tier_index=0,
                tier_type=TierType.RETURN_OF_CAPITAL,
                available_cash_before_tier_keur=5000.0,
                allocated_amount_keur=5000.0,
                allocated_per_sponsor_keur=(("SPONSOR-1", 5000.0),),
                remaining_cash_after_tier_keur=0.0,
            ),
            TierAllocationEntry(
                tier_index=1,
                tier_type=TierType.PREFERRED_RETURN,
                available_cash_before_tier_keur=0.0,
                allocated_amount_keur=0.0,
                allocated_per_sponsor_keur=(("SPONSOR-1", 0.0),),
                remaining_cash_after_tier_keur=0.0,
            ),
        ]
        period_result = PeriodWaterfallResult(
            period_index=0,
            available_cash_keur=5000.0,
            tier_entries=tuple(tier_entries),
            total_allocated_keur=5000.0,
            total_remaining_cash_keur=0.0,
            cumulative_distributions_by_sponsor_keur=(("SPONSOR-1", 5000.0),),
        )
        wf_result = WaterfallAllocationResult(
            investor_id="WATERFALL",
            period_results=(period_result,),
            total_allocated_keur=5000.0,
            total_distributions_by_sponsor_keur=(("SPONSOR-1", 5000.0),),
        )
        tmp = _make_tmp_wb()
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a") as writer:
            write_sponsor_waterfall_audit_sheets(writer, waterfall_result=wf_result)

        wb = load_workbook(tmp)
        ws = wb["Sponsor Waterfall Allocation"]
        # Row 3: tier 0, Row 4: tier 1
        assert ws.cell(3, 2).value == 0   # tier_index col
        assert ws.cell(4, 2).value == 1   # tier_index col
        # Tier types
        tier_type_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(2, col).value == "tier_type":
                tier_type_col = col
                break
        assert tier_type_col is not None
        assert "RETURN_OF_CAPITAL" in str(ws.cell(3, tier_type_col).value)
        assert "PREFERRED_RETURN" in str(ws.cell(4, tier_type_col).value)

class TestCumulativeDistributionColumns:
    def test_cumdist_column_for_sponsor_only_in_early_period(self):
        """A sponsor appearing only in an earlier period still gets a cumdist_ column."""
        # SPONSOR-1 appears in period 0 and 1; LP-1 only in period 0
        per_sp = 5000.0 / 2  # 2500 each
        period0_entries = [
            TierAllocationEntry(
                tier_index=0,
                tier_type=TierType.RETURN_OF_CAPITAL,
                available_cash_before_tier_keur=5000.0,
                allocated_amount_keur=5000.0,
                allocated_per_sponsor_keur=(("SPONSOR-1", per_sp), ("LP-1", per_sp)),
                remaining_cash_after_tier_keur=0.0,
            ),
        ]
        period1_entries = [
            TierAllocationEntry(
                tier_index=0,
                tier_type=TierType.RETURN_OF_CAPITAL,
                available_cash_before_tier_keur=5000.0,
                allocated_amount_keur=5000.0,
                allocated_per_sponsor_keur=(("SPONSOR-1", 5000.0),),  # LP-1 NOT here
                remaining_cash_after_tier_keur=0.0,
            ),
        ]
        period_results = [
            PeriodWaterfallResult(
                period_index=0,
                available_cash_keur=5000.0,
                tier_entries=tuple(period0_entries),
                total_allocated_keur=5000.0,
                total_remaining_cash_keur=0.0,
                cumulative_distributions_by_sponsor_keur=(
                    ("SPONSOR-1", per_sp), ("LP-1", per_sp)
                ),
            ),
            PeriodWaterfallResult(
                period_index=1,
                available_cash_keur=5000.0,
                tier_entries=tuple(period1_entries),
                total_allocated_keur=5000.0,
                total_remaining_cash_keur=0.0,
                cumulative_distributions_by_sponsor_keur=(
                    ("SPONSOR-1", per_sp + 5000.0),  # LP-1 drops out
                ),
            ),
        ]
        result = WaterfallAllocationResult(
            investor_id="WATERFALL",
            period_results=tuple(period_results),
            total_allocated_keur=10000.0,
            total_distributions_by_sponsor_keur=(("SPONSOR-1", 7500.0), ("LP-1", 2500.0)),
        )
        df = build_waterfall_allocation_table(result)
        # LP-1 only appears in period 0 but its cumdist column must still exist
        assert "cumdist_LP-1_keur" in df.columns
        # And it should have values (0.0 before period 0, per_sp in period 0, then unchanged)
        lp_cum = df["cumdist_LP-1_keur"].tolist()
        # Period 0 has LP-1 cumdist = 2500
        assert 2500.0 in lp_cum
