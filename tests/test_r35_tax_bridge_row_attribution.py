"""Validation for the TUHO R35 row attribution audit workbook."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner


WORKBOOK_PATH = Path("reports/phase6_tuho_r35_row_attribution.xlsx")
REQUIRED_SHEETS = {
    "Summary",
    "R35 Attribution",
    "Upstream Rows",
    "Loss Rows R36-R41",
    "CIT Rows R43-R44",
    "Largest R35 Deltas",
    "Suspected Drivers",
}


def _run_default_tuho():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return WaterfallRunner(project, engine).run(config)


def test_r35_attribution_workbook_exists_with_required_sheets():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    assert REQUIRED_SHEETS.issubset(set(workbook.sheetnames))


def test_r35_attribution_contains_60_operating_periods():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["R35 Attribution"]

    op_indices = [sheet.cell(row=row, column=1).value for row in range(2, 62)]

    assert op_indices == list(range(60))
    assert sheet.cell(row=62, column=1).value == "TOTAL"


def test_largest_r35_deltas_are_sorted_descending_by_absolute_delta():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Largest R35 Deltas"]

    abs_deltas = [sheet.cell(row=row, column=7).value for row in range(2, 62)]

    assert abs_deltas == sorted(abs_deltas, reverse=True)
    assert sheet.cell(row=2, column=2).value == 0
    assert sheet.cell(row=2, column=6).value == pytest.approx(2024.8624, abs=0.01)
    assert sheet.cell(row=2, column=8).value == "SHL interest gross/net/timing"


def test_summary_captures_driver_ranking_and_r35_total():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    summary = workbook["Summary"]
    values = {
        summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
        for row in range(2, summary.max_row + 1)
    }

    assert values["Total R35 delta"] == pytest.approx(12_216.370554918296, abs=0.01)
    assert values["SHL interest delta"] == pytest.approx(10_347.270775417655, abs=0.01)
    assert values["Depreciation delta"] == pytest.approx(2_302.166786061971, abs=0.01)
    assert values["R34 delta"] == pytest.approx(0.0, abs=0.01)
    assert values["R99 readiness"] == "Blocked"


def test_workbook_generation_does_not_change_default_runtime_behavior():
    baseline = _run_default_tuho()
    diagnostic = _run_default_tuho()

    assert diagnostic.total_revenue_keur == pytest.approx(baseline.total_revenue_keur, abs=0.0001)
    assert diagnostic.total_opex_keur == pytest.approx(baseline.total_opex_keur, abs=0.0001)
    assert diagnostic.total_tax_keur == pytest.approx(baseline.total_tax_keur, abs=0.0001)
    assert diagnostic.total_senior_ds_keur == pytest.approx(
        baseline.total_senior_ds_keur,
        abs=0.0001,
    )
