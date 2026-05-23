"""Tests for the Phase 10 calibration reconciliation pack."""

from __future__ import annotations

import csv
import os
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
GAP_REGISTER = REPORTS / "phase10_calibration_gap_register.csv"
SOURCE_INVENTORY = REPORTS / "phase10_calibration_source_inventory.csv"
SUMMARY = REPORTS / "phase10_calibration_reconciliation_summary.csv"
NAVIGATION_MAP = REPORTS / "phase10_review_pack_navigation_map.csv"
SIGNOFF_MATRIX = REPORTS / "phase10_review_signoff_matrix.csv"
DOC = DOCS / "phase10_calibration_gap_reconciliation_pack.md"
NAV_DOC = DOCS / "phase10_review_pack_polish_and_lender_navigation.md"
EXEC_DOC = DOCS / "phase10_executive_dashboard_and_signoff_flow.md"


def _generate_if_missing() -> None:
    if all(path.exists() for path in (WORKBOOK, GAP_REGISTER, SOURCE_INVENTORY, SUMMARY, NAVIGATION_MAP, SIGNOFF_MATRIX)):
        return
    write_calibration_reconciliation_pack()


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def test_workbook_generates():
    write_calibration_reconciliation_pack()
    assert WORKBOOK.exists()


def test_required_sheets_exist():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    expected = [
        "Navigation",
        "Executive Dashboard",
        "Executive Summary",
        "Review Signoff",
        "Governance",
        "Governance Timeline",
        "Readiness Matrix",
        "Runtime Summary",
        "Revenue Reconciliation",
        "OPEX Reconciliation",
        "Senior Debt Reconciliation",
        "SHL Reconciliation",
        "Tax R35-R67-R69",
        "CFADS Waterfall",
        "Distributions Sponsor",
        "Returns Reconciliation",
        "Gap Register",
        "Source Inventory",
        "Accepted Conventions",
        "Reviewer Notes",
    ]
    assert wb.sheetnames == expected


def test_navigation_sheet_exists_and_has_hyperlinks():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Navigation"]
    hyperlinks = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                hyperlinks.append(str(cell.hyperlink.target))
    assert any("Executive Dashboard" in link for link in hyperlinks)
    assert any("Gap Register" in link for link in hyperlinks)


def test_horizontal_periods_exist():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Revenue Reconciliation"]
    headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
    assert "P1" in headers
    assert "P61" in headers


def test_classification_and_root_cause_columns_exist():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Revenue Reconciliation"]
    headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
    assert "Classification" in headers
    assert "Root Cause" in headers


def test_severity_legend_exists():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Navigation")
    assert "Status Legend" in text
    assert "ACCEPTED_CONVENTION" in text
    assert "GOVERNANCE_BLOCKER" in text


def test_non_pass_rows_contain_explanations():
    _generate_if_missing()
    rows = _csv_rows(SUMMARY)
    non_pass = [row for row in rows if row["classification"] != "PASS"]
    assert non_pass
    for row in non_pass:
        assert row["root_cause"], row
        assert row["recommended_action"], row


def test_source_inventory_exists_and_has_required_columns():
    _generate_if_missing()
    rows = _csv_rows(SOURCE_INVENTORY)
    assert rows
    required = {
        "metric",
        "source_file",
        "period_level_available",
        "total_available",
        "evidence_side",
        "confidence_level",
        "usable_for_reconciliation",
        "notes",
    }
    assert required.issubset(rows[0].keys())


def test_governance_sheet_contains_gate_status():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Governance")
    assert "G20 status BLOCKED" in text
    assert "R99/R102 status NOT APPROVED" in text
    assert "Governance Warning" in text


def test_runtime_vs_preview_labels_exist():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Runtime Summary") + " " + _sheet_text(wb, "Governance")
    assert "Runtime / Preview" in text
    assert "review" in text.lower() or "runtime" in text.lower()


def test_executive_summary_contains_classification_counts():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Executive Summary")
    assert "Classification Counts" in text
    assert "Top 10 Material Gaps" in text
    assert "What Is Engineering vs Governance" in text


def test_executive_dashboard_exists_and_contains_counts():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Executive Dashboard")
    assert "Classification Counts" in text
    assert "Top 5 Material Gaps" in text
    assert "Top Governance Blockers" in text
    assert "Review recommendation summary" in text


def test_missing_evidence_rows_are_not_zero_filled():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Revenue Reconciliation"]
    target_row = None
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label == "CO2 Revenue - Excel":
            target_row = row
            break
    assert target_row is not None
    first_period_value = ws.cell(row=target_row, column=3).value
    assert isinstance(first_period_value, str)
    assert first_period_value.startswith("MISSING_EVIDENCE")


def test_gap_register_contains_reviewer_columns():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Gap Register"]
    headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
    assert "Owner" in headers
    assert "Requires Stakeholder Decision?" in headers
    assert "Requires Runtime Change?" in headers
    assert "Expected Roadmap Phase" in headers


def test_reviewer_notes_expanded():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Reviewer Notes")
    assert "Already runtime-verified" in text
    assert "Governance-only blockers" in text
    assert "Roadmap view" in text
    assert "IC reviewer focus" in text
    assert "Lender reviewer focus" in text
    assert "Audit reviewer focus" in text


def test_review_signoff_contains_workflow_statuses():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Review Signoff")
    assert "review_status" in text
    assert "GOVERNANCE_PENDING" in text or "READY_FOR_SIGNOFF" in text or "IN_REVIEW" in text
    rows = _csv_rows(SIGNOFF_MATRIX)
    statuses = {row["current_status"] for row in rows}
    assert "GOVERNANCE_PENDING" in statuses or "IN_REVIEW" in statuses


def test_governance_timeline_and_readiness_matrix_exist():
    _generate_if_missing()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    timeline_text = _sheet_text(wb, "Governance Timeline")
    readiness_text = _sheet_text(wb, "Readiness Matrix")
    assert "Phase 10 export foundation" in timeline_text
    assert "review_area" in readiness_text
    assert "runtime_complete" in readiness_text
    assert "governance_status" in readiness_text


def test_doc_states_governance_limits():
    content = DOC.read_text(encoding="utf-8")
    nav_content = NAV_DOC.read_text(encoding="utf-8")
    exec_content = EXEC_DOC.read_text(encoding="utf-8")
    assert "G20 remains `BLOCKED`" in content
    assert "R99/R102 remain `NOT APPROVED`" in content
    assert "no runtime formula changes" in content
    assert "navigation philosophy" in nav_content.lower()
    assert "governance workflow philosophy" in exec_content.lower()
    assert "no runtime changes statement" in exec_content.lower()


def test_navigation_map_report_exists_and_has_required_columns():
    _generate_if_missing()
    rows = _csv_rows(NAVIGATION_MAP)
    assert rows
    required = {
        "sheet_name",
        "reviewer_role",
        "primary_purpose",
        "navigation_priority",
        "governance_sensitive",
        "notes",
    }
    assert required.issubset(rows[0].keys())


def test_signoff_matrix_report_exists_and_has_required_columns():
    _generate_if_missing()
    rows = _csv_rows(SIGNOFF_MATRIX)
    assert rows
    required = {
        "review_area",
        "reviewer_type",
        "current_status",
        "runtime_ready",
        "evidence_ready",
        "governance_ready",
        "blocker_type",
        "recommended_next_step",
    }
    assert required.issubset(rows[0].keys())


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
