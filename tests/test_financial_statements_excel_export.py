"""Excel audit export tests for offline financial statements."""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner
from domain.financial_statements import (
    assemble_financial_statements,
    export_financial_statements_audit_workbook,
)


def _run_tuho():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return project, WaterfallRunner(project, engine).run(config)


def _waterfall_snapshot(result):
    return tuple(
        (
            period.revenue_keur,
            period.opex_keur,
            period.tax_keur,
            period.senior_ds_keur,
            period.shl_service_keur,
            period.distribution_keur,
            period.r99_fcf_for_distribution_keur,
            period.r102_fcf_for_shl_keur,
        )
        for period in result.periods
    )


def _statements_snapshot(statements):
    return {
        "pnl_periods": tuple(asdict(period) for period in statements.pnl.periods),
        "tax_periods": tuple(asdict(period) for period in statements.tax_bridge.periods),
        "bs_periods": tuple(asdict(period) for period in statements.balance_sheet.periods),
        "cf_periods": tuple(asdict(period) for period in statements.pf_cash_waterfall.periods),
    }


def _column_values(sheet, column: int) -> list[object]:
    return [sheet.cell(row=row, column=column).value for row in range(1, sheet.max_row + 1)]


def _output_path(filename: str) -> Path:
    output_dir = Path(__file__).parents[1] / "test_output_dir"
    output_dir.mkdir(exist_ok=True)
    return output_dir / filename


def test_financial_statements_excel_audit_workbook_is_created_and_readable():
    _, waterfall = _run_tuho()
    statements = assemble_financial_statements(waterfall)
    output_path = _output_path("financial_statements_audit.xlsx")

    returned_path = export_financial_statements_audit_workbook(
        statements,
        output_path,
        project_name="TUHO-WIND-1",
    )

    assert returned_path == output_path
    assert output_path.exists()
    workbook = load_workbook(output_path, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "P&L",
        "Balance Sheet",
        "PF Cash Waterfall",
        "Source Mapping",
        "Known Gaps",
    ]


def test_summary_contains_key_totals():
    _, waterfall = _run_tuho()
    statements = assemble_financial_statements(waterfall)
    output_path = export_financial_statements_audit_workbook(
        statements,
        _output_path("summary.xlsx"),
        project_name="TUHO-WIND-1",
    )
    sheet = load_workbook(output_path, data_only=True)["Summary"]

    labels = _column_values(sheet, 1)
    assert "Project name" in labels
    assert "Total revenue" in labels
    assert "Total OPEX" in labels
    assert "Total EBITDA" in labels
    assert "Total senior DS" in labels
    assert "Total cash tax" in labels
    assert "Total SHL service" in labels
    assert "Total distributions" in labels
    assert "Max BS balance check" in labels
    assert "R99/R102 identity status" in labels


def test_statement_sheets_contain_required_rows():
    _, waterfall = _run_tuho()
    statements = assemble_financial_statements(waterfall)
    output_path = export_financial_statements_audit_workbook(
        statements,
        _output_path("rows.xlsx"),
    )
    workbook = load_workbook(output_path, data_only=True)

    pnl_labels = set(_column_values(workbook["P&L"], 2))
    assert {"Revenues", "Operating expenses", "EBIT", "Net income", "Net dividends"} <= pnl_labels

    balance_sheet_labels = set(_column_values(workbook["Balance Sheet"], 2))
    assert "Balance Check" in balance_sheet_labels
    assert "Cash" in balance_sheet_labels

    cf_codes = set(_column_values(workbook["PF Cash Waterfall"], 1))
    assert {"CF_R99", "CF_R102"} <= cf_codes


def test_source_mapping_and_known_gaps_show_placeholder_and_residual_statuses():
    _, waterfall = _run_tuho()
    statements = assemble_financial_statements(waterfall)
    output_path = export_financial_statements_audit_workbook(
        statements,
        _output_path("source_mapping.xlsx"),
    )
    workbook = load_workbook(output_path, data_only=True)

    source_statuses = set(_column_values(workbook["Source Mapping"], 6))
    assert "placeholder" in source_statuses
    assert "residual" in source_statuses
    assert "audit-only" in source_statuses
    assert "runtime output" in source_statuses

    gaps = set(_column_values(workbook["Known Gaps"], 1))
    assert "gross fixed assets" not in gaps
    assert "Gross fixed assets" in gaps
    assert "R99/R102" in gaps


def test_export_does_not_mutate_statements_or_waterfall():
    _, waterfall = _run_tuho()
    waterfall_before = _waterfall_snapshot(waterfall)
    statements = assemble_financial_statements(waterfall)
    statements_before = _statements_snapshot(statements)

    export_financial_statements_audit_workbook(
        statements,
        _output_path("mutation_guard.xlsx"),
        project_name="TUHO-WIND-1",
    )

    assert _waterfall_snapshot(waterfall) == waterfall_before
    assert _statements_snapshot(statements) == statements_before
