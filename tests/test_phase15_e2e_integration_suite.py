from __future__ import annotations

import os
import sys
import uuid
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.api.project_runner import run_project
from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.input_adapter import build_projectinputs
from app.input_schema import DebtInput, OpexInput, ProjectInputsSchema, RevenueInput
from app.persistence.repository import (
    bind_workspace_to_scenario,
    compare_scenarios,
    get_workspace_state,
    list_exports,
    record_export,
    record_workspace_runtime,
    runtime_guard_for_snapshot,
    save_project,
    save_run,
    save_scenario,
    save_workspace_state,
)


def _governance_state() -> dict[str, str]:
    return {"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"}


def _default_snapshot(project_code: str) -> dict[str, str]:
    code = (project_code or "tuho").strip().lower()
    return {
        "active_project": code,
        "project_type": "Solar" if code == "oborovo" else "Wind",
        "scenario": "Base",
        "capacity_mw": "",
        "tariff_eur_mwh": "",
        "p50_hours": "",
        "total_capex_keur": "",
        "opex_y1_keur": "",
        "gearing_pct": "",
        "target_dscr": "",
        "interest_rate_pct": "",
        "tenor_years": "",
        "cod_date": "",
        "construction_months": "",
        "horizon_years": "",
        "capacity_factor": "",
        "ppa_term_years": "",
    }


def _runtime_result_for_snapshot(snapshot: dict[str, str]) -> dict:
    def _float(value: str) -> float | None:
        return None if value in ("", None) else float(value)

    def _int(value: str) -> int | None:
        return None if value in ("", None) else int(float(value))

    schema = ProjectInputsSchema(
        project_type=snapshot["project_type"],
        scenario=snapshot["scenario"],
        capacity_mw=_float(snapshot.get("capacity_mw", "")),
        revenue=RevenueInput(
            tariff_eur_mwh=_float(snapshot.get("tariff_eur_mwh", "")),
            p50_hours=_float(snapshot.get("p50_hours", "")),
        ),
        opex=OpexInput(
            opex_y1_keur=_float(snapshot.get("opex_y1_keur", "")),
        ),
        debt=DebtInput(
            gearing_pct=_float(snapshot.get("gearing_pct", "")),
            target_dscr=_float(snapshot.get("target_dscr", "")),
            interest_rate_pct=_float(snapshot.get("interest_rate_pct", "")),
            tenor_years=_int(snapshot.get("tenor_years", "")),
        ),
    )
    override = build_projectinputs(schema)
    project_key = "TUHO" if snapshot["active_project"] == "tuho" else "Oborovo"
    return run_project(project_key, snapshot["scenario"], project_inputs_override=override)


def _compare_display(value) -> str:
    if value in (None, "", "NOT_AVAILABLE"):
        return "pending / unavailable"
    return str(value)


def _compare_delta_display(value) -> str:
    if value is None:
        return "not_applicable"
    return str(value)


@pytest.fixture
def test_db():
    import app.persistence.db as db_mod

    old_path = os.environ.get("FINCO_DB_PATH")
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_file = os.path.join(base_dir, f"phase15_e2e_{uuid.uuid4().hex[:8]}.db")
    os.environ["FINCO_DB_PATH"] = db_file
    db_mod.DB_PATH = db_file

    yield db_file

    if os.path.exists(db_file):
        os.remove(db_file)
    if old_path:
        os.environ["FINCO_DB_PATH"] = old_path
        db_mod.DB_PATH = old_path
    else:
        os.environ.pop("FINCO_DB_PATH", None)


def test_guided_internal_pilot_workflow_end_to_end(test_db):
    user_id = f"u{uuid.uuid4().hex[:8]}"
    governance = _governance_state()
    project = save_project(user_id, "tuho", "TUHO Wind 1", "tuho", governance_state=governance)

    baseline_snapshot = _default_snapshot("tuho")
    baseline_snapshot["tariff_eur_mwh"] = "72"
    baseline_snapshot["opex_y1_keur"] = "2100"
    baseline_snapshot["target_dscr"] = "1.15"
    baseline_snapshot["interest_rate_pct"] = "4.8"

    baseline_saved = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Baseline",
        project_code="tuho",
        source_project_template="tuho",
        snapshot=baseline_snapshot,
        governance_state=governance,
        replay_metadata={"runtime_economic_mode": False, "audit_economic_mode": False},
    )
    workspace = bind_workspace_to_scenario(user_id, project.project_id, "tuho", baseline_saved, governance_state=governance)
    assert workspace.dirty is False
    assert workspace.active_scenario_id == baseline_saved.scenario_id

    edited_snapshot = dict(workspace.saved_snapshot)
    edited_snapshot["tariff_eur_mwh"] = "88"
    edited_snapshot["opex_y1_keur"] = "2450"
    edited_snapshot["target_dscr"] = "1.23"
    dirty_workspace = save_workspace_state(
        user_id=user_id,
        project_id=project.project_id,
        project_code="tuho",
        active_scenario_id=workspace.active_scenario_id,
        active_scenario_name=workspace.active_scenario_name,
        draft_snapshot=edited_snapshot,
        saved_snapshot=workspace.saved_snapshot,
        dirty=True,
        governance_state=governance,
        replay_metadata={"draft_only": True},
    )

    assert dirty_workspace.dirty is True
    allow_run, runtime_origin, guard_message = runtime_guard_for_snapshot(dirty_workspace, edited_snapshot)
    assert allow_run is False
    assert runtime_origin == "preview_only"
    assert "Unsaved edits" in guard_message
    assert dirty_workspace.last_runtime_snapshot == {}

    saved_after_edit = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Pilot Edited",
        project_code="tuho",
        source_project_template="tuho",
        snapshot=edited_snapshot,
        governance_state=governance,
        replay_metadata={"runtime_economic_mode": False, "audit_economic_mode": False},
    )
    rebound = bind_workspace_to_scenario(user_id, project.project_id, "tuho", saved_after_edit, governance_state=governance)

    assert rebound.dirty is False
    assert rebound.active_scenario_id == saved_after_edit.scenario_id
    assert rebound.last_runtime_snapshot == {}
    assert rebound.last_runtime_snapshot_id is None

    allow_run, runtime_origin, guard_message = runtime_guard_for_snapshot(rebound, rebound.saved_snapshot)
    assert allow_run is True
    assert runtime_origin == "saved_state"
    assert guard_message == ""

    runtime_result = _runtime_result_for_snapshot(rebound.saved_snapshot)
    runtime_snapshot_id = f"phase15-run-{uuid.uuid4().hex[:8]}"
    runtime_state = record_workspace_runtime(
        user_id=user_id,
        project_id=project.project_id,
        project_code="tuho",
        runtime_snapshot=dict(rebound.saved_snapshot),
        runtime_summary=runtime_result["kpis"],
        runtime_snapshot_id=runtime_snapshot_id,
        runtime_origin="saved_state",
        governance_state=governance,
        active_scenario_id=saved_after_edit.scenario_id,
        active_scenario_name=saved_after_edit.scenario_name,
        replay_metadata={"runtime_economic_mode": False, "audit_economic_mode": False},
    )
    run_record = save_run(
        user_id=user_id,
        project_type="TUHO",
        scenario=rebound.saved_snapshot["scenario"],
        inputs=dict(rebound.saved_snapshot),
        kpis=runtime_result["kpis"],
        replay_metadata={
            "runtime_snapshot_id": runtime_snapshot_id,
            "runtime_economic_mode": False,
            "audit_economic_mode": False,
        },
    )

    assert runtime_state.last_runtime_snapshot_id == runtime_snapshot_id
    assert runtime_state.last_runtime_origin == "saved_state"
    assert runtime_state.last_runtime_snapshot["tariff_eur_mwh"] == "88"
    assert run_record.replay_metadata["runtime_snapshot_id"] == runtime_snapshot_id
    assert run_record.replay_metadata["runtime_economic_mode"] is False
    assert run_record.replay_metadata["audit_economic_mode"] is False

    workbook_dir = Path(os.path.dirname(test_db)) / f"phase15_export_{uuid.uuid4().hex[:8]}"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbook_dir / "phase15_runtime_workbook.xlsx"
    workbook_path.write_bytes(export_institutional_workbook_skeleton("tuho"))
    workbook = load_workbook(BytesIO(workbook_path.read_bytes()), data_only=True)
    assert "Runtime Summary" in workbook.sheetnames
    assert "Cover" in workbook.sheetnames

    export_record = record_export(
        user_id=user_id,
        project_code="tuho",
        export_type="institutional_workbook",
        artifact_name=workbook_path.name,
        artifact_path=str(workbook_path),
        project_id=project.project_id,
        scenario_id=saved_after_edit.scenario_id,
        governance_state=governance,
        runtime_snapshot_id=runtime_snapshot_id,
        replay_metadata={
            "runtime_authority": "backend_only",
            "runtime_snapshot_id": runtime_snapshot_id,
            "runtime_economic_mode": False,
            "audit_economic_mode": False,
        },
    )
    exports = list_exports(user_id, project_id=project.project_id)
    assert any(item.export_id == export_record.export_id for item in exports)
    assert export_record.runtime_snapshot_id == runtime_snapshot_id
    assert export_record.governance_state["g20_status"] == "BLOCKED"
    assert export_record.governance_state["r99_r102_status"] == "NOT APPROVED"
    assert export_record.replay_metadata["runtime_authority"] == "backend_only"
    assert export_record.replay_metadata["runtime_economic_mode"] is False
    assert export_record.replay_metadata["audit_economic_mode"] is False

    second_snapshot = dict(baseline_snapshot)
    second_snapshot["tariff_eur_mwh"] = "70"
    second_snapshot["opex_y1_keur"] = "2200"
    second_saved = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Compare Target",
        project_code="tuho",
        source_project_template="tuho",
        snapshot=second_snapshot,
        governance_state=governance,
        last_run_summary={"total_revenue_keur": 115000.0, "project_irr": 0.089, "equity_irr": None, "avg_dscr": 1.24},
        replay_metadata={"runtime_economic_mode": False, "audit_economic_mode": False},
    )

    compare_result = compare_scenarios(user_id, second_saved.scenario_id, saved_after_edit.scenario_id)
    assert compare_result is not None
    assert compare_result["left"].scenario_id == second_saved.scenario_id
    assert compare_result["right"].scenario_id == saved_after_edit.scenario_id

    revenue_row = next(row for row in compare_result["metrics"] if row["metric"] == "Revenue")
    equity_irr_row = next(row for row in compare_result["metrics"] if row["metric"] == "Equity IRR")
    compare_display_left = _compare_display(equity_irr_row["left_value"])
    compare_display_right = _compare_display(equity_irr_row["right_value"])
    compare_delta = _compare_delta_display(equity_irr_row["delta"])

    assert revenue_row["left_value"] == 115000.0
    assert compare_display_left == "pending / unavailable"
    assert compare_display_right == "pending / unavailable"
    assert compare_delta == "not_applicable"

    refreshed = get_workspace_state(user_id, project.project_id)
    assert refreshed is not None
    assert refreshed.dirty is False
    assert refreshed.last_runtime_snapshot_id == runtime_snapshot_id
    assert refreshed.last_runtime_origin == "saved_state"

    assert governance["g20_status"] == "BLOCKED"
    assert governance["r99_r102_status"] == "NOT APPROVED"

    workbook_path.unlink(missing_ok=True)
    workbook_dir.rmdir()


def test_phase15_docs_and_reports_capture_guardrails():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    doc = Path(base) / "docs" / "phase15_e2e_integration_suite.md"
    workflow = Path(base) / "reports" / "phase15_e2e_workflow_matrix.csv"
    authority = Path(base) / "reports" / "phase15_e2e_authority_boundary_matrix.csv"
    gaps = Path(base) / "reports" / "phase15_e2e_remaining_gaps.csv"

    assert doc.exists()
    assert workflow.exists()
    assert authority.exists()
    assert gaps.exists()

    doc_text = doc.read_text(encoding="utf-8")
    workflow_text = workflow.read_text(encoding="utf-8")
    authority_text = authority.read_text(encoding="utf-8")
    gaps_text = gaps.read_text(encoding="utf-8")

    assert "save scenario → run model → export workbook → compare scenarios" in doc_text
    assert "Save creates a persisted scenario boundary but does not run the model." in doc_text
    assert "Export does not auto-run the model." in doc_text
    assert "Compare does not auto-save and does not auto-run." in doc_text
    assert "`audit_economic_mode` remains audit/reconciliation-only." in doc_text
    assert "`runtime_economic_mode` remains the only explicit runtime staging path." in doc_text
    assert "`G20` remains `BLOCKED`." in doc_text
    assert "`R99/R102` remain `NOT APPROVED`." in doc_text

    assert "compare_scenarios,yes" in workflow_text
    assert "save_does_not_auto_run,confirmed" in authority_text
    assert "export_does_not_auto_run,confirmed" in authority_text
    assert "compare_does_not_auto_save_or_run,confirmed" in authority_text
    assert "pending_values_not_zero_filled,confirmed" in authority_text
    assert "browser_automation," in gaps_text
    assert "external_model_review," in gaps_text
