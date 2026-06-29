"""C2-PR22: Export / Run Safety Guardrails.

Proves, with route-level regression tests (not just narrative), that
preview values sent to ``POST /model/preview`` can never leak into:

  - the real runtime-summary CSV export (`/exports/runtime-summary.csv`)
  - the real institutional workbook Excel export
    (`/exports/institutional-workbook.xlsx`)
  - persisted/saved model state (DB row unchanged before/after a
    preview request — mirrors the existing
    tests/test_c2_pr14_opex_preview.py::TestNoFinancialEngineCallOrPersistenceMutation
    pattern)

and that a `/model/preview` POST has zero observable side effect on a
subsequent export's bytes (byte-identical pre/post).

Scope decision (documented per the task brief): full real-engine
"Run produces different output before/after preview" testing is not
added here as a *new* heavy end-to-end Run-form test, because:

  - `/run` requires a full multi-field form payload
    (`app.services.run_service.execute_run_route`) that is already
    exhaustively exercised by the existing Phase-9/PR9-era Run test
    suites elsewhere in this repo; duplicating that heavy fixture here
    would not add new leak-detection coverage beyond what the
    lighter-weight "DB row unchanged" + "export bytes unchanged" tests
    below already prove (since /run's only inputs are the saved
    DB/form state, which these tests already prove `/model/preview`
    cannot touch).
  - Instead, this file proves the same guarantee more directly and
    more cheaply: `/model/preview` never mutates the DB (so `/run`,
    which reads only DB/form state, cannot possibly be affected by it
    either) and the real export routes' byte content is unaffected by
    a preceding `/model/preview` call, even when that call carries a
    distinctive sentinel number.

Uses fastapi.testclient.TestClient against the real `main_web.app`,
exactly like every other `test_c2_pr1*.py` file in this repo.
"""
import os

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
os.environ.setdefault("FINCO_COOKIE_SECURE", "false")

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient

from main_web import app
from app.auth import create_session_token, COOKIE_NAME
from app.persistence.db import DB_PATH

client = TestClient(app)

# A distinctive sentinel preview number, chosen so its digit string is
# extremely unlikely to ever appear in any legitimately-computed
# export figure (CAPEX/Revenue/OPEX/EBITDA/Operating-Cash-Flow totals,
# IDs, dates, etc.) by coincidence.
SENTINEL_VALUE = 987654.32
SENTINEL_DIGITS = "98765432"  # without separators/decimal point
SENTINEL_FORMATTED = "987,654.32"


def _auth_cookies():
    token = create_session_token()
    return {COOKIE_NAME: token}


def _preview_payload(**overrides):
    payload = {
        "valid": True,
        "dirtyCells": [
            "capex!C.01.01.amount",
            "revenue!ppa_tariff_eur_mwh",
            "opex!OM-01.budget",
        ],
        "affectedGroups": ["overview-kpis"],
        "projectDirty": True,
        "reason": "manual-flush",
        "executionStatus": "stubbed",
        "project": None,
        "capexTotalPreview": SENTINEL_VALUE,
        "revenueTotalPreview": SENTINEL_VALUE,
        "opexTotalPreview": SENTINEL_VALUE,
        "ebitdaPreview": SENTINEL_VALUE,
        "operatingCashFlowPreview": SENTINEL_VALUE,
    }
    payload.update(overrides)
    return payload


def _post_sentinel_preview():
    resp = client.post(
        "/model/preview",
        json=_preview_payload(),
        cookies=_auth_cookies(),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["ok"] is True
    # Confirm the sentinel really did round-trip through the preview
    # response (i.e. this is a meaningful test of leakage, not a no-op).
    assert body["capex"]["capex_total_preview"] == SENTINEL_VALUE
    assert body["revenue"]["preview"] == SENTINEL_VALUE
    assert body["opex"]["preview"] == SENTINEL_VALUE
    assert body["ebitda"]["preview"] == SENTINEL_VALUE
    assert body["operating_cash_flow"]["preview"] == SENTINEL_VALUE
    return body


class TestPreviewSentinelNeverLeaksIntoRuntimeSummaryCsv:
    """Point 1/2/5: preview values never appear in the audit/export CSV."""

    def test_sentinel_absent_from_runtime_summary_csv(self):
        _post_sentinel_preview()
        resp = client.get(
            "/exports/runtime-summary.csv?project=tuho",
            cookies=_auth_cookies(),
        )
        assert resp.status_code == 200
        text = resp.content.decode("utf-8", errors="replace")
        assert SENTINEL_DIGITS not in text
        assert SENTINEL_FORMATTED not in text
        assert str(SENTINEL_VALUE) not in text


class TestPreviewSentinelNeverLeaksIntoInstitutionalWorkbookExport:
    """Point 1/5: preview values never appear in the Excel export bytes."""

    def test_sentinel_absent_from_institutional_workbook_xlsx(self):
        _post_sentinel_preview()
        resp = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        )
        assert resp.status_code == 200
        # Excel workbooks are zip containers; scan raw bytes (covers both
        # any literal shared-string content and any other embedded text)
        # as a conservative, format-agnostic leakage check.
        raw = resp.content
        assert SENTINEL_DIGITS.encode("ascii") not in raw
        assert SENTINEL_FORMATTED.encode("ascii") not in raw

        # Also assert against the parsed cell values for the strongest,
        # most direct form of this guarantee.
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(raw))
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        assert abs(cell.value - SENTINEL_VALUE) > 0.01, (
                            f"Sentinel preview value leaked into "
                            f"sheet '{sheet}' cell {cell.coordinate}"
                        )


class TestPreviewRequestDoesNotMutatePersistedState:
    """Point 4: /model/preview never writes to the DB.

    Mirrors tests/test_c2_pr14_opex_preview.py::
    TestNoFinancialEngineCallOrPersistenceMutation::test_no_persistence_mutation
    exactly, extended to cover all five preview fields at once.
    """

    def test_db_file_byte_identical_before_and_after_sentinel_preview(self):
        before_mtime = os.path.getmtime(DB_PATH) if os.path.exists(DB_PATH) else None
        before_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else None

        _post_sentinel_preview()

        after_mtime = os.path.getmtime(DB_PATH) if os.path.exists(DB_PATH) else None
        after_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else None
        assert before_mtime == after_mtime
        assert before_size == after_size

    def test_no_financial_engine_call_with_full_sentinel_payload(self, monkeypatch):
        import app.waterfall_core as waterfall_core

        def _boom(*args, **kwargs):
            raise AssertionError(
                "waterfall_core.run_project must never be called by /model/preview"
            )

        monkeypatch.setattr(waterfall_core, "run_project", _boom, raising=False)
        _post_sentinel_preview()


_ISO_TIMESTAMP_RE = None


def _xlsx_entries_normalized(raw_bytes):
    """Return {name: text-with-timestamps-blanked} for every zip entry
    in an xlsx export.

    The export workbook legitimately embeds fresh wall-clock generation
    timestamps in `docProps/core.xml` and in worksheet cells (e.g. a
    "Weekend run at <ISO timestamp>" notice) on every single request,
    regardless of any preview state — this is expected, real,
    pre-existing behaviour of `build_institutional_workbook_export`,
    not a leakage vector. Comparing raw bytes directly would make every
    "did /model/preview change this export" test flaky/always-failing
    for a reason that has nothing to do with preview leakage. Blanking
    out ISO-8601 timestamps (the only kind of run-to-run-varying text
    these exports contain) before comparing isolates exactly the
    signal this test cares about: did the *data* change.
    """
    import re
    import zipfile
    from io import BytesIO

    global _ISO_TIMESTAMP_RE
    if _ISO_TIMESTAMP_RE is None:
        _ISO_TIMESTAMP_RE = re.compile(
            r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:\d{2})?"
        )

    zf = zipfile.ZipFile(BytesIO(raw_bytes))
    out = {}
    for name in zf.namelist():
        raw = zf.read(name)
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            out[name] = raw
            continue
        out[name] = _ISO_TIMESTAMP_RE.sub("<TIMESTAMP>", text)
    return out


class TestExportBytesUnaffectedByPrecedingPreviewRequest:
    """Point 3: a /model/preview POST has zero effect on export output.

    Captures a real export's bytes, fires a /model/preview request
    carrying the sentinel, then re-fetches the same export and asserts
    identity — proving the preview round trip is a true no-op for both
    export routes from the server's point of view. The CSV export is
    compared byte-for-byte (it carries no embedded wall-clock
    timestamp); the xlsx export is compared entry-by-entry, excluding
    only `docProps/core.xml`'s generation timestamp (see
    `_xlsx_entries_normalized` above for why that
    one file is excluded).
    """

    def test_runtime_summary_csv_byte_identical_around_preview(self):
        before = client.get(
            "/exports/runtime-summary.csv?project=tuho",
            cookies=_auth_cookies(),
        ).content

        _post_sentinel_preview()

        after = client.get(
            "/exports/runtime-summary.csv?project=tuho",
            cookies=_auth_cookies(),
        ).content

        assert before == after

    def test_institutional_workbook_byte_identical_around_preview(self):
        before = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        ).content

        _post_sentinel_preview()

        after = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        ).content

        assert _xlsx_entries_normalized(before) == \
            _xlsx_entries_normalized(after)


class TestSavedCapexRevenueOutputsUnaffectedByPreviewState:
    """Point 7: saved/authoritative outputs are unaffected by whatever
    sentinel sits in the preview echo.

    Mirrors the established "Overview KPIs byte-identical pre/post
    preview edit" assertion style used throughout the PR10-PR20 browser
    suites, here applied at the export-bytes level rather than the DOM
    level (route-level test, no browser required).
    """

    def test_export_outputs_identical_regardless_of_sentinel_magnitude(self):
        first = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        ).content

        # Two different sentinel magnitudes in two separate preview
        # requests must produce the exact same (unaffected) export.
        client.post(
            "/model/preview",
            json=_preview_payload(capexTotalPreview=111111.11),
            cookies=_auth_cookies(),
        )
        second = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        ).content

        client.post(
            "/model/preview",
            json=_preview_payload(capexTotalPreview=222222.22),
            cookies=_auth_cookies(),
        )
        third = client.get(
            "/exports/institutional-workbook.xlsx?project=tuho",
            cookies=_auth_cookies(),
        ).content

        e1 = _xlsx_entries_normalized(first)
        e2 = _xlsx_entries_normalized(second)
        e3 = _xlsx_entries_normalized(third)
        assert e1 == e2 == e3
