from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.excel_export import build_excel_export
from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.persistence.provenance import build_replay_metadata
from app.ui_runner import run_demo_project


ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "phase14_reviewer_cover_notes.md"
NOTE_MATRIX = ROOT / "reports" / "phase14_reviewer_cover_note_matrix.csv"
INTERPRETATION_REPORT = ROOT / "reports" / "phase14_workbook_interpretation_notes.csv"
NO_CLAIMS_REPORT = ROOT / "reports" / "phase14_cover_note_no_claims_matrix.csv"
PHASE9_DOC = ROOT / "docs" / "phase9_audit_economic_mode_contract_reconciliation.md"


def _sheet_text(sheet) -> str:
    parts: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return "\n".join(parts)


def _first_numeric_values(sheet, limit: int = 8) -> list[float]:
    values: list[float] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                values.append(float(cell.value))
                if len(values) >= limit:
                    return values
    raise AssertionError(f"No numeric values found in sheet {sheet.title!r}")


def test_institutional_workbook_includes_reviewer_cover_notes():
    workbook = load_workbook(BytesIO(export_institutional_workbook_skeleton("tuho")), data_only=True)
    cover_text = _sheet_text(workbook["Cover"])
    governance_text = _sheet_text(workbook["Governance"])

    assert "financial source of truth" in cover_text
    assert "do not mutate the artifact already exported" in cover_text
    assert "accepted conventions" in cover_text.lower()
    assert "G20 remains BLOCKED" in cover_text or "G20 remains BLOCKED" in governance_text
    assert "R99/R102 remain NOT APPROVED" in cover_text or "R99/R102 remain NOT APPROVED" in governance_text
    assert "unavailable" in governance_text
    assert "not_applicable" in governance_text
    assert "not lender-ready without external model review" in governance_text.lower()
    assert "not audit-certified" in governance_text.lower()


def test_values_only_excel_notes_include_reviewer_cover_notes_without_numeric_drift():
    demo = run_demo_project("Solar")
    provenance_metadata = build_replay_metadata(
        project_key="oborovo",
        project_inputs=demo.project_inputs,
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"},
        export_type="excel_model_export",
        workbook_type="values_only_excel_export",
        active_project="oborovo",
        scenario_name="Base",
        runtime_origin="factory_base_runtime",
    )
    plain_wb = load_workbook(
        BytesIO(build_excel_export(result=demo.result, project_inputs=demo.project_inputs)),
        data_only=True,
    )
    noted_wb = load_workbook(
        BytesIO(build_excel_export(result=demo.result, project_inputs=demo.project_inputs, provenance_metadata=provenance_metadata)),
        data_only=True,
    )

    notes_text = _sheet_text(noted_wb["Notes"])
    assert "Reviewer Cover Notes" in notes_text
    assert "financial source of truth" in notes_text
    assert "saved/exported snapshot context" in notes_text
    assert "G20 remains BLOCKED" in notes_text
    assert "R99/R102 remain NOT APPROVED" in notes_text
    assert "must not be read as zero" in notes_text
    assert "not lender-ready without external model review" in notes_text.lower()
    assert "not audit-certified" in notes_text.lower()

    assert _first_numeric_values(plain_wb["Dashboard"]) == _first_numeric_values(noted_wb["Dashboard"])
    assert _first_numeric_values(plain_wb["Revenue"]) == _first_numeric_values(noted_wb["Revenue"])


def test_reports_and_docs_exist_with_guardrail_statements():
    assert DOC.exists()
    assert NOTE_MATRIX.exists()
    assert INTERPRETATION_REPORT.exists()
    assert NO_CLAIMS_REPORT.exists()

    text = DOC.read_text(encoding="utf-8")
    phase9_text = PHASE9_DOC.read_text(encoding="utf-8")
    assert "No runtime/model formulas were changed" in text
    assert "No workbook calculations were changed" in text
    assert "No persistence authority promotion occurred" in text
    assert "No replay engine behavior was added" in text
    assert "`G20` remains `BLOCKED`" in text
    assert "`R99/R102` remain `NOT APPROVED`" in text
    assert "`audit_economic_mode` remains audit/reconciliation-only." in text
    assert "`runtime_economic_mode` remains the only explicit runtime staging path." in text
    assert "Conclusion" in phase9_text
