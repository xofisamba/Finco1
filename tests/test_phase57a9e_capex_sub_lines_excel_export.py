from __future__ import annotations

import os
import sqlite3
import subprocess
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
try:
    from fastapi.testclient import TestClient
except ModuleNotFoundError:  # pragma: no cover - environment-dependent
    TestClient = None

from app.auth import COOKIE_NAME, create_session_token

REPO_ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture
def test_db(monkeypatch):
    temp_root = REPO_ROOT / "app" / "data"
    temp_root.mkdir(exist_ok=True)
    db_file = temp_root / f"phase57a9e_{uuid.uuid4().hex[:8]}.db"
    monkeypatch.setenv("FINCO_DB_PATH", str(db_file))
    import app.persistence.db as db_mod

    monkeypatch.setattr(db_mod, "DB_PATH", str(db_file))
    db_mod.init_db()
    try:
        yield db_file
    finally:
        for suffix in ("", "-wal", "-shm"):
            try:
                (Path(str(db_file) + suffix)).unlink(missing_ok=True)
            except PermissionError:
                pass


@pytest.fixture
def client(test_db):
    if TestClient is None:
        pytest.skip("FastAPI/TestClient not installed in this environment")
    import main_web

    tc = TestClient(main_web.app)
    tc.cookies.set(COOKIE_NAME, create_session_token())
    return tc


def _load_wb(excel_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)


def _sheet_rows_by_business_code(ws) -> list[dict[str, object]]:
    header_row_idx = None
    headers = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        values = list(row)
        if "Business Code" in values:
            header_row_idx = values.index("Business Code")
            headers = values
            break
    assert headers is not None, "Business Code header row not found"

    rows = []
    found_header = False
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not found_header:
            if values == list(headers):
                found_header = True
            continue
        if all(v in (None, "") for v in values):
            continue
        if values[0] == "Note":
            continue
        record = {
            str(headers[idx]): values[idx] if idx < len(values) else None
            for idx in range(len(headers))
            if headers[idx] is not None
        }
        if record.get("Business Code") in (None, ""):
            continue
        rows.append(record)
    return rows


def _notes_dict(ws) -> dict[str, str]:
    notes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = row[1] if len(row) > 1 else None
        if key == "Parent Category":
            break
        if key:
            notes[str(key)] = "" if value is None else str(value)
    return notes


def _mapping_rows(ws) -> list[tuple[str, str]]:
    found = False
    out = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if values[:2] == ["Parent Category", "Mapped Capex Field"]:
            found = True
            continue
        if not found:
            continue
        if "Business Code" in values:
            break
        if values[0] and values[1]:
            out.append((str(values[0]), str(values[1])))
    return out


def _make_user_project_with_sub_lines(user_id: str = "u57a9e"):
    from app.persistence.repository import create_project_record

    return create_project_record(
        user_id=user_id,
        project_code="pilot-capex-audit",
        project_name="Capex Audit Project",
        project_type="Solar",
        project_origin="user_project",
        template_source="generic_solar",
        baseline_snapshot=_snapshot(),
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
    )


def _create_sub_line(project_id: str, *, parent: str, label: str, amount: float):
    from app.persistence.capex_sub_lines import create_sub_line

    with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
        conn.row_factory = sqlite3.Row
        sub = create_sub_line(
            conn.cursor(),
            project_id=project_id,
            parent_category_code=parent,
            label=label,
            amount_keur=amount,
            comments=f"{label} comment",
        )
        conn.commit()
        return sub


def _soft_delete(project_id: str, sub_line_id: str):
    from app.persistence.capex_sub_lines import soft_delete_sub_line

    with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
        conn.row_factory = sqlite3.Row
        assert soft_delete_sub_line(
            conn.cursor(),
            project_id=project_id,
            sub_line_id=sub_line_id,
        )
        conn.commit()


def _make_non_base_scenario(project, user_id: str, overrides: dict):
    from app.persistence.repository import add_scenario, seed_scenarios_if_needed

    baseline = _snapshot()
    base = seed_scenarios_if_needed(
        user_id=user_id,
        project_id=project.project_id,
        project_code=project.project_code,
        project_type="Solar",
        source_project_template="generic_solar",
        baseline_snapshot=baseline,
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
        template_origin="generic_solar",
    )
    scenario = add_scenario(
        user_id=user_id,
        project_id=project.project_id,
        project_code=project.project_code,
        scenario_name="Capex Override",
        parent_scenario_id=base.scenario_id,
        base_input_set=base.base_input_set,
        overrides=overrides,
    )
    return base, scenario


def _snapshot(**overrides):
    data = {
        "active_project": "pilot-capex-audit",
        "project_name": "Capex Audit Project",
        "project_type": "Solar",
        "project_origin": "user_created",
        "template_source": "generic_solar",
        "country_market": "Croatia",
        "scenario": "Base",
        "capacity_mw": "50",
        "cod_date": "2027-01-01",
        "construction_months": "12",
        "horizon_years": "25",
        "tariff_eur_mwh": "60",
        "ppa_term_years": "15",
        "p50_hours": "1400",
        "opex_y1_keur": "1000",
        "total_capex_keur": "50000",
        "gearing_pct": "70",
        "interest_rate_pct": "5",
        "tenor_years": "15",
        "target_dscr": "1.30",
    }
    data.update(overrides)
    return data


class TestPhase57A9EExcelExport:
    def test_factory_tuho_export_unchanged(self):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        demo = run_demo_project("TUHO", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={},
            )
        )
        assert "CapEx_SubLines_Audit" not in wb.sheetnames

    def test_factory_oborovo_export_unchanged(self):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        demo = run_demo_project("Oborovo", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={},
            )
        )
        assert "CapEx_SubLines_Audit" not in wb.sheetnames

    def test_user_project_export_includes_cnn_u_rows(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        c02 = _create_sub_line(project.project_id, parent="C.02", label="Extra EPC", amount=1000.0)
        c08 = _create_sub_line(project.project_id, parent="C.08", label="Extra DD", amount=200.0)
        c11 = _create_sub_line(project.project_id, parent="C.11", label="Extra Legal", amount=300.0)

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        ws = wb["CapEx_SubLines_Audit"]
        rows = _sheet_rows_by_business_code(ws)
        codes = {row["Business Code"] for row in rows}
        assert c02.business_code in codes
        assert c08.business_code in codes
        assert c11.business_code in codes

    def test_soft_deleted_rows_excluded_from_default_export(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        keep = _create_sub_line(project.project_id, parent="C.02", label="Keep", amount=100.0)
        drop = _create_sub_line(project.project_id, parent="C.03", label="Drop", amount=200.0)
        _soft_delete(project.project_id, drop.sub_line_id)

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        codes = {row["Business Code"] for row in rows}
        assert keep.business_code in codes
        assert drop.business_code not in codes

    def test_scenario_override_amount_exported_as_effective_amount(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        sub = _create_sub_line(project.project_id, parent="C.02", label="Override EPC", amount=1000.0)
        _base, scenario = _make_non_base_scenario(
            project,
            "u57a9e",
            {"_capex_sub_line_overrides": {sub.sub_line_id: 1750.0}},
        )

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={
                    "project_id": project.project_id,
                    "active_scenario_id": scenario.scenario_id,
                    "active_scenario_name": scenario.scenario_name,
                },
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        row = next(r for r in rows if r["Business Code"] == sub.business_code)
        assert row["Default Amount (kEUR)"] == 1000
        assert row["Effective Amount (kEUR)"] == 1750
        assert row["Scenario Override Applied"] is True
        assert row["Scenario Override Amount (kEUR)"] == 1750

    def test_override_replaces_default_not_delta_in_export(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        sub = _create_sub_line(project.project_id, parent="C.02", label="Replace Me", amount=1200.0)
        _base, scenario = _make_non_base_scenario(
            project,
            "u57a9e",
            {"_capex_sub_line_overrides": {sub.sub_line_id: 300.0}},
        )

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={
                    "project_id": project.project_id,
                    "active_scenario_id": scenario.scenario_id,
                    "active_scenario_name": scenario.scenario_name,
                },
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        row = next(r for r in rows if r["Business Code"] == sub.business_code)
        assert row["Effective Amount (kEUR)"] == 300
        assert row["Effective Amount (kEUR)"] != 1500

    def test_c08_and_c11_mapping_visible_as_audit_legal(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        c08 = _create_sub_line(project.project_id, parent="C.08", label="DD", amount=150.0)
        c11 = _create_sub_line(project.project_id, parent="C.11", label="Legal", amount=250.0)

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        row08 = next(r for r in rows if r["Business Code"] == c08.business_code)
        row11 = next(r for r in rows if r["Business Code"] == c11.business_code)
        assert row08["Mapped Capex Field"] == "audit_legal"
        assert row11["Mapped Capex Field"] == "audit_legal"
        mapping = dict(_mapping_rows(wb["CapEx_SubLines_Audit"]))
        assert mapping["C.08"] == "audit_legal"
        assert mapping["C.11"] == "audit_legal"

    def test_no_c17_c18_rows_exported(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        _create_sub_line(project.project_id, parent="C.02", label="Allowed", amount=100.0)

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        assert not any(
            str(row["Business Code"]).startswith("C.17.") or str(row["Business Code"]).startswith("C.18.")
            for row in rows
        )

    def test_stale_override_does_not_crash_export(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project_with_sub_lines()
        _create_sub_line(project.project_id, parent="C.02", label="Live", amount=100.0)
        _base, scenario = _make_non_base_scenario(
            project,
            "u57a9e",
            {"_capex_sub_line_overrides": {"stale-uuid-1": 999.0}},
        )

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={
                    "project_id": project.project_id,
                    "active_scenario_id": scenario.scenario_id,
                    "active_scenario_name": scenario.scenario_name,
                },
            )
        )
        notes = _notes_dict(wb["CapEx_SubLines_Audit"])
        assert "stale-uuid-1" in notes["Stale Override Handling"]

    def test_download_route_user_project_export_includes_audit_sheet(self, client, test_db):
        from app.persistence.repository import bind_workspace_to_scenario

        user_id = "1"
        project = _make_user_project_with_sub_lines(user_id=user_id)
        _create_sub_line(project.project_id, parent="C.02", label="Extra EPC", amount=1000.0)
        base_case, _scenario = _make_non_base_scenario(project, user_id, {})
        bind_workspace_to_scenario(
            user_id,
            project.project_id,
            project.project_code,
            base_case,
            governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
        )

        response = client.post("/download", data=_snapshot(active_project=project.project_code))
        assert response.status_code == 200
        wb = _load_wb(response.content)
        assert "CapEx_SubLines_Audit" in wb.sheetnames
        notes = _notes_dict(wb["CapEx_SubLines_Audit"])
        assert notes["Project ID"] == project.project_id
        assert notes["Audit Mode"] == "active_only"

    def test_download_route_factory_exports_do_not_include_audit_sheet(self, client):
        tuho = client.post("/download", data={"project_type": "Wind", "scenario": "Base"})
        oborovo = client.post("/download", data={"project_type": "Solar", "scenario": "Base", "active_project": "oborovo"})

        assert tuho.status_code == 200
        assert oborovo.status_code == 200
        assert "CapEx_SubLines_Audit" not in _load_wb(tuho.content).sheetnames
        assert "CapEx_SubLines_Audit" not in _load_wb(oborovo.content).sheetnames

    def test_download_route_propagates_active_scenario_override_metadata(self, client, test_db):
        from app.persistence.repository import bind_workspace_to_scenario, list_exports, select_scenario

        user_id = "1"
        project = _make_user_project_with_sub_lines(user_id=user_id)
        sub = _create_sub_line(project.project_id, parent="C.02", label="Override EPC", amount=1000.0)
        base_case, scenario = _make_non_base_scenario(
            project,
            user_id,
            {"_capex_sub_line_overrides": {sub.sub_line_id: 1750.0}},
        )
        bind_workspace_to_scenario(
            user_id,
            project.project_id,
            project.project_code,
            base_case,
            governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
        )
        assert select_scenario(user_id, project.project_id, scenario.scenario_id) is True

        response = client.post("/download", data=_snapshot(active_project=project.project_code))
        assert response.status_code == 200

        wb = _load_wb(response.content)
        row = next(
            r for r in _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
            if r["Business Code"] == sub.business_code
        )
        assert row["Default Amount (kEUR)"] == 1000
        assert row["Effective Amount (kEUR)"] == 1750
        assert row["Scenario Override Applied"] is True

        export = list_exports(user_id, project_id=project.project_id, limit=1)[0]
        replay = export.replay_metadata
        assert replay["project_id"] == project.project_id
        assert replay["active_scenario_id"] == scenario.scenario_id
        assert replay["active_scenario_name"] == scenario.scenario_name
        assert replay["project_origin"] == project.project_origin
        assert replay["capex_sub_lines_audit_mode"] == "active_only"


class TestGuardrails:
    def test_no_run_path_changes(self):
        changed = _changed_files()
        assert "app/services/run_service.py" not in changed

    def test_no_ui_or_static_changes(self):
        changed = _changed_files()
        assert not any(path.startswith("app/templates/") for path in changed)
        assert "static/app.js" not in changed
        assert "static/styles.css" not in changed

    def test_no_schema_or_persistence_changes(self):
        changed = _changed_files()
        assert not any(path.startswith("app/persistence/") for path in changed)

    def test_rc1_untouched(self):
        result = subprocess.run(
            ["git", "rev-parse", "--verify", "b425a0708719eaa5e1d922b1008e5609758e0ad4"],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0


def _changed_files() -> set[str]:
    result = subprocess.run(
        ["git", "diff", "origin/main", "--name-only"],
        cwd=str(REPO_ROOT),
        capture_output=True,
        text=True,
    )
    return {line.strip() for line in result.stdout.splitlines() if line.strip()}
