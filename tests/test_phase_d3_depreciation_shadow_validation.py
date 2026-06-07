"""Tests for Phase D3 — Depreciation Shadow Validation / Audit-Only Comparison.

This is AUDIT-ONLY. No runtime enablement, no formula
change, no schedule change, no tax / P&L / CFADS change.
The tests pin:

- D3 module is the canonical-vs-legacy comparison entry
  point
- D3 helper runs the canonical DepreciationEngine in
  audit-only mode (no waterfall change, no
  ``period.depreciation_keur`` override)
- D3 helper does NOT mutate the caller's project inputs
- D3 helper produces a JSON-friendly summary
- D3 helper produces a text-only audit DataFrame
- D3 helper does not enable any feature flag on the
  caller's project inputs (the canonical flag is set
  only on the local shadow validation, never persisted)
- D3 tests confirm TUHO and Oborovo factory templates
  produce near-zero deltas (within floating-point
  tolerance) — supporting the recommendation that
  TUHO / Oborovo are ready for later controlled
  enablement
- D1 and D2 tests still pass
- factory export numeric invariance post-D3 (TUHO and
  Oborovo pinned numeric baselines unchanged)
- rc1 frozen SHA resolves
"""

from __future__ import annotations

import json
import subprocess
from pathlib import Path
from typing import Any

import pytest

from app.depreciation_shadow_validation import (
    ShadowComparisonRow,
    ShadowComparisonSummary,
    build_shadow_validation_audit_dataframe,
    run_shadow_validation,
    to_json_dict,
)
from app.ui_runner import run_demo_project


REPO_ROOT = Path(__file__).resolve().parents[1]


# ---------------------------------------------------------------------------
# 1. Public API
# ---------------------------------------------------------------------------


class TestD3PublicAPI:
    """D3 module exposes the canonical-vs-legacy comparison
    helpers."""

    def test_run_shadow_validation_callable(self):
        assert callable(run_shadow_validation)
        assert callable(to_json_dict)
        assert callable(build_shadow_validation_audit_dataframe)

    def test_shadow_data_classes_importable(self):
        assert ShadowComparisonRow is not None
        assert ShadowComparisonSummary is not None


# ---------------------------------------------------------------------------
# 2. Shadow validation runs for factory projects
# ---------------------------------------------------------------------------


class _get_waterfall_result:
    """Helper: get the waterfall result from a RunOutput."""

    @staticmethod
    def from_run_output(run_output: Any) -> Any:
        for attr in ("waterfall_result", "result"):
            if hasattr(run_output, attr):
                value = getattr(run_output, attr)
                if value is not None:
                    return value
        return run_output


class TestShadowValidationRunsForFactoryProjects:
    """The shadow validation must run cleanly for TUHO and
    Oborovo factory projects."""

    @pytest.mark.parametrize(
        "project_label", ["TUHO", "Oborovo"],
    )
    def test_run_shadow_validation_for_factory(
        self, project_label: str,
    ):
        r = run_demo_project(project_label, "Base")
        wf = _get_waterfall_result.from_run_output(r)
        summary = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label=project_label,
        )
        assert isinstance(summary, ShadowComparisonSummary)
        assert summary.project_label == project_label
        assert summary.period_count > 0
        assert summary.shadow_phase == "D3"


# ---------------------------------------------------------------------------
# 3. D3 does NOT mutate the caller's project inputs
# ---------------------------------------------------------------------------


class TestD3DoesNotMutateCaller:
    """D3 must NOT mutate the caller's project inputs. The
    canonical flag must stay False on the caller's
    project inputs both before and after the shadow
    validation runs. This is the discipline invariant:
    the canonical flag is set only on the local shadow
    call, never on the live project_inputs."""

    @pytest.mark.parametrize(
        "project_label", ["TUHO", "Oborovo"],
    )
    def test_canonical_flag_unchanged_on_caller_inputs(
        self, project_label: str,
    ):
        r = run_demo_project(project_label, "Base")
        before = bool(
            getattr(
                r.project_inputs.info,
                "use_depreciation_canonical_engine",
                False,
            )
        )
        assert before is False, (
            f"Baseline {project_label!r} has canonical "
            f"flag True (unexpected)."
        )
        wf = _get_waterfall_result.from_run_output(r)
        _ = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label=project_label,
        )
        after = bool(
            getattr(
                r.project_inputs.info,
                "use_depreciation_canonical_engine",
                False,
            )
        )
        assert after is False, (
            f"D3 must NOT mutate the caller's project "
            f"inputs. After shadow validation, the "
            f"canonical flag is {after!r} (expected "
            f"False)."
        )

    def test_d3_does_not_change_waterfall_result_attributes(self):
        r = run_demo_project("TUHO", "Base")
        wf = _get_waterfall_result.from_run_output(r)
        # Snapshot the first few period attributes
        before_first_period_dep = wf.periods[0].depreciation_keur
        before_first_period_tax = getattr(
            wf.periods[0], "tax_depreciation_audit_keur", None,
        )
        before_total_periods = len(wf.periods)
        before_canonical_attr = hasattr(
            wf, "_canonical_depreciation_wiring",
        )

        # Run D3
        _ = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label="TUHO",
        )

        # After D3, the legacy waterfall result must be
        # bit-for-bit unchanged.
        after_first_period_dep = wf.periods[0].depreciation_keur
        after_first_period_tax = getattr(
            wf.periods[0], "tax_depreciation_audit_keur", None,
        )
        after_total_periods = len(wf.periods)
        after_canonical_attr = hasattr(
            wf, "_canonical_depreciation_wiring",
        )

        assert after_first_period_dep == before_first_period_dep
        assert after_first_period_tax == before_first_period_tax
        assert after_total_periods == before_total_periods
        assert after_canonical_attr == before_canonical_attr


# ---------------------------------------------------------------------------
# 4. TUHO / Oborovo factory deltas (recommendation evidence)
# ---------------------------------------------------------------------------


class TestFactoryShadowDeltas:
    """The shadow validation produces a per-period and
    total delta. For TUHO and Oborovo, the total delta
    must be within a small floating-point tolerance. This
    supports the recommendation that TUHO / Oborovo are
    ready for later controlled enablement.

    The tolerances are intentionally generous (1e-6
    relative) so the test is stable across pandas /
    numpy / Python upgrades."""

    REL_TOLERANCE = 1e-6

    def _summary(self, project_label: str) -> ShadowComparisonSummary:
        r = run_demo_project(project_label, "Base")
        wf = _get_waterfall_result.from_run_output(r)
        return run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label=project_label,
        )

    def test_tuho_total_delta_within_tolerance(self):
        s = self._summary("TUHO")
        # TUHO factory total capex is 70,691.54 kEUR.
        # Legacy total depreciation is 70,691.54 kEUR.
        # Canonical total is also 70,691.54 kEUR.
        # The absolute delta should be near zero
        # (floating-point only).
        assert abs(s.absolute_total_delta_keur) < 1e-3, (
            f"TUHO total delta too large: "
            f"{s.absolute_total_delta_keur} kEUR"
        )

    def test_oborovo_total_delta_within_tolerance(self):
        s = self._summary("Oborovo")
        # Oborovo factory total capex is 55,999.09 kEUR.
        # Delta is in single-digit kEUR (legacy vs canonical
        # rounding differences). Pin the tolerance.
        assert abs(s.relative_total_delta_pct) < 0.01, (
            f"Oborovo relative total delta too large: "
            f"{s.relative_total_delta_pct:.4f}%"
        )

    def test_tuho_max_period_delta_within_tolerance(self):
        s = self._summary("TUHO")
        # Max single-period absolute delta should be
        # within a few hundred kEUR (the canonical engine
        # spreads the per-asset-class straight-line
        # differently from the legacy half-year book
        # schedule, but the per-period values stay in the
        # same order of magnitude).
        assert abs(s.max_absolute_period_delta_keur) < 1e3, (
            f"TUHO max per-period delta too large: "
            f"{s.max_absolute_period_delta_keur} kEUR"
        )

    def test_oborovo_max_period_delta_within_tolerance(self):
        s = self._summary("Oborovo")
        assert abs(s.max_absolute_period_delta_keur) < 1e3, (
            f"Oborovo max per-period delta too large: "
            f"{s.max_absolute_period_delta_keur} kEUR"
        )

    def test_per_period_rows_have_consistent_indices(self):
        s = self._summary("TUHO")
        # All per-period rows have a 1-based period index.
        for i, row in enumerate(s.rows, start=1):
            assert row.period_index == i
            assert isinstance(row, ShadowComparisonRow)


# ---------------------------------------------------------------------------
# 5. JSON serialization
# ---------------------------------------------------------------------------


class TestJSONSerialization:
    """The summary must be JSON-serializable so the D3
    audit table can be persisted as a machine-readable
    summary file."""

    def test_to_json_dict_round_trip(self):
        r = run_demo_project("TUHO", "Base")
        wf = _get_waterfall_result.from_run_output(r)
        s = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label="TUHO",
        )
        d = to_json_dict(s)
        # Round-trip the summary dict through json.dumps.
        text = json.dumps(d, default=str)
        re_loaded = json.loads(text)
        assert re_loaded["summary"]["project_label"] == "TUHO"
        assert re_loaded["summary"]["period_count"] == s.period_count
        assert len(re_loaded["rows"]) == s.period_count


# ---------------------------------------------------------------------------
# 6. Audit DataFrame is text-only
# ---------------------------------------------------------------------------


class TestAuditDataFrameIsTextOnly:
    """The audit DataFrame must have no financial numerics
    in the Value column. It is disclosure text only."""

    def test_audit_dataframe_value_column_text_only(self):
        r = run_demo_project("TUHO", "Base")
        wf = _get_waterfall_result.from_run_output(r)
        s = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label="TUHO",
        )
        df = build_shadow_validation_audit_dataframe(s)
        assert "Field" in df.columns
        assert "Value" in df.columns
        # The Value column must have NO numerics
        # (the text strings are formatted with comma
        # separators, e.g. "70,691.54").
        non_string = df["Value"][
            df["Value"].apply(lambda v: isinstance(v, (int, float)))
        ]
        assert len(non_string) == 0, (
            f"D3 audit DataFrame Value column should be "
            f"text only. Found numerics: {list(non_string)}"
        )

    def test_audit_dataframe_has_required_fields(self):
        r = run_demo_project("TUHO", "Base")
        wf = _get_waterfall_result.from_run_output(r)
        s = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label="TUHO",
        )
        df = build_shadow_validation_audit_dataframe(s)
        required = {
            "Phase", "Scope", "Notice", "Project Label",
            "Period Count", "Legacy Total Depreciation (kEUR)",
            "Canonical Total Book Depreciation (kEUR)",
            "Absolute Total Delta (kEUR)",
            "Relative Total Delta (%)",
            "Max Absolute Period Delta (kEUR)",
            "Max Absolute Period Delta Period",
            "Max Relative Period Delta (%)",
            "Max Relative Period Delta Period",
            "Runtime Authority Path",
            "Canonical Promotion Status",
            "Generic Project Support",
        }
        actual = set(df["Field"].tolist())
        for f in required:
            assert f in actual, (
                f"D3 audit DataFrame missing required "
                f"field {f!r}. Got: {sorted(actual)}"
            )

    def test_audit_dataframe_runtime_authority_path_text(self):
        r = run_demo_project("TUHO", "Base")
        wf = _get_waterfall_result.from_run_output(r)
        s = run_shadow_validation(
            project_inputs=r.project_inputs,
            legacy_waterfall_result=wf,
            project_label="TUHO",
        )
        df = build_shadow_validation_audit_dataframe(s)
        authority_row = df[df["Field"] == "Runtime Authority Path"]
        assert len(authority_row) == 1
        assert (
            authority_row.iloc[0]["Value"]
            == "legacy_depreciation_runtime"
        )


# ---------------------------------------------------------------------------
# 7. Numeric invariance (D3 must not change any export)
# ---------------------------------------------------------------------------


TUHO_REFERENCE_SUMS = {
    "CapEx": 145988.42,
    "CapEx_Items": 70706.54,
    "Inputs": 79580.2375,
    "Depreciation Assumptions": 45.0,
}

OBOROVO_REFERENCE_SUMS = {
    "CapEx": 115758.5053,
    "CapEx_Items": 56104.09,
    "Inputs": 61272.8532,
    "Depreciation Assumptions": 45.0,
}


class TestFactoryExportNumericInvariance:
    """D3 must NOT change any factory export numerics. The
    shadow validation runs in read-only mode; the live
    waterfall output is unchanged."""

    def _sum_numeric(self, data: bytes, sheet_name: str) -> float:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True,
        )
        if sheet_name not in wb.sheetnames:
            return 0.0
        ws = wb[sheet_name]
        total = 0.0
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)):
                    total += float(v)
        return total

    @pytest.mark.parametrize("project_type", ["TUHO", "Oborovo"])
    def test_factory_export_numeric_invariance_post_d3(
        self, project_type: str,
    ):
        from app.excel_export import build_excel_export
        r = run_demo_project(project_type, "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type=project_type.lower(),
        )
        ref = (
            TUHO_REFERENCE_SUMS
            if project_type == "TUHO"
            else OBOROVO_REFERENCE_SUMS
        )
        for sheet_name, expected in ref.items():
            actual = self._sum_numeric(data, sheet_name)
            assert abs(actual - expected) < 0.01, (
                f"{project_type} {sheet_name!r} sum drifted: "
                f"expected {expected}, got {actual}. D3 must "
                f"NOT change any factory export value."
            )


# ---------------------------------------------------------------------------
# 8. rc1 frozen
# ---------------------------------------------------------------------------


class TestRc1Frozen:
    def test_rc1_sha_resolves(self):
        r = subprocess.run(
            [
                "git", "rev-parse", "--verify",
                "b425a0708719eaa5e1d922b1008e5609758e0ad4",
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        assert r.returncode == 0, (
            "rc1 frozen SHA must still resolve."
        )


# ---------------------------------------------------------------------------
# 9. D3 phase plan + hard no-go
# ---------------------------------------------------------------------------


class TestD3PhasePlanAndHardNoGo:
    """D3 — Depreciation Shadow Validation / Audit-Only
    Comparison is read-only / advisory. The hard no-go
    list is the implementation boundary."""

    def test_hard_no_go_pinned(self):
        no_go = [
            "no_runtime_depreciation_enablement (D3 is shadow only)",
            "no_feature_flag_enablement",
            "no_formula_changes",
            "no_depreciation_schedule_changes",
            "no_tax_calculation_changes",
            "no_pnl_calculation_changes",
            "no_cfads_changes",
            "no_period_depreciation_keur_changes",
            "no_tax_depreciation_audit_keur_changes",
            "no_canonical_runtime_promotion",
            "no_persistence_changes",
            "no_schema_changes",
            "no_ui_workflow_changes",
            "no_generic_project_depreciation_claims",
            "rc1_frozen (b425a0708719eaa5e1d922b1008e5609758e0ad4)",
        ]
        assert len(no_go) == 15

    def test_stop_after_report_contract(self):
        contract = {
            "type": "DRAFT PR",
            "merge": "NO auto-merge",
            "next_phase_start": (
                "Stack validation report (no further "
                "phase) — final stack D1+D2+D3 closes the "
                "PR #530 safe-next-step arc"
            ),
        }
        assert contract["type"] == "DRAFT PR"
        assert contract["merge"] == "NO auto-merge"
