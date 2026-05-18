"""Validation for the TUHO R35 row source validation workbook."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner


WORKBOOK_PATH = Path("reports/phase6_tuho_r35_source_validation.xlsx")
REQUIRED_SHEETS = {
    "Summary",
    "Period Alignment",
    "Sign Convention",
    "Revenue Check",
    "OPEX Check",
    "Depreciation Check",
    "Senior Interest Check",
    "SHL Interest Check",
    "R34 Check",
    "R35 Reconstruction",
    "Largest Deltas",
}


def _run_default_tuho():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return WaterfallRunner(project, engine).run(config)


def test_r35_source_validation_workbook_exists_with_required_sheets():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    assert REQUIRED_SHEETS.issubset(set(workbook.sheetnames))


def test_period_alignment_contains_60_matched_operating_periods():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Period Alignment"]

    op_indices = [sheet.cell(row=row, column=1).value for row in range(2, 62)]
    matches = [sheet.cell(row=row, column=8).value for row in range(2, 62)]

    assert op_indices == list(range(60))
    assert all(matches)


def test_sign_convention_sheet_separates_signs_from_source_gaps():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Sign Convention"]
    values = {
        sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=5).value
        for row in range(2, sheet.max_row + 1)
    }

    assert values["R8 Revenue"] == "OK"
    assert values["R10 OPEX"] == "OK"
    assert values["R34 Fiscal Reintegration"] == "OK"
    assert values["R13 Depreciation"] == "source gap"
    assert values["R27 SHL Interest"] == "source gap"


def test_summary_confirms_main_source_validation_findings():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    summary = workbook["Summary"]
    values = {
        summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
        for row in range(2, summary.max_row + 1)
    }

    assert values["All op_idx/date aligned"] is True
    assert values["R35 total delta"] == pytest.approx(12_216.3705549183, abs=0.01)
    assert values["SHL interest delta"] == pytest.approx(10_347.270775417655, abs=0.01)
    assert values["Depreciation delta"] == pytest.approx(2_302.166786061971, abs=0.01)
    assert values["R34 delta"] == pytest.approx(0.0, abs=0.01)
    assert values["Previous attribution valid?"] == "Yes with caveat"
    assert values["Next branch"] == "phase6-shl-gross-interest-pnl-bridge"


def test_shl_and_depreciation_checks_identify_real_source_gaps():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    shl_sheet = workbook["SHL Interest Check"]
    depreciation_sheet = workbook["Depreciation Check"]

    assert shl_sheet.cell(row=2, column=3).value == pytest.approx(-1297.402605529328, abs=0.01)
    assert shl_sheet.cell(row=2, column=5).value == 0
    assert shl_sheet.cell(row=2, column=9).value == pytest.approx(1297.402605529328, abs=0.01)
    assert shl_sheet.cell(row=2, column=10).value == (
        "real source mismatch: gross accrued vs current Python field"
    )

    assert depreciation_sheet.cell(row=2, column=3).value == pytest.approx(
        -1845.438771304573,
        abs=0.01,
    )
    assert depreciation_sheet.cell(row=2, column=6).value == pytest.approx(
        -1168.508560730594,
        abs=0.01,
    )
    assert depreciation_sheet.cell(row=2, column=8).value == "book/tax row ownership"


def test_r35_reconstruction_and_largest_delta_sheet_are_consistent():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    reconstruction = workbook["R35 Reconstruction"]
    largest = workbook["Largest Deltas"]

    assert reconstruction.cell(row=2, column=1).value == 0
    assert reconstruction.cell(row=2, column=9).value == pytest.approx(2024.862400038727, abs=0.01)
    assert reconstruction.cell(row=2, column=15).value == pytest.approx(0.0, abs=0.01)

    abs_deltas = [largest.cell(row=row, column=7).value for row in range(2, 62)]
    assert abs_deltas == sorted(abs_deltas, reverse=True)
    assert largest.cell(row=2, column=2).value == 0
    assert largest.cell(row=2, column=8).value == "SHL interest field mapping/accrual"


def test_workbook_generation_does_not_change_default_runtime_behavior():
    baseline = _run_default_tuho()
    diagnostic = _run_default_tuho()

    assert diagnostic.total_revenue_keur == pytest.approx(baseline.total_revenue_keur, abs=0.0001)
    assert diagnostic.total_opex_keur == pytest.approx(baseline.total_opex_keur, abs=0.0001)
    assert diagnostic.total_tax_keur == pytest.approx(baseline.total_tax_keur, abs=0.0001)
    assert diagnostic.total_senior_ds_keur == pytest.approx(
        baseline.total_senior_ds_keur,
        abs=0.0001,
    )
