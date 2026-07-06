"""ENGINE-1A — Pilot-blocking wiring fixes.

E1-1: IC Report XLSX includes P&L Statement, Balance Sheet, Cash Flow Statement sheets.
E1-2: RuntimeSummary exposes total_tax_keur and effective_tax_rate_pct.

No waterfall math changed. No financial formulas changed.
"""
from __future__ import annotations

import io
import pytest


# ─── Shared fixtures ──────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def tuho_proj_result():
    from app.project_factories import create_default_tuho_wind1
    from app.ui_runner import _build_period_engine
    from app.waterfall_runner import WaterfallRunner, WaterfallRunConfig
    proj = create_default_tuho_wind1()
    eng = _build_period_engine(proj)
    result = WaterfallRunner(proj, eng).run(WaterfallRunConfig.from_inputs(proj, eng))
    return proj, result


# ─── E1-1: IC Report XLSX financial statement sheets ─────────────────────────

class TestIcReportXlsxFinancialSheets:
    @pytest.fixture(scope="class")
    def xlsx_bytes(self, tuho_proj_result):
        from app.services.ic_report_service import export_report_xlsx
        proj, result = tuho_proj_result
        return export_report_xlsx(proj, result, "Base", covenant_periods=[])

    @pytest.fixture(scope="class")
    def workbook(self, xlsx_bytes):
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    # Existing sheets still present
    def test_existing_sheets_unchanged(self, workbook):
        expected = {
            "Executive Summary", "Covenant Schedule",
            "Distribution Schedule", "Debt Schedule", "Tax Summary",
        }
        assert expected.issubset(set(workbook.sheetnames))

    # New sheets present
    def test_pnl_sheet_present(self, workbook):
        assert "P&L Statement" in workbook.sheetnames

    def test_balance_sheet_present(self, workbook):
        assert "Balance Sheet" in workbook.sheetnames

    def test_cash_flow_sheet_present(self, workbook):
        assert "Cash Flow Statement" in workbook.sheetnames

    # New sheets have data
    def test_pnl_sheet_has_header_and_data(self, workbook):
        ws = workbook["P&L Statement"]
        assert ws.max_row >= 2
        assert ws.cell(1, 1).value is not None

    def test_balance_sheet_has_header_and_data(self, workbook):
        ws = workbook["Balance Sheet"]
        assert ws.max_row >= 2
        assert ws.cell(1, 1).value is not None

    def test_cash_flow_sheet_has_header_and_data(self, workbook):
        ws = workbook["Cash Flow Statement"]
        assert ws.max_row >= 2
        assert ws.cell(1, 1).value is not None

    # Spot-check: P&L row labels present
    def test_pnl_contains_revenue_row(self, workbook):
        ws = workbook["P&L Statement"]
        labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert any("Revenue" in str(l) for l in labels if l)

    def test_pnl_contains_net_income_row(self, workbook):
        ws = workbook["P&L Statement"]
        labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert any("Net Income" in str(l) for l in labels if l)

    # Spot-check: Balance Sheet has numeric values
    def test_balance_sheet_has_numeric_values(self, workbook):
        ws = workbook["Balance Sheet"]
        numeric_found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numeric_found = True
                    break
        assert numeric_found

    # Spot-check: Cash Flow Statement has numeric values
    def test_cash_flow_has_numeric_values(self, workbook):
        ws = workbook["Cash Flow Statement"]
        numeric_found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numeric_found = True
                    break
        assert numeric_found


# ─── E1-2: RuntimeSummary tax KPIs ───────────────────────────────────────────

class TestRuntimeSummaryTaxKpis:
    @pytest.fixture(scope="class")
    def run_result_dict(self, tuho_proj_result):
        """Simulate the kpis dict that project_runner produces."""
        proj, result = tuho_proj_result
        return {
            "kpis": {
                "project_irr": result.project_irr,
                "equity_irr": result.equity_irr,
                "avg_dscr": result.actual_avg_dscr,
                "min_dscr": result.actual_min_dscr,
                "total_revenue_keur": result.total_revenue_keur,
                "total_ebitda_keur": result.total_ebitda_keur,
                "total_opex_keur": result.total_opex_keur,
                "total_capex_keur": proj.capex.total_capex,
                "total_distributions_keur": result.total_distribution_keur,
                "total_tax_keur": result.total_tax_keur,
            },
            "derivation_evidence": {},
            "tables": {},
        }

    @pytest.fixture(scope="class")
    def summary(self, run_result_dict):
        from app.ui.runtime_summary import build_runtime_summary
        return build_runtime_summary(run_result_dict, "tuho", "TUHO Wind 1")

    def test_dataclass_has_total_tax_keur_field(self, summary):
        assert hasattr(summary, "total_tax_keur")

    def test_dataclass_has_effective_tax_rate_pct_field(self, summary):
        assert hasattr(summary, "effective_tax_rate_pct")

    def test_total_tax_keur_not_available(self, summary):
        assert summary.total_tax_keur != "NOT_AVAILABLE"

    def test_effective_tax_rate_pct_not_available(self, summary):
        assert summary.effective_tax_rate_pct != "NOT_AVAILABLE"

    def test_total_tax_keur_is_numeric_string(self, summary):
        # Should be formatted as "X,XXX kEUR"
        assert "kEUR" in summary.total_tax_keur

    def test_effective_tax_rate_pct_is_percentage_string(self, summary):
        assert "%" in summary.effective_tax_rate_pct

    def test_to_dict_includes_tax_fields(self, summary):
        d = summary.to_dict()
        assert "total_tax_keur" in d
        assert "effective_tax_rate_pct" in d

    def test_to_dict_tax_fields_not_available(self, summary):
        d = summary.to_dict()
        assert d["total_tax_keur"] != "NOT_AVAILABLE"
        assert d["effective_tax_rate_pct"] != "NOT_AVAILABLE"

    def test_golden_tuho_total_tax_keur_positive(self, tuho_proj_result):
        """TUHO Wind 1 has positive total tax — verify it surfaces correctly."""
        proj, result = tuho_proj_result
        assert result.total_tax_keur is not None
        assert result.total_tax_keur > 0

    def test_golden_tuho_effective_tax_rate_reasonable(self, run_result_dict):
        """Effective tax rate on TUHO should be a small positive fraction of EBITDA."""
        from app.ui.runtime_summary import build_runtime_summary
        summary = build_runtime_summary(run_result_dict, "tuho", "TUHO Wind 1")
        # Rate is formatted as "X.XX%" — parse and verify it's between 0 and 100
        rate_str = summary.effective_tax_rate_pct.replace("%", "").strip()
        rate = float(rate_str)
        assert 0 < rate < 100
