"""Tests for the Phase 10 final calibration residual closeout pack."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOCS = ROOT / "docs"

WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
FINAL_REGISTRY = REPORTS / "phase10_final_residual_registry.csv"
BLOCKERS = REPORTS / "phase10_governance_blocker_summary.csv"
CONVENTIONS = REPORTS / "phase10_accepted_conventions_registry.csv"
PHASE11 = REPORTS / "phase10_phase11_carry_forward_items.csv"
DOC = DOCS / "phase10_final_calibration_residual_closeout_pack.md"


def _generate() -> None:
    write_calibration_reconciliation_pack()


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def test_closeout_sheet_exists():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "Calibration Residual Closeout" in wb.sheetnames


def test_final_reports_exist():
    _generate()
    assert FINAL_REGISTRY.exists()
    assert BLOCKERS.exists()
    assert CONVENTIONS.exists()
    assert PHASE11.exists()


def test_final_classification_framework_exists():
    _generate()
    rows = _csv_rows(FINAL_REGISTRY)
    classes = {row["classification"] for row in rows}
    assert "RESOLVED" in classes
    assert "ACCEPTED_CONVENTION" in classes
    assert "EVIDENCE_LIMITATION" in classes
    assert "GOVERNANCE_BLOCKER" in classes
    assert "ENGINEERING_FOLLOWUP" in classes
    assert "STAKEHOLDER_DECISION" in classes


def test_governance_blocker_and_engineering_followup_sections_exist():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Calibration Residual Closeout")
    assert "Governance blockers" in text or "Governance blockers".lower() in text.lower()
    assert "Remaining engineering work" in text or "engineering" in text.lower()


def test_executive_dashboard_contains_final_residual_summary():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Executive Dashboard")
    assert "Final residual posture summary" in text
    assert "Engineering follow-up count" in text


def test_review_signoff_contains_ownership_columns():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Review Signoff"]
    headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
    assert "governance_owner" in headers
    assert "roadmap_owner" in headers


def test_reviewer_notes_are_finalized():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Reviewer Notes")
    assert "How to interpret remaining residuals" in text
    assert "What should not be treated as runtime defects" in text
    assert "Phase 11 transition view" in text


def test_runtime_vs_governance_distinction_preserved():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance = _sheet_text(wb, "Governance")
    closeout = _sheet_text(wb, "Calibration Residual Closeout")
    assert "G20 status BLOCKED" in governance
    assert "R99/R102 status NOT APPROVED" in governance
    assert "Safe for review" in closeout


def test_doc_states_scope_and_limits():
    content = DOC.read_text(encoding="utf-8")
    assert "residual closeout philosophy" in content.lower()
    assert "governance interpretation framework" in content.lower()
    assert "Phase 11 transition framing".lower() in content.lower()
    assert "G20 remains `BLOCKED`" in content
    assert "R99/R102 remain `NOT APPROVED`" in content
    assert "no runtime changes statement" in content.lower()


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
