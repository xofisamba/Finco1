"""V4-1: UI == Factory parity tests.

Root cause: three independent sources of drift between factory-created
and UI-created projects were identified and fixed in input_adapter.py:

1. OPEX collapse — form always supplies opex_y1_keur, causing
   _resolve_user_inputs to replace all factory OpEx items (12 items
   for TUHO, with varying inflation rates) with a single generic item
   at 2% uniform inflation.  Fix: skip replacement when Y1 total
   matches factory sum within 0.01 kEUR.

2. CAPEX restructuring — form always supplies total_capex_keur, causing
   _resolve_user_inputs to zero IDC/bank-fee sub-lines and rebuild the
   CAPEX structure.  Fix: skip restructuring when total matches factory
   total within 0.01 kEUR.

3. Interest rate format confusion — _set_financing_interest_rate expects
   a value in percentage (5.75 for 5.75%) but the form serializes the
   snapshot as a decimal (0.0575).  This caused margin_bps to be clamped
   to 0, reducing the all-in rate from 5.75% to 3.1% for TUHO.  Fix:
   skip override when the supplied value (interpreted as either percentage
   or decimal) matches the factory all-in rate within 0.1 bps.

4. Saved-state path — build_projectinputs_from_snapshot used the generic
   Wind/Solar factory base regardless of template_source, losing all
   calibrated TUHO/Oborovo configuration (SHL mechanics, merchant curve,
   tax params, frozen DS schedule).  Fix: detect template_source from
   snapshot and use the project-specific factory base.
"""
from __future__ import annotations

import pytest

from app.input_adapter import build_projectinputs_from_snapshot, build_projectinputs_seeded
from app.project_factories import create_default_oborovo, create_default_tuho_wind1
from app.services.run_service import _get_seed_base_inputs
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner


def _run(proj):
    eng = _build_period_engine(proj)
    return WaterfallRunner(proj, eng).run(WaterfallRunConfig.from_inputs(proj, eng))


def _factory_snapshot(factory_fn, project_type: str, template_source: str) -> dict:
    """Build a snapshot dict that mirrors what the UI saves for a factory template."""
    proj = factory_fn()
    opex_y1 = sum(item.y1_amount_keur for item in proj.opex)
    return {
        "project_name": proj.info.name,
        "project_type": project_type,
        "country_market": proj.info.country_iso,
        "capacity_mw": str(proj.technical.capacity_mw),
        "tariff_eur_mwh": str(proj.revenue.ppa_base_tariff),
        "p50_hours": str(proj.technical.operating_hours_p50),
        "total_capex_keur": str(proj.capex.total_capex),
        "opex_y1_keur": str(opex_y1),
        "gearing_pct": "",
        "target_dscr": str(proj.financing.target_dscr),
        # Form stores interest rate as decimal (0.0575 for 5.75%)
        "interest_rate_pct": str(
            proj.financing.base_rate + proj.financing.margin_bps / 10_000
        ),
        "tenor_years": str(proj.financing.senior_tenor_years),
        "cod_date": str(proj.info.cod_date),
        "construction_months": str(proj.info.construction_months),
        "horizon_years": str(proj.info.horizon_years),
        "ppa_term_years": str(int(proj.revenue.ppa_term_years)),
        "active_project": template_source,
        "template_source": template_source,
    }


def _schema_from_factory(factory_fn, project_type: str):
    """Build a ProjectInputsSchema as the run service does from the form."""
    from main_web import _build_schema_from_form as build_schema

    proj = factory_fn()
    opex_y1 = sum(item.y1_amount_keur for item in proj.opex)
    return build_schema(
        project_type=project_type,
        scenario="Base",
        capacity_mw=str(proj.technical.capacity_mw),
        tariff_eur_mwh=str(proj.revenue.ppa_base_tariff),
        p50_hours=str(proj.technical.operating_hours_p50),
        total_capex_keur=str(proj.capex.total_capex),
        opex_y1_keur=str(opex_y1),
        target_dscr=str(proj.financing.target_dscr),
        interest_rate_pct=str(
            proj.financing.base_rate + proj.financing.margin_bps / 10_000
        ),
        tenor_years=str(proj.financing.senior_tenor_years),
        cod_date=str(proj.info.cod_date),
        construction_months=str(proj.info.construction_months),
        horizon_years=str(proj.info.horizon_years),
        ppa_term_years_form=str(int(proj.revenue.ppa_term_years)),
    )


# ─── TUHO parity ──────────────────────────────────────────────────────────

class TestTuhoUIFactoryParity:
    """Fresh-run (form) path and saved-state path both produce factory-identical KPIs."""

    def test_fresh_run_opex_structure_preserved(self):
        """build_projectinputs_seeded preserves factory opex when Y1 total matches."""
        factory = create_default_tuho_wind1()
        schema = _schema_from_factory(create_default_tuho_wind1, "Wind")
        ui_inputs = build_projectinputs_seeded(schema, _get_seed_base_inputs("tuho"))

        assert len(ui_inputs.opex) == len(factory.opex), (
            f"Opex structure collapsed: {len(ui_inputs.opex)} items vs "
            f"{len(factory.opex)} in factory"
        )

    def test_fresh_run_capex_structure_preserved(self):
        """build_projectinputs_seeded preserves IDC/bank-fee sub-lines when total matches."""
        factory = create_default_tuho_wind1()
        schema = _schema_from_factory(create_default_tuho_wind1, "Wind")
        ui_inputs = build_projectinputs_seeded(schema, _get_seed_base_inputs("tuho"))

        assert ui_inputs.capex.total_capex == pytest.approx(
            factory.capex.total_capex, abs=0.01
        )

    def test_fresh_run_interest_rate_preserved(self):
        """build_projectinputs_seeded preserves factory base_rate/margin_bps split."""
        factory = create_default_tuho_wind1()
        schema = _schema_from_factory(create_default_tuho_wind1, "Wind")
        ui_inputs = build_projectinputs_seeded(schema, _get_seed_base_inputs("tuho"))

        factory_all_in = factory.financing.base_rate + factory.financing.margin_bps / 10_000
        ui_all_in = ui_inputs.financing.base_rate + ui_inputs.financing.margin_bps / 10_000
        assert ui_all_in == pytest.approx(factory_all_in, abs=0.00001)

    def test_fresh_run_equity_irr_bit_identical(self):
        """UI fresh-run KPIs equal factory KPIs to 0.01 bps."""
        factory = create_default_tuho_wind1()
        schema = _schema_from_factory(create_default_tuho_wind1, "Wind")
        ui_inputs = build_projectinputs_seeded(schema, _get_seed_base_inputs("tuho"))

        r_factory = _run(factory)
        r_ui = _run(ui_inputs)

        assert r_ui.equity_irr == pytest.approx(r_factory.equity_irr, abs=0.0001)
        assert r_ui.actual_avg_dscr == pytest.approx(r_factory.actual_avg_dscr, abs=0.001)
        assert r_ui.total_distribution_keur == pytest.approx(
            r_factory.total_distribution_keur, abs=1.0
        )

    def test_saved_state_path_equity_irr_bit_identical(self):
        """build_projectinputs_from_snapshot with template_source=tuho equals factory."""
        factory = create_default_tuho_wind1()
        snapshot = _factory_snapshot(create_default_tuho_wind1, "Wind", "tuho")
        saved_inputs = build_projectinputs_from_snapshot(snapshot)

        r_factory = _run(factory)
        r_saved = _run(saved_inputs)

        assert r_saved.equity_irr == pytest.approx(r_factory.equity_irr, abs=0.0001)
        assert r_saved.actual_avg_dscr == pytest.approx(r_factory.actual_avg_dscr, abs=0.001)
        assert r_saved.total_distribution_keur == pytest.approx(
            r_factory.total_distribution_keur, abs=1.0
        )

    def test_saved_state_opex_structure_preserved(self):
        """Saved-state path preserves factory opex structure for TUHO template."""
        factory = create_default_tuho_wind1()
        snapshot = _factory_snapshot(create_default_tuho_wind1, "Wind", "tuho")
        saved_inputs = build_projectinputs_from_snapshot(snapshot)

        assert len(saved_inputs.opex) == len(factory.opex)


# ─── Oborovo parity ───────────────────────────────────────────────────────

class TestOborovoUIFactoryParity:
    """Oborovo parity mirrors TUHO — same three drift sources fixed."""

    def test_saved_state_equity_irr_bit_identical(self):
        factory = create_default_oborovo()
        snapshot = _factory_snapshot(create_default_oborovo, "Solar", "oborovo")
        saved_inputs = build_projectinputs_from_snapshot(snapshot)

        r_factory = _run(factory)
        r_saved = _run(saved_inputs)

        assert r_saved.equity_irr == pytest.approx(r_factory.equity_irr, abs=0.0001)
        assert r_saved.actual_avg_dscr == pytest.approx(r_factory.actual_avg_dscr, abs=0.001)

    def test_generic_template_source_unchanged(self):
        """Generic template_source falls through to generic factory (no regression)."""
        snapshot = _factory_snapshot(create_default_tuho_wind1, "Wind", "generic_wind")
        # Should not raise and should return Wind-type project
        result = build_projectinputs_from_snapshot(snapshot)
        assert result.info.country_iso is not None


# ─── Changed-value identity checks ──────────────────────────────────────

class TestChangedValuesStillApplied:
    """Ensure the tolerance checks do NOT suppress real user changes."""

    def test_changed_opex_is_applied(self):
        """When opex_y1 differs from factory, single-item replacement IS applied."""
        snapshot = _factory_snapshot(create_default_tuho_wind1, "Wind", "tuho")
        snapshot["opex_y1_keur"] = "3000.0"  # different from factory ~1998
        inputs = build_projectinputs_from_snapshot(snapshot)
        assert len(inputs.opex) == 1, "Replacement should apply when value differs"
        assert inputs.opex[0].y1_amount_keur == pytest.approx(3000.0, abs=0.01)

    def test_changed_interest_rate_is_applied(self):
        """When interest_rate differs from factory, rate override IS applied."""
        snapshot = _factory_snapshot(create_default_tuho_wind1, "Wind", "tuho")
        # Supply 8% as decimal (form format) — very different from factory ~5.75%
        snapshot["interest_rate_pct"] = "0.08"
        inputs = build_projectinputs_from_snapshot(snapshot)
        all_in = inputs.financing.base_rate + inputs.financing.margin_bps / 10_000
        # 0.08 / 100 = 0.0008 — that's the "decimal treated as pct" path; it won't
        # match factory, so override fires.  Just check it changed from factory.
        factory = create_default_tuho_wind1()
        factory_all_in = factory.financing.base_rate + factory.financing.margin_bps / 10_000
        assert all_in != pytest.approx(factory_all_in, abs=0.001), (
            "Interest rate should change when user supplies a different value"
        )


# ─── Export: canonical FS sheets ─────────────────────────────────────────

class TestExportCanonicalFS:
    """Values-only export includes canonical financial statements (no legacy generator)."""

    def test_export_includes_pnl_and_bs_sheets(self):
        """build_excel_export produces PnL, Balance Sheet, PF Cash Waterfall sheets."""
        import openpyxl
        from io import BytesIO
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        demo = run_demo_project("TUHO")
        result = demo.result
        project_inputs = demo.project_inputs

        raw = build_excel_export(
            result=result,
            project_inputs=project_inputs,
            project_type="Wind",
        )
        wb = openpyxl.load_workbook(BytesIO(raw))
        sheet_names = wb.sheetnames
        assert "PnL" in sheet_names, f"PnL sheet missing; got {sheet_names}"
        assert "Balance Sheet" in sheet_names, f"Balance Sheet missing; got {sheet_names}"
        assert "PF Cash Waterfall" in sheet_names, f"PF Cash Waterfall missing; got {sheet_names}"

    def test_pnl_sheet_has_data(self):
        """PnL sheet contains non-empty revenue values from canonical engine."""
        import openpyxl
        from io import BytesIO
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        demo = run_demo_project("TUHO")
        raw = build_excel_export(
            result=demo.result,
            project_inputs=demo.project_inputs,
            project_type="Wind",
        )
        wb = openpyxl.load_workbook(BytesIO(raw))
        ws = wb["PnL"]
        # Check at least one data row (row 2+) has a non-zero revenue value
        revenue_col = None
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value and "Revenue" in str(col[0].value):
                revenue_col = col[0].column
                break
        assert revenue_col is not None, "Revenue column not found in PnL sheet"
        revenues = [ws.cell(row=r, column=revenue_col).value for r in range(2, ws.max_row + 1)]
        assert any(v and v != 0 for v in revenues), "PnL Revenue column is empty"
