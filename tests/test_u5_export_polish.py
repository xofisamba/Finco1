"""U5 — Export Polish.

Presentation-only improvements to the institutional export
workbook (app/export/institutional_workbook.py and
app/export/workbook_index.py). No formula changes, no export
value changes, no workbook calculation changes, no sheet
reordering, no parity changes.

These tests pin:
  - Internal "Phase 11" / "factory" wording is removed from the
    Cover, OPEX, and Audit sheets, and the print header.
  - The Workbook_Index sheet inventory now documents the
    Validation Status sheet (previously missing).
  - Cover, Workbook_Index, and Validation Status tabs are
    visually highlighted with a tab color.
  - Sheet creation order is unchanged (18 + Validation Status = 19
    sheets, in the same sequence as before).
"""

from __future__ import annotations

import io
import pathlib

import openpyxl
import pytest

from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.export.workbook_index import INSTITUTIONAL_SHEET_INVENTORY

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]

EXPECTED_SHEET_ORDER = (
    "Export_Metadata",
    "Workbook_Index",
    "Cover",
    "Governance",
    "Runtime Summary",
    "Inputs",
    "Construction",
    "OPEX",
    "CAPEX",
    "Revenue",
    "Senior Debt",
    "SHL",
    "Tax",
    "P&L",
    "Cash Flow",
    "Balance Sheet",
    "Audit",
    "Gap Register",
    "Validation Status",
)


@pytest.fixture(scope="module")
def workbook():
    data = export_institutional_workbook_skeleton("tuho")
    return openpyxl.load_workbook(io.BytesIO(data))


class TestSheetOrderUnchanged:
    def test_sheet_order_matches_pinned_sequence(self, workbook):
        assert tuple(workbook.sheetnames) == EXPECTED_SHEET_ORDER


class TestNoInternalCodenames:
    def test_institutional_workbook_source_has_no_phase11_codename(self):
        text = (
            REPO_ROOT / "app/export/institutional_workbook.py"
        ).read_text()
        assert "Phase 11" not in text

    def test_cover_sheet_has_no_phase_codename(self, workbook):
        cover_text = " ".join(
            str(c.value)
            for row in workbook["Cover"].iter_rows()
            for c in row
            if c.value
        )
        assert "Phase 11" not in cover_text

    def test_print_header_has_no_phase_codename(self, workbook):
        for sheet in workbook.worksheets:
            assert "Phase 11" not in (sheet.oddHeader.right.text or "")

    def test_opex_sheet_has_no_factory_assumptions_wording(self, workbook):
        opex_text = " ".join(
            str(c.value)
            for row in workbook["OPEX"].iter_rows()
            for c in row
            if c.value
        )
        assert "factory assumptions" not in opex_text.lower()

    def test_audit_sheet_has_no_factory_bound_wording(self, workbook):
        audit_text = " ".join(
            str(c.value)
            for row in workbook["Audit"].iter_rows()
            for c in row
            if c.value
        )
        assert "factory-bound" not in audit_text.lower()


class TestWorkbookIndexCoversValidationStatus:
    def test_inventory_includes_validation_status(self):
        names = [row[0] for row in INSTITUTIONAL_SHEET_INVENTORY]
        assert "Validation Status" in names

    def test_inventory_row_is_well_formed(self):
        row = next(
            r for r in INSTITUTIONAL_SHEET_INVENTORY
            if r[0] == "Validation Status"
        )
        assert len(row) == 5
        assert row[1]


class TestTabColorHighlighting:
    def test_cover_tab_is_colored(self, workbook):
        assert workbook["Cover"].sheet_properties.tabColor is not None

    def test_workbook_index_tab_is_colored(self, workbook):
        assert workbook["Workbook_Index"].sheet_properties.tabColor is not None

    def test_validation_status_tab_is_colored(self, workbook):
        assert (
            workbook["Validation Status"].sheet_properties.tabColor is not None
        )


class TestCoverMentionsValidationStatus:
    def test_cover_points_to_validation_status_sheet(self, workbook):
        cover_text = " ".join(
            str(c.value)
            for row in workbook["Cover"].iter_rows()
            for c in row
            if c.value
        )
        assert "Validation Status" in cover_text
