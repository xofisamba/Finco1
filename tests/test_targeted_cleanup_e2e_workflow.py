"""Targeted cleanup end-to-end workflow coverage."""

from __future__ import annotations

import os
import sys
import uuid
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.api.project_runner import run_project
from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.input_adapter import build_projectinputs
from app.input_schema import DebtInput, OpexInput, ProjectInputsSchema, RevenueInput
from app.persistence.repository import (
    bind_workspace_to_scenario,
    get_workspace_state,
    list_exports,
    record_export,
    record_workspace_runtime,
    runtime_guard_for_snapshot,
    save_project,
    save_run,
    save_scenario,
    save_workspace_state,
)


def _governance_state() -> dict[str, str]:
    return {"g20_status": "BLOCKED", "r99_r102_status": "NOT APPROVED"}


def _default_snapshot(project_code: str) -> dict[str, str]:
    code = (project_code or "tuho").strip().lower()
    return {
        "active_project": code,
        "project_type": "Solar" if code == "oborovo" else "Wind",
        "scenario": "Base",
        "capacity_mw": "",
        "tariff_eur_mwh": "",
        "p50_hours": "",
        "total_capex_keur": "",
        "opex_y1_keur": "",
        "gearing_pct": "",
        "target_dscr": "",
        "interest_rate_pct": "",
        "tenor_years": "",
        "cod_date": "",
        "construction_months": "",
        "horizon_years": "",
        "capacity_factor": "",
        "ppa_term_years": "",
    }


def _runtime_result_for_snapshot(snapshot: dict[str, str]) -> dict:
    def _float(value: str) -> float | None:
        return None if value in ("", None) else float(value)

    def _int(value: str) -> int | None:
        return None if value in ("", None) else int(float(value))

    schema = ProjectInputsSchema(
        project_type=snapshot["project_type"],
        scenario=snapshot["scenario"],
        capacity_mw=_float(snapshot.get("capacity_mw", "")),
        revenue=RevenueInput(
            tariff_eur_mwh=_float(snapshot.get("tariff_eur_mwh", "")),
            p50_hours=_float(snapshot.get("p50_hours", "")),
        ),
        opex=OpexInput(
            opex_y1_keur=_float(snapshot.get("opex_y1_keur", "")),
        ),
        debt=DebtInput(
            gearing_pct=_float(snapshot.get("gearing_pct", "")),
            target_dscr=_float(snapshot.get("target_dscr", "")),
            interest_rate_pct=_float(snapshot.get("interest_rate_pct", "")),
            tenor_years=_int(snapshot.get("tenor_years", "")),
        ),
    )
    override = build_projectinputs(schema)
    project_key = "TUHO" if snapshot["active_project"] == "tuho" else "Oborovo"
    return run_project(project_key, snapshot["scenario"], project_inputs_override=override)


@pytest.fixture
def test_db():
    import app.persistence.db as db_mod

    old_path = os.environ.get("FINCO_DB_PATH")
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_file = os.path.join(base_dir, f"targeted_cleanup_e2e_{uuid.uuid4().hex[:8]}.db")
    os.environ["FINCO_DB_PATH"] = db_file
    db_mod.DB_PATH = db_file

    yield db_file

    if os.path.exists(db_file):
        os.remove(db_file)
    if old_path:
        os.environ["FINCO_DB_PATH"] = old_path
        db_mod.DB_PATH = old_path
    else:
        os.environ.pop("FINCO_DB_PATH", None)


def test_project_select_edit_save_run_and_export_workflow(test_db):
    user_id = f"u{uuid.uuid4().hex[:8]}"
    governance = _governance_state()
    project = save_project(user_id, "tuho", "TUHO Wind 1", "tuho", governance_state=governance)

    base_snapshot = _default_snapshot("tuho")
    base_snapshot["tariff_eur_mwh"] = "72"
    base_snapshot["opex_y1_keur"] = "2100"
    base_snapshot["target_dscr"] = "1.15"

    saved = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Base Saved",
        project_code="tuho",
        source_project_template="tuho",
        snapshot=base_snapshot,
        governance_state=governance,
    )
    workspace = bind_workspace_to_scenario(user_id, project.project_id, "tuho", saved, governance_state=governance)
    assert workspace.dirty is False

    edited_snapshot = dict(workspace.saved_snapshot)
    edited_snapshot["tariff_eur_mwh"] = "88"
    edited_snapshot["opex_y1_keur"] = "2450"
    workspace = save_workspace_state(
        user_id=user_id,
        project_id=project.project_id,
        project_code="tuho",
        active_scenario_id=workspace.active_scenario_id,
        active_scenario_name=workspace.active_scenario_name,
        draft_snapshot=edited_snapshot,
        saved_snapshot=workspace.saved_snapshot,
        dirty=True,
        governance_state=governance,
    )

    assert workspace.dirty is True
    allowed, runtime_origin, message = runtime_guard_for_snapshot(workspace, edited_snapshot)
    assert allowed is False
    assert runtime_origin == "preview_only"
    assert "Unsaved edits" in message
    assert workspace.last_runtime_snapshot == {}

    saved_after_edit = save_scenario(
        user_id=user_id,
        project_id=project.project_id,
        scenario_name="TUHO Edited Saved",
        project_code="tuho",
        source_project_template="tuho",
        snapshot=edited_snapshot,
        governance_state=governance,
    )
    rebound = bind_workspace_to_scenario(user_id, project.project_id, "tuho", saved_after_edit, governance_state=governance)
    assert rebound.dirty is False
    assert rebound.active_scenario_id == saved_after_edit.scenario_id
    assert saved_after_edit.scenario_id != saved.scenario_id

    allowed, runtime_origin, message = runtime_guard_for_snapshot(rebound, rebound.saved_snapshot)
    assert allowed is True
    assert runtime_origin == "saved_state"
    assert message == ""

    runtime_result = _runtime_result_for_snapshot(rebound.saved_snapshot)
    runtime_snapshot_id = f"cleanup-run-{uuid.uuid4().hex[:8]}"
    recorded = record_workspace_runtime(
        user_id=user_id,
        project_id=project.project_id,
        project_code="tuho",
        runtime_snapshot=dict(rebound.saved_snapshot),
        runtime_summary=runtime_result["kpis"],
        runtime_snapshot_id=runtime_snapshot_id,
        runtime_origin="saved_state",
        governance_state=governance,
        active_scenario_id=saved_after_edit.scenario_id,
        active_scenario_name=saved_after_edit.scenario_name,
    )
    run_record = save_run(
        user_id=user_id,
        project_type="TUHO",
        scenario=rebound.saved_snapshot["scenario"],
        inputs=dict(rebound.saved_snapshot),
        kpis=runtime_result["kpis"],
        replay_metadata={"runtime_snapshot_id": runtime_snapshot_id},
    )

    assert recorded.last_runtime_snapshot_id == runtime_snapshot_id
    assert recorded.last_runtime_origin == "saved_state"
    assert recorded.last_runtime_snapshot["tariff_eur_mwh"] == "88"
    assert run_record.replay_metadata["runtime_snapshot_id"] == runtime_snapshot_id

    post_runtime_draft = dict(recorded.saved_snapshot)
    post_runtime_draft["tariff_eur_mwh"] = "91"
    stale_state = save_workspace_state(
        user_id=user_id,
        project_id=project.project_id,
        project_code="tuho",
        active_scenario_id=recorded.active_scenario_id,
        active_scenario_name=recorded.active_scenario_name,
        draft_snapshot=post_runtime_draft,
        saved_snapshot=recorded.saved_snapshot,
        dirty=True,
        governance_state=recorded.governance_state,
        last_runtime_snapshot=recorded.last_runtime_snapshot,
        last_runtime_summary=recorded.last_runtime_summary,
        last_runtime_snapshot_id=recorded.last_runtime_snapshot_id,
        last_runtime_origin=recorded.last_runtime_origin,
        last_runtime_scenario_id=recorded.last_runtime_scenario_id,
        last_runtime_at=recorded.last_runtime_at,
    )
    assert stale_state.last_runtime_snapshot["tariff_eur_mwh"] == "88"
    assert stale_state.draft_snapshot["tariff_eur_mwh"] == "91"

    workbook_dir = Path(os.path.dirname(test_db)) / f"targeted_cleanup_export_{uuid.uuid4().hex[:8]}"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbook_dir / "targeted_cleanup_runtime_workbook.xlsx"
    workbook_path.write_bytes(export_institutional_workbook_skeleton("tuho"))
    workbook = load_workbook(workbook_path, data_only=True)
    assert "Runtime Summary" in workbook.sheetnames

    export_record = record_export(
        user_id=user_id,
        project_code="tuho",
        export_type="institutional_workbook",
        artifact_name=workbook_path.name,
        artifact_path=str(workbook_path),
        project_id=project.project_id,
        scenario_id=saved_after_edit.scenario_id,
        governance_state=governance,
        runtime_snapshot_id=runtime_snapshot_id,
        replay_metadata={"runtime_authority": "backend_only"},
    )

    exports = list_exports(user_id, project_id=project.project_id)
    assert any(item.export_id == export_record.export_id for item in exports)
    assert export_record.governance_state["g20_status"] == "BLOCKED"
    assert export_record.governance_state["r99_r102_status"] == "NOT APPROVED"

    reloaded = get_workspace_state(user_id, project.project_id)
    assert reloaded is not None
    assert reloaded.dirty is True
    assert reloaded.last_runtime_snapshot["tariff_eur_mwh"] == "88"
    assert reloaded.draft_snapshot["tariff_eur_mwh"] == "91"

    workbook_path.unlink(missing_ok=True)
    workbook_dir.rmdir()
