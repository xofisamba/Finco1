"""Phase 6B.5 — Tests for SPV tax Excel export helper."""
from __future__ import annotations

import os
import tempfile

import pytest

from domain.tax.templates.inputs import (
    TaxTemplate,
    TaxDepreciationRule,
    CITTier,
)
from domain.tax.templates.resolver import resolve_tax_template
from domain.tax.engine_inputs import SPVTaxEngineInputs
from domain.tax.engine_runner import run_spv_tax_engine
from app.tax_excel_export import write_spv_tax_audit_sheets


# ── Fixtures ──────────────────────────────────────────────────────────────────

def make_hr_template() -> TaxTemplate:
    return TaxTemplate(
        country_code="HR",
        template_name="HR Flat 10%",
        tax_year=2026,
        cit_tiers=(CITTier(0.0, None, 0.10),),
        depreciation_rules=(
            TaxDepreciationRule(
                asset_category="infra",
                method="straight_line",
                annual_rate=None,
                useful_life_years=20.0,
                max_deductible_rate=None,
                bonus_depreciation_pct=0.0,
                deductible=True,
                notes="HR straight-line",
            ),
        ),
        withholding_tax_dividends=0.0,
        withholding_tax_interest=0.0,
    )


def resolved_config(template: TaxTemplate):
    return resolve_tax_template(template, ())


def make_result(entity_code="HR-SPV-001", n_periods=5):
    template = make_hr_template()
    config = resolved_config(template)

    inputs = SPVTaxEngineInputs(
        entity_code=entity_code,
        resolved_tax_config=config,
        ebitda_by_period_keur=tuple(1000.0 for _ in range(n_periods)),
        deductible_interest_by_period_keur=tuple(0.0 for _ in range(n_periods)),
        book_depreciation_by_period_keur=tuple(500.0 for _ in range(n_periods)),
        asset_cost_keur=10_000.0,
        depreciation_rule_asset_category="infra",
        non_deductible_addbacks_by_period_keur=(0.0,) * n_periods,
    )
    return run_spv_tax_engine(inputs)


# ── Test: empty results ──────────────────────────────────────────────────────

class TestEmptyResults:
    def test_empty_tax_results_no_sheets(self):
        """Empty tax_results tuple creates no Tax* sheets."""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        # Create a valid workbook with one visible sheet
        wb_init = Workbook()
        ws_init = wb_init.active
        ws_init.title = "PreExisting"
        ws_init["A1"] = "keep"
        wb_init.save(tmp_path)
        wb_init.close()

        # Call write with empty results — function is no-op
        writer = pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a")
        write_spv_tax_audit_sheets(writer, ())
        writer.close()

        # Verify: no Tax sheets added
        wb = load_workbook(tmp_path)
        tax_sheets = [s for s in wb.sheetnames if s.startswith("Tax")]
        assert len(tax_sheets) == 0, f"Expected no Tax* sheets, got: {tax_sheets}"
        wb.close()
        os.unlink(tmp_path)



# ── Test: Tax Summary sheet ──────────────────────────────────────────────────

class TestTaxSummarySheet:
    def test_writer_creates_tax_summary_sheet(self):
        """Tax Summary sheet is created when tax_results is non-empty."""
        result = make_result()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))

            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            assert "Tax Summary" in wb.sheetnames
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_tax_summary_contains_entity_code(self):
        """Tax Summary sheet contains the entity code from result."""
        result = make_result(entity_code="HR-SPV-001")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))

            import openpyxl
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb["Tax Summary"]
            # First data row (after header) should have entity code
            # Find it by searching in first column
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "HR-SPV-001":
                    found = True
                    break
            assert found, "Entity code HR-SPV-001 not found in Tax Summary sheet"
            wb.close()
        finally:
            os.unlink(tmp_path)


# ── Test: per-SPV tax sheets ──────────────────────────────────────────────────

class TestPerSpvTaxSheets:
    def test_writer_creates_per_spv_tax_sheet(self):
        """Per-SPV sheet named Tax_{entity_code} is created."""
        result = make_result(entity_code="HR-SPV-001")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))

            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            assert "Tax_HR-SPV-001" in wb.sheetnames
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_long_entity_code_sheet_name_truncated(self):
        """Sheet name is truncated to Excel 31-char limit."""
        long_code = "A" * 50  # 50 chars
        result = make_result(entity_code=long_code)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))

            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            for sheet_name in wb.sheetnames:
                assert len(sheet_name) <= 31, f"Sheet name '{sheet_name}' exceeds 31 chars"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_audit_note_in_first_row_of_per_spv_sheet(self):
        """Per-SPV sheet first row contains the audit-only notice."""
        result = make_result(entity_code="HR-SPV-001")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))

            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb["Tax_HR-SPV-001"]
            # Row 1 = DataFrame headers, Row 2 = audit note, Row 3+ = data
            header_cell = ws.cell(row=1, column=1).value
            assert header_cell == "Period", f"Expected 'Period' header in row 1, got: {header_cell}"
            second_cell = ws.cell(row=2, column=1).value
            assert "AUDIT-ONLY" in str(second_cell), f"Expected AUDIT-ONLY in row 2, got: {second_cell}"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_multiple_spv_results_creates_multiple_sheets(self):
        """Multiple SPV results create multiple per-SPV sheets."""
        result1 = make_result(entity_code="HR-SPV-001")
        result2 = make_result(entity_code="HR-SPV-002")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result1, result2))

            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            assert "Tax_HR-SPV-001" in wb.sheetnames
            assert "Tax_HR-SPV-002" in wb.sheetnames
            # Also has summary
            assert "Tax Summary" in wb.sheetnames
            wb.close()
        finally:
            os.unlink(tmp_path)


# ── Test: no mutation ─────────────────────────────────────────────────────────

class TestNoMutation:
    def test_no_mutation_of_spv_tax_result(self):
        """write_spv_tax_audit_sheets does not mutate the SPVTaxResult."""
        result = make_result(entity_code="HR-SPV-001")
        original_cit = result.total_cit_payable_keur

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            import pandas as pd
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                write_spv_tax_audit_sheets(writer, (result,))
        finally:
            os.unlink(tmp_path)

        assert result.total_cit_payable_keur == original_cit