"""Stack U: Pilot Trust Polish.

Three targeted fixes:

U1 — Export IRR Scaling
    The dashboard sheet in excel_export.py divided IRR values by 100 before
    writing them to Excel cells that already had "0.0%" number format (which
    multiplies by 100 on display). This caused equity_irr=0.1159 to be stored
    as 0.001159, displayed as 0.1% instead of 11.6%.
    Fix: remove the erroneous /100 in _write_dashboard_sheet(); store the raw
    decimal fraction; Excel's "0.0%" format handles the display.

U2 — OPEX Template 500 Guard
    sheet_opex_detail.html crashed on "%.1f"|format(value) when value is None.
    Two sites:
      - child inflation_pct: "if child.inflation_pct == 0" did not catch None
      - flat legacy items: escalation_pct accessed without "or 0" guard.
    Fix: "not child.inflation_pct" covers None and 0; "item.escalation_pct or 0"
    guards the legacy flat table.

U3 — Report Artefact Hygiene
    Pin SHA-256 of generated reports that must not be modified in non-report PRs.
    Recent PRs accidentally included phase10 and phase12 artefacts in diffs.
    Tests here detect unintended modifications.

No engine changes.  No parity numbers move.
"""
from __future__ import annotations

import hashlib
import io
import os
import sys
import tempfile
from pathlib import Path

import pytest

os.environ.setdefault("FINCO_SECRET_KEY", "test-secret-for-pytest-only")
sys.path.insert(0, str(Path(__file__).parent.parent))

REPO_ROOT = Path(__file__).parent.parent


# ── Module-scoped heavy fixtures ─────────────────────────────────────────────

@pytest.fixture(scope="module")
def tuho_result():
    from app.ui_runner import run_demo_project
    return run_demo_project("TUHO").result


@pytest.fixture(scope="module")
def oborovo_result():
    from app.ui_runner import run_demo_project
    return run_demo_project("Oborovo").result


def _build_wb(result):
    """Run build_excel_export and return the openpyxl workbook."""
    import openpyxl
    from app.excel_export import build_excel_export
    from app.project_factories import create_default_tuho_wind1
    buf = build_excel_export(result=result, project_inputs=create_default_tuho_wind1())
    return openpyxl.load_workbook(io.BytesIO(buf))


# ── U1: Export IRR Scaling ────────────────────────────────────────────────────

class TestU1IRRScaling:
    """Dashboard sheet IRR cells must store decimal fractions, not /100 values."""

    @pytest.fixture(scope="class")
    def tuho_wb(self, tuho_result):
        return _build_wb(tuho_result)

    @pytest.fixture(scope="class")
    def oborovo_wb(self, oborovo_result):
        from app.excel_export import build_excel_export
        from app.project_factories import create_default_oborovo
        import openpyxl
        buf = build_excel_export(result=oborovo_result, project_inputs=create_default_oborovo())
        return openpyxl.load_workbook(io.BytesIO(buf))

    def _dashboard_irr_cells(self, wb):
        """Return {label: cell} for IRR rows in the Dashboard sheet."""
        ws = wb["Dashboard"]
        cells = {}
        for row in ws.iter_rows():
            if row[0].value and "irr" in str(row[0].value).lower():
                cells[str(row[0].value)] = row[1]
        return cells

    def test_tuho_dashboard_exists(self, tuho_wb):
        assert "Dashboard" in tuho_wb.sheetnames

    def test_tuho_irr_cells_are_numeric(self, tuho_wb):
        irr_cells = self._dashboard_irr_cells(tuho_wb)
        assert irr_cells, "No IRR rows found in Dashboard sheet"
        for label, cell in irr_cells.items():
            assert isinstance(cell.value, float), (
                f"IRR cell '{label}' should be numeric float, got {type(cell.value)}: {cell.value}"
            )

    def test_tuho_project_irr_not_divided_twice(self, tuho_result, tuho_wb):
        """Stored value must be the raw decimal (e.g. 0.0941), not 0.000941."""
        irr_cells = self._dashboard_irr_cells(tuho_wb)
        for label, cell in irr_cells.items():
            if "project" in label.lower():
                stored = cell.value
                expected = tuho_result.project_irr
                assert abs(stored - expected) < 0.001, (
                    f"Project IRR cell stores {stored:.6f}; expected ~{expected:.6f}. "
                    f"If {stored:.6f} ≈ {expected/100:.6f} the /100 bug is back."
                )

    def test_tuho_equity_irr_not_divided_twice(self, tuho_result, tuho_wb):
        irr_cells = self._dashboard_irr_cells(tuho_wb)
        for label, cell in irr_cells.items():
            if "equity" in label.lower():
                stored = cell.value
                expected = tuho_result.equity_irr
                assert abs(stored - expected) < 0.001, (
                    f"Equity IRR cell stores {stored:.6f}; expected ~{expected:.6f}. "
                    f"If {stored:.6f} ≈ {expected/100:.6f} the /100 bug is back."
                )

    def test_tuho_irr_cells_have_pct_format(self, tuho_wb):
        """IRR cells should carry the '0.0%' Excel number format."""
        irr_cells = self._dashboard_irr_cells(tuho_wb)
        for label, cell in irr_cells.items():
            fmt = cell.number_format or ""
            assert "%" in fmt, (
                f"IRR cell '{label}' has number_format='{fmt}'; expected percentage format"
            )

    def test_tuho_irr_in_plausible_range(self, tuho_wb):
        """Decimal IRR values should be in (0.01, 0.30) — not near zero (over-divided)."""
        irr_cells = self._dashboard_irr_cells(tuho_wb)
        for label, cell in irr_cells.items():
            v = cell.value
            assert 0.01 < v < 0.30, (
                f"IRR cell '{label}' = {v:.6f} is outside [0.01, 0.30]. "
                f"Values near 0.001 indicate the /100 division bug."
            )

    def test_oborovo_irr_cells_are_numeric(self, oborovo_wb):
        irr_cells = self._dashboard_irr_cells(oborovo_wb)
        assert irr_cells
        for label, cell in irr_cells.items():
            assert isinstance(cell.value, float), (
                f"Oborovo IRR cell '{label}' should be float, got {type(cell.value)}: {cell.value}"
            )

    def test_oborovo_equity_irr_not_divided_twice(self, oborovo_result, oborovo_wb):
        irr_cells = self._dashboard_irr_cells(oborovo_wb)
        for label, cell in irr_cells.items():
            if "equity" in label.lower():
                stored = cell.value
                expected = oborovo_result.equity_irr
                assert abs(stored - expected) < 0.001, (
                    f"Oborovo equity IRR stores {stored:.6f}; expected ~{expected:.6f}"
                )

    def test_oborovo_irr_in_plausible_range(self, oborovo_wb):
        irr_cells = self._dashboard_irr_cells(oborovo_wb)
        for label, cell in irr_cells.items():
            v = cell.value
            assert 0.01 < v < 0.30, (
                f"Oborovo IRR cell '{label}' = {v:.6f} outside [0.01, 0.30]"
            )


# ── U2: OPEX Template 500 Guard ───────────────────────────────────────────────

class TestU2OpexTemplate500Guard:
    """sheet_opex_detail.html must not 500 when OPEX values are missing/None."""

    @pytest.fixture(scope="class")
    def jinja_env(self):
        from jinja2 import Environment, FileSystemLoader
        templates_dir = REPO_ROOT / "app" / "templates"
        return Environment(loader=FileSystemLoader(str(templates_dir)))

    def _minimal_project_ctx(self, **overrides):
        """Return a minimal project_ctx dict sufficient to render sheet_opex_detail."""
        ctx = {
            "name": "Test Project",
            "horizon_years": 5,
            "opex_y1_total_keur": 100.0,
            "opex_contingency_pct": 0,
            "opex_contingency_method": "none",
            "opex_detail_items": [],
            "opex_items": [],
        }
        ctx.update(overrides)
        return ctx

    def _render(self, jinja_env, project_ctx, is_user_project=False):
        tmpl = jinja_env.get_template("partials/sheet_opex_detail.html")
        return tmpl.render(project_ctx=project_ctx, is_user_project=is_user_project)

    def test_renders_with_empty_opex_detail_items(self, jinja_env):
        ctx = self._minimal_project_ctx()
        html = self._render(jinja_env, ctx)
        assert "No detailed OPEX items available" in html

    def test_renders_with_none_opex_y1_total(self, jinja_env):
        ctx = self._minimal_project_ctx(opex_y1_total_keur=None)
        html = self._render(jinja_env, ctx)
        assert html  # must not raise

    def test_renders_with_child_inflation_pct_none(self, jinja_env):
        """Template must not crash when child.inflation_pct is None."""
        child = {
            "code": "B.01.01",
            "name": "Asset Mgmt",
            "budget_y1_keur": 60.0,
            "inflation_pct": None,   # ← the crash trigger
            "wth_rate": 0.0,
            "source": "factory",
            "notes": "",
            "yearly_values": [60.0] * 5,
            "active_flags": [1] * 5,
        }
        cat = {
            "code": "B.01",
            "name": "Technical Management",
            "is_contingency": False,
            "contingency_pct": 0.0,
            "children": [child],
            "yearly_totals": [60.0] * 5,
        }
        ctx = self._minimal_project_ctx(opex_detail_items=[cat])
        html = self._render(jinja_env, ctx)
        assert "flat" in html  # None triggers the fallback label

    def test_renders_with_child_inflation_pct_zero(self, jinja_env):
        """zero inflation_pct should show 'flat' label, not crash."""
        child = {
            "code": "B.01.01",
            "name": "Asset Mgmt",
            "budget_y1_keur": 60.0,
            "inflation_pct": 0.0,
            "wth_rate": 0.0,
            "source": "factory",
            "notes": "",
            "yearly_values": [60.0] * 5,
            "active_flags": [1] * 5,
        }
        cat = {
            "code": "B.01",
            "name": "Technical Management",
            "is_contingency": False,
            "contingency_pct": 0.0,
            "children": [child],
            "yearly_totals": [60.0] * 5,
        }
        ctx = self._minimal_project_ctx(opex_detail_items=[cat])
        html = self._render(jinja_env, ctx)
        assert "flat" in html

    def test_renders_with_child_inflation_pct_nonzero(self, jinja_env):
        """Non-zero inflation_pct should show numeric value, not crash."""
        child = {
            "code": "B.01.01",
            "name": "Asset Mgmt",
            "budget_y1_keur": 60.0,
            "inflation_pct": 2.0,
            "wth_rate": 0.0,
            "source": "factory",
            "notes": "",
            "yearly_values": [60.0, 61.2, 62.4, 63.7, 64.9],
            "active_flags": [1] * 5,
        }
        cat = {
            "code": "B.01",
            "name": "Technical Management",
            "is_contingency": False,
            "contingency_pct": 0.0,
            "children": [child],
            "yearly_totals": [60.0, 61.2, 62.4, 63.7, 64.9],
        }
        ctx = self._minimal_project_ctx(opex_detail_items=[cat])
        html = self._render(jinja_env, ctx)
        assert "2.0%" in html

    def test_renders_legacy_opex_items_with_none_escalation(self, jinja_env):
        """Legacy flat items with None escalation_pct must not crash."""
        item = {
            "code": "b-01",
            "name": "Asset Mgmt",
            "y1_keur": 60.0,
            "unit": "kEUR",
            "fixed_variable": "Fixed",
            "recurring_oneoff": "Recurring",
            "escalation_pct": None,   # ← crash trigger in legacy table
            "start_year": 1,
            "end_year": 5,
            "notes": "",
        }
        ctx = self._minimal_project_ctx(opex_items=[item])
        html = self._render(jinja_env, ctx)
        assert "Asset Mgmt" in html
        assert "0.0%" in html  # None → 0.0 via "or 0" guard

    def test_renders_legacy_opex_items_with_normal_escalation(self, jinja_env):
        """Normal escalation_pct renders correctly after guard is applied."""
        item = {
            "code": "b-01",
            "name": "Asset Mgmt",
            "y1_keur": 60.0,
            "unit": "kEUR",
            "fixed_variable": "Fixed",
            "recurring_oneoff": "Recurring",
            "escalation_pct": 2.0,
            "start_year": 1,
            "end_year": 5,
            "notes": "",
        }
        ctx = self._minimal_project_ctx(opex_items=[item])
        html = self._render(jinja_env, ctx)
        assert "2.0%" in html


# ── U3: Report Artefact Hygiene ───────────────────────────────────────────────

# SHA-256 pins for generated report artefacts.
# These files must not be modified in non-report PRs.
# If a report legitimately needs regeneration, update the SHA here in a
# dedicated report-update commit with a clear commit message.
_REPORT_ARTEFACT_SHA_PINS: dict[str, str] = {
    "reports/phase10_calibration_reconciliation_pack.xlsx": (
        # Pinned at Stack U base (d7b5767 — Stack S squash merge).
        "112cf501deda9ec030189851d51dec09f7e98abdb628d99bb2c7e04584c24f15"
    ),
    "reports/phase12_governance_label_usage_matrix.csv": (
        # Pinned at Stack U base (d7b5767 — Stack S squash merge).
        "62e4d7fce9dd2e22ca66c21ba1751c3994a8ddfadc56e246d9445f18fe2764d6"
    ),
}


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        h.update(f.read())
    return h.hexdigest()


class TestU3ReportArtefactHygiene:
    """Generated report artefacts must not be unintentionally modified."""

    @pytest.mark.parametrize("relpath,expected_sha", list(_REPORT_ARTEFACT_SHA_PINS.items()))
    def test_report_artefact_sha256_unchanged(self, relpath, expected_sha):
        path = REPO_ROOT / relpath
        assert path.exists(), f"Report artefact not found: {relpath}"
        actual = _sha256_file(path)
        assert actual == expected_sha, (
            f"Generated report artefact '{relpath}' has been modified.\n"
            f"  Expected SHA: {expected_sha}\n"
            f"  Actual SHA:   {actual}\n"
            f"If regeneration is intentional, update the SHA pin in "
            f"tests/test_stack_u_pilot_trust_polish.py::_REPORT_ARTEFACT_SHA_PINS."
        )
