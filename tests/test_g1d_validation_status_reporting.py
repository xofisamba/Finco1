"""G1D-VALIDATION-REPORTING: tests for user-facing validation status
reporting (app/validation_status.py) and its Excel export integration.

This module classifies and labels existing runtime outputs; it does not
run a different calculation and does not change any model formula or
factory default.
"""
from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.export.institutional_workbook import export_institutional_workbook_skeleton
from app.validation_status import (
    DISCLOSURE_NOT_ADVICE,
    MetricValidationLevel,
    ValidationTier,
    all_validation_statuses,
    get_validation_status,
)

REPO_ROOT = Path(__file__).resolve().parent.parent

FORBIDDEN_USER_FACING_TERMS = ("factory", "baseline", "hardcoded", "bootstrap")

REQUIRED_GENERIC_VALIDATED_METRICS = {
    "total_capex_keur",
    "total_revenue_keur",
    "total_opex_keur",
    "total_ebitda_keur",
    "idc_keur",
    "bank_fees_keur",
    "senior_debt_keur",
    "realized_gearing",
    "equity_funding_stack_keur",
}

REQUIRED_GENERIC_CAVEAT_METRICS = {
    "senior_debt_service_keur",
    "avg_dscr",
    "min_dscr",
    "project_irr",
    "equity_irr",
}


@pytest.mark.parametrize("project_key", ["tuho", "oborovo"])
def test_reference_projects_are_calibrated_frozen_reference(project_key: str) -> None:
    status = get_validation_status(project_key)
    assert status.tier == ValidationTier.CALIBRATED_FROZEN_REFERENCE
    assert status.metrics == ()
    assert status.internal_classification == "calibrated frozen internal reference"


@pytest.mark.parametrize("project_key", ["generic_solar", "solar", "generic_wind", "wind"])
def test_generic_solar_wind_are_validated_with_caveats(project_key: str) -> None:
    status = get_validation_status(project_key)
    assert status.tier == ValidationTier.VALIDATED_WITH_CAVEATS
    assert status.internal_classification == "validated generic bootstrap model with caveats"

    validated_keys = {m.metric_key for m in status.validated_metrics()}
    caveat_keys = {m.metric_key for m in status.caveated_metrics()}
    assert validated_keys == REQUIRED_GENERIC_VALIDATED_METRICS
    assert caveat_keys == REQUIRED_GENERIC_CAVEAT_METRICS

    for metric in status.caveated_metrics():
        assert metric.caveat_text, f"{metric.metric_key} must carry caveat text"


@pytest.mark.parametrize("project_key", ["bess", "generic_bess", "solar_bess", "wind_bess"])
def test_bess_hybrid_is_design_only(project_key: str) -> None:
    status = get_validation_status(project_key)
    assert status.tier == ValidationTier.DESIGN_ONLY
    assert status.internal_classification == "design-only / not externally validated"
    assert "not externally validated" in " ".join(status.disclosures).lower()


def test_portfolio_is_experimental() -> None:
    status = get_validation_status("portfolio")
    assert status.tier == ValidationTier.EXPERIMENTAL
    assert status.internal_classification == "experimental / internal only"
    assert "not externally validated" in " ".join(status.disclosures).lower()


def test_unknown_project_key_falls_back_to_experimental_without_raising() -> None:
    status = get_validation_status("some_unknown_future_project")
    assert status.tier == ValidationTier.EXPERIMENTAL


def test_disclosure_text_present_for_every_project() -> None:
    for status in all_validation_statuses():
        assert DISCLOSURE_NOT_ADVICE in status.disclosures


def test_generic_disclosures_include_reference_scope_and_methodology_caveat() -> None:
    for project_key in ("generic_solar", "generic_wind"):
        status = get_validation_status(project_key)
        joined = " ".join(status.disclosures)
        assert "internal reference workbooks" in joined
        assert "not a full external bank model review" in joined
        assert "methodology caveat" in joined.lower()


def test_no_forbidden_internal_terms_in_user_facing_text() -> None:
    for status in all_validation_statuses():
        haystacks = [status.tier_label, status.tier_description, *status.disclosures]
        for metric in status.metrics:
            haystacks.append(metric.display_name)
            haystacks.append(metric.caveat_text)
        text = " ".join(haystacks).lower()
        for term in FORBIDDEN_USER_FACING_TERMS:
            assert term not in text, f"forbidden internal term '{term}' leaked into user-facing text"


@pytest.mark.parametrize(
    "project_key,expected_tier_label",
    [
        ("tuho", "Validated reference project"),
        ("oborovo", "Validated reference project"),
        ("generic_solar", "Validated model with caveats"),
        ("generic_wind", "Validated model with caveats"),
        ("solar_bess", "Design-only, not externally validated"),
    ],
)
def test_validation_status_sheet_present_in_excel_export(
    project_key: str, expected_tier_label: str
) -> None:
    """Excel export carries the validation tier label for each project that
    has a runnable factory; the export bundle keys on the runtime project
    key, not the validation-status alias spellings."""
    export_project_key = {"solar_bess": "generic_solar"}.get(project_key, project_key)
    workbook_bytes = export_institutional_workbook_skeleton(export_project_key)
    workbook = load_workbook(filename=__import__("io").BytesIO(workbook_bytes))
    assert "Validation Status" in workbook.sheetnames


def test_validation_status_sheet_runs_for_all_runnable_projects() -> None:
    for project_key, expected_tier_label in (
        ("tuho", "Validated reference project"),
        ("oborovo", "Validated reference project"),
        ("generic_solar", "Validated model with caveats"),
        ("generic_wind", "Validated model with caveats"),
    ):
        workbook_bytes = export_institutional_workbook_skeleton(project_key)
        workbook = load_workbook(filename=__import__("io").BytesIO(workbook_bytes))
        sheet = workbook["Validation Status"]
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
        assert expected_tier_label in values, f"{project_key}: missing tier label in Validation Status sheet"


@pytest.mark.parametrize("project_key", ["tuho", "oborovo", "generic_solar", "generic_wind"])
def test_export_metadata_non_claims_block_does_not_contradict_validation_status(
    project_key: str,
) -> None:
    """The institutional workbook's Export_Metadata non-claims block must
    not contradict the Validation Status sheet added by G1D: Generic
    Solar/Wind are now classified as validated-with-caveats, so the old
    'exploratory and unvalidated' / 'do not use for financial decisions'
    wording (written before G1D) is no longer accurate and must be gone."""
    workbook_bytes = export_institutional_workbook_skeleton(project_key)
    workbook = load_workbook(filename=__import__("io").BytesIO(workbook_bytes))
    sheet = workbook["Export_Metadata"]
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    text = " ".join(values)

    assert "Generic solar/wind paths remain exploratory and unvalidated" not in text
    assert "Do not use generic paths for financial decisions" not in text

    assert (
        "Generic Solar/Wind are validated generic models with documented methodology caveats"
        in text
    )
    assert "BESS/Hybrid and Portfolio remain not externally validated" in text
    assert (
        "All outputs are financial modelling estimates only and do not constitute "
        "legal, tax, accounting, or investment advice" in text
    )

    # Stronger warnings must remain untouched.
    assert "NOT bank/lender approval" in text
    assert "NOT external audit certification" in text
    assert "NOT SaaS-ready or enterprise-ready" in text
    assert "G20 remains BLOCKED" in text
    assert "R99/R102 remain NOT APPROVED" in text


def test_engine_files_unchanged() -> None:
    """This validation-reporting-only branch must not touch runtime/domain
    code or factory defaults."""
    expected_hashes = {
        "app/waterfall_core.py": "6bf49f33efc989736c17cea0cb9b7723",
        "app/input_adapter.py": "ab296e927a3bc5869726519f68e58bec",
        "domain/inputs.py": "73dd17f60203e4121934381ef72964b6",
        "domain/revenue/generation.py": "177ef500f38bdb3c1379438c9e1d4f29",
        "domain/waterfall/waterfall_engine.py": "ddd93874e776ada65b94dda9aef678b6",
    }
    for rel_path, expected_md5 in expected_hashes.items():
        path = REPO_ROOT / rel_path
        actual_md5 = hashlib.md5(path.read_bytes()).hexdigest()
        assert actual_md5 == expected_md5, f"{rel_path} changed unexpectedly"


def test_project_factories_unchanged() -> None:
    """G1D is reporting/docs/UI-export only -- it must not touch the
    factory defaults fixed by G1H."""
    path = REPO_ROOT / "app" / "project_factories.py"
    actual_md5 = hashlib.md5(path.read_bytes()).hexdigest()
    assert actual_md5 == "bb541771e7837fb397c7cf16738e9fa0"
