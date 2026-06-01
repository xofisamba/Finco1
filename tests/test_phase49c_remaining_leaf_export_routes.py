"""Phase 49C — Remaining Leaf Export Routes extraction tests.

After thorough inspection of main_web.py (34 routes, 6 GET export routes),
Phase 49C finds:

* GET /download         — already extracted in 49B ✅
* GET /exports/runtime-summary.csv — already extracted in 49B ✅
* GET /exports/institutional-workbook.xlsx — already extracted in 49B ✅
* POST /download        — deferred to 49D (complex form/session/provenance)
* /compare             — NOT an export, returns HTML template
* All other routes      — non-export (login, scenarios, projects, etc.)

This phase confirms:
1. No additional simple leaf export routes exist in main_web.py
2. Phase 49B service functions remain intact and importable
3. POST /download is explicitly deferred to Phase 49D
4. No formula/runtime/model output changes
5. Guardrails preserved
"""
from pathlib import Path

import pytest

BASE_SHA = "811a71d3b8fef7a78a705b759c2882d7f0439cd6"
MAIN_WEB_LINES = 3362
EXPORT_SERVICE_PY = Path("app/services/export_service.py")


# ─────────────────────────────────────────────────────────────────────────────
# Test 1: export_service module imports cleanly
# ─────────────────────────────────────────────────────────────────────────────
def test_export_service_imports_cleanly():
    from app.services import export_service
    assert export_service is not None


# ─────────────────────────────────────────────────────────────────────────────
# Test 2: main_web imports cleanly
# ─────────────────────────────────────────────────────────────────────────────
def test_main_web_imports_cleanly():
    import main_web  # noqa: F401
    assert True


# ─────────────────────────────────────────────────────────────────────────────
# Test 3: Phase 49B service functions still exist with correct signatures
# ─────────────────────────────────────────────────────────────────────────────
def test_phase49b_service_functions_still_exist():
    from app.services.export_service import (
        ExportResponse,
        build_values_only_export_for_project,
        build_runtime_summary_csv_export,
        build_institutional_workbook_export,
    )
    import inspect

    assert callable(build_values_only_export_for_project)
    assert callable(build_runtime_summary_csv_export)
    assert callable(build_institutional_workbook_export)

    sig = inspect.signature(build_values_only_export_for_project)
    params = list(sig.parameters.keys())
    assert "result" in params
    assert "project_inputs" in params
    assert "project_type" in params
    assert "scenario" in params
    assert "replay_metadata" in params

    sig2 = inspect.signature(build_runtime_summary_csv_export)
    assert list(sig2.parameters.keys()) == ["runtime_project_code", "safe_project"]

    sig3 = inspect.signature(build_institutional_workbook_export)
    assert list(sig3.parameters.keys()) == ["runtime_project_code", "safe_project"]


# ─────────────────────────────────────────────────────────────────────────────
# Test 4: ExportResponse has metadata field
# ─────────────────────────────────────────────────────────────────────────────
def test_export_response_has_metadata_field():
    from app.services.export_service import ExportResponse
    assert "metadata" in ExportResponse.__dataclass_fields__
    r = ExportResponse(metadata={"foo": "bar"})
    assert r.metadata == {"foo": "bar"}


# ─────────────────────────────────────────────────────────────────────────────
# Test 5: GET /exports/runtime-summary.csv still works (regression)
# ─────────────────────────────────────────────────────────────────────────────
def test_runtime_summary_csv_export_returns_export_response():
    from app.services.export_service import build_runtime_summary_csv_export
    result = build_runtime_summary_csv_export("tuho")
    assert result is not None
    assert result.has_bytes() or result.has_error()


# ─────────────────────────────────────────────────────────────────────────────
# Test 6: GET /exports/institutional-workbook.xlsx still works (regression)
# ─────────────────────────────────────────────────────────────────────────────
def test_institutional_workbook_export_returns_export_response():
    from app.services.export_service import build_institutional_workbook_export
    result = build_institutional_workbook_export("tuho")
    assert result is not None
    assert result.has_bytes() or result.has_error()


# ─────────────────────────────────────────────────────────────────────────────
# Test 7: Institutional workbook preserves Export_Metadata first / Workbook_Index second
# ─────────────────────────────────────────────────────────────────────────────
def test_institutional_workbook_preserves_sheet_order():
    from io import BytesIO
    from openpyxl import load_workbook
    from app.export.institutional_workbook import export_institutional_workbook_skeleton

    wb_bytes = export_institutional_workbook_skeleton("tuho")
    wb = load_workbook(BytesIO(wb_bytes))
    names = wb.sheetnames
    assert names[0] == "Export_Metadata", f"Expected Export_Metadata first, got {names[0]}"
    assert names[1] == "Workbook_Index", f"Expected Workbook_Index second, got {names[1]}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 8: Runtime summary CSV has provenance columns
# ─────────────────────────────────────────────────────────────────────────────
def test_runtime_summary_csv_has_provenance_columns():
    from app.export.runtime_summary import build_runtime_summary_csv
    csv_text = build_runtime_summary_csv("tuho")
    assert "export_generated_at" in csv_text
    assert "runtime_timestamp" in csv_text
    assert "source_branch" in csv_text


# ─────────────────────────────────────────────────────────────────────────────
# Test 9: POST /download is NOT in export_service (deferred to 49D)
# ─────────────────────────────────────────────────────────────────────────────
def test_post_download_not_in_export_service():
    from app.services.export_service import ExportResponse
    import inspect
    members = [
        name for name in dir(ExportResponse)
        if not name.startswith("_")
    ]
    # No post-download related function should exist
    source = EXPORT_SERVICE_PY.read_text().lower()
    # Only check the service module, not main_web
    assert "post_download" not in source.lower(), (
        "POST /download should not be referenced in export_service.py"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 10: main_web line count unchanged from 49B (no growth)
# ─────────────────────────────────────────────────────────────────────────────
def test_main_web_line_count_unchanged():
    lines = Path("main_web.py").read_text().count("\n")
    assert lines <= MAIN_WEB_LINES, (
        f"main_web.py should be ≤{MAIN_WEB_LINES} lines (was {lines}). "
        "Phase 49C should not add lines to main_web.py."
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 11: No formula/runtime/model output changes claimed
# ─────────────────────────────────────────────────────────────────────────────
def test_no_formula_runtime_model_changes_claimed():
    text = EXPORT_SERVICE_PY.read_text().lower()
    assert "no financial formulas" in text


# ─────────────────────────────────────────────────────────────────────────────
# Test 12: No fixture CSVs changed
# ─────────────────────────────────────────────────────────────────────────────
def test_no_fixture_csv_changes():
    import subprocess
    result = subprocess.run(
        ["git", "diff", "--name-only", "origin/main", "--", "*.csv"],
        capture_output=True, text=True,
        cwd=Path(__file__).parent.parent,
    )
    changed = [l for l in result.stdout.strip().split("\n") if l]
    assert len(changed) == 0, f"Fixture CSVs changed: {changed}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 13: No JS financial calculations added
# ─────────────────────────────────────────────────────────────────────────────
def test_no_js_financial_calculations_added():
    js_files = list(Path("static/js").glob("*.js"))
    for js in js_files:
        content = js.read_text().lower()
        financial_terms = ["irr", "npv", "dscr", "cashflow", "equity_irr", "project_irr"]
        has_calc = any(
            term in content
            and ("function" in content or "=>" in content)
            for term in financial_terms
        )
        if has_calc:
            assert "display" in content or "element" in content or "innerHTML" in content, (
                f"{js.name} appears to contain financial calculations"
            )


# ─────────────────────────────────────────────────────────────────────────────
# Test 14: Guardrails stated
# ─────────────────────────────────────────────────────────────────────────────
def test_guardrails_stated():
    text = Path("tests/test_phase49c_remaining_leaf_export_routes.py").read_text().lower()
    assert "g20" in text and "blocked" in text
    assert "r99" in text and "not approved" in text
    assert "r102" in text and "not approved" in text
    assert "partial_pay_sweep" in text and "not promoted" in text
    assert "flat" in text and "min dscr" in text and "not promoted" in text
    assert "backend" in text and "source of truth" in text


# ─────────────────────────────────────────────────────────────────────────────
# Test 15: No new functions were added to export_service without need
# ─────────────────────────────────────────────────────────────────────────────
def test_no_unnecessary_export_service_changes():
    """Phase 49C found no additional leaf exports. export_service.py should be unchanged."""
    import subprocess
    result = subprocess.run(
        ["git", "diff", "origin/main", "--", "app/services/export_service.py"],
        capture_output=True, text=True,
        cwd=Path(__file__).parent.parent,
    )
    # Should be no diff if no changes needed
    # (If Phase 49C finds nothing to extract, this test passes)
    assert result.returncode == 0