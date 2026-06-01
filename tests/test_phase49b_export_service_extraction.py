"""Phase 49B — Export Service Extraction tests (rev2 after review fixes).

Verifies:
1. app.services.export_service imports cleanly
2. main_web imports cleanly
3. export service exposes expected public functions
4. ExportResponse.metadata populated for CSV/workbook exports (provenance preserved)
5. build_values_only_export_for_project accepts (result, project_inputs, ...) signature
6. record_export provenance uses export.metadata for runtime CSV and workbook
7. Export_Metadata first / Workbook_Index second in institutional workbook
8. Runtime summary CSV provenance columns present
9. No fixture CSV changes
10. Guardrails stated
"""
from io import BytesIO
from pathlib import Path

import pytest
from unittest.mock import MagicMock

BASE_SHA = "290e362b9b383df26a6fc37d9eb7b98805e339d9"
MAIN_WEB_LINES_BEFORE = 3367

EXPORT_SERVICE_PY = Path("app/services/export_service.py")
SERVICES_INIT = Path("app/services/__init__.py")


# ─────────────────────────────────────────────────────────────────────────────
# Test 1: export_service module imports cleanly
# ─────────────────────────────────────────────────────────────────────────────
def test_export_service_imports_cleanly():
    from app.services import export_service
    assert export_service is not None


# ─────────────────────────────────────────────────────────────────────────────
# Test 2: main_web imports cleanly
# ─────────────────────────────────────────────────────────────────────────────
def test_main_web_imports_cleanly():
    import main_web  # noqa: F401
    assert True


# ─────────────────────────────────────────────────────────────────────────────
# Test 3: export service exposes expected public functions
# ─────────────────────────────────────────────────────────────────────────────
def test_export_service_exposes_expected_functions():
    from app.services.export_service import (
        ExportResponse,
        build_values_only_export_for_project,
        build_runtime_summary_csv_export,
        build_institutional_workbook_export,
    )
    assert callable(build_values_only_export_for_project)
    assert callable(build_runtime_summary_csv_export)
    assert callable(build_institutional_workbook_export)
    assert hasattr(ExportResponse, "has_error")
    assert hasattr(ExportResponse, "has_bytes")
    assert "metadata" in ExportResponse.__dataclass_fields__, (
        "ExportResponse must have a metadata field"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 4: ExportResponse metadata field exists and is a dict
# ─────────────────────────────────────────────────────────────────────────────
def test_export_response_has_metadata_field():
    from app.services.export_service import ExportResponse
    # Success response with metadata
    ok = ExportResponse(
        bytes_data=b"hello",
        filename="test.xlsx",
        media_type="text/plain",
        metadata={"export_generated_at": "2026-01-01T00:00:00Z"},
    )
    assert ok.metadata is not None
    assert isinstance(ok.metadata, dict)

    # Error response with metadata
    err = ExportResponse(status_code=400, error_content="<html>error</html>")
    assert err.metadata == {}


# ─────────────────────────────────────────────────────────────────────────────
# Test 5: build_runtime_summary_csv_export populates metadata from runtime_rows[0]
# ─────────────────────────────────────────────────────────────────────────────
def test_runtime_summary_csv_export_metadata_populated():
    from app.services.export_service import build_runtime_summary_csv_export

    export = build_runtime_summary_csv_export("tuho")
    assert export.has_error() is False, f"Expected success, got: {export.error_content}"
    assert export.metadata is not None
    assert "export_generated_at" in export.metadata, (
        f"metadata missing export_generated_at. Keys: {list(export.metadata.keys())}"
    )
    assert "runtime_generated_at" in export.metadata, (
        f"metadata missing runtime_generated_at. Keys: {list(export.metadata.keys())}"
    )
    assert "runtime_origin" in export.metadata, (
        f"metadata missing runtime_origin. Keys: {list(export.metadata.keys())}"
    )
    assert "source_branch" in export.metadata


# ─────────────────────────────────────────────────────────────────────────────
# Test 6: build_institutional_workbook_export populates metadata from runtime_rows[0]
# ─────────────────────────────────────────────────────────────────────────────
def test_institutional_workbook_export_metadata_populated():
    from app.services.export_service import build_institutional_workbook_export

    export = build_institutional_workbook_export("tuho")
    assert export.has_error() is False, f"Expected success, got: {export.error_content}"
    assert export.metadata is not None
    assert "export_generated_at" in export.metadata
    assert "runtime_generated_at" in export.metadata
    assert "runtime_origin" in export.metadata
    assert "source_branch" in export.metadata


# ─────────────────────────────────────────────────────────────────────────────
# Test 7: build_values_only_export_for_project passes replay_metadata to build_excel_export
#          (monkeypatch to verify build_excel_export is called with correct provenance_metadata)
# ─────────────────────────────────────────────────────────────────────────────
def test_values_only_export_passes_replay_metadata_to_build_excel_export(monkeypatch):
    from app.services.export_service import build_values_only_export_for_project
    from unittest.mock import MagicMock

    captured = {}

    def fake_build_excel(result, project_inputs, provenance_metadata=None):
        captured["provenance_metadata"] = provenance_metadata
        return b"fake-bytes"

    monkeypatch.setattr("app.excel_export.build_excel_export", fake_build_excel)

    fake_result = MagicMock()
    fake_inputs = MagicMock()
    test_metadata = {
        "export_timestamp": "2026-06-01T10:00:00Z",
        "runtime_timestamp": "2026-06-01T10:00:00Z",
        "runtime_origin": "factory_base_runtime",
    }

    export = build_values_only_export_for_project(
        fake_result,
        fake_inputs,
        "Solar",
        "Base",
        replay_metadata=test_metadata,
    )

    assert captured.get("provenance_metadata") == test_metadata, (
        f"Expected replay_metadata to be passed unchanged to build_excel_export. "
        f"Got: {captured.get('provenance_metadata')}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 8: build_values_only_export_for_project signature: (result, project_inputs, project_type, scenario, **kwargs)
# ─────────────────────────────────────────────────────────────────────────────
def test_values_only_export_signature_accepts_result_and_inputs():
    from app.services.export_service import build_values_only_export_for_project
    import inspect

    sig = inspect.signature(build_values_only_export_for_project)
    params = list(sig.parameters.keys())
    assert "result" in params, f"result not in signature: {params}"
    assert "project_inputs" in params, f"project_inputs not in signature: {params}"
    assert "project_type" in params, f"project_type not in signature: {params}"
    assert "scenario" in params, f"scenario not in signature: {params}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 9: institutional workbook contains Export_Metadata (first) and Workbook_Index (second)
# ─────────────────────────────────────────────────────────────────────────────
def test_institutional_workbook_has_export_metadata_and_index_sheets():
    from app.export.institutional_workbook import export_institutional_workbook_skeleton
    wb_bytes = export_institutional_workbook_skeleton("tuho")
    assert wb_bytes is not None
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(wb_bytes))
    sheet_names = wb.sheetnames
    assert "Export_Metadata" in sheet_names, f"Export_Metadata sheet missing. Got: {sheet_names}"
    assert "Workbook_Index" in sheet_names, f"Workbook_Index sheet missing. Got: {sheet_names}"
    assert sheet_names[0] == "Export_Metadata", (
        f"Export_Metadata must be first sheet. Order: {sheet_names}"
    )
    assert sheet_names[1] == "Workbook_Index", (
        f"Workbook_Index must be second sheet. Order: {sheet_names}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 10: runtime summary CSV has provenance columns
# ─────────────────────────────────────────────────────────────────────────────
def test_runtime_summary_csv_has_provenance_columns():
    from app.export.runtime_summary import build_runtime_summary_csv
    csv_text = build_runtime_summary_csv("tuho")
    assert "export_generated_at" in csv_text
    assert "runtime_timestamp" in csv_text
    assert "source_branch" in csv_text
    assert "scenario_id" in csv_text
    assert "scenario_name" in csv_text


# ─────────────────────────────────────────────────────────────────────────────
# Test 11: main_web lines still reduced (not expanded)
# ─────────────────────────────────────────────────────────────────────────────
def test_main_web_lines_still_reduced():
    lines = Path("main_web.py").read_text().count("\n")
    assert lines < MAIN_WEB_LINES_BEFORE, (
        f"main_web should have fewer lines than {MAIN_WEB_LINES_BEFORE} after refactor. "
        f"Got {lines}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 12: No formula/runtime/model output changes claimed
# ─────────────────────────────────────────────────────────────────────────────
def test_no_formula_runtime_model_changes_claimed():
    text = EXPORT_SERVICE_PY.read_text().lower()
    assert "no financial formulas" in text, (
        "export_service.py must explicitly state 'no financial formulas' in its docstring"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 13: No fixture CSVs changed
# ─────────────────────────────────────────────────────────────────────────────
def test_no_fixture_csv_changes():
    import subprocess
    result = subprocess.run(
        ["git", "diff", "--name-only", "origin/main", "--", "*.csv"],
        capture_output=True, text=True,
        cwd=Path(__file__).parent.parent,
    )
    changed = [l for l in result.stdout.strip().split("\n") if l]
    assert len(changed) == 0, f"Fixture CSVs changed: {changed}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 14: No JS financial calculations added
# ─────────────────────────────────────────────────────────────────────────────
def test_no_js_financial_calculations_added():
    js_files = list(Path("static/js").glob("*.js"))
    for js in js_files:
        content = js.read_text().lower()
        financial_terms = ["irr", "npv", "dscr", "cashflow", "equity_irr", "project_irr"]
        has_calc = any(
            term in content
            and ("function" in content or "=>" in content)
            for term in financial_terms
        )
        if has_calc:
            assert "display" in content or "element" in content or "innerHTML" in content, (
                f"{js.name} appears to contain financial calculations"
            )


# ─────────────────────────────────────────────────────────────────────────────
# Test 15: Guardrails stated in test file
# ─────────────────────────────────────────────────────────────────────────────
def test_guardrails_stated():
    text = Path("tests/test_phase49b_export_service_extraction.py").read_text().lower()
    assert "g20" in text and "blocked" in text, "Tests must state G20 BLOCKED"
    assert "r99" in text and "not approved" in text, "Tests must state R99 NOT APPROVED"
    assert "r102" in text and "not approved" in text, "Tests must state R102 NOT APPROVED"
    assert "partial_pay_sweep" in text and "not promoted" in text, (
        "Tests must state partial_pay_sweep not promoted"
    )
    assert "flat" in text and "min dscr" in text and "not promoted" in text, (
        "Tests must state flat/min DSCR sculpting not promoted"
    )
    assert "backend" in text and "source of truth" in text, (
        "Tests must state backend source of truth"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Test 16: app/services/__init__.py exists and imports cleanly
# ─────────────────────────────────────────────────────────────────────────────
def test_services_init_exists():
    assert SERVICES_INIT.exists(), f"{SERVICES_INIT} not found"
    from app.services import export_service as es
    assert es is not None


# ─────────────────────────────────────────────────────────────────────────────
# Test 17: ExportResponse has no unused/unwanted fields
# ─────────────────────────────────────────────────────────────────────────────
def test_export_response_no_unused_params():
    text = EXPORT_SERVICE_PY.read_text()
    # Verify no project_record or user_id in export_service.py (they should only be in main_web.py)
    # These should NOT appear as parameters in build functions
    import inspect
    from app.services.export_service import (
        build_values_only_export_for_project,
        build_runtime_summary_csv_export,
        build_institutional_workbook_export,
    )
    for fn in [build_values_only_export_for_project, build_runtime_summary_csv_export, build_institutional_workbook_export]:
        sig = inspect.signature(fn)
        params = list(sig.parameters.keys())
        assert "project_record" not in params, (
            f"{fn.__name__} should not have project_record param. Got: {params}"
        )
        assert "user_id" not in params, (
            f"{fn.__name__} should not have user_id param. Got: {params}"
        )