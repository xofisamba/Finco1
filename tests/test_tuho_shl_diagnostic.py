"""TUHO SHL period-by-period reconciliation: Excel vs Python.

Diagnostic: Build complete Excel-vs-Python table for P28-P36 (Excel operating idx 26-36).

Excel data extracted directly from tuho_excel_1.xlsm DS and CF sheets.
Python data from WaterfallRunner.run() result.

Run: python tests/test_tuho_shl_diagnostic.py
"""

import openpyxl
from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunner, WaterfallRunConfig


# ─── Excel extraction ─────────────────────────────────────────────────────────

WB_PATH = "/root/.openclaw/workspace/oborovo_model/tuho_excel_1.xlsm"


def _extract_excel_tuho():
    wb = openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
    ds = wb["DS"]
    cf = wb["CF"]

    # Header row (row 2 in sheet, col E onwards = operating period 0)
    ds_header = list(ds.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cf_header = list(cf.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    def _col_for_op_idx(sheet_header, op_idx):
        """Return column index (0-based) for given operating period index."""
        # sheet col 5 = operating period 0 (Excel col E = idx 4 in 0-based)
        return 5 + op_idx

    def _row_vals(sheet, row_num, col_start=5, col_end=None):
        rows = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        if not rows:
            return []
        row = rows[0]
        end = col_end if col_end else len(row)
        return row[col_start:end]

    # Extract for op_idx 26-36 (Excel P28-P36)
    data = {}
    for op_idx in range(26, 37):
        col = _col_for_op_idx(ds_header, op_idx)
        date = ds_header[col] if col < len(ds_header) else None
        data[op_idx] = {"date": date}

        # DS sheet rows
        # Row 19: DSCR
        dscr_row = _row_vals(ds, 19)
        data[op_idx]["dscr"] = dscr_row[col] if col < len(dscr_row) else None

        # Row 127: Debt Service incl WHT (total)
        ds_row = _row_vals(ds, 127)
        data[op_idx]["total_ds"] = ds_row[col] if col < len(ds_row) else None

        # Row 120-126: SHL balance schedule
        shl_begin = _row_vals(ds, 120)
        shl_int = _row_vals(ds, 122)
        shl_prin = _row_vals(ds, 124)
        shl_end = _row_vals(ds, 126)
        data[op_idx]["shl_open"] = shl_begin[col] if col < len(shl_begin) else None
        data[op_idx]["shl_int"] = shl_int[col] if col < len(shl_int) else None
        data[op_idx]["shl_prin"] = shl_prin[col] if col < len(shl_prin) else None
        data[op_idx]["shl_close"] = shl_end[col] if col < len(shl_end) else None

        # Row 47: Senior beginning balance
        sr_begin = _row_vals(ds, 47)
        sr_end = _row_vals(ds, 53)
        sr_prin = _row_vals(ds, 49)
        sr_int = _row_vals(ds, 50)
        data[op_idx]["sr_open"] = sr_begin[col] if col < len(sr_begin) else None
        data[op_idx]["sr_close"] = sr_end[col] if col < len(sr_end) else None
        data[op_idx]["sr_prin"] = sr_prin[col] if col < len(sr_prin) else None
        data[op_idx]["sr_int"] = sr_int[col] if col < len(sr_int) else None

        # Row 54: Senior Debt Service
        sr_ds_row = _row_vals(ds, 54)
        data[op_idx]["sr_ds"] = sr_ds_row[col] if col < len(sr_ds_row) else None

        # Row 117: FCF for SHL Principal
        fcf_prin = _row_vals(ds, 117)
        data[op_idx]["fcf_shl_prin"] = fcf_prin[col] if col < len(fcf_prin) else None

        # CF sheet rows
        # Row 99: FCF for Distribution (R99)
        r99_row = _row_vals(cf, 99)
        data[op_idx]["r99"] = r99_row[col] if col < len(r99_row) else None

        # Row 102: FCF for SHL (R102)
        r102_row = _row_vals(cf, 102)
        data[op_idx]["r102"] = r102_row[col] if col < len(r102_row) else None

        # Row 104: Net SHL (R104)
        r104_row = _row_vals(cf, 104)
        data[op_idx]["r104"] = r104_row[col] if col < len(r104_row) else None

        # Row 106: FCF for dividends (R106)
        r106_row = _row_vals(cf, 106)
        data[op_idx]["r106"] = r106_row[col] if col < len(r106_row) else None

        # Row 119: Net Dividends (R119)
        r119_row = _row_vals(cf, 119)
        data[op_idx]["r119"] = r119_row[col] if col < len(r119_row) else None

    return data


# ─── Python extraction ─────────────────────────────────────────────────────────

def _run_python_tuho():
    proj = create_default_tuho_wind1()
    engine = _build_period_engine(proj)
    config = WaterfallRunConfig.from_inputs(proj, engine)
    runner = WaterfallRunner(inputs=proj, engine=engine)
    result = runner.run(config)
    return result


def _extract_python_periods(result):
    """Map: Python result index N = Excel operating period N.
    (First result period = Excel operating period 0 = 2030-06-30)
    """
    # Python result.periods[0] = Excel op_idx 0
    out = {}
    for py_idx, period in enumerate(result.periods):
        out[py_idx] = {
            "date": getattr(period, "date", None),
            "cf_after_tax": getattr(period, "cf_after_tax_keur", None),
            "senior_ds": getattr(period, "senior_ds_keur", None),
            "senior_interest": getattr(period, "senior_interest_keur", None),
            "senior_principal": getattr(period, "senior_principal_keur", None),
            "remaining_senior_balance": getattr(period, "senior_balance_keur", None),
            "dsra_contrib": getattr(period, "dsra_contribution_keur", None),
            "cf_after_reserves": getattr(period, "cf_after_reserves_keur", None),
            "shl_interest": getattr(period, "shl_interest_keur", None),
            "shl_principal": getattr(period, "shl_principal_keur", None),
            "shl_service": getattr(period, "shl_service_keur", None),
            "shl_pik": getattr(period, "shl_pik_keur", None),
            "shl_balance": getattr(period, "shl_balance_keur", None),
            "distribution": getattr(period, "distribution_keur", None),
            "dscr": getattr(period, "dscr", None),
            "lockup": getattr(period, "lockup_active", None),
            "cash_sweep": getattr(period, "cash_sweep_keur", None),
            # _cf_for_shl not directly available — reconstruct
            "cf_after_tax": getattr(period, "cf_after_tax_keur", None),
        }
    return out


# ─── Reconciliation table ─────────────────────────────────────────────────────

def fmt(v, dp=1):
    if v is None:
        return "?"
    if isinstance(v, float):
        return f"{v:.{dp}f}"
    return str(v)


def main():
    print("=== Loading Excel TUHO workbook ===")
    excel = _extract_excel_tuho()
    print(f"Excel data extracted: {len(excel)} periods")

    print("\n=== Running Python TUHO waterfall ===")
    result = _run_python_tuho()
    python_all = _extract_python_periods(result)
    print(f"Python periods: {len(python_all)}")

    # Build reconciliation table
    # Excel op_idx 26-36 → Python result index 26-36
    print("\n" + "=" * 200)
    header = [
        "Date", "ExOp", "PyIdx",
        # Excel CF rows
        "Ex_R99", "Ex_R102", "Ex_R104", "Ex_R106", "Ex_R119",
        # Excel SHL
        "Ex_SHL_open", "Ex_SHL_int", "Ex_SHL_prin", "Ex_SHL_close",
        # Excel Senior
        "Ex_Sr_open", "Ex_Sr_prin", "Ex_Sr_ds",
        # Excel DSCR
        "Ex_DSCR", "Ex_totalDS",
        # Python CF
        "Py_cf_after_tax", "Py_sr_ds", "Py_sr_int", "Py_sr_prin",
        "Py_sr_bal", "Py_dsra", "Py_cf_reserves",
        # Python SHL
        "Py_SHL_int", "Py_SHL_prin", "Py_SHL_close",
        # Python dist
        "Py_dist", "Py_DSCR", "Py_lockup", "Py_sweep",
        # Delta
        "Delta_SHL_close",
    ]

    print("\t".join(header))
    print("-" * 200)

    for op_idx in range(26, 37):
        py_idx = op_idx  # Python index == Excel operating index

        ex = excel.get(op_idx, {})
        py = python_all.get(py_idx, {})

        # Reconstruct _cf_for_shl for Python (before cap, if any)
        # In current code: _cf_for_shl = max(0, cf_after_tax - senior_ds) since dsra_contrib=0 at that point
        py_cf = py.get("cf_after_tax")
        py_sr_ds = py.get("senior_ds")
        py_raw_cf = (py_cf - py_sr_ds) if (py_cf is not None and py_sr_ds is not None) else None

        ex_date = ex.get("date")
        ex_date_str = str(ex_date)[:10] if ex_date else "?"

        delta_close = None
        ex_close = ex.get("shl_close")
        py_close = py.get("shl_balance")
        if ex_close is not None and py_close is not None:
            delta_close = py_close - ex_close

        row = [
            ex_date_str,
            str(op_idx),
            str(py_idx),
            # Excel CF
            fmt(ex.get("r99")),
            fmt(ex.get("r102")),
            fmt(ex.get("r104")),
            fmt(ex.get("r106")),
            fmt(ex.get("r119")),
            # Excel SHL
            fmt(ex.get("shl_open")),
            fmt(ex.get("shl_int")),
            fmt(ex.get("shl_prin")),
            fmt(ex.get("shl_close")),
            # Excel Senior
            fmt(ex.get("sr_open")),
            fmt(ex.get("sr_prin")),
            fmt(ex.get("sr_ds")),
            # Excel DSCR
            fmt(ex.get("dscr"), 2),
            fmt(ex.get("total_ds")),
            # Python CF
            fmt(py.get("cf_after_tax")),
            fmt(py.get("senior_ds")),
            fmt(py.get("senior_interest")),
            fmt(py.get("senior_principal")),
            fmt(py.get("remaining_senior_balance")),
            fmt(py.get("dsra_contrib")),
            fmt(py.get("cf_after_reserves")),
            # Python SHL
            fmt(py.get("shl_interest")),
            fmt(py.get("shl_principal")),
            fmt(py.get("shl_balance")),
            # Python dist
            fmt(py.get("distribution")),
            fmt(py.get("dscr"), 3),
            fmt(py.get("lockup")),
            fmt(py.get("cash_sweep")),
            # Delta
            fmt(delta_close),
        ]
        print("\t".join(row))

    print("\n" + "=" * 200)

    # Analysis questions
    print("\n=== Analysis ===\n")

    # Q1: Is Excel SHL repayment capped by R99, R102, R99 minus senior sweep?
    print("Q1: Excel cash ceiling for SHL")
    for op_idx in range(26, 37):
        ex = excel[op_idx]
        r99 = ex.get("r99", 0) or 0
        r102 = ex.get("r102", 0) or 0
        r104_net = ex.get("r104", 0) or 0  # negative = outflow
        shl_prin = ex.get("shl_prin", 0) or 0
        sr_ds = ex.get("sr_ds", 0) or 0
        total_ds = ex.get("total_ds", 0) or 0
        shl_int = ex.get("shl_int", 0) or 0
        ex_date = str(ex.get("date", ""))[:10]
        print(f"  OpIdx={op_idx} {ex_date}: R99={r99:.1f} R102={r102:.1f} SHL_prin={shl_prin:.1f} SHL_int={shl_int:.1f} SR_DS={sr_ds:.1f} R104={r104_net:.1f} TotalDS={total_ds:.1f}")

    print("\nQ2/Q3: Lockup behavior in Excel")
    for op_idx in range(26, 37):
        ex = excel[op_idx]
        dscr = ex.get("dscr")
        shl_int = ex.get("shl_int", 0) or 0
        shl_prin = ex.get("shl_prin", 0) or 0
        ex_date = str(ex.get("date", ""))[:10]
        print(f"  OpIdx={op_idx} {ex_date}: DSCR={dscr} SHL_int={shl_int:.1f} SHL_prin={shl_prin:.1f}")

    print("\nQ5: R104 Net SHL = interest + principal?")
    for op_idx in range(26, 37):
        ex = excel[op_idx]
        r104 = ex.get("r104", 0) or 0
        shl_int = ex.get("shl_int", 0) or 0
        shl_prin = ex.get("shl_prin", 0) or 0
        check = -(shl_int + shl_prin)  # R104 is negative
        ex_date = str(ex.get("date", ""))[:10]
        print(f"  OpIdx={op_idx} {ex_date}: R104={r104:.1f} -(int+prin)={check:.1f} match={abs(r104-check)<1:.0f}")

    print("\nQ6: Python over/under paying analysis")
    print("  Comparing Python SHL principal vs Excel SHL principal per period:")
    for op_idx in range(26, 37):
        ex = excel[op_idx]
        py = python_all.get(op_idx, {})
        ex_prin = ex.get("shl_prin", 0) or 0
        py_prin = py.get("shl_principal", 0) or 0
        ex_date = str(ex.get("date", ""))[:10]
        delta = py_prin - ex_prin
        print(f"  OpIdx={op_idx} {ex_date}: Excel_prin={ex_prin:.1f} Py_prin={py_prin:.1f} Delta={delta:.1f}")


if __name__ == "__main__":
    main()