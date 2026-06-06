from __future__ import annotations

import os
import sqlite3
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest


REPO_ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture
def test_db(monkeypatch):
    temp_root = REPO_ROOT / "app" / "data"
    temp_root.mkdir(exist_ok=True)
    db_file = temp_root / f"phase57a10a_{uuid.uuid4().hex[:8]}.db"
    monkeypatch.setenv("FINCO_DB_PATH", str(db_file))
    import app.persistence.db as db_mod

    monkeypatch.setattr(db_mod, "DB_PATH", str(db_file))
    db_mod.init_db()
    try:
        yield db_file
    finally:
        for suffix in ("", "-wal", "-shm"):
            try:
                (Path(str(db_file) + suffix)).unlink(missing_ok=True)
            except PermissionError:
                pass


def _snapshot(**overrides):
    data = {
        "active_project": "pilot-capex-comments",
        "project_name": "Capex Comments Project",
        "project_type": "Solar",
        "project_origin": "user_created",
        "template_source": "generic_solar",
        "country_market": "Croatia",
        "scenario": "Base",
        "capacity_mw": "50",
        "cod_date": "2027-01-01",
        "construction_months": "12",
        "horizon_years": "25",
        "tariff_eur_mwh": "60",
        "ppa_term_years": "15",
        "p50_hours": "1400",
        "opex_y1_keur": "1000",
        "total_capex_keur": "50000",
        "gearing_pct": "70",
        "interest_rate_pct": "5",
        "tenor_years": "15",
        "target_dscr": "1.30",
    }
    data.update(overrides)
    return data


def _load_wb(excel_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)


def _sheet_rows_by_business_code(ws) -> list[dict[str, object]]:
    headers = None
    found_header = False
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not found_header:
            if "Business Code" in values:
                headers = values
                found_header = True
            continue
        if all(v in (None, "") for v in values):
            continue
        if values[0] == "Note":
            continue
        record = {
            str(headers[idx]): values[idx] if idx < len(values) else None
            for idx in range(len(headers))
            if headers[idx] is not None
        }
        if record.get("Business Code") in (None, ""):
            continue
        rows.append(record)
    assert headers is not None
    return rows


def _make_user_project(user_id: str = "u57a10a"):
    from app.persistence.repository import create_project_record

    return create_project_record(
        user_id=user_id,
        project_code="pilot-capex-comments",
        project_name="Capex Comments Project",
        project_type="Solar",
        project_origin="user_project",
        template_source="generic_solar",
        baseline_snapshot=_snapshot(),
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
    )


def _create_sub_line(
    project_id: str,
    *,
    parent: str = "C.02",
    label: str = "Extra EPC",
    amount: float = 1000.0,
    comments: str = "",
):
    from app.persistence.capex_sub_lines import create_sub_line

    with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
        conn.row_factory = sqlite3.Row
        sub = create_sub_line(
            conn.cursor(),
            project_id=project_id,
            parent_category_code=parent,
            label=label,
            amount_keur=amount,
            comments=comments,
        )
        conn.commit()
        return sub


def _list_active_sub_lines(project_id: str):
    from app.persistence.capex_sub_lines import get_active_sub_lines_for_project

    return tuple(get_active_sub_lines_for_project(project_id))


def _make_scenario_a(project, user_id: str = "u57a10a"):
    from app.persistence.repository import add_scenario, seed_scenarios_if_needed

    base = seed_scenarios_if_needed(
        user_id=user_id,
        project_id=project.project_id,
        project_code=project.project_code,
        project_type="Solar",
        source_project_template="generic_solar",
        baseline_snapshot=_snapshot(),
        governance_state={"g20_status": "BLOCKED", "r99_r102_status": "NOT_APPROVED"},
        template_origin="generic_solar",
    )
    scenario = add_scenario(
        user_id=user_id,
        project_id=project.project_id,
        project_code=project.project_code,
        scenario_name="Scenario A",
        parent_scenario_id=base.scenario_id,
        base_input_set=base.base_input_set,
        overrides={},
    )
    return base, scenario


def _capex_amounts_tuple(capex) -> tuple[tuple[str, float], ...]:
    return tuple(
        (field, getattr(capex, field).amount_keur)
        for field in capex._CAPEX_ITEM_FIELDS
    )


class TestPhase57A10ACapexMetadataSafeColumns:
    def test_comments_already_persist_and_load_back_unchanged(self, test_db):
        project = _make_user_project()
        sub = _create_sub_line(
            project.project_id,
            comments="Owner note: cable trench scope still under review.",
        )

        loaded = _list_active_sub_lines(project.project_id)
        assert len(loaded) == 1
        assert loaded[0].sub_line_id == sub.sub_line_id
        assert loaded[0].comments == "Owner note: cable trench scope still under review."

    def test_comments_survive_scenario_duplication_without_new_sub_line_rows(self, test_db):
        from app.persistence.repository import duplicate_scenario

        project = _make_user_project()
        sub = _create_sub_line(
            project.project_id,
            comments="Scenario-independent project comment.",
        )
        _base, scenario_a = _make_scenario_a(project)

        scenario_b = duplicate_scenario("u57a10a", scenario_a.scenario_id, "Scenario B")

        assert scenario_b is not None
        loaded = _list_active_sub_lines(project.project_id)
        assert len(loaded) == 1
        assert loaded[0].sub_line_id == sub.sub_line_id
        assert loaded[0].comments == "Scenario-independent project comment."

    def test_comments_appear_in_capex_sub_lines_audit_export(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        project = _make_user_project()
        sub = _create_sub_line(
            project.project_id,
            comments="Comment carried into export audit.",
        )

        demo = run_demo_project("Solar", "Base")
        wb = _load_wb(
            build_excel_export(
                result=demo.result,
                project_inputs=demo.project_inputs,
                provenance_metadata={"project_id": project.project_id},
            )
        )
        rows = _sheet_rows_by_business_code(wb["CapEx_SubLines_Audit"])
        row = next(r for r in rows if r["Business Code"] == sub.business_code)
        assert row["Comments"] == "Comment carried into export audit."

    def test_comments_do_not_affect_run_materialization_or_capex_totals(self, test_db):
        from app.persistence.capex_sub_lines import CapexSubLine, replace_sub_lines_for_project
        from app.services.capex_sub_lines_integration import _apply_user_sub_lines_to_capex
        from app.ui_runner import run_demo_project

        project = _make_user_project()
        sub = _create_sub_line(
            project.project_id,
            comments="First comment.",
        )

        demo = run_demo_project("Solar", "Base")
        folded_a = _apply_user_sub_lines_to_capex(
            demo.project_inputs.capex,
            project_id=project.project_id,
            scenario_overrides={},
        )

        updated = CapexSubLine(
            id=sub.id,
            sub_line_id=sub.sub_line_id,
            project_id=sub.project_id,
            parent_category_code=sub.parent_category_code,
            business_code=sub.business_code,
            display_order=sub.display_order,
            label=sub.label,
            amount_keur=sub.amount_keur,
            comments="Second comment only; amount unchanged.",
            schedule_json=sub.schedule_json,
            source=sub.source,
            is_active=True,
            governance_state=sub.governance_state,
            replay_metadata=sub.replay_metadata,
            created_at=sub.created_at,
            updated_at=sub.updated_at,
        )
        with sqlite3.connect(os.environ["FINCO_DB_PATH"]) as conn:
            conn.row_factory = sqlite3.Row
            replace_sub_lines_for_project(
                conn.cursor(),
                project_id=project.project_id,
                sub_lines=[updated],
            )
            conn.commit()

        folded_b = _apply_user_sub_lines_to_capex(
            demo.project_inputs.capex,
            project_id=project.project_id,
            scenario_overrides={},
        )

        assert _capex_amounts_tuple(folded_a) == _capex_amounts_tuple(folded_b)

    def test_factory_tuho_and_oborovo_exports_remain_unchanged(self, test_db):
        from app.excel_export import build_excel_export
        from app.ui_runner import run_demo_project

        tuho = run_demo_project("TUHO", "Base")
        oborovo = run_demo_project("Oborovo", "Base")

        tuho_wb = _load_wb(
            build_excel_export(
                result=tuho.result,
                project_inputs=tuho.project_inputs,
                provenance_metadata={},
            )
        )
        oborovo_wb = _load_wb(
            build_excel_export(
                result=oborovo.result,
                project_inputs=oborovo.project_inputs,
                provenance_metadata={},
            )
        )

        assert "CapEx_SubLines_Audit" not in tuho_wb.sheetnames
        assert "CapEx_SubLines_Audit" not in oborovo_wb.sheetnames
