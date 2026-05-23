"""Tests for Phase 11 institutional export product polish."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

import openpyxl

from app.export.calibration_reconciliation import write_calibration_reconciliation_pack


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
DOC = ROOT / "docs" / "phase11_institutional_export_product_polish.md"
WORKBOOK = REPORTS / "phase10_calibration_reconciliation_pack.xlsx"
CHECKLIST = REPORTS / "phase11_export_polish_checklist.csv"
FORMATTING = REPORTS / "phase11_workbook_formatting_matrix.csv"
PATHS = REPORTS / "phase11_reviewer_navigation_paths.csv"


def _generate() -> None:
    write_calibration_reconciliation_pack()


def _sheet_text(workbook, sheet_name: str) -> str:
    ws = workbook[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def test_cover_sheet_enhanced():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Cover"]
    text = _sheet_text(wb, "Cover")
    assert ws["A1"].value == "Institutional Review Pack"
    assert "Workbook Version" in text
    assert "Important Reviewer Notes" in text


def test_navigation_links_exist():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Navigation"]
    links = [str(cell.hyperlink.target) for row in ws.iter_rows() for cell in row if cell.hyperlink]
    assert any("Cover" in link for link in links)
    assert any("Executive Dashboard" in link for link in links)


def test_workbook_metadata_standardized():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK)
    cover = _sheet_text(wb, "Cover")
    runtime = _sheet_text(wb, "Runtime Summary")
    assert "Workbook Version" in cover
    assert "Runtime / Review" in cover
    assert "Runtime / Preview" in runtime


def test_print_settings_applied():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK)
    for sheet_name in ["Cover", "Navigation", "Executive Dashboard", "Gap Register"]:
        ws = wb[sheet_name]
        assert ws.print_title_rows is not None
        assert ws.page_setup.fitToWidth == 1


def test_freeze_panes_standardized():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK)
    assert wb["Cover"].freeze_panes == "A5"
    assert wb["Navigation"].freeze_panes == "A20"
    assert wb["Executive Dashboard"].freeze_panes == "A8"


def test_severity_styling_framework_exists():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Navigation")
    assert "PASS" in text
    assert "ACCEPTED_CONVENTION" in text
    assert "GOVERNANCE_BLOCKER" in text


def test_reviewer_workflow_sheets_still_exist():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "Review Signoff" in wb.sheetnames
    assert "Governance Timeline" in wb.sheetnames
    assert "Readiness Matrix" in wb.sheetnames


def test_executive_dashboard_still_exists():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "Executive Dashboard" in wb.sheetnames
    text = _sheet_text(wb, "Executive Dashboard")
    assert "Review recommendation summary" in text


def test_governance_banners_preserved():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    governance = _sheet_text(wb, "Governance")
    assert "G20 status BLOCKED" in governance
    assert "R99/R102 status NOT APPROVED" in governance


def test_runtime_vs_preview_distinction_preserved():
    _generate()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    text = _sheet_text(wb, "Runtime Summary")
    assert "Runtime / Preview" in text


def test_phase11_reports_exist():
    _generate()
    assert CHECKLIST.exists()
    assert FORMATTING.exists()
    assert PATHS.exists()
    assert _csv_rows(CHECKLIST)
    assert _csv_rows(FORMATTING)
    assert _csv_rows(PATHS)


def test_docs_and_scope_constraints_hold():
    text = DOC.read_text(encoding="utf-8")
    assert "no runtime changes statement" in text.lower()
    assert "G20" in text and "BLOCKED" in text
    assert "R99/R102" in text and "NOT APPROVED" in text
    assert "No runtime/model formulas are changed" in text


def test_no_runtime_model_formula_files_changed():
    result = subprocess.run(
        ["git", "diff", "--name-only"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    changed = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    forbidden_prefixes = (
        "domain/waterfall",
        "domain/shl",
        "domain/tax",
        "app/waterfall",
        "app/ui_runner.py",
        "app/project_factories.py",
    )
    offenders = [path for path in changed if path.replace("\\", "/").startswith(forbidden_prefixes)]
    assert not offenders, offenders
