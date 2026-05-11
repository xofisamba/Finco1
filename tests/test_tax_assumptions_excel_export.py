"""Phase 6D.1 — Tests for tax assumptions Excel export helper."""
import os
import tempfile

import pytest

from domain.tax.templates.inputs import (
    CITTier,
    TaxDepreciationRule,
    TaxTemplate,
    TaxTemplateOverride,
    ResolvedTaxConfig,
)
from domain.tax.templates.resolver import resolve_tax_template
from app.tax_assumptions_excel_export import write_tax_assumptions_audit_sheets


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def flat_template():
    return TaxTemplate(
        country_code="HR",
        template_name="HR Flat 18%",
        tax_year=2026,
        cit_tiers=(CITTier(min_profit_keur=0.0, max_profit_keur=None, tax_rate=0.18),),
        depreciation_rules=(
            TaxDepreciationRule(
                asset_category="buildings", method="straight_line",
                annual_rate=0.05, useful_life_years=20.0, max_deductible_rate=0.025,
                bonus_depreciation_pct=0.0, deductible=True, notes="20yr buildings",
            ),
        ),
        withholding_tax_dividends=0.12,
        withholding_tax_interest=0.0,
        loss_carryforward_years=5,
        thin_cap_ratio=4.0,
        interest_limitation_pct_ebitda=0.30,
        metadata=(("notes", "standard HR"),),
    )


@pytest.fixture
def progressive_template():
    return TaxTemplate(
        country_code="ME",
        template_name="ME Progressive",
        tax_year=2026,
        cit_tiers=(
            CITTier(min_profit_keur=0.0, max_profit_keur=500.0, tax_rate=0.10),
            CITTier(min_profit_keur=500.0, max_profit_keur=None, tax_rate=0.15),
        ),
        depreciation_rules=(),
        withholding_tax_dividends=0.0,
        withholding_tax_interest=0.0,
        loss_carryforward_years=None,
        thin_cap_ratio=None,
        interest_limitation_pct_ebitda=None,
        metadata=(),
    )


@pytest.fixture
def resolved_config(flat_template):
    override = TaxTemplateOverride(
        override_name="custom_wht",
        field_path="withholding_tax_dividends",
        override_value=0.10,
        reason="Treaty rate",
    )
    return resolve_tax_template(flat_template, (override,))


# ── Tests ────────────────────────────────────────────────────────────────────

class TestNoOpBehavior:
    def test_all_empty_is_noop(self):
        """Empty templates + configs + overrides creates no sheets."""
        import pandas as pd
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        wb = Workbook()
        wb.active.title = "PreExisting"
        wb.save(tmp_path)
        wb.close()

        writer = pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a")
        write_tax_assumptions_audit_sheets(writer, templates=(), resolved_configs=(), overrides=())
        writer.close()

        from openpyxl import load_workbook as _load; wb2 = _load(tmp_path)
        tax_sheets = [s for s in wb2.sheetnames if "Tax" in s]
        assert len(tax_sheets) == 0
        wb2.close()
        os.unlink(tmp_path)

    def test_empty_templates_only_is_noop(self):
        """Empty templates tuple with no overrides/configs is no-op."""
        import pandas as pd
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        wb = Workbook()
        wb.active.title = "PreExisting"
        wb.save(tmp_path)
        wb.close()

        writer = pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a")
        write_tax_assumptions_audit_sheets(writer, templates=())
        writer.close()

        from openpyxl import load_workbook as _load; wb2 = _load(tmp_path)
        tax_sheets = [s for s in wb2.sheetnames if "Tax" in s]
        assert len(tax_sheets) == 0
        wb2.close()
        os.unlink(tmp_path)


class TestTaxTemplatesSheet:
    def test_creates_tax_templates_sheet(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Tax Templates" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_audit_note_row_1(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb["Tax Templates"]
        first_cell = ws.cell(row=1, column=1).value
        assert "AUDIT-ONLY" in str(first_cell)
        wb.close()
        os.unlink(tmp_path)

    def test_headers_row_2(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb["Tax Templates"]
        header = ws.cell(row=2, column=1).value
        assert "Template Code" in str(header)
        wb.close()
        os.unlink(tmp_path)


class TestTaxTiersSheet:
    def test_creates_tax_tiers_sheet(self, progressive_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(progressive_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Tax Tiers" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_tiers_data_rows(self, progressive_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(progressive_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb["Tax Tiers"]
        # Should have audit note row + headers row + 2 data rows = 4 rows
        assert ws.max_row >= 4
        wb.close()
        os.unlink(tmp_path)


class TestTaxDepRulesSheet:
    def test_creates_tax_dep_rules_sheet(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Tax Dep Rules" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)


class TestTaxOverridesSheet:
    def test_creates_overrides_sheet_when_overrides_provided(self, flat_template):
        import pandas as pd
        override = TaxTemplateOverride(
            override_name="custom_wht",
            field_path="withholding_tax_dividends",
            override_value=0.10,
            reason="Treaty rate",
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,), overrides=(override,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Tax Overrides" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_omits_overrides_sheet_when_no_overrides(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,), overrides=())

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Tax Overrides" not in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)


class TestResolvedTaxConfigSheet:
    def test_creates_resolved_config_sheet(self, resolved_config):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, resolved_configs=(resolved_config,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Resolved Tax Config" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_omits_resolved_config_sheet_when_empty(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(flat_template,), resolved_configs=())

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Resolved Tax Config" not in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)



@pytest.fixture
def template_no_tiers():
    """Template with no CIT tiers."""
    return TaxTemplate(
        country_code="XX",
        template_name="XX No Tiers",
        tax_year=2026,
        cit_tiers=(),
        depreciation_rules=(
            TaxDepreciationRule(
                asset_category="equipment", method="straight_line",
                annual_rate=0.20, useful_life_years=5.0, max_deductible_rate=None,
                bonus_depreciation_pct=0.0, deductible=True, notes="equipment only",
            ),
        ),
        withholding_tax_dividends=0.0,
        withholding_tax_interest=0.0,
        loss_carryforward_years=None,
        thin_cap_ratio=None,
        interest_limitation_pct_ebitda=None,
        metadata=(),
    )


@pytest.fixture
def template_no_dep_rules():
    """Template with no depreciation rules."""
    return TaxTemplate(
        country_code="YY",
        template_name="YY No Dep",
        tax_year=2026,
        cit_tiers=(CITTier(min_profit_keur=0.0, max_profit_keur=None, tax_rate=0.15),),
        depreciation_rules=(),
        withholding_tax_dividends=0.0,
        withholding_tax_interest=0.0,
        loss_carryforward_years=None,
        thin_cap_ratio=None,
        interest_limitation_pct_ebitda=None,
        metadata=(),
    )


class TestEdgeCases:
    """Edge-case coverage for tax assumptions export."""

    def test_template_no_cit_tiers_does_not_crash(self, template_no_tiers):
        """Template with no CIT tiers does not crash export."""
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(template_no_tiers,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        # Tax Templates sheet should exist (it has no tiers)
        assert "Tax Templates" in wb.sheetnames
        # Tax Tiers should NOT be created (all templates have empty tiers)
        assert "Tax Tiers" not in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_template_no_depreciation_rules_does_not_crash(self, template_no_dep_rules):
        """Template with no depreciation rules does not crash export."""
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(writer, templates=(template_no_dep_rules,))

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        # Tax Templates and Tax Tiers should exist
        assert "Tax Templates" in wb.sheetnames
        assert "Tax Tiers" in wb.sheetnames
        # Tax Dep Rules should NOT be created
        assert "Tax Dep Rules" not in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_mixed_templates_produces_tiers_sheet(self, flat_template, template_no_tiers):
        """When at least one template has tiers, tiers sheet is created."""
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(
                writer,
                templates=(flat_template, template_no_tiers),
            )

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        # flat_template has tiers → Tax Tiers IS created
        assert "Tax Tiers" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_mixed_templates_produces_dep_sheet(self, flat_template, template_no_dep_rules):
        """When at least one template has dep rules, dep sheet is created."""
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(
                writer,
                templates=(flat_template, template_no_dep_rules),
            )

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        # flat_template has dep rules → Tax Dep Rules IS created
        assert "Tax Dep Rules" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_empty_resolved_frames_do_not_crash(self, flat_template):
        """Empty resolved config frames do not crash export."""
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        # resolved_config but with no resolved frames collected
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(
                writer,
                templates=(flat_template,),
                resolved_configs=(),  # empty
            )

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert "Resolved Tax Config" not in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)



class TestExcelValidity:
    def test_excel_opens_successfully(self, flat_template, progressive_template, resolved_config):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        override = TaxTemplateOverride(
            override_name="custom_wht",
            field_path="withholding_tax_dividends",
            override_value=0.10,
            reason="Treaty",
        )

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(
                writer,
                templates=(flat_template, progressive_template),
                resolved_configs=(resolved_config,),
                overrides=(override,),
            )

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        assert len(wb.sheetnames) > 0
        wb.close()
        os.unlink(tmp_path)

    def test_optional_sheets_all_present_when_data_provided(self, flat_template):
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        override = TaxTemplateOverride(
            override_name="custom_wht",
            field_path="withholding_tax_dividends",
            override_value=0.10,
            reason="Treaty",
        )
        resolved = resolve_tax_template(flat_template, (override,))

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_tax_assumptions_audit_sheets(
                writer,
                templates=(flat_template,),
                resolved_configs=(resolved,),
                overrides=(override,),
            )

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        expected = {"Tax Templates", "Tax Tiers", "Tax Dep Rules", "Tax Overrides", "Resolved Tax Config"}
        found = set(wb.sheetnames) & expected
        assert found == expected, f"Missing sheets: {expected - found}"
        wb.close()
        os.unlink(tmp_path)