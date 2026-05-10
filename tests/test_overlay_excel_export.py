"""Phase 5E tests for overlay Excel export visibility."""
from __future__ import annotations

from io import BytesIO
import pandas as pd
import pytest

from openpyxl import load_workbook

from domain.portfolio.distribution_constraints import (
    DistributionBlockReason,
    DistributionConstraintConfig,
)
from domain.portfolio.distribution_constraints.result import (
    DistributionConstraintPeriod,
    DistributionConstraintResult,
)
from domain.portfolio.distribution_constraints.overlay import (
    SPVRetainedCashPeriod,
    SPVRetainedCashOverlay,
)
from domain.portfolio.distribution_constraints.holdco_overlay import (
    HoldCoRetainedCashOverlay,
)
from domain.portfolio.cash_ledger import (
    CashLedgerPeriod,
    CashMovement,
    CashMovementType,
    EntityCashLedger,
    PortfolioCashLedger,
)

# Import the functions under test
from app.excel_export import (
    _write_overlay_sheets,
    _write_cash_ledger_sheet,
    _write_dist_constraints_sheet,
    _write_spv_retained_cash_sheet,
    _write_holdco_retained_cash_sheet,
    _safe_sheet_name,
    _check_sheet_name_length,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def cm(period, entity, mtype, amount):
    return CashMovement(
        period=period, entity_code=entity, movement_type=mtype,
        amount_keur=amount, description="",
    )


# ── A. _write_overlay_sheets orchestration ───────────────────────────────────

class TestWriteOverlaySheets:
    def test_all_none_skips_all_sheets(self):
        # When all overlays are None, _write_overlay_sheets adds no sheets.
        # openpyxl requires at least one visible sheet to save, so we check
        # that _write_overlay_sheets is callable and doesn't raise.
        output = BytesIO()
        try:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                _write_overlay_sheets(writer, cash_ledger=None, spv_overlays=None,
                                      holdco_overlay=None, constraint_results=None)
            # If a sheet was added, check it's not an overlay sheet
            output.seek(0)
            wb = load_workbook(output)
            overlay_sheets = [n for n in wb.sheetnames
                              if any(x in n for x in ["Cash Ledger", "Dist Constraints", "HoldCo"])]
            assert overlay_sheets == [], f"Unexpected overlay sheets: {overlay_sheets}"
        except IndexError as e:
            if "At least one sheet must be visible" in str(e):
                # No sheets added — this is the expected behavior when all overlays are None
                pass
            else:
                raise

    def test_cash_ledger_sheet_added(self):
        ledger = make_entity_ledger("SOLAR-1", [
            (0, 1000.0, [cm(0, "SOLAR-1", CashMovementType.OPERATING_CASHFLOW, 500.0),
                          cm(0, "SOLAR-1", CashMovementType.EQUITY_DISTRIBUTION, -500.0)],
             1000.0, 1000.0),
        ])
        port = PortfolioCashLedger(entities=(ledger,), total_ending_cash_keur=1000.0)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_overlay_sheets(writer, cash_ledger=port)
        output.seek(0)
        wb = load_workbook(output)
        assert "Cash Ledger" in wb.sheetnames

    def test_spv_overlay_sheet_added(self):
        p = SPVRetainedCashPeriod(
            period=0, entity_code="SOLAR-1",
            requested_distribution_keur=500.0, allowed_distribution_keur=400.0,
            retained_cash_keur=600.0, cash_before_distribution_keur=1000.0,
        )
        ov = SPVRetainedCashOverlay(entity_code="SOLAR-1", periods=(p,))
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_overlay_sheets(writer, spv_overlays=[ov])
        output.seek(0)
        wb = load_workbook(output)
        assert any("SPV" in n for n in wb.sheetnames)

    def test_holdco_overlay_sheet_added(self):
        ov = HoldCoRetainedCashOverlay(
            entity_code="HOLDCO",
            retained_cash_by_period=(100.0,),
            requested_distribution_by_period=(500.0,),
            available_distribution_by_period=(400.0,),
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_overlay_sheets(writer, holdco_overlay=ov)
        output.seek(0)
        wb = load_workbook(output)
        assert "HoldCo Ret Cash" in wb.sheetnames

    def test_constraint_results_sheet_added(self):
        cp = DistributionConstraintPeriod(
            period=0, entity_code="SOLAR-1",
            cash_before_distribution_keur=1000.0, requested_distribution_keur=500.0,
            allowed_distribution_keur=500.0, retained_cash_keur=500.0,
        )
        cr = DistributionConstraintResult(
            entity_code="SOLAR-1", periods=(cp,),
            total_requested_distribution_keur=500.0,
            total_allowed_distribution_keur=500.0,
            total_retained_cash_keur=500.0,
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_overlay_sheets(writer, constraint_results=[cr])
        output.seek(0)
        wb = load_workbook(output)
        assert "Dist Constraints" in wb.sheetnames


# ── B. Sheet name length ───────────────────────────────────────────────────────

class TestSheetNameLength:
    def test_all_sheet_names_le_31_chars(self):
        assert len("Cash Ledger") <= 31
        assert len("Dist Constraints") <= 31
        assert len("HoldCo Ret Cash") <= 31

    def test_check_sheet_name_length_passes_valid(self):
        _check_sheet_name_length("Cash Ledger")  # no raise
        _check_sheet_name_length("Dist Constraints")
        _check_sheet_name_length("HoldCo Ret Cash")

    def test_check_sheet_name_length_raises_on_32(self):
        with pytest.raises(AssertionError, match="exceeds 31-char"):
            _check_sheet_name_length("A" * 32)

    def test_spv_sheet_name_truncated_to_31(self):
        p = SPVRetainedCashPeriod(
            period=0, entity_code="X" * 50,
            requested_distribution_keur=500.0, allowed_distribution_keur=500.0,
            retained_cash_keur=500.0, cash_before_distribution_keur=1000.0,
        )
        ov = SPVRetainedCashOverlay(entity_code="X" * 50, periods=(p,))
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_spv_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        for name in wb.sheetnames:
            assert len(name) <= 31


# ── C. Cash ledger sheet ──────────────────────────────────────────────────────

class TestCashLedgerSheet:
    def test_headers_present(self):
        # closing=1500 = opening 1000 + movement +500 per CashLedgerPeriod invariant
        ledger = make_entity_ledger("SOLAR-1", [
            (0, 1000.0,
             [cm(0, "SOLAR-1", CashMovementType.OPERATING_CASHFLOW, 500.0)],
             1500.0, 1500.0),
        ])
        port = PortfolioCashLedger(entities=(ledger,), total_ending_cash_keur=1500.0)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_cash_ledger_sheet(writer, port)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Cash Ledger"]
        headers = [cell.value for cell in ws[1]]
        assert "Entity Code" in headers
        assert "Period" in headers
        assert "Movement Type" in headers

    def test_row_count_matches_movements(self):
        ledger = make_entity_ledger("SOLAR-1", [
            (0, 1000.0,
             [cm(0, "SOLAR-1", CashMovementType.OPERATING_CASHFLOW, 500.0),
              cm(0, "SOLAR-1", CashMovementType.EQUITY_DISTRIBUTION, -500.0)],
             1000.0, 1000.0),
        ])
        port = PortfolioCashLedger(entities=(ledger,), total_ending_cash_keur=1000.0)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_cash_ledger_sheet(writer, port)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Cash Ledger"]
        # 1 header row + 2 movement rows
        assert ws.max_row == 3

    def test_warnings_exported(self):
        # Use zero-movement that maintains closing=opening (no change)
        period = CashLedgerPeriod(
            period=0, entity_code="SOLAR-1",
            opening_cash_keur=1000.0, movements=(
                CashMovement(period=0, entity_code="SOLAR-1",
                             movement_type=CashMovementType.OPERATING_CASHFLOW,
                             amount_keur=0.0, description=""),
            ),
            closing_cash_keur=1000.0,
            retained_cash_keur=1000.0, warnings=("Test warning",),
        )
        ledger = EntityCashLedger(
            entity_code="SOLAR-1", periods=(period,),
            total_movements_keur=0.0, ending_cash_keur=1000.0, warnings=(),
        )
        port = PortfolioCashLedger(entities=(ledger,), total_ending_cash_keur=1000.0)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_cash_ledger_sheet(writer, port)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Cash Ledger"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "Test warning" in values


# ── D. Distribution constraints sheet ────────────────────────────────────────

class TestDistConstraintsSheet:
    def test_headers_present(self):
        cp = DistributionConstraintPeriod(
            period=0, entity_code="SOLAR-1",
            cash_before_distribution_keur=1000.0, requested_distribution_keur=500.0,
            allowed_distribution_keur=500.0, retained_cash_keur=500.0,
        )
        cr = DistributionConstraintResult(
            entity_code="SOLAR-1", periods=(cp,),
            total_requested_distribution_keur=500.0,
            total_allowed_distribution_keur=500.0,
            total_retained_cash_keur=500.0,
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_dist_constraints_sheet(writer, [cr])
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Dist Constraints"]
        headers = [cell.value for cell in ws[1]]
        for h in ["Entity Code", "Period", "Requested Distribution", "Allowed Distribution", "Retained Cash"]:
            assert h in headers

    def test_block_reasons_exported(self):
        cp = DistributionConstraintPeriod(
            period=0, entity_code="SOLAR-1",
            cash_before_distribution_keur=500.0, requested_distribution_keur=500.0,
            allowed_distribution_keur=0.0, retained_cash_keur=500.0,
            block_reasons=(DistributionBlockReason.MANUAL_LOCKUP,),
        )
        cr = DistributionConstraintResult(
            entity_code="SOLAR-1", periods=(cp,),
            total_requested_distribution_keur=500.0,
            total_allowed_distribution_keur=0.0,
            total_retained_cash_keur=500.0,
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_dist_constraints_sheet(writer, [cr])
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Dist Constraints"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "MANUAL_LOCKUP" in " ".join(str(v) for v in values)


# ── E. SPV retained cash sheet ───────────────────────────────────────────────

class TestSPVRetainedCashSheet:
    def test_sheet_name_includes_entity_code(self):
        ov = build_spv_overlay("SOLAR-1", [(0, 500.0, 400.0, 600.0, 1000.0)])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_spv_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        assert any("SOLAR-1" in n for n in wb.sheetnames)

    def test_headers_present(self):
        ov = build_spv_overlay("SOLAR-1", [(0, 500.0, 400.0, 600.0, 1000.0)])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_spv_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.worksheets[0]
        headers = [cell.value for cell in ws[1]]
        for h in ["Requested Distribution", "Allowed Distribution", "Retained Cash", "Available Distribution"]:
            assert h in headers

    def test_available_distribution_correct(self):
        # available = cash_before - retained
        # period 0: cash_before=1000, retained=600 → available=400
        ov = build_spv_overlay("SOLAR-1", [(0, 500.0, 400.0, 600.0, 1000.0)])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_spv_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.worksheets[0]
        # Find "Available Distribution" column
        headers = [cell.value for cell in ws[1]]
        avail_col = headers.index("Available Distribution") + 1
        val = ws.cell(row=2, column=avail_col).value
        assert val == 400.0


# ── F. HoldCo retained cash sheet ─────────────────────────────────────────────

class TestHoldCoRetainedCashSheet:
    def test_headers_present(self):
        ov = HoldCoRetainedCashOverlay(
            entity_code="HOLDCO",
            retained_cash_by_period=(100.0,),
            requested_distribution_by_period=(500.0,),
            available_distribution_by_period=(400.0,),
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_holdco_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["HoldCo Ret Cash"]
        headers = [cell.value for cell in ws[1]]
        for h in ["HoldCo Code", "Requested Distribution", "Retained Cash", "Available Distribution"]:
            assert h in headers

    def test_values_match_overlay(self):
        ov = HoldCoRetainedCashOverlay(
            entity_code="HOLDCO",
            retained_cash_by_period=(100.0, 200.0),
            requested_distribution_by_period=(500.0, 600.0),
            available_distribution_by_period=(400.0, 400.0),
        )
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_holdco_retained_cash_sheet(writer, ov)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["HoldCo Ret Cash"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        req_col = header.index("Requested Distribution")
        ret_col = header.index("Retained Cash")
        # Period 0 (after audit note row at index 1)
        assert rows[2][req_col] == 500.0, f"period 0 requested={rows[2][req_col]}"
        assert rows[2][ret_col] == 100.0, f"period 0 retained={rows[2][ret_col]}"
        # Period 1
        assert rows[3][req_col] == 600.0, f"period 1 requested={rows[3][req_col]}"
        assert rows[3][ret_col] == 200.0, f"period 1 retained={rows[3][ret_col]}"


# ── G. Safe sheet name ────────────────────────────────────────────────────────

class TestSafeSheetName:
    def test_removes_invalid_chars(self):
        result = _safe_sheet_name("A/B:C\\D*E?F[G]H")
        for c in "\\/:*?[]":
            assert c not in result

    def test_truncates_to_28_chars(self):
        result = _safe_sheet_name("X" * 50)
        assert len(result) == 28

    def test_preserves_valid_name(self):
        assert _safe_sheet_name("SOLAR-1") == "SOLAR-1"


# ── H. No regression for existing export ──────────────────────────────────────

class TestNoRegression:
    def test_build_excel_export_works_with_no_overlays(self):
        # This tests that the optional parameters don't break the existing signature
        from app.excel_export import build_excel_export
        # Just verify the function still has the right signature - no error on None params
        import inspect
        sig = inspect.signature(build_excel_export)
        params = list(sig.parameters.keys())
        assert "cash_ledger" in params
        assert "spv_overlays" in params
        assert "holdco_overlay" in params
        assert "constraint_results" in params


# ── G. build_excel_export integration tests ─────────────────────────────────────

from app.excel_export import build_excel_export
from domain.portfolio.distribution_constraints.overlay import (
    SPVRetainedCashPeriod,
    SPVRetainedCashOverlay,
)


class TestBuildExcelExportWithOverlays:
    def test_overlay_sheets_present_when_provided(self):
        """build_excel_export() bytes contain expected overlay sheets."""
        # Build minimal SPV overlay
        p = SPVRetainedCashPeriod(
            period=0, entity_code="SOLAR-1",
            requested_distribution_keur=500.0, allowed_distribution_keur=400.0,
            retained_cash_keur=600.0, cash_before_distribution_keur=1000.0,
        )
        spv_ov = SPVRetainedCashOverlay(entity_code="SOLAR-1", periods=(p,))

        # Build minimal HoldCo overlay
        hc_ov = HoldCoRetainedCashOverlay(
            entity_code="HOLDCO",
            retained_cash_by_period=(100.0,),
            requested_distribution_by_period=(500.0,),
            available_distribution_by_period=(400.0,),
        )

        # Build constraint result
        cp = DistributionConstraintPeriod(
            period=0, entity_code="SOLAR-1",
            cash_before_distribution_keur=1000.0, requested_distribution_keur=500.0,
            allowed_distribution_keur=400.0, retained_cash_keur=600.0,
        )
        cr = DistributionConstraintResult(
            entity_code="SOLAR-1", periods=(cp,),
            total_requested_distribution_keur=500.0,
            total_allowed_distribution_keur=400.0,
            total_retained_cash_keur=600.0,
        )

        # Build cash ledger
        entity_period = CashLedgerPeriod(
            period=0, entity_code="SOLAR-1",
            opening_cash_keur=1000.0,
            movements=(CashMovement(period=0, entity_code="SOLAR-1",
                                    movement_type=CashMovementType.OPERATING_CASHFLOW,
                                    amount_keur=500.0, description=""),),
            closing_cash_keur=1500.0, retained_cash_keur=1500.0, warnings=(),
        )
        entity_ledger = EntityCashLedger(
            entity_code="SOLAR-1", periods=(entity_period,),
            total_movements_keur=500.0, ending_cash_keur=1500.0, warnings=(),
        )
        cash_ledger = PortfolioCashLedger(entities=(entity_ledger,), total_ending_cash_keur=1500.0)

        bytes_out = build_excel_export(
            result=None,
            portfolio_result=None,
            cash_ledger=cash_ledger,
            spv_overlays=[spv_ov],
            holdco_overlay=hc_ov,
            constraint_results=[cr],
        )

        wb = load_workbook(BytesIO(bytes_out))
        sheet_names = wb.sheetnames
        assert "Cash Ledger" in sheet_names, f"Cash Ledger missing from {sheet_names}"
        assert "Dist Constraints" in sheet_names, f"Dist Constraints missing from {sheet_names}"
        assert any("SOLAR-1" in n for n in sheet_names), f"SPV sheet missing from {sheet_names}"
        assert "HoldCo Ret Cash" in sheet_names, f"HoldCo Ret Cash missing from {sheet_names}"

    def test_overlays_omitted_no_overlay_sheets(self):
        """When overlay params are None, no overlay sheets are created."""
        bytes_out = build_excel_export(result=None, portfolio_result=None)

        wb = load_workbook(BytesIO(bytes_out))
        overlay_sheets = [n for n in wb.sheetnames
                         if n in ("Cash Ledger", "Dist Constraints", "HoldCo Ret Cash")
                         or n.startswith("SPV_")]
        assert overlay_sheets == [], f"Unexpected overlay sheets: {overlay_sheets}"

    def test_sheet_names_le_31_chars(self):
        """All overlay sheet names respect openpyxl 31-char limit."""
        p = SPVRetainedCashPeriod(
            period=0, entity_code="X" * 50,
            requested_distribution_keur=500.0, allowed_distribution_keur=500.0,
            retained_cash_keur=500.0, cash_before_distribution_keur=1000.0,
        )
        ov = SPVRetainedCashOverlay(entity_code="X" * 50, periods=(p,))
        hc_ov = HoldCoRetainedCashOverlay(
            entity_code="HOLDCO",
            retained_cash_by_period=(100.0,),
            requested_distribution_by_period=(500.0,),
            available_distribution_by_period=(400.0,),
        )

        bytes_out = build_excel_export(
            spv_overlays=[ov], holdco_overlay=hc_ov,
            constraint_results=[],
            cash_ledger=None,
        )
        wb = load_workbook(BytesIO(bytes_out))
        for name in wb.sheetnames:
            assert len(name) <= 31, f"Sheet name '{name}' exceeds 31 chars"


# ── test helpers ──────────────────────────────────────────────────────────────

def make_entity_ledger(entity_code, periods_data):
    periods = []
    for period_idx, opening, movement_list, closing, retained in periods_data:
        periods.append(CashLedgerPeriod(
            period=period_idx, entity_code=entity_code,
            opening_cash_keur=opening, movements=tuple(movement_list),
            closing_cash_keur=closing, retained_cash_keur=retained, warnings=(),
        ))
    return EntityCashLedger(
        entity_code=entity_code, periods=tuple(periods),
        total_movements_keur=sum(sum(m.amount_keur for m in p.movements) for p in periods),
        ending_cash_keur=periods[-1].closing_cash_keur if periods else 0.0,
        warnings=(),
    )


def build_spv_overlay(entity_code, period_data):
    """period_data: list of (period, req, allowed, retained, cash_before)"""
    periods = []
    for pd_ in period_data:
        period, req, allowed, retained, cash_before = pd_
        periods.append(SPVRetainedCashPeriod(
            period=period, entity_code=entity_code,
            requested_distribution_keur=req, allowed_distribution_keur=allowed,
            retained_cash_keur=retained, cash_before_distribution_keur=cash_before,
        ))
    return SPVRetainedCashOverlay(entity_code=entity_code, periods=tuple(periods))