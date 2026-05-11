"""Phase 6C.4 — Tests for HoldCo tax Excel export helper."""
import os
import tempfile

import pytest

from domain.tax.holdco_inputs import (
    HoldCoTaxInputs,
    WithholdingTaxConfig,
    InterestDeductibilityConfig,
)
from domain.tax.holdco_runner import run_holdco_tax_engine
from app.holdco_tax_excel_export import write_holdco_tax_audit_sheets


# ── Fixtures ──────────────────────────────────────────────────────────────────

def make_inputs(entity_code="HR-HoldCo-001", n_periods=3, **overrides):
    defaults = dict(
        entity_code=entity_code,
        country_code="HR",
        tax_year=2026,
        dividend_income_by_period_keur=tuple(1000.0 for _ in range(n_periods)),
        shl_interest_income_by_period_keur=tuple(200.0 for _ in range(n_periods)),
        shl_principal_received_by_period_keur=tuple(500.0 for _ in range(n_periods)),
        holdco_opex_by_period_keur=tuple(50.0 for _ in range(n_periods)),
        withholding_tax_config=WithholdingTaxConfig(rate_dividends=0.12, rate_interest=0.0),
        interest_deductibility=InterestDeductibilityConfig(thin_cap_ratio=4.0, interest_limitation_pct_ebitda=0.30),
        metadata=(("source", "test"),),
    )
    defaults.update(overrides)
    return HoldCoTaxInputs(**defaults)


def make_result(**overrides):
    valid_overrides = {}
    for k, v in overrides.items():
        if k in ["entity_code", "n_periods",
                 "dividend_income_by_period_keur", "shl_interest_income_by_period_keur",
                 "shl_principal_received_by_period_keur", "holdco_opex_by_period_keur"]:
            valid_overrides[k] = v
    inputs = make_inputs(**valid_overrides)
    return run_holdco_tax_engine(inputs)


# ── Tests ────────────────────────────────────────────────────────────────────

class TestNoOpBehavior:
    def test_none_is_noop(self):
        """None holdco_tax_results creates no sheets and no error."""
        import pandas as pd
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        wb = Workbook()
        wb.active.title = "PreExisting"
        wb.save(tmp_path)
        wb.close()

        writer = pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a")
        write_holdco_tax_audit_sheets(writer, None)
        writer.close()

        from openpyxl import load_workbook as _load; wb2 = _load(tmp_path)
        tax_sheets = [s for s in wb2.sheetnames if "HoldCo" in s]
        assert len(tax_sheets) == 0
        wb2.close()
        os.unlink(tmp_path)

    def test_empty_tuple_is_noop(self):
        """Empty tuple holdco_tax_results creates no sheets and no error."""
        import pandas as pd
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        wb = Workbook()
        wb.active.title = "PreExisting"
        wb.save(tmp_path)
        wb.close()

        writer = pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a")
        write_holdco_tax_audit_sheets(writer, ())
        writer.close()

        from openpyxl import load_workbook as _load; wb2 = _load(tmp_path)
        tax_sheets = [s for s in wb2.sheetnames if "HoldCo" in s]
        assert len(tax_sheets) == 0
        wb2.close()
        os.unlink(tmp_path)


class TestHoldCoTaxSummarySheet:
    def test_creates_holdco_tax_summary_sheet(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        assert "HoldCo Tax Summary" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_summary_audit_note_row_1(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        ws = wb["HoldCo Tax Summary"]
        first_cell = ws.cell(row=1, column=1).value
        assert "AUDIT-ONLY" in str(first_cell)
        wb.close()
        os.unlink(tmp_path)

    def test_summary_headers_row_2(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        ws = wb["HoldCo Tax Summary"]
        header_cell = ws.cell(row=2, column=1).value
        assert "Entity Code" in str(header_cell)
        wb.close()
        os.unlink(tmp_path)


class TestHoldCoTaxPerEntitySheet:
    def test_creates_per_entity_sheet(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        assert "HoldCoTax_HR-HoldCo-001" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)

    def test_audit_note_row_1_on_per_entity_sheet(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        ws = wb["HoldCoTax_HR-HoldCo-001"]
        first_cell = ws.cell(row=1, column=1).value
        assert "AUDIT-ONLY" in str(first_cell)
        wb.close()
        os.unlink(tmp_path)

    def test_entity_code_with_colon_sanitized(self):
        """Colon in entity code is sanitized, sheet created successfully."""
        result = make_result(entity_code="HR:HC:001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        # Colon should be replaced with underscore
        assert any("HoldCoTax_" in s and ":" not in s for s in wb.sheetnames)
        wb.close()
        os.unlink(tmp_path)

    def test_duplicate_long_names_get_unique_suffix(self):
        """Two entity codes that truncate to the same base get unique names."""
        # Both would truncate to "HoldCoTax_AAAAAAAAAA..." (31 chars)
        code1 = "A" * 25  # "HoldCoTax_" + 25 = 34 → truncated to 31
        code2 = "A" * 24 + "B"  # different last char

        result1 = make_result(entity_code=code1)
        result2 = make_result(entity_code=code2)
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result1, result2))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        holdco_sheets = [s for s in wb.sheetnames if s.startswith("HoldCoTax_")]
        assert len(holdco_sheets) == 2
        assert holdco_sheets[0] != holdco_sheets[1]
        # One should end with _2
        assert any(s.endswith("_2") for s in holdco_sheets)
        wb.close()
        os.unlink(tmp_path)


class TestExcelValidity:
    def test_excel_opens_successfully_after_export(self):
        result = make_result(entity_code="HR-HoldCo-001")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result,))

        # Should not raise
        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        assert len(wb.sheetnames) > 0
        wb.close()
        os.unlink(tmp_path)

    def test_multiple_entities_all_present(self):
        result1 = make_result(entity_code="HR-HoldCo-001")
        result2 = make_result(entity_code="HR-HoldCo-002")
        import pandas as pd

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            write_holdco_tax_audit_sheets(writer, (result1, result2))

        from openpyxl import load_workbook as _load; wb = _load(tmp_path)
        assert "HoldCoTax_HR-HoldCo-001" in wb.sheetnames
        assert "HoldCoTax_HR-HoldCo-002" in wb.sheetnames
        wb.close()
        os.unlink(tmp_path)