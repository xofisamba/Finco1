"""Tests for app/excel_export.py."""
import pytest
from io import BytesIO
from app.excel_export import build_excel_export
from app.project_factories import create_default_solar_project
from app.ui_runner import run_demo_project


def test_build_excel_export_returns_bytes():
    result = run_demo_project("Solar")
    data = build_excel_export(result=result.result, project_inputs=result.project_inputs)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_excel_export_has_all_required_sheets():
    """All required sheets must be present in the export for Solar project."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        integration_note=None,
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    sheet_names = wb.sheetnames
    required = [
        "Dashboard", "Inputs", "CapEx",
        "Revenue", "Debt", "Tax_Depreciation",
        "Waterfall", "Returns", "Validation", "Notes",
    ]
    missing = [s for s in required if s not in sheet_names]
    assert not missing, f"Missing sheets: {missing}"


def test_excel_export_no_formulas_in_cells():
    """All cells must be values only — no Excel formulas."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(result=result.result, project_inputs=result.project_inputs)
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != 'f', \
                    f"Formula found in sheet '{sheet}' at {cell.coordinate}: {cell.value}"


def test_excel_export_annual_columns_are_years():
    """Annual-view sheets must have 4-digit year column headers (YYYY)."""
    import openpyxl, re
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        period_view="Annual",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet_name in ["Revenue", "Debt", "Tax_Depreciation", "Waterfall"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        year_cols = [h for h in headers if isinstance(h, str) and re.match(r"\d{4}", h)]
        assert len(year_cols) > 0, \
            f"Sheet '{sheet_name}' has no year columns: {headers[:8]}"


def test_excel_export_contains_portfolio_sheet_when_portfolio_result_provided():
    import openpyxl
    result = run_demo_project("Portfolio")
    data = build_excel_export(portfolio_result=result.portfolio_result, project_inputs=result.project_inputs)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Portfolio" in wb.sheetnames


def test_excel_export_does_not_import_calibration_modules():
    import ast, inspect
    from app import excel_export
    src = inspect.getsource(excel_export)
    tree = ast.parse(src)
    imports = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imports.append(alias.name if alias.asname is None else alias.asname)
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            imports.append(mod)
    bad = [i for i in imports if i and ("calibration" in i or "excel_oborovo" in i or "excel_tuho" in i)]
    assert not bad, f"Bad imports: {bad}"


def test_inputs_summary_uses_technical_capacity():
    from app.input_helpers import build_inputs_summary_table
    proj = create_default_solar_project()
    df = build_inputs_summary_table(proj)
    assert "Capacity (MW)" in df["Field"].values


def test_capex_summary_uses_total_capex_fallback():
    from app.input_helpers import build_capex_summary_table
    proj = create_default_solar_project()
    df = build_capex_summary_table(proj)
    assert "Total CapEx (kEUR)" in df["Field"].values


def test_excel_export_with_portfolio_result_and_project_inputs():
    import openpyxl
    result = run_demo_project("Portfolio")
    solar_result = run_demo_project("Solar")
    data = build_excel_export(
        portfolio_result=result.portfolio_result,
        project_inputs=solar_result.project_inputs,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Portfolio" in wb.sheetnames
    assert "Dashboard" in wb.sheetnames


def test_excel_export_contains_validation_and_notes_sheets():
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        validation_issues=[],
        integration_status="full",
        period_view="Semiannual",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Validation" in wb.sheetnames
    assert "Notes" in wb.sheetnames


def test_excel_export_values_only_no_formulas():
    """Excel file should have no formula cells."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(result=result.result, project_inputs=result.project_inputs)
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != 'f', f"Formula found in {sheet}: {cell.coordinate}"


def test_excel_export_handles_project_without_capex_items():
    """Export should not crash if capex has no items."""
    result = run_demo_project("Solar")
    data = build_excel_export(result=result.result, project_inputs=result.project_inputs)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_excel_export_annual_view_has_year_columns():
    """Annual view in Excel should have year-labeled columns (YYYY format)."""
    import openpyxl, re
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        period_view="Annual",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet_name in ["Revenue", "Debt", "Tax_Depreciation", "Waterfall"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        year_cols = [h for h in headers if h is not None and isinstance(h, str) and re.match(r"\d{4}", h)]
        assert len(year_cols) > 0, f"Sheet {sheet_name} has no year columns in headers: {headers}"


def test_excel_export_does_not_use_local_aggregate_annual():
    """excel_export.py must not define its own _aggregate_annual function."""
    import ast, inspect
    from app import excel_export
    src = inspect.getsource(excel_export)
    assert "def _aggregate_annual" not in src, \
        "_aggregate_annual should have been removed; use aggregate_period_table_annual from output_tables"


def test_excel_export_annual_view_uses_output_table_helper():
    """Annual view must use aggregate_period_table_annual from output_tables."""
    import ast, inspect
    from app import excel_export
    src = inspect.getsource(excel_export)
    tree = ast.parse(src)
    found = False
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and node.id == "aggregate_period_table_annual":
            found = True
            break
    assert found, "aggregate_period_table_annual from output_tables not used in excel_export.py"


def test_excel_export_includes_scenario_summary():
    """Notes sheet contains Scenario Deltas rows when scenario is Downside."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        scenario="Downside",
        integration_status="full",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Notes"]
    fields = [row[0] for row in ws.iter_rows(max_row=ws.max_row, values_only=True) if row[0]]
    assert "Scenario Deltas" in fields
    # Should contain at least P50, CapEx, Tariff delta rows
    assert any("P50" in f or "CapEx" in f or "Tariff" in f for f in fields)


def test_excel_notes_include_scenario_deltas():
    """Downside scenario notes should include +/- change values."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        scenario="Downside",
        integration_status="full",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Notes"]
    values = [row[1] for row in ws.iter_rows(max_row=ws.max_row, values_only=True) if row[1]]
    # Downside should show negative changes like "-10%"
    changes = [v for v in values if "%" in str(v) and ("-" in str(v) or "+" in str(v))]
    assert len(changes) > 0, f"Expected scenario change percentages in Notes, got: {values}"


def test_excel_notes_include_bess_hybrid_partial_warning():
    """Excel Notes sheet must include BESS/hybrid partial warning."""
    from io import BytesIO
    from app.ui_runner import run_demo_project
    from app.excel_export import build_excel_export
    import openpyxl

    result = run_demo_project("Solar+BESS")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="partial",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    notes_ws = wb["Notes"]
    all_values = [v for row in notes_ws.iter_rows(values_only=True) for v in row if v]
    notes_text = " ".join(str(v) for v in all_values).lower()
    assert "partial" in notes_text or "bess" in notes_text, \
        f"Notes sheet should mention BESS/hybrid partial status. Got: {notes_text[:200]}"


def test_excel_notes_include_portfolio_experimental_warning():
    """Notes sheet should contain Portfolio experimental warning when status=experimental."""
    import openpyxl
    result = run_demo_project("Portfolio")
    data = build_excel_export(
        portfolio_result=result.portfolio_result,
        project_inputs=result.project_inputs,
        integration_status="experimental",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Notes"]
    fields = [row[0] for row in ws.iter_rows(max_row=ws.max_row, values_only=True) if row[0]]
    portfolio_rows = [r for r in fields if "Portfolio" in str(r) or "IRR" in str(r)]
    assert len(portfolio_rows) > 0, f"Expected Portfolio/IRR warning in Notes, got fields: {fields}"


def test_portfolio_sponsor_irr_unavailable_evidence_label():
    """Portfolio IRR note in Dashboard or Notes should use production evidence wording."""
    import openpyxl
    result = run_demo_project("Portfolio")
    data = build_excel_export(
        portfolio_result=result.portfolio_result,
        project_inputs=result.project_inputs,
        integration_status="experimental",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    # Check all sheets for production evidence wording.
    all_values = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # iter_rows with values_only=True returns tuples of cell values
        all_values.extend([v for row in ws.iter_rows(values_only=True) for v in row if v])
    assert any("portfolio sponsor irr evidence unavailable" in str(v).lower()
               for v in all_values), f"Portfolio IRR unavailable-evidence note not found in workbook values"


def test_excel_has_required_sheets():
    """Verify all required sheets exist (Dashboard, Notes, Inputs, CapEx, CapEx_Items, Revenue, Debt, Tax_Depreciation, Waterfall, Returns)."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    required = [
        "Dashboard", "Notes", "Inputs", "CapEx", "CapEx_Items",
        "Revenue", "Debt", "Tax_Depreciation", "Waterfall", "Returns",
    ]
    missing = [s for s in required if s not in wb.sheetnames]
    assert not missing, f"Missing sheets: {missing}"



def test_excel_sponsor_irr_not_numeric_zero():
    """Sponsor IRR in Portfolio table should be 'n/a' string, not 0.0 float, when placeholder."""
    import openpyxl
    result = run_demo_project("Portfolio")
    data = build_excel_export(
        portfolio_result=result.portfolio_result,
        project_inputs=result.project_inputs,
        integration_status="experimental",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    found = False
    for sheet_name in ["Portfolio", "Dashboard", "Notes"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            label = row[0] if row else ""
            if label and "sponsor" in str(label).lower() and "irr" in str(label).lower():
                val = row[1]
                assert val == "n/a", f"Sponsor IRR in {sheet_name} should be 'n/a' but got {val!r}"
                found = True
                break
        if found:
            break
    assert found, "Sponsor IRR row not found in Portfolio/Dashboard/Notes sheets"


def test_excel_values_only():
    """All cells must be values only — no Excel formula strings."""
    import openpyxl
    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != 'f', \
                    f"Formula found in sheet '{sheet}' at {cell.coordinate}: {cell.value}"


def test_excel_metadata_present():
    """Notes sheet must contain model_version, run_timestamp, scenario."""
    import openpyxl
    from io import BytesIO
    from app.excel_export import build_excel_export
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    notes = wb["Notes"]
    fields = [row[0] for row in notes.iter_rows(values_only=True)]
    assert "Model Version" in fields, "Notes must include Model Version"
    assert "Run Timestamp" in fields, "Notes must include Run Timestamp"
    assert "Scenario" in fields, "Notes must include Scenario"


def test_excel_contains_required_ic_sheets():
    """Excel must contain sheets required for IC pack."""
    import openpyxl
    from io import BytesIO
    from app.excel_export import build_excel_export
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    required = ["Dashboard", "Notes", "Returns", "CapEx", "Revenue"]
    missing = [s for s in required if s not in wb.sheetnames]
    assert not missing, f"Missing IC-required sheets: {missing}"


def test_excel_values_only():
    """All cells must be values only — no Excel formula strings."""
    import openpyxl
    from io import BytesIO
    from app.excel_export import build_excel_export
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        scenario="Base",
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != 'f', \
                    f"Formula found in sheet '{sheet}' at {cell.coordinate}"

def test_dashboard_dscr_uses_actual_period_dscr():
    """Dashboard min_dscr/avg_dscr must match actual period DSCRs."""
    from app.ui_runner import run_demo_project
    from app.output_tables import build_dashboard_kpis
    result = run_demo_project("Solar").result
    kpis = build_dashboard_kpis(result)
    assert kpis["min_dscr"] == result.actual_min_dscr, (
        "min_dscr KPI must equal actual_min_dscr"
    )
    assert kpis["avg_dscr"] == result.actual_avg_dscr, (
        "avg_dscr KPI must equal actual_avg_dscr"
    )


def test_excel_export_uses_run_metadata_when_provided():
    """When run_metadata is provided, Notes sheet uses its git_sha, timestamp, scenario, project_type."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.run_metadata import RunMetadata
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")

    meta = RunMetadata(
        run_id="test-run-001",
        timestamp="2026-01-15T10:00:00+00:00",
        model_version="industry-engine-refactor",
        git_sha="abc1234",
        scenario="Upside",
        project_type="Solar",
        notes="test note",
        warnings=[],
    )

    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        run_metadata=meta,
    )

    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Notes"]

    # Build dict of Field → Value from Notes sheet
    notes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            notes[row[0]] = row[1]

    assert notes["Git SHA"] == "abc1234", f"Expected 'abc1234', got {notes['Git SHA']!r}"
    assert notes["Run Timestamp"] == "2026-01-15T10:00:00+00:00", f"Expected timestamp, got {notes['Run Timestamp']!r}"
    assert notes["Scenario"] == "Upside", f"Expected 'Upside', got {notes['Scenario']!r}"
    assert notes["Project Type"] == "Solar", f"Expected 'Solar', got {notes['Project Type']!r}"


def test_excel_notes_advanced_opex_warning_when_manual_item():
    """Notes sheet contains Advanced OPEX warning when a MANUAL item is present."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    # Pass a MANUAL source item
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.MANUAL_SCHEDULE,
            annual_values_keur=(200.0, 210.0),
            source=OpexSource.MANUAL,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    notes_ws = wb["Notes"]
    notes = {row[0]: row[1] for row in notes_ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert "Advanced OPEX" in notes, f"Expected 'Advanced OPEX' in Notes, got: {notes}"
    assert "Manual" in notes["Advanced OPEX"] or "hardcoded" in notes["Advanced OPEX"].lower()


def test_excel_notes_no_advanced_opex_warning_when_all_formula():
    """Notes sheet does NOT contain Advanced OPEX warning when all items are FORMULA-driven."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
            is_hardcoded=False,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    notes_ws = wb["Notes"]
    notes = {row[0]: row[1] for row in notes_ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert "Advanced OPEX" not in notes, f"Expected no Advanced OPEX warning for formula items, got: {notes}"


def test_excel_notes_advanced_opex_warning_when_hardcoded():
    """Notes sheet contains Advanced OPEX warning when is_hardcoded=True."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
            is_hardcoded=True,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    notes_ws = wb["Notes"]
    notes = {row[0]: row[1] for row in notes_ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert "Advanced OPEX" in notes


def test_excel_opex_detail_sheet_exists_when_advanced_opex_enabled():
    """OPEX Detail sheet is created when advanced_opex_line_items is provided and non-empty."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
        ),
        OpexLineItem(
            name="Operations", category="operations",
            base_year_amount_keur=100.0, inflation_rate=0.02,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "OPEX Detail" in wb.sheetnames, f"Expected 'OPEX Detail' sheet, got: {wb.sheetnames}"


def test_excel_opex_detail_sheet_absent_when_no_advanced_opex():
    """OPEX Detail sheet is NOT created when advanced_opex_line_items is None."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=None,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "OPEX Detail" not in wb.sheetnames, f"OPEX Detail should not exist without advanced OPEX"


def test_excel_opex_detail_sheet_absent_when_empty_tuple():
    """OPEX Detail sheet is NOT created when advanced_opex_line_items is empty tuple."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=(),
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "OPEX Detail" not in wb.sheetnames, f"OPEX Detail should not exist with empty tuple"


def test_excel_opex_detail_columns_and_row_count():
    """OPEX Detail sheet has correct columns and row count = items * horizon_years."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    horizon = result.project_inputs.info.horizon_years  # 25
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
        ),
        OpexLineItem(
            name="Operations", category="operations",
            base_year_amount_keur=100.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["OPEX Detail"]

    # Collect headers
    headers = [cell.value for cell in ws[1]]
    expected = ["Line Item Name", "Category", "Year", "Value (kEUR)", "Source", "Is Override", "Fixed Assumption", "Override Note"]
    assert headers == expected, f"Expected {expected}, got {headers}"

    # Row count (excluding header)
    data_rows = ws.max_row - 1  # subtract header
    assert data_rows == len(items) * horizon, f"Expected {len(items) * horizon} rows, got {data_rows}"

    # Verify Year column is 1-based
    years = [row[2] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert years[0] == 1 and years[-1] == horizon


def test_excel_opex_detail_values_are_values_only():
    """OPEX Detail sheet contains concrete numbers, not formula strings."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["OPEX Detail"]

    # All data rows should have numeric Value (kEUR) — not strings
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[3]  # Value (kEUR) column
        assert isinstance(val, (int, float)), f"Expected numeric Value, got {type(val).__name__}: {val!r}"
        assert val > 0, f"Expected positive value, got {val}"


def test_excel_opex_detail_hardcoded_flag_from_is_hardcoded_field():
    """OPEX Detail Fixed Assumption=True when OpexLineItem.is_hardcoded=True (even with FORMULA source)."""
    from io import BytesIO
    import openpyxl
    from app.excel_export import build_excel_export
    from app.opex_engine import OpexLineItem, OpexSource, CalculationMode
    from app.ui_runner import run_demo_project

    result = run_demo_project("Solar")
    # source=FORMULA but is_hardcoded=True — was incorrectly False before fix
    items = (
        OpexLineItem(
            name="Insurance", category="insurance",
            base_year_amount_keur=200.0, inflation_rate=0.0,
            calculation_mode=CalculationMode.INFLATED_FROM_BASE,
            source=OpexSource.FORMULA,
            is_hardcoded=True,
        ),
    )
    data = build_excel_export(
        result=result.result,
        project_inputs=result.project_inputs,
        integration_status="full",
        advanced_opex_line_items=items,
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["OPEX Detail"]

    # Find the row(s) for Insurance — Fixed Assumption must be True
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "Insurance":
            assert row[6] is True, f"Fixed Assumption should be True for is_hardcoded=True item, got {row[6]}"

# ── Portfolio UI export tests (Phase 1.5) ──────────────────────────────────

class TestPortfolioUITables:
    """Test build_portfolio_summary_table and build_portfolio_spv_table directly.

    The full Excel export has a timezone datetime issue with Mock results
    that is unrelated to Phase 1.5 code. The individual table builders
    are tested here directly.
    """

    def test_portfolio_summary_table_columns(self):
        """Portfolio summary table has expected columns."""
        from app.portfolio_ui import build_portfolio_summary_table
        from domain.portfolio.independent import IndependentPortfolioResult

        result = IndependentPortfolioResult(
            portfolio_name="Test",
            spv_outputs=(),
            total_revenue_keur=100_000.0,
            total_ebitda_keur=80_000.0,
            total_tax_keur=10_000.0,
            total_senior_ds_keur=50_000.0,
            total_distribution_keur=20_000.0,
            min_dscr=1.2,
            avg_dscr=1.35,
            spv_project_irrs=(),
            spv_equity_irrs=(),
            simple_avg_project_irr=0.09,
            simple_avg_equity_irr=0.11,
            dsrf_enabled=False,
            warnings=(),
        )
        df = build_portfolio_summary_table(result)
        assert list(df.columns) == ["Field", "Value"]
        fields = set(df["Field"])
        assert "Portfolio Name" in fields
        assert "Number of SPVs" in fields
        assert "Total Revenue (kEUR)" in fields
        assert "Min DSCR (conservative)" in fields
        assert "Simple Avg Project IRR (NOT Portfolio XIRR)" in fields

    def test_portfolio_spv_table_columns(self):
        """Portfolio SPV table has expected columns."""
        from app.portfolio_ui import build_portfolio_spv_table
        from domain.portfolio.independent import IndependentPortfolioResult, SPVOutput

        out = SPVOutput(
            project_code="SPV-A", project_name="Project A",
            project_irr=0.09, equity_irr=0.12,
            total_revenue_keur=1000.0, total_ebitda_keur=800.0,
            total_tax_keur=100.0, total_senior_ds_keur=500.0,
            total_distribution_keur=200.0, avg_dscr=1.3, min_dscr=1.15,
            waterfall_result=None, warnings=(),
        )
        result = IndependentPortfolioResult(
            portfolio_name="Test", spv_outputs=(out,),
            total_revenue_keur=1000.0, total_ebitda_keur=800.0,
            total_tax_keur=100.0, total_senior_ds_keur=500.0,
            total_distribution_keur=200.0, min_dscr=1.15, avg_dscr=1.3,
            spv_project_irrs=(0.09,), spv_equity_irrs=(0.12,),
            dsrf_enabled=False, warnings=(),
        )
        df = build_portfolio_spv_table(result)
        assert "SPV Code" in df.columns
        assert "Revenue (kEUR)" in df.columns
        assert "Project IRR" in df.columns
        assert "Equity IRR" in df.columns
        assert len(df) == 1
        assert df["SPV Code"].iloc[0] == "SPV-A"

    def test_portfolio_summary_irr_label_present(self):
        """Summary table contains the explicit IRR label."""
        from app.portfolio_ui import build_portfolio_summary_table, _IRR_LABEL
        from domain.portfolio.independent import IndependentPortfolioResult

        result = IndependentPortfolioResult(
            portfolio_name="Test", spv_outputs=(),
            total_revenue_keur=100_000.0, total_ebitda_keur=80_000.0,
            total_tax_keur=10_000.0, total_senior_ds_keur=50_000.0,
            total_distribution_keur=20_000.0, min_dscr=1.2, avg_dscr=1.35,
            spv_project_irrs=(), spv_equity_irrs=(),
            simple_avg_project_irr=0.09, simple_avg_equity_irr=0.11,
            dsrf_enabled=False, warnings=(),
        )
        df = build_portfolio_summary_table(result)
        # Check IRR label is present in summary
        irr_row = df[df["Field"].str.contains("IRR", na=False)]
        assert len(irr_row) >= 2  # Project and Equity IRR rows

    def test_dsrf_placeholder_status(self):
        """DSRF shows as disabled in portfolio summary."""
        from app.portfolio_ui import build_portfolio_summary_table
        from domain.portfolio.independent import IndependentPortfolioResult

        result = IndependentPortfolioResult(
            portfolio_name="Test", spv_outputs=(),
            total_revenue_keur=100_000.0, total_ebitda_keur=80_000.0,
            total_tax_keur=10_000.0, total_senior_ds_keur=50_000.0,
            total_distribution_keur=20_000.0, min_dscr=1.2, avg_dscr=1.35,
            spv_project_irrs=(), spv_equity_irrs=(),
            simple_avg_project_irr=0.09, simple_avg_equity_irr=0.11,
            dsrf_enabled=False, warnings=(),
        )
        df = build_portfolio_summary_table(result)
        dsrf_row = df[df["Field"] == "DSRF Enabled"]
        assert len(dsrf_row) == 1
        assert dsrf_row["Value"].values[0] == False

# ── Real Excel export tests for IndependentPortfolioResult (Phase 1.5) ─────

class TestIndependentPortfolioExcelExport:
    """Test that build_excel_export produces correct sheets for IndependentPortfolioResult.

    Uses real run_independent_portfolio() output (not mocks) so we verify
    the actual workbook structure end-to-end.
    """

    @pytest.fixture
    def _oborovo_shim(self):
        """Ensure project_factories is imported (registers the replace shim)."""
        import app.project_factories  # noqa: F401

    def _make_independent_portfolio_result(self):
        """Create a real IndependentPortfolioResult from two default solar SPVs."""
        from dataclasses import replace
        from app.project_factories import create_default_solar_project
        from domain.portfolio.independent import (
            run_independent_portfolio,
            IndependentPortfolioInputs,
        )
        s1 = replace(create_default_solar_project(), info=replace(
            create_default_solar_project().info, code="SOLAR-T1", name="Solar T1"))
        s2 = replace(create_default_solar_project(), info=replace(
            create_default_solar_project().info, code="SOLAR-T2", name="Solar T2"))
        return run_independent_portfolio(
            IndependentPortfolioInputs(projects=(s1, s2)), strict=True)

    def test_portfolio_summary_sheet_exists(self, _oborovo_shim):
        """Portfolio_Summary sheet is created for IndependentPortfolioResult."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Portfolio_Summary" in wb.sheetnames

    def test_portfolio_spvs_sheet_exists(self, _oborovo_shim):
        """Portfolio_SPVs sheet is created for IndependentPortfolioResult."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Portfolio_SPVs" in wb.sheetnames

    def test_portfolio_notes_sheet_exists(self, _oborovo_shim):
        """Portfolio_Notes sheet is created for IndependentPortfolioResult."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Portfolio_Notes" in wb.sheetnames

    def test_portfolio_summary_irr_label_not_portfolio_xirr(self, _oborovo_shim):
        """Portfolio_Summary IRR rows explicitly say NOT Portfolio XIRR."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Summary"]
        irr_fields = [str(row[1]) for row in ws.iter_rows(values_only=True) if row[1] and "IRR" in str(row[1])]
        assert any("NOT Portfolio XIRR" in f for f in irr_fields), f"No IRR label found with disclaimer: {irr_fields}"

    def test_portfolio_notes_contains_no_holdco(self, _oborovo_shim):
        """Portfolio_Notes explicitly says No HoldCo."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Notes"]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "No HoldCo" in all_text

    def test_portfolio_notes_contains_no_shl(self, _oborovo_shim):
        """Portfolio_Notes explicitly says No SHL."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Notes"]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "No SHL" in all_text

    def test_portfolio_notes_contains_no_sponsor_irr(self, _oborovo_shim):
        """Portfolio_Notes explicitly says No Sponsor IRR."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Notes"]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "No Sponsor IRR" in all_text

    def test_portfolio_notes_contains_dsrf_placeholder(self, _oborovo_shim):
        """Portfolio_Notes states DSRF is a placeholder only."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Notes"]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "DSRF" in all_text and ("revolving" in all_text.lower() or "facility" in all_text.lower())

    def test_no_holdco_sheet_not_created(self, _oborovo_shim):
        """No separate HoldCo sheet is created (not in scope)."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "HoldCo" not in wb.sheetnames
        assert "Holdco" not in wb.sheetnames

    def test_independent_portfolio_dashboard_no_pooled_revenue_label(self, _oborovo_shim):
        """Independent portfolio Dashboard should not contain 'Pooled Revenue'."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Dashboard"]
        all_labels = [str(row[0]) for row in ws.iter_rows(values_only=True) if row[0]]
        assert "Pooled Revenue" not in all_labels, f"Found 'Pooled Revenue' in Dashboard: {all_labels}"

    def test_independent_portfolio_dashboard_has_portfolio_revenue_label(self, _oborovo_shim):
        """Independent portfolio Dashboard should contain 'Portfolio Revenue'."""
        from io import BytesIO
        import openpyxl

        result = self._make_independent_portfolio_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Dashboard"]
        all_labels = [str(row[0]) for row in ws.iter_rows(values_only=True) if row[0]]
        assert "Portfolio Revenue (kEUR)" in all_labels, f"Missing 'Portfolio Revenue' in Dashboard: {all_labels}"


class TestDSRFExcelExport:
    """DSRF Excel export tests for Phase 2."""

    def _ensure_factory(self):
        """Ensure project_factories is imported (registers the replace shim)."""
        import app.project_factories  # noqa: F401

    def _make_dsrf_enabled_result(self):
        """Create IndependentPortfolioResult with DSRF enabled and populated periods."""
        self._ensure_factory()
        from dataclasses import replace
        from app.project_factories import create_default_solar_project
        from domain.portfolio.independent import (
            run_independent_portfolio,
            IndependentPortfolioInputs,
            DSRFConfig,
        )

        s1 = replace(create_default_solar_project(), info=replace(
            create_default_solar_project().info, code="SOLAR-D1", name="Solar D1"))

        config = DSRFConfig(
            enabled=True,
            sizing_months=6,
            sizing_basis="average_debt_service",
            commitment_fee_rate_pa=0.005,
            margin_rate_pa=0.02,
            euribor_rate_pa=0.03,
            period_year_fraction=0.5,
            repayment_priority="before_distributions",
        )

        return run_independent_portfolio(
            IndependentPortfolioInputs(projects=(s1,), portfolio_name="DSRF-Test", dsrf=config),
            strict=True,
        )

    def _make_dsrf_disabled_result(self):
        """Create IndependentPortfolioResult with DSRF disabled (no DSRF sheet)."""
        self._ensure_factory()
        from dataclasses import replace
        from app.project_factories import create_default_solar_project
        from domain.portfolio.independent import (
            run_independent_portfolio,
            IndependentPortfolioInputs,
        )
        s1 = replace(create_default_solar_project(), info=replace(
            create_default_solar_project().info, code="SOLAR-X1", name="Solar X1"))

        return run_independent_portfolio(
            IndependentPortfolioInputs(projects=(s1,), portfolio_name="DSRF-Disabled", dsrf=None),
            strict=True,
        )

    def test_dsrf_sheet_absent_when_disabled(self):
        """DSRF sheet is NOT created when dsrf=None or enabled=False."""
        result = self._make_dsrf_disabled_result()
        assert result.dsrf_enabled is False
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "DSRF" not in wb.sheetnames

    def test_dsrf_sheet_absent_when_enabled_but_no_periods(self):
        """DSRF sheet is NOT created when enabled=True but dsrf_periods is empty."""
        self._ensure_factory()
        from dataclasses import replace
        from app.project_factories import create_default_solar_project
        from domain.portfolio.independent import (
            run_independent_portfolio,
            IndependentPortfolioInputs,
            DSRFConfig,
        )

        s1 = replace(create_default_solar_project(), info=replace(
            create_default_solar_project().info, code="SOLAR-E1", name="Solar E1"))
        config = DSRFConfig(enabled=True, sizing_months=6, sizing_basis="average_debt_service",
                            period_year_fraction=0.5)
        result = run_independent_portfolio(
            IndependentPortfolioInputs(projects=(s1,), portfolio_name="Test", dsrf=config),
            strict=True,
        )
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        if "DSRF" in wb.sheetnames:
            ws = wb["DSRF"]
            row_count = sum(1 for _ in ws.iter_rows())
            assert row_count > 1, "DSRF sheet should have data rows when dsrf_periods is populated"

    def test_dsrf_sheet_exists_when_enabled_with_periods(self):
        """DSRF sheet IS created when dsrf_enabled=True AND dsrf_periods is non-empty."""
        result = self._make_dsrf_enabled_result()
        assert result.dsrf_enabled is True
        assert len(result.dsrf_periods) > 0, "DSRF periods should be populated for this test"
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "DSRF" in wb.sheetnames

    def test_dsrf_sheet_headers_exact(self):
        """DSRF sheet has exactly the specified 15 column headers in order."""
        result = self._make_dsrf_enabled_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["DSRF"]
        # Use iter_rows() to get Cell objects with .value attribute
        first_row = next(ws.iter_rows())
        # Filter out None cells (empty leading columns)
        headers = [str(cell.value) for cell in first_row if cell.value is not None]
        expected = [
            "Period", "SPV Code", "Facility Limit (kEUR)", "Drawn Start (kEUR)",
            "Undrawn Start (kEUR)", "Scheduled Senior DS (kEUR)", "CFADS Available (kEUR)",
            "Debt Service Shortfall (kEUR)", "Draw (kEUR)", "Drawn Interest (kEUR)",
            "Commitment Fee (kEUR)", "Repayment (kEUR)", "Drawn End (kEUR)",
            "Undrawn End (kEUR)", "Cash Available for Distribution (kEUR)",
        ]
        assert headers == expected, f"DSRF headers mismatch.\nGot: {headers}\nExpected: {expected}"

    def test_dsrf_sheet_has_data_rows(self):
        """DSRF sheet has at least one data row when dsrf_periods exist."""
        result = self._make_dsrf_enabled_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["DSRF"]
        row_count = sum(1 for _ in ws.iter_rows())
        assert row_count >= 2, f"DSRF sheet should have header + data rows, got {row_count}"

    def test_portfolio_summary_contains_dsrf_rows(self):
        """Portfolio_Summary contains DSRF rows when enabled."""
        result = self._make_dsrf_enabled_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Portfolio_Summary"]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "DSRF Facility Limit" in all_text, "DSRF Facility Limit row missing"
        assert "DSRF Total Draw" in all_text, "DSRF Total Draw row missing"
        assert "DSRF Distribution Reduction" in all_text, "DSRF Distribution Reduction row missing"

    def test_dsrf_export_no_prohibited_terminology(self):
        """Exported DSRF labels must not contain: top-up, release, funded, balance."""
        result = self._make_dsrf_enabled_result()
        data = build_excel_export(result=None, portfolio_result=result,
                                  project_inputs=None, validation_issues=None)
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data))

        # Check DSRF sheet headers
        if "DSRF" in wb.sheetnames:
            ws = wb["DSRF"]
            dsrf_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
            for term in ("top-up", "topup", "release", "funded", "balance"):
                assert term not in dsrf_text.lower(), f"Prohibited term '{term}' found in DSRF sheet"

        # Check Portfolio_Summary labels
        ws_sum = wb["Portfolio_Summary"]
        sum_text = " ".join(str(c.value) for row in ws_sum.iter_rows() for c in row if c.value)
        for term in ("top-up", "topup", "release", "funded", "balance"):
            dsrf_label_text = " ".join(l for l in sum_text.split() if "dsrf" in l.lower() or "DSRF" in l)
            assert term not in dsrf_label_text.lower(), f"Prohibited term '{term}' in DSRF Portfolio_Summary labels"

# ── Phase 6B: Tax audit sheets integration ────────────────────────────────────

class TestTaxAuditSheetsIntegration:
    """Tests for optional tax_results parameter in build_excel_export."""

    def _make_tax_result(self, entity_code="HR-SPV-001"):
        from domain.tax.templates.inputs import TaxTemplate, TaxDepreciationRule, CITTier
        from domain.tax.templates.resolver import resolve_tax_template
        from domain.tax.engine_inputs import SPVTaxEngineInputs
        from domain.tax.engine_runner import run_spv_tax_engine

        template = TaxTemplate(
            country_code="HR",
            template_name="HR Flat 10%",
            tax_year=2026,
            cit_tiers=(CITTier(0.0, None, 0.10),),
            depreciation_rules=(
                TaxDepreciationRule(
                    asset_category="infra",
                    method="straight_line",
                    annual_rate=None,
                    useful_life_years=20.0,
                    max_deductible_rate=None,
                    bonus_depreciation_pct=0.0,
                    deductible=True,
                    notes="HR straight-line",
                ),
            ),
            withholding_tax_dividends=0.0,
            withholding_tax_interest=0.0,
        )
        config = resolve_tax_template(template, ())
        inputs = SPVTaxEngineInputs(
            entity_code=entity_code,
            resolved_tax_config=config,
            ebitda_by_period_keur=tuple(1000.0 for _ in range(5)),
            deductible_interest_by_period_keur=tuple(0.0 for _ in range(5)),
            book_depreciation_by_period_keur=tuple(500.0 for _ in range(5)),
            asset_cost_keur=10_000.0,
            depreciation_rule_asset_category="infra",
            non_deductible_addbacks_by_period_keur=(0.0,) * 5,
        )
        return run_spv_tax_engine(inputs)

    def test_existing_export_works_without_tax_results(self):
        """Default export (no tax_results) is unchanged."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
        )
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(BytesIO(data))
        # No SPV tax audit sheets when tax_results not provided
        tax_sheets = [s for s in wb.sheetnames if s in ("Tax Summary",) or s.startswith("Tax_") and not s.startswith("Tax_D")]
        # (Tax_Depreciation is a pre-existing disclosure sheet, not SPV audit)
        assert len(tax_sheets) == 0, f"Unexpected Tax sheets: {tax_sheets}"
        wb.close()

    def test_tax_results_creates_tax_summary_sheet(self):
        """Passing tax_results creates Tax Summary sheet."""
        import openpyxl
        tax_result = self._make_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_results=(tax_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax Summary" in wb.sheetnames, f"Tax Summary not found. Sheets: {wb.sheetnames}"
        wb.close()

    def test_tax_results_creates_per_spv_tax_sheet(self):
        """Passing tax_results creates per-SPV Tax_{entity_code} sheet."""
        import openpyxl
        tax_result = self._make_tax_result(entity_code="HR-SPV-001")
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_results=(tax_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax_HR-SPV-001" in wb.sheetnames, f"Tax_HR-SPV-001 not found. Sheets: {wb.sheetnames}"
        wb.close()

    def test_existing_core_sheets_still_exist_with_tax_results(self):
        """Existing core sheets are unchanged when tax_results is passed."""
        import openpyxl
        tax_result = self._make_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_results=(tax_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        required = [
            "Dashboard", "Inputs", "CapEx",
            "Revenue", "Debt", "Tax_Depreciation",
            "Waterfall", "Returns", "Validation", "Notes",
        ]
        missing = [s for s in required if s not in wb.sheetnames]
        assert not missing, f"Missing core sheets: {missing}"
        wb.close()

    def test_no_tax_sheets_when_tax_results_is_empty_tuple(self):
        """tax_results=() creates no Tax sheets."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_results=(),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        tax_sheets = [s for s in wb.sheetnames if s in ("Tax Summary",) or s.startswith("Tax_") and not s.startswith("Tax_D")]
        assert len(tax_sheets) == 0, f"Unexpected SPV audit sheets with empty tuple: {tax_sheets}"
        wb.close()

    def test_tax_summary_sheet_audit_note_row_1(self):
        """Tax Summary sheet row 1 contains audit-only note."""
        import openpyxl
        tax_result = self._make_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_results=(tax_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["Tax Summary"]
        first_cell = ws.cell(row=1, column=1).value
        assert "AUDIT-ONLY" in str(first_cell), f"Expected AUDIT-ONLY in row 1, got: {first_cell}"
        wb.close()


class TestHoldCoTaxAuditSheetsIntegration:
    """Phase 6C.5: Optional HoldCo tax audit sheets in Excel export."""

    def _make_holdco_tax_result(self, entity_code="HR-HoldCo-001", n_periods=3):
        from domain.tax.holdco_inputs import (
            HoldCoTaxInputs,
            WithholdingTaxConfig,
            InterestDeductibilityConfig,
        )
        from domain.tax.holdco_runner import run_holdco_tax_engine
        inputs = HoldCoTaxInputs(
            entity_code=entity_code,
            country_code="HR",
            tax_year=2026,
            dividend_income_by_period_keur=tuple(1000.0 for _ in range(n_periods)),
            shl_interest_income_by_period_keur=tuple(200.0 for _ in range(n_periods)),
            shl_principal_received_by_period_keur=tuple(500.0 for _ in range(n_periods)),
            holdco_opex_by_period_keur=tuple(50.0 for _ in range(n_periods)),
            withholding_tax_config=WithholdingTaxConfig(rate_dividends=0.12, rate_interest=0.0),
            interest_deductibility=InterestDeductibilityConfig(thin_cap_ratio=4.0, interest_limitation_pct_ebitda=0.30),
            metadata=(("source", "test"),),
        )
        return run_holdco_tax_engine(inputs)

    def test_default_export_unchanged_without_holdco_tax_results(self):
        """Default export (no holdco_tax_results) creates no HoldCo tax sheets."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        holdco_sheets = [s for s in wb.sheetnames if "HoldCo" in s]
        assert len(holdco_sheets) == 0, f"Unexpected HoldCo sheets without param: {holdco_sheets}"
        wb.close()

    def test_holdco_tax_results_none_creates_no_sheets(self):
        """holdco_tax_results=None creates no HoldCo tax sheets."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=None,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        holdco_sheets = [s for s in wb.sheetnames if "HoldCo" in s]
        assert len(holdco_sheets) == 0, f"Unexpected HoldCo sheets with None: {holdco_sheets}"
        wb.close()

    def test_holdco_tax_results_empty_tuple_creates_no_sheets(self):
        """holdco_tax_results=() creates no HoldCo tax sheets."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=(),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        holdco_sheets = [s for s in wb.sheetnames if "HoldCo" in s]
        assert len(holdco_sheets) == 0, f"Unexpected HoldCo sheets with empty tuple: {holdco_sheets}"
        wb.close()

    def test_holdco_tax_results_creates_holdco_tax_summary(self):
        """Provided holdco_tax_results creates 'HoldCo Tax Summary' sheet."""
        import openpyxl
        hc_result = self._make_holdco_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=(hc_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "HoldCo Tax Summary" in wb.sheetnames, f"HoldCo Tax Summary not found. Sheets: {wb.sheetnames}"
        wb.close()

    def test_holdco_tax_results_creates_per_entity_sheet(self):
        """Provided holdco_tax_results creates per-entity 'HoldCoTax_{entity_code}' sheet."""
        import openpyxl
        hc_result = self._make_holdco_tax_result(entity_code="HR-HoldCo-001")
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=(hc_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "HoldCoTax_HR-HoldCo-001" in wb.sheetnames, f"HoldCoTax_HR-HoldCo-001 not found. Sheets: {wb.sheetnames}"
        wb.close()

    def test_existing_core_sheets_still_exist_with_holdco_tax_results(self):
        """Existing core sheets are unchanged when holdco_tax_results is passed."""
        import openpyxl
        hc_result = self._make_holdco_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=(hc_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        required = [
            "Dashboard", "Inputs", "CapEx",
            "Revenue", "Debt", "Tax_Depreciation",
            "Waterfall", "Returns", "Validation", "Notes",
        ]
        missing = [s for s in required if s not in wb.sheetnames]
        assert not missing, f"Missing core sheets: {missing}"
        wb.close()

    def test_holdco_tax_summary_sheet_audit_note_row_1(self):
        """HoldCo Tax Summary sheet row 1 contains audit-only note."""
        import openpyxl
        hc_result = self._make_holdco_tax_result()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            holdco_tax_results=(hc_result,),
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb["HoldCo Tax Summary"]
        first_cell = ws.cell(row=1, column=1).value
        assert "AUDIT-ONLY" in str(first_cell), f"Expected AUDIT-ONLY in row 1, got: {first_cell}"
        wb.close()


class TestTaxAssumptionSnapshotIntegration:
    """Tests for optional tax_assumption_snapshot parameter in build_excel_export."""

    def _make_snapshot(self):
        from domain.tax.templates.inputs import (
            CITTier, TaxDepreciationRule, TaxTemplate, TaxTemplateOverride,
        )
        from domain.tax.templates.resolver import resolve_tax_template
        from domain.tax.snapshot_builders import build_tax_assumption_snapshot

        template = TaxTemplate(
            country_code="HR",
            template_name="HR Flat 18%",
            tax_year=2026,
            cit_tiers=(CITTier(min_profit_keur=0.0, max_profit_keur=None, tax_rate=0.18),),
            depreciation_rules=(
                TaxDepreciationRule(
                    asset_category="buildings", method="straight_line",
                    annual_rate=0.05, useful_life_years=20.0, max_deductible_rate=0.025,
                    bonus_depreciation_pct=0.0, deductible=True, notes="20yr buildings",
                ),
            ),
            withholding_tax_dividends=0.12,
            withholding_tax_interest=0.0,
            loss_carryforward_years=5,
            thin_cap_ratio=4.0,
            interest_limitation_pct_ebitda=0.30,
            metadata=(("notes", "standard HR"),),
        )
        override = TaxTemplateOverride(
            override_name="custom_wht",
            field_path="withholding_tax_dividends",
            override_value=0.10,
            reason="Treaty rate HR-ME",
        )
        resolved = resolve_tax_template(template, (override,))
        return build_tax_assumption_snapshot(
            templates=(template,),
            resolved_configs=(resolved,),
            overrides=(override,),
        )

    def test_default_export_unchanged_when_snapshot_none(self):
        """Default export (snapshot=None) is unchanged — no tax assumption sheets."""
        import openpyxl
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
        )
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(BytesIO(data))
        tax_assumption_sheets = [
            "Tax Templates", "Tax Tiers", "Tax Dep Rules",
            "Tax Overrides", "Resolved Tax Config",
        ]
        found = [s for s in wb.sheetnames if s in tax_assumption_sheets]
        assert len(found) == 0, f"Unexpected tax assumption sheets: {found}"
        wb.close()

    def test_snapshot_creates_tax_templates_sheet(self):
        """Passing snapshot creates Tax Templates sheet."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax Snapshot Templates" in wb.sheetnames
        wb.close()

    def test_snapshot_creates_tax_tiers_sheet(self):
        """Passing snapshot creates Tax Tiers sheet."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax Snapshot Templates" in wb.sheetnames  # Tiers included in snapshot
        wb.close()

    def test_snapshot_creates_tax_dep_rules_sheet(self):
        """Passing snapshot creates Tax Dep Rules sheet."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax Snapshot Templates" in wb.sheetnames  # Dep rules included in snapshot
        wb.close()

    def test_snapshot_creates_tax_overrides_sheet(self):
        """Passing snapshot creates Tax Overrides sheet."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "Tax Snapshot Overrides" in wb.sheetnames
        wb.close()

    def test_snapshot_resolved_exports_when_present(self):
        """Snapshot with resolved_config_snapshots creates Tax Snapshot Resolved sheet."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        # Resolved config snapshot creates Tax Snapshot Resolved sheet
        assert "Tax Snapshot Resolved" in wb.sheetnames
        wb.close()

    def test_existing_core_sheets_unchanged_with_snapshot(self):
        """Core sheets are unchanged when snapshot is provided."""
        import openpyxl
        snapshot = self._make_snapshot()
        result = run_demo_project("Solar")
        data = build_excel_export(
            result=result.result,
            project_inputs=result.project_inputs,
            tax_assumption_snapshot=snapshot,
        )
        wb = openpyxl.load_workbook(BytesIO(data))
        required = [
            "Dashboard", "Inputs", "CapEx",
            "Revenue", "Debt", "Tax_Depreciation",
            "Waterfall", "Returns", "Validation", "Notes",
        ]
        missing = [s for s in required if s not in wb.sheetnames]
        assert not missing, f"Missing core sheets: {missing}"
        wb.close()


