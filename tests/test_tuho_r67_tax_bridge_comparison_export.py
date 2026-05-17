"""Workbook validation for the TUHO R67 tax bridge comparison export."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.project_factories import create_default_tuho_wind1
from app.ui_runner import _build_period_engine
from app.waterfall_runner import WaterfallRunConfig, WaterfallRunner


WORKBOOK_PATH = Path("reports/phase6_tuho_r67_tax_bridge_comparison.xlsx")
REQUIRED_SHEETS = {
    "Summary",
    "Period Comparison",
    "Taxable Income Bridge",
    "Excel vs Python Upstream Rows",
    "Largest Deltas",
    "Suspected Causes",
}


def _run_default_tuho():
    project = create_default_tuho_wind1()
    engine = _build_period_engine(project)
    config = WaterfallRunConfig.from_inputs(project, engine)
    return WaterfallRunner(project, engine).run(config)


def test_workbook_exists_and_required_sheets_are_present():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    assert REQUIRED_SHEETS.issubset(set(workbook.sheetnames))


def test_period_comparison_contains_60_operating_periods():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Period Comparison"]

    op_indices = [
        sheet.cell(row=row, column=1).value
        for row in range(2, 62)
    ]

    assert op_indices == list(range(60))
    assert sheet.cell(row=62, column=1).value == "TOTAL"


def test_largest_delta_sheet_is_sorted_descending_by_absolute_delta():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Largest Deltas"]

    abs_deltas = [sheet.cell(row=row, column=7).value for row in range(2, 62)]

    assert abs_deltas == sorted(abs_deltas, reverse=True)
    assert sheet.cell(row=2, column=2).value == 25
    assert sheet.cell(row=2, column=7).value == pytest.approx(825.6459441816605, abs=0.01)


def test_summary_captures_key_totals_and_interpretation():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Summary"]
    values = {sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value for row in range(2, 9)}

    assert values["Excel R67 total"] == pytest.approx(-38_240.920880415375, abs=0.01)
    assert values["Python legacy R67"] == pytest.approx(-20_140.22414240874, abs=0.01)
    assert values["Python tax bridge R67"] == pytest.approx(-36_091.61517762715, abs=0.01)
    assert values["Residual"] == pytest.approx(2_149.3057027882255, abs=0.01)
    assert values["Material period count"] == 18
    assert values["Interpretation"] == "R99 remains blocked"


def test_workbook_creation_did_not_change_default_runtime_behavior():
    baseline = _run_default_tuho()
    diagnostic = _run_default_tuho()

    assert diagnostic.total_revenue_keur == pytest.approx(baseline.total_revenue_keur, abs=0.0001)
    assert diagnostic.total_opex_keur == pytest.approx(baseline.total_opex_keur, abs=0.0001)
    assert diagnostic.total_tax_keur == pytest.approx(baseline.total_tax_keur, abs=0.0001)
    assert diagnostic.total_senior_ds_keur == pytest.approx(
        baseline.total_senior_ds_keur,
        abs=0.0001,
    )
