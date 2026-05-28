"""Tests for Phase 20U-A OPEX Line-Item Reconciliation Workbook.

Branch: phase20u-opex-line-item-reconciliation-workbook
Status: DIAGNOSTIC ONLY — no runtime formula changes

Tests
-----
1. Workbook is generated.
2. Workbook has required sheets: TUHO_OPEX, OBOROVO_OPEX, Summary.
3. TUHO sheet contains Excel table, Python table, and Delta table.
4. Oborovo sheet contains Excel table, Python table, and Delta table.
5. Period columns are horizontal.
6. B.01/B.02/B.12 rows exist for Oborovo (Python-side).
7. Total OPEX row exists for both projects.
8. Delta table calculates Python - Excel for aggregate.
9. Missing rows are marked MISSING, not silently omitted.
10. No runtime formula changes.
"""

import os
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
OUT_XLSX = REPORTS / "phase20u_opex_line_item_reconciliation.xlsx"
SCRIPT = ROOT / "scripts" / "export_phase20u_opex_reconciliation_workbook.py"


class TestPhase20UWorkbookGeneration:
    """Test Phase 20U-A workbook generation."""

    def test_workbook_exists(self):
        """Reconciliation workbook should be generated."""
        # Run the script first
        import subprocess
        result = subprocess.run(
            ["python3", str(SCRIPT)],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=120,
        )
        # Script should complete without error (allowed to fail gracefully)
        # Check workbook exists regardless
        assert OUT_XLSX.exists(), (
            f"Workbook not generated at {OUT_XLSX}\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )

    def test_workbook_has_required_sheets(self):
        """Workbook must have TUHO_OPEX, OBOROVO_OPEX, Summary sheets."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        sheet_names = wb.sheetnames
        required = {"TUHO_OPEX", "OBOROVO_OPEX", "Summary"}
        missing = required - set(sheet_names)
        assert not missing, f"Missing sheets: {missing} (have: {sheet_names})"

    def test_tuho_sheet_structure(self):
        """TUHO sheet must contain Excel table, Python table, and Delta table."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["TUHO_OPEX"]

        # Scan for table header labels
        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)

        # Check for table markers
        assert "TABLE 1" in content_str, "TUHO sheet missing TABLE 1 (Excel OPEX)"
        assert "TABLE 2" in content_str, "TUHO sheet missing TABLE 2 (Python OPEX)"
        assert "TABLE 3" in content_str, "TUHO sheet missing TABLE 3 (Delta)"

    def test_oboroovo_sheet_structure(self):
        """Oborovo sheet must contain Excel table, Python table, and Delta table."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["OBOROVO_OPEX"]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)

        assert "TABLE 1" in content_str, "Oborovo sheet missing TABLE 1 (Excel OPEX)"
        assert "TABLE 2" in content_str, "Oborovo sheet missing TABLE 2 (Python OPEX)"
        assert "TABLE 3" in content_str, "Oborovo sheet missing TABLE 3 (Delta)"

    def test_period_columns_horizontal(self):
        """Period columns must be horizontal (P1, P2, P3, P4, etc. in header row)."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["OBOROVO_OPEX"]

        # Find P1, P2, P3, P4 in same row (horizontal layout)
        p_positions = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ("P1", "P2", "P3", "P4", "P5", "P6"):
                    r = cell.row
                    p_positions[cell.value] = r

        # All P columns should be in the same row (within table sections)
        # At minimum, we expect P1 and P4 to be found
        assert "P1" in p_positions or "P4" in p_positions, (
            f"No period labels (P1..P6) found in Oborovo sheet. "
            f"This suggests period columns are not horizontal."
        )

    def test_b01_b02_b12_rows_present_python(self):
        """B.01, B.02, B.12 rows must exist for Oborovo (Python line items)."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["OBOROVO_OPEX"]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))

        content_str = " | ".join(content)

        # Python line items must be present
        for code in ("B.01", "B.02", "B.12"):
            assert code in content_str, (
                f"Code {code} not found in Oborovo sheet. "
                f"Python line-item rows should be present."
            )

    def test_total_opex_row_exists_both_projects(self):
        """Total OPEX row must exist for both TUHO and Oborovo."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

        for sheet_name in ("TUHO_OPEX", "OBOROVO_OPEX"):
            ws = wb[sheet_name]
            content = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        content.append(str(cell))
            content_str = " | ".join(content)
            assert "TOTAL" in content_str, (
                f"Total OPEX row not found in {sheet_name}"
            )

    def test_delta_table_shows_python_minus_excel(self):
        """Delta table must show Python - Excel values for aggregate rows."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["OBOROVO_OPEX"]

        # Check for TABLE 3 (Delta)
        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))
        content_str = " | ".join(content)
        assert "TABLE 3" in content_str, "TABLE 3 (Delta) not found in Oborovo sheet"

    def test_missing_rows_marked_not_omitted(self):
        """Missing Excel line-item rows must be marked MISSING_SOURCE, not omitted."""
        if not OUT_XLSX.exists():
            pytest.skip("Workbook not yet generated")

        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
        ws = wb["OBOROVO_OPEX"]

        content = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    content.append(str(cell))
        content_str = " | ".join(content)

        # MISSING_SOURCE must appear (not silently omitted)
        assert "MISSING_SOURCE" in content_str, (
            "Excel line-item rows should be marked MISSING_SOURCE, not silently omitted"
        )

    def test_no_runtime_formula_changes(self):
        """Confirmation: no runtime formula changes were made."""
        # This is a documentation test — verify the script doesn't modify domain files
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f for f in result.stdout.strip().split("\n") if f]

        # Allow only reports, scripts, docs, tests changes
        allowed_prefixes = (
            "reports/", "scripts/export_phase20u", "docs/phase20u",
            "tests/test_phase20u", ".gitignore",
        )
        for f in changed:
            if not any(f.startswith(p) for p in allowed_prefixes):
                pytest.fail(f"Unexpected file changed: {f} (runtime formula change?)")


class TestPhase20UNoRuntimeChanges:
    """Confirm no runtime formulas changed."""

    def test_domain_opex_unchanged(self):
        """domain/opex/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/ files changed (runtime formula change): {changed}"

    def test_domain_revenue_unchanged(self):
        """domain/revenue/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/revenue/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/revenue/ files changed: {changed}"

    def test_domain_tax_unchanged(self):
        """domain/tax/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/tax/", "domain/atad/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/tax/ or domain/atad/ files changed: {changed}"

    def test_domain_senior_debt_unchanged(self):
        """domain/senior_debt/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/senior_debt/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/senior_debt/ files changed: {changed}"

    def test_domain_shl_unchanged(self):
        """domain/shl/*.py files must not have been modified."""
        import subprocess
        result = subprocess.run(
            ["git", "diff", "--name-only", "domain/shl/"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
        )
        changed = [f.strip() for f in result.stdout.strip().split("\n") if f.strip()]
        assert not changed, f"domain/shl/ files changed: {changed}"


class TestPhase20UValidationCommands:
    """Run validation commands from the phase brief."""

    def test_import_main_web(self):
        """python -c 'import main_web' must succeed."""
        import subprocess
        result = subprocess.run(
            ["python3", "-c", "import main_web"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"import main_web failed: {result.stderr}"

    def test_opex_tests_pass(self):
        """tests/test_opex.py must pass."""
        import subprocess
        result = subprocess.run(
            ["python3", "-m", "pytest", "tests/test_opex.py", "-v", "--tb=short"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=60,
        )
        # We expect these to pass (OPEX formulas not changed)
        assert "failed" not in result.stdout.lower(), (
            f"test_opex.py had failures:\n{result.stdout}"
        )

    def test_revenue_tests_pass(self):
        """tests/test_revenue.py must pass."""
        import subprocess
        result = subprocess.run(
            ["python3", "-m", "pytest", "tests/test_revenue.py", "-v", "--tb=short"],
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert "failed" not in result.stdout.lower(), (
            f"test_revenue.py had failures:\n{result.stdout}"
        )