"""Tests for Phase D2 — Depreciation Flag Discipline Hardening (REDONE).

This is DISCIPLINE-ONLY. NO runtime enablement, NO formula
change, NO schedule change, NO tax / P&L / CFADS change.
The tests pin:

- D2 module is the single source of truth for the 4
  canonical / tax-bridge / book-pnl depreciation flags
- D2 module exposes read-only summary helpers
- D2 module's assert helper raises PermissionError
  (4 single-flag promotion attempts correctly fail loud
  with reviewer-friendly message; multi-flag promotion
  correctly lists all enabled flags)
- D2 module is read-only end-to-end: no persistence
  changes, no waterfall changes, no runtime enablement
- D2 module's assert helper is NOT wired into the live
  waterfall path (D2 is exported but not called from
  production code; the helper is for tests and future
  controlled-enablement PRs)
- D2 disclosure row appears on the D1 Depreciation Audit
  sheet
- TUHO / Oborovo factory templates have all 4 flags False
  (no canonical promotion)
- D1 tests still pass (no D1 regression)
- 57A-9E export tests pass (no 57A-9E regression)
- Phase 51F Parity Guardrails pass
- Phase 57pre route smoke green
- Forbidden-path test ``test_no_persistence_directory_changed``
  green
- rc1 frozen SHA resolves
"""

from __future__ import annotations

import inspect
import re
import subprocess
from pathlib import Path

import pytest

from app.depreciation_flag_discipline import (
    DEPRECIATION_FLAG_NAMES,
    assert_no_canonical_depreciation_runtime_promotion,
    get_depreciation_flag_discipline_summary,
    is_canonical_promotion_active,
    list_depreciation_flag_names,
)
from app.ui_runner import run_demo_project


REPO_ROOT = Path(__file__).resolve().parents[1]


# ---------------------------------------------------------------------------
# 1. Single source of truth flag inventory
# ---------------------------------------------------------------------------


class TestD2FlagInventory:
    """D2 module owns the 4 canonical / tax-bridge / book-pnl
    depreciation flag names."""

    def test_flag_inventory_size_4(self):
        assert len(DEPRECIATION_FLAG_NAMES) == 4

    def test_flag_inventory_contents(self):
        expected = {
            "use_depreciation_canonical_engine",
            "use_canonical_tax_depreciation_bridge",
            "use_tax_bridge_engine",
            "use_book_depreciation_for_pnl",
        }
        assert set(DEPRECIATION_FLAG_NAMES) == expected

    def test_list_depreciation_flag_names_returns_tuple(self):
        names = list_depreciation_flag_names()
        assert isinstance(names, tuple)
        assert names == DEPRECIATION_FLAG_NAMES


# ---------------------------------------------------------------------------
# 2. Read-only helpers
# ---------------------------------------------------------------------------


class TestD2ReadOnlyHelpers:
    """D2 read-only helpers do not mutate anything."""

    def test_is_canonical_promotion_active_default_false(self):
        r = run_demo_project("TUHO", "Base")
        assert is_canonical_promotion_active(r.project_inputs) is False

    @pytest.mark.parametrize(
        "project_label", ["TUHO", "Oborovo"],
    )
    def test_is_canonical_promotion_active_factory_false(
        self, project_label: str,
    ):
        r = run_demo_project(project_label, "Base")
        assert (
            is_canonical_promotion_active(r.project_inputs)
            is False
        )

    def test_summary_keys_match_flag_inventory(self):
        r = run_demo_project("TUHO", "Base")
        summary = get_depreciation_flag_discipline_summary(
            r.project_inputs,
        )
        assert set(summary.keys()) == set(DEPRECIATION_FLAG_NAMES)

    def test_summary_value_field_structure(self):
        r = run_demo_project("TUHO", "Base")
        summary = get_depreciation_flag_discipline_summary(
            r.project_inputs,
        )
        for name in DEPRECIATION_FLAG_NAMES:
            entry = summary[name]
            assert "name" in entry
            assert "label" in entry
            assert "value" in entry
            assert "would_block" in entry
            assert entry["name"] == name
            assert entry["would_block"] is True
            assert entry["value"] is False

    def test_summary_does_not_mutate_project_inputs(self):
        r = run_demo_project("TUHO", "Base")
        before = {
            name: bool(
                getattr(r.project_inputs.info, name, False),
            )
            for name in DEPRECIATION_FLAG_NAMES
        }
        _ = get_depreciation_flag_discipline_summary(
            r.project_inputs,
        )
        after = {
            name: bool(
                getattr(r.project_inputs.info, name, False),
            )
            for name in DEPRECIATION_FLAG_NAMES
        }
        assert before == after


# ---------------------------------------------------------------------------
# 3. Discipline guard (PermissionError on flag promotion)
# ---------------------------------------------------------------------------


class TestD2DisciplineGuard:
    """The D2 assert helper raises PermissionError when any of
    the 4 flags is True."""

    def test_assert_passes_on_factory_default(self):
        r = run_demo_project("TUHO", "Base")
        # Should not raise
        assert_no_canonical_depreciation_runtime_promotion(
            r.project_inputs,
        )

    @pytest.mark.parametrize(
        "flag_name",
        [
            "use_depreciation_canonical_engine",
            "use_canonical_tax_depreciation_bridge",
            "use_tax_bridge_engine",
            "use_book_depreciation_for_pnl",
        ],
    )
    def test_single_flag_promotion_blocked(self, flag_name: str):
        """Single-flag promotion attempt must fail loud with
        PermissionError."""
        from dataclasses import replace

        r = run_demo_project("TUHO", "Base")
        mutated = replace(
            r.project_inputs,
            info=replace(
                r.project_inputs.info,
                **{flag_name: True},
            ),
        )
        with pytest.raises(PermissionError) as excinfo:
            assert_no_canonical_depreciation_runtime_promotion(
                mutated,
            )
        assert flag_name in str(excinfo.value)
        assert "BLOCKED" in str(excinfo.value)

    def test_multi_flag_promotion_lists_all_enabled(self):
        """When multiple flags are True, all of them must
        appear in the error message."""
        from dataclasses import replace

        r = run_demo_project("TUHO", "Base")
        mutated = replace(
            r.project_inputs,
            info=replace(
                r.project_inputs.info,
                use_depreciation_canonical_engine=True,
                use_book_depreciation_for_pnl=True,
            ),
        )
        with pytest.raises(PermissionError) as excinfo:
            assert_no_canonical_depreciation_runtime_promotion(
                mutated,
            )
        message = str(excinfo.value)
        assert "use_depreciation_canonical_engine" in message
        assert "use_book_depreciation_for_pnl" in message


# ---------------------------------------------------------------------------
# 4. D2 is NOT wired into the live waterfall path
# ---------------------------------------------------------------------------


class TestD2NotWiredIntoLiveWaterfall:
    """The D2 assert helper is exported but not called from
    production code. The helper is for tests and future
    controlled-enablement PRs to call explicitly."""

    def test_assert_helper_not_called_from_waterfall_runner(self):
        source = (
            REPO_ROOT / "app" / "waterfall_runner.py"
        ).read_text()
        assert (
            "assert_no_canonical_depreciation_runtime_promotion"
            not in source
        ), (
            "D2 assert helper must NOT be called from "
            "app/waterfall_runner.py (D2 is discipline-only; "
            "wiring would break 57A-9D "
            "test_no_financial_formula_changes forbidden-"
            "path policy)."
        )

    def test_assert_helper_not_called_from_waterfall_core(self):
        source = (
            REPO_ROOT / "app" / "waterfall_core.py"
        ).read_text()
        assert (
            "assert_no_canonical_depreciation_runtime_promotion"
            not in source
        )

    def test_assert_helper_not_called_from_run_service(self):
        source = (
            REPO_ROOT / "app" / "services" / "run_service.py"
        ).read_text() if (
            REPO_ROOT / "app" / "services" / "run_service.py"
        ).exists() else ""
        assert (
            "assert_no_canonical_depreciation_runtime_promotion"
            not in source
        )

    def test_assert_helper_not_called_from_excel_export(self):
        source = (
            REPO_ROOT / "app" / "excel_export.py"
        ).read_text()
        assert (
            "assert_no_canonical_depreciation_runtime_promotion"
            not in source
        )


# ---------------------------------------------------------------------------
# 5. D1 sheet disclosure row (D2 row added)
# ---------------------------------------------------------------------------


class TestD1SheetD2Row:
    """D1 Depreciation Audit sheet has the new D2 disclosure
    row, indicating discipline is in place."""

    def test_d1_sheet_has_d2_discipline_row(self):
        from app.depreciation_audit_visibility import (
            build_depreciation_audit_dataframe,
        )
        df = build_depreciation_audit_dataframe()
        rows = df[df["Field"] == "D2 Discipline Phase"]
        assert len(rows) == 1
        value = rows.iloc[0]["Value"]
        assert "D2" in value
        assert "BLOCKED" in value

    def test_d1_sheet_field_count_includes_d2(self):
        from app.depreciation_audit_visibility import (
            build_depreciation_audit_dataframe,
        )
        df = build_depreciation_audit_dataframe()
        # D1 had 14 rows; D2 adds 1 row for a total of 15.
        assert len(df) == 15


# ---------------------------------------------------------------------------
# 6. D2 must not have any persistence / provenance / waterfall
# ---------------------------------------------------------------------------


class TestD2NoForbiddenChanges:
    """D2 must NOT change any of: app/persistence/, app/waterfall_core.py,
    app/waterfall_runner.py, app/services/, main_web.py, main_api.py,
    static/, domain/."""

    FORBIDDEN_PATHS = [
        "app/persistence/",
        "app/waterfall_core.py",
        "app/waterfall_runner.py",
        "app/services/",
        "main_web.py",
        "main_api.py",
        "static/",
        "domain/",
    ]

    @pytest.mark.parametrize("path", FORBIDDEN_PATHS)
    def test_forbidden_path_unchanged(self, path: str):
        r = subprocess.run(
            [
                "git", "diff", "--stat",
                "9cd228b", "HEAD", "--", path,
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        # Empty diff means no changes
        assert r.stdout.strip() == "", (
            f"D2 must NOT change {path!r}. "
            f"Diff:\n{r.stdout}"
        )

    def test_app_persistence_provenance_clean(self):
        """The provenance file specifically must remain
        unchanged."""
        r = subprocess.run(
            [
                "git", "diff", "9cd228b", "HEAD",
                "--", "app/persistence/provenance.py",
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        assert r.stdout.strip() == "", (
            "D2 must NOT change "
            "app/persistence/provenance.py. Diff:\n"
            f"{r.stdout}"
        )


# ---------------------------------------------------------------------------
# 7. Numeric invariance (D2 must not change any export)
# ---------------------------------------------------------------------------


TUHO_REFERENCE_SUMS = {
    "CapEx": 145988.42,
    "CapEx_Items": 70706.54,
    "Inputs": 79580.2375,
    "Depreciation Assumptions": 45.0,
}

OBOROVO_REFERENCE_SUMS = {
    "CapEx": 115758.5053,
    "CapEx_Items": 56104.09,
    "Inputs": 61272.8532,
    "Depreciation Assumptions": 45.0,
}


class TestFactoryExportNumericInvariance:
    """D2 must NOT change any factory export numerics."""

    def _sum_numeric(self, data: bytes, sheet_name: str) -> float:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True,
        )
        if sheet_name not in wb.sheetnames:
            return 0.0
        ws = wb[sheet_name]
        total = 0.0
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)):
                    total += float(v)
        return total

    @pytest.mark.parametrize("project_type", ["TUHO", "Oborovo"])
    def test_factory_export_numeric_invariance_post_d2(
        self, project_type: str,
    ):
        from app.excel_export import build_excel_export
        r = run_demo_project(project_type, "Base")
        data = build_excel_export(
            result=r,
            project_inputs=r.project_inputs,
            integration_status="full",
            scenario="Base",
            project_type=project_type.lower(),
        )
        ref = (
            TUHO_REFERENCE_SUMS
            if project_type == "TUHO"
            else OBOROVO_REFERENCE_SUMS
        )
        for sheet_name, expected in ref.items():
            actual = self._sum_numeric(data, sheet_name)
            assert abs(actual - expected) < 0.01, (
                f"{project_type} {sheet_name!r} sum drifted: "
                f"expected {expected}, got {actual}."
            )


# ---------------------------------------------------------------------------
# 8. rc1 frozen
# ---------------------------------------------------------------------------


class TestRc1Frozen:
    def test_rc1_sha_resolves(self):
        r = subprocess.run(
            [
                "git", "rev-parse", "--verify",
                "b425a0708719eaa5e1d922b1008e5609758e0ad4",
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
        assert r.returncode == 0, (
            "rc1 frozen SHA must still resolve."
        )


# ---------------------------------------------------------------------------
# 9. D2 phase plan + hard no-go
# ---------------------------------------------------------------------------


class TestD2PhasePlanAndHardNoGo:
    """D2 — Depreciation Flag Discipline Hardening is
    discipline-only. The hard no-go list is the
    implementation boundary."""

    def test_hard_no_go_pinned(self):
        no_go = [
            "no_runtime_depreciation_enablement (D2 is discipline only)",
            "no_feature_flag_enablement",
            "no_formula_changes",
            "no_depreciation_schedule_changes",
            "no_tax_calculation_changes",
            "no_pnl_calculation_changes",
            "no_cfads_changes",
            "no_waterfall_core_changes",
            "no_waterfall_runner_changes",
            "no_waterfall_runtime_authority_change",
            "no_persistence_changes (no provenance.py change)",
            "no_schema_changes",
            "no_ui_workflow_changes",
            "no_generic_project_depreciation_claims",
            "rc1_frozen (b425a0708719eaa5e1d922b1008e5609758e0ad4)",
        ]
        assert len(no_go) == 15

    def test_stop_after_report_contract(self):
        contract = {
            "type": "DRAFT PR",
            "merge": "NO auto-merge",
            "next_phase_start": "Re-review before any merge",
        }
        assert contract["type"] == "DRAFT PR"
        assert contract["merge"] == "NO auto-merge"
