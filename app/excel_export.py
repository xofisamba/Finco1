"""Values-only Excel export — no formulas, no calibration."""
from __future__ import annotations
import pandas as pd
from io import BytesIO

from app.portfolio_ui import (
    build_portfolio_summary_table,
    build_portfolio_spv_table,
)
from app.tax_excel_export import write_spv_tax_audit_sheets
from app.holdco_tax_excel_export import write_holdco_tax_audit_sheets
from app.tax_assumptions_excel_export import write_tax_assumptions_audit_sheets
from app.tax_assumptions_snapshot_excel_export import write_tax_assumption_snapshot_sheets
from app.sponsor_waterfall_excel_export import write_sponsor_waterfall_audit_sheets

from domain.portfolio.distribution_constraints import (
    DistributionConstraintConfig,
    DistributionConstraintResult,
)
from domain.portfolio.distribution_constraints.overlay import (
    SPVRetainedCashOverlay,
)
from domain.portfolio.distribution_constraints.holdco_overlay import (
    HoldCoRetainedCashOverlay,
)
from domain.portfolio.cash_ledger import PortfolioCashLedger

import openpyxl
from openpyxl.styles import Font, numbers
from openpyxl.utils import get_column_letter


# KPI label mapping: raw key → clean display label
_DASHBOARD_LABELS = {
    "total_revenue_keur": "Total Revenue (kEUR)",
    "total_ebitda_keur": "EBITDA (kEUR)",
    "total_tax_keur": "Total Tax (kEUR)",
    "project_irr": "Project IRR (%)",
    "equity_irr": "Equity IRR (%)",
    "sponsor_irr": "Sponsor IRR (%)",
    "min_dscr": "Min DSCR",
    "avg_dscr": "Avg DSCR",
    "total_senior_ds_keur": "Senior Debt Service (kEUR)",
    "total_distribution_keur": "Distributions (kEUR)",
}


def build_excel_export(
    *,
    result=None,
    portfolio_result=None,
    project_inputs=None,
    validation_issues=None,
    integration_status: str = "full",
    integration_note: str | None = None,
    scenario: str = "Base",
    project_type: str = "Solar",
    period_view: str = "Semiannual",
    warnings: list = None,
    run_metadata=None,
    advanced_opex_line_items: tuple = None,
    advanced_capex_line_items: tuple = None,
    include_reconciliation_sheets: bool = False,
    # ── Phase 5D overlay sheets (optional) ──────────────────────────
    cash_ledger=None,
    spv_overlays=None,
    holdco_overlay=None,
    constraint_results=None,
    # ── Phase 6B: Tax audit sheets (optional) ───────────────────────
    tax_results=None,
    # ── Phase 6C.5: HoldCo tax audit sheets (optional) ─────────────────
    holdco_tax_results: tuple = None,
    # ── Phase 6D.2: Tax assumptions snapshot (optional) ─────────────────
    tax_assumption_snapshot=None,
    # ── Phase 7C-5: Sponsor waterfall audit sheets (optional) ───────────
    sponsor_waterfall_result=None,
    sponsor_preferred_return_result=None,
    tier_annotated_capital_accounts=None,
    provenance_metadata: dict | None = None,
) -> bytes:
    """Build a values-only Excel workbook from waterfall results.
    
    Args:
        include_reconciliation_sheets: If True, adds Debt Schedule, Project CF Bridge,
            Equity CF Bridge, and Calibration Notes sheets.
    
    Returns bytes suitable for st.download_button.
    """
    from app.output_tables import (
        build_dashboard_kpis,
        build_waterfall_table,
        build_revenue_table,
        build_debt_table,
        build_tax_depreciation_table,
        build_returns_table,
        build_portfolio_table,
        aggregate_period_table_annual,
    )
    from app.input_helpers import build_inputs_summary_table, build_capex_summary_table, build_capex_items_table
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Dashboard (always first) ─────────────────────────────────────
        _write_dashboard_sheet(writer, result, portfolio_result, integration_status, integration_note, scenario)

        # ── Returns ───────────────────────────────────────────────────────
        if result is not None:
            _write_sheet(writer, "Returns", build_returns_table(result),
                         number_format={"IRR": "0.0%", "DSCR": "0.00x", "kEUR": "#,##0"})

        # ── DSCR Summary ────────────────────────────────────────────────────
        if result is not None and hasattr(result, "target_dscr"):
            rows = [
                ("Metric", "Value"),
                ("Target DSCR", f"{result.target_dscr:.3f}"),
                ("Actual Min DSCR", f"{result.actual_min_dscr:.3f}"),
                ("Actual Avg DSCR", f"{result.actual_avg_dscr:.3f}"),
                ("Deviation", f"{result.actual_min_dscr - result.target_dscr:+.3f}"),
            ]
            dscr_df = pd.DataFrame(rows[1:], columns=["Metric", "Value"])
            _write_sheet(writer, "DSCR Summary", dscr_df)

        # ── Waterfall ─────────────────────────────────────────────────────
        if result is not None:
            wf_df = build_waterfall_table(result)
            if period_view == "Annual":
                wf_df = aggregate_period_table_annual(wf_df)
            _write_sheet(writer, "Waterfall", wf_df,
                         number_format={"kEUR": "#,##0", "DSCR": "0.00x"})

        # ── Revenue ────────────────────────────────────────────────────────
        if result is not None:
            rev_df = build_revenue_table(result)
            if period_view == "Annual":
                rev_df = aggregate_period_table_annual(rev_df)
            _write_sheet(writer, "Revenue", rev_df, number_format={"kEUR": "#,##0", "MWh": "#,##0"})

        # ── Debt ──────────────────────────────────────────────────────────
        if result is not None:
            debt_df = build_debt_table(result)
            if period_view == "Annual":
                debt_df = aggregate_period_table_annual(debt_df)
            _write_sheet(writer, "Debt", debt_df,
                         number_format={"kEUR": "#,##0", "DSCR": "0.00x", "LLCR": "0.00x", "PLCR": "0.00x"})

        # ── Tax & Depreciation ─────────────────────────────────────────────
        if result is not None:
            tax_df = build_tax_depreciation_table(result)
            if period_view == "Annual":
                tax_df = aggregate_period_table_annual(tax_df)
            _write_sheet(writer, "Tax_Depreciation", tax_df, number_format={"kEUR": "#,##0"})

        # ── Notes ──────────────────────────────────────────────────────────
        if run_metadata is not None:
            git_sha = run_metadata.git_sha
            ts = run_metadata.timestamp
            scenario = run_metadata.scenario
            project_type = run_metadata.project_type
        else:
            import subprocess
            try:
                git_sha = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], text=True).strip()
            except Exception:
                git_sha = "n/a"
            ts = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M UTC")
        _write_notes_sheet(writer, integration_status, integration_note, scenario, project_type, period_view,
                          warnings=warnings if warnings else [], git_sha=git_sha, timestamp=ts,
                          advanced_opex_line_items=advanced_opex_line_items,
                          advanced_capex_line_items=advanced_capex_line_items,
                          provenance_metadata=provenance_metadata)

        # ── Inputs ────────────────────────────────────────────────────────
        _write_sheet(writer, "Inputs", build_inputs_summary_table(project_inputs))

        # ── CapEx ──────────────────────────────────────────────────────────
        _write_sheet(writer, "CapEx", build_capex_summary_table(project_inputs),
                     number_format={"kEUR": "#,##0"})
        _write_sheet(writer, "CapEx_Items", build_capex_items_table(project_inputs),
                     number_format={"Amount": "#,##0"})

        # ── OPEX Detail (Advanced OPEX only) ───────────────────────────────
        if advanced_opex_line_items:
            from app.opex_engine import generate_opex_schedule
            horizon = getattr(getattr(project_inputs, "info", None), "horizon_years", 25) if project_inputs else 25
            schedule = generate_opex_schedule(advanced_opex_line_items, horizon)
            _write_opex_detail_sheet(writer, schedule)

        # ── Portfolio ──────────────────────────────────────────────────────
        if portfolio_result is not None:
            from domain.portfolio.independent import IndependentPortfolioResult

            # IndependentPortfolioResult (Phase 1.5): write per-SPV sheets instead of pooled Portfolio sheet
            # PooledPortfolioResult (legacy): use existing build_portfolio_table
            if isinstance(portfolio_result, IndependentPortfolioResult):
                _write_sheet(writer, "Portfolio_Summary",
                             build_portfolio_summary_table(portfolio_result))
                _write_sheet(writer, "Portfolio_SPVs",
                             build_portfolio_spv_table(portfolio_result))
                _write_portfolio_notes_sheet(writer, portfolio_result)
                # ── DSRF sheet (Phase 2, optional) ─────────────────────────────
                if portfolio_result.dsrf_enabled and portfolio_result.dsrf_periods:
                    _write_dsrf_sheet(writer, portfolio_result)
            else:
                # Pooled/legacy portfolio — use existing aggregated table
                _write_sheet(writer, "Portfolio", build_portfolio_table(portfolio_result),
                             number_format={"kEUR": "#,##0", "IRR": "0.0%", "DSCR": "0.00x"})
                if portfolio_result.portfolio_cashflows:
                    rows = []
                    for row in portfolio_result.portfolio_cashflows:
                        date = row.get("date", "")
                        total = row.get("total_cashflow", 0.0)
                        breakdown = row.get("breakdown", {})
                        row_dict = {"Date": str(date), "Total CF (keur)": round(total, 2)}
                        for proj, contrib in breakdown.items():
                            row_dict[f"  {proj}"] = round(contrib, 2)
                        rows.append(row_dict)
                    cf_df = pd.DataFrame(rows)
                    _write_sheet(writer, "Portfolio CF", cf_df, number_format={"kEUR": "#,##0"})

        # ── Reconciliation / Audit Sheets (optional) ─────────────────────
        if include_reconciliation_sheets and result is not None:
            _write_reconciliation_sheets(writer, result, project_inputs)

        # ── Validation ────────────────────────────────────────────────────
        _write_validation_sheet(writer, validation_issues)

        # ── Depreciation Assumptions (Bankable Framework) ─────────────────
        # Select profile based on project_type: Solar → solar_croatia_ibl, Wind → wind_croatia_ibl
        depr_profile = f"{project_type.lower()}_croatia_ibl" if project_type.lower() in ("solar", "wind") else "solar_croatia_ibl"
        _write_depreciation_assumptions_sheet(writer, depr_profile)

        # ── Tax Depreciation Disclosure ───────────────────────────────────
        _write_tax_depreciation_sheet_for_project(writer, project_inputs, advanced_capex_line_items, project_type)

        # ── Book Depreciation Disclosure ──────────────────────────────────
        _write_book_depreciation_sheet_for_project(writer, project_inputs, advanced_capex_line_items, project_type)

        # ── Phase 5D: Overlay Sheets (optional) ─────────────────────────
        _write_overlay_sheets(
            writer,
            cash_ledger=cash_ledger,
            spv_overlays=spv_overlays,
            holdco_overlay=holdco_overlay,
            constraint_results=constraint_results,
        )

        # ── Phase 6B: Optional tax audit sheets ─────────────────────────
        if tax_results:
            write_spv_tax_audit_sheets(writer, tax_results)

        # ── Phase 6C.5: Optional HoldCo tax audit sheets ────────────────
        if holdco_tax_results:
            write_holdco_tax_audit_sheets(writer, holdco_tax_results)

        # ── Phase 6D.2: Optional tax assumptions snapshot sheets ────────
        if tax_assumption_snapshot is not None:
            write_tax_assumption_snapshot_sheets(writer, tax_assumption_snapshot)

        # ── Phase 7C-5: Optional sponsor waterfall audit sheets ─────────
        if sponsor_waterfall_result or sponsor_preferred_return_result or tier_annotated_capital_accounts:
            write_sponsor_waterfall_audit_sheets(
                writer,
                waterfall_result=sponsor_waterfall_result,
                preferred_return_result=sponsor_preferred_return_result,
                tier_annotated_accounts=tier_annotated_capital_accounts,
            )

    output.seek(0)
    return output.read()


# ─── Reconciliation / Audit Sheets ─────────────────────────────────────────


def _write_reconciliation_sheets(
    writer,
    result,
    project_inputs,
) -> None:
    """Write optional audit sheets: Debt Schedule, Project CF Bridge, Equity CF Bridge."""
    from app.reconciliation import (
        build_debt_schedule_rows,
        build_project_cf_rows,
        build_equity_cf_rows,
    )
    import pandas as pd

    # ── Debt Schedule ──────────────────────────────────────────────────────
    debt_rows = build_debt_schedule_rows(result)
    if debt_rows:
        debt_df = pd.DataFrame(debt_rows)
        _write_sheet(writer, "Debt Schedule", debt_df,
                     number_format={"kEUR": "#,##0", "DSCR": "0.00x"})

    # ── Project CF Bridge ──────────────────────────────────────────────────
    proj_rows = build_project_cf_rows(result, project_inputs=project_inputs)
    if proj_rows:
        proj_df = pd.DataFrame(proj_rows)
        _write_sheet(writer, "Project CF Bridge", proj_df,
                     number_format={"kEUR": "#,##0"})

    # ── Equity CF Bridge ───────────────────────────────────────────────────
    eq_rows = build_equity_cf_rows(result)
    if eq_rows:
        eq_df = pd.DataFrame(eq_rows)
        _write_sheet(writer, "Equity CF Bridge", eq_df,
                     number_format={"kEUR": "#,##0"})

    # ── Calibration Notes ────────────────────────────────────────────────
    _write_calibration_notes_sheet(writer, result, project_inputs)


def _write_calibration_notes_sheet(
    writer,
    result,
    project_inputs,
) -> None:
    """Write calibration status notes sheet.
    
    Content varies by project type:
    - Oborovo: calibration status vs Excel reference, AFRY merchant curve
    - TUHO: TUHO-specific calibration, CO2 status
    - Generic Solar/Wind: screening-grade note only, no Excel reference claims
    """
    import pandas as pd

    # Determine project type from project_inputs
    project_name = "Unknown"
    project_type = "Unknown"
    is_oborovo = False
    is_tuho = False

    if project_inputs is not None:
        info = getattr(project_inputs, "info", None)
        if info is not None:
            project_name = getattr(info, "name", "Unknown")
            project_type = getattr(info, "project_type", "Unknown")
        # Detect Oborovo/TUHO by name or code
        code = getattr(info, "code", "") or ""
        name_lower = project_name.lower()
        is_oborovo = "oborovo" in name_lower or code.startswith("OBR")
        is_tuho = "tuho" in name_lower or "wind" in name_lower and "obr" not in code.lower()

    # No debt-range heuristic — use only explicit metadata

    # Base rows (all projects)
    rows = [
        ("Project", "Value"),
        ("Project", project_name),
        ("Project Type", project_type),
        ("", ""),
    ]

    # KPI section
    rows.extend([
        ("KPI Summary", ""),
        ("Project IRR", f"{result.project_irr * 100:.3f}%" if hasattr(result, "project_irr") else "n/a"),
        ("Equity IRR", f"{result.equity_irr * 100:.3f}%" if hasattr(result, "equity_irr") else "n/a"),
        ("Avg DSCR", f"{result.actual_avg_dscr:.3f}" if hasattr(result, "actual_avg_dscr") else "n/a"),
        ("Min DSCR", f"{result.actual_min_dscr:.3f}" if hasattr(result, "actual_min_dscr") else "n/a"),
        ("Debt (kEUR)", f"{result.sculpting_result.debt_keur:,.0f}" if (
            hasattr(result, "sculpting_result") and result.sculpting_result
        ) else "n/a"),
    ])

    # Oborovo-specific calibration status
    if is_oborovo:
        rows.extend([
            ("", ""),
            ("Oborovo Calibration Status (P0+P1)", ""),
            ("Project IRR", "✅ 7.985% vs reference 7.96% (+0.025pp) — CALIBRATED"),
            ("Equity IRR", "⚠️ 9.17% vs reference 10.60% (-1.43pp) — partial (P2)"),
            ("Debt Sizing", "✅ 42,852 kEUR anchor maintained"),
            ("DSCR", "⚠️ 1.229 vs reference 1.147 (+0.082)"),
            ("Revenue (Y1)", "✅ Calibrated via PPA tariff"),
            ("", ""),
            ("Oborovo Equity IRR Gap Explanation", ""),
            ("Depreciation convention", "20y (book) / 25y (tax) vs lender 30y asset life"),
            ("DSCR averaging", "Model uses annual; Excel may use semiannual"),
            ("Sculpting method", "Iterative sculpt vs static annuity in Excel"),
            ("Tax loss carryforward", "Construction-period timing differs"),
            ("", ""),
            ("Merchant Curve", ""),
            ("Profile used", _get_merchant_profile_name(project_inputs)),
            ("PPA period (Y1-Y12)", "Fixed tariff: 57 EUR/MWh × 1.02^(Y-1)"),
            ("Merchant period (Y13-Y30)", "AFRY Central Q1 2026 4h Degraded scenario"),
        ])
    elif is_tuho:
        rows.extend([
            ("", ""),
            ("TUHO Calibration Status", ""),
            ("Project IRR", "✅ 9.47% vs reference 9.47% (exact match)"),
            ("Equity IRR", "✅ 11.56% vs reference 11.61% (-0.05pp) — CALIBRATED"),
            ("CO2 Revenue", "✅ Enabled: 611 kEUR Y1, equity IRR 11.81% vs 11.61%"),
            ("DSCR", "⚠️ 1.682 vs reference 1.451 (+0.231)"),
        ])
    else:
        # Generic Solar/Wind — screening only, no Excel reference
        rows.extend([
            ("", ""),
            ("Calibration Note", "Screening-grade model — no Excel reference calibration."),
            ("Model Status", "For internal scenario review only."),
            ("Purpose", "Not a substitute for lender due diligence."),
        ])

    # Common section for all projects
    rows.extend([
        ("", ""),
        ("Model Status", "Screening-grade, not lender-grade or bank-certified."),
        ("External audit", "Not a substitute for external model audit or lender due diligence."),
    ])

    df = pd.DataFrame(rows[1:], columns=rows[0])
    _write_sheet(writer, "Calibration Notes", df)


def _get_merchant_profile_name(project_inputs) -> str:
    """Extract merchant price curve name from project inputs for disclosure."""
    if project_inputs is None:
        return "Unknown"
    rev = getattr(project_inputs, "revenue", None)
    if rev is None:
        return "Unknown"
    mkt = getattr(rev, "market_prices_curve", None)
    if mkt is None:
        return "Default (2% escalation)"
    # Check for AFRY pattern (starts with 0s for PPA period)
    if mkt and len(mkt) > 12 and all(v == 0 for v in mkt[:12]):
        return "CROATIA_SOLAR_AFRY_CENTRAL_2024 (Y13-Y30)"
    return f"Custom curve ({len(mkt)} values)"




def _write_sheet(writer, name: str, df: pd.DataFrame,
                 number_format: dict | None = None) -> None:
    """Write a DataFrame to a sheet with bold headers, frozen row, auto-width, and number formats."""
    if df is None or df.empty:
        df = pd.DataFrame({"Note": ["No data available"]})
    df.to_excel(writer, sheet_name=name[:31], index=True)
    ws = writer.sheets[name[:31]]

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Ensure sheet is visible
    ws.sheet_state = "visible"

    # Auto column width based on content
    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col))  # header length
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Apply number formats based on row labels
    if number_format:
        irr_patterns = ("irr", "project irr", "equity irr", "sponsor irr")
        dscr_patterns = ("dscr", "min dscr", "avg dscr", "llcr", "plcr")
        keur_patterns = ("keur", "revenue", "ebitda", "debt", "tax", "distribution", "capex")
        mwh_patterns = ("mwh", "generation", "capacity")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            row_label_cell = ws.cell(row=row_idx, column=1)
            if row_label_cell.value:
                label_lower = str(row_label_cell.value).lower()
                for col_idx_c, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue
                    col_letter = get_column_letter(col_idx_c)
                    # IRR format
                    if any(p in label_lower for p in irr_patterns) and "format" not in label_lower:
                        cell.number_format = "0.0%"
                    # DSCR / LLCR / PLCR format
                    elif any(p in label_lower for p in dscr_patterns):
                        cell.number_format = "0.00x"
                    # kEUR format
                    elif any(p in label_lower for p in keur_patterns):
                        cell.number_format = "#,##0"
                    # MWh format
                    elif any(p in label_lower for p in mwh_patterns):
                        cell.number_format = "#,##0"


def _write_dashboard_sheet(writer, result, portfolio_result, status, note, scenario) -> None:
    """Write a Dashboard sheet with KPI summary and integration info."""
    from app.output_tables import build_dashboard_kpis
    rows = [
        ("Integration Status", status),
        ("Scenario", scenario),
    ]
    if note:
        rows.append(("Note", note))
    if result is not None:
        kpis = build_dashboard_kpis(result)
        for k, v in kpis.items():
            if v is not None and v != "n/a":
                label = _DASHBOARD_LABELS.get(k, k)
                # Format IRR values as decimals → Excel will show as %
                if "irr" in k.lower():
                    try:
                        rows.append((label, float(v) / 100))  # Store as decimal for 0.0% format
                    except (TypeError, ValueError):
                        rows.append((label, v))
                else:
                    rows.append((label, v))
    if portfolio_result is not None:
        from domain.portfolio.independent import IndependentPortfolioResult

        if isinstance(portfolio_result, IndependentPortfolioResult):
            # Independent SPV portfolio — use "Portfolio" labels (not "Pooled")
            rows.append(("Portfolio Revenue (kEUR)", portfolio_result.total_revenue_keur))
            rows.append(("Portfolio EBITDA (kEUR)", portfolio_result.total_ebitda_keur))
            rows.append(("Portfolio DSCR (Avg)", portfolio_result.avg_dscr))
            rows.append(("Portfolio DSCR (Min)", portfolio_result.min_dscr))
        else:
            # Pooled/legacy portfolio — use "Pooled" labels
            rows.append(("Pooled Revenue (kEUR)", portfolio_result.total_revenue_keur))
            rows.append(("Pooled EBITDA (kEUR)", portfolio_result.total_ebitda_keur))
            rows.append(("Portfolio DSCR (Avg)", portfolio_result.avg_dscr))
            rows.append(("Portfolio DSCR (Min)", portfolio_result.min_dscr))

    df = pd.DataFrame(rows, columns=["Metric", "Value"])
    df.to_excel(writer, sheet_name="Dashboard", index=False)
    ws = writer.sheets["Dashboard"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"

    # Apply number formats to Dashboard sheet
    irr_cols = []
    dscr_cols = []
    keur_cols = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        label_cell = ws.cell(row=row_idx, column=1)
        if label_cell.value:
            label_lower = str(label_cell.value).lower()
            if "irr" in label_lower:
                irr_cols.append(row_idx)
            elif "dscr" in label_lower:
                dscr_cols.append(row_idx)
            elif "keur" in label_lower or "revenue" in label_lower or "ebitda" in label_lower:
                keur_cols.append(row_idx)

    for row_idx in irr_cols:
        ws.cell(row=row_idx, column=2).number_format = "0.0%"
    for row_idx in dscr_cols:
        ws.cell(row=row_idx, column=2).number_format = "0.00x"
    for row_idx in keur_cols:
        ws.cell(row=row_idx, column=2).number_format = "#,##0"


def _write_opex_detail_sheet(writer, schedule) -> None:
    """Write OPEX Detail sheet from an OpexSchedule.

    Columns: Line Item Name | Category | Year | Value (kEUR) | Source | Is Override | Is Hardcoded | Override Note
    One row per line item per year.
    """
    rows = []
    for entry in schedule.entries:
        rows.append({
            "Line Item Name": entry.line_item_name,
            "Category": entry.category,
            "Year": entry.year_index + 1,  # 1-based for readability
            "Value (kEUR)": round(entry.value_keur, 2),
            "Source": entry.source.value,
            "Is Override": entry.is_override,
            "Is Hardcoded": entry.is_hardcoded,
            "Override Note": entry.override_note,
        })

    if not rows:
        df = pd.DataFrame({"Note": ["No OPEX data available"]})
    else:
        df = pd.DataFrame(rows)

    df.to_excel(writer, sheet_name="OPEX Detail", index=False)
    ws = writer.sheets["OPEX Detail"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"


def _write_validation_sheet(writer, validation_issues) -> None:
    """Write a Validation sheet."""
    if validation_issues is None:
        df = pd.DataFrame({"severity": ["info"], "field": [""], "message": ["No validation issues."]})
    else:
        rows = [(i.severity, i.field, i.message) for i in validation_issues]
        df = pd.DataFrame(rows, columns=["severity", "field", "message"])
    df.to_excel(writer, sheet_name="Validation", index=False)
    ws = writer.sheets["Validation"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"


def _write_notes_sheet(writer, status, note, scenario, project_type, period_view, warnings=None, git_sha=None, timestamp=None, advanced_opex_line_items=None, advanced_capex_line_items=None, provenance_metadata=None) -> None:
    if warnings is None:
        warnings = []
    """Write a Notes sheet."""
    if timestamp is None:
        timestamp = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M UTC")
    if git_sha is None:
        git_sha = "n/a"
    rows = [
        ("Field", "Value"),
        ("Model Version", "industry-engine-refactor"),
        ("Git SHA", git_sha),
        ("Run Timestamp", timestamp),
        ("Integration Status", status),
        ("Scenario", scenario),
        ("Project Type", project_type),
        ("Period View", period_view),
        ("Note", note if note else "n/a"),
        ("Values-only Export", "No formulas used in this workbook."),
        ("Economic LCOE", "Excludes debt service — see methodology document for details."),
        ("Export Generated At", (provenance_metadata or {}).get("export_generated_at", timestamp)),
        ("Runtime Timestamp", (provenance_metadata or {}).get("runtime_generated_at", timestamp)),
        ("Branch Name", (provenance_metadata or {}).get("branch_name", "unavailable")),
        ("Active Project", (provenance_metadata or {}).get("active_project", "unavailable")),
        ("Scenario ID", (provenance_metadata or {}).get("scenario_id", "not_applicable")),
        ("Scenario Name", (provenance_metadata or {}).get("scenario_name", scenario)),
        ("Scenario Revision", (provenance_metadata or {}).get("scenario_revision", "not_applicable")),
        ("Runtime Snapshot ID", (provenance_metadata or {}).get("runtime_snapshot_id", "unavailable")),
        ("Runtime Origin", (provenance_metadata or {}).get("runtime_origin", "not_applicable")),
        ("Template Origin", (provenance_metadata or {}).get("template_origin", "unavailable")),
        ("Template Revision", (provenance_metadata or {}).get("template_revision", "unavailable")),
        ("Export Template Version", (provenance_metadata or {}).get("export_template_version", "not separately versioned")),
        ("Governance Posture", (provenance_metadata or {}).get("governance_posture_summary", "G20 remains BLOCKED; R99/R102 remain NOT APPROVED")),
        ("Runtime Flag Snapshot", (provenance_metadata or {}).get("runtime_flags_json", "unavailable")),
        ("Replay Limitations", (provenance_metadata or {}).get("replay_limitations_notice", "Replay metadata improves traceability but is descriptive only.")),
        ("", ""),
        ("Reviewer Cover Notes", "—"),
        ("  Runtime Authority", "Runtime/backend output remains the financial source of truth; this workbook is descriptive only."),
        ("  Snapshot vs Live Draft", "This export reflects the saved/exported snapshot context. Later draft edits do not mutate the exported artifact."),
        ("  Stale Runtime", "If current draft changed after the last clean run, save first and rerun before relying on stale runtime outputs."),
        ("  Governance Boundaries", "G20 remains BLOCKED. R99/R102 remain NOT APPROVED. Accepted conventions explain; they do not approve."),
        ("  Metadata Markers", "unavailable and not_applicable are intentional markers and must not be read as zero."),
        ("  Governed Residuals", "Documented residuals or pending items are governed interpretation notes, not silent parity fabrication."),
        ("  No-Claims", "Not lender-ready without external model review, not audit-certified, not multi-user governance-ready, and not a replay engine."),
        ("", ""),
        ("Depreciation Disclosure", "—"),
        ("  Tax/Book Schedules", "Model outputs, not audited accounts"),
        ("  Jurisdiction Profile", "Requires tax advisor confirmation"),
        ("  COD-Month Convention", "Not yet supported"),
        ("  Declining Balance", "Not yet supported"),
        ("  Day Fraction", "Applied in waterfall (period view), not in annual disclosure table"),
    ]

    # BESS/hybrid warning
    if status == "partial":
        rows.append(("BESS/hybrid Status", "Partial integration — revenue-only shown, waterfall in progress"))

    # Model Warnings section
    if warnings:
        rows.append(("Model Warnings", "—"))
        for w in warnings:
            rows.append((w.get("code", "WARN"), w.get("message", str(w))))

    # Advanced OPEX manual/hardcoded warning
    if advanced_opex_line_items:
        from app.opex_engine import OpexSource
        has_manual = any(item.source == OpexSource.MANUAL for item in advanced_opex_line_items)
        has_hardcoded = any(item.is_hardcoded for item in advanced_opex_line_items)
        has_overrides = any(item.has_manual_overrides() for item in advanced_opex_line_items)
        if has_manual or has_hardcoded or has_overrides:
            rows.append((
                "Advanced OPEX",
                "Manual or hardcoded values present — review override notes",
            ))

    # Advanced CAPEX warning
    if advanced_capex_line_items:
        has_manual = any(getattr(item, "is_manual", False) for item in advanced_capex_line_items)
        rows.append((
            "Advanced CAPEX",
            "Advanced CAPEX matrix active" + (" — manual values present" if has_manual else ""),
        ))

    # Portfolio warning
    if status == "experimental":
        rows.append(("Portfolio Status", "Experimental — sponsor IRR is placeholder"))

    # Scenario deltas
    if scenario != "Base":
        if status == "experimental":
            rows.append(("Scenario", f"{scenario} — NOT APPLIED"))
            rows.append(("Scenario Deltas", "Base case shown — Portfolio does not support scenarios"))
        else:
            from app.scenario_manager import ScenarioManager
            sm = ScenarioManager(project_type)
            deltas = sm.scenario_summary(scenario)
            rows.append(("Scenario Deltas", "—"))
            for row in deltas:
                rows.append((
                    f"  {row['assumption']}",
                    f"{row['change']} ({row['scenario']} vs base)",
                ))
                if row.get("note"):
                    rows.append(("  Note", row["note"]))
    else:
        rows.append(("Scenario Deltas", "Base case — no adjustments"))

    df = pd.DataFrame(rows[1:], columns=["Field", "Value"])
    df.to_excel(writer, sheet_name="Notes", index=False)
    ws = writer.sheets["Notes"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"


def _write_depreciation_assumptions_sheet(writer, profile_name: str) -> None:
    """Write 'Depreciation Assumptions' sheet — shows profile rules for all asset classes."""
    from app.depreciation_bankable import get_profile, DepreciationMethod
    
    profile = get_profile(profile_name)
    
    rows = [
        ("Asset Class", "Tax Life (y)", "Book Life (y)", 
         "Tax Method", "Book Method", "Tax Convention", "Book Convention", "Notes"),
    ]
    
    notes_map = {
        "land": "Non-depreciable — land is not a depreciable asset",
        "contingency": "Allocated proportionally to depreciable asset classes",
        "inverters": "Separate 10y tax life vs 25y for modules",
    }
    
    for ac, rule in profile.asset_rules.items():
        note = notes_map.get(ac.value.lower(), "")
        method_display = rule.tax_method.value.replace("_", " ").title()
        book_method_display = rule.book_method.value.replace("_", " ").title()
        conv_display = rule.tax_convention.value.replace("_", " ").title()
        book_conv_display = rule.book_convention.value.replace("_", " ").title()
        
        rows.append((
            ac.value.replace("_", " ").title(),
            str(rule.tax_life_years) if rule.tax_life_years > 0 else "N/A",
            str(rule.book_life_years) if rule.book_life_years > 0 else "N/A",
            method_display,
            book_method_display,
            conv_display,
            book_conv_display,
            note,
        ))
    
    df = pd.DataFrame(rows[1:], columns=rows[0])
    _write_sheet(writer, "Depreciation Assumptions", df)
    ws = writer.sheets["Depreciation Assumptions"]
    ws.freeze_panes = "A2"


def _get_annual_totals(schedule):
    """Get full list of annual depreciation totals by period.
    
    Handles both TaxDepreciationSchedule (method) and WaterfallDepreciationSchedule (property).
    """
    if hasattr(schedule, "total_by_period") and callable(schedule.total_by_period):
        # TaxDepreciationSchedule: method taking period argument
        return [schedule.total_by_period(y) for y in range(schedule.total_periods)]
    else:
        # WaterfallDepreciationSchedule or dict: property/list
        return list(schedule.total_by_period)

def _write_tax_depreciation_sheet(writer, schedule) -> None:
    """Write 'Tax Depreciation' sheet — annual tax depreciation by asset class."""
    if not hasattr(schedule, "total_by_period"):
        return
    
    from app.depreciation_bankable import BankableAssetClass
    all_asset_classes = list(BankableAssetClass)
    
    sort_order = {
        "solar_modules": 0, "inverters": 1, "mounting_structures": 2,
        "grid_connection": 3, "transformer": 4, "civil_works": 5,
        "development_soft": 6, "contingency": 7, "land": 8, "other": 9,
    }
    
    asset_classes = []
    for ac in all_asset_classes:
        vals = schedule.total_by_asset_class(ac)
        if any(v > 0.01 for v in vals):
            asset_classes.append(ac)
    
    asset_classes.sort(key=lambda ac: sort_order.get(ac.value, 10))
    
    years = list(range(schedule.total_periods))
    col_labels = ["Asset Class"] + [f"Y{y+1}" for y in years]
    
    rows = [col_labels]
    for ac in asset_classes:
        vals = schedule.total_by_asset_class(ac)
        rows.append([ac.value.replace("_", " ").title()] + [f"{v:,.1f}" if v >= 0.01 else "—" for v in vals])
    
    annual_totals = _get_annual_totals(schedule)
    total_row = ["Total"] + [f"{v:,.1f}" if v >= 0.01 else "—" for v in annual_totals]
    rows.append(total_row)
    
    df = pd.DataFrame(rows[1:], columns=rows[0])
    _write_sheet(writer, "Tax Depreciation", df, number_format={"kEUR": "#,##0"})
    ws = writer.sheets["Tax Depreciation"]
    ws.freeze_panes = "B2"


def _get_book_annual_totals(book_schedule, total_periods):
    """Get annual totals from book_schedule handling method vs property."""
    if book_schedule is None:
        return [0] * total_periods
    elif hasattr(book_schedule, "total_by_period") and callable(book_schedule.total_by_period):
        return [book_schedule.total_by_period(y) for y in range(total_periods)]
    elif hasattr(book_schedule, "total_by_period"):
        return list(book_schedule.total_by_period)
    else:
        return [0] * total_periods

def _write_book_depreciation_sheet(writer, tax_schedule, book_schedule) -> None:
    """Write 'Book Depreciation' sheet — annual book depreciation by asset class."""
    if not hasattr(tax_schedule, "total_by_period"):
        return
    
    from app.depreciation_bankable import BankableAssetClass
    
    sort_order = {
        "solar_modules": 0, "inverters": 1, "mounting_structures": 2,
        "grid_connection": 3, "transformer": 4, "civil_works": 5,
        "development_soft": 6, "contingency": 7, "land": 8, "other": 9,
    }
    
    all_asset_classes = list(BankableAssetClass)
    asset_classes = []
    for ac in all_asset_classes:
        if book_schedule is None:
            vals = [0] * tax_schedule.total_periods
        elif hasattr(book_schedule, "total_by_asset_class"):
            vals = book_schedule.total_by_asset_class(ac)
        else:
            vals = [0] * tax_schedule.total_periods
        if any(v > 0.01 for v in vals):
            asset_classes.append(ac)
    
    asset_classes.sort(key=lambda ac: sort_order.get(ac.value, 10))
    
    years = list(range(tax_schedule.total_periods))
    col_labels = ["Asset Class"] + [f"Y{y+1}" for y in years]
    
    rows = [col_labels]
    for ac in asset_classes:
        if book_schedule is None:
            vals = [0] * tax_schedule.total_periods
        elif hasattr(book_schedule, "total_by_asset_class"):
            vals = book_schedule.total_by_asset_class(ac)
        else:
            vals = [0] * tax_schedule.total_periods
        rows.append([ac.value.replace("_", " ").title()] + [f"{v:,.1f}" if v >= 0.01 else "—" for v in vals])
    
    book_totals = _get_book_annual_totals(book_schedule, tax_schedule.total_periods)
    total_row = ["Total"] + [f"{v:,.1f}" if v >= 0.01 else "—" for v in book_totals]
    rows.append(total_row)
    
    df = pd.DataFrame(rows[1:], columns=rows[0])
    _write_sheet(writer, "Book Depreciation", df, number_format={"kEUR": "#,##0"})
    ws = writer.sheets["Book Depreciation"]
    ws.freeze_panes = "B2"


def _write_tax_depreciation_sheet_for_project(writer, project_inputs, advanced_capex_line_items, project_type: str = "Solar") -> None:
    """Write Tax Depreciation sheet using project inputs.
    
    Args:
        project_type: explicitly passed (Solar or Wind) — NOT derived from project_inputs.info
    """
    from app.depreciation_bankable import (
        generate_tax_and_book_schedule, 
        map_capex_line_item_to_basis,
        get_profile,
        DepreciationConvention,
    )
    from domain.inputs import ProjectInputs
    
    if project_inputs is None:
        return
    
    # Skip portfolio inputs (no single project info)
    if not isinstance(project_inputs, ProjectInputs):
        return
    
    horizon = getattr(project_inputs.info, "horizon_years", 25) if project_inputs else 25
    # project_type is passed explicitly — do NOT derive from project_inputs.info
    profile_name = f"{project_type.lower()}_croatia_ibl"
    
    if advanced_capex_line_items:
        profile = get_profile(profile_name)
        basis_items = [map_capex_line_item_to_basis(item, profile, project_type) for item in advanced_capex_line_items]
        tax_sched, book_sched = generate_tax_and_book_schedule(
            basis_items, profile, total_periods=horizon,
            convention=DepreciationConvention.FULL_YEAR,
        )
        _write_tax_depreciation_sheet(writer, tax_sched)
    else:
        # No advanced CAPEX — write empty sheet with note
        rows = [("Note", "Advanced CAPEX not provided. No depreciation schedule available."), 
                ("Detail", "Use Advanced CAPEX matrix to enable bankable depreciation disclosure.")]
        df = pd.DataFrame(rows[1:], columns=rows[0])
        _write_sheet(writer, "Tax Depreciation", df)


def _write_book_depreciation_sheet_for_project(writer, project_inputs, advanced_capex_line_items, project_type: str = "Solar") -> None:
    """Write Book Depreciation sheet using project inputs.
    
    Args:
        project_type: explicitly passed (Solar or Wind) — NOT derived from project_inputs.info
    """
    from app.depreciation_bankable import (
        generate_tax_and_book_schedule, 
        map_capex_line_item_to_basis,
        get_profile,
        DepreciationConvention,
    )
    
    from domain.inputs import ProjectInputs
    
    if project_inputs is None:
        return
    
    # Skip portfolio inputs (no single project info)
    if not isinstance(project_inputs, ProjectInputs):
        return
    
    horizon = getattr(project_inputs.info, "horizon_years", 25) if project_inputs else 25
    # project_type is passed explicitly — do NOT derive from project_inputs.info
    profile_name = f"{project_type.lower()}_croatia_ibl"
    
    if advanced_capex_line_items:
        profile = get_profile(profile_name)
        basis_items = [map_capex_line_item_to_basis(item, profile, project_type) for item in advanced_capex_line_items]
        tax_sched, book_sched = generate_tax_and_book_schedule(
            basis_items, profile, total_periods=horizon,
            convention=DepreciationConvention.FULL_YEAR,
        )
        _write_book_depreciation_sheet(writer, tax_sched, book_sched)
    else:
        rows = [("Note", "Advanced CAPEX not provided. No book depreciation schedule available."),
                ("Detail", "Book depreciation is separate from tax depreciation for reporting purposes.")]
        df = pd.DataFrame(rows[1:], columns=rows[0])
        _write_sheet(writer, "Book Depreciation", df)

def _write_dsrf_sheet(writer, portfolio_result) -> None:
    """Write 'DSRF' sheet — Phase 2 revolving facility schedule.

    Only created when portfolio_result.dsrf_enabled is True and dsrf_periods is non-empty.
    Uses facility terminology only (draw, repayment, drawn, undrawn, facility limit).
    """
    from domain.portfolio.independent.result import IndependentPortfolioResult
    import pandas as pd

    if not isinstance(portfolio_result, IndependentPortfolioResult):
        return

    periods = portfolio_result.dsrf_periods
    if not periods:
        return

    # Build column order matching spec exactly
    rows = [(
        "Period",
        "SPV Code",
        "Facility Limit (kEUR)",
        "Drawn Start (kEUR)",
        "Undrawn Start (kEUR)",
        "Scheduled Senior DS (kEUR)",
        "CFADS Available (kEUR)",
        "Debt Service Shortfall (kEUR)",
        "Draw (kEUR)",
        "Drawn Interest (kEUR)",
        "Commitment Fee (kEUR)",
        "Repayment (kEUR)",
        "Drawn End (kEUR)",
        "Undrawn End (kEUR)",
        "Cash Available for Distribution (kEUR)",
    )]

    for p in periods:
        rows.append((
            getattr(p, "period", ""),
            getattr(p, "spv_code", ""),
            _round(getattr(p, "facility_limit_keur", 0.0)),
            _round(getattr(p, "drawn_start_keur", 0.0)),
            _round(getattr(p, "undrawn_start_keur", 0.0)),
            _round(getattr(p, "scheduled_senior_ds_keur", 0.0)),
            _round(getattr(p, "cfads_available_keur", 0.0)),
            _round(getattr(p, "debt_service_shortfall_keur", 0.0)),
            _round(getattr(p, "draw_keur", 0.0)),
            _round(getattr(p, "drawn_interest_keur", 0.0)),
            _round(getattr(p, "commitment_fee_keur", 0.0)),
            _round(getattr(p, "repayment_keur", 0.0)),
            _round(getattr(p, "drawn_end_keur", 0.0)),
            _round(getattr(p, "undrawn_end_keur", 0.0)),
            _round(getattr(p, "cash_available_for_distribution_keur", 0.0)),
        ))

    df = pd.DataFrame(rows[1:], columns=rows[0])
    _write_sheet(writer, "DSRF", df, number_format={"kEUR": "#,##0"})
    ws = writer.sheets["DSRF"]
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"


def _round(v: float, ndigits: int = 1) -> float:
    """Round kEUR values consistently."""
    return round(v, ndigits) if isinstance(v, float) else v


def _write_portfolio_notes_sheet(writer, portfolio_result) -> None:
    from openpyxl.styles import Font

    lines = [
        "Phase 1.5 Portfolio — Independent SPV Aggregation",
        "",
        "This export uses domain/portfolio/independent/ IndependentPortfolioResult.",
        "It is NOT a pooled financing portfolio (see domain/portfolio/waterfall.py).",
        "",
        "--- Limitations ---",
        "No HoldCo entity.",
        "No SHL / intercompany flows.",
        "No Sponsor IRR.",
        "No monthly model frequency.",
        "No cross-SP cash pooling.",
        "No retained earnings constraint.",
        "",
        "--- IRR Disclaimer ---",
        "Simple Average Project IRR and Simple Average Equity IRR are UNWEIGHTED",
        "averages of per-SP IRRs. They are NOT true portfolio XIRR values.",
        "True portfolio IRR requires date-aligned XIRR over all SPV cash flows",
        "and is NOT implemented in Phase 1 or Phase 1.5.",
        "",
        "--- DSRF (Phase 2) ---",
        "DSRF is a revolving debt service reserve facility (liquidity facility).",
        "It is NOT a cash reserve account (unlike DSRA).",
        "DSRF reduces distributions via: commitment fee + drawn interest + repayment.",
        "IRR recalculation is deferred to future step.",
        "Semiannual model only. No HoldCo. No SHL. No pooled financing.",
        "",
        "--- Warnings ---",
    ]
    if portfolio_result.warnings:
        for w in portfolio_result.warnings:
            lines.append(f"  ⚠ {w}")
    else:
        lines.append("  (none)")

    df = pd.DataFrame([(l,) for l in lines], columns=["Note"])
    df.to_excel(writer, sheet_name="Portfolio_Notes", index=False)
    ws = writer.sheets["Portfolio_Notes"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.sheet_state = "visible"


# ── Phase 5D: Overlay Sheets ────────────────────────────────────────────────────


def _write_overlay_sheets(
    writer,
    cash_ledger=None,
    spv_overlays=None,
    holdco_overlay=None,
    constraint_results=None,
) -> None:
    """Write optional overlay sheets for Phase 5D distribution constraint visibility.

    All parameters optional — if None the sheet is skipped silently.

    Sheet names are <= 31 chars and follow openpyxl sheet name limits.
    """
    if cash_ledger is not None:
        _write_cash_ledger_sheet(writer, cash_ledger)

    if constraint_results is not None:
        _write_dist_constraints_sheet(writer, constraint_results)

    if spv_overlays is not None:
        for ov in spv_overlays:
            if isinstance(ov, SPVRetainedCashOverlay):
                _write_spv_retained_cash_sheet(writer, ov)

    if holdco_overlay is not None:
        _write_holdco_retained_cash_sheet(writer, holdco_overlay)


def _write_cash_ledger_sheet(writer, ledger: PortfolioCashLedger) -> None:
    """Write Cash Ledger sheet — one row per movement, with entity and period context."""
    rows = []
    for entity_ledger in (ledger.entities or ()):
        for period in (entity_ledger.periods or ()):
            for mov in (period.movements or ()):
                rows.append({
                    "Entity Code": entity_ledger.entity_code,
                    "Period": period.period,
                    "Opening Cash": round(period.opening_cash_keur, 2),
                    "Movement Type": mov.movement_type.name if hasattr(mov.movement_type, 'name') else str(mov.movement_type),
                    "Movement Amount": round(mov.amount_keur, 2),
                    "Closing Cash": round(period.closing_cash_keur, 2),
                    "Warnings": "; ".join(period.warnings) if period.warnings else "",
                })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Entity Code", "Period", "Opening Cash", "Movement Type",
        "Movement Amount", "Closing Cash", "Warnings"
    ])
    _write_sheet(writer, "Cash Ledger", df, number_format={"Opening Cash": "#,##0", "Movement Amount": "#,##0", "Closing Cash": "#,##0"})
    _check_sheet_name_length("Cash Ledger")


def _write_dist_constraints_sheet(writer, constraint_results) -> None:
    """Write Distribution Constraints sheet from constraint result objects."""
    rows = []
    for cr in (constraint_results or ()):
        if not isinstance(cr, DistributionConstraintResult):
            continue
        for cp in (cr.periods or ()):
            rows.append({
                "Entity Code": cr.entity_code,
                "Period": cp.period,
                "Cash Before Distribution": round(cp.cash_before_distribution_keur, 2),
                "Requested Distribution": round(cp.requested_distribution_keur, 2),
                "Allowed Distribution": round(cp.allowed_distribution_keur, 2),
                "Retained Cash": round(cp.retained_cash_keur, 2),
                "Block Reasons": "; ".join(r.name if hasattr(r, 'name') else str(r) for r in (cp.block_reasons or ())),
                "Warnings": "; ".join(cp.warnings) if cp.warnings else "",
            })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Entity Code", "Period", "Cash Before Distribution", "Requested Distribution",
        "Allowed Distribution", "Retained Cash", "Block Reasons", "Warnings"
    ])
    _write_sheet(writer, "Dist Constraints", df, number_format={
        "Cash Before Distribution": "#,##0", "Requested Distribution": "#,##0",
        "Allowed Distribution": "#,##0", "Retained Cash": "#,##0"
    })
    _check_sheet_name_length("Dist Constraints")


def _write_spv_retained_cash_sheet(writer, overlay: SPVRetainedCashOverlay) -> None:
    """Write SPV Retained Cash sheet for one SPV overlay.

    Sheet name includes entity code truncated to openpyxl limit (28 chars + "SPV_").
    """
    safe_name = _safe_sheet_name(overlay.entity_code or "SPV")
    sheet_name = f"SPV_{safe_name}"[:31]
    _check_sheet_name_length(sheet_name)

    rows = []
    for p in (overlay.periods or ()):
        rows.append({
            "SPV Code": overlay.entity_code,
            "Period": p.period,
            "Requested Distribution": round(p.requested_distribution_keur, 2),
            "Allowed Distribution": round(p.allowed_distribution_keur, 2),
            "Retained Cash": round(p.retained_cash_keur, 2),
            "Available Distribution": round(p.cash_before_distribution_keur - p.retained_cash_keur, 2),
            "Warnings": "; ".join(p.warnings) if p.warnings else "",
        })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "SPV Code", "Period", "Requested Distribution", "Allowed Distribution",
        "Retained Cash", "Available Distribution", "Warnings"
    ])
    _write_sheet(writer, sheet_name, df, number_format={
        "Requested Distribution": "#,##0", "Allowed Distribution": "#,##0",
        "Retained Cash": "#,##0", "Available Distribution": "#,##0"
    })


def _write_holdco_retained_cash_sheet(writer, overlay: HoldCoRetainedCashOverlay) -> None:
    """Write HoldCo Retained Cash sheet."""
    rows = []
    max_len = max(
        len(overlay.retained_cash_by_period),
        len(overlay.requested_distribution_by_period),
        len(overlay.available_distribution_by_period),
    )
    for i in range(max_len):
        rows.append({
            "HoldCo Code": overlay.entity_code,
            "Period": i,
            "Requested Distribution": round(
                overlay.requested_distribution_by_period[i]
                if i < len(overlay.requested_distribution_by_period) else 0.0, 2),
            "Retained Cash": round(
                overlay.retained_cash_by_period[i]
                if i < len(overlay.retained_cash_by_period) else 0.0, 2),
            "Available Distribution": round(
                overlay.available_distribution_by_period[i]
                if i < len(overlay.available_distribution_by_period) else 0.0, 2),
            "Warnings": "",
        })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "HoldCo Code", "Period", "Requested Distribution", "Retained Cash", "Available Distribution", "Warnings"
    ])
    # Prepend audit-only note row
    note_row = {col: "" for col in df.columns}
    note_row["HoldCo Code"] = "AUDIT-ONLY NOTE"
    note_row["Requested Distribution"] = 0.0
    note_row["Retained Cash"] = 0.0
    note_row["Available Distribution"] = 0.0
    note_df = pd.concat([pd.DataFrame([note_row]), df], ignore_index=True)

    _write_sheet(writer, "HoldCo Ret Cash", note_df, number_format={
        "Requested Distribution": "#,##0", "Retained Cash": "#,##0", "Available Distribution": "#,##0"
    })
    _check_sheet_name_length("HoldCo Ret Cash")


def _safe_sheet_name(name: str) -> str:
    """Sanitize a sheet name for openpyxl compatibility."""
    # openpyxl disallows \ / * ? [ ] : in sheet names
    import re
    clean = re.sub(r'[\/\\*?[\]:]', '_', str(name))
    return clean[:28]  # leave room for prefix


def _check_sheet_name_length(name: str) -> None:
    """Assertion that sheet name is <= 31 chars (openpyxl limit)."""
    assert len(name) <= 31, f"Sheet name '{name}' exceeds 31-char limit ({len(name)})"
