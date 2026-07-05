"""V4-3: Sensitivity analysis service.

Pure computation layer — no persistence side effects, no model changes.
Applies configurable shocks to ProjectInputs and re-runs the canonical engine.
"""
from __future__ import annotations

import csv
import io
from dataclasses import replace, fields
from typing import Any, Optional

# Shock type registry (key -> human label, unit)
SHOCK_REGISTRY: dict[str, tuple[str, str]] = {
    "capex": ("CAPEX", "%"),
    "opex": ("OPEX", "%"),
    "ppa_price": ("PPA Price", "%"),
    "merchant_price": ("Merchant Price", "%"),
    "yield": ("Yield (P50 Hours)", "%"),
    "availability": ("Availability", "%"),
    "interest_rate": ("Interest Rate", "bps"),
    "tax_rate": ("Tax Rate", "%"),
}

# Default shock levels (percentage points applied as multipliers for % shocks)
DEFAULT_SHOCK_LEVELS = [-15.0, -10.0, -5.0, 5.0, 10.0, 15.0]

# KPI definitions: (attr_on_WaterfallResult, display_label, format_hint)
KPI_DEFS: list[tuple[str, str, str]] = [
    ("total_revenue_keur", "Revenue (kEUR)", "keur"),
    ("total_ebitda_keur", "EBITDA (kEUR)", "keur"),
    ("_cfads_keur", "CFADS (kEUR)", "keur"),
    ("total_tax_keur", "Corp. Tax (kEUR)", "keur"),
    ("total_senior_ds_keur", "Senior DS (kEUR)", "keur"),
    ("total_distribution_keur", "Equity Distribution (kEUR)", "keur"),
    ("project_irr", "Project IRR", "pct"),
    ("equity_irr", "Equity IRR", "pct"),
    ("equity_npv", "Equity NPV (kEUR)", "keur"),
    ("actual_avg_dscr", "Avg DSCR", "x"),
    ("min_dscr", "Min DSCR", "x"),
    ("min_llcr", "Min LLCR", "x"),
]


def _extract_kpis(result: Any) -> dict[str, Optional[float]]:
    """Extract all KPIs from a WaterfallResult into a flat dict."""
    kpis: dict[str, Optional[float]] = {}
    for attr, label, _ in KPI_DEFS:
        if attr == "_cfads_keur":
            # CFADS = r69_fcf_banks_keur (FCF to banks, the DSCR numerator)
            try:
                kpis[attr] = sum(
                    getattr(p, "r69_fcf_banks_keur", 0.0) or 0.0
                    for p in result.periods
                )
            except Exception:
                kpis[attr] = None
        else:
            v = getattr(result, attr, None)
            kpis[attr] = float(v) if v is not None else None
    return kpis


def _apply_shock(proj: Any, shock_type: str, level_pct: float) -> Any:
    """Return a new ProjectInputs with the given shock applied.

    level_pct is a signed percentage: +10.0 = +10%, -5.0 = -5%.
    For interest_rate shock the level is treated as basis-point change:
    level_pct=10.0 → +10 bps added to base_rate.
    """
    factor = 1.0 + level_pct / 100.0

    if shock_type == "capex":
        from finco_core.inputs._models import CapexItem
        capex = proj.capex
        updates: dict[str, Any] = {}
        for f in fields(capex):
            v = getattr(capex, f.name)
            if isinstance(v, CapexItem):
                updates[f.name] = replace(v, amount_keur=v.amount_keur * factor)
            elif isinstance(v, (int, float)):
                updates[f.name] = v * factor
        return replace(proj, capex=replace(capex, **updates))

    elif shock_type == "opex":
        new_opex = tuple(
            replace(o, y1_amount_keur=o.y1_amount_keur * factor)
            for o in proj.opex
        )
        return replace(proj, opex=new_opex)

    elif shock_type == "ppa_price":
        return replace(
            proj,
            revenue=replace(proj.revenue, ppa_base_tariff=proj.revenue.ppa_base_tariff * factor),
        )

    elif shock_type == "merchant_price":
        orig = proj.revenue.market_prices_curve
        if orig:
            new_curve = tuple(v * factor for v in orig)
            return replace(
                proj,
                revenue=replace(proj.revenue, market_prices_curve=new_curve),
            )
        return proj

    elif shock_type == "yield":
        return replace(
            proj,
            technical=replace(
                proj.technical,
                operating_hours_p50=proj.technical.operating_hours_p50 * factor,
            ),
        )

    elif shock_type == "availability":
        new_avail = min(1.0, max(0.0, proj.technical.plant_availability * factor))
        return replace(
            proj,
            technical=replace(proj.technical, plant_availability=new_avail),
        )

    elif shock_type == "interest_rate":
        # level_pct treated as basis points (e.g. +10.0 → +10 bps)
        delta_rate = level_pct / 10_000.0
        new_base = max(0.0, proj.financing.base_rate + delta_rate)
        return replace(
            proj,
            financing=replace(proj.financing, base_rate=new_base),
        )

    elif shock_type == "tax_rate":
        new_rate = max(0.0, min(1.0, proj.tax.corporate_rate * factor))
        return replace(
            proj,
            tax=replace(proj.tax, corporate_rate=new_rate),
        )

    return proj


def _run_once(proj: Any) -> dict[str, Optional[float]]:
    from app.ui_runner import _build_period_engine
    from app.waterfall_runner import WaterfallRunner, WaterfallRunConfig

    eng = _build_period_engine(proj)
    result = WaterfallRunner(proj, eng).run(WaterfallRunConfig.from_inputs(proj, eng))
    return _extract_kpis(result)


def run_sensitivity(
    proj: Any,
    shock_types: list[str],
    shock_levels: list[float] | None = None,
) -> dict[str, Any]:
    """Run full sensitivity matrix.

    Returns:
      base_kpis: dict[str, float]   — KPIs at base
      rows: list[SensitivityRow]    — one per (shock_type, level) combination
      kpi_defs: list                — ordered KPI metadata
      shock_registry: dict          — human labels for shock types
    """
    if shock_levels is None:
        shock_levels = DEFAULT_SHOCK_LEVELS

    base_kpis = _run_once(proj)

    rows = []
    for stype in shock_types:
        for level in shock_levels:
            try:
                shocked_proj = _apply_shock(proj, stype, level)
                shocked_kpis = _run_once(shocked_proj)
                deltas = {
                    k: (shocked_kpis[k] - base_kpis[k])
                    if shocked_kpis[k] is not None and base_kpis[k] is not None
                    else None
                    for k in base_kpis
                }
            except Exception as exc:
                shocked_kpis = {k: None for k in base_kpis}
                deltas = {k: None for k in base_kpis}
                rows.append({
                    "shock_type": stype,
                    "shock_label": SHOCK_REGISTRY.get(stype, (stype, "%"))[0],
                    "level_pct": level,
                    "kpis": shocked_kpis,
                    "deltas": deltas,
                    "error": str(exc),
                })
                continue
            rows.append({
                "shock_type": stype,
                "shock_label": SHOCK_REGISTRY.get(stype, (stype, "%"))[0],
                "level_pct": level,
                "kpis": shocked_kpis,
                "deltas": deltas,
                "error": None,
            })

    return {
        "base_kpis": base_kpis,
        "rows": rows,
        "kpi_defs": KPI_DEFS,
        "shock_registry": SHOCK_REGISTRY,
    }


def build_tornado_data(
    sensitivity_result: dict[str, Any],
    kpi_key: str = "equity_irr",
    n_top: int = 10,
) -> list[dict[str, Any]]:
    """Compute tornado chart data: range of impact per shock type.

    For each shock type, computes max_delta and min_delta across all levels,
    then ranks by absolute range (|max - min|), largest first.

    Returns a list of dicts sorted descending by impact range.
    """
    base_kpis = sensitivity_result["base_kpis"]
    base_val = base_kpis.get(kpi_key)
    rows = sensitivity_result["rows"]

    impact_by_shock: dict[str, dict[str, Any]] = {}
    for row in rows:
        stype = row["shock_type"]
        delta = row["deltas"].get(kpi_key)
        if delta is None:
            continue
        if stype not in impact_by_shock:
            impact_by_shock[stype] = {
                "label": row["shock_label"],
                "min_delta": delta,
                "max_delta": delta,
                "min_level": row["level_pct"],
                "max_level": row["level_pct"],
            }
        else:
            if delta < impact_by_shock[stype]["min_delta"]:
                impact_by_shock[stype]["min_delta"] = delta
                impact_by_shock[stype]["min_level"] = row["level_pct"]
            if delta > impact_by_shock[stype]["max_delta"]:
                impact_by_shock[stype]["max_delta"] = delta
                impact_by_shock[stype]["max_level"] = row["level_pct"]

    tornado = []
    for stype, d in impact_by_shock.items():
        impact_range = d["max_delta"] - d["min_delta"]
        tornado.append({
            "shock_type": stype,
            "label": d["label"],
            "min_delta": d["min_delta"],
            "max_delta": d["max_delta"],
            "min_level": d["min_level"],
            "max_level": d["max_level"],
            "impact_range": impact_range,
            "abs_range": abs(impact_range),
        })

    tornado.sort(key=lambda x: x["abs_range"], reverse=True)
    return tornado[:n_top]


def export_sensitivity_csv(sensitivity_result: dict[str, Any]) -> str:
    """Export sensitivity table to CSV string."""
    kpi_defs = sensitivity_result["kpi_defs"]
    base_kpis = sensitivity_result["base_kpis"]

    buf = io.StringIO()
    writer = csv.writer(buf)

    kpi_attrs = [k for k, _, _ in kpi_defs]
    kpi_labels = [lbl for _, lbl, _ in kpi_defs]

    # Header
    writer.writerow(["Shock Type", "Level (%)", "Error"] + kpi_labels + [f"Δ {lbl}" for lbl in kpi_labels])

    # Base row
    writer.writerow(
        ["Base", "0", ""]
        + [_fmt_csv(base_kpis.get(k)) for k in kpi_attrs]
        + [""] * len(kpi_attrs)
    )

    for row in sensitivity_result["rows"]:
        writer.writerow(
            [row["shock_label"], row["level_pct"], row["error"] or ""]
            + [_fmt_csv(row["kpis"].get(k)) for k in kpi_attrs]
            + [_fmt_csv(row["deltas"].get(k)) for k in kpi_attrs]
        )

    return buf.getvalue()


def export_sensitivity_xlsx(sensitivity_result: dict[str, Any]) -> bytes:
    """Export sensitivity table to Excel bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensitivity"

    kpi_defs = sensitivity_result["kpi_defs"]
    base_kpis = sensitivity_result["base_kpis"]
    kpi_attrs = [k for k, _, _ in kpi_defs]
    kpi_labels = [lbl for _, lbl, _ in kpi_defs]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    base_fill = PatternFill("solid", fgColor="D9E1F2")
    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")

    # Header row
    headers = ["Shock Type", "Level", "Error"] + kpi_labels + [f"Δ {lbl}" for lbl in kpi_labels]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Base row
    base_row = 2
    ws.cell(row=base_row, column=1, value="Base")
    ws.cell(row=base_row, column=2, value=0)
    for col, k in enumerate(kpi_attrs, 4):
        v = base_kpis.get(k)
        cell = ws.cell(row=base_row, column=col, value=round(v, 6) if v is not None else None)
        cell.fill = base_fill

    # Sensitivity rows
    for r_idx, row in enumerate(sensitivity_result["rows"], base_row + 1):
        ws.cell(row=r_idx, column=1, value=row["shock_label"])
        ws.cell(row=r_idx, column=2, value=row["level_pct"])
        ws.cell(row=r_idx, column=3, value=row["error"] or "")
        # KPI values
        for col, k in enumerate(kpi_attrs, 4):
            v = row["kpis"].get(k)
            ws.cell(row=r_idx, column=col, value=round(v, 6) if v is not None else None)
        # Delta values
        delta_start = 4 + len(kpi_attrs)
        for col, k in enumerate(kpi_attrs, delta_start):
            d = row["deltas"].get(k)
            cell = ws.cell(row=r_idx, column=col, value=round(d, 6) if d is not None else None)
            if d is not None:
                cell.fill = pos_fill if d > 0 else (neg_fill if d < 0 else PatternFill())

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_csv(v: Any) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):.6f}"
    except (TypeError, ValueError):
        return str(v)
