"""Institutional workbook export for Phase 10.

This module binds existing runtime outputs, project context assumptions, and
offline financial statement assembly into a standardized review workbook.
It does not change runtime authority or formulas.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.export_metadata import build_export_metadata, metadata_rows
from app.export.workbook_index import (
    INSTITUTIONAL_SHEET_INVENTORY,
    write_workbook_index_sheet_full,
)
from app.export.runtime_summary import _run_project, build_runtime_summary_rows
from app.input_helpers import (
    build_capex_items_table,
    build_capex_summary_table,
    build_inputs_summary_table,
)
from app.output_tables import build_debt_table, build_revenue_table
from app.ui.project_context import ProjectContext, get_project_context
from domain.financial_statements import assemble_financial_statements


@dataclass(frozen=True)
class WorkbookSheetDefinition:
    sheet_order: int
    sheet_name: str
    implemented_level: str
    runtime_or_preview: str
    governance_sensitive: bool
    notes: str


@dataclass(frozen=True)
class WorkbookExportBundle:
    project_key: str
    project_name: str
    generated_at: str
    branch: str
    commit_sha: str
    runtime_timestamp: str
    active_project: str
    scenario_id: str
    scenario_name: str
    scenario_revision: str
    runtime_snapshot_id: str
    runtime_origin: str
    template_origin: str
    template_revision: str
    export_template_version: str
    runtime_flag_count: str
    runtime_flags_json: str
    governance_posture_summary: str
    replay_limitations: str
    context: ProjectContext
    project_inputs: object
    runtime_result: object
    runtime_rows: list[dict[str, str]]
    statements: object
    inputs_summary: object
    capex_summary: object
    capex_items: object
    revenue_table: object
    debt_table: object


def _resolve_export_senior_debt_keur(bundle: WorkbookExportBundle) -> float:
    """Return the senior debt value to display in the export.

    Phase S1-A (low-risk, presentation-only):

    - For TUHO and Oborovo, ``bundle.context.senior_debt_keur`` is the
      frozen Excel-derived senior debt amount (input field). The
      runtime uses this value as the ``fixed_debt_keur`` override, so
      the runtime result equals the input. We return the input value
      directly to preserve bit-identical parity with the prior
      export.

    - For Generic projects, ``bundle.context.senior_debt_keur`` is 0.0
      because the Generic factory does not set
      ``financing.fixed_debt_keur``. The runtime computes the senior
      debt amount from the DSCR-sculpt formula (with optional gearing
      cap). We return ``runtime_result.sculpting_result.debt_keur``
      in this case so the export shows the actual runtime value
      instead of 0.

    The function never invents a value: it returns either the input
    (when non-zero) or the runtime result. For TUHO and Oborovo the
    two are bit-identical, so this preserves the parity contract.
    For Generic, the runtime result is the authoritative value.
    """
    ctx_debt = bundle.context.senior_debt_keur or 0.0
    if ctx_debt > 0.0:
        # TUHO, Oborovo, or any future project with explicit
        # fixed_debt_keur. Input is authoritative (runtime
        # result is bit-identical to input because fixed_debt_keur
        # overrides the runtime calculation).
        return float(ctx_debt)
    # Generic (or any project without fixed_debt_keur): fall
    # back to the runtime result. sculpting_result is always
    # present when the runtime completes successfully.
    rt = bundle.runtime_result
    sculpting = getattr(rt, "sculpting_result", None) if rt is not None else None
    if sculpting is not None and getattr(sculpting, "debt_keur", None) is not None:
        return float(sculpting.debt_keur)
    return 0.0


INSTITUTIONAL_SHEET_DEFINITIONS = (
    WorkbookSheetDefinition(0, "Export_Metadata", "implemented", "review", True, "Export provenance, trust hygiene, and non-claims. Added Phase 47."),
    WorkbookSheetDefinition(1, "Workbook_Index", "implemented", "review", True, "Sheet inventory and workbook guide. Added Phase 48."),
    WorkbookSheetDefinition(2, "Cover", "implemented", "review", True, "Institutional cover sheet with provenance."),
    WorkbookSheetDefinition(2, "Governance", "implemented", "review", True, "Governance metadata and approval status."),
    WorkbookSheetDefinition(3, "Runtime Summary", "implemented", "runtime", True, "Existing runtime summary values only."),
    WorkbookSheetDefinition(4, "Inputs", "runtime_bound", "runtime + template", True, "Project metadata and assumption binding from existing context."),
    WorkbookSheetDefinition(5, "Construction", "summary_bound", "template + runtime", True, "Construction timing and funding summary from existing inputs and runtime anchors."),
    WorkbookSheetDefinition(6, "OPEX", "runtime_bound", "runtime + template", True, "OPEX assumptions plus runtime total binding."),
    WorkbookSheetDefinition(7, "CAPEX", "summary_bound", "template + runtime", True, "CAPEX summary, funding split anchors, and capex item binding."),
    WorkbookSheetDefinition(8, "Revenue", "runtime_bound", "runtime + template", True, "Revenue assumptions and period revenue table from existing runtime outputs."),
    WorkbookSheetDefinition(9, "Senior Debt", "runtime_bound", "runtime + template", True, "Senior debt assumptions and period debt service schedule from existing runtime outputs."),
    WorkbookSheetDefinition(10, "SHL", "runtime_bound", "runtime + template", True, "SHL assumptions and runtime SHL cashflow / balance schedule."),
    WorkbookSheetDefinition(11, "Tax", "runtime_bound", "runtime + template", True, "Tax assumptions and tax bridge periods from existing assembled statements."),
    WorkbookSheetDefinition(12, "P&L", "runtime_bound", "runtime", True, "Offline assembled P&L using existing runtime result as source."),
    WorkbookSheetDefinition(13, "Cash Flow", "runtime_bound", "runtime", True, "Offline PF cash waterfall using existing runtime result as source."),
    WorkbookSheetDefinition(14, "Balance Sheet", "runtime_bound", "runtime", True, "Offline balance sheet assembly from existing runtime result."),
    WorkbookSheetDefinition(15, "Audit", "runtime_bound", "review", True, "Runtime source notes, provenance, and audit boundary statements."),
    WorkbookSheetDefinition(16, "Gap Register", "runtime_bound", "review", True, "Known gaps and accepted conventions reused from existing documentation."),
)


THIN_BORDER = Border(
    left=Side(style="thin", color="D5D9E2"),
    right=Side(style="thin", color="D5D9E2"),
    top=Side(style="thin", color="D5D9E2"),
    bottom=Side(style="thin", color="D5D9E2"),
)
TITLE_FILL = PatternFill("solid", fgColor="0F274A")
SECTION_FILL = PatternFill("solid", fgColor="DCE6F2")
META_FILL = PatternFill("solid", fgColor="EEF3F8")
RUNTIME_FILL = PatternFill("solid", fgColor="E7F3E8")
TEMPLATE_FILL = PatternFill("solid", fgColor="FFF4D6")
REVIEW_FILL = PatternFill("solid", fgColor="F2F4F8")
STATUS_BLOCKED_FILL = PatternFill("solid", fgColor="FCE4E4")
STATUS_NOT_APPROVED_FILL = PatternFill("solid", fgColor="FDEBD0")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
HEADER_FONT = Font(bold=True, size=11, name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
K_EUR_FORMAT = '#,##0.0;[Red](#,##0.0);"-"'
RATIO_FORMAT = "0.00%"
MULTIPLE_FORMAT = "0.000x"


RUNTIME_SUMMARY_LABELS = {
    "active_project": ("Active project", ""),
    "project_irr": ("Project IRR", RATIO_FORMAT),
    "equity_irr": ("Equity IRR", RATIO_FORMAT),
    "total_revenue_keur": ("Total revenue", K_EUR_FORMAT),
    "total_ebitda_keur": ("Total EBITDA", K_EUR_FORMAT),
    "total_opex_keur": ("Total OPEX", K_EUR_FORMAT),
    "avg_dscr": ("Average DSCR", MULTIPLE_FORMAT),
    "min_dscr": ("Minimum DSCR", MULTIPLE_FORMAT),
    "total_distributions_keur": ("Total distributions", K_EUR_FORMAT),
    "total_shl_service_keur": ("Total SHL service", K_EUR_FORMAT),
    "g20_status": ("G20 status", ""),
    "r99_r102_status": ("R99/R102 status", ""),
}

REVIEWER_COVER_NOTES = (
    "Runtime/backend output remains the financial source of truth; this workbook is descriptive and reviewer-facing only.",
    "This export reflects a saved/exported snapshot context. Later draft edits do not mutate the artifact already exported.",
    "If the active draft changed after the last clean run, treat the runtime snapshot as stale and rerun only after saving.",
    "Accepted conventions explain governed residuals or timing choices; they do not imply approval or runtime promotion.",
    "Unavailable/not_applicable markers are intentional metadata markers and must not be read as zero values.",
    "G20 remains BLOCKED and R99/R102 remain NOT APPROVED in this export branch.",
    "Not lender-ready without external model review, not audit-certified, not multi-user governance-ready, and not a replay engine.",
)


def export_institutional_workbook_skeleton(project: str) -> bytes:
    bundle = _build_export_bundle(project)
    workbook = Workbook()

    # Export_Metadata sheet — first sheet, trust hygiene, non-claims
    export_meta = workbook.active
    export_meta.title = "Export_Metadata"
    _write_export_metadata_sheet(export_meta, bundle)

    # Workbook_Index sheet — second sheet, sheet inventory and guide (Phase 48)
    index_sheet = workbook.create_sheet("Workbook_Index")
    write_workbook_index_sheet_full(
        index_sheet,
        workbook_name="Institutional Runtime Workbook",
        project_name=bundle.project_name,
        export_type="institutional_workbook",
        inventory=INSTITUTIONAL_SHEET_INVENTORY,
        is_generic=bundle.active_project not in {"tuho", "oborovo"},
    )

    cover = workbook.create_sheet("Cover")
    _write_cover_sheet(cover, bundle)

    _write_governance_sheet(workbook.create_sheet("Governance"), bundle)
    _write_runtime_summary_sheet(workbook.create_sheet("Runtime Summary"), bundle)
    _write_inputs_sheet(workbook.create_sheet("Inputs"), bundle)
    _write_construction_sheet(workbook.create_sheet("Construction"), bundle)
    _write_opex_sheet(workbook.create_sheet("OPEX"), bundle)
    _write_capex_sheet(workbook.create_sheet("CAPEX"), bundle)
    _write_revenue_sheet(workbook.create_sheet("Revenue"), bundle)
    _write_senior_debt_sheet(workbook.create_sheet("Senior Debt"), bundle)
    _write_shl_sheet(workbook.create_sheet("SHL"), bundle)
    _write_tax_sheet(workbook.create_sheet("Tax"), bundle)
    _write_pnl_sheet(workbook.create_sheet("P&L"), bundle)
    _write_cash_flow_sheet(workbook.create_sheet("Cash Flow"), bundle)
    _write_balance_sheet(workbook.create_sheet("Balance Sheet"), bundle)
    _write_audit_sheet(workbook.create_sheet("Audit"), bundle)
    _write_gap_register_sheet(workbook.create_sheet("Gap Register"), bundle)

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_institutional_workbook_skeleton(
    project: str,
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(export_institutional_workbook_skeleton(project))
    return path


def write_workbook_sheet_map_csv(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sheet_order",
                "sheet_name",
                "implemented_level",
                "runtime_or_preview",
                "governance_sensitive",
                "notes",
            ],
        )
        writer.writeheader()
        for definition in INSTITUTIONAL_SHEET_DEFINITIONS:
            writer.writerow(
                {
                    "sheet_order": definition.sheet_order,
                    "sheet_name": definition.sheet_name,
                    "implemented_level": definition.implemented_level,
                    "runtime_or_preview": definition.runtime_or_preview,
                    "governance_sensitive": "yes" if definition.governance_sensitive else "no",
                    "notes": definition.notes,
                }
            )
    return output_path


def write_runtime_workbook_binding_status_csv(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            "sheet_name": definition.sheet_name,
            "runtime_binding_level": definition.implemented_level,
            "runtime_or_preview": definition.runtime_or_preview,
            "implemented_fields": _binding_fields_for_sheet(definition.sheet_name),
            "remaining_placeholders": _remaining_placeholders_for_sheet(definition.sheet_name),
            "governance_sensitive": "yes" if definition.governance_sensitive else "no",
            "notes": definition.notes,
        }
        for definition in INSTITUTIONAL_SHEET_DEFINITIONS
    ]
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sheet_name",
                "runtime_binding_level",
                "runtime_or_preview",
                "implemented_fields",
                "remaining_placeholders",
                "governance_sensitive",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def _build_export_bundle(project: str) -> WorkbookExportBundle:
    project_key = (project or "tuho").strip().lower()
    runtime_rows = build_runtime_summary_rows(project_key)
    project_inputs, runtime_result = _run_project(project_key)
    statements = assemble_financial_statements(runtime_result)
    context = get_project_context(project_key)
    return WorkbookExportBundle(
        project_key=project_key,
        project_name=runtime_rows[0]["project"],
        generated_at=runtime_rows[0]["export_generated_at"],
        branch=runtime_rows[0]["branch_name"],
        commit_sha=runtime_rows[0]["commit_sha"],
        runtime_timestamp=runtime_rows[0]["runtime_timestamp"],
        active_project=runtime_rows[0]["active_project"],
        scenario_id=runtime_rows[0]["scenario_id"],
        scenario_name=runtime_rows[0]["scenario_name"],
        scenario_revision=runtime_rows[0]["scenario_revision"],
        runtime_snapshot_id=runtime_rows[0]["runtime_snapshot_id"],
        runtime_origin=runtime_rows[0]["runtime_origin"],
        template_origin=runtime_rows[0]["template_origin"],
        template_revision=runtime_rows[0]["template_revision"],
        export_template_version=runtime_rows[0]["export_template_version"],
        runtime_flag_count=runtime_rows[0]["runtime_flag_count"],
        runtime_flags_json=runtime_rows[0]["runtime_flags_json"],
        governance_posture_summary=runtime_rows[0]["governance_posture_summary"],
        replay_limitations=runtime_rows[0]["replay_limitations"],
        context=context,
        project_inputs=project_inputs,
        runtime_result=runtime_result,
        runtime_rows=runtime_rows,
        statements=statements,
        inputs_summary=build_inputs_summary_table(project_inputs),
        capex_summary=build_capex_summary_table(project_inputs),
        capex_items=build_capex_items_table(project_inputs),
        revenue_table=build_revenue_table(runtime_result),
        debt_table=build_debt_table(runtime_result),
    )


def _source_branch() -> str:
    return os.getenv("GIT_BRANCH") or os.getenv("BRANCH_NAME") or "unknown"


def _write_cover_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    sheet["A1"] = "Institutional Runtime Workbook"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18, name="Calibri")
    sheet["A2"] = bundle.project_name
    sheet["A2"].font = Font(size=14, bold=True, name="Calibri")
    sheet["A3"] = "Institutional lender, IC, and audit workbook built from existing runtime outputs and documented assumptions."
    sheet["A3"].alignment = Alignment(wrap_text=True)
    rows = [
        ("Project", bundle.project_name),
        ("Workbook type", "institutional_workbook_runtime_binding"),
        ("Export generated at", bundle.generated_at),
        ("Runtime timestamp", bundle.runtime_timestamp),
        ("Source branch", bundle.branch or _source_branch()),
        ("Commit SHA", bundle.commit_sha),
        ("Active project", bundle.active_project),
        ("Scenario ID", bundle.scenario_id),
        ("Scenario name", bundle.scenario_name),
        ("Scenario revision", bundle.scenario_revision),
        ("Runtime snapshot ID", bundle.runtime_snapshot_id),
        ("Runtime origin", bundle.runtime_origin),
        ("Template origin", bundle.template_origin),
        ("Template revision", bundle.template_revision),
        ("Export template version", bundle.export_template_version),
        ("Workbook version", "Phase 11 institutional export polish"),
        ("Runtime / preview", "runtime-bound workbook with explicit assumption sections"),
        ("Governance status", bundle.governance_posture_summary),
        ("Runtime flag snapshot", f"{bundle.runtime_flag_count} flags captured for replay provenance"),
        ("Purpose", "Institutional review workbook fed from existing runtime outputs and documented assumptions."),
        ("Status", "Phase 11 product polish layer - not final bankability"),
    ]
    for row_idx, (label, value) in enumerate(rows, start=5):
        sheet.cell(row=row_idx, column=1, value=label)
        sheet.cell(row=row_idx, column=2, value=value)
        sheet.cell(row=row_idx, column=1).fill = META_FILL
        sheet.cell(row=row_idx, column=1).font = HEADER_FONT
        sheet.cell(row=row_idx, column=1).border = THIN_BORDER
        sheet.cell(row=row_idx, column=2).border = THIN_BORDER
        sheet.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
    sheet["D5"] = "Reviewer readiness"
    sheet["D5"].font = HEADER_FONT
    for row_idx, note in enumerate(REVIEWER_COVER_NOTES, start=6):
        sheet.cell(row=row_idx, column=4, value=note)
    sheet["D13"] = bundle.replay_limitations
    for row_idx in range(6, 14):
        sheet.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 44
    sheet.column_dimensions["D"].width = 46


def _write_governance_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "Governance review")
    rows = [
        ("Governance status", bundle.runtime_rows[0]["governance_status"], "review", "Export/workbook branch only."),
        ("G20 status", bundle.runtime_rows[0]["g20_status"], "review", "Remains blocked in this branch."),
        ("R99/R102 status", bundle.runtime_rows[0]["r99_r102_status"], "review", "Runtime promotion remains not approved."),
        ("Governance posture summary", bundle.governance_posture_summary, "review", "Descriptive governance posture only; no approval implied."),
        ("Commit SHA", bundle.commit_sha, "review", "Code version that generated this workbook."),
        ("Branch name", bundle.branch, "review", "Branch provenance for the generating code path."),
        ("Export generated at", bundle.generated_at, "review", "Timestamp when this workbook artifact was generated."),
        ("Runtime timestamp", bundle.runtime_timestamp, "review", "Timestamp for the runtime execution behind this export."),
        ("Active project", bundle.active_project, "review", "Canonical active project label for this export path."),
        ("Scenario ID", bundle.scenario_id, "review", "Explicitly marked not_applicable when no saved scenario boundary exists."),
        ("Scenario name", bundle.scenario_name, "review", "Explicitly marked not_applicable when no saved scenario boundary exists."),
        ("Scenario revision", bundle.scenario_revision, "review", "Saved snapshot or scenario boundary marker where available."),
        ("Runtime snapshot ID", bundle.runtime_snapshot_id, "review", "Explicitly marked unavailable when this export path has no persisted runtime snapshot record."),
        ("Runtime origin", bundle.runtime_origin, "review", "Descriptive runtime boundary label only."),
        ("Template origin", bundle.template_origin, "template assumption", "Factory/template source for the project context."),
        ("Template revision", bundle.template_revision, "template assumption", "Lightweight template provenance; no separate semantic version exists."),
        ("Export template version", bundle.export_template_version, "review", "Export packaging/version metadata only."),
        ("Runtime flags captured", bundle.runtime_flag_count, "review", "Replay provenance only; flags remain non-authoritative metadata."),
        ("Project code", bundle.context.code, "template assumption", "Read-only project context."),
        ("Scenario label", "Factory-bound base runtime", "runtime", "No scenario overrides introduced here."),
        ("Data provenance", "Project context + runtime result + offline financial statements", "review", "No workbook-only formulas."),
        ("Replay limitations", bundle.replay_limitations, "review", "Traceability improved without implying a full replay engine."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Governance and provenance", rows)
    reviewer_rows = [
        (
            "Runtime authority",
            "Runtime/backend output remains the financial source of truth.",
            "review",
            "This workbook is descriptive only and is not the calculation engine.",
        ),
        (
            "Snapshot vs live draft",
            "The workbook reflects the exported runtime/saved snapshot context.",
            "review",
            "Later draft edits do not mutate the exported artifact already in the reviewer's hands.",
        ),
        (
            "Stale runtime rule",
            "If current draft differs from the last clean run, rerun only after saving.",
            "review",
            "A stale runtime snapshot is a workflow signal, not a silent recalculation.",
        ),
        (
            "Governed residuals",
            "Documented residuals and accepted conventions are explanatory, not hidden parity fabrication.",
            "review",
            "Pending or unavailable metrics should be read as pending/not available, never as zero.",
        ),
        (
            "No-claims",
            "Not lender-ready without external model review; not audit-certified; not a replay engine.",
            "review",
            "Governance blockers remain explicit: G20 BLOCKED and R99/R102 NOT APPROVED.",
        ),
    ]
    _write_key_value_section(sheet, next_row, "Reviewer interpretation notes", reviewer_rows)


def _write_runtime_summary_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "Runtime summary")
    rows = []
    for runtime_row in bundle.runtime_rows:
        metric = runtime_row["metric"]
        if metric not in RUNTIME_SUMMARY_LABELS:
            continue
        label, number_format = RUNTIME_SUMMARY_LABELS[metric]
        rows.append((label, _coerce_value(runtime_row["value"], number_format), runtime_row["runtime_or_preview"], runtime_row["notes"], number_format))
    _write_key_value_section(sheet, 6, "Runtime summary values", rows, include_format=True)


def _write_inputs_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    rows = [
        ("Project name", bundle.context.name, "template assumption", "Factory-bound project context."),
        ("Project company", bundle.context.company, "template assumption", "Factory-bound project context."),
        ("Project code", bundle.context.code, "template assumption", "Factory-bound project context."),
        ("Technology", bundle.context.technology, "template assumption", "Read-only project context."),
        ("Country", bundle.context.country_iso, "template assumption", "Read-only project context."),
        ("Capacity MW", bundle.context.capacity_mw, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("COD", bundle.context.cod_date, "template assumption", "Read-only project context."),
        ("Financial close", bundle.context.financial_close, "template assumption", "Read-only project context."),
        ("Construction months", bundle.context.construction_months, "template assumption", "Read-only project context."),
        ("Horizon years", bundle.context.horizon_years, "template assumption", "Read-only project context."),
        ("Period frequency", bundle.context.period_frequency, "template assumption", "Read-only project context."),
        ("Runtime project IRR", bundle.runtime_result.project_irr, "runtime", "Existing runtime output.", RATIO_FORMAT),
        ("Runtime equity IRR", bundle.runtime_result.equity_irr, "runtime", "Existing runtime output.", RATIO_FORMAT),
        ("G20 status", bundle.context.g20_status, "review", "Governance label."),
        ("R99/R102 status", bundle.context.r99_r102_status, "review", "Governance label."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Project metadata", rows, include_format=True)

    inputs_rows = [
        (record["Field"], record["Value"], "template assumption", "Existing input summary helper.")
        for record in bundle.inputs_summary.to_dict(orient="records")
    ]
    _write_key_value_section(sheet, next_row, "Existing input summary table", inputs_rows)


def _write_construction_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "template + runtime")
    financing = bundle.project_inputs.financing
    rows = [
        ("Construction start proxy", bundle.context.financial_close, "template assumption", "Financial close used as workbook anchor."),
        ("COD", bundle.context.cod_date, "template assumption", "Project context COD."),
        ("Construction duration months", bundle.context.construction_months, "template assumption", "Project context."),
        ("Total CAPEX", bundle.context.total_capex_keur, "template assumption", "Factory-bound CAPEX total.", K_EUR_FORMAT),
        ("IDC", bundle.context.idc_keur, "template assumption", "Factory-bound financing cost assumption.", K_EUR_FORMAT),
        ("Bank fees", bundle.context.bank_fees_keur, "template assumption", "Factory-bound financing cost assumption.", K_EUR_FORMAT),
        ("Senior debt anchor", _resolve_export_senior_debt_keur(bundle), "template assumption + runtime", "For TUHO/Oborovo: frozen Excel-derived anchor. For Generic: runtime DSCR-sculpted value.", K_EUR_FORMAT),
        ("SHL anchor", bundle.context.shl_amount_keur + bundle.context.shl_idc_keur, "template assumption", "Opening SHL assumption including IDC.", K_EUR_FORMAT),
        ("Share capital", getattr(financing, "share_capital_keur", 0.0), "template assumption", "Existing financing input.", K_EUR_FORMAT),
        ("Share premium", getattr(financing, "share_premium_keur", 0.0), "template assumption", "Existing financing input.", K_EUR_FORMAT),
        ("Runtime status note", "Detailed construction schedule not exported in this branch", "preview", "Workbook uses summary binding only."),
    ]
    _write_key_value_section(sheet, 6, "Construction summary", rows, include_format=True)


def _write_opex_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    rows = [
        ("Runtime total OPEX", bundle.runtime_result.total_opex_keur, "runtime", "Existing runtime output.", K_EUR_FORMAT),
        ("Factory Y1 total OPEX", bundle.context.opex_y1_total_keur, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("Contingency method", bundle.context.opex_contingency_method, "template assumption", "Read-only project context."),
        ("Contingency percent", bundle.context.opex_contingency_pct / 100.0, "template assumption", "Read-only project context.", RATIO_FORMAT),
        ("Runtime/preview boundary", "Line items below are factory assumptions; runtime total above is authoritative", "review", "No workbook-only OPEX calculations."),
    ]
    next_row = _write_key_value_section(sheet, 6, "OPEX summary", rows, include_format=True)
    item_rows = [
        (item["name"], item["y1_keur"], item["inflation_pct"], "template assumption", "Factory OPEX item")
        for item in bundle.context.opex_items
    ]
    _write_simple_table(
        sheet,
        next_row,
        "OPEX line items",
        ("Line item", "Y1 kEUR", "Inflation", "Source", "Notes"),
        item_rows,
        format_columns={2: K_EUR_FORMAT, 3: RATIO_FORMAT},
    )


def _write_capex_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "template + runtime")
    financing = bundle.project_inputs.financing
    total_capex = bundle.context.total_capex_keur or 0.0
    senior = _resolve_export_senior_debt_keur(bundle)
    shl = (bundle.context.shl_amount_keur or 0.0) + (bundle.context.shl_idc_keur or 0.0)
    share_capital = getattr(financing, "share_capital_keur", None) or 0.0
    share_premium = getattr(financing, "share_premium_keur", None) or 0.0
    implied_gap = total_capex - senior - shl - share_capital - share_premium
    rows = [
        ("Total CAPEX", total_capex, "template assumption", "Project context total CAPEX.", K_EUR_FORMAT),
        ("Senior debt funding", senior, "template assumption", "Financing anchor from project context.", K_EUR_FORMAT),
        ("SHL funding incl. IDC", shl, "template assumption", "Opening SHL assumption.", K_EUR_FORMAT),
        ("Share capital", share_capital, "template assumption", "Existing financing input.", K_EUR_FORMAT),
        ("Share premium", share_premium, "template assumption", "Existing financing input.", K_EUR_FORMAT),
        ("Implied funding remainder", implied_gap, "review", "Residual shown for workbook transparency only.", K_EUR_FORMAT),
    ]
    next_row = _write_key_value_section(sheet, 6, "CAPEX and funding summary", rows, include_format=True)

    capex_rows = [
        (record["Field"], record["Value"], "template assumption", "Existing capex summary helper.")
        for record in bundle.capex_summary.to_dict(orient="records")
    ]
    next_row = _write_key_value_section(sheet, next_row, "Existing capex summary table", capex_rows)

    item_rows = [
        (
            record["Name"],
            record["Asset Class"],
            record["Amount (kEUR)"],
            record["Life (years)"],
            "template assumption",
            "Existing capex item helper.",
        )
        for record in bundle.capex_items.to_dict(orient="records")
    ]
    _write_simple_table(
        sheet,
        next_row,
        "CAPEX items",
        ("Line item", "Asset class", "Amount kEUR", "Life years", "Source", "Notes"),
        item_rows,
        format_columns={3: K_EUR_FORMAT},
    )


def _write_revenue_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    revenue = bundle.project_inputs.revenue
    rows = [
        ("P50 operating hours", bundle.context.operating_hours_p50, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("Plant availability", bundle.context.plant_availability, "template assumption", "Read-only project context.", RATIO_FORMAT),
        ("Grid availability", bundle.context.grid_availability, "template assumption", "Read-only project context.", RATIO_FORMAT),
        ("PPA tariff EUR/MWh", bundle.context.ppa_tariff_eur_mwh, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("PPA term years", bundle.context.ppa_term_years, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("PPA index", bundle.context.ppa_index_pct, "template assumption", "Read-only project context.", RATIO_FORMAT),
        ("CO2 enabled", "yes" if bundle.context.co2_enabled else "no", "template assumption", "Read-only project context."),
        ("Runtime total revenue", bundle.runtime_result.total_revenue_keur, "runtime", "Existing runtime output.", K_EUR_FORMAT),
        ("Merchant start operating period", getattr(revenue, "first_merchant_operating_period_index", "NOT_AVAILABLE"), "template assumption", "Existing revenue input where available."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Revenue assumptions and totals", rows, include_format=True)
    _write_dataframe_section(sheet, next_row, "Revenue period table", bundle.revenue_table, "runtime", "Existing revenue table builder.")


def _write_senior_debt_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    financing = bundle.project_inputs.financing
    rows = [
        ("Senior debt amount", _resolve_export_senior_debt_keur(bundle), "template assumption + runtime", "For TUHO/Oborovo: frozen Excel-derived anchor. For Generic: runtime DSCR-sculpted value.", K_EUR_FORMAT),
        ("Senior tenor years", bundle.context.senior_tenor_years, "template assumption", "Read-only project context.", K_EUR_FORMAT),
        ("Interest assumption", bundle.context.interest_rate_pct, "template assumption", "Base rate plus margin from project context.", RATIO_FORMAT),
        ("Target DSCR", bundle.context.target_dscr, "template assumption", "Read-only project context.", MULTIPLE_FORMAT),
        ("Runtime avg DSCR", bundle.runtime_result.actual_avg_dscr, "runtime", "Existing runtime output.", MULTIPLE_FORMAT),
        ("Runtime min DSCR", bundle.runtime_result.actual_min_dscr, "runtime", "Existing runtime output.", MULTIPLE_FORMAT),
        ("Runtime total senior debt service", bundle.runtime_result.total_senior_ds_keur, "runtime", "Existing runtime output.", K_EUR_FORMAT),
        ("Debt sizing method", getattr(financing, "debt_sizing_method", "NOT_AVAILABLE"), "template assumption", "Read-only financing input."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Senior debt summary", rows, include_format=True)
    _write_dataframe_section(sheet, next_row, "Senior debt period table", bundle.debt_table, "runtime", "Existing debt table builder.")


def _write_shl_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    financing = bundle.project_inputs.financing
    periods = list(bundle.runtime_result.periods)
    opening_balances = [bundle.context.shl_amount_keur + bundle.context.shl_idc_keur]
    opening_balances.extend(float(periods[idx - 1].shl_balance_keur) for idx in range(1, len(periods)))
    rows = [
        ("Opening SHL incl. IDC", bundle.context.shl_amount_keur + bundle.context.shl_idc_keur, "template assumption", "Project context anchor.", K_EUR_FORMAT),
        ("SHL nominal amount", bundle.context.shl_amount_keur, "template assumption", "Project context anchor.", K_EUR_FORMAT),
        ("SHL rate", bundle.context.shl_rate_pct, "template assumption", "Project context anchor.", RATIO_FORMAT),
        ("SHL repayment method", getattr(financing, "shl_repayment_method", "NOT_AVAILABLE"), "template assumption", "Existing financing input."),
        ("Total SHL cash interest", sum(period.shl_interest_keur for period in periods), "runtime", "Existing runtime period field.", K_EUR_FORMAT),
        ("Total SHL PIK", sum(period.shl_pik_keur for period in periods), "runtime", "Existing runtime period field.", K_EUR_FORMAT),
        ("Total SHL principal", sum(period.shl_principal_keur for period in periods), "runtime", "Existing runtime period field.", K_EUR_FORMAT),
    ]
    next_row = _write_key_value_section(sheet, 6, "SHL summary", rows, include_format=True)
    _write_horizontal_metric_table(
        sheet,
        next_row,
        "SHL period schedule",
        [_period_label(period) for period in periods],
        [
            ("SHL Opening Balance", opening_balances, "runtime + template", "Opening balance uses prior closing after period 1."),
            ("SHL Cash Interest", [period.shl_interest_keur for period in periods], "runtime", "Existing runtime period field."),
            ("SHL PIK Capitalized", [period.shl_pik_keur for period in periods], "runtime", "Existing runtime period field."),
            ("SHL Principal Repaid", [period.shl_principal_keur for period in periods], "runtime", "Existing runtime period field."),
            ("SHL Closing Balance", [period.shl_balance_keur for period in periods], "runtime", "Existing runtime period field."),
        ],
        number_format=K_EUR_FORMAT,
    )


def _write_tax_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime + template assumptions")
    tax = bundle.project_inputs.tax
    periods = list(bundle.statements.tax_bridge.periods)
    rows = [
        ("Corporate income tax rate", bundle.context.cit_rate_pct, "template assumption", "Project context tax assumption.", RATIO_FORMAT),
        ("Loss carryforward years", bundle.context.loss_carryforward_years, "template assumption", "Project context tax assumption.", K_EUR_FORMAT),
        ("ATAD EBITDA limit", getattr(tax, "atad_ebitda_limit", "NOT_AVAILABLE"), "template assumption", "Existing tax input.", RATIO_FORMAT),
        ("ATAD minimum interest kEUR", getattr(tax, "atad_min_interest_keur", "NOT_AVAILABLE"), "template assumption", "Existing tax input.", K_EUR_FORMAT),
        ("Runtime total cash tax", sum(period.cash_tax_current_period_keur for period in periods), "runtime", "Existing assembled tax bridge.", K_EUR_FORMAT),
        ("Runtime total CIT accrual", sum(period.cit_accrual_keur for period in periods), "runtime", "Existing assembled tax bridge.", K_EUR_FORMAT),
        ("Known limitation", "R99/R102 remains governance-only, not accepted as runtime source", "review", "No runtime authority change in this branch."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Tax summary", rows, include_format=True)
    _write_horizontal_metric_table(
        sheet,
        next_row,
        "Tax bridge periods",
        [period.date.isoformat() for period in periods],
        [
            ("Tax depreciation", [period.tax_depreciation_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Fiscal reintegration", [period.fiscal_reintegration_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Taxable income before losses", [period.taxable_income_before_losses_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Opening tax losses", [period.tax_loss_opening_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Losses used", [period.tax_loss_used_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Closing tax losses", [period.tax_loss_closing_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Taxable profit after losses", [period.taxable_profit_after_losses_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("CIT accrual", [period.cit_accrual_keur for period in periods], "runtime", "Existing assembled tax bridge."),
            ("Cash tax current period", [period.cash_tax_current_period_keur for period in periods], "runtime", "Existing assembled tax bridge."),
        ],
        number_format=K_EUR_FORMAT,
    )


def _write_pnl_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime")
    periods = list(bundle.statements.pnl.periods)
    _write_horizontal_metric_table(
        sheet,
        6,
        "P&L period table",
        [period.date.isoformat() for period in periods],
        [
            ("Revenue", [period.revenues_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("OPEX", [period.operating_expenses_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Depreciation", [period.depreciation_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("EBIT", [period.ebit_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Senior Interest", [period.senior_interest_expense_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("SHL Interest", [period.shl_interest_expense_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("EBT", [period.earnings_before_tax_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Fiscal Reintegration", [period.fiscal_reintegration_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Taxable Income", [period.taxable_income_before_losses_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Net Income", [period.net_income_keur for period in periods], "runtime", "Offline assembled P&L."),
            ("Net Dividends", [period.net_dividends_keur for period in periods], "runtime", "Offline assembled P&L."),
        ],
        number_format=K_EUR_FORMAT,
    )


def _write_cash_flow_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime")
    periods = list(bundle.statements.pf_cash_waterfall.periods)
    _write_horizontal_metric_table(
        sheet,
        6,
        "Project finance cash waterfall",
        [period.date.isoformat() for period in periods],
        [
            ("Revenue Cash", [period.revenue_cash_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("OPEX Cash", [period.opex_cash_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("EBITDA Cash", [period.ebitda_cash_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("Cash Tax", [period.cash_tax_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("FCF Banks", [period.fcf_banks_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("Senior Debt Service", [period.senior_total_ds_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("FCF for SHL", [period.fcf_for_shl_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("SHL Cash Interest", [period.shl_cash_interest_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("SHL Principal", [period.shl_principal_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
            ("Net Dividends", [period.net_dividends_keur for period in periods], "runtime", "Offline assembled PF cash waterfall."),
        ],
        number_format=K_EUR_FORMAT,
    )


def _write_balance_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "runtime")
    periods = list(bundle.statements.balance_sheet.periods)
    _write_horizontal_metric_table(
        sheet,
        6,
        "Balance sheet period table",
        [period.date.isoformat() for period in periods],
        [
            ("Gross Fixed Assets", [period.gross_fixed_assets_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Accumulated Depreciation", [period.accumulated_depreciation_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Net Fixed Assets", [period.net_fixed_assets_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Cash", [period.cash_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Senior Debt Balance", [period.senior_balance_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("SHL Balance", [period.shl_balance_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Retained Earnings", [period.retained_earnings_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Total Assets", [period.total_assets_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Total Liabilities + Equity", [period.total_liabilities_equity_keur for period in periods], "runtime", "Offline assembled balance sheet."),
            ("Balance Check", [period.balance_check_keur for period in periods], "runtime", "Offline assembled balance sheet residual check."),
        ],
        number_format=K_EUR_FORMAT,
    )


def _write_audit_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "review")
    rows = [
        ("Runtime source", "WaterfallRunner(project_inputs, engine).run(config)", "runtime", "Existing runtime authority."),
        ("Project assumption source", "ProjectContext + factory-bound ProjectInputs", "template assumption", "Read-only context layer."),
        ("P&L / tax / BS source", "assemble_financial_statements(runtime_result)", "runtime", "Offline assembly from existing runtime result."),
        ("Revenue period source", "build_revenue_table(runtime_result)", "runtime", "Existing output table builder."),
        ("Debt period source", "build_debt_table(runtime_result)", "runtime", "Existing output table builder."),
        ("Workbook-only formulas", "none", "review", "Presentation-only export branch."),
        ("Runtime / preview rule", "Sheets label runtime-derived sections separately from template assumptions", "review", "No blurred authority."),
        ("Governance note", "G20 remains blocked; R99/R102 remains not approved", "review", "No approval implied."),
    ]
    next_row = _write_key_value_section(sheet, 6, "Audit source map", rows)
    audit_rows = [
        ("Inputs", "project context + input helpers", "runtime + template", "Bound"),
        ("OPEX", "project context + runtime total", "runtime + template", "Bound"),
        ("CAPEX", "project context + capex helpers", "template + runtime", "Bound"),
        ("Revenue", "runtime result + output table", "runtime + template", "Bound"),
        ("Senior Debt", "runtime result + output table", "runtime + template", "Bound"),
        ("SHL", "runtime period fields", "runtime + template", "Bound"),
        ("Tax", "assembled tax bridge", "runtime + template", "Bound"),
        ("P&L", "assembled P&L", "runtime", "Bound"),
        ("Cash Flow", "assembled PF cash waterfall", "runtime", "Bound"),
        ("Balance Sheet", "assembled balance sheet", "runtime", "Bound"),
    ]
    _write_simple_table(
        sheet,
        next_row,
        "Sheet binding coverage",
        ("Sheet", "Primary source", "Runtime/preview", "Status"),
        audit_rows,
    )


def _write_gap_register_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    _write_metadata_block(sheet, bundle, "review")
    rows = [
        ("GAP-01", "Detailed construction schedule", "RUNTIME_BINDING_PENDING", "Construction", "Detailed runtime binding pending; summary only in this branch."),
        ("GAP-02", "Detailed OPEX escalation schedule", "WARN", "OPEX", "Workbook shows line items and total OPEX, not a full escalator roll-forward."),
        ("GAP-03", "Detailed CAPEX spend curve", "RUNTIME_BINDING_PENDING", "CAPEX", "CAPEX items are bound; full period spend curve is not exported here."),
        ("GAP-04", "R99/R102 runtime authority", "BLOCKER", "Tax / Cash Flow", "R99/R102 remains governance-only and not approved as runtime source."),
        ("GAP-05", "G20 sign-off", "BLOCKER", "Governance", "Workbook does not change governance status."),
        ("GAP-06", "Construction date and XIRR conventions", "ACCEPTED_CONVENTION", "Returns / governance", "Handled as documented convention, not a workbook formula change."),
        ("GAP-07", "Detailed runtime binding for remaining sub-lines", "WARN", "Multiple sheets", "Runtime summary and assembled statements are bound first; sub-line expansion is future work."),
    ]
    _write_simple_table(
        sheet,
        6,
        "Known gaps and accepted conventions",
        ("Gap ID", "Item", "Classification", "Section", "Notes"),
        rows,
        fill_column=3,
    )


def _write_export_metadata_sheet(sheet, bundle: WorkbookExportBundle) -> None:
    """Write the Export_Metadata sheet as the first sheet (Phase 47).

    This sheet provides trust hygiene, non-claims, and provenance information.
    It does not change any financial formulas or model outputs.
    """
    sheet["A1"] = "Export Metadata"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
    sheet["A2"] = "Trust hygiene, provenance, and non-claims — Phase 47"
    sheet["A2"].font = Font(size=11, italic=True, name="Calibri")

    # Build export metadata from bundle fields
    meta = build_export_metadata(
        project_id=bundle.project_key,
        project_name=bundle.project_name,
        scenario_id=bundle.scenario_id,
        scenario_name=bundle.scenario_name,
        active_project=bundle.active_project,
        run_at=bundle.runtime_timestamp,
        export_type="institutional_workbook",
    )

    rows_data = metadata_rows(meta)

    start_row = 4
    sheet.cell(row=start_row, column=1, value="Field")
    sheet.cell(row=start_row, column=2, value="Value")
    sheet.cell(row=start_row, column=1).fill = META_FILL
    sheet.cell(row=start_row, column=1).font = HEADER_FONT
    sheet.cell(row=start_row, column=2).fill = META_FILL
    sheet.cell(row=start_row, column=2).font = HEADER_FONT

    for offset, (label, value) in enumerate(rows_data, start=start_row + 1):
        sheet.cell(row=offset, column=1, value=label)
        sheet.cell(row=offset, column=1).fill = META_FILL
        sheet.cell(row=offset, column=1).font = HEADER_FONT
        cell = sheet.cell(row=offset, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True)

    # Non-claims block — prominent warning
    warning_row = start_row + len(rows_data) + 3
    sheet.cell(row=warning_row, column=1, value="NON-CLAIMS")
    sheet.cell(row=warning_row, column=1).fill = STATUS_BLOCKED_FILL
    sheet.cell(row=warning_row, column=1).font = HEADER_FONT
    sheet.merge_cells(
        start_row=warning_row, start_column=1,
        end_row=warning_row, end_column=2,
    )
    sheet.cell(row=warning_row, column=1).alignment = Alignment(horizontal="center")

    warnings = [
        "INTERNAL REVIEW EVIDENCE ONLY.",
        "NOT bank/lender approval. NOT external audit certification.",
        "NOT SaaS-ready or enterprise-ready. Controlled trusted pilot scope only.",
        "TUHO and Oborovo are within controlled trusted pilot scope.",
        "Generic solar/wind paths remain exploratory and unvalidated.",
        "Do not use generic paths for financial decisions.",
        "Backend is source of truth. Exports reflect the last clean backend run.",
        "G20 remains BLOCKED. R99/R102 remain NOT APPROVED.",
        "partial_pay_sweep not promoted. flat/min DSCR sculpting not promoted.",
        "Paid pilot remains NOT READY. Enterprise SaaS remains NOT READY.",
    ]
    for i, text in enumerate(warnings, start=warning_row + 1):
        sheet.cell(row=i, column=1, value=text)
        sheet.cell(row=i, column=1).fill = STATUS_BLOCKED_FILL
        sheet.cell(row=i, column=1).font = Font(size=10, bold=True, name="Calibri")
        sheet.merge_cells(
            start_row=i, start_column=1,
            end_row=i, end_column=2,
        )

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 72


def _write_metadata_block(sheet, bundle: WorkbookExportBundle, marker: str) -> None:
    metadata = (
        ("Project", bundle.project_name),
        ("Export generated at", bundle.generated_at),
        ("Runtime / preview", marker),
        ("Runtime boundary", f"{bundle.runtime_timestamp} | {bundle.runtime_origin} | {bundle.runtime_snapshot_id}"),
        ("Provenance", f"{bundle.branch or _source_branch()} | {bundle.commit_sha} | {bundle.template_origin} | {bundle.scenario_name}"),
    )
    for row_idx, (label, value) in enumerate(metadata, start=1):
        sheet.cell(row=row_idx, column=1, value=label)
        sheet.cell(row=row_idx, column=2, value=value)
        sheet.cell(row=row_idx, column=1).fill = META_FILL
        sheet.cell(row=row_idx, column=1).font = HEADER_FONT


def _write_key_value_section(
    sheet,
    start_row: int,
    title: str,
    rows: list[tuple],
    *,
    include_format: bool = False,
) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    _style_section_title(sheet, start_row)
    headers = ("Field", "Value", "Source", "Notes")
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=start_row + 1, column=col_idx, value=header)
        _style_header_cell(sheet.cell(row=start_row + 1, column=col_idx))
    for offset, row in enumerate(rows, start=2):
        label, value, source, notes = row[:4]
        sheet.cell(row=start_row + offset, column=1, value=label)
        value_cell = sheet.cell(row=start_row + offset, column=2, value=value)
        sheet.cell(row=start_row + offset, column=3, value=source)
        sheet.cell(row=start_row + offset, column=4, value=notes)
        _fill_source_row(sheet, start_row + offset, source)
        if include_format and len(row) > 4 and row[4]:
            value_cell.number_format = row[4]
    return start_row + len(rows) + 4


def _write_simple_table(
    sheet,
    start_row: int,
    title: str,
    headers: tuple[str, ...],
    rows: list[tuple],
    *,
    format_columns: dict[int, str] | None = None,
    fill_column: int | None = None,
) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    _style_section_title(sheet, start_row)
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=start_row + 1, column=col_idx, value=header)
        _style_header_cell(sheet.cell(row=start_row + 1, column=col_idx))
    for row_offset, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=start_row + row_offset, column=col_idx, value=value)
            if format_columns and col_idx in format_columns:
                cell.number_format = format_columns[col_idx]
        if fill_column and len(row) >= fill_column:
            _fill_classification_row(sheet, start_row + row_offset, str(row[fill_column - 1]))
    return start_row + len(rows) + 4


def _write_dataframe_section(sheet, start_row: int, title: str, dataframe, source: str, notes: str) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    _style_section_title(sheet, start_row)
    headers = ["Line Item"]
    headers.extend(str(column) for column in dataframe.columns)
    headers.extend(["Source", "Notes"])
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=start_row + 1, column=col_idx, value=header)
        _style_header_cell(sheet.cell(row=start_row + 1, column=col_idx))
    for row_offset, (index_label, row_values) in enumerate(dataframe.iterrows(), start=2):
        sheet.cell(row=start_row + row_offset, column=1, value=str(index_label))
        for col_idx, value in enumerate(row_values.tolist(), start=2):
            cell = sheet.cell(row=start_row + row_offset, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = K_EUR_FORMAT if str(index_label) != "DSCR" else MULTIPLE_FORMAT
        sheet.cell(row=start_row + row_offset, column=2 + len(dataframe.columns), value=source)
        sheet.cell(row=start_row + row_offset, column=3 + len(dataframe.columns), value=notes)
        _fill_source_row(sheet, start_row + row_offset, source)
    return start_row + len(dataframe.index) + 4


def _write_horizontal_metric_table(
    sheet,
    start_row: int,
    title: str,
    period_labels: list[str],
    rows: list[tuple[str, list[float], str, str]],
    *,
    number_format: str,
) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    _style_section_title(sheet, start_row)
    headers = ["Metric"]
    headers.extend(period_labels)
    headers.extend(["Total", "Source", "Notes"])
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=start_row + 1, column=col_idx, value=header)
        _style_header_cell(sheet.cell(row=start_row + 1, column=col_idx))

    for row_offset, (label, values, source, notes) in enumerate(rows, start=2):
        sheet.cell(row=start_row + row_offset, column=1, value=label)
        total = 0.0
        for col_idx, value in enumerate(values, start=2):
            cell = sheet.cell(row=start_row + row_offset, column=col_idx, value=value)
            cell.number_format = MULTIPLE_FORMAT if number_format == MULTIPLE_FORMAT else number_format
            total += float(value or 0.0)
        total_cell = sheet.cell(row=start_row + row_offset, column=2 + len(period_labels), value=total)
        total_cell.number_format = number_format
        sheet.cell(row=start_row + row_offset, column=3 + len(period_labels), value=source)
        sheet.cell(row=start_row + row_offset, column=4 + len(period_labels), value=notes)
        _fill_source_row(sheet, start_row + row_offset, source)
    return start_row + len(rows) + 4


def _style_section_title(sheet, row_idx: int) -> None:
    cell = sheet.cell(row=row_idx, column=1)
    cell.fill = SECTION_FILL
    cell.font = HEADER_FONT


def _style_header_cell(cell) -> None:
    cell.fill = SECTION_FILL
    cell.font = HEADER_FONT


def _fill_source_row(sheet, row_idx: int, source: str) -> None:
    fill = REVIEW_FILL
    if "runtime" in source:
        fill = RUNTIME_FILL
    elif "template" in source:
        fill = TEMPLATE_FILL
    for cell in sheet[row_idx]:
        if cell.column <= 4 or cell.value is not None:
            cell.fill = fill


def _fill_classification_row(sheet, row_idx: int, classification: str) -> None:
    normalized = classification.upper()
    fill = REVIEW_FILL
    if normalized == "BLOCKER":
        fill = STATUS_BLOCKED_FILL
    elif normalized in {"WARN", "MISSING_EVIDENCE", "MISSING_EXCEL_EVIDENCE", "MISSING_REVIEW_SCALAR", "SOURCE_NOT_AVAILABLE", "RUNTIME_BINDING_PENDING"}:
        fill = WARN_FILL
    elif normalized == "ACCEPTED_CONVENTION":
        fill = TEMPLATE_FILL
    for cell in sheet[row_idx]:
        cell.fill = fill


def _format_sheet(sheet) -> None:
    sheet.freeze_panes = "A7"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "$1:$6"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddHeader.left.text = "Institutional runtime workbook"
    sheet.oddHeader.center.text = sheet.title
    sheet.oddHeader.right.text = "Phase 11 polish"
    for letter, width in (("A", 28), ("B", 22), ("C", 18), ("D", 48), ("E", 16)):
        sheet.column_dimensions[letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _coerce_value(value: str, number_format: str) -> object:
    if not number_format:
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _period_label(period: object) -> str:
    if hasattr(period, "end_date") and getattr(period, "end_date") is not None:
        return getattr(period, "end_date").isoformat()
    if hasattr(period, "date") and getattr(period, "date") is not None:
        return getattr(period, "date").isoformat()
    if hasattr(period, "period") and getattr(period, "period") is not None:
        return str(getattr(period, "period"))
    return "unknown-period"


def _binding_fields_for_sheet(sheet_name: str) -> str:
    mapping = {
        "Cover": "project, timestamp, export type, branch, runtime/preview, purpose, status",
        "Governance": "governance status, G20, R99/R102, provenance",
        "Runtime Summary": "IRR, revenue, EBITDA, OPEX, DSCR, distributions, SHL service",
        "Inputs": "project metadata, technical assumptions, governance labels, input summary",
        "Construction": "dates, duration, capex anchors, financing anchors",
        "OPEX": "runtime total OPEX, contingency method, line items",
        "CAPEX": "total capex, funding anchors, capex summary, capex items",
        "Revenue": "PPA assumptions, availability, runtime revenue table",
        "Senior Debt": "debt assumptions, DSCR, debt service table",
        "SHL": "opening balance, rate, repayment method, SHL schedule",
        "Tax": "CIT assumptions, tax bridge periods, cash tax",
        "P&L": "assembled revenues, opex, depreciation, EBIT, EBT, taxable income, net income",
        "Cash Flow": "assembled PF cash waterfall, FCF, debt service, SHL, dividends",
        "Balance Sheet": "assembled assets, liabilities, equity, balance check",
        "Audit": "runtime sources, provenance, sheet coverage",
        "Gap Register": "known gaps, accepted conventions, governance blockers",
    }
    return mapping[sheet_name]


def _remaining_placeholders_for_sheet(sheet_name: str) -> str:
    mapping = {
        "Cover": "none",
        "Governance": "none",
        "Runtime Summary": "none",
        "Inputs": "detailed scenario override matrix",
        "Construction": "full construction spend curve",
        "OPEX": "full escalator schedule by period",
        "CAPEX": "full spend curve by construction period",
        "Revenue": "separate CO2/balancing sub-line sheets",
        "Senior Debt": "expanded covenant analytics",
        "SHL": "separate accrued-versus-cash reviewer view",
        "Tax": "expanded phase 6 residual diagnostics",
        "P&L": "extra sub-line mapping beyond assembled statement",
        "Cash Flow": "distribution account detail sheet",
        "Balance Sheet": "full capital accounts breakout",
        "Audit": "none",
        "Gap Register": "none",
    }
    return mapping[sheet_name]
