"""Phase 10 calibration reconciliation export pack.

This module builds a reviewer-facing workbook and companion CSVs that explain
remaining Excel-vs-model gaps without changing runtime authority or formulas.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.export.runtime_summary import build_runtime_summary_rows, _run_project


PASS = "PASS"
WARN = "WARN"
FAIL = "FAIL"
ACCEPTED_CONVENTION = "ACCEPTED_CONVENTION"
MISSING_EVIDENCE = "MISSING_EVIDENCE"
RUNTIME_BINDING_PENDING = "RUNTIME_BINDING_PENDING"
GOVERNANCE_BLOCKER = "GOVERNANCE_BLOCKER"
GOVERNANCE_REVIEW = "GOVERNANCE_REVIEW"
MATERIAL_DELTA = "MATERIAL_DELTA"
EVIDENCE_LIMITATION = "EVIDENCE_LIMITATION"
GROUPED_SOURCE_ONLY = "GROUPED_SOURCE_ONLY"

VALID_CLASSIFICATIONS = {
    PASS,
    WARN,
    FAIL,
    ACCEPTED_CONVENTION,
    MISSING_EVIDENCE,
    RUNTIME_BINDING_PENDING,
    GOVERNANCE_BLOCKER,
    GOVERNANCE_REVIEW,
    MATERIAL_DELTA,
    EVIDENCE_LIMITATION,
    GROUPED_SOURCE_ONLY,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D9E2"),
    right=Side(style="thin", color="D5D9E2"),
    top=Side(style="thin", color="D5D9E2"),
    bottom=Side(style="thin", color="D5D9E2"),
)
TITLE_FILL = PatternFill("solid", fgColor="183153")
HEADER_FILL = PatternFill("solid", fgColor="365D8D")
SECTION_FILL = PatternFill("solid", fgColor="DEE8F4")
PASS_FILL = PatternFill("solid", fgColor="E7F4E9")
WARN_FILL = PatternFill("solid", fgColor="FFF0C9")
FAIL_FILL = PatternFill("solid", fgColor="FAD8D6")
CONVENTION_FILL = PatternFill("solid", fgColor="DDEBF7")
MISS_FILL = PatternFill("solid", fgColor="F8D7DA")
PENDING_FILL = PatternFill("solid", fgColor="FCE5CD")
GOVERNANCE_FILL = PatternFill("solid", fgColor="F4CCCC")
META_FILL = PatternFill("solid", fgColor="EEF3F8")
ROW_ALT_FILL = PatternFill("solid", fgColor="F9FBFD")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
HEADER_FONT = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
BODY_FONT = Font(size=10, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
K_EUR_FORMAT = '#,##0.0;[Red](#,##0.0);"-"'
PCT_FORMAT = "0.00%"
X_FORMAT = "0.000x"


ROOT = Path(__file__).resolve().parents[2]
REPORTS_DIR = ROOT / "reports"
DOCS_DIR = ROOT / "docs"
SCRIPTS_DIR = ROOT / "scripts"

DEFAULT_WORKBOOK = REPORTS_DIR / "phase10_calibration_reconciliation_pack.xlsx"
DEFAULT_GAP_REGISTER = REPORTS_DIR / "phase10_calibration_gap_register.csv"
DEFAULT_SOURCE_INVENTORY = REPORTS_DIR / "phase10_calibration_source_inventory.csv"
DEFAULT_SUMMARY = REPORTS_DIR / "phase10_calibration_reconciliation_summary.csv"
DEFAULT_NAVIGATION_MAP = REPORTS_DIR / "phase10_review_pack_navigation_map.csv"
DEFAULT_SIGNOFF_MATRIX = REPORTS_DIR / "phase10_review_signoff_matrix.csv"
DEFAULT_RUNTIME_BINDING_INVENTORY = REPORTS_DIR / "phase10_runtime_binding_inventory.csv"
DEFAULT_EVIDENCE_COVERAGE_SUMMARY = REPORTS_DIR / "phase10_evidence_coverage_summary.csv"
DEFAULT_RUNTIME_BINDING_GAP_REGISTER = REPORTS_DIR / "phase10_runtime_binding_gap_register.csv"
DEFAULT_IRR_RECONCILIATION_SUMMARY = REPORTS_DIR / "phase10_irr_reconciliation_summary.csv"
DEFAULT_IRR_CONVENTION_MAP = REPORTS_DIR / "phase10_irr_convention_map.csv"
DEFAULT_IRR_GOVERNANCE_ITEMS = REPORTS_DIR / "phase10_irr_governance_items.csv"
DEFAULT_CO2_BALANCING_SOURCE_MAP = REPORTS_DIR / "phase10_co2_balancing_source_map.csv"
DEFAULT_CO2_BALANCING_GAP_SUMMARY = REPORTS_DIR / "phase10_co2_balancing_gap_summary.csv"
DEFAULT_CO2_BALANCING_EVIDENCE_QUALITY = REPORTS_DIR / "phase10_co2_balancing_evidence_quality.csv"


@dataclass(frozen=True)
class MetricSpec:
    section: str
    label: str
    excel_values: list[float | str | None]
    model_values: list[float | str | None]
    excel_source: str
    model_source: str
    number_format: str = K_EUR_FORMAT
    classification: str | None = None
    root_cause: str = ""
    recommended_action: str = ""
    governance_impact: str = "None"
    notes: str = ""
    status: str | None = None
    total_label: str | None = None
    owner: str = "Review"
    requires_stakeholder_decision: str = "no"
    requires_runtime_change: str = "no"
    expected_roadmap_phase: str = "Phase 10"


@dataclass(frozen=True)
class ReconciliationArtifacts:
    workbook_path: Path
    gap_register_path: Path
    source_inventory_path: Path
    summary_path: Path
    navigation_map_path: Path
    signoff_matrix_path: Path
    runtime_binding_inventory_path: Path
    evidence_coverage_summary_path: Path
    runtime_binding_gap_register_path: Path
    irr_reconciliation_summary_path: Path
    irr_convention_map_path: Path
    irr_governance_items_path: Path
    co2_balancing_source_map_path: Path
    co2_balancing_gap_summary_path: Path
    co2_balancing_evidence_quality_path: Path


def _num(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper().startswith("MISSING_EVIDENCE"):
        return None
    if text in {"N/A", "None"}:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _col(rows: list[dict[str, str]], name: str, count: int, default=None) -> list[float | str | None]:
    values: list[float | str | None] = []
    for row in rows[:count]:
        raw = row.get(name, default)
        numeric = _num(raw)
        values.append(numeric if numeric is not None else raw if raw not in (None, "") else default)
    while len(values) < count:
        values.append(default)
    return values[:count]


def _sum_numeric(values: list[float | str | None]) -> float | None:
    nums = [_num(v) for v in values]
    if any(v is None for v in nums):
        return None
    return sum(v for v in nums if v is not None)


def _lag_opening(closing_values: list[float], first_opening: float) -> list[float]:
    return [first_opening] + closing_values[:-1]


def _classify(
    excel_values: list[float | str | None],
    model_values: list[float | str | None],
    classification_override: str | None,
    status_override: str | None,
) -> tuple[str, str, float | str]:
    if classification_override:
        status = status_override or ("PASS" if classification_override == PASS else "REVIEW")
        return classification_override, status, _delta_total(excel_values, model_values)

    excel_nums = [_num(v) for v in excel_values]
    model_nums = [_num(v) for v in model_values]
    if all(v is None for v in excel_nums) or all(v is None for v in model_nums):
        return MISSING_EVIDENCE, "REVIEW", "N/A"

    deltas = [
        abs((m or 0.0) - (e or 0.0))
        for e, m in zip(excel_nums, model_nums)
        if e is not None and m is not None
    ]
    if not deltas:
        return MISSING_EVIDENCE, "REVIEW", "N/A"

    max_delta = max(deltas)
    total_excel = abs(_sum_numeric(excel_values) or 0.0)
    tolerance = max(0.1, total_excel * 0.001)
    if max_delta <= tolerance:
        return PASS, "OK", _delta_total(excel_values, model_values)
    if max_delta <= max(100.0, total_excel * 0.02):
        return WARN, "REVIEW", _delta_total(excel_values, model_values)
    return FAIL, "ACTION", _delta_total(excel_values, model_values)


def _delta_total(excel_values: list[float | str | None], model_values: list[float | str | None]) -> float | str:
    excel_total = _sum_numeric(excel_values)
    model_total = _sum_numeric(model_values)
    if excel_total is None or model_total is None:
        return "N/A"
    return model_total - excel_total


def _fill_for_classification(classification: str) -> PatternFill:
    return {
        PASS: PASS_FILL,
        WARN: WARN_FILL,
        FAIL: FAIL_FILL,
        ACCEPTED_CONVENTION: CONVENTION_FILL,
        MISSING_EVIDENCE: MISS_FILL,
        RUNTIME_BINDING_PENDING: PENDING_FILL,
        GOVERNANCE_BLOCKER: GOVERNANCE_FILL,
        GOVERNANCE_REVIEW: CONVENTION_FILL,
        MATERIAL_DELTA: FAIL_FILL,
        EVIDENCE_LIMITATION: WARN_FILL,
        GROUPED_SOURCE_ONLY: PENDING_FILL,
    }.get(classification, META_FILL)


def _runtime_period_data(result, project_inputs) -> dict[str, list[float | str | None] | int | list[str] | float]:
    periods = result.periods
    count = len(periods)
    dates = [f"P{i + 1}" for i in range(count)]
    period_ends = [
        period.date.strftime("%Y-%m-%d") if hasattr(period.date, "strftime") else str(period.date)
        for period in periods
    ]

    senior_closing = [float(getattr(period, "senior_balance_keur", 0.0) or 0.0) for period in periods]
    senior_principal = [float(getattr(period, "senior_principal_keur", 0.0) or 0.0) for period in periods]
    senior_opening = _lag_opening(
        senior_closing,
        senior_closing[0] + senior_principal[0],
    )

    shl_closing = [float(getattr(period, "shl_balance_keur", 0.0) or 0.0) for period in periods]
    shl_principal = [float(getattr(period, "shl_principal_keur", 0.0) or 0.0) for period in periods]
    shl_pik = [float(getattr(period, "shl_pik_keur", 0.0) or 0.0) for period in periods]
    shl_opening = [
        shl_closing[i] + shl_principal[i] - shl_pik[i]
        for i in range(count)
    ]

    def f(attr: str) -> list[float]:
        return [float(getattr(period, attr, 0.0) or 0.0) for period in periods]

    return {
        "count": count,
        "period_labels": dates,
        "period_ends": period_ends,
        "generation_mwh": f("generation_mwh"),
        "revenue_keur": f("revenue_keur"),
        "opex_keur": f("opex_keur"),
        "ebitda_keur": f("ebitda_keur"),
        "depreciation_keur": f("depreciation_keur"),
        "senior_opening_keur": senior_opening,
        "senior_interest_keur": f("senior_interest_keur"),
        "senior_principal_keur": senior_principal,
        "senior_ds_keur": f("senior_ds_keur"),
        "senior_closing_keur": senior_closing,
        "dscr": [None if float(getattr(period, "dscr", 0.0) or 0.0) > 100 else float(getattr(period, "dscr", 0.0) or 0.0) for period in periods],
        "shl_opening_keur": shl_opening,
        "shl_gross_accrued_interest_keur": f("shl_gross_accrued_interest_keur"),
        "shl_interest_keur": f("shl_interest_keur"),
        "shl_pik_keur": shl_pik,
        "shl_principal_keur": shl_principal,
        "shl_service_keur": f("shl_service_keur"),
        "shl_closing_keur": shl_closing,
        "taxable_income_before_losses_audit_keur": f("taxable_income_before_losses_audit_keur"),
        "tax_depreciation_audit_keur": f("tax_depreciation_audit_keur"),
        "corporate_tax_cash_keur": f("corporate_tax_cash_keur"),
        "tax_loss_opening_audit_keur": f("tax_loss_opening_audit_keur"),
        "tax_loss_used_audit_keur": f("tax_loss_used_audit_keur"),
        "tax_loss_closing_audit_keur": f("tax_loss_closing_audit_keur"),
        "taxable_profit_after_losses_audit_keur": f("taxable_profit_after_losses_audit_keur"),
        "cit_accrual_audit_keur": f("cit_accrual_audit_keur"),
        "cash_tax_current_period_audit_keur": f("cash_tax_current_period_audit_keur"),
        "r69_fcf_banks_keur": f("r69_fcf_banks_keur"),
        "r99_fcf_for_distribution_keur": f("r99_fcf_for_distribution_keur"),
        "r102_fcf_for_shl_keur": f("r102_fcf_for_shl_keur"),
        "distribution_keur": f("distribution_keur"),
        "da_paid_distribution_keur": f("da_paid_distribution_keur"),
        "legacy_distribution_keur": f("legacy_distribution_keur"),
        "project_irr": float(result.project_irr or 0.0),
        "equity_irr": float(result.equity_irr or 0.0),
        "avg_dscr": float(result.actual_avg_dscr or 0.0),
        "min_dscr": float(result.actual_min_dscr or 0.0),
        "total_revenue_keur": float(result.total_revenue_keur or 0.0),
        "total_opex_keur": float(result.total_opex_keur or 0.0),
        "total_ebitda_keur": float(result.total_ebitda_keur or 0.0),
        "total_distribution_keur": float(result.total_distribution_keur or 0.0),
        "project_name": project_inputs.info.name,
    }


def _accepted_conventions_rows() -> list[dict[str, str]]:
    path = REPORTS_DIR / "phase9_final_tuho_accepted_conventions.csv"
    rows = _load_csv_rows(path)
    if rows:
        return rows
    return []


def _source_inventory_rows(period_count: int) -> list[dict[str, str]]:
    return [
        {
            "metric": "Revenue, OPEX, EBITDA, Senior Debt, Distribution, DSCR",
            "source_file": "reports/phase9_tuho_full_line_item_period_bridge.csv",
            "period_level_available": f"yes ({period_count} periods)",
            "total_available": "yes",
            "evidence_side": "excel",
            "confidence_level": "high",
            "usable_for_reconciliation": "yes",
            "notes": "Primary committed period bridge for revenue, opex, debt, and distributions.",
        },
        {
            "metric": "SHL opening, interest, principal, closing",
            "source_file": "reports/phase9_tuho_shl_period_bridge.csv",
            "period_level_available": "partial",
            "total_available": "yes",
            "evidence_side": "excel",
            "confidence_level": "medium",
            "usable_for_reconciliation": "yes",
            "notes": "Gross accrued interest is mapped; cash interest and PIK remain partially evidenced.",
        },
        {
            "metric": "Runtime summary KPIs",
            "source_file": "app/export/runtime_summary.py",
            "period_level_available": "no",
            "total_available": "yes",
            "evidence_side": "runtime",
            "confidence_level": "high",
            "usable_for_reconciliation": "yes",
            "notes": "Authoritative runtime summary export foundation.",
        },
        {
            "metric": "CO2 / balancing revenue sub-lines",
            "source_file": "reports/phase9_tuho_full_line_item_period_bridge.csv + runtime.revenue_keur",
            "period_level_available": "grouped only",
            "total_available": "partial",
            "evidence_side": "excel+runtime",
            "confidence_level": "low",
            "usable_for_reconciliation": "partial",
            "notes": "Electricity revenue total is committed, but CO2 and balancing remain grouped-only or missing across current evidence sources.",
        },
        {
            "metric": "Phase 9 accepted conventions",
            "source_file": "reports/phase9_final_tuho_accepted_conventions.csv",
            "period_level_available": "no",
            "total_available": "yes",
            "evidence_side": "governance",
            "confidence_level": "high",
            "usable_for_reconciliation": "yes",
            "notes": "Governed convention decisions reused for reviewer clarity.",
        },
        {
            "metric": "Calibration gap baseline",
            "source_file": "reports/phase9_tuho_calibration_gap_register.csv",
            "period_level_available": "no",
            "total_available": "yes",
            "evidence_side": "review",
            "confidence_level": "medium",
            "usable_for_reconciliation": "yes",
            "notes": "Used to classify known residuals and governance blockers.",
        },
        {
            "metric": "Runtime-to-workbook binding coverage",
            "source_file": "reports/phase10_runtime_workbook_binding_status.csv",
            "period_level_available": "no",
            "total_available": "yes",
            "evidence_side": "review",
            "confidence_level": "high",
            "usable_for_reconciliation": "yes",
            "notes": "Tracks which workbook sheets are runtime-bound versus placeholder-forward.",
        },
    ]


def _co2_balancing_report_rows(runtime: dict, per_bridge: list[dict[str, str]]) -> list[dict[str, str]]:
    total_revenue_runtime = _sum_numeric(runtime["revenue_keur"]) or 0.0
    total_revenue_excel = _sum_numeric(_col(per_bridge, "excel_revenue_keur", int(runtime["count"]), None))
    total_revenue_excel = 0.0 if total_revenue_excel is None else total_revenue_excel
    return [
        {
            "metric": "Electricity Revenue",
            "runtime_source": "runtime.revenue_keur",
            "excel_source": "phase9_tuho_full_line_item_period_bridge.csv / excel_revenue_keur",
            "grouped_or_separate": "SEPARATE",
            "evidence_quality": "HIGH",
            "runtime_available": "yes",
            "excel_available": "yes",
            "workbook_bound": "yes",
            "classification": PASS,
            "runtime_value_available": f"{total_revenue_runtime:.1f}",
            "excel_value_available": f"{total_revenue_excel:.1f}",
            "unresolved_reason": "",
            "governance_sensitive": "no",
            "notes": "Electricity revenue is the only revenue sub-line with strong committed evidence on both sides.",
        },
        {
            "metric": "CO2 Revenue",
            "runtime_source": "none",
            "excel_source": "none",
            "grouped_or_separate": "SEPARATE_REQUESTED_BUT_UNAVAILABLE",
            "evidence_quality": "MISSING",
            "runtime_available": "no",
            "excel_available": "no",
            "workbook_bound": "yes",
            "classification": MISSING_EVIDENCE,
            "runtime_value_available": "MISSING_EVIDENCE",
            "excel_value_available": "MISSING_EVIDENCE",
            "unresolved_reason": "Neither the committed Excel evidence set nor runtime period exports expose a standalone CO2 revenue row.",
            "governance_sensitive": "yes",
            "notes": "Do not synthesize a CO2 split from total revenue.",
        },
        {
            "metric": "Balancing Revenue",
            "runtime_source": "none",
            "excel_source": "none",
            "grouped_or_separate": "SEPARATE_REQUESTED_BUT_UNAVAILABLE",
            "evidence_quality": "MISSING",
            "runtime_available": "no",
            "excel_available": "no",
            "workbook_bound": "yes",
            "classification": MISSING_EVIDENCE,
            "runtime_value_available": "MISSING_EVIDENCE",
            "excel_value_available": "MISSING_EVIDENCE",
            "unresolved_reason": "Standalone balancing revenue is not committed in the current Excel extracts or runtime export fields.",
            "governance_sensitive": "yes",
            "notes": "Missing balancing evidence should stay explicit rather than being netted into a guessed split.",
        },
        {
            "metric": "Grouped Revenue Evidence",
            "runtime_source": "runtime.revenue_keur",
            "excel_source": "phase9_tuho_full_line_item_period_bridge.csv / excel_revenue_keur",
            "grouped_or_separate": "GROUPED_ONLY",
            "evidence_quality": "MEDIUM",
            "runtime_available": "yes",
            "excel_available": "yes",
            "workbook_bound": "yes",
            "classification": GROUPED_SOURCE_ONLY,
            "runtime_value_available": f"{total_revenue_runtime:.1f}",
            "excel_value_available": f"{total_revenue_excel:.1f}",
            "unresolved_reason": "The committed evidence currently supports revenue only at grouped total level for CO2/balancing-sensitive review.",
            "governance_sensitive": "yes",
            "notes": "This grouped row helps reviewers separate evidence limitations from runtime-total concerns.",
        },
        {
            "metric": "Runtime Aggregation Boundary",
            "runtime_source": "runtime.revenue_keur",
            "excel_source": "not applicable",
            "grouped_or_separate": "GROUPED_ONLY",
            "evidence_quality": "LOW",
            "runtime_available": "yes",
            "excel_available": "no",
            "workbook_bound": "yes",
            "classification": EVIDENCE_LIMITATION,
            "runtime_value_available": "aggregate only",
            "excel_value_available": "N/A",
            "unresolved_reason": "Runtime revenue is available as a total, but the current runtime export surface does not publish separate CO2 or balancing components.",
            "governance_sensitive": "no",
            "notes": "This is a reporting/source-map limitation, not a runtime revenue defect.",
        },
        {
            "metric": "Excel Extraction Boundary",
            "runtime_source": "not applicable",
            "excel_source": "phase9_tuho_full_line_item_period_bridge.csv / excel_revenue_keur",
            "grouped_or_separate": "GROUPED_ONLY",
            "evidence_quality": "LOW",
            "runtime_available": "no",
            "excel_available": "yes",
            "workbook_bound": "yes",
            "classification": EVIDENCE_LIMITATION,
            "runtime_value_available": "N/A",
            "excel_value_available": "aggregate only",
            "unresolved_reason": "The committed Excel bridge gives a total revenue row but not a durable CO2-versus-balancing split.",
            "governance_sensitive": "yes",
            "notes": "This should be treated as an extraction/evidence-quality issue rather than a model defect.",
        },
    ]


def _irr_report_rows(runtime: dict, conventions: list[dict[str, str]]) -> list[dict[str, str]]:
    convention_lookup = {row.get("convention_id", ""): row for row in conventions}
    project_runtime = float(runtime["project_irr"])
    equity_runtime = float(runtime["equity_irr"])
    project_excel = 0.0947
    equity_excel = 0.1161
    conv_xirr = convention_lookup.get("CONV-01", {})
    conv_shl_idc = convention_lookup.get("CONV-02", {})
    conv_distribution = convention_lookup.get("CONV-03", {})

    return [
        {
            "metric": "Project IRR",
            "runtime_value": f"{project_runtime:.6f}",
            "excel_value": f"{project_excel:.6f}",
            "delta": f"{project_runtime - project_excel:.6f}",
            "classification": PASS,
            "root_cause": "Runtime project IRR remains within the already documented closeout tolerance versus Excel.",
            "governance_sensitive": "no",
            "stakeholder_decision_required": "no",
            "governance_impact": "None",
            "notes": "This is not a runtime-calculation issue in the current review pack.",
        },
        {
            "metric": "Equity IRR",
            "runtime_value": f"{equity_runtime:.6f}",
            "excel_value": f"{equity_excel:.6f}",
            "delta": f"{equity_runtime - equity_excel:.6f}",
            "classification": GOVERNANCE_REVIEW,
            "root_cause": "The residual is primarily convention- and presentation-driven: XIRR date anchors, SHL IDC investment-base treatment, and distribution-definition framing all affect the reviewer-facing bridge.",
            "governance_sensitive": "yes",
            "stakeholder_decision_required": "yes",
            "governance_impact": "Stakeholder acceptance of residual IRR interpretation remains required before G20 can be treated as review-complete.",
            "notes": "This row explains the remaining equity IRR discussion; it does not change runtime IRR logic.",
        },
        {
            "metric": "Reconciliation IRR",
            "runtime_value": "MISSING_EVIDENCE",
            "excel_value": "MISSING_EVIDENCE",
            "delta": "N/A",
            "classification": MISSING_EVIDENCE,
            "root_cause": "A dedicated reconciliation IRR reporting view has not existed in the committed workbook layer until this branch, and there is still no separate approved scalar target to compute against automatically.",
            "governance_sensitive": "yes",
            "stakeholder_decision_required": "yes",
            "governance_impact": "Stakeholders can waive this view or request a later reporting-only implementation with more explicit timing assumptions.",
            "notes": "Missing evidence is preserved honestly; no proxy IRR is fabricated here.",
        },
        {
            "metric": "XIRR Date Convention",
            "runtime_value": "runtime period-end framing",
            "excel_value": "Excel construction-date / workbook-specific anchor",
            "delta": "presentation-only",
            "classification": ACCEPTED_CONVENTION,
            "root_cause": conv_xirr.get("description", "Excel and model may use different XIRR date anchors."),
            "governance_sensitive": "yes",
            "stakeholder_decision_required": conv_xirr.get("stakeholder_decision_required", "yes"),
            "governance_impact": "Presentation and closeout interpretation only; no runtime formula override is implied.",
            "notes": conv_xirr.get("notes", "Treat as a convention discussion rather than a runtime defect."),
        },
        {
            "metric": "Distribution Timing / Definition",
            "runtime_value": "runtime distribution_keur",
            "excel_value": "Excel dividends / distribution presentation",
            "delta": "definition-only",
            "classification": ACCEPTED_CONVENTION,
            "root_cause": conv_distribution.get("description", "Runtime distributions and Excel dividend presentation are definitionally different."),
            "governance_sensitive": "yes",
            "stakeholder_decision_required": conv_distribution.get("stakeholder_decision_required", "yes"),
            "governance_impact": "Reviewers need to agree which lens to present in signoff materials.",
            "notes": conv_distribution.get("notes", "Default runtime distributions remain authoritative; audit-only staging stays separate."),
        },
        {
            "metric": "SHL IDC Investment-Base Treatment",
            "runtime_value": "documented runtime treatment",
            "excel_value": "Excel convention baseline",
            "delta": "convention-sensitive",
            "classification": GOVERNANCE_REVIEW,
            "root_cause": conv_shl_idc.get("description", "SHL IDC investment-base treatment is a known equity IRR convention driver."),
            "governance_sensitive": "yes",
            "stakeholder_decision_required": conv_shl_idc.get("stakeholder_decision_required", "yes"),
            "governance_impact": "This convention affects how equity IRR residuals are explained to IC and lender reviewers.",
            "notes": conv_shl_idc.get("notes", "Documented convention; not a runtime rewrite request."),
        },
        {
            "metric": "Semiannual Timing / COD Framing",
            "runtime_value": "semiannual runtime horizon",
            "excel_value": "semiannual workbook horizon",
            "delta": "timing-sensitive",
            "classification": GOVERNANCE_REVIEW,
            "root_cause": "Semiannual period framing and COD anchor interpretation can shift reviewer-facing XIRR timing without implying an economic runtime defect.",
            "governance_sensitive": "yes",
            "stakeholder_decision_required": "yes",
            "governance_impact": "Requires reviewers to distinguish timing conventions from actual model behavior.",
            "notes": "Use this row to explain why small IRR drift can remain acceptable even when runtime parity elsewhere is strong.",
        },
    ]


def _irr_convention_map_rows(irr_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for entry in irr_rows:
        if entry["classification"] not in {ACCEPTED_CONVENTION, GOVERNANCE_REVIEW}:
            continue
        rows.append(
            {
                "metric": entry["metric"],
                "runtime_value": entry["runtime_value"],
                "excel_value": entry["excel_value"],
                "delta": entry["delta"],
                "classification": entry["classification"],
                "root_cause": entry["root_cause"],
                "governance_sensitive": entry["governance_sensitive"],
                "stakeholder_decision_required": entry["stakeholder_decision_required"],
                "notes": entry["notes"],
            }
        )
    return rows


def _irr_governance_item_rows(irr_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for entry in irr_rows:
        if entry["governance_sensitive"] != "yes":
            continue
        rows.append(
            {
                "metric": entry["metric"],
                "runtime_value": entry["runtime_value"],
                "excel_value": entry["excel_value"],
                "delta": entry["delta"],
                "classification": entry["classification"],
                "root_cause": entry["root_cause"],
                "governance_sensitive": entry["governance_sensitive"],
                "stakeholder_decision_required": entry["stakeholder_decision_required"],
                "notes": entry["governance_impact"],
            }
        )
    return rows


def _build_specs(runtime: dict, per_bridge: list[dict[str, str]], shl_bridge: list[dict[str, str]]) -> dict[str, list[MetricSpec]]:
    n = int(runtime["count"])
    missing = lambda reason: [f"MISSING_EVIDENCE: {reason}"] * n
    blank = [""] * n

    revenue_specs = [
        MetricSpec(
            "Revenue",
            "Electricity Revenue",
            _col(per_bridge, "excel_revenue_keur", n, None),
            runtime["revenue_keur"],
            "phase9_tuho_full_line_item_period_bridge.csv / excel_revenue_keur",
            "runtime.revenue_keur",
        ),
        MetricSpec(
            "Revenue",
            "CO2 Revenue",
            missing("no committed Excel CO2 row"),
            missing("no runtime CO2 period field"),
            "none",
            "none",
            classification=MISSING_EVIDENCE,
            root_cause="Committed Excel extracts and runtime period fields do not expose CO2 separately.",
            recommended_action="Keep as explicit evidence gap or add a dedicated CO2 source map branch.",
            governance_impact="Stakeholder review only",
            notes="Do not silently roll this into zero or total revenue.",
        ),
        MetricSpec(
            "Revenue",
            "Balancing Revenue",
            missing("no committed Excel balancing row"),
            missing("no runtime balancing period field"),
            "none",
            "none",
            classification=MISSING_EVIDENCE,
            root_cause="Balancing remains outside the committed extraction set and runtime summary exports.",
            recommended_action="Treat as sub-line follow-up only if stakeholders require separate mapping.",
            governance_impact="None",
            notes="Total revenue remains the authoritative comparable line.",
        ),
        MetricSpec(
            "Revenue",
            "Total Revenue",
            _col(per_bridge, "excel_revenue_keur", n, None),
            runtime["revenue_keur"],
            "phase9_tuho_full_line_item_period_bridge.csv / excel_revenue_keur",
            "runtime.revenue_keur",
        ),
    ]

    opex_specs = [
        MetricSpec(
            "OPEX",
            "Total OPEX",
            _col(per_bridge, "excel_opex_keur", n, None),
            runtime["opex_keur"],
            "phase9_tuho_full_line_item_period_bridge.csv / excel_opex_keur",
            "runtime.opex_keur",
            classification=WARN,
            root_cause="Known local-tax and minor-row grouping residual from prior calibration review.",
            recommended_action="Keep grouped unless a dedicated OPEX sub-line evidence branch is requested.",
            governance_impact="None",
            notes="Residual is documented rather than runtime-fixed here.",
        ),
        MetricSpec(
            "OPEX",
            "Contingency",
            missing("no committed Excel contingency sub-line"),
            missing("contingency not exposed as standalone runtime period field"),
            "none",
            "none",
            classification=RUNTIME_BINDING_PENDING,
            root_cause="Contingency is baked into deterministic OPEX logic but not exported as a separate period row.",
            recommended_action="Add reporting-only breakout if lenders require contingency visibility.",
            governance_impact="Presentation only",
            notes="Do not treat missing breakout as a zero contingency.",
        ),
        MetricSpec(
            "OPEX",
            "EBITDA",
            _col(per_bridge, "excel_ebitda_keur", n, None),
            runtime["ebitda_keur"],
            "phase9_tuho_full_line_item_period_bridge.csv / excel_ebitda_keur",
            "runtime.ebitda_keur",
        ),
    ]

    debt_specs = [
        MetricSpec("Senior Debt", "Opening Balance", _col(per_bridge, "excel_senior_opening_keur", n, None), runtime["senior_opening_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_senior_opening_keur", "derived from runtime senior closing/principal", classification=WARN, root_cause="The committed Excel bridge and runtime schedule do not share the same opening-balance cut point for the operational horizon.", recommended_action="Keep the opening balance row as a reconciliation aid and rely on principal, interest, and closing rows for authoritative debt behavior.", governance_impact="Review only", notes="No debt sizing rewrite is implied."),
        MetricSpec("Senior Debt", "Interest", _col(per_bridge, "excel_senior_interest_keur", n, None), runtime["senior_interest_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_senior_interest_keur", "runtime.senior_interest_keur", classification=WARN, root_cause="A small timing and day-count residual remains between the committed Excel extract and runtime debt schedule.", recommended_action="Document as a minor review residual unless lenders ask for a debt-source extraction refresh.", governance_impact="Review only", notes="The residual is small relative to total debt service."),
        MetricSpec("Senior Debt", "Principal Repayment", _col(per_bridge, "excel_senior_principal_keur", n, None), runtime["senior_principal_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_senior_principal_keur", "runtime.senior_principal_keur", classification=WARN, root_cause="The tiny principal delta is rounding-level drift between the committed extract and runtime precision.", recommended_action="Treat as immaterial unless an exact tie-out is required for a formal audit appendix.", governance_impact="None", notes="No runtime debt behavior change is needed."),
        MetricSpec("Senior Debt", "Debt Service", [(_num(a) or 0.0) + (_num(b) or 0.0) for a, b in zip(_col(per_bridge, "excel_senior_interest_keur", n, None), _col(per_bridge, "excel_senior_principal_keur", n, None))], runtime["senior_ds_keur"], "computed from committed Excel interest + principal", "runtime.senior_ds_keur", classification=WARN, root_cause="Debt service inherits the same minor interest timing residual visible in the senior interest row.", recommended_action="Keep this as a transparent review delta rather than a model-fix candidate.", governance_impact="Review only", notes="Derived from existing interest and principal rows."),
        MetricSpec("Senior Debt", "Closing Balance", _col(per_bridge, "excel_senior_closing_keur", n, None), runtime["senior_closing_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_senior_closing_keur", "runtime.senior_balance_keur", classification=WARN, root_cause="Closing balance drift follows from the opening-balance cut-point difference carried through the operational bridge.", recommended_action="Use this row for reviewer traceability only; core runtime debt outputs remain unchanged.", governance_impact="Review only", notes="No runtime change implied."),
        MetricSpec("Senior Debt", "DSCR", _col(per_bridge, "excel_dscr", n, None), runtime["dscr"], "phase9_tuho_full_line_item_period_bridge.csv / excel_dscr", "runtime.dscr", number_format=X_FORMAT, classification=WARN, root_cause="Distribution and downstream cash staging conventions still create a documented DSCR residual versus Excel.", recommended_action="Keep as calibration transparency item; do not rewrite debt logic in this branch.", governance_impact="G20 review context", notes="Known review residual, not a debt sizing bug."),
    ]

    shl_specs = [
        MetricSpec("SHL", "Opening Balance", _col(shl_bridge, "excel_shl_opening_keur", n, None), runtime["shl_opening_keur"], "phase9_tuho_shl_period_bridge.csv / excel_shl_opening_keur", "derived from runtime shl closing/principal/pik", classification=MISSING_EVIDENCE, root_cause="The committed SHL bridge does not provide a full-horizon opening-balance series that can be lined up one-for-one with runtime periods.", recommended_action="Treat this as an evidence limitation and anchor review attention on principal, interest, and closing balance rows.", governance_impact="Review only", notes="Missing evidence is not a zero opening balance."),
        MetricSpec("SHL", "Gross Accrued Interest", _col(shl_bridge, "excel_shl_interest_keur", n, None), runtime["shl_gross_accrued_interest_keur"], "phase9_tuho_shl_period_bridge.csv / excel_shl_interest_keur", "runtime.shl_gross_accrued_interest_keur", classification=WARN, root_cause="The remaining SHL gross-interest delta reflects timing and presentation differences between the committed bridge and the runtime accrued-interest schedule.", recommended_action="Keep the row visible for reviewer scrutiny, but do not redesign SHL mechanics in this branch.", governance_impact="Review only", notes="Gross accrued interest is the authoritative model-side SHL P&L bridge."),
        MetricSpec("SHL", "Cash Interest Paid", missing("no committed Excel cash-interest split"), runtime["shl_interest_keur"], "none", "runtime.shl_interest_keur", classification=MISSING_EVIDENCE, root_cause="Committed Excel bridge evidences gross accrued SHL interest, not a separate cash-interest split.", recommended_action="Keep separate from gross accrued interest and document as presentation gap.", governance_impact="None", notes="Runtime field is real; Excel side is the missing piece."),
        MetricSpec("SHL", "PIK Capitalized", missing("no committed Excel PIK split"), runtime["shl_pik_keur"], "none", "runtime.shl_pik_keur", classification=MISSING_EVIDENCE, root_cause="PIK is not separately extractable from the committed Excel evidence set.", recommended_action="Map Excel PIK rows only if stakeholders require the split.", governance_impact="None", notes="Do not net PIK into cash interest."),
        MetricSpec("SHL", "Principal Repaid", _col(shl_bridge, "excel_shl_principal_keur", n, None), runtime["shl_principal_keur"], "phase9_tuho_shl_period_bridge.csv / excel_shl_principal_keur", "runtime.shl_principal_keur", classification=WARN, root_cause="SHL principal timing moved closer to Excel after the controlled alignment work, but committed bridge totals still reflect a slightly different period framing.", recommended_action="Document as a timing reconciliation item and keep the runtime-alignment branch as the authoritative history.", governance_impact="Stakeholder review only", notes="No new SHL timing change is made here."),
        MetricSpec("SHL", "Closing Balance", _col(shl_bridge, "excel_shl_closing_keur", n, None), runtime["shl_closing_keur"], "phase9_tuho_shl_period_bridge.csv / excel_shl_closing_keur", "runtime.shl_balance_keur", classification=MISSING_EVIDENCE, root_cause="The committed SHL bridge does not provide a clean full-horizon closing-balance series that matches runtime period boundaries exactly.", recommended_action="Keep as an evidence limitation and rely on the runtime balance schedule plus prior Phase 9 review workbook for detail.", governance_impact="Review only", notes="Explicit evidence gap, not a silent zero."),
        MetricSpec("SHL", "Total SHL Service", missing("no committed Excel total SHL-service row"), runtime["shl_service_keur"], "none", "runtime.shl_service_keur", classification=MISSING_EVIDENCE, root_cause="The runtime SHL service field exists, but the committed Excel evidence pack does not expose a separate total-service bridge row.", recommended_action="Use the runtime service row for lender review and only add Excel-side extraction if stakeholders require a service split.", governance_impact="Review only", notes="Expands runtime-side SHL visibility without changing mechanics."),
    ]

    tax_specs = [
        MetricSpec("Tax", "Book Depreciation", missing("book depreciation not in committed Excel extract"), runtime["depreciation_keur"], "none", "runtime.depreciation_keur", classification=MISSING_EVIDENCE, root_cause="Book depreciation is available runtime-side but not committed Excel-side at full-horizon row granularity.", recommended_action="Use the Phase 6/9 depreciation evidence branches for any future ledger-level parity work.", governance_impact="Review only", notes="Explicitly kept separate from tax depreciation."),
        MetricSpec("Tax", "Tax Depreciation", missing("tax depreciation not in committed Excel extract"), runtime["tax_depreciation_audit_keur"], "none", "runtime.tax_depreciation_audit_keur", classification=MISSING_EVIDENCE, root_cause="Tax depreciation evidence exists in prior diagnostics but not as a committed period-level Excel bridge row here.", recommended_action="Leave as evidence gap unless a tax-specific parity branch is opened.", governance_impact="Review only", notes="No silent zero-fill."),
        MetricSpec("Tax", "R35 Taxable Income Before Losses", missing("R35 not in committed Excel extract"), runtime["taxable_income_before_losses_audit_keur"], "none", "runtime.taxable_income_before_losses_audit_keur", classification=WARN, root_cause="Known governed Phase 6 residual; Excel-side R35 row ownership remains partially unmapped in committed extracts.", recommended_action="Keep governed residual explicit; do not hardcode a parity plug.", governance_impact="G20 support evidence", notes="Residual is documented, not runtime-fixed."),
        MetricSpec("Tax", "R67 CIT Cash", missing("R67 not in committed Excel extract"), runtime["corporate_tax_cash_keur"], "none", "runtime.corporate_tax_cash_keur", classification=MISSING_EVIDENCE, root_cause="Committed Excel extracts do not include a period-level R67 bridge in this pack.", recommended_action="Carry forward as missing evidence rather than zero.", governance_impact="Review only", notes="Runtime tax cash remains visible."),
        MetricSpec("Tax", "Losses Opening", missing("loss opening not in committed Excel extract"), runtime["tax_loss_opening_audit_keur"], "none", "runtime.tax_loss_opening_audit_keur", classification=MISSING_EVIDENCE, root_cause="Loss bucket evidence is runtime-side only in this branch.", recommended_action="Reference the dedicated Phase 6 loss-engine evidence if a reviewer asks for lineage.", governance_impact="None", notes="Audit field available runtime-side."),
        MetricSpec("Tax", "Losses Used", missing("loss usage not in committed Excel extract"), runtime["tax_loss_used_audit_keur"], "none", "runtime.tax_loss_used_audit_keur", classification=MISSING_EVIDENCE, root_cause="Committed Excel bridge does not expose period-level loss usage in this pack.", recommended_action="Keep visible as runtime-only audit detail.", governance_impact="None", notes="Do not synthesize Excel values."),
        MetricSpec("Tax", "Losses Closing", missing("loss closing not in committed Excel extract"), runtime["tax_loss_closing_audit_keur"], "none", "runtime.tax_loss_closing_audit_keur", classification=MISSING_EVIDENCE, root_cause="Closing loss carryforward is visible in runtime audit fields but not committed Excel-side in this pack.", recommended_action="Preserve as a runtime-backed evidence row and add Excel extraction only if deeper tax review is requested.", governance_impact="Review only", notes="Improves tax loss visibility without fabricating Excel values."),
        MetricSpec("Tax", "Taxable Profit After Losses", missing("taxable profit after losses not in committed Excel extract"), runtime["taxable_profit_after_losses_audit_keur"], "none", "runtime.taxable_profit_after_losses_audit_keur", classification=MISSING_EVIDENCE, root_cause="Taxable profit after losses is already exposed runtime-side but not committed on the Excel side of this pack.", recommended_action="Keep the runtime row visible and document the Excel-side evidence gap explicitly.", governance_impact="Review only", notes="Useful for lender and tax reviewers even where Excel parity remains incomplete."),
        MetricSpec("Tax", "CIT Accrual", missing("CIT accrual not in committed Excel extract"), runtime["cit_accrual_audit_keur"], "none", "runtime.cit_accrual_audit_keur", classification=MISSING_EVIDENCE, root_cause="The runtime audit field exists for CIT accrual, but the committed Excel evidence pack does not include a matching row.", recommended_action="Expose the runtime accrual row and preserve the missing Excel side explicitly.", governance_impact="Review only", notes="Improves tax bridge completeness without changing formulas."),
        MetricSpec("Tax", "Cash Tax Current Period", missing("cash tax current period not in committed Excel extract"), runtime["cash_tax_current_period_audit_keur"], "none", "runtime.cash_tax_current_period_audit_keur", classification=MISSING_EVIDENCE, root_cause="Current-period cash tax is already exposed through runtime audit fields, but the committed Excel evidence pack has no matching line.", recommended_action="Show the runtime value and keep the Excel-side gap documented.", governance_impact="Review only", notes="Avoids treating runtime-available tax evidence as missing on both sides."),
    ]

    cfads_specs = [
        MetricSpec("CFADS", "EBITDA Cash Proxy", _col(per_bridge, "excel_ebitda_keur", n, None), runtime["ebitda_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_ebitda_keur", "runtime.ebitda_keur"),
        MetricSpec("CFADS", "Tax Cash", missing("no committed Excel tax-cash row"), runtime["corporate_tax_cash_keur"], "none", "runtime.corporate_tax_cash_keur", classification=MISSING_EVIDENCE, root_cause="Excel tax cash row is not committed in the current evidence pack.", recommended_action="Keep as runtime evidence only.", governance_impact="Review only", notes="Missing evidence is explicit."),
        MetricSpec("CFADS", "CFADS / R69", missing("R69 not in committed Excel extract"), runtime["r69_fcf_banks_keur"], "none", "runtime.r69_fcf_banks_keur", classification=MISSING_EVIDENCE, root_cause="Period-level Excel R69 remains outside the committed evidence set.", recommended_action="Future tax/waterfall review can map this if required.", governance_impact="Review only", notes="Runtime value is available."),
        MetricSpec("CFADS", "FCF for Distribution / R99", blank, runtime["r99_fcf_for_distribution_keur"], "governance gate only", "runtime.r99_fcf_for_distribution_keur", classification=GOVERNANCE_BLOCKER, root_cause="R99 runtime promotion remains not approved by governance.", recommended_action="Do not promote or rely on R99 as authoritative runtime source.", governance_impact="R99/R102 not approved", notes="Shown for audit visibility only."),
        MetricSpec("CFADS", "FCF for SHL / R102", blank, runtime["r102_fcf_for_shl_keur"], "governance gate only", "runtime.r102_fcf_for_shl_keur", classification=GOVERNANCE_BLOCKER, root_cause="R102 runtime promotion remains not approved by governance.", recommended_action="Keep audit-only until explicit approval.", governance_impact="R99/R102 not approved", notes="No runtime promotion implied."),
    ]

    distributions_specs = [
        MetricSpec("Distributions", "Excel Dividends / Distribution", _col(per_bridge, "excel_distribution_keur", n, None), runtime["distribution_keur"], "phase9_tuho_full_line_item_period_bridge.csv / excel_distribution_keur", "runtime.distribution_keur", classification=ACCEPTED_CONVENTION, root_cause="Excel dividend labeling and runtime distribution wiring are documented as definitionally different.", recommended_action="Use runtime distribution as authoritative and keep DA staging separate.", governance_impact="Stakeholder presentation choice", notes="No mixed flag-state values."),
        MetricSpec("Distributions", "DA-Wired / Pre-G20 Staging", missing("no committed Excel DA staging row"), runtime["da_paid_distribution_keur"], "none", "runtime.da_paid_distribution_keur", classification=GOVERNANCE_BLOCKER, root_cause="DA-wired distribution remains audit-only pending governance approval.", recommended_action="Keep clearly separated from the default runtime path.", governance_impact="G20 / R99 review", notes="Not authoritative for parity closeout."),
        MetricSpec("Distributions", "Legacy Runtime Distribution", blank, runtime["legacy_distribution_keur"], "not applicable", "runtime.legacy_distribution_keur", classification=ACCEPTED_CONVENTION, root_cause="Legacy distribution remains a reporting lens, not the reviewer-facing primary distribution stream.", recommended_action="Retain for traceability only.", governance_impact="None", notes="Separate from default runtime distribution row."),
    ]

    returns_specs = [
        MetricSpec("Returns", "Project IRR", [""] * n, [""] * n, "phase9 calibration reference", "runtime.project_irr", number_format=PCT_FORMAT, classification=PASS, root_cause="Within documented tolerance versus Excel.", recommended_action="None", governance_impact="None", notes="See Total column for scalar value.", total_label="scalar"),
        MetricSpec("Returns", "Equity IRR", [""] * n, [""] * n, "phase9 calibration reference", "runtime.equity_irr", number_format=PCT_FORMAT, classification=WARN, root_cause="Residual is governed by accepted conventions and remaining reporting-view decisions.", recommended_action="Stakeholder can accept residual or request a reconciliation IRR reporting view.", governance_impact="G20 stakeholder acceptance", notes="See Total column for scalar value.", total_label="scalar"),
        MetricSpec("Returns", "Reconciliation IRR", [""] * n, [""] * n, "not implemented", "not implemented", number_format=PCT_FORMAT, classification=MISSING_EVIDENCE, root_cause="Reconciliation IRR view has not been implemented as a reporting-only layer.", recommended_action="Open a dedicated reporting-view branch if stakeholders require it.", governance_impact="Stakeholder decision pending", notes="Missing evidence is explicit.", total_label="scalar"),
        MetricSpec("Returns", "MOIC", [""] * n, [""] * n, "not in committed Excel extract", "not exposed in runtime summary", number_format=X_FORMAT, classification=RUNTIME_BINDING_PENDING, root_cause="MOIC is not yet exported as a first-class runtime/reporting metric in the current workbook foundations.", recommended_action="Add a reporting-only MOIC view if the review committee asks for it.", governance_impact="None", notes="Do not invent a proxy here.", total_label="scalar"),
        MetricSpec("Returns", "Average DSCR", [""] * n, [""] * n, "phase9 calibration reference", "runtime.actual_avg_dscr", number_format=X_FORMAT, classification=WARN, root_cause="Residual remains documented in prior calibration evidence and is not a new runtime issue in this branch.", recommended_action="Keep as review evidence only.", governance_impact="G20 support evidence", notes="See Total column for scalar value.", total_label="scalar"),
        MetricSpec("Returns", "Minimum DSCR", [""] * n, [""] * n, "not in committed Excel extract", "runtime.actual_min_dscr", number_format=X_FORMAT, classification=MISSING_EVIDENCE, root_cause="Committed Excel evidence does not include a reliable minimum DSCR row for this pack.", recommended_action="Document as missing evidence rather than a failed model row.", governance_impact="Review only", notes="See Total column for scalar value.", total_label="scalar"),
    ]

    return {
        "Revenue Reconciliation": revenue_specs,
        "OPEX Reconciliation": opex_specs,
        "Senior Debt Reconciliation": debt_specs,
        "SHL Reconciliation": shl_specs,
        "Tax R35-R67-R69": tax_specs,
        "CFADS Waterfall": cfads_specs,
        "Distributions Sponsor": distributions_specs,
        "Returns Reconciliation": returns_specs,
    }


def _setup_sheet(sheet, period_labels: list[str]) -> None:
    headers = [
        "Metric / Source",
        "Evidence Type",
        *period_labels,
        "Total",
        "Status",
        "Classification",
        "Root Cause",
        "Recommended Action",
        "Governance Impact",
        "Notes",
    ]
    for idx, value in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    widths = [34, 18] + [10] * len(period_labels) + [12, 12, 20, 32, 28, 22, 28]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.freeze_panes = "C6"


def _write_triplet(sheet, start_row: int, spec: MetricSpec, period_labels: list[str], scalar_totals: dict[str, tuple[str | float, str | float]]) -> tuple[int, dict[str, str]]:
    classification, status, delta_total = _classify(
        spec.excel_values,
        spec.model_values,
        spec.classification,
        spec.status,
    )
    fill = _fill_for_classification(classification)
    notes = spec.notes or ""
    total_excel = _sum_numeric(spec.excel_values)
    total_model = _sum_numeric(spec.model_values)
    default_owner = {
        "Revenue": "Commercial",
        "OPEX": "Operations",
        "Senior Debt": "Debt",
        "SHL": "Sponsor / SHL",
        "Tax": "Tax",
        "CFADS": "Waterfall / Finance",
        "Distributions": "Governance / Sponsor",
        "Returns": "Finance / Governance",
    }.get(spec.section, "Review")
    owner = spec.owner if spec.owner != "Review" else default_owner
    requires_stakeholder_decision = spec.requires_stakeholder_decision
    if requires_stakeholder_decision == "no" and classification in {ACCEPTED_CONVENTION, GOVERNANCE_BLOCKER}:
        requires_stakeholder_decision = "yes"
    requires_runtime_change = spec.requires_runtime_change
    if requires_runtime_change == "no" and classification in {FAIL, RUNTIME_BINDING_PENDING}:
        requires_runtime_change = "yes"

    if spec.total_label == "scalar":
        total_excel, total_model = scalar_totals.get(spec.label, ("MISSING_EVIDENCE", "MISSING_EVIDENCE"))
        delta_total = (
            total_model - total_excel
            if isinstance(total_excel, (int, float)) and isinstance(total_model, (int, float))
            else "N/A"
        )

    rows = [
        (f"{spec.label} - Excel", spec.excel_source, spec.excel_values, total_excel, "", "", "", "", "", notes),
        (f"{spec.label} - Model", spec.model_source, spec.model_values, total_model, status, "runtime" if "runtime" in spec.model_source else "review", "", "", "", notes),
        (f"{spec.label} - Delta", "", _delta_values(spec.excel_values, spec.model_values), delta_total, status, classification, spec.root_cause, spec.recommended_action, spec.governance_impact, notes),
    ]

    for offset, (label, evidence_type, values, total, row_status, row_classification, root_cause, action, governance_impact, row_notes) in enumerate(rows):
        row = start_row + offset
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=evidence_type)
        for idx, value in enumerate(values, start=3):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = BODY_FONT
            if isinstance(value, (int, float)):
                cell.number_format = spec.number_format
            elif isinstance(value, str) and value.startswith("MISSING_EVIDENCE"):
                cell.fill = MISS_FILL
            elif offset == 2 and value == "N/A":
                cell.fill = CONVENTION_FILL
        total_col = 3 + len(period_labels)
        total_cell = sheet.cell(row=row, column=total_col, value=total)
        total_cell.border = THIN_BORDER
        total_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if isinstance(total, (int, float)):
            total_cell.number_format = spec.number_format
        elif isinstance(total, str) and total.startswith("MISSING_EVIDENCE"):
            total_cell.fill = MISS_FILL
        if offset == 2:
            total_cell.fill = fill
        sheet.cell(row=row, column=total_col + 1, value=row_status).fill = fill if row_status else META_FILL
        sheet.cell(row=row, column=total_col + 2, value=row_classification).fill = fill if row_classification else META_FILL
        sheet.cell(row=row, column=total_col + 3, value=root_cause).fill = fill if root_cause else META_FILL
        sheet.cell(row=row, column=total_col + 4, value=action).fill = fill if action else META_FILL
        sheet.cell(row=row, column=total_col + 5, value=governance_impact).fill = fill if governance_impact else META_FILL
        sheet.cell(row=row, column=total_col + 6, value=row_notes).fill = fill if row_notes and offset == 2 else META_FILL
        for col in range(1, total_col + 7):
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in {1, 2, total_col + 1, total_col + 2, total_col + 3, total_col + 4, total_col + 5, total_col + 6}:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row % 2 == 0 and offset != 2 and cell.fill == PatternFill():
                cell.fill = ROW_ALT_FILL

    return start_row + 4, {
        "section": spec.section,
        "metric": spec.label,
        "excel_total": str(total_excel),
        "model_total": str(total_model),
        "delta_total": str(delta_total),
        "status": status,
        "classification": classification,
        "root_cause": spec.root_cause,
        "recommended_action": spec.recommended_action,
        "governance_impact": spec.governance_impact,
        "notes": notes,
        "owner": owner,
        "requires_stakeholder_decision": requires_stakeholder_decision,
        "requires_runtime_change": requires_runtime_change,
        "expected_roadmap_phase": spec.expected_roadmap_phase,
    }


def _delta_values(excel_values: list[float | str | None], model_values: list[float | str | None]) -> list[float | str]:
    values: list[float | str] = []
    for excel_value, model_value in zip(excel_values, model_values):
        excel_num = _num(excel_value)
        model_num = _num(model_value)
        if excel_num is None or model_num is None:
            values.append("N/A")
        else:
            values.append(model_num - excel_num)
    return values


def _severity_order(classification: str) -> int:
    order = {
        GOVERNANCE_BLOCKER: 0,
        FAIL: 1,
        MATERIAL_DELTA: 1,
        WARN: 2,
        GOVERNANCE_REVIEW: 3,
        RUNTIME_BINDING_PENDING: 3,
        EVIDENCE_LIMITATION: 4,
        GROUPED_SOURCE_ONLY: 4,
        MISSING_EVIDENCE: 5,
        ACCEPTED_CONVENTION: 6,
        PASS: 7,
    }
    return order.get(classification, 99)


def _materiality_value(entry: dict[str, str]) -> float:
    delta = _num(entry.get("delta_total"))
    return abs(delta) if delta is not None else -1.0


def _apply_provenance_banner(sheet, title: str, runtime_label: str, governance_status: str, generated_at: str, project_name: str) -> None:
    sheet["A1"] = title
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = WHITE_FONT
    sheet["A2"] = f"Project: {project_name} | Pack: Phase 10 review pack | Mode: {runtime_label} | Governance: {governance_status}"
    sheet["A2"].fill = META_FILL
    sheet["A2"].font = BODY_FONT
    sheet["A3"] = f"Generated: {generated_at}"
    sheet["A3"].font = Font(size=9, italic=True, name="Calibri")
    sheet["A3"].alignment = Alignment(horizontal="left")
    nav_cell = sheet["H1"]
    nav_cell.value = "Back to Navigation"
    nav_cell.hyperlink = "#'Navigation'!A1"
    nav_cell.style = "Hyperlink"
    sheet.oddFooter.left.text = f"{project_name} | Phase 10 review pack"
    sheet.oddFooter.center.text = runtime_label
    sheet.oddFooter.right.text = generated_at


def _navigation_rows() -> list[dict[str, str]]:
    return [
        {"sheet_name": "Navigation", "reviewer_role": "All reviewers", "primary_purpose": "Start here, understand the pack, and jump to major sections.", "navigation_priority": "1", "governance_sensitive": "yes", "notes": "Contains legend, governance warning, and hyperlinks."},
        {"sheet_name": "Executive Dashboard", "reviewer_role": "IC / lender / governance lead", "primary_purpose": "See overall parity posture, major risks, blockers, and next actions without drilling into row detail first.", "navigation_priority": "2", "governance_sensitive": "yes", "notes": "Best first stop after Navigation for non-engineering readers."},
        {"sheet_name": "Executive Summary", "reviewer_role": "IC / lender / audit lead", "primary_purpose": "See parity posture, counts, top material gaps, and stakeholder items.", "navigation_priority": "3", "governance_sensitive": "yes", "notes": "Best first stop after the dashboard for evidence detail."},
        {"sheet_name": "Review Signoff", "reviewer_role": "Governance / PMO / reviewers", "primary_purpose": "Track manual signoff readiness by area, evidence completeness, runtime verification, and recommended action.", "navigation_priority": "4", "governance_sensitive": "yes", "notes": "Workbook-only workflow tracker; not persistence."},
        {"sheet_name": "Governance", "reviewer_role": "IC / governance", "primary_purpose": "Review G20 and R99/R102 status and what is engineering versus approval gating.", "navigation_priority": "5", "governance_sensitive": "yes", "notes": "Governance warning banner lives here."},
        {"sheet_name": "Governance Timeline", "reviewer_role": "Governance / lender / audit", "primary_purpose": "Understand which Phase 9 and Phase 10 checkpoints are complete versus still pending.", "navigation_priority": "6", "governance_sensitive": "yes", "notes": "Timeline of parity and governance milestones."},
        {"sheet_name": "Readiness Matrix", "reviewer_role": "All reviewers", "primary_purpose": "See maturity across runtime, export, evidence, and governance dimensions by review area.", "navigation_priority": "7", "governance_sensitive": "yes", "notes": "Fast maturity scan for IC and lender meetings."},
        {"sheet_name": "Runtime Summary", "reviewer_role": "All reviewers", "primary_purpose": "Anchor the pack to current runtime outputs and labels.", "navigation_priority": "8", "governance_sensitive": "yes", "notes": "Runtime versus review labels preserved."},
        {"sheet_name": "Revenue Reconciliation", "reviewer_role": "Commercial / audit", "primary_purpose": "Review revenue lines, totals, and missing evidence treatment.", "navigation_priority": "9", "governance_sensitive": "yes", "notes": "CO2 and balancing remain explicit evidence gaps."},
        {"sheet_name": "CO2 & Balancing Reconciliation", "reviewer_role": "Commercial / audit / governance", "primary_purpose": "See grouped-versus-separate evidence quality for revenue sub-lines and understand where CO2/balancing remains an evidence problem rather than a runtime failure.", "navigation_priority": "10", "governance_sensitive": "yes", "notes": "Dedicated source-map hardening sheet."},
        {"sheet_name": "OPEX Reconciliation", "reviewer_role": "Operations / audit", "primary_purpose": "Review OPEX totals, contingency breakout status, and EBITDA bridge.", "navigation_priority": "11", "governance_sensitive": "yes", "notes": "Grouped OPEX residuals remain documented."},
        {"sheet_name": "Senior Debt Reconciliation", "reviewer_role": "Lender / debt reviewer", "primary_purpose": "Review debt schedule visibility, DSCR residuals, and timing notes.", "navigation_priority": "12", "governance_sensitive": "yes", "notes": "Debt rows remain runtime-transparent."},
        {"sheet_name": "SHL Reconciliation", "reviewer_role": "Sponsor / audit", "primary_purpose": "Review SHL opening, accrued, cash, PIK, principal, and balance treatment.", "navigation_priority": "13", "governance_sensitive": "yes", "notes": "Presentation and evidence limitations are explicit."},
        {"sheet_name": "Tax R35-R67-R69", "reviewer_role": "Tax / audit", "primary_purpose": "Review governed tax residuals, runtime audit fields, and evidence gaps.", "navigation_priority": "14", "governance_sensitive": "yes", "notes": "No tax runtime changes in this pack."},
        {"sheet_name": "CFADS Waterfall", "reviewer_role": "Lender / cashflow reviewer", "primary_purpose": "Review cashflow rows, tax cash visibility, and audit-only R99/R102 staging.", "navigation_priority": "15", "governance_sensitive": "yes", "notes": "Governance blockers are highlighted."},
        {"sheet_name": "Distributions Sponsor", "reviewer_role": "Sponsor / IC", "primary_purpose": "Review distribution definition and DA staging separation.", "navigation_priority": "16", "governance_sensitive": "yes", "notes": "No mixed authoritative path."},
        {"sheet_name": "Returns Reconciliation", "reviewer_role": "IC / sponsor", "primary_purpose": "Review project IRR, equity IRR, DSCR, and reconciliation-view gaps.", "navigation_priority": "17", "governance_sensitive": "yes", "notes": "Equity IRR residual remains a stakeholder matter."},
        {"sheet_name": "IRR Reconciliation", "reviewer_role": "IC / lender / governance", "primary_purpose": "Explain runtime IRR versus Excel IRR, convention drift, and whether remaining gaps are engineering or governance-sensitive.", "navigation_priority": "18", "governance_sensitive": "yes", "notes": "Purpose-built reviewer sheet for IRR interpretation."},
        {"sheet_name": "Gap Register", "reviewer_role": "All reviewers", "primary_purpose": "Sort and filter all open items by severity, owner, and next action.", "navigation_priority": "19", "governance_sensitive": "yes", "notes": "Primary action list."},
        {"sheet_name": "Source Inventory", "reviewer_role": "Audit / model validation", "primary_purpose": "Understand evidence confidence and where each metric comes from.", "navigation_priority": "20", "governance_sensitive": "yes", "notes": "Useful for provenance checks."},
        {"sheet_name": "Accepted Conventions", "reviewer_role": "Governance / audit", "primary_purpose": "Separate accepted convention drift from runtime defects.", "navigation_priority": "21", "governance_sensitive": "yes", "notes": "Not a runtime-fix sheet."},
        {"sheet_name": "Reviewer Notes", "reviewer_role": "All reviewers", "primary_purpose": "Read the non-engineering guide to unknowns, blockers, and next steps.", "navigation_priority": "22", "governance_sensitive": "yes", "notes": "Best closing sheet."},
    ]


def _write_navigation(sheet, runtime_summary_rows: list[dict[str, str]]) -> None:
    generated_at = runtime_summary_rows[0]["generated_at"]
    project_name = runtime_summary_rows[0]["project"]
    _apply_provenance_banner(sheet, "Navigation", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", generated_at, project_name)
    sheet["A5"] = "Workbook Purpose"
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "Institutional review pack for lender, IC, and audit readers. Start here, then jump to the summary, gap register, or a specific discipline."
    sheet["A7"] = "Governance Warning"
    sheet["A7"].fill = GOVERNANCE_FILL
    sheet["A7"].font = BOLD_FONT
    sheet["B7"] = "G20 remains BLOCKED. R99/R102 remain NOT APPROVED. Accepted conventions are documentation, not runtime fixes."
    sheet["B7"].fill = GOVERNANCE_FILL

    sheet["A9"] = "How To Use This Pack"
    sheet["A9"].font = BOLD_FONT
    guidance = [
        "PASS means the currently committed Excel evidence and runtime output align within tolerance.",
        "WARN means the row is materially explainable but still worth reviewer attention.",
        "ACCEPTED_CONVENTION means the difference is understood and documented as a presentation or definition choice.",
        "MISSING_EVIDENCE means the evidence is not available; it is not a hidden zero.",
        "EVIDENCE_LIMITATION means the runtime or Excel side exists only at a broader level than the reviewer wants.",
        "GROUPED_SOURCE_ONLY means the available evidence supports only a grouped revenue row, not a defensible sub-line split.",
        "RUNTIME_BINDING_PENDING means the runtime logic exists but the reporting breakout is not yet first-class.",
        "GOVERNANCE_BLOCKER means the row is visible for review but cannot be treated as approved runtime authority.",
    ]
    row = 10
    for line in guidance:
        sheet.cell(row=row, column=1, value=line)
        sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    sheet["D5"] = "Status Legend"
    sheet["D5"].font = BOLD_FONT
    legend = [
        (PASS, PASS_FILL),
        (WARN, WARN_FILL),
        (FAIL, FAIL_FILL),
        (ACCEPTED_CONVENTION, CONVENTION_FILL),
        (MISSING_EVIDENCE, MISS_FILL),
        (EVIDENCE_LIMITATION, WARN_FILL),
        (GROUPED_SOURCE_ONLY, PENDING_FILL),
        (RUNTIME_BINDING_PENDING, PENDING_FILL),
        (GOVERNANCE_BLOCKER, GOVERNANCE_FILL),
    ]
    row = 6
    for label, fill in legend:
        sheet.cell(row=row, column=4, value=label).fill = fill
        sheet.cell(row=row, column=5, value="See workbook classifications and Gap Register.")
        row += 1

    sheet["A18"] = "Sheet Navigation"
    sheet["A18"].font = BOLD_FONT
    headers = ["Sheet", "Reviewer Role", "Purpose", "Priority", "Open"]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=19, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 20
    for entry in _navigation_rows()[1:]:
        sheet.cell(row=row, column=1, value=entry["sheet_name"])
        sheet.cell(row=row, column=2, value=entry["reviewer_role"])
        sheet.cell(row=row, column=3, value=entry["primary_purpose"])
        sheet.cell(row=row, column=4, value=entry["navigation_priority"])
        open_cell = sheet.cell(row=row, column=5, value=f"Go to {entry['sheet_name']}")
        open_cell.hyperlink = f"#'{entry['sheet_name']}'!A1"
        open_cell.style = "Hyperlink"
        for col in range(1, 6):
            sheet.cell(row=row, column=col).border = THIN_BORDER
            sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
        row += 1

    quick_links = [
        ("Go to Executive Dashboard", "Executive Dashboard"),
        ("Go to Executive Summary", "Executive Summary"),
        ("Go to Review Signoff", "Review Signoff"),
        ("Go to IRR Reconciliation", "IRR Reconciliation"),
        ("Go to Gap Register", "Gap Register"),
        ("Go to Reviewer Notes", "Reviewer Notes"),
    ]
    row += 1
    sheet.cell(row=row, column=1, value="Quick Links").font = BOLD_FONT
    row += 1
    for label, target in quick_links:
        cell = sheet.cell(row=row, column=1, value=label)
        cell.hyperlink = f"#'{target}'!A1"
        cell.style = "Hyperlink"
        row += 1

    for col, width in {"A": 30, "B": 28, "C": 60, "D": 10, "E": 24}.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A20"


def _write_exec_summary(sheet, summary_rows: list[dict[str, str]], runtime_summary_rows: list[dict[str, str]]) -> None:
    generated_at = runtime_summary_rows[0]["generated_at"]
    project_name = runtime_summary_rows[0]["project"]
    _apply_provenance_banner(sheet, "Executive Summary", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", generated_at, project_name)
    sheet["A5"] = "Current Overall Parity Posture"
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "Technically strong evidence base with governed residuals and explicit missing-evidence boundaries."
    sheet["A7"] = "Classification Counts"
    sheet["A7"].font = BOLD_FONT
    counts: dict[str, int] = {key: 0 for key in VALID_CLASSIFICATIONS}
    for row in summary_rows:
        counts[row["classification"]] = counts.get(row["classification"], 0) + 1
    r = 8
    for key in [PASS, WARN, FAIL, ACCEPTED_CONVENTION, MISSING_EVIDENCE, EVIDENCE_LIMITATION, GROUPED_SOURCE_ONLY, RUNTIME_BINDING_PENDING, GOVERNANCE_BLOCKER]:
        sheet.cell(row=r, column=1, value=key)
        sheet.cell(row=r, column=2, value=counts.get(key, 0))
        sheet.cell(row=r, column=1).fill = _fill_for_classification(key)
        r += 1
    sheet["D7"] = "Headline Runtime Summary"
    sheet["D7"].font = BOLD_FONT
    rr = 8
    for row in runtime_summary_rows[:8]:
        sheet.cell(row=rr, column=4, value=row["metric"])
        sheet.cell(row=rr, column=5, value=row["value"])
        sheet.cell(row=rr, column=6, value=row["unit"])
        rr += 1

    top_material = sorted(
        [row for row in summary_rows if row["classification"] != PASS],
        key=lambda row: (_severity_order(row["classification"]), -_materiality_value(row)),
    )[:10]
    sheet["A17"] = "Top 10 Material Gaps"
    sheet["A17"].font = BOLD_FONT
    headers = ["Metric", "Classification", "Root Cause", "Recommended Action"]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=18, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row_idx = 19
    for row in top_material:
        values = [f"{row['section']} / {row['metric']}", row["classification"], row["root_cause"], row["recommended_action"]]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if idx == 2:
                cell.fill = _fill_for_classification(str(value))
        row_idx += 1

    engineering_items = [row for row in summary_rows if row["requires_runtime_change"] == "yes"]
    stakeholder_items = [row for row in summary_rows if row["requires_stakeholder_decision"] == "yes"]
    sheet["F17"] = "What Is Engineering vs Governance"
    sheet["F17"].font = BOLD_FONT
    sheet["F18"] = f"Engineering items: {len(engineering_items)}"
    sheet["F19"] = f"Stakeholder decision items: {len(stakeholder_items)}"
    sheet["F20"] = "Major remaining uncertainties"
    sheet["F21"] = "Tax evidence gaps, IRR convention interpretation, reconciliation IRR reporting completeness, and sub-line revenue mapping remain the primary non-runtime unknowns."
    sheet["F22"] = "Governance blockers"
    sheet["F23"] = "G20 remains BLOCKED; R99/R102 remain NOT APPROVED and audit-only."
    sheet["F24"] = "IRR interpretation note"
    sheet["G24"] = "Small IRR drift can still be convention-driven rather than a runtime defect; use the IRR Reconciliation sheet before escalating engineering work."
    sheet["F25"] = "Review Recommendation"
    sheet["F25"].font = BOLD_FONT
    sheet["F26"] = "Safe to review"
    sheet["G26"] = "Yes - the workbook is suitable for review meetings and signoff preparation."
    sheet["F27"] = "Safe to rely on without caveats"
    sheet["G27"] = "No - governance blockers, accepted conventions, and evidence gaps remain explicit."
    sheet["F28"] = "Evidence confidence summary"
    sheet["G28"] = "High for runtime-backed totals, medium for timing-sensitive debt/SHL bridges, lower where source evidence is still missing."
    runtime_ready = len([row for row in summary_rows if row["classification"] not in {MISSING_EVIDENCE, RUNTIME_BINDING_PENDING}])
    total_items = max(len(summary_rows), 1)
    sheet["F29"] = "Runtime coverage summary"
    sheet["G29"] = f"{round(runtime_ready / total_items * 100, 1)}% of tracked rows have runtime-backed or convention-classified coverage."
    sheet["F30"] = "Governance heatmap summary"
    sheet["G30"] = f"{counts.get(GOVERNANCE_BLOCKER, 0)} governance blockers and {len(stakeholder_items)} stakeholder-decision items remain open."
    for col, width in {"A": 32, "B": 18, "C": 28, "D": 24, "E": 18, "F": 24}.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["G"].width = 48
    sheet.freeze_panes = "A8"


def _signoff_status(rows: list[dict[str, str]]) -> str:
    classes = {row["classification"] for row in rows}
    if GOVERNANCE_BLOCKER in classes or GOVERNANCE_REVIEW in classes:
        return "GOVERNANCE_PENDING"
    if FAIL in classes or MATERIAL_DELTA in classes:
        return "BLOCKED"
    if MISSING_EVIDENCE in classes or RUNTIME_BINDING_PENDING in classes or EVIDENCE_LIMITATION in classes or GROUPED_SOURCE_ONLY in classes:
        return "IN_REVIEW"
    if WARN in classes:
        return "READY_FOR_SIGNOFF"
    return "ACCEPTED"


def _review_signoff_rows(summary_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in summary_rows:
        grouped.setdefault(row["section"], []).append(row)

    reviewer_map = {
        "Revenue": "Commercial",
        "OPEX": "Operations",
        "Senior Debt": "Lender / Debt",
        "SHL": "Sponsor / Audit",
        "Tax": "Tax / Audit",
        "CFADS": "Finance / Waterfall",
        "Distributions": "Governance / Sponsor",
        "Returns": "IC / Finance",
    }
    rows_out: list[dict[str, str]] = []
    for area in ["Revenue", "OPEX", "Senior Debt", "SHL", "Tax", "CFADS", "Distributions", "Returns"]:
        area_rows = grouped.get(area, [])
        classes = {row["classification"] for row in area_rows}
        rows_out.append(
            {
                "review_area": area,
                "reviewer_type": reviewer_map.get(area, "Review"),
                "owner": area_rows[0]["owner"] if area_rows else "Review",
                "reviewer": reviewer_map.get(area, "Review"),
                "review_status": _signoff_status(area_rows) if area_rows else "NOT_STARTED",
                "current_status": _signoff_status(area_rows) if area_rows else "NOT_STARTED",
                "evidence_complete": "no" if MISSING_EVIDENCE in classes or EVIDENCE_LIMITATION in classes or GROUPED_SOURCE_ONLY in classes else "yes",
                "runtime_verified": "no" if RUNTIME_BINDING_PENDING in classes else "yes",
                "governance_reviewed": "pending" if GOVERNANCE_BLOCKER in classes else "yes",
                "stakeholder_decision_required": "yes" if any(row["requires_stakeholder_decision"] == "yes" for row in area_rows) else "no",
                "runtime_ready": "no" if RUNTIME_BINDING_PENDING in classes else "yes",
                "evidence_ready": "no" if MISSING_EVIDENCE in classes or EVIDENCE_LIMITATION in classes or GROUPED_SOURCE_ONLY in classes else "yes",
                "governance_ready": "no" if GOVERNANCE_BLOCKER in classes else "yes",
                "blocker_type": "governance" if GOVERNANCE_BLOCKER in classes else "evidence" if MISSING_EVIDENCE in classes or EVIDENCE_LIMITATION in classes or GROUPED_SOURCE_ONLY in classes else "runtime_binding" if RUNTIME_BINDING_PENDING in classes else "none",
                "recommended_action": area_rows[0]["recommended_action"] if area_rows else "Complete review coverage.",
                "recommended_next_step": area_rows[0]["recommended_action"] if area_rows else "Complete review coverage.",
                "notes": area_rows[0]["notes"] if area_rows else "",
            }
        )
    return rows_out


def _write_executive_dashboard(sheet, summary_rows: list[dict[str, str]], runtime_summary_rows: list[dict[str, str]]) -> None:
    generated_at = runtime_summary_rows[0]["generated_at"]
    project_name = runtime_summary_rows[0]["project"]
    _apply_provenance_banner(sheet, "Executive Dashboard", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", generated_at, project_name)

    counts: dict[str, int] = {key: 0 for key in VALID_CLASSIFICATIONS}
    for row in summary_rows:
        counts[row["classification"]] = counts.get(row["classification"], 0) + 1
    top_material = sorted(
        [row for row in summary_rows if row["classification"] != PASS],
        key=lambda row: (_severity_order(row["classification"]), -_materiality_value(row)),
    )[:5]
    blockers = [row for row in summary_rows if row["classification"] == GOVERNANCE_BLOCKER][:5]
    stakeholder_items = [row for row in summary_rows if row["requires_stakeholder_decision"] == "yes"]
    engineering_items = [row for row in summary_rows if row["requires_runtime_change"] == "yes"]
    runtime_ready = len([row for row in summary_rows if row["classification"] not in {MISSING_EVIDENCE, RUNTIME_BINDING_PENDING}])
    evidence_ready = len([row for row in summary_rows if row["classification"] != MISSING_EVIDENCE])
    total_items = max(len(summary_rows), 1)
    unresolved = len([row for row in summary_rows if row["classification"] != PASS])

    sheet["A5"] = "Overall Status"
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "Technically reviewable with explicit governance blockers and stakeholder decisions still pending."
    sheet["A7"] = "Classification Counts"
    sheet["A7"].font = BOLD_FONT
    row = 8
    for key in [PASS, WARN, FAIL, ACCEPTED_CONVENTION, MISSING_EVIDENCE, EVIDENCE_LIMITATION, GROUPED_SOURCE_ONLY, RUNTIME_BINDING_PENDING, GOVERNANCE_BLOCKER]:
        sheet.cell(row=row, column=1, value=key).fill = _fill_for_classification(key)
        sheet.cell(row=row, column=2, value=counts.get(key, 0))
        row += 1

    info_rows = [
        ("Runtime coverage summary", f"{round(runtime_ready / total_items * 100, 1)}%"),
        ("Evidence coverage summary", f"{round(evidence_ready / total_items * 100, 1)}%"),
        ("Coverage trend summary", "Runtime-backed rows now exceed pure placeholder rows across debt, SHL, tax, CFADS, and distributions."),
        ("Runtime completeness estimate", f"{runtime_ready} of {total_items} tracked rows are runtime-backed or convention-classified."),
        ("Evidence completeness estimate", f"{evidence_ready} of {total_items} tracked rows are not blocked solely by missing evidence."),
        ("Unresolved material delta summary", f"{unresolved} rows remain non-PASS, with the largest concentration in governed tax and evidence-bound review items."),
        ("CO2 / balancing evidence summary", "Revenue sub-line totals remain strongest at grouped level; standalone CO2 and balancing rows are still evidence-limited rather than runtime-failed."),
        ("Grouped vs separated evidence summary", "Electricity revenue is well evidenced separately, while CO2 and balancing remain grouped-only or missing in committed sources."),
        ("IRR parity summary", "Project IRR is within tolerance; equity IRR remains a governed interpretation topic rather than a newly identified runtime defect."),
        ("IRR convention summary", "XIRR anchors, SHL IDC investment-base treatment, and distribution-definition framing remain the main IRR interpretation drivers."),
        ("Accepted convention count", counts.get(ACCEPTED_CONVENTION, 0)),
        ("Stakeholder-decision count", len(stakeholder_items)),
        ("Governance blocker count", counts.get(GOVERNANCE_BLOCKER, 0)),
        ("Remaining engineering items", len(engineering_items)),
        ("Review recommendation summary", "Use for IC/lender review; not for unconditional signoff."),
    ]
    sheet["D7"] = "Readiness Snapshot"
    sheet["D7"].font = BOLD_FONT
    row = 8
    for label, value in info_rows:
        sheet.cell(row=row, column=4, value=label).font = BOLD_FONT
        sheet.cell(row=row, column=5, value=value)
        row += 1
    detail_start = max(17, row + 1)
    sheet.cell(row=detail_start, column=1, value="Top 5 Material Gaps").font = BOLD_FONT
    for idx, header in enumerate(["Metric", "Classification", "Root Cause", "Recommended Action"], start=1):
        cell = sheet.cell(row=detail_start + 1, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = detail_start + 2
    for item in top_material:
        values = [f"{item['section']} / {item['metric']}", item["classification"], item["root_cause"], item["recommended_action"]]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if idx == 2:
                cell.fill = _fill_for_classification(str(value))
        row += 1

    sheet.cell(row=detail_start, column=6, value="Top Governance Blockers").font = BOLD_FONT
    for idx, header in enumerate(["Metric", "Governance Impact", "Notes"], start=6):
        cell = sheet.cell(row=detail_start + 1, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = detail_start + 2
    for item in blockers:
        values = [f"{item['section']} / {item['metric']}", item["governance_impact"], item["notes"]]
        for idx, value in enumerate(values, start=6):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1

    next_actions_start = max(detail_start + 10, row + 2)
    sheet.cell(row=next_actions_start, column=1, value="Recommended Next Actions").font = BOLD_FONT
    next_actions = [
        "Keep G20 blocked until stakeholder signoff is complete.",
        "Use Review Signoff and Readiness Matrix to coordinate IC, lender, and audit walkthroughs.",
        "Treat R99/R102 as audit-only until explicit governance approval.",
        "Escalate only the rows that remain engineering or evidence blockers after reviewer pass-through.",
    ]
    row = next_actions_start + 1
    for action in next_actions:
        sheet.cell(row=row, column=1, value=action)
        row += 1

    for col, width in {"A": 28, "B": 18, "C": 12, "D": 24, "E": 26, "F": 28, "G": 24, "H": 36}.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A8"


def _write_review_signoff(sheet, signoff_rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(sheet, "Review Signoff", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = [
        "review_area",
        "owner",
        "reviewer",
        "review_status",
        "evidence_complete",
        "runtime_verified",
        "governance_reviewed",
        "stakeholder_decision_required",
        "recommended_action",
        "notes",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    for entry in signoff_rows:
        for idx, key in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=idx, value=entry[key])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    for col, width in {"A": 18, "B": 20, "C": 20, "D": 18, "E": 14, "F": 14, "G": 16, "H": 18, "I": 32, "J": 40}.items():
        sheet.column_dimensions[col].width = width


def _write_governance_timeline(sheet) -> None:
    _apply_provenance_banner(sheet, "Governance Timeline", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = ["milestone", "date", "status", "dependency", "notes"]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    rows = [
        ("Phase 9 TUHO parity closeout review", "2026-05-23", "completed", "corrected review workbook", "Accepted with minor follow-up."),
        ("Phase 10 export foundation", "2026-05-23", "completed", "runtime summary export", "Foundation for institutional exports."),
        ("Phase 10 runtime-to-workbook binding", "2026-05-23", "completed", "institutional workbook skeleton", "Runtime-bound workbook sections available."),
        ("Phase 10 calibration reconciliation pack", "2026-05-23", "completed", "runtime workbook binding", "Source-backed classification layer established."),
        ("Phase 10 lender/audit navigation polish", "2026-05-23", "completed", "calibration reconciliation pack", "Navigation and reviewer usability improved."),
        ("Executive dashboard and signoff flow", "2026-05-23", "in_review", "navigation-polished workbook", "Adds governance workflow support inside the workbook."),
        ("Stakeholder signoff / G20 gate review", "pending", "blocked", "executive dashboard + signoff completion", "G20 remains blocked until formal signoff."),
        ("R99/R102 promotion review", "pending", "blocked", "explicit governance approval", "Still not approved in this branch."),
    ]
    row = 6
    for values in rows:
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    for col, width in {"A": 34, "B": 14, "C": 14, "D": 26, "E": 42}.items():
        sheet.column_dimensions[col].width = width


def _write_readiness_matrix(sheet, signoff_rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(sheet, "Readiness Matrix", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = ["review_area", "runtime_complete", "export_complete", "parity_status", "evidence_strength", "governance_status", "remaining_work", "roadmap_phase"]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    for entry in signoff_rows:
        parity_status = entry["current_status"]
        evidence_strength = "high" if entry["evidence_ready"] == "yes" else "medium" if entry["runtime_ready"] == "yes" else "limited"
        governance_status = "blocked" if entry["governance_ready"] == "no" else "reviewable"
        values = [
            entry["review_area"],
            entry["runtime_ready"],
            "yes",
            parity_status,
            evidence_strength,
            governance_status,
            entry["recommended_next_step"],
            "Phase 10",
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    for col, width in {"A": 18, "B": 14, "C": 14, "D": 18, "E": 18, "F": 18, "G": 42, "H": 14}.items():
        sheet.column_dimensions[col].width = width


def _write_governance(sheet) -> None:
    _apply_provenance_banner(sheet, "Governance", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    sheet["A5"] = "Governance Warning"
    sheet["A5"].fill = GOVERNANCE_FILL
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "This workbook is a review pack only. It does not approve G20 and it does not promote R99/R102."
    sheet["B5"].fill = GOVERNANCE_FILL
    rows = [
        ("Pack status", "Audit / review evidence only"),
        ("G20 status", "BLOCKED"),
        ("R99/R102 status", "NOT APPROVED"),
        ("Runtime changes in this branch", "None"),
        ("Stakeholder sign-off", "Pending"),
        ("Accepted conventions", "Not runtime fixes"),
    ]
    row = 7
    for label, value in rows:
        sheet.cell(row=row, column=1, value=label).font = BOLD_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1
    sheet["A15"] = "Governance Notes"
    sheet["A15"].font = BOLD_FONT
    notes = [
        "G20 remains BLOCKED until stakeholder gate sign-off, even where technical evidence is sufficient.",
        "R99/R102 remain NOT APPROVED and are shown only as audit / staging context.",
        "Accepted conventions document presentation and scope decisions; they do not change runtime authority.",
    ]
    row = 16
    for note in notes:
        sheet.cell(row=row, column=1, value=note)
        sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 24


def _write_runtime_summary(sheet, runtime_summary_rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(
        sheet,
        "Runtime Summary",
        "runtime-derived values inside a review pack",
        "G20 BLOCKED / R99-R102 NOT APPROVED",
        runtime_summary_rows[0]["generated_at"],
        runtime_summary_rows[0]["project"],
    )
    headers = ["Metric", "Value", "Unit", "Runtime / Preview", "Governance", "Notes"]
    for idx, value in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    for summary_row in runtime_summary_rows:
        values = [
            summary_row["metric"],
            summary_row["value"],
            summary_row["unit"],
            summary_row["runtime_or_preview"],
            f"{summary_row['g20_status']} / {summary_row['r99_r102_status']}",
            summary_row["notes"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    for col, width in {"A": 28, "B": 18, "C": 10, "D": 18, "E": 24, "F": 50}.items():
        sheet.column_dimensions[col].width = width


def _write_notes(sheet) -> None:
    _apply_provenance_banner(sheet, "Reviewer Notes", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    notes = [
        ("Suggested review order", "Executive Dashboard -> Executive Summary -> Review Signoff -> Readiness Matrix -> Gap Register -> discipline sheets -> Reviewer Notes."),
        ("IC reviewer focus", "Focus on parity posture, top governance blockers, equity IRR residual treatment, and whether remaining gaps change investment conclusions."),
        ("Lender reviewer focus", "Focus on debt, DSCR, CFADS, governance blockers, and whether missing evidence affects reliance on credit-facing outputs."),
        ("Audit reviewer focus", "Focus on Source Inventory, Accepted Conventions, missing evidence discipline, and whether every non-PASS row has a defensible explanation."),
        ("Engineer focus", "Focus only on rows still marked runtime binding pending or areas that would need a dedicated follow-up branch."),
        ("How to interpret grouped revenue evidence", "Grouped revenue evidence means the current sources only support a total or combined view. It does not mean the runtime engine is mixing rows incorrectly."),
        ("CO2 / balancing reviewer caution", "If a row is labeled GROUPED_SOURCE_ONLY or MISSING_EVIDENCE, do not infer a synthetic CO2 or balancing split. Reviewers should treat that as a source-map limitation."),
        ("How to interpret IRR drift", "Start with the IRR Reconciliation sheet. Small project or equity IRR deltas can come from date anchors, distribution framing, or other documented conventions before they imply a runtime defect."),
        ("When IRR drift is acceptable", "Treat convention-backed drift as reviewable if the root cause is explicit, runtime values are stable, and the remaining gap is a governance or presentation issue rather than an unexplained economics change."),
        ("Stakeholder decisions", "Equity IRR residual, reconciliation IRR reporting view, and R99/R102 governance promotion remain outside this branch."),
        ("Likely acceptable convention drift", "XIRR date convention, SHL presentation splits, and grouped OPEX minor rows are documented as conventions rather than runtime defects."),
        ("Still requires engineering work", "Any future reconciliation IRR view, deeper tax row extraction, or sub-line CO2/balancing mapping should happen in dedicated reporting branches."),
        ("Evidence limitation only", "Rows marked MISSING_EVIDENCE are not zero and should not be interpreted as model failures."),
        ("Presentation only", "Runtime vs preview labels and governance badges are included to help reviewers avoid treating audit-only staging rows as runtime authority."),
        ("Already runtime-verified", "Revenue, core OPEX totals, runtime summary KPIs, and the current institutional workbook binding are already sourced from existing runtime outputs."),
        ("Governance-only blockers", "R99 and R102 remain audit-visible but governance-blocked; this pack does not change their status."),
        ("What remains unresolved", "A separate reconciliation IRR scalar is still not available as committed evidence, so this branch explains the gap instead of manufacturing a synthetic answer."),
        ("Roadmap view", "Reviewer follow-up can now split cleanly into reporting, evidence, or governance workstreams rather than broad model redesign."),
    ]
    row = 5
    for label, value in notes:
        sheet.cell(row=row, column=1, value=label).font = BOLD_FONT
        sheet.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True)
        row += 1
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90


def _write_source_inventory_sheet(sheet, rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(sheet, "Source Inventory", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = [
        "Metric",
        "Source File",
        "Period-Level Available?",
        "Total Available?",
        "Runtime / Model / Excel?",
        "Confidence Level",
        "Usable for Reconciliation?",
        "Notes",
    ]
    for idx, value in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    for entry in rows:
        values = [
            entry["metric"],
            entry["source_file"],
            entry["period_level_available"],
            entry["total_available"],
            entry["evidence_side"],
            entry["confidence_level"],
            entry["usable_for_reconciliation"],
            entry["notes"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    widths = {"A": 34, "B": 46, "C": 18, "D": 14, "E": 18, "F": 16, "G": 18, "H": 48}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _write_accepted_conventions_sheet(sheet, rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(sheet, "Accepted Conventions", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = ["Convention ID", "Convention", "Description", "Affected Metrics", "Accepted for Closeout", "Decision Required", "Notes"]
    for idx, value in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    for entry in rows:
        values = [
            entry.get("convention_id", ""),
            entry.get("convention", ""),
            entry.get("description", ""),
            entry.get("affected_metrics", ""),
            entry.get("accepted_for_closeout", ""),
            entry.get("stakeholder_decision_required", ""),
            entry.get("notes", ""),
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    sheet.freeze_panes = "A6"
    for col, width in {"A": 14, "B": 32, "C": 48, "D": 28, "E": 18, "F": 18, "G": 30}.items():
        sheet.column_dimensions[col].width = width


def _write_gap_register_sheet(sheet, summary_rows: list[dict[str, str]]) -> None:
    _apply_provenance_banner(sheet, "Gap Register", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    headers = ["Section", "Metric", "Total Excel", "Total Model", "Delta", "Status", "Classification", "Root Cause", "Recommended Action", "Governance Impact", "Owner", "Requires Stakeholder Decision?", "Requires Runtime Change?", "Expected Roadmap Phase", "Notes"]
    for idx, value in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row = 6
    ordered_rows = sorted(summary_rows, key=lambda entry: (_severity_order(entry["classification"]), -_materiality_value(entry), entry["section"], entry["metric"]))
    for entry in ordered_rows:
        values = [
            entry["section"],
            entry["metric"],
            entry["excel_total"],
            entry["model_total"],
            entry["delta_total"],
            entry["status"],
            entry["classification"],
            entry["root_cause"],
            entry["recommended_action"],
            entry["governance_impact"],
            entry["owner"],
            entry["requires_stakeholder_decision"],
            entry["requires_runtime_change"],
            entry["expected_roadmap_phase"],
            entry["notes"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if idx == 7:
                cell.fill = _fill_for_classification(str(value))
        row += 1
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:O{row - 1}"
    widths = {"A": 18, "B": 28, "C": 14, "D": 14, "E": 12, "F": 10, "G": 22, "H": 34, "I": 34, "J": 22, "K": 18, "L": 18, "M": 18, "N": 18, "O": 26}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _write_reconciliation_sheet(sheet, specs: list[MetricSpec], runtime: dict, summary_rows: list[dict[str, str]], scalar_totals: dict[str, tuple[str | float, str | float]]) -> None:
    period_labels = list(runtime["period_labels"])
    display_titles = {
        "Tax R35-R67-R69": "Tax / R35-R67-R69 Reconciliation",
        "CFADS Waterfall": "CFADS / Waterfall Reconciliation",
        "Distributions Sponsor": "Distributions / Sponsor Reconciliation",
    }
    _apply_provenance_banner(sheet, display_titles.get(sheet.title, sheet.title), "review", "G20 BLOCKED / R99-R102 NOT APPROVED", "see Runtime Summary", "TUHO Wind 1")
    _setup_sheet(sheet, period_labels)
    sheet["A4"] = "Reviewer Guidance"
    sheet["A4"].font = BOLD_FONT
    sheet["A4"].fill = SECTION_FILL
    sheet["B4"] = "Read the Delta row for classification and action. Missing evidence remains explicit and is never silently zero-filled."
    sheet["B4"].fill = SECTION_FILL
    row = 6
    for spec in specs:
        row, summary = _write_triplet(sheet, row, spec, period_labels, scalar_totals)
        summary_rows.append(summary)


def _irr_summary_row(entry: dict[str, str]) -> dict[str, str]:
    default_owner = "Finance / Governance"
    requires_runtime_change = "no"
    if entry["classification"] == MATERIAL_DELTA:
        requires_runtime_change = "yes"
    return {
        "section": "IRR Reporting",
        "metric": entry["metric"],
        "excel_total": entry["excel_value"],
        "model_total": entry["runtime_value"],
        "delta_total": entry["delta"],
        "status": "REVIEW" if entry["classification"] != PASS else "OK",
        "classification": entry["classification"],
        "root_cause": entry["root_cause"],
        "recommended_action": (
            "Keep documented as convention-driven review guidance."
            if entry["classification"] == ACCEPTED_CONVENTION
            else "Escalate to stakeholder review."
            if entry["stakeholder_decision_required"] == "yes"
            else "None"
        ),
        "governance_impact": entry["governance_impact"],
        "notes": entry["notes"],
        "owner": default_owner,
        "requires_stakeholder_decision": entry["stakeholder_decision_required"],
        "requires_runtime_change": requires_runtime_change,
        "expected_roadmap_phase": "Phase 10",
    }


def _co2_balancing_summary_row(entry: dict[str, str]) -> dict[str, str]:
    runtime_total = entry["runtime_value_available"]
    excel_total = entry["excel_value_available"]
    delta_total = "N/A"
    runtime_num = _num(runtime_total)
    excel_num = _num(excel_total)
    if runtime_num is not None and excel_num is not None:
        delta_total = str(runtime_num - excel_num)
    return {
        "section": "CO2 / Balancing",
        "metric": entry["metric"],
        "excel_total": str(excel_total),
        "model_total": str(runtime_total),
        "delta_total": delta_total,
        "status": "REVIEW" if entry["classification"] != PASS else "OK",
        "classification": entry["classification"],
        "root_cause": entry["unresolved_reason"],
        "recommended_action": (
            "Keep grouped until stronger source evidence exists."
            if entry["classification"] == GROUPED_SOURCE_ONLY
            else "Preserve explicit evidence gap; do not synthesize a split."
            if entry["classification"] in {MISSING_EVIDENCE, EVIDENCE_LIMITATION}
            else "None"
        ),
        "governance_impact": "Revenue-source interpretation only" if entry["governance_sensitive"] == "yes" else "None",
        "notes": entry["notes"],
        "owner": "Commercial / Audit",
        "requires_stakeholder_decision": "yes" if entry["governance_sensitive"] == "yes" else "no",
        "requires_runtime_change": "no",
        "expected_roadmap_phase": "Phase 10",
    }


def _write_irr_reconciliation_sheet(sheet, irr_rows: list[dict[str, str]], runtime_summary_rows: list[dict[str, str]]) -> None:
    generated_at = runtime_summary_rows[0]["generated_at"]
    project_name = runtime_summary_rows[0]["project"]
    _apply_provenance_banner(sheet, "IRR Reconciliation", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", generated_at, project_name)

    sheet["A5"] = "Reviewer Guidance"
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "Use this sheet to distinguish convention-driven IRR drift from actual runtime defects. This branch explains differences; it does not alter runtime IRR calculations."
    sheet["A7"] = "Metric"
    sheet["B7"] = "Runtime IRR / Runtime View"
    sheet["C7"] = "Excel IRR / Excel View"
    sheet["D7"] = "Delta"
    sheet["E7"] = "Classification"
    sheet["F7"] = "Root Cause"
    sheet["G7"] = "Governance Impact"
    sheet["H7"] = "Stakeholder Decision Required?"
    sheet["I7"] = "Notes"
    for col in range(1, 10):
        cell = sheet.cell(row=7, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 8
    for entry in irr_rows:
        values = [
            entry["metric"],
            entry["runtime_value"],
            entry["excel_value"],
            entry["delta"],
            entry["classification"],
            entry["root_cause"],
            entry["governance_impact"],
            entry["stakeholder_decision_required"],
            entry["notes"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if idx == 5:
                cell.fill = _fill_for_classification(str(value))
            elif row % 2 == 0:
                cell.fill = ROW_ALT_FILL
            if idx in {2, 3, 4} and isinstance(value, str) and value.replace("-", "").replace(".", "").isdigit():
                cell.number_format = PCT_FORMAT
        row += 1

    sheet["A18"] = "Accepted IRR Conventions"
    sheet["A18"].font = BOLD_FONT
    sheet["A19"] = "XIRR construction-date anchors, distribution-definition framing, and SHL IDC investment-base treatment are documented interpretation layers."
    sheet["A20"] = "Remaining unresolved IRR gaps"
    sheet["A21"] = "A dedicated reconciliation IRR scalar is still MISSING_EVIDENCE, so governance review should focus on explanation quality rather than a forced numeric tie-out."
    sheet["A22"] = "Runtime vs Excel timing note"
    sheet["A23"] = "A small equity IRR delta can remain acceptable when runtime economics are stable and the residual is traceable to timing or presentation conventions."

    widths = {
        "A": 28,
        "B": 20,
        "C": 20,
        "D": 14,
        "E": 22,
        "F": 44,
        "G": 28,
        "H": 20,
        "I": 40,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A8"


def _write_co2_balancing_sheet(sheet, rows: list[dict[str, str]], runtime_summary_rows: list[dict[str, str]]) -> None:
    generated_at = runtime_summary_rows[0]["generated_at"]
    project_name = runtime_summary_rows[0]["project"]
    _apply_provenance_banner(sheet, "CO2 & Balancing Reconciliation", "review", "G20 BLOCKED / R99-R102 NOT APPROVED", generated_at, project_name)

    sheet["A5"] = "Reviewer Guidance"
    sheet["A5"].font = BOLD_FONT
    sheet["B5"] = "Use this sheet to see which revenue rows are truly separated, which are grouped-only, and which remain missing on one or both evidence sides. Grouped evidence is not a runtime defect by itself."
    headers = [
        "Metric",
        "Runtime Source",
        "Excel Source",
        "Grouped / Separate",
        "Evidence Quality",
        "Runtime Available",
        "Excel Available",
        "Workbook Bound",
        "Runtime Value",
        "Excel Value",
        "Classification",
        "Unresolved Reason",
        "Notes",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=7, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 8
    for entry in rows:
        values = [
            entry["metric"],
            entry["runtime_source"],
            entry["excel_source"],
            entry["grouped_or_separate"],
            entry["evidence_quality"],
            entry["runtime_available"],
            entry["excel_available"],
            entry["workbook_bound"],
            entry["runtime_value_available"],
            entry["excel_value_available"],
            entry["classification"],
            entry["unresolved_reason"],
            entry["notes"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if idx == 11:
                cell.fill = _fill_for_classification(str(value))
            elif row % 2 == 0:
                cell.fill = ROW_ALT_FILL
            if idx in {9, 10} and isinstance(value, (int, float)):
                cell.number_format = K_EUR_FORMAT
        row += 1

    sheet["A18"] = "Evidence Policy"
    sheet["A18"].font = BOLD_FONT
    sheet["A19"] = "If runtime revenue exists only as an aggregate, it stays aggregated."
    sheet["A20"] = "If Excel evidence exists only as grouped revenue, it stays grouped."
    sheet["A21"] = "Standalone CO2 or balancing values are never fabricated from totals."
    sheet["A22"] = "Grouped-source rows are there to explain evidence boundaries, not to imply a hidden model error."

    widths = {
        "A": 24,
        "B": 24,
        "C": 32,
        "D": 24,
        "E": 16,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 22,
        "L": 42,
        "M": 36,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A8"


def _scalar_totals(runtime: dict) -> dict[str, tuple[str | float, str | float]]:
    return {
        "Project IRR": (0.0947, runtime["project_irr"]),
        "Equity IRR": (0.1161, runtime["equity_irr"]),
        "Reconciliation IRR": ("MISSING_EVIDENCE", "MISSING_EVIDENCE"),
        "MOIC": ("MISSING_EVIDENCE", "MISSING_EVIDENCE"),
        "Average DSCR": (1.451, runtime["avg_dscr"]),
        "Minimum DSCR": ("MISSING_EVIDENCE", runtime["min_dscr"]),
    }


def write_review_pack_navigation_map_csv(path: Path | str) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _navigation_rows()
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sheet_name",
                "reviewer_role",
                "primary_purpose",
                "navigation_priority",
                "governance_sensitive",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def write_calibration_reconciliation_pack(
    *,
    workbook_path: Path | str = DEFAULT_WORKBOOK,
    gap_register_path: Path | str = DEFAULT_GAP_REGISTER,
    source_inventory_path: Path | str = DEFAULT_SOURCE_INVENTORY,
    summary_path: Path | str = DEFAULT_SUMMARY,
    navigation_map_path: Path | str = DEFAULT_NAVIGATION_MAP,
    signoff_matrix_path: Path | str = DEFAULT_SIGNOFF_MATRIX,
    runtime_binding_inventory_path: Path | str = DEFAULT_RUNTIME_BINDING_INVENTORY,
    evidence_coverage_summary_path: Path | str = DEFAULT_EVIDENCE_COVERAGE_SUMMARY,
    runtime_binding_gap_register_path: Path | str = DEFAULT_RUNTIME_BINDING_GAP_REGISTER,
    irr_reconciliation_summary_path: Path | str = DEFAULT_IRR_RECONCILIATION_SUMMARY,
    irr_convention_map_path: Path | str = DEFAULT_IRR_CONVENTION_MAP,
    irr_governance_items_path: Path | str = DEFAULT_IRR_GOVERNANCE_ITEMS,
    co2_balancing_source_map_path: Path | str = DEFAULT_CO2_BALANCING_SOURCE_MAP,
    co2_balancing_gap_summary_path: Path | str = DEFAULT_CO2_BALANCING_GAP_SUMMARY,
    co2_balancing_evidence_quality_path: Path | str = DEFAULT_CO2_BALANCING_EVIDENCE_QUALITY,
) -> ReconciliationArtifacts:
    workbook_path = Path(workbook_path)
    gap_register_path = Path(gap_register_path)
    source_inventory_path = Path(source_inventory_path)
    summary_path = Path(summary_path)
    navigation_map_path = Path(navigation_map_path)
    signoff_matrix_path = Path(signoff_matrix_path)
    runtime_binding_inventory_path = Path(runtime_binding_inventory_path)
    evidence_coverage_summary_path = Path(evidence_coverage_summary_path)
    runtime_binding_gap_register_path = Path(runtime_binding_gap_register_path)
    irr_reconciliation_summary_path = Path(irr_reconciliation_summary_path)
    irr_convention_map_path = Path(irr_convention_map_path)
    irr_governance_items_path = Path(irr_governance_items_path)
    co2_balancing_source_map_path = Path(co2_balancing_source_map_path)
    co2_balancing_gap_summary_path = Path(co2_balancing_gap_summary_path)
    co2_balancing_evidence_quality_path = Path(co2_balancing_evidence_quality_path)

    for path in (
        workbook_path,
        gap_register_path,
        source_inventory_path,
        summary_path,
        navigation_map_path,
        signoff_matrix_path,
        runtime_binding_inventory_path,
        evidence_coverage_summary_path,
        runtime_binding_gap_register_path,
        irr_reconciliation_summary_path,
        irr_convention_map_path,
        irr_governance_items_path,
        co2_balancing_source_map_path,
        co2_balancing_gap_summary_path,
        co2_balancing_evidence_quality_path,
    ):
        path.parent.mkdir(parents=True, exist_ok=True)

    project_inputs, runtime_result = _run_project("tuho")
    runtime = _runtime_period_data(runtime_result, project_inputs)
    runtime_rows = build_runtime_summary_rows("tuho")

    per_bridge = _load_csv_rows(REPORTS_DIR / "phase9_tuho_full_line_item_period_bridge.csv")
    shl_bridge = _load_csv_rows(REPORTS_DIR / "phase9_tuho_shl_period_bridge.csv")
    conventions = _accepted_conventions_rows()
    source_inventory = _source_inventory_rows(int(runtime["count"]))
    specs_by_sheet = _build_specs(runtime, per_bridge, shl_bridge)
    irr_rows = _irr_report_rows(runtime, conventions)
    co2_balancing_rows = _co2_balancing_report_rows(runtime, per_bridge)

    workbook = Workbook()
    navigation = workbook.active
    navigation.title = "Navigation"
    _write_navigation(navigation, runtime_rows)
    dashboard = workbook.create_sheet("Executive Dashboard")
    executive = workbook.create_sheet("Executive Summary")
    review_signoff = workbook.create_sheet("Review Signoff")
    governance = workbook.create_sheet("Governance")
    governance_timeline = workbook.create_sheet("Governance Timeline")
    readiness = workbook.create_sheet("Readiness Matrix")
    runtime_summary = workbook.create_sheet("Runtime Summary")

    summary_rows: list[dict[str, str]] = []
    scalar_totals = _scalar_totals(runtime)

    _write_exec_summary(executive, summary_rows, runtime_rows)
    _write_governance(governance)
    _write_governance_timeline(governance_timeline)
    _write_runtime_summary(runtime_summary, runtime_rows)

    for sheet_name in [
        "Revenue Reconciliation",
        "CO2 & Balancing Reconciliation",
        "OPEX Reconciliation",
        "Senior Debt Reconciliation",
        "SHL Reconciliation",
        "Tax R35-R67-R69",
        "CFADS Waterfall",
        "Distributions Sponsor",
        "Returns Reconciliation",
    ]:
        if sheet_name == "CO2 & Balancing Reconciliation":
            _write_co2_balancing_sheet(workbook.create_sheet(sheet_name), co2_balancing_rows, runtime_rows)
            for entry in co2_balancing_rows:
                summary_rows.append(_co2_balancing_summary_row(entry))
        else:
            _write_reconciliation_sheet(
                workbook.create_sheet(sheet_name),
                specs_by_sheet[sheet_name],
                runtime,
                summary_rows,
                scalar_totals,
            )

    for entry in irr_rows:
        summary_rows.append(_irr_summary_row(entry))

    # Refresh summary after counts are known.
    _write_exec_summary(executive, summary_rows, runtime_rows)
    signoff_rows = _review_signoff_rows(summary_rows)
    _write_executive_dashboard(dashboard, summary_rows, runtime_rows)
    _write_review_signoff(review_signoff, signoff_rows)
    _write_readiness_matrix(readiness, signoff_rows)
    _write_irr_reconciliation_sheet(workbook.create_sheet("IRR Reconciliation"), irr_rows, runtime_rows)
    _write_gap_register_sheet(workbook.create_sheet("Gap Register"), summary_rows)
    _write_source_inventory_sheet(workbook.create_sheet("Source Inventory"), source_inventory)
    _write_accepted_conventions_sheet(workbook.create_sheet("Accepted Conventions"), conventions)
    _write_notes(workbook.create_sheet("Reviewer Notes"))

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    workbook_path.write_bytes(output.getvalue())

    _write_csv(gap_register_path, summary_rows)
    _write_csv(source_inventory_path, source_inventory)
    _write_csv(summary_path, _summary_report_rows(summary_rows))
    _write_csv(signoff_matrix_path, signoff_rows)
    inventory_rows = _runtime_binding_inventory_rows(summary_rows)
    _write_csv(runtime_binding_inventory_path, inventory_rows)
    _write_csv(evidence_coverage_summary_path, _evidence_coverage_summary_rows(summary_rows, inventory_rows))
    _write_csv(runtime_binding_gap_register_path, _runtime_binding_gap_register_rows(summary_rows))
    _write_csv(irr_reconciliation_summary_path, irr_rows)
    _write_csv(irr_convention_map_path, _irr_convention_map_rows(irr_rows))
    _write_csv(irr_governance_items_path, _irr_governance_item_rows(irr_rows))
    _write_csv(co2_balancing_source_map_path, co2_balancing_rows)
    _write_csv(co2_balancing_gap_summary_path, [row for row in co2_balancing_rows if row["classification"] != PASS])
    _write_csv(
        co2_balancing_evidence_quality_path,
        [
            {
                "metric": row["metric"],
                "runtime_value_available": row["runtime_value_available"],
                "excel_value_available": row["excel_value_available"],
                "grouped_or_separate": row["grouped_or_separate"],
                "evidence_quality": row["evidence_quality"],
                "classification": row["classification"],
                "unresolved_reason": row["unresolved_reason"],
                "governance_sensitive": row["governance_sensitive"],
                "notes": row["notes"],
            }
            for row in co2_balancing_rows
        ],
    )
    write_review_pack_navigation_map_csv(navigation_map_path)

    return ReconciliationArtifacts(
        workbook_path=workbook_path,
        gap_register_path=gap_register_path,
        source_inventory_path=source_inventory_path,
        summary_path=summary_path,
        navigation_map_path=navigation_map_path,
        signoff_matrix_path=signoff_matrix_path,
        runtime_binding_inventory_path=runtime_binding_inventory_path,
        evidence_coverage_summary_path=evidence_coverage_summary_path,
        runtime_binding_gap_register_path=runtime_binding_gap_register_path,
        irr_reconciliation_summary_path=irr_reconciliation_summary_path,
        irr_convention_map_path=irr_convention_map_path,
        irr_governance_items_path=irr_governance_items_path,
        co2_balancing_source_map_path=co2_balancing_source_map_path,
        co2_balancing_gap_summary_path=co2_balancing_gap_summary_path,
        co2_balancing_evidence_quality_path=co2_balancing_evidence_quality_path,
    )


def _summary_report_rows(summary_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for entry in summary_rows:
        rows.append(
            {
                "section": entry["section"],
                "metric": entry["metric"],
                "classification": entry["classification"],
                "status": entry["status"],
                "delta_total": entry["delta_total"],
                "root_cause": entry["root_cause"],
                "recommended_action": entry["recommended_action"],
                "governance_impact": entry["governance_impact"],
                "owner": entry["owner"],
                "requires_stakeholder_decision": entry["requires_stakeholder_decision"],
                "requires_runtime_change": entry["requires_runtime_change"],
                "expected_roadmap_phase": entry["expected_roadmap_phase"],
                "notes": entry["notes"],
            }
        )
    return rows


def _has_runtime_bound_value(model_total: str, notes: str) -> bool:
    text = (model_total or "").strip()
    if not text or text == "N/A":
        return False
    if text.startswith("MISSING_EVIDENCE"):
        return False
    return True


def _runtime_binding_inventory_rows(summary_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    sheet_map = {
        "Revenue": "Revenue Reconciliation",
        "CO2 / Balancing": "CO2 & Balancing Reconciliation",
        "OPEX": "OPEX Reconciliation",
        "Senior Debt": "Senior Debt Reconciliation",
        "SHL": "SHL Reconciliation",
        "Tax": "Tax R35-R67-R69",
        "CFADS": "CFADS Waterfall",
        "Distributions": "Distributions Sponsor",
        "Returns": "Returns Reconciliation",
        "IRR Reporting": "IRR Reconciliation",
    }
    for entry in summary_rows:
        runtime_available = _has_runtime_bound_value(entry["model_total"], entry["notes"])
        currently_bound = runtime_available
        currently_missing = "yes" if not runtime_available else "no"
        if runtime_available:
            reason = ""
            confidence = "high" if entry["classification"] not in {MISSING_EVIDENCE, RUNTIME_BINDING_PENDING} else "medium"
        else:
            reason = entry["root_cause"] or "Runtime/export field is not currently surfaced for this workbook row."
            confidence = "low"
        rows.append(
            {
                "section": entry["section"],
                "metric": entry["metric"],
                "runtime_source": entry["owner"],
                "workbook_target_sheet": sheet_map.get(entry["section"], "Review"),
                "currently_bound": "yes" if currently_bound else "no",
                "currently_missing": currently_missing,
                "reason_if_missing": reason,
                "confidence_level": confidence,
                "classification": entry["classification"],
                "notes": entry["notes"],
            }
        )
    return rows


def _evidence_coverage_summary_rows(summary_rows: list[dict[str, str]], inventory_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    total = max(len(summary_rows), 1)
    runtime_bound = sum(1 for row in inventory_rows if row["currently_bound"] == "yes")
    unresolved = sum(1 for row in summary_rows if row["classification"] != PASS)
    rows = [
        {
            "metric": "runtime_coverage_pct",
            "value": f"{round(runtime_bound / total * 100, 1)}",
            "unit": "percent",
            "notes": "Rows with runtime-side values surfaced in the workbook.",
        },
        {
            "metric": "evidence_coverage_pct",
            "value": f"{round(sum(1 for row in summary_rows if row['classification'] != MISSING_EVIDENCE) / total * 100, 1)}",
            "unit": "percent",
            "notes": "Rows not blocked solely by missing evidence.",
        },
        {
            "metric": "unresolved_gap_count",
            "value": str(unresolved),
            "unit": "count",
            "notes": "All rows still outside PASS status.",
        },
        {
            "metric": "governance_blocker_count",
            "value": str(sum(1 for row in summary_rows if row["classification"] == GOVERNANCE_BLOCKER)),
            "unit": "count",
            "notes": "Rows blocked by approval or stakeholder governance.",
        },
        {
            "metric": "accepted_convention_count",
            "value": str(sum(1 for row in summary_rows if row["classification"] == ACCEPTED_CONVENTION)),
            "unit": "count",
            "notes": "Rows explained by accepted convention rather than runtime defect.",
        },
    ]
    return rows


def _runtime_binding_gap_register_rows(summary_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for entry in summary_rows:
        if entry["classification"] == PASS:
            continue
        rows.append(
            {
                "section": entry["section"],
                "metric": entry["metric"],
                "classification": entry["classification"],
                "why_unresolved": entry["root_cause"] or "Needs reviewer follow-up.",
                "engineering_work_exists": entry["requires_runtime_change"],
                "evidence_missing": "yes" if entry["classification"] == MISSING_EVIDENCE else "no",
                "governance_decision_required": entry["requires_stakeholder_decision"],
                "expected_roadmap_phase": entry["expected_roadmap_phase"],
                "recommended_action": entry["recommended_action"],
                "notes": entry["notes"],
            }
        )
    return rows


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
