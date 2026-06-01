"""Workbook Index / README sheet builder for Phase 48.

Provides sheet-inventory and workbook summary content for pilot-facing
Excel exports. Presentation-only; does not change financial formulas or
model outputs.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
META_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_FILL = PatternFill("solid", fgColor="EEF3F8")
TABLE_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
TABLE_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
WARNING_FILL = PatternFill("solid", fgColor="FCE4E4")
HEADER_FONT = Font(bold=True, size=10, name="Calibri")
SECTION_FONT = Font(bold=True, size=11, name="Calibri")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
WARNING_FONT = Font(size=10, bold=True, name="Calibri")


# Sheet inventory rows for institutional workbook.
INSTITUTIONAL_SHEET_INVENTORY = [
    ("Export_Metadata", "Provenance, trust hygiene, non-claims", "Reviewers", "Review", "First sheet — Phase 47 added"),
    ("Workbook_Index", "This sheet — sheet inventory and workbook guide", "All users", "Review", "Phase 48 added"),
    ("Cover", "Cover page with project name and workbook type", "All users", "Review", ""),
    ("Governance", "Governance metadata and approval status", "Reviewers", "Review", ""),
    ("Runtime Summary", "Summary KPIs: Project IRR, Equity IRR, Avg DSCR, etc.", "All users", "Runtime", "Factory-bound runtime values"),
    ("Inputs", "Project metadata and assumption binding", "Reviewers", "Runtime+Template", ""),
    ("Construction", "Construction timing and funding summary", "Reviewers", "Template+Runtime", ""),
    ("OPEX", "OPEX assumptions and runtime total", "Reviewers", "Runtime+Template", ""),
    ("CAPEX", "CAPEX summary and funding split", "Reviewers", "Template+Runtime", ""),
    ("Revenue", "Revenue assumptions and period revenue table", "Reviewers", "Runtime+Template", ""),
    ("Senior Debt", "Senior debt assumptions and debt service schedule", "Reviewers", "Runtime+Template", ""),
    ("SHL", "Subordinated hybrid loan assumptions and cashflow/balance", "Reviewers", "Runtime+Template", ""),
    ("Tax", "Tax assumptions and tax bridge periods", "Reviewers", "Runtime+Template", ""),
    ("P&L", "Offline assembled P&L from runtime result", "Reviewers", "Runtime", ""),
    ("Cash Flow", "Offline PF cash waterfall from runtime result", "Reviewers", "Runtime", ""),
    ("Balance Sheet", "Offline balance sheet assembly from runtime result", "Reviewers", "Runtime", ""),
    ("Audit", "Runtime source notes, provenance, audit boundary", "Reviewers", "Review", ""),
    ("Gap Register", "Known gaps and accepted conventions", "Reviewers", "Review", ""),
]

CALIBRATION_SHEET_INVENTORY = [
    ("Export_Metadata", "Provenance, trust hygiene, non-claims", "Reviewers", "Review", "First sheet — Phase 47 added"),
    ("Workbook_Index", "This sheet — sheet inventory and workbook guide", "All users", "Review", "Phase 48 added"),
    ("Navigation", "Navigation links to all sheets", "All users", "Review", ""),
    ("Executive Dashboard", "High-level summary and KPIs", "Executives", "Review", ""),
    ("Executive Summary", "Summary of calibration findings", "Executives", "Review", ""),
    ("Review Signoff", "Manual signoff tracker by area", "PMO/Reviewers", "Review", "Not persistence layer"),
    ("Governance", "Governance metadata and approval status", "Reviewers", "Review", ""),
    ("Governance Timeline", "Governance timeline and milestones", "Reviewers", "Review", ""),
    ("Readiness Matrix", "Signoff readiness by area", "PMO/Reviewers", "Review", ""),
    ("Runtime Summary", "Runtime KPIs from existing values", "Reviewers", "Runtime", ""),
    ("Revenue Reconciliation", "Revenue calibration reconciliation", "Reviewers", "Runtime+Template", ""),
    ("CO2 & Balancing Reconciliation", "CO2 and balancing calibration", "Reviewers", "Runtime+Template", ""),
    ("OPEX Reconciliation", "OPEX calibration reconciliation", "Reviewers", "Runtime+Template", ""),
    ("Senior Debt Reconciliation", "Senior debt calibration", "Reviewers", "Runtime+Template", ""),
    ("SHL Reconciliation", "SHL calibration", "Reviewers", "Runtime+Template", ""),
    ("Tax R35-R67-R69", "Tax calibration", "Reviewers", "Runtime+Template", ""),
    ("CFADS Waterfall", "CFADS waterfall", "Reviewers", "Runtime", ""),
    ("Distributions Sponsor", "Sponsor distributions", "Reviewers", "Runtime", ""),
    ("Returns Reconciliation", "Returns calibration", "Reviewers", "Runtime+Template", ""),
]


def build_workbook_index_rows(
    inventory: list[tuple[str, str, str, str, str]],
    workbook_name: str,
    project_name: str = "",
    export_type: str = "",
) -> list[list[str]]:
    """Build rows for Workbook_Index sheet content.

    Returns rows ready for writing to a sheet.
    """
    rows = []
    # Section header: Workbook purpose
    rows.append(["Workbook Index"])
    rows.append([f"Workbook: {workbook_name}"])
    if project_name:
        rows.append([f"Project: {project_name}"])
    if export_type:
        rows.append([f"Export type: {export_type}"])
    rows.append([])  # blank separator

    # Section: Sheet Inventory
    rows.append(["Sheet Inventory"])
    rows.append(["Sheet Name", "Purpose", "Audience", "Status", "Notes"])
    for row in inventory:
        rows.append(list(row))
    rows.append([])

    return rows


def write_workbook_index_sheet(sheet, rows: list[list[str]]) -> None:
    """Write the Workbook_Index sheet from pre-built rows.

    Styling: title row, section headers, table headers,
    alternating rows, column widths.
    """
    # Row 1: title
    sheet["A1"] = rows[0][0]
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:E1")

    current_row = 2
    for row in rows[1:]:
        if not row:  # blank separator
            current_row += 1
            continue

        # Section header (non-table rows that look like headers)
        if len(row) == 1 or (len(row) == 2 and ":" in row[0]):
            # Section header
            sheet.cell(row=current_row, column=1, value=row[0])
            sheet.cell(row=current_row, column=1).font = SECTION_FONT
            sheet.cell(row=current_row, column=1).fill = SECTION_FILL
            sheet.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=5,
            )
            current_row += 1
        elif row[0] == "Sheet Name":
            # Table header row
            for col_idx, val in enumerate(row, start=1):
                cell = sheet.cell(row=current_row, column=col_idx, value=val)
                cell.fill = TABLE_HEADER_FILL
                cell.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            current_row += 1
        elif len(row) == 5 and row[0] != "Sheet Name":
            # Data row — alternate fill
            fill = TABLE_ALT_FILL if (current_row % 2 == 0) else PatternFill()
            for col_idx, val in enumerate(row, start=1):
                cell = sheet.cell(row=current_row, column=col_idx, value=val)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 1
        else:
            # Metadata row (e.g. "Workbook: ...")
            for col_idx, val in enumerate(row, start=1):
                sheet.cell(row=current_row, column=col_idx, value=val)
                sheet.cell(row=current_row, column=col_idx).font = BODY_FONT
            current_row += 1

    # Column widths
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 30


def write_workbook_index_sheet_full(
    sheet,
    *,
    workbook_name: str,
    project_name: str = "",
    export_type: str = "",
    inventory: list[tuple[str, str, str, str, str]],
    is_generic: bool = False,
) -> None:
    """Write a complete Workbook_Index sheet with all sections.

    Includes: title, workbook metadata, sheet inventory table,
    trusted pilot scope, generic boundary, export hygiene,
    internal review notice, non-claims, guardrails.
    """
    WARNING_FILL_LOCAL = PatternFill("solid", fgColor="FCE4E4")
    GUARDRAIL_FILL = PatternFill("solid", fgColor="FDEBD0")
    SCOPE_FILL = PatternFill("solid", fgColor="E7F3E8")

    sheet["A1"] = "Workbook Index"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:E1")

    r = 2
    # Metadata
    meta = [
        ("Workbook", workbook_name),
        ("Project", project_name or "Not specified"),
        ("Export type", export_type or "Not specified"),
        ("Added", "Phase 48 — Export Index / README Polish"),
    ]
    for label, val in meta:
        sheet.cell(row=r, column=1, value=label).font = HEADER_FONT
        sheet.cell(row=r, column=1).fill = META_FILL
        sheet.cell(row=r, column=2, value=val).font = BODY_FONT
        sheet.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        r += 1

    r += 1  # blank

    # ── Sheet Inventory ──────────────────────────────────────────
    sheet.cell(row=r, column=1, value="Sheet Inventory").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = SECTION_FILL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    # Table header
    for col_idx, header in enumerate(["Sheet Name", "Purpose", "Audience", "Status", "Notes"], start=1):
        cell = sheet.cell(row=r, column=col_idx, value=header)
        cell.fill = TABLE_HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    # Data rows
    for i, (sheet_name, purpose, audience, status, notes) in enumerate(inventory):
        fill = TABLE_ALT_FILL if (i % 2 == 0) else PatternFill()
        for col_idx, val in enumerate([sheet_name, purpose, audience, status, notes], start=1):
            cell = sheet.cell(row=r, column=col_idx, value=val)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1  # blank

    # ── Trusted Pilot Scope ──────────────────────────────────────
    sheet.cell(row=r, column=1, value="Trusted Pilot Scope").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = SCOPE_FILL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    scope_lines = [
        "TUHO frozen-template path",
        "Oborovo frozen-template path",
        "TUHO CO2 revenue treatment",
        "Oborovo OpEx",
        "Senior debt / DSCR / SHL frozen path",
        "Scenario / export / audit evidence as pilot support material",
    ]
    for line in scope_lines:
        sheet.cell(row=r, column=1, value=f"✅ {line}").font = BODY_FONT
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1

    # ── Generic Boundary ──────────────────────────────────────────
    sheet.cell(row=r, column=1, value="Generic Boundary").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = WARNING_FILL_LOCAL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    generic_lines = [
        "⚠️  Generic solar and wind paths are exploratory and unvalidated.",
        "⚠️  Not for financial decisions.",
        "⚠️  Do not use for lender, bank, audit, or certification purposes.",
    ]
    for line in generic_lines:
        sheet.cell(row=r, column=1, value=line).font = WARNING_FONT
        sheet.cell(row=r, column=1).fill = WARNING_FILL_LOCAL
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1

    # ── Export Hygiene ────────────────────────────────────────────
    sheet.cell(row=r, column=1, value="Export Hygiene").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = SECTION_FILL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    hygiene_lines = [
        "Exports reflect the last clean backend run, not unsaved draft edits.",
        "Re-run model after any input change before exporting.",
        "Treat runtime snapshot as stale if active draft changed after last clean run.",
        "Timestamp filenames to avoid stale export confusion.",
    ]
    for line in hygiene_lines:
        sheet.cell(row=r, column=1, value=f"— {line}").font = BODY_FONT
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1

    # ── Internal Review Notice ────────────────────────────────────
    sheet.cell(row=r, column=1, value="Audit Interpretation").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = SECTION_FILL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    audit_lines = [
        "Internal review evidence only.",
        "Not certified external audit.",
        "Not bank/lender approval.",
        "Not a replay engine.",
    ]
    for line in audit_lines:
        sheet.cell(row=r, column=1, value=f"— {line}").font = BODY_FONT
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1

    # ── NON-CLAIMS ────────────────────────────────────────────────
    sheet.cell(row=r, column=1, value="NON-CLAIMS").font = Font(bold=True, size=12, name="Calibri")
    sheet.cell(row=r, column=1).fill = WARNING_FILL_LOCAL
    sheet.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    non_claims_lines = [
        "NOT bank/lender approval.",
        "NOT external audit certification.",
        "NOT SaaS-ready or enterprise-ready.",
        "Controlled trusted pilot scope only.",
        "Paid pilot remains NOT READY.",
        "Enterprise SaaS remains NOT READY.",
    ]
    for line in non_claims_lines:
        sheet.cell(row=r, column=1, value=line).font = WARNING_FONT
        sheet.cell(row=r, column=1).fill = WARNING_FILL_LOCAL
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1

    # ── Guardrails ────────────────────────────────────────────────
    sheet.cell(row=r, column=1, value="Guardrails").font = SECTION_FONT
    sheet.cell(row=r, column=1).fill = GUARDRAIL_FILL
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    guardrail_lines = [
        "G20: BLOCKED",
        "R99: NOT APPROVED",
        "R102: NOT APPROVED",
        "partial_pay_sweep: not promoted",
        "flat/min DSCR sculpting: not promoted",
        "Backend: source of truth",
    ]
    for line in guardrail_lines:
        sheet.cell(row=r, column=1, value=line).font = BODY_FONT
        sheet.cell(row=r, column=1).fill = GUARDRAIL_FILL
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    # Column widths
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 30